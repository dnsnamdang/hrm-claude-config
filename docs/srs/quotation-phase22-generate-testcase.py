import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Quotation Phase 22"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

even_fill = PatternFill(start_color='FFF0F4FF', end_color='FFF0F4FF', fill_type='solid')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ In Báo Giá + Sửa Tỷ Suất LN (Assign/Quotation - Phase 22)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'Báo giá'
PREFIX = 'QP'


def write_section_row(title):
    global current_row
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=13)
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(vertical='center')
    for col in range(1, 14):
        ws.cell(row=current_row, column=col).border = thin_border
        if col != 3:
            ws.cell(row=current_row, column=col).fill = section_fill


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    current_row += 1
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority, precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    is_even = (current_row % 2 == 0)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if is_even:
            cell.fill = even_fill


# === Data Validation cho cột Status ===
dv = DataValidation(type='list', formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
dv.error = 'Chọn giá trị hợp lệ'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# =====================================================================
# SECTION 1: Modal cấu hình in
# =====================================================================
write_section_row('1. MODAL CẤU HÌNH IN BÁO GIÁ')

write_tc(1, 1, 'Hiển thị modal cấu hình in', 'P0',
    'Đã mở trang chi tiết báo giá',
    '1. Click button "In"',
    'Báo giá có sản phẩm',
    'Modal mở: title "Cấu hình in báo giá", checkbox "Hiện hàng hoá cấp con" (mặc định bật), 14 checkbox cột (tất cả check), checkbox "Chọn tất cả", 2 button "Huỷ" + "Xem trước"')

write_tc(1, 2, '14 cột hiển thị đúng tên', 'P1',
    'Modal cấu hình đã mở',
    '1. Kiểm tra danh sách checkbox cột',
    '',
    'STT, Tên hàng hoá, Mã hàng hoá, Model, Thương hiệu, Xuất xứ, ĐVT, Thông số kỹ thuật, Số lượng, Đơn giá bán, Thành tiền bán, VAT (%), Tiền VAT, Thành tiền sau VAT')

write_tc(1, 3, '"Chọn tất cả" toggle bỏ chọn', 'P0',
    'Tất cả cột đang check',
    '1. Bỏ check "Chọn tất cả"',
    '',
    'Tất cả 14 checkbox cột bị bỏ check')

write_tc(1, 4, '"Chọn tất cả" toggle chọn lại', 'P0',
    'Đã bỏ chọn vài cột',
    '1. Check "Chọn tất cả"',
    '',
    'Tất cả 14 checkbox cột được check lại')

write_tc(1, 5, 'Auto uncheck "Chọn tất cả" khi bỏ 1 cột', 'P1',
    'Tất cả cột đang check',
    '1. Bỏ check 1 cột (vd: Model)',
    '',
    'Checkbox "Chọn tất cả" tự bỏ check')

write_tc(1, 6, 'Auto check "Chọn tất cả" khi check đủ 14', 'P1',
    'Đã bỏ 1 cột → "Chọn tất cả" đang unchecked',
    '1. Check lại cột vừa bỏ',
    '',
    'Checkbox "Chọn tất cả" tự check lại')

write_tc(1, 7, 'Không chọn cột nào → warning', 'P0',
    'Tất cả cột đã bỏ check',
    '1. Click "Xem trước"',
    '',
    'Hiện text đỏ: "Vui lòng chọn ít nhất 1 cột để in". Modal KHÔNG đóng')

write_tc(1, 8, 'Warning ẩn khi chọn lại cột', 'P2',
    'Đang hiện warning (QP_001.007)',
    '1. Check 1 cột bất kỳ',
    '',
    'Thông báo đỏ biến mất')

write_tc(1, 9, 'Mặc định chọn tất cả khi mở lại', 'P1',
    'Đã mở modal, bỏ vài cột, đóng modal',
    '1. Click "In" mở modal lại',
    '',
    'Tất cả 14 cột được check lại (reset về mặc định)')

# =====================================================================
# SECTION 2: Preview + In
# =====================================================================
write_section_row('2. MODAL XEM TRƯỚC BÁO GIÁ')

write_tc(2, 1, 'Layout preview đúng cấu trúc', 'P0',
    'Đã cấu hình chọn tất cả cột',
    '1. Click "Xem trước"',
    'Báo giá có sản phẩm + dịch vụ',
    'Modal XL: Banner → Title "BÁO GIÁ HÀNG HÓA" → Header 2 cột → Bảng sản phẩm → Tổng VAT/sau VAT → Điều khoản → Ghi chú → Chữ ký')

write_tc(2, 2, 'Header info hiển thị đúng', 'P0',
    'Đã mở preview',
    '1. Kiểm tra 2 cột header info',
    'BG có đầy đủ KH, liên hệ, dự án',
    'Trái: Kính gửi, Tên đơn vị, Địa chỉ, ĐT/Email, Hiệu lực, Bảo hành. Phải: Mã BG, Dự án, Ngày, ĐDKD, Giao hàng, Loại tiền tệ')

write_tc(2, 3, 'Chỉ hiện cột đã chọn', 'P0',
    'Modal config đã mở',
    '1. Bỏ chọn: Model, Thương hiệu, Xuất xứ\n2. Click "Xem trước"',
    '',
    'Bảng chỉ hiện 11 cột. Header + data + tổng đều khớp 11 cột')

write_tc(2, 4, 'Nhóm La Mã hiển thị đúng', 'P0',
    'BG có sản phẩm phân nhóm',
    '1. Xem preview',
    '',
    'Header nhóm: "I. Tên nhóm", "II. ...". Background xanh nhạt, bold, colspan toàn bộ')

write_tc(2, 5, 'Hiện cấp con khi bật toggle', 'P0',
    'Check "Hiện hàng hoá cấp con"',
    '1. Xem trước BG có cha+con',
    '',
    'Dưới cha có row con: STT 1.1, 1.2... Text xám. Tên lùi 20px. Giá bán/VAT/thành tiền hiện "—"')

write_tc(2, 6, 'Ẩn cấp con khi tắt toggle', 'P0',
    'Bỏ check "Hiện hàng hoá cấp con"',
    '1. Xem trước',
    '',
    'Chỉ hiện cha + orphan. Không có row con, không có STT dạng x.y')

write_tc(2, 7, 'Dịch vụ bổ sung cuối bảng', 'P0',
    'BG có dịch vụ bổ sung',
    '1. Xem preview',
    '',
    'Sau nhóm SP: header "Dịch vụ bổ sung" (xanh nhạt, bold) + row DV với STT riêng (1,2...), mã DV-xxx, tên, giá, VAT...')

write_tc(2, 8, 'Row TỔNG CỘNG đúng', 'P0',
    'Xem preview',
    '1. Kiểm tra row cuối bảng',
    '',
    'Bold, border-top 2px, bg #f9f9f9. Cột tên hiện "TỔNG CỘNG". Cột thành tiền/VAT/sau VAT hiện tổng (SP + DV)')

write_tc(2, 9, 'Button In trên header modal', 'P0',
    'Preview đã mở',
    '1. Kiểm tra vị trí button "In"',
    '',
    'Button In (icon ri-printer-line) nằm cạnh title "Xem trước báo giá" ở header. KHÔNG có footer buttons')

write_tc(2, 10, 'In mở cửa sổ mới + print dialog', 'P0',
    'Preview đã mở',
    '1. Click "In"',
    '',
    'Mở cửa sổ trình duyệt mới chứa HTML báo giá. Sau ~300ms tự mở print dialog. Chỉ nội dung BG (không sidebar/topbar)')

write_tc(2, 11, 'Bản in A4 landscape đúng', 'P1',
    'Print dialog đã mở',
    '1. Kiểm tra print preview của trình duyệt',
    '',
    'A4 ngang, font Times New Roman, bảng font 10px, background nhóm/header in ra')

write_tc(2, 12, 'Bảng dài lặp header khi sang trang', 'P1',
    'BG có >20 sản phẩm',
    '1. In → kiểm tra trang 2+',
    '',
    'Header bảng lặp ở mỗi trang (display: table-header-group). Không cắt ngang row')

# =====================================================================
# SECTION 3: In từ trang danh sách
# =====================================================================
write_section_row('3. IN TỪ TRANG DANH SÁCH BÁO GIÁ')

write_tc(3, 1, 'Icon máy in trên row', 'P0',
    'Mở /assign/quotations',
    '1. Kiểm tra cột actions mỗi row',
    '',
    'Có icon máy in (ri-printer-line) trong cột actions')

write_tc(3, 2, 'Click In → load API → mở modal config', 'P0',
    'Trang danh sách',
    '1. Click icon máy in trên 1 row',
    '',
    'Gọi API GET /assign/quotations/{id}. Sau khi load xong → mở QuotationPrintConfigModal')

write_tc(3, 3, 'Flow in hoạt động đúng từ danh sách', 'P0',
    'Modal config đã mở từ danh sách',
    '1. Chọn cột → "Xem trước"\n2. Verify preview\n3. Click "In"',
    '',
    'Giống in từ trang chi tiết: preview đúng data, print dialog mở')

write_tc(3, 4, 'API lỗi → thông báo lỗi', 'P1',
    'Trang danh sách',
    '1. Click In trên BG đã bị xoá (hoặc ngắt mạng)',
    '',
    'Toast: "Không tải được chi tiết báo giá". Không crash, không mở modal')

write_tc(3, 5, 'Trang không crash khi chưa chọn BG', 'P0',
    'Load trang danh sách',
    '1. Browse bình thường, không click In',
    '',
    'Trang hoạt động bình thường. Không lỗi console (v-if="item" ngăn render khi null)')

# =====================================================================
# SECTION 4: Công thức tỷ suất LN
# =====================================================================
write_section_row('4. CÔNG THỨC TỶ SUẤT LỢI NHUẬN — (bán-nhập)/nhập')

write_tc(4, 1, 'marginPercent tổng dùng /totalImport (edit)', 'P0',
    'Mở sửa báo giá',
    '1. Nhập giá sao cho totalSale=120, totalImport=100\n2. Xem tỷ suất tổng',
    'totalSale=120, totalImport=100',
    'Tỷ suất = (120-100)/100 × 100 = 20%. KHÔNG phải 16.67% (/sale)')

write_tc(4, 2, 'totalImport = 0 → 0% (edit)', 'P0',
    'BG mới, chưa có giá nhập',
    '1. Nhập giá bán cho orphan\n2. Xem tỷ suất',
    'totalImport=0, totalSale=100',
    'Tỷ suất = 0%. Không lỗi chia cho 0')

write_tc(4, 3, 'marginPercent tổng dùng /totalImport (show)', 'P0',
    'Xem chi tiết BG',
    '1. Kiểm tra tỷ suất tổng',
    'totalSale=200, totalImport=150',
    'Tỷ suất = (200-150)/150 × 100 = 33.33%')

write_tc(4, 4, 'totalImport = 0 → 0% (show)', 'P1',
    'BG không có giá nhập',
    '1. Xem chi tiết',
    '',
    'Tỷ suất = 0%')

write_tc(4, 5, 'lineMarginPercent per-row dùng /import (edit)', 'P0',
    'Sửa BG có orphan',
    '1. Orphan: giá nhập=80, SL=2, giá bán=100, SL=2\n2. Xem cột tỷ suất dòng',
    'import=160, sale=200',
    'Tỷ suất dòng = (200-160)/160 × 100 = 25%')

write_tc(4, 6, 'lineMarginPercent cho CHA (roll-up import)', 'P0',
    'BG có cha + con',
    '1. Con: giá nhập 50, SL 3 → import con=150\n2. Cha SL=1 → giá nhập cha=150\n3. Cha giá bán=200',
    'import cha=150, sale cha=200',
    'Tỷ suất CHA = (200-150)/150 × 100 = 33.33%')

write_tc(4, 7, 'lineMarginPercent per-row (show)', 'P0',
    'Xem chi tiết BG',
    '1. Kiểm tra cột tỷ suất từng row',
    '',
    'Công thức giống edit: (thành_tiền_bán - thành_tiền_nhập) / thành_tiền_nhập × 100')

write_tc(4, 8, 'import = 0 → hiện "—"', 'P0',
    'SP có giá nhập = 0',
    '1. Kiểm tra cột tỷ suất dòng đó',
    'estimated_price=0, quoted_price=100',
    'Hiện "—" (null → text-muted). KHÔNG hiện NaN/Infinity')

write_tc(4, 9, 'BE computeTotals dùng /totalImport', 'P0',
    'Lưu BG có giá bán + nhập',
    '1. Lưu BG\n2. Kiểm tra DB field profit_margin',
    '',
    'profit_margin = (totalSale-totalImport)/totalImport × 100, guard totalImport > 0')

write_tc(4, 10, 'BE per-line profit_margin dùng /import', 'P1',
    'GET /assign/quotations/{id}',
    '1. Kiểm tra mỗi product trong response',
    '',
    'product.profit_margin = (quoted_total-est_total)/est_total × 100, guard est>0, else null')

# =====================================================================
# SECTION 5: Cảnh báo màu tỷ suất
# =====================================================================
write_section_row('5. CẢNH BÁO MÀU TỶ SUẤT LỢI NHUẬN')

write_tc(5, 1, 'Tổng tỷ suất đỏ khi < threshold (edit)', 'P0',
    'Cấu hình profit_margin_threshold = 15%',
    '1. Sửa BG sao cho tỷ suất tổng = 10%\n2. Xem màu cột tỷ suất',
    'threshold=15%, margin=10%',
    '10% hiện màu ĐỎ (text-danger) tại row tổng + ô thống kê')

write_tc(5, 2, 'Tổng tỷ suất xanh khi ≥ threshold (edit)', 'P0',
    'threshold = 15%',
    '1. Sửa BG sao cho tỷ suất = 20%',
    'threshold=15%, margin=20%',
    '20% hiện màu XANH (text-success)')

write_tc(5, 3, 'Per-row đỏ/xanh theo threshold', 'P0',
    'threshold = 15%',
    '1. SP A: margin=10%\n2. SP B: margin=25%\n3. Xem cột tỷ suất',
    '',
    'SP A: đỏ. SP B: xanh')

write_tc(5, 4, 'Dịch vụ bổ sung — cảnh báo đỏ', 'P0',
    'threshold = 15%',
    '1. DV: giá nhập=100, giá bán=105 → margin=5%',
    'imp=100, sale=105',
    'Cột tỷ suất DV hiện ĐỎ (5% < 15%). Có :class="marginColorClass" trên td')

write_tc(5, 5, 'Dịch vụ bổ sung — xanh khi đạt', 'P0',
    'threshold = 15%',
    '1. DV: giá nhập=100, giá bán=150 → margin=50%',
    'imp=100, sale=150',
    'Cột tỷ suất DV hiện XANH (50% ≥ 15%)')

write_tc(5, 6, 'Margin null → xám', 'P1',
    'SP/DV có giá nhập = 0',
    '1. Xem cột tỷ suất',
    '',
    '"—" màu XÁM (text-muted)')

write_tc(5, 7, 'Threshold = 0 → margin 0% = xanh', 'P2',
    'Cấu hình threshold = 0%',
    '1. SP có margin = 0%',
    'threshold=0%, margin=0%',
    'margin 0% ≥ threshold 0% → xanh')

write_tc(5, 8, 'Trang chi tiết: tổng + row + DV có cảnh báo', 'P0',
    'Xem chi tiết BG',
    '1. Kiểm tra cột tỷ suất: tổng, từng SP, từng DV',
    '',
    'Tất cả cell tỷ suất đều có màu đỏ/xanh/xám theo threshold')

write_tc(5, 9, 'svcMarginPercent dùng /import cho DV (show)', 'P0',
    'Xem chi tiết BG có DV',
    '1. DV: giá nhập=80, giá bán=100, SL=1',
    '',
    'Tỷ suất DV = (100-80)/80 × 100 = 25%')

# =====================================================================
# SECTION 6: Tính tổng nhập/bán
# =====================================================================
write_section_row('6. TÍNH TỔNG NHẬP / BÁN')

write_tc(6, 1, 'totalSale skip CON', 'P0',
    'BG có Cha (giá bán 200, SL 1) + Con',
    '1. Kiểm tra totalSale',
    'Cha sale=200×1',
    'totalSale = 200. KHÔNG cộng con (tránh đếm đôi)')

write_tc(6, 2, 'totalImport skip CHA (có con)', 'P0',
    'Cha + Con A (imp 50×2=100) + Con B (imp 30×1=30)',
    '1. Kiểm tra totalImport',
    '',
    'totalImport từ con: 100+30=130. KHÔNG cộng thêm cha')

write_tc(6, 3, 'DV bổ sung cộng cả 2 tổng', 'P0',
    'Thêm DV: giá nhập=80, giá bán=100, SL=1',
    '1. Kiểm tra totalImport + totalSale',
    '',
    'totalImport += 80. totalSale += 100')

write_tc(6, 4, 'Orphan tính cả import và sale', 'P0',
    'Orphan: giá nhập=60, giá bán=90, SL=2',
    '1. Kiểm tra totalImport + totalSale',
    '',
    'totalImport += 120. totalSale += 180')

# =====================================================================
# SECTION 7: E2E Flow
# =====================================================================
write_section_row('7. E2E — FLOW IN BÁO GIÁ ĐẦY ĐỦ')

write_tc(7, 1, 'E2E: In từ trang chi tiết', 'P0',
    'BG có 2 nhóm, cha+con, orphan, DV',
    '1. Mở chi tiết → Click "In"\n2. Bỏ chọn Model, Thương hiệu → "Xem trước"\n3. Verify: nhóm, SP, con (1.1), DV, tổng\n4. Click "In" → print dialog\n5. Huỷ in',
    '',
    'Mọi bước OK. 12 cột. Tổng = SP + DV. A4 ngang')

write_tc(7, 2, 'E2E: In từ trang danh sách', 'P0',
    'Trang danh sách BG',
    '1. Click icon In trên row\n2. Đợi load → modal config\n3. Chọn tất cả, bật cấp con → "Xem trước"\n4. Click "In"',
    '',
    'Giống E2E chi tiết. Data load từ API. Không crash')

write_tc(7, 3, 'In BG chỉ có DV (không có SP)', 'P1',
    'BG chỉ có dịch vụ bổ sung',
    '1. In từ chi tiết',
    '',
    'Hiện DV + tổng. Không có nhóm La Mã. Không lỗi')

write_tc(7, 4, 'In BG không có DV', 'P1',
    'BG chỉ có SP, không DV',
    '1. Xem preview',
    '',
    'Không hiện section "Dịch vụ bổ sung". Tổng = chỉ SP')

write_tc(7, 5, 'BG thiếu thông tin KH', 'P2',
    'BG không có liên hệ/địa chỉ/email',
    '1. Xem preview',
    '',
    'Fields thiếu hiện trống (không null/undefined). Layout không vỡ')


# === SAVE ===
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, 'quotation-phase22-testcases.xlsx')
wb.save(output_path)
print(f'Saved to {output_path}')
print(f'Total test cases: {current_row - 7 - 7}')  # subtract header rows + section rows
