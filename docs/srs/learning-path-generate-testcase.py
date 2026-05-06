import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Lo trinh hoc"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Lộ trình học (Training/LearningPath)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Đào tạo'
GROUP = 'Lộ trình học'
PREFIX = 'LTH'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. DANH SÁCH LỘ TRÌNH HỌC (/training/learning-path)
# =============================================
write_section_row('I. DANH SÁCH LỘ TRÌNH HỌC (/training/learning-path)')

write_tc(1, 1,
    'Hiển thị trang danh sách khi truy cập',
    'P0',
    'User đã đăng nhập, có quyền "Quản lý lộ trình học"',
    '1. Vào menu Đào tạo > Lộ trình học',
    '',
    'Hiển thị bộ lọc nhanh (tìm theo mã, tên)\nBảng có cột: STT, Mã+Tên lộ trình, Loại đào tạo, Người tạo, Số khoá học, Tổng bài học, Điều kiện đạt, Trạng thái, Cập nhật\nNút "Thêm mới" hiện')

write_tc(1, 2,
    'Hiển thị badge trạng thái đúng màu',
    'P0',
    'Có lộ trình ở đủ 3 trạng thái',
    '1. Quan sát cột Trạng thái',
    '3 lộ trình status 1, 2, 3',
    'Nháp = xám (#64748B)\nĐang dùng = xanh lá (#16A34A)\nKhoá = đỏ (#DC2626)')

write_tc(1, 3,
    'Tìm kiếm nhanh theo mã lộ trình',
    'P0',
    'Lộ trình LP-2026-00001 tồn tại',
    '1. Nhập "LP-2026" vào ô tìm kiếm nhanh\n2. Enter',
    'Keyword: LP-2026',
    'Hiển thị các lộ trình có code chứa "LP-2026"')

write_tc(1, 4,
    'Tìm kiếm nhanh theo tên lộ trình',
    'P0',
    'Có lộ trình tên "Onboarding Dev"',
    '1. Nhập "Onboarding" vào ô tìm kiếm nhanh\n2. Enter',
    'Keyword: Onboarding',
    'Hiển thị lộ trình có name chứa "Onboarding"')

write_tc(1, 5,
    'Filter theo Trạng thái',
    'P0',
    'Có lộ trình ở cả 3 trạng thái',
    '1. Mở bộ lọc nâng cao\n2. Chọn Trạng thái = "Đang dùng"\n3. Nhấn Tìm kiếm',
    'status = 2',
    'Chỉ hiển thị lộ trình status = 2')

write_tc(1, 6,
    'Filter theo Loại đào tạo',
    'P0',
    'Có nhiều loại đào tạo',
    '1. Mở bộ lọc nâng cao\n2. Chọn Loại đào tạo\n3. Tìm kiếm',
    'training_type_id = X',
    'Chỉ hiển thị lộ trình thuộc loại đào tạo X')

write_tc(1, 7,
    'Filter theo Điều kiện đạt',
    'P1',
    'Có lộ trình với các result_rule khác nhau',
    '1. Chọn "Tất cả khoá bắt buộc ĐẠT"\n2. Tìm kiếm',
    'result_rule = REQUIRED_ONLY',
    'Chỉ hiển thị lộ trình result_rule = REQUIRED_ONLY')

write_tc(1, 8,
    'Filter theo Người tạo',
    'P1',
    'Nhiều lộ trình do nhiều user tạo',
    '1. Chọn người tạo cụ thể\n2. Tìm kiếm',
    'created_by = user A',
    'Chỉ hiển thị lộ trình do user A tạo')

write_tc(1, 9,
    'Filter theo Người cập nhật gần nhất',
    'P1',
    'Nhiều lộ trình',
    '1. Chọn người cập nhật\n2. Tìm kiếm',
    'updated_by = user B',
    'Chỉ hiển thị lộ trình updated_by = B')

write_tc(1, 10,
    'Filter theo khoảng Ngày cập nhật',
    'P1',
    'Có lộ trình cập nhật trong nhiều ngày',
    '1. Nhập "Ngày cập nhật từ" và "Ngày cập nhật đến"\n2. Tìm kiếm',
    'Từ 01/04/2026, Đến 30/04/2026',
    'Chỉ hiển thị lộ trình có updated_at nằm trong khoảng')

write_tc(1, 11,
    'Kết hợp nhiều filter',
    'P1',
    'Dữ liệu phong phú',
    '1. Chọn Trạng thái = Đang dùng\n2. Chọn Loại đào tạo = X\n3. Tìm kiếm',
    '',
    'Kết quả thoả cả 2 điều kiện (AND)')

write_tc(1, 12,
    'Reset bộ lọc',
    'P1',
    'Đã filter, bảng đang lọc',
    '1. Nhấn nút Reset',
    '',
    'Tất cả filter trở về mặc định\nBảng hiển thị toàn bộ dữ liệu')

write_tc(1, 13,
    'Sort theo Mã lộ trình',
    'P1',
    'Nhiều lộ trình',
    '1. Click header cột "Mã lộ trình" 2 lần',
    '',
    'Lần 1: sort tăng dần code\nLần 2: sort giảm dần code')

write_tc(1, 14,
    'Sort theo Ngày cập nhật',
    'P1',
    'Nhiều lộ trình',
    '1. Click header cột "Cập nhật"',
    '',
    'Sort theo updated_at tăng/giảm dần')

write_tc(1, 15,
    'Phân trang - Chuyển trang',
    'P1',
    'Có > 20 lộ trình',
    '1. Click trang 2',
    '',
    'Bảng load dữ liệu trang 2, pagination cập nhật')

write_tc(1, 16,
    'Phân trang - Đổi số dòng/trang',
    'P1',
    'Có nhiều lộ trình',
    '1. Chọn 50 trong dropdown số dòng/trang',
    '',
    'Bảng hiển thị tối đa 50 lộ trình/trang')

write_tc(1, 17,
    'Action - Xem chi tiết',
    'P0',
    'Có lộ trình trong danh sách',
    '1. Click icon Xem ở dòng lộ trình',
    '',
    'Điều hướng tới /training/learning-path/{id}\nHiển thị chi tiết read-only')

write_tc(1, 18,
    'Action - Sửa chỉ hiện cho lộ trình Nháp',
    'P0',
    'Có lộ trình Nháp và Đang dùng',
    '1. Quan sát cột Hành động',
    '',
    'Lộ trình Nháp: icon Sửa hiện\nLộ trình Đang dùng / Khoá: icon Sửa ẩn')

write_tc(1, 19,
    'Action - Xoá chỉ hiện cho lộ trình Nháp',
    'P0',
    'Có lộ trình Nháp',
    '1. Click icon Xoá\n2. Confirm',
    'status = 1',
    'Lộ trình bị xoá khỏi DB, biến mất khỏi bảng\nLộ trình Đang dùng/Khoá: nút Xoá không hiện')

write_tc(1, 20,
    'Toggle Khoá từ Nháp → Khoá',
    'P0',
    'Lộ trình status = Nháp',
    '1. Click toggle khoá trên dòng\n2. Xác nhận',
    '',
    'Status chuyển thành Khoá (#DC2626)\nToast: "Đã khoá lộ trình học"')

write_tc(1, 21,
    'Toggle Mở khoá từ Khoá → Nháp',
    'P0',
    'Lộ trình status = Khoá',
    '1. Click toggle khoá\n2. Xác nhận',
    '',
    'Status chuyển thành Nháp (#64748B)\nToast: "Đã mở khoá lộ trình học"')

write_tc(1, 22,
    'Toggle Khoá cho lộ trình Đang dùng → bị chặn',
    'P0',
    'Lộ trình status = Đang dùng',
    '1. Click toggle khoá',
    '',
    'Báo lỗi: "Lộ trình đang dùng, không thể khoá"\nStatus không đổi')

write_tc(1, 23,
    'Hiển thị số khoá học và tổng bài học',
    'P0',
    'Lộ trình có 3 khoá, tổng 15 bài',
    '1. Quan sát cột "Số khoá học" và "Tổng bài học"',
    '',
    'Số khoá học = 3\nTổng bài học = 15 (tổng lessons của tất cả subjects)')


# =============================================
# II. TẠO LỘ TRÌNH HỌC (/training/learning-path/add)
# =============================================
write_section_row('II. TẠO LỘ TRÌNH HỌC (/training/learning-path/add)')

write_tc(2, 1,
    'Form hiển thị 4 tab đầy đủ',
    'P0',
    'User có quyền "Quản lý lộ trình học"',
    '1. Click nút "Thêm mới" từ danh sách',
    '',
    'Form có 4 tab: Thông tin lộ trình, Cấu hình kết quả, Cấu hình người học, Chứng chỉ\nTab "Thông tin lộ trình" active mặc định')

write_tc(2, 2,
    'Tab Thông tin - Mã lộ trình tự sinh',
    'P0',
    'Vào trang tạo mới',
    '1. Quan sát field "Mã lộ trình"',
    '',
    'Tự sinh format LP-YYYY-NNNNN (ví dụ: LP-2026-00001)\nField read-only, không cho sửa')

write_tc(2, 3,
    'Tab Thông tin - Validate bắt buộc Tên lộ trình',
    'P0',
    'Form trống',
    '1. Để trống Tên lộ trình\n2. Click Lưu',
    '',
    'Báo lỗi: field name bắt buộc\nKhông gọi API')

write_tc(2, 4,
    'Tab Thông tin - Validate max length Tên (255 ký tự)',
    'P1',
    '',
    '1. Nhập tên > 255 ký tự\n2. Lưu',
    'Chuỗi 260 ký tự',
    'Báo lỗi validate max:255')

write_tc(2, 5,
    'Tab Thông tin - Chọn Loại đào tạo',
    'P0',
    'Có nhiều training_types trong hệ thống',
    '1. Dropdown Loại đào tạo\n2. Chọn 1 loại',
    '',
    'Dropdown load danh sách training_types\nGiá trị được gán vào form')

write_tc(2, 6,
    'Tab Thông tin - Status mặc định là Nháp',
    'P0',
    'Vào trang tạo mới',
    '1. Quan sát field Status',
    '',
    'Mặc định = Nháp (1)\nCó thể đổi sang Đang dùng khi tạo')

write_tc(2, 7,
    'Tab Thông tin - Switch "Học tuần tự" (linear_required)',
    'P1',
    '',
    '1. Bật switch "Yêu cầu học tuần tự"',
    '',
    'linear_required = true\nGợi ý: người học phải hoàn thành khoá trước mới mở khoá sau')

write_tc(2, 8,
    'Tab Thông tin - Switch "Công khai" (is_public)',
    'P1',
    '',
    '1. Bật switch "Công khai"',
    '',
    'is_public = true')

write_tc(2, 9,
    'Tab Thông tin - Nhập Mô tả, Mục tiêu, Ghi chú quản trị',
    'P1',
    '',
    '1. Nhập text vào 3 field: Mô tả, Mục tiêu, Ghi chú quản trị',
    'admin_note max 500 ký tự',
    'Dữ liệu được lưu đúng\nadmin_note > 500 ký tự → báo lỗi')

write_tc(2, 10,
    'Tab Thông tin - Thêm khoá học vào lộ trình',
    'P0',
    'Có subjects trong hệ thống',
    '1. Click "Thêm khoá học"\n2. Modal hiện danh sách subjects\n3. Chọn 2 khoá\n4. Xác nhận',
    '',
    '2 khoá hiện trong danh sách với sort_order tự gán\nHiển thị tên khoá + số bài học')

write_tc(2, 11,
    'Tab Thông tin - Kéo thả sắp xếp khoá học',
    'P0',
    'Đã thêm 3+ khoá',
    '1. Kéo khoá thứ 3 lên vị trí 1',
    '',
    'Thứ tự cập nhật ngay trên UI\nsort_order được gán lại tương ứng')

write_tc(2, 12,
    'Tab Thông tin - Cấu hình khoá: bắt buộc/tuỳ chọn + ghi chú',
    'P1',
    'Đã thêm khoá',
    '1. Click icon cấu hình trên 1 khoá\n2. Toggle "Bắt buộc"\n3. Nhập ghi chú\n4. OK',
    'is_required = true, note max 500',
    'Khoá được đánh dấu bắt buộc\nGhi chú lưu đúng')

write_tc(2, 13,
    'Tab Thông tin - Xoá khoá khỏi lộ trình',
    'P1',
    'Đã thêm khoá',
    '1. Click icon xoá trên 1 khoá\n2. Xác nhận',
    '',
    'Khoá biến mất khỏi danh sách\nsort_order các khoá còn lại reindex')

write_tc(2, 14,
    'Tab Kết quả - Chọn rule "Tất cả khoá bắt buộc ĐẠT"',
    'P0',
    '',
    '1. Chuyển tab "Cấu hình kết quả"\n2. Chọn "Tất cả khoá bắt buộc ĐẠT"',
    'result_rule = REQUIRED_ONLY',
    'Chọn thành công\nKhông hiện field % tối thiểu')

write_tc(2, 15,
    'Tab Kết quả - Chọn rule "Tất cả khoá trong lộ trình ĐẠT"',
    'P0',
    '',
    '1. Chọn "Tất cả khoá trong lộ trình ĐẠT"',
    'result_rule = ALL_COURSES',
    'Chọn thành công\nKhông hiện field % tối thiểu')

write_tc(2, 16,
    'Tab Kết quả - Chọn rule "Đạt tối thiểu X%" + nhập %',
    'P0',
    '',
    '1. Chọn "Đạt tối thiểu X% số khoá"\n2. Nhập 70',
    'result_rule = MIN_PERCENT, result_min_pass_percent = 70',
    'Field % tối thiểu hiện ra\nGiá trị 70 được set')

write_tc(2, 17,
    'Tab Kết quả - Validate % phải 1-100 khi rule MIN_PERCENT',
    'P0',
    'result_rule = MIN_PERCENT',
    '1. Nhập 0 → lưu\n2. Nhập 101 → lưu\n3. Để trống → lưu',
    '',
    'Cả 3 trường hợp báo lỗi validate\nChỉ chấp nhận 1-100')

write_tc(2, 18,
    'Tab Kết quả - Hiển thị thống kê khoá học',
    'P1',
    'Đã thêm 5 khoá (3 bắt buộc, 2 tuỳ chọn)',
    '1. Xem panel thống kê',
    '',
    'Hiển thị: Tổng khoá = 5, Bắt buộc = 3, Tuỳ chọn = 2')

write_tc(2, 19,
    'Tab Người học - Thêm assignee theo Phòng ban (mandatory)',
    'P0',
    'Có phòng ban trong hệ thống',
    '1. Chuyển tab "Cấu hình người học"\n2. Cột "Bắt buộc" → pill "Phòng ban"\n3. Chọn 2 phòng ban',
    'assignee_type=department, assignment_mode=mandatory',
    '2 phòng ban hiện trong cột Bắt buộc')

write_tc(2, 20,
    'Tab Người học - Thêm assignee theo Vị trí (recommended)',
    'P0',
    'Có positions',
    '1. Cột "Khuyến nghị" → pill "Vị trí"\n2. Chọn 1 vị trí',
    'assignee_type=position, assignment_mode=recommended',
    '1 vị trí hiện trong cột Khuyến nghị')

write_tc(2, 21,
    'Tab Người học - Thêm assignee theo Năng lực',
    'P1',
    'Có capabilities',
    '1. Pill "Năng lực"\n2. Chọn 1 capability',
    'assignee_type=capability',
    'Capability hiện trong danh sách assignees')

write_tc(2, 22,
    'Tab Chứng chỉ - Bật/tắt chứng chỉ',
    'P1',
    '',
    '1. Chuyển tab "Chứng chỉ"\n2. Bật switch "Cấp chứng chỉ"',
    'certificate_enabled = true',
    'Panel cấu hình chứng chỉ hiện ra\nKhi tắt: panel ẩn đi')

write_tc(2, 23,
    'Tab Chứng chỉ - Upload template',
    'P1',
    'certificate_enabled = true',
    '1. Click "Upload template"\n2. Chọn file ảnh (1600x900)',
    '',
    'File được upload, preview hiện trên canvas\ncertificate_template_url được set')

write_tc(2, 24,
    'Tab Chứng chỉ - Cấu hình 4 field vị trí',
    'P2',
    'Đã upload template',
    '1. Cấu hình: Tên khoá, Họ tên, Ngày cấp, Người ký\n2. Nhập x, y, size, weight cho từng field',
    '',
    'certificate_fields lưu đúng 4 field config\nPreview hiện text đúng vị trí')

write_tc(2, 25,
    'Tab Chứng chỉ - Render preview + Download PDF',
    'P2',
    'Đã cấu hình đầy đủ',
    '1. Click "Render"\n2. Click "Download PDF"',
    '',
    'Render: canvas hiện template + text đúng vị trí\nDownload: file .pdf được tải về')

write_tc(2, 26,
    'Lưu lộ trình thành công (status Nháp)',
    'P0',
    'Đã nhập tên + chọn result_rule + thêm ≥1 khoá',
    '1. Click Lưu',
    '',
    'Tạo thành công, toast "Tạo lộ trình học thành công"\nĐiều hướng về danh sách\nCode format LP-YYYY-NNNNN')

write_tc(2, 27,
    'Lưu lộ trình với status Đang dùng',
    'P0',
    'Đã nhập đầy đủ + chọn status = Đang dùng',
    '1. Chọn Status = "Đang dùng"\n2. Click Lưu',
    'status = 2',
    'Tạo thành công với status = 2 ngay lập tức')

write_tc(2, 28,
    'Validate: result_rule bắt buộc',
    'P0',
    '',
    '1. Không chọn result_rule\n2. Lưu',
    '',
    'Báo lỗi validate: result_rule là bắt buộc')


# =============================================
# III. SỬA LỘ TRÌNH HỌC (/training/learning-path/{id}/edit)
# =============================================
write_section_row('III. SỬA LỘ TRÌNH HỌC (/training/learning-path/{id}/edit)')

write_tc(3, 1,
    'Load dữ liệu đầy đủ khi vào trang sửa',
    'P0',
    'Lộ trình Nháp có 3 khoá + 2 assignees + certificate',
    '1. Vào /training/learning-path/{id}/edit',
    '',
    'Tất cả 4 tab hiển thị dữ liệu đã lưu:\n- Tab Info: tên, loại ĐT, code, subjects\n- Tab Kết quả: rule + %\n- Tab Người học: assignees\n- Tab Chứng chỉ: template + fields')

write_tc(3, 2,
    'Sửa tên + lưu thành công',
    'P0',
    'Lộ trình Nháp',
    '1. Đổi tên\n2. Lưu',
    '',
    'Toast: "Cập nhật lộ trình học thành công"\nTên mới hiện ở danh sách')

write_tc(3, 3,
    'Thêm/bớt khoá khi sửa',
    'P0',
    'Lộ trình có 3 khoá',
    '1. Xoá 1 khoá\n2. Thêm 2 khoá mới\n3. Lưu',
    '',
    'Sau lưu: lộ trình có 4 khoá\nsort_order cập nhật đúng')

write_tc(3, 4,
    'Đổi result_rule từ REQUIRED_ONLY sang MIN_PERCENT',
    'P1',
    'Lộ trình có result_rule = REQUIRED_ONLY',
    '1. Đổi sang MIN_PERCENT\n2. Nhập 60%\n3. Lưu',
    '',
    'result_rule = MIN_PERCENT\nresult_min_pass_percent = 60')

write_tc(3, 5,
    'Thay đổi assignees khi sửa',
    'P1',
    'Lộ trình có 2 assignees',
    '1. Xoá 1 assignee\n2. Thêm 1 assignee mới\n3. Lưu',
    '',
    'Assignees được sync đúng (xoá cũ, thêm mới)')

write_tc(3, 6,
    'Không cho sửa lộ trình Đang dùng',
    'P0',
    'Lộ trình status = 2',
    '1. Thử truy cập /training/learning-path/{id}/edit',
    '',
    'Không hiện nút Sửa trên detail\nNếu vào URL trực tiếp: form ở mode show (read-only) hoặc redirect')

write_tc(3, 7,
    'Không cho sửa lộ trình Khoá',
    'P0',
    'Lộ trình status = 3',
    '1. Thử truy cập edit',
    '',
    'Tương tự: form read-only hoặc redirect')


# =============================================
# IV. XEM CHI TIẾT (/training/learning-path/{id})
# =============================================
write_section_row('IV. XEM CHI TIẾT (/training/learning-path/{id})')

write_tc(4, 1,
    'Hiển thị đầy đủ thông tin 4 tab (read-only)',
    'P0',
    'Lộ trình có đầy đủ data',
    '1. Vào /training/learning-path/{id}',
    '',
    'Tất cả tab hiện đúng data\nKhông có field editable\nTab subjects hiện danh sách khoá kèm lessons + exams')

write_tc(4, 2,
    'Nút Sửa chỉ hiện cho Nháp',
    'P0',
    'Lộ trình Nháp',
    '1. Quan sát nút hành động',
    '',
    'Nút "Sửa" hiện ở trang chi tiết\nLộ trình Đang dùng/Khoá: nút Sửa ẩn')

write_tc(4, 3,
    'Hiển thị danh sách khoá + bài học lồng nhau',
    'P0',
    'Lộ trình có 3 khoá, mỗi khoá có lessons + exams',
    '1. Xem tab Thông tin',
    '',
    'Mỗi khoá hiện: tên, bắt buộc/tuỳ chọn, ghi chú\nMỗi khoá expand ra danh sách lessons + exams')

write_tc(4, 4,
    'Hiển thị thông tin người tạo + cập nhật',
    'P1',
    '',
    '1. Xem phần footer/meta',
    '',
    'Hiện: Người tạo + thời gian, Người cập nhật + thời gian')

write_tc(4, 5,
    'Lộ trình không tồn tại → 404',
    'P1',
    'id = 99999 không tồn tại',
    '1. Mở /training/learning-path/99999',
    '',
    'Hiển thị 404 hoặc báo "Lộ trình không tồn tại"')


# =============================================
# V. XOÁ LỘ TRÌNH HỌC
# =============================================
write_section_row('V. XOÁ LỘ TRÌNH HỌC')

write_tc(5, 1,
    'Xoá lộ trình Nháp thành công',
    'P0',
    'Lộ trình status = 1',
    '1. Click icon Xoá\n2. Modal xác nhận hiện\n3. Confirm',
    '',
    'Lộ trình bị xoá\nRelated subjects + assignees cũng bị xoá\nToast: "Xoá lộ trình học thành công"')

write_tc(5, 2,
    'Xoá lộ trình Đang dùng → bị chặn',
    'P0',
    'Lộ trình status = 2',
    '1. Gọi API DELETE /training/learning-paths/{id}',
    '',
    'API trả 400: "Chỉ xoá được lộ trình ở trạng thái Nháp"\nFE không hiện nút Xoá')

write_tc(5, 3,
    'Xoá lộ trình Khoá → bị chặn',
    'P0',
    'Lộ trình status = 3',
    '1. Gọi API DELETE /training/learning-paths/{id}',
    '',
    'API trả 400: "Chỉ xoá được lộ trình ở trạng thái Nháp"')

write_tc(5, 4,
    'Huỷ modal xác nhận xoá',
    'P1',
    'Click Xoá, modal hiện',
    '1. Click "Huỷ" trên modal',
    '',
    'Modal đóng, lộ trình vẫn còn')


# =============================================
# VI. API - VALIDATION & SECURITY
# =============================================
write_section_row('VI. API - VALIDATION & SECURITY')

write_tc(6, 1,
    'API store - thiếu name → 422',
    'P0',
    '',
    '1. POST /training/learning-paths body thiếu name',
    '',
    'Response 422: name is required')

write_tc(6, 2,
    'API store - status invalid → 422',
    'P0',
    '',
    '1. POST body status = 5',
    'status = 5',
    'Response 422: status must be in 1,2,3')

write_tc(6, 3,
    'API store - result_rule invalid → 422',
    'P0',
    '',
    '1. POST body result_rule = "INVALID"',
    '',
    'Response 422: result_rule invalid')

write_tc(6, 4,
    'API store - MIN_PERCENT thiếu percent → 422',
    'P0',
    '',
    '1. POST body result_rule = MIN_PERCENT, không có result_min_pass_percent',
    '',
    'Response 422: required_if validation fail')

write_tc(6, 5,
    'API store - subjects.*.subject_id bắt buộc',
    'P0',
    '',
    '1. POST body subjects = [{"sort_order": 1}]',
    '',
    'Response 422: subjects.0.subject_id required')

write_tc(6, 6,
    'API store - assignees.*.assignee_type phải in:department,position,capability',
    'P0',
    '',
    '1. POST body assignees = [{"assignee_type": "invalid", "assignee_id": 1, "assignment_mode": "mandatory"}]',
    '',
    'Response 422: assignee_type invalid')

write_tc(6, 7,
    'API store - assignees.*.assignment_mode phải in:mandatory,recommended',
    'P0',
    '',
    '1. POST body assignees = [{"assignee_type": "department", "assignee_id": 1, "assignment_mode": "optional"}]',
    '',
    'Response 422: assignment_mode invalid')

write_tc(6, 8,
    'API store - training_type_id không tồn tại → 422',
    'P1',
    '',
    '1. POST body training_type_id = 99999',
    '',
    'Response 422: training_type_id not exists')

write_tc(6, 9,
    'API getAll - chỉ trả lộ trình Active',
    'P0',
    'Có lộ trình Nháp + Active + Khoá',
    '1. GET /training/learning-paths/getAll',
    '',
    'Response chỉ chứa lộ trình status = 2\nDùng cho dropdown chọn lộ trình ở nơi khác')

write_tc(6, 10,
    'API getNextCode trả code format đúng',
    'P0',
    '',
    '1. GET /training/learning-paths/get-next-code',
    '',
    'Response: {"code": "LP-2026-XXXXX"}\nCode match regex ^LP-\\d{4}-\\d{5}$')

write_tc(6, 11,
    'API toggleLock - lộ trình Active → 400',
    'P0',
    'Lộ trình status = 2',
    '1. POST /training/learning-paths/{id}/toggle-lock',
    '',
    'Response 400: "Lộ trình đang dùng, không thể khoá"')

write_tc(6, 12,
    'API delete - lộ trình không phải Nháp → 400',
    'P0',
    'Lộ trình status = 2',
    '1. DELETE /training/learning-paths/{id}',
    '',
    'Response 400: "Chỉ xoá được lộ trình ở trạng thái Nháp"')

write_tc(6, 13,
    'API - User không có permission → 403',
    'P0',
    'User không có quyền "Quản lý lộ trình học"',
    '1. Gọi bất kỳ API learning-paths',
    '',
    'Response 403 Forbidden')


# =============================================
# VII. BUSINESS RULES & LUỒNG E2E
# =============================================
write_section_row('VII. BUSINESS RULES & LUỒNG E2E')

write_tc(7, 1,
    'Happy path: Tạo → Kích hoạt → Khoá → Mở khoá',
    'P0',
    'User có quyền',
    '1. Tạo lộ trình Nháp với 3 khoá + 2 assignees\n2. Sửa: đổi status = Đang dùng, Lưu\n3. Toggle Khoá (thất bại vì Đang dùng)\n4. Sửa status về Nháp\n5. Toggle Khoá (thành công)\n6. Toggle Mở khoá',
    '',
    'Luồng trạng thái: 1 → 2 → (chặn khoá) → 1 → 3 → 1\nMỗi bước toast message đúng')

write_tc(7, 2,
    'E2E: Tạo đầy đủ với certificate + assignees',
    'P0',
    '',
    '1. Tạo lộ trình: tên, loại ĐT, 5 khoá (3 bắt buộc)\n2. Tab Kết quả: MIN_PERCENT 80%\n3. Tab Người học: 2 PB mandatory + 1 position recommended\n4. Tab Chứng chỉ: upload template + config fields\n5. Lưu status = Đang dùng',
    '',
    'Tạo thành công\nXem chi tiết: tất cả data hiện đúng 4 tab')

write_tc(7, 3,
    'Mã code auto-increment đúng',
    'P0',
    '',
    '1. Tạo lộ trình A → code LP-2026-00001\n2. Tạo lộ trình B → code LP-2026-00002\n3. Xoá lộ trình A\n4. Tạo lộ trình C',
    '',
    'Code C = LP-2026-00003 (không tái dùng code cũ, dựa trên max id)')

write_tc(7, 4,
    'Sync subjects: xoá hết → thêm mới khi update',
    'P0',
    'Lộ trình có 3 khoá',
    '1. Sửa: xoá hết, thêm 2 khoá khác\n2. Lưu\n3. Xem chi tiết',
    '',
    'Chỉ hiện 2 khoá mới\nKhoá cũ bị xoá khỏi bảng learning_path_subjects')

write_tc(7, 5,
    'Sync assignees: thay đổi khi update',
    'P0',
    'Lộ trình có 2 assignees',
    '1. Sửa: bỏ 1, thêm 3 mới\n2. Lưu\n3. Xem chi tiết',
    '',
    'Tab Người học hiện 4 assignees (1 cũ + 3 mới)\nBảng learning_path_assignees sync đúng')

write_tc(7, 6,
    'Xoá lộ trình cascade subjects + assignees',
    'P0',
    'Lộ trình Nháp có 3 khoá + 2 assignees',
    '1. Xoá lộ trình\n2. Kiểm tra DB',
    '',
    'Record learning_paths bị xoá\nRecords learning_path_subjects bị xoá\nRecords learning_path_assignees bị xoá')


# =============================================
# VIII. EDGE CASES
# =============================================
write_section_row('VIII. EDGE CASES')

write_tc(8, 1,
    'Tạo lộ trình không có khoá nào (subjects rỗng)',
    'P1',
    '',
    '1. Nhập tên + result_rule\n2. Không thêm khoá\n3. Lưu',
    'subjects = []',
    'API cho phép (subjects nullable)\nNhưng lộ trình rỗng khoá → cần warning UI hoặc cho phép tuỳ nghiệp vụ')

write_tc(8, 2,
    'Tạo lộ trình không có assignees',
    'P1',
    '',
    '1. Không thêm assignees\n2. Lưu',
    'assignees = []',
    'API cho phép (assignees nullable)\nLộ trình tạo thành công, chưa gán người học')

write_tc(8, 3,
    'Thêm cùng 1 subject 2 lần',
    'P1',
    '',
    '1. Thêm subject A\n2. Thêm subject A lần nữa',
    '',
    'FE chặn: không cho thêm khoá đã có trong list\nHoặc API xử lý deduplicate')

write_tc(8, 4,
    'admin_note vượt 500 ký tự',
    'P1',
    '',
    '1. Nhập admin_note 501+ ký tự\n2. Lưu',
    'Chuỗi 510 ký tự',
    'API trả 422: admin_note max 500')

write_tc(8, 5,
    'certificate_template_url quá dài (>500)',
    'P2',
    '',
    '1. Upload file tạo URL rất dài',
    '',
    'API trả 422 nếu URL > 500 ký tự')

write_tc(8, 6,
    'Sort + Filter + Pagination kết hợp',
    'P1',
    '>40 lộ trình',
    '1. Filter status = 1\n2. Sort by name asc\n3. Chuyển trang 2',
    '',
    'Trang 2 vẫn giữ filter + sort\nDữ liệu nhất quán')

write_tc(8, 7,
    'subject_id không tồn tại trong hệ thống',
    'P1',
    '',
    '1. Gọi API subjects = [{"subject_id": 99999, "sort_order": 1}]',
    '',
    'API xử lý lỗi graceful (422 hoặc bỏ qua tuỳ logic)\nKhông crash server 500')

write_tc(8, 8,
    'Concurrent update cùng 1 lộ trình',
    'P2',
    '2 user mở cùng lộ trình Nháp',
    '1. User A sửa tên + lưu\n2. User B sửa khoá + lưu (gần cùng lúc)',
    '',
    'Cả 2 save thành công\nDữ liệu cuối = save cuối cùng (last-write-wins)\nKhông crash')

write_tc(8, 9,
    'Toggle lock cho lộ trình không tồn tại',
    'P2',
    '',
    '1. POST /training/learning-paths/99999/toggle-lock',
    '',
    'API trả 404')

write_tc(8, 10,
    'result_min_pass_percent khi rule không phải MIN_PERCENT',
    'P1',
    '',
    '1. POST body result_rule = ALL_COURSES, result_min_pass_percent = 80',
    '',
    'API bỏ qua giá trị % (nullable)\nKhông báo lỗi, nhưng không lưu % vô nghĩa')


# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# === SAVE ===
output_path = r'D:\laragon\www\hrm\docs\srs\learning-path-testcases.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Total testcases: {current_row - 7 - 8}')  # trừ section rows
