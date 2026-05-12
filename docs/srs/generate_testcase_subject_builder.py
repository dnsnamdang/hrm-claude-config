import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Subject Builder"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 25, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Quản lý Khoá học — Subject Builder (Module Đào tạo)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i + 1, column=10, value=label).font = summary_font
    ws.cell(row=i + 1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER FUNCTIONS ===
current_row = 7

MODULE = 'Đào tạo'
GROUP = 'Khoá học (Subject Builder)'
PREFIX = 'TRSB'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:02d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. DANH SÁCH KHOÁ HỌC
# =============================================
write_section_row('I. DANH SÁCH KHOÁ HỌC (/training/subjects)')

write_tc(1, 1,
    'Hiển thị danh sách khoá học',
    'P0',
    'Có ít nhất 3 khoá học trong DB (status: 1, 2, 3)',
    '1. Đăng nhập\n2. Vào menu Đào tạo > Khoá học\n3. Quan sát bảng danh sách',
    '',
    '- Bảng hiển thị đúng dữ liệu (mã, tên, loại, trạng thái)\n- Badge màu đúng: Hoạt động=xanh, Khoá=đỏ, Nháp=xám\n- Sidebar menu hiển thị "Khoá học" (không phải "Môn học")')

write_tc(1, 2,
    'Filter theo trạng thái',
    'P1',
    'Có khoá học ở cả 3 trạng thái (1, 2, 3)',
    '1. Chọn filter Trạng thái = "Nháp"\n2. Submit tìm kiếm',
    'Trạng thái: Nháp (status=3)',
    'Chỉ hiển thị khoá học status=3\nCác status 1, 2 bị ẩn khỏi danh sách')

write_tc(1, 3,
    'Filter theo tên/mã',
    'P1',
    'Có ít nhất 3 khoá học khác tên',
    '1. Nhập từ khoá vào ô tìm kiếm\n2. Submit',
    'Keyword: "An toàn"',
    'Chỉ hiển thị khoá học khớp tên hoặc mã với từ khoá đã nhập')

write_tc(1, 4,
    'Nút sửa — hiển thị đúng theo trạng thái',
    'P0',
    '3 khoá học có status khác nhau: 1, 2, 3',
    '1. Quan sát cột hành động với 3 khoá học có status 1, 2, 3',
    '',
    '- Status=1 (Hoạt động): nút ✏ sửa hiện\n- Status=3 (Nháp): nút ✏ sửa hiện\n- Status=2 (Khoá): nút ✏ sửa ẩn hoặc disabled')


# =============================================
# II. TẠO MỚI — LƯU TẠM (DRAFT)
# =============================================
write_section_row('II. TẠO MỚI KHOÁ HỌC — LƯU TẠM (DRAFT, status=3)')

write_tc(2, 1,
    'Lưu tạm chỉ với tên',
    'P0',
    'Đang ở màn tạo mới khoá học',
    '1. Click "+ Tạo mới"\n2. Chỉ nhập "Tên khoá học test"\n3. Click "Lưu tạm"',
    'Tên: "Khoá học test TRSB"',
    '- API POST /store-builder gọi với status=3\n- Khoá học được tạo với status=Nháp\n- Mã được auto-gen dạng SUB-YYYY-NNNNN\n- Toast thành công hiển thị\n- Redirect về danh sách')

write_tc(2, 2,
    'Lưu tạm khi tên rỗng',
    'P0',
    'Đang ở màn tạo mới khoá học',
    '1. Vào tạo mới\n2. Để trống field tên\n3. Click "Lưu tạm"',
    'Tên: (để trống)',
    '- FE validate: hiển thị lỗi inline dưới field tên\n- Không gọi API\n- Tab tự chuyển về Tab 1 (Thông tin) nếu đang ở tab khác')

write_tc(2, 3,
    'Lưu tạm với mã trùng',
    'P1',
    'Đã có khoá học với mã "SUB-2026-00001"',
    '1. Nhập tên và nhập mã đã tồn tại\n2. Click "Lưu tạm"',
    'Mã: "SUB-2026-00001" (đã tồn tại)',
    '- API trả về 422\n- Lỗi inline "Mã khoá học đã tồn tại" hiển thị dưới field mã\n- Tab tự chuyển về Tab 1 chứa field lỗi')


# =============================================
# III. TẠO MỚI — LƯU ĐẦY ĐỦ (status=1)
# =============================================
write_section_row('III. TẠO MỚI KHOÁ HỌC — LƯU ĐẦY ĐỦ (status=1 Hoạt động)')

write_tc(3, 1,
    'Lưu đầy đủ — evaluation_mode = completion',
    'P0',
    'Có ít nhất 1 bài học trong ngân hàng',
    '1. Tab 1: nhập tên, chọn loại đào tạo, thêm ≥1 bài học\n2. Tab 2: chọn "Hoàn thành bài học", rule = "Tất cả bắt buộc"\n3. Click "Lưu" (status=1)',
    '',
    '- API gọi thành công 200\n- Khoá học được tạo status=1\n- evaluation_config.rule = "all_required" lưu đúng\n- evaluation_mode = "completion"\n- Không có row nào trong subject_exams')

write_tc(3, 2,
    'Lưu đầy đủ — evaluation_mode = exam',
    'P0',
    'Có đề thi trong hệ thống',
    '1. Tab 1: nhập tên, thêm ≥1 bài học\n2. Tab 2: chọn "Thi đề", chọn 1 đề thi\n3. Nhập: time_limit=60, pass_score=70, attempt_limit=3, score_rule=highest\n4. Click "Lưu"',
    'time_limit=60, pass_score=70, attempt_limit=3',
    '- evaluation_mode = "exam"\n- subject_exams có 1 row với đúng exam_id, time_limit_min=60, pass_score_percent=70\n- exam_attempt_limit=3, exam_score_rule="highest" lưu trên bảng subjects')

write_tc(3, 3,
    'Validation — exam mode thiếu đề thi',
    'P0',
    'Đang ở màn tạo mới',
    '1. Tab 2: chọn "Thi đề"\n2. Không chọn đề nào\n3. Click "Lưu"',
    '',
    '- Lỗi validation: "Phải cấu hình ít nhất một bài thi"\n- Tab tự chuyển về Tab 2')

write_tc(3, 4,
    'Validation — đề thi có tự luận + không chọn người chấm',
    'P0',
    'Có đề thi chứa câu tự luận (essay_count > 0)',
    '1. Tab 2: chọn đề thi có câu tự luận\n2. Không chọn người chấm\n3. Click "Lưu"',
    'Đề thi có essay_count > 0',
    '- FE validate: lỗi "Phải chọn ít nhất 1 người chấm" dưới grader_ids của đề đó\n- Tab chuyển về Tab 2')

write_tc(3, 5,
    'Lưu với ảnh banner',
    'P1',
    'Có file ảnh JPEG/PNG để upload',
    '1. Tab 1: click vùng BannerUploader, chọn file ảnh\n2. Chờ preview xuất hiện ngay (local) và spinner upload\n3. Sau upload: ảnh hiển thị URL server\n4. Click "Lưu"',
    'File: banner.jpg (~500KB)',
    '- banner_url lưu đúng URL S3\n- Khi load lại màn sửa: ảnh vẫn hiển thị')

write_tc(3, 6,
    'Lưu với chứng chỉ đầy đủ',
    'P1',
    'Có file ảnh nền chứng chỉ',
    '1. Tab 4: bật toggle chứng chỉ\n2. Upload ảnh nền\n3. Điền cả 4 trường (course_name, employee_name, issued_date, signer) với X, Y, font_size\n4. Click "Lưu"',
    '',
    '- certificate_enabled=1 lưu đúng\n- certificate_template_url có giá trị URL\n- 4 row trong subject_certificate_fields với đúng field_key và tọa độ')

write_tc(3, 7,
    'Validation — chứng chỉ bật nhưng thiếu ảnh nền',
    'P1',
    'Đang ở màn tạo mới',
    '1. Tab 4: bật toggle chứng chỉ\n2. KHÔNG upload ảnh nền\n3. Click "Lưu"',
    '',
    '- Lỗi "Chưa chọn ảnh mẫu chứng chỉ"\n- Tab tự chuyển về Tab 4')


# =============================================
# IV. CẤU HÌNH CHƯƠNG & BÀI HỌC (TAB 1)
# =============================================
write_section_row('IV. CẤU HÌNH CHƯƠNG & BÀI HỌC (Tab 1 — use_chapters)')

write_tc(4, 1,
    'Thêm chương và bài học',
    'P0',
    'Có bài học trong ngân hàng bài học',
    '1. Toggle "Có chương"\n2. Click "Thêm chương", nhập tên chương\n3. Trong chương, click "Thêm bài học", chọn bài từ ngân hàng\n4. Click "Lưu"',
    'Tên chương: "Chương 1: Giới thiệu"',
    '- Chương tạo trong subject_chapters với đúng sort_order\n- Bài học trong subject_lessons với đúng chapter_id và sort_order')

write_tc(4, 2,
    'Sửa tên chương — không tạo chương mới (bug fix Phase 16)',
    'P0',
    'Đang có ≥2 chương chưa lưu DB (chỉ tồn tại local state)',
    '1. Có 2 chương: "Chương 1" và "Chương 2"\n2. Click icon sửa chương đầu tiên\n3. Nhập tên mới "Chương 1 — Updated"\n4. Lưu modal chương',
    '',
    '- Chương đầu tiên được CẬP NHẬT tên mới (không tạo thêm chương mới)\n- Chương thứ 2 giữ nguyên\n- Tổng vẫn là 2 chương')

write_tc(4, 3,
    'Đổi use_chapters từ "Có chương" sang "Không chương"',
    'P0',
    'Khoá học đang có use_chapters=1 với 2 chương + bài học đã lưu',
    '1. Vào sửa khoá học\n2. Toggle sang "Không có chương" (use_chapters=2)\n3. Click "Lưu"',
    '',
    '- subject_chapters bị xoá sạch\n- subject_lessons vẫn tồn tại nhưng chapter_id = null')

write_tc(4, 4,
    'Prerequisite — bật điều kiện mở khoá',
    'P1',
    'Có ≥2 bài học trong khoá',
    '1. Mở modal info bài thứ 2\n2. Bật "Bật điều kiện mở khoá"\n3. Chọn bài thứ 1 làm prerequisite\n4. Lưu modal → Lưu khoá học',
    '',
    '- prerequisite_enabled=1 lưu đúng\n- prerequisite_subject_lesson_ids chứa ID bài thứ 1')

write_tc(4, 5,
    'Xoá bài học đang là prerequisite',
    'P1',
    'Bài A là prerequisite của bài B',
    '1. Xoá bài A khỏi danh sách\n2. Lưu khoá học',
    '',
    '- prerequisite_subject_lesson_ids của bài B không còn ID bài A\n- Không bị lỗi khi lưu')

write_tc(4, 6,
    'Prerequisite persist sau khi đóng/mở modal',
    'P1',
    'Đã cấu hình prerequisite cho bài học',
    '1. Cấu hình prerequisite cho bài B (bài A là điều kiện)\n2. Đóng modal bài B\n3. Mở lại modal bài B',
    '',
    '- Danh sách prerequisite vẫn hiển thị đúng bài A (không bị reset về trống)')

write_tc(4, 7,
    'Drag-drop sắp xếp thứ tự chương',
    'P1',
    'Có ≥2 chương trong khoá học',
    '1. Kéo thả đổi vị trí chương (chương 2 lên trước chương 1)\n2. Lưu khoá học',
    '',
    '- sort_order cập nhật theo thứ tự kéo thả\n- Khi load lại: thứ tự chương vẫn đúng')


# =============================================
# V. CẤU HÌNH ĐÁNH GIÁ (TAB 2)
# =============================================
write_section_row('V. CẤU HÌNH ĐÁNH GIÁ (Tab 2 — evaluation_mode)')

write_tc(5, 1,
    'Chuyển exam → completion → xoá subject_exams',
    'P0',
    'Khoá học đã lưu với eval_mode=exam, có 1 đề thi + grader',
    '1. Vào sửa khoá học\n2. Tab 2: chuyển sang eval_mode="Hoàn thành bài học"\n3. Click "Lưu"',
    '',
    '- subject_exams bị xoá sạch\n- subject_exam_graders cũng bị xoá\n- evaluation_mode = "completion" lưu đúng')

write_tc(5, 2,
    'exam_participation_required — bật và nhập %',
    'P1',
    'Eval mode = exam',
    '1. Eval mode = exam\n2. Bật "Yêu cầu hoàn thành bài học trước khi thi"\n3. Nhập % = 80\n4. Lưu',
    'exam_min_required_percent: 80',
    '- exam_participation_required=1\n- exam_min_required_percent=80 lưu đúng')

write_tc(5, 3,
    'Validation participation bật nhưng không nhập %',
    'P1',
    'Eval mode = exam, bật participation',
    '1. Bật "Yêu cầu hoàn thành bài học trước khi thi"\n2. Không nhập % tối thiểu\n3. Click "Lưu"',
    '',
    '- Lỗi "Chưa nhập % bài học tối thiểu" hiện dưới field\n- Tab chuyển về Tab 2')

write_tc(5, 4,
    'Nhiều đề thi — mỗi đề cấu hình riêng',
    'P1',
    'Có ≥2 đề thi trong hệ thống',
    '1. Chọn 2 đề thi\n2. Đề 1: time=60, pass=70\n3. Đề 2: time=90, pass=80\n4. Lưu',
    'Đề 1: 60p/70%, Đề 2: 90p/80%',
    '- subject_exams có 2 row với cấu hình riêng biệt cho từng đề')


# =============================================
# VI. CẤU HÌNH NGƯỜI HỌC (TAB 3)
# =============================================
write_section_row('VI. CẤU HÌNH NGƯỜI HỌC (Tab 3 — onboarding & assignees)')

write_tc(6, 1,
    'Onboarding — bật và điền số ngày',
    'P1',
    'Đang ở Tab 3 của khoá học',
    '1. Bật toggle Onboarding\n2. Nhập onboarding_new_employee_days = 30\n3. Nhập onboarding_must_finish_days = 60\n4. Lưu',
    'new_employee_days=30, must_finish_days=60',
    '- onboarding_enabled=1\n- onboarding_new_employee_days=30\n- onboarding_must_finish_days=60 lưu đúng')

write_tc(6, 2,
    'Onboarding bật nhưng không nhập ngày',
    'P1',
    'Đang ở Tab 3',
    '1. Bật toggle Onboarding\n2. Không nhập số ngày\n3. Click "Lưu"',
    '',
    '- Lỗi validation: "Chưa nhập số ngày được coi là nhân viên mới"\n- Tab chuyển về Tab 3')

write_tc(6, 3,
    'Đối tượng bắt buộc — theo phòng ban',
    'P1',
    'Có ít nhất 2 phòng ban trong hệ thống',
    '1. Tab 3: chọn pill "Phòng ban" ở khối bắt buộc\n2. Chọn 2 phòng ban\n3. Lưu',
    '',
    '- mandatory_assignee_type = "department"\n- subject_assignees có 2 row với assignee_type="department", is_mandatory=1')

write_tc(6, 4,
    'Pill selector giữ data khi chuyển type',
    'P1',
    'Đang ở Tab 3',
    '1. Chọn pill "Phòng ban", chọn 2 phòng ban\n2. Switch sang pill "Chức vụ", chọn 1 chức vụ\n3. Switch lại "Phòng ban"',
    '',
    '- Dữ liệu 2 phòng ban đã chọn vẫn còn (không bị reset khi switch qua lại)')

write_tc(6, 5,
    'Load lại đúng pill đã chọn khi vào sửa',
    'P1',
    'Khoá học đã lưu với mandatory_assignee_type="position" (chức vụ)',
    '1. Vào sửa khoá học\n2. Chuyển sang Tab 3',
    '',
    '- Pill "Chức vụ" đã được chọn sẵn (active)\n- Danh sách chức vụ đã chọn hiển thị đúng')


# =============================================
# VII. CẤU HÌNH CHỨNG CHỈ (TAB 4)
# =============================================
write_section_row('VII. CẤU HÌNH CHỨNG CHỈ (Tab 4 — certificate)')

write_tc(7, 1,
    'Canvas preview real-time khi nhập text',
    'P1',
    'Đã upload ảnh nền chứng chỉ',
    '1. Nhập text cho course_name\n2. Nhập text cho signer\n3. Thay đổi X, Y, font_size của course_name',
    '',
    '- Canvas preview cập nhật real-time theo mỗi thay đổi\n- Text overlay hiển thị đúng vị trí và kích thước font')

write_tc(7, 2,
    'Download PDF preview chứng chỉ',
    'P2',
    'Cấu hình cert đầy đủ (ảnh nền + 4 trường)',
    '1. Click nút "Download PDF"',
    '',
    '- File PDF được tải về\n- PDF chứa ảnh nền + text overlay (không phải ảnh trắng/lỗi CORS)')

write_tc(7, 3,
    'Tắt chứng chỉ → xoá certificate_fields',
    'P1',
    'Khoá học đang có certificate_enabled=1 và 4 field rows',
    '1. Vào sửa khoá học\n2. Tab 4: tắt toggle chứng chỉ\n3. Click "Lưu"',
    '',
    '- certificate_enabled=0\n- subject_certificate_fields bị xoá sạch (0 row)')


# =============================================
# VIII. UPLOAD ẢNH BANNER (BannerUploader)
# =============================================
write_section_row('VIII. UPLOAD ẢNH BANNER (BannerUploader component)')

write_tc(8, 1,
    'Upload và preview ngay (local FileReader)',
    'P0',
    'Đang ở Tab 1 màn tạo mới',
    '1. Click vùng BannerUploader\n2. Chọn file ảnh JPEG\n3. Quan sát preview',
    'File: banner.jpg (~500KB)',
    '- Preview ảnh hiển thị NGAY (local FileReader, trước khi upload xong)\n- Spinner loading hiển thị trong khi đang upload lên S3\n- Sau upload: ảnh hiển thị URL server (không phải local blob)')

write_tc(8, 2,
    'Đổi ảnh khi đã có ảnh',
    'P1',
    'Đã có ảnh banner',
    '1. Hover vào vùng ảnh → hiện overlay\n2. Click "Đổi ảnh"\n3. Chọn ảnh mới',
    '',
    '- File picker mở ra\n- Chọn ảnh mới → ảnh cũ bị thay thế bằng ảnh mới\n- Preview cập nhật')

write_tc(8, 3,
    'Xoá ảnh',
    'P1',
    'Đang có ảnh banner',
    '1. Hover vào vùng ảnh\n2. Click nút xóa (icon đỏ)',
    '',
    '- Ảnh biến mất, quay về trạng thái rỗng (placeholder)\n- banner_url = null khi lưu khoá học')

write_tc(8, 4,
    'Disabled state — không upload được khi đang xem',
    'P2',
    'Đang ở màn xem chi tiết (isShow=true)',
    '1. Quan sát vùng BannerUploader',
    '',
    '- BannerUploader ở trạng thái disabled (opacity thấp hơn, cursor không phải pointer)\n- Không thể click để upload hoặc xoá ảnh')


# =============================================
# IX. KHOÁ / MỞ KHOÁ
# =============================================
write_section_row('IX. KHOÁ / MỞ KHOÁ KHOÁ HỌC')

write_tc(9, 1,
    'Khoá khoá học đang hoạt động',
    'P1',
    'Khoá học status=1 (Hoạt động)',
    '1. Tìm khoá học status=1 trong danh sách\n2. Click icon khoá (🔒)\n3. Confirm dialog hiện với thông báo phù hợp\n4. Click xác nhận',
    '',
    '- Status chuyển 1→2\n- Badge đổi sang màu đỏ "Khoá"\n- Nút ✏ sửa biến mất trong danh sách')

write_tc(9, 2,
    'Mở khoá',
    'P1',
    'Khoá học status=2 (Khoá)',
    '1. Tìm khoá học status=2\n2. Click icon mở khoá (🔓)\n3. Xác nhận',
    '',
    '- Status chuyển 2→1\n- Badge đổi sang màu xanh "Hoạt động"\n- Nút ✏ sửa xuất hiện lại')


# =============================================
# X. XOÁ KHOÁ HỌC
# =============================================
write_section_row('X. XOÁ KHOÁ HỌC')

write_tc(10, 1,
    'Xoá DRAFT — luôn được phép',
    'P0',
    'Khoá học status=3 (DRAFT), chưa được dùng ở đâu',
    '1. Tìm khoá học status=3\n2. Click xoá, xác nhận',
    '',
    '- is_can_delete=true từ API\n- Khoá học bị xoá thành công\n- Không còn hiển thị trong danh sách')

write_tc(10, 2,
    'Xoá khoá học có downstream reference',
    'P0',
    'Khoá học status=1 đã được dùng trong 1 course_request hoặc training_program',
    '1. Tìm khoá học có reference\n2. Quan sát nút xoá',
    '',
    '- is_can_delete=false từ API\n- Nút xoá bị disable hoặc ẩn đi\n- Không thể thực hiện xoá')

write_tc(10, 3,
    'Xoá khoá học không có reference',
    'P1',
    'Khoá học status=1 chưa được dùng ở bất kỳ đâu',
    '1. Tìm khoá học chưa được dùng\n2. Click xoá, xác nhận',
    '',
    '- is_can_delete=true\n- Khoá học bị xoá thành công')


# =============================================
# XI. CHỈNH SỬA — LOAD LẠI ĐÚNG DỮ LIỆU
# =============================================
write_section_row('XI. CHỈNH SỬA — LOAD LẠI ĐẦY ĐỦ DỮ LIỆU KHI VÀO SỬA')

write_tc(11, 1,
    'Load lại subject cũ đã migrate (backfill)',
    'P0',
    'Có khoá học cũ tạo từ trước rebuild (data backfill từ evaluation_config JSON cũ)',
    '1. Tìm khoá học cũ (trước rebuild)\n2. Click vào sửa',
    '',
    '- Tab 2 hiển thị đúng evaluation_mode + config từ JSON cũ\n- Tab 3 hiển thị đúng assignees (working_position/capability backfill)\n- Không lỗi JS, không trang trắng')

write_tc(11, 2,
    'Load đầy đủ 4 tab khi sửa khoá học mới tạo',
    'P0',
    'Khoá học mới tạo đầy đủ 4 tab (chapters, exams, assignees, cert)',
    '1. Vào sửa khoá học đã tạo đầy đủ',
    '',
    '- Tab 1: chapters + subject_lessons đúng thứ tự sort_order\n- Tab 2: eval config đúng (mode + đề thi + graders)\n- Tab 3: onboarding + assignees đúng\n- Tab 4: cert template URL + 4 fields đúng tọa độ')

write_tc(11, 3,
    'Override completion persist khi sửa',
    'P1',
    'Khoá học đã lưu với 1 bài có override_completion=2',
    '1. Vào sửa khoá học\n2. Mở modal info bài học đó',
    '',
    '- Modal info bài học hiển thị "Đang ghi đè" (trạng thái màu xanh)\n- Tiêu chí hoàn thành hiển thị theo override, không phải mặc định của bài học')


# =============================================
# XII. PERMISSION
# =============================================
write_section_row('XII. PHÂN QUYỀN (Permission)')

write_tc(12, 1,
    'User không có quyền — bị từ chối truy cập',
    'P0',
    'Tài khoản không được phân quyền "Quản lý khoá học"',
    '1. Đăng nhập với tài khoản không có quyền\n2. Truy cập /training/subjects',
    '',
    '- Bị từ chối truy cập (403 hoặc redirect về trang lỗi)\n- Không hiển thị nội dung trang')

write_tc(12, 2,
    'Permission rename không ảnh hưởng user cũ',
    'P0',
    'User đã được phân quyền "Quản lý môn học" (tên cũ) trước khi chạy migration rename',
    '1. Chạy migration rename permission "Quản lý môn học" → "Quản lý khoá học"\n2. User đăng nhập lại\n3. Truy cập /training/subjects',
    '',
    '- User vẫn truy cập được (permission_id không đổi, role_has_permissions giữ nguyên)\n- Không bị mất quyền sau khi rename')


# =============================================
# XIII. BULLETLISTEDITOR
# =============================================
write_section_row('XIII. BULLETLISTEDITOR (what_includes, for_who, will_learn)')

write_tc(13, 1,
    'Nhập và lưu nội dung dạng bullet list',
    'P1',
    'Đang ở Tab 1, section hiển thị thông tin',
    '1. Click vào editor "Khoá học này có gì?"\n2. Nhập text\n3. Format bullet list\n4. Lưu khoá học',
    'Nội dung: "• Nội dung 1\n• Nội dung 2"',
    '- what_includes lưu dạng HTML <ul><li>...</li></ul>\n- Khi load lại: nội dung render đúng trong editor, không bị mất')

write_tc(13, 2,
    '3 BulletListEditor độc lập nhau',
    'P2',
    'Đang ở Tab 1',
    '1. Nhập nội dung khác nhau vào 3 editor:\n   - "Khoá học này có gì?" (what_includes)\n   - "Dành cho ai?" (for_who)\n   - "Bạn sẽ học được gì?" (will_learn)\n2. Lưu',
    '',
    '- 3 field lưu độc lập: what_includes, for_who, will_learn\n- Không ảnh hưởng lẫn nhau\n- Khi load lại: 3 field đúng nội dung riêng biệt')


# =============================================
# XIV. REGRESSION — DOWNSTREAM
# =============================================
write_section_row('XIV. REGRESSION — DOWNSTREAM KHÔNG BỊ ẢNH HƯỞNG')

write_tc(14, 1,
    'Màn Chương trình đào tạo vẫn hoạt động đúng',
    'P0',
    'Có chương trình đào tạo chứa khoá học đã được rebuild',
    '1. Vào Đào tạo > Chương trình đào tạo\n2. Xem danh sách khoá học trong chương trình',
    '',
    '- Khoá học hiển thị đúng tên, mã\n- Không lỗi JS console\n- Không lỗi API (500, 404)')

write_tc(14, 2,
    'Màn Lớp học (courses) vẫn load đúng',
    'P0',
    'Có lớp học chứa khoá học đã migrate',
    '1. Vào Đào tạo > Lớp học\n2. Xem chi tiết lớp học có khoá học đã migrate',
    '',
    '- Thông tin khoá học hiển thị đúng trong lớp học\n- Không lỗi API, không trang trắng')


# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# === SAVE ===
output_path = r'd:\CompanyProject\hrm\hrm-claude-config\docs\srs\Testcase_Subject_Builder.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Total rows written: {current_row - 7}')
section_count = 14
print(f'Total testcases: {current_row - 7 - section_count}')
