import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "My To Do"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ My To Do (Assign/MyTodo)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'My To Do'
PREFIX = 'MTD'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. MÀN HÌNH CHÍNH (/assign/my-todo)
# =============================================
write_section_row('I. MÀN HÌNH CHÍNH (/assign/my-todo)')

write_tc(1, 1,
    'Hiển thị trang My To Do',
    'P0',
    'User đã đăng nhập',
    '1. Menu Giao việc → My To Do',
    '',
    'Trang hiển thị 2 cột: trái (danh sách items) + phải (calendar + danh sách cá nhân)\nTitle "My To Do" trên topbar')

write_tc(1, 2,
    'Stats bar hiển thị đúng',
    'P0',
    'Có items ở nhiều trạng thái',
    '1. Xem phần stats phía trên',
    '',
    '4 ô stats: Quá hạn (đỏ), Hôm nay (xanh), Tuần này (teal), Tổng (xám)\nSố đếm chính xác theo dữ liệu')

write_tc(1, 3,
    'Filter bar hiển thị đầy đủ',
    'P0',
    'Đang ở trang My To Do',
    '1. Xem phần filter',
    '',
    'Dropdown Loại, dropdown Trạng thái, ô tìm kiếm, nút "Tạo todo"')

write_tc(1, 4,
    'Sidebar calendar hiển thị',
    'P0',
    '',
    '1. Xem cột phải',
    '',
    'Mini calendar tháng hiện tại\nNút ◀ / ▶ chuyển tháng\nDot indicator trên ngày có items')

write_tc(1, 5,
    'Danh sách cá nhân trên sidebar',
    'P0',
    'User có ≥1 danh sách',
    '1. Xem phần dưới calendar',
    '',
    'Tiêu đề "DANH SÁCH CÁ NHÂN"\nDot màu + tên list + "completed/total"\nNút "+ Tạo danh sách mới"')

write_tc(1, 6,
    'Item Task hiển thị đúng format',
    'P0',
    'User có task được giao',
    '1. Xem 1 item Task trong danh sách',
    '',
    'Viền trái xanh dương\nDòng 1: title + badge "Được giao" (xanh pill)\nDòng 2: status badge (border+bg) · tên DA · người phụ trách\nGóc phải: ngày hạn DD/MM')

write_tc(1, 7,
    'Item Issue hiển thị đúng format',
    'P0',
    'User có issue được giao',
    '1. Xem 1 item Issue',
    '',
    'Viền trái đỏ\nBadge "Được giao"\nStatus badge (Mới/Đang xử lý/...)')

write_tc(1, 8,
    'Item Meeting hiển thị đúng format',
    'P0',
    'User có meeting sắp tới',
    '1. Xem item Meeting',
    '',
    'Viền trái xanh dương nhạt\nBadge "Tham gia" (xanh lá pill)\nSố thành viên + địa điểm')

write_tc(1, 9,
    'Item Personal Todo hiển thị đúng format',
    'P0',
    'User có todo cá nhân',
    '1. Xem item Personal',
    '',
    'Viền trái xám\nCheckbox (thay icon loại)\nBadge "Cá nhân" (xám pill)\nStatus "Việc cần làm" (secondary)')

write_tc(1, 10,
    'Role badge hiển thị đúng màu theo source',
    'P0',
    'Có items với nhiều source khác nhau',
    '1. So sánh badge giữa các item',
    '',
    'Được giao = xanh dương pill (#4a7cfb, bg #EFF6FF)\nTham gia = xanh lá (#22c55e)\nCần duyệt = vàng (#d4a03c)\nTheo dõi = tím (#8b5cf6)\nCá nhân = xám (#6b7280)')

write_tc(1, 11,
    'Status badge có border + background',
    'P0',
    '',
    '1. Xem dòng 2 của item',
    '',
    'Badge nhỏ (10px), viền + background nhạt\n"Đang làm" = viền xanh + nền xanh nhạt\n"Quá hạn" = viền đỏ + nền đỏ nhạt')

write_tc(1, 12,
    'Dot separator giữa meta items',
    'P1',
    'Item có nhiều thông tin trên dòng 2',
    '1. Xem dòng 2',
    '',
    'Giữa status, project, assignee có chấm tròn nhỏ (3px xám) ngăn cách')

write_tc(1, 13,
    'Ngày hạn hôm nay hiện giờ thay ngày',
    'P1',
    'Item có due_date = hôm nay, due_time = 14:00',
    '1. Xem góc phải item',
    'due_date = today, due_time = 14:00',
    'Góc phải hiện "14:00" thay vì "DD/MM"')

write_tc(1, 14,
    'Ngày hạn quá hạn hiện chữ đỏ',
    'P0',
    'Item quá hạn',
    '1. Xem góc phải item quá hạn',
    '',
    'Ngày hạn chữ đỏ\nTitle KHÔNG đỏ (giữ màu bình thường)')

write_tc(1, 15,
    'Sub-task hiển thị dưới todo cha',
    'P0',
    'Todo cá nhân có sub-tasks',
    '1. Xem todo có sub-tasks',
    '',
    'Sub-tasks hiện bên dưới, thụt vào\nMỗi sub-task có checkbox + title')


# =============================================
# II. NHÓM THỜI GIAN
# =============================================
write_section_row('II. NHÓM THỜI GIAN')

write_tc(2, 1,
    'Nhóm "Hôm nay" mở mặc định',
    'P0',
    'Có items hôm nay',
    '1. Mở trang My To Do',
    '',
    'Nhóm Hôm nay expand\nHeader xanh #4a7cfb\nHiện ngày DD/MM + count')

write_tc(2, 2,
    'Nhóm "Ngày mai" thu gọn mặc định',
    'P0',
    'Có items ngày mai',
    '1. Xem nhóm Ngày mai',
    '',
    'Thu gọn (collapse)\nHeader vàng #d4a03c\nClick → expand hiện items')

write_tc(2, 3,
    'Nhóm "Tuần này" thu gọn mặc định',
    'P0',
    'Có items tuần này',
    '1. Xem nhóm Tuần này',
    '',
    'Thu gọn\nHeader teal #14b8a6')

write_tc(2, 4,
    'Nhóm "Quá hạn" hiển thị items quá hạn',
    'P0',
    'Có items due_date < today',
    '1. Xem nhóm Quá hạn',
    '',
    'Header đỏ #e25c5c\nThu gọn mặc định\nChứa đúng items quá hạn')

write_tc(2, 5,
    'Nhóm "Không hạn" cho items không có due_date',
    'P1',
    'Có todo không chọn ngày hạn',
    '1. Tạo todo không due_date\n2. Xem nhóm Không hạn',
    '',
    'Item xuất hiện trong nhóm "Không hạn"')

write_tc(2, 6,
    'Toggle collapse/expand nhóm',
    'P0',
    '',
    '1. Click header nhóm đang collapse\n2. Click lại header',
    '',
    'Click 1: mở ra hiện items\nClick 2: thu gọn')

write_tc(2, 7,
    'Nhóm rỗng vẫn hiển thị (Hôm nay, Ngày mai, Tuần này, Quá hạn)',
    'P1',
    'Không có item nào hôm nay',
    '1. Xem nhóm Hôm nay',
    '',
    'Nhóm vẫn hiện với count = 0')


# =============================================
# III. BỘ LỌC & TÌM KIẾM
# =============================================
write_section_row('III. BỘ LỌC & TÌM KIẾM')

write_tc(3, 1,
    'Lọc theo loại: Task',
    'P0',
    'Có items nhiều loại',
    '1. Chọn "Task" trong dropdown Loại',
    'type = task',
    'Chỉ hiện items type = task\nStats cập nhật theo filter')

write_tc(3, 2,
    'Lọc theo loại: Issue',
    'P0',
    '',
    '1. Chọn "Issue"',
    'type = issue',
    'Chỉ hiện issues')

write_tc(3, 3,
    'Lọc theo loại: Meeting',
    'P1',
    '',
    '1. Chọn "Meeting"',
    'type = meeting',
    'Chỉ hiện meetings')

write_tc(3, 4,
    'Lọc theo loại: Cá nhân',
    'P0',
    '',
    '1. Chọn "Cá nhân"',
    'type = personal',
    'Chỉ hiện personal todos')

write_tc(3, 5,
    'Lọc trạng thái: Chưa xong (mặc định)',
    'P0',
    '',
    '1. Xem dropdown Trạng thái',
    'status = pending',
    'Mặc định = "Chưa xong"\nChỉ hiện items chưa hoàn thành')

write_tc(3, 6,
    'Lọc trạng thái: Đã xong',
    'P0',
    'Có todos đã hoàn thành',
    '1. Chọn "Đã xong"',
    'status = done',
    'Chỉ hiện personal todos đã hoàn thành (BR-10)\nBỏ qua entity hệ thống')

write_tc(3, 7,
    'Lọc trạng thái: Tất cả',
    'P1',
    '',
    '1. Chọn "Tất cả"',
    '',
    'Hiện cả xong + chưa xong')

write_tc(3, 8,
    'Tìm kiếm theo title (client-side)',
    'P0',
    'Có item title chứa "báo cáo"',
    '1. Gõ "báo cáo" vào ô tìm kiếm',
    'keyword = "báo cáo"',
    'Filter client-side, chỉ hiện items match\nKhông phân biệt hoa thường')

write_tc(3, 9,
    'Tìm kiếm không có kết quả',
    'P1',
    '',
    '1. Gõ "XYZNOTEXIST"',
    '',
    'Danh sách trống\nCác nhóm hiện header nhưng không có items')

write_tc(3, 10,
    'Kết hợp filter loại + trạng thái',
    'P0',
    '',
    '1. Chọn Loại = Task\n2. Chọn Trạng thái = Chưa xong',
    '',
    'Chỉ hiện tasks chưa hoàn thành')


# =============================================
# IV. CALENDAR
# =============================================
write_section_row('IV. CALENDAR')

write_tc(4, 1,
    'Click ngày trên calendar → flat mode',
    'P0',
    'Có items ngày 03/05',
    '1. Click ngày 03/05 trên mini calendar',
    'date = 2026-05-03',
    'Chuyển flat mode, chỉ hiện items due_date = 03/05\nNgày được highlight trên calendar')

write_tc(4, 2,
    'Click lại cùng ngày → quay về grouped mode',
    'P0',
    'Đang chọn ngày 03/05',
    '1. Click lại ngày 03/05',
    '',
    'Bỏ highlight\nQuay về grouped mode (nhóm theo thời hạn)')

write_tc(4, 3,
    'Dot indicator trên ngày có items',
    'P0',
    'Items có due_date 01/05 và 02/05',
    '1. Xem mini calendar',
    '',
    'Ngày 01/05 và 02/05 có dot indicator (chấm nhỏ)')

write_tc(4, 4,
    'Chuyển tháng trước/sau',
    'P1',
    '',
    '1. Click ◀ (tháng trước)\n2. Click ▶ (tháng sau)',
    '',
    'Calendar chuyển tháng\nDot indicator cập nhật theo dữ liệu tháng mới')

write_tc(4, 5,
    'Click ngày không có items',
    'P1',
    '',
    '1. Click 1 ngày không có item nào',
    '',
    'Chuyển flat mode, danh sách trống')


# =============================================
# V. CRUD TODO CÁ NHÂN
# =============================================
write_section_row('V. CRUD TODO CÁ NHÂN')

write_tc(5, 1,
    'Mở modal tạo todo',
    'P0',
    '',
    '1. Click "Tạo todo" trên filter bar',
    '',
    'Modal hiện: title (*), description, dropdown danh sách, date picker, time picker')

write_tc(5, 2,
    'Tạo todo thành công',
    'P0',
    '',
    '1. Điền title "Gọi NCC"\n2. Chọn list\n3. Chọn ngày hạn\n4. Click Lưu',
    'title = "Gọi NCC"',
    'Modal đóng\nTodo mới xuất hiện đúng nhóm thời gian\nToast thành công')

write_tc(5, 3,
    'Tạo todo không có ngày hạn',
    'P0',
    '',
    '1. Điền title\n2. KHÔNG chọn ngày hạn\n3. Lưu',
    '',
    'Tạo thành công\nItem ở nhóm "Không hạn"')

write_tc(5, 4,
    'Tạo todo thiếu title (validation)',
    'P0',
    '',
    '1. Để trống title\n2. Lưu',
    'title = ""',
    'Hiện lỗi validation "Tiêu đề là bắt buộc"\nKhông gọi API')

write_tc(5, 5,
    'Hủy tạo todo',
    'P1',
    '',
    '1. Mở modal, điền thông tin\n2. Bấm Hủy',
    '',
    'Modal đóng, không tạo todo mới')

write_tc(5, 6,
    'Đánh dấu hoàn thành todo',
    'P0',
    'Có todo chưa xong',
    '1. Click checkbox todo',
    '',
    'Checkbox checked\nTitle gạch ngang (line-through)\nStatus = "Hoàn thành" (success)\nAPI PATCH /toggle gọi đúng')

write_tc(5, 7,
    'Bỏ đánh dấu hoàn thành',
    'P0',
    'Có todo đã xong',
    '1. Click checkbox todo đã xong',
    '',
    'Checkbox unchecked\nBỏ gạch ngang\nStatus = "Việc cần làm" (secondary)')

write_tc(5, 8,
    'Toggle sub-task',
    'P0',
    'Todo có sub-task',
    '1. Click checkbox sub-task',
    '',
    'Sub-task toggle hoàn thành\nTitle gạch/bỏ gạch')

write_tc(5, 9,
    'Completed_at ghi đúng khi toggle (BR-07)',
    'P1',
    '',
    '1. Đánh hoàn thành → kiểm tra DB\n2. Bỏ hoàn thành → kiểm tra DB',
    '',
    'Hoàn thành: completed_at = now()\nBỏ: completed_at = null')

write_tc(5, 10,
    'Sửa title todo',
    'P0',
    'Có todo cá nhân',
    '1. Mở modal sửa → đổi title → Lưu',
    '',
    'Title cập nhật trên danh sách')

write_tc(5, 11,
    'Đổi ngày hạn todo',
    'P1',
    '',
    '1. Mở modal sửa → chọn ngày mới → Lưu',
    '',
    'Item chuyển sang nhóm thời gian tương ứng')

write_tc(5, 12,
    'Chuyển todo sang list khác',
    'P1',
    'Có ≥2 lists',
    '1. Mở modal sửa → đổi danh sách → Lưu',
    '',
    'Todo chuyển sang list mới\nCount cũ giảm, count mới tăng')

write_tc(5, 13,
    'Xóa todo cá nhân',
    'P0',
    '',
    '1. Hover todo → click xóa → xác nhận',
    '',
    'Todo biến mất, count giảm')

write_tc(5, 14,
    'Xóa todo có sub-tasks (cascade BR-04)',
    'P0',
    'Todo cha có 2 sub-tasks',
    '1. Xóa todo cha → xác nhận',
    '',
    'Todo cha + 2 sub-tasks đều bị xóa')

write_tc(5, 15,
    'Hủy xóa',
    'P1',
    '',
    '1. Click xóa → bấm Hủy',
    '',
    'Dialog đóng, todo vẫn còn')


# =============================================
# VI. SUB-TASK
# =============================================
write_section_row('VI. SUB-TASK')

write_tc(6, 1,
    'Tạo sub-task',
    'P0',
    'Đang xem chi tiết list',
    '1. Click "Thêm bước con" dưới todo cha\n2. Nhập title\n3. Enter',
    '',
    'Sub-task xuất hiện dưới todo cha, thụt vào')

write_tc(6, 2,
    'Không cho tạo sub-task cấp 2 (BR-01)',
    'P0',
    'Todo đã là sub-task (parent_id != null)',
    '1. Cố tạo sub-task cho 1 sub-task',
    'parent.parent_id != null',
    'API trả lỗi 400\nKhông tạo được')

write_tc(6, 3,
    'Xóa sub-task',
    'P0',
    'Todo cha có sub-task',
    '1. Hover sub-task → click xóa',
    '',
    'Sub-task bị xóa\nTodo cha vẫn còn')


# =============================================
# VII. CRUD DANH SÁCH CÁ NHÂN
# =============================================
write_section_row('VII. CRUD DANH SÁCH CÁ NHÂN')

write_tc(7, 1,
    'Tạo danh sách mới',
    'P0',
    '',
    '1. Click "+ Tạo danh sách mới"\n2. Điền tên → Lưu',
    'name = "Follow up"',
    'Modal đóng\nList mới xuất hiện trên sidebar với count "0/0"')

write_tc(7, 2,
    'Tạo danh sách thiếu tên (validation)',
    'P0',
    '',
    '1. Mở modal → để trống name → Lưu',
    'name = ""',
    'Lỗi: "Tên danh sách là bắt buộc"')

write_tc(7, 3,
    'Tạo danh sách với mô tả',
    'P1',
    '',
    '1. Điền name + description → Lưu',
    '',
    'Tạo thành công, description lưu đúng')

write_tc(7, 4,
    'Sort order tự tăng (BR-08)',
    'P1',
    'Đã có 2 lists',
    '1. Tạo list thứ 3',
    '',
    'sort_order = max + 1\nList mới ở cuối')

write_tc(7, 5,
    'Sửa tên danh sách',
    'P0',
    'Đang xem chi tiết list',
    '1. Click "Sửa" → đổi tên → Lưu',
    '',
    'Tên cập nhật trên sidebar + header')

write_tc(7, 6,
    'Xóa danh sách rỗng',
    'P0',
    'List không có todo',
    '1. Xóa → xác nhận',
    '',
    'List biến mất khỏi sidebar')

write_tc(7, 7,
    'Xóa danh sách có todos (cascade BR-03)',
    'P0',
    'List có 3 todos',
    '1. Xóa → xác nhận',
    '',
    'List + tất cả 3 todos bị xóa\nQuay về trang chính nếu đang xem chi tiết list đó')

write_tc(7, 8,
    'Hủy xóa danh sách',
    'P1',
    '',
    '1. Click xóa → bấm Hủy',
    '',
    'Confirm đóng, list vẫn còn')

write_tc(7, 9,
    'Tự tạo list mặc định khi chưa có (BR-02)',
    'P0',
    'User mới, chưa có list nào',
    '1. Truy cập My To Do lần đầu',
    '',
    'Hệ thống tự tạo list "Mac dinh"\nSidebar hiện 1 list')

write_tc(7, 10,
    'Không tạo thêm nếu đã có list',
    'P1',
    'User đã có 2 lists',
    '1. Truy cập lại My To Do',
    '',
    'Không tạo list mới, giữ 2 lists cũ')


# =============================================
# VIII. KÉO THẢ SẮP XẾP DANH SÁCH
# =============================================
write_section_row('VIII. KÉO THẢ SẮP XẾP DANH SÁCH')

write_tc(8, 1,
    'Drag handle hiện khi hover',
    'P0',
    'Có ≥2 lists',
    '1. Hover vào 1 list trên sidebar',
    '',
    'Icon drag handle (≡) xuất hiện bên trái')

write_tc(8, 2,
    'Kéo thả đổi vị trí list',
    'P0',
    'Có ≥2 lists',
    '1. Kéo list "Follow up" lên trên "Việc cần làm"',
    '',
    'UI cập nhật ngay (optimistic update)\nThứ tự mới hiển thị đúng')

write_tc(8, 3,
    'Kéo thả persist sau reload',
    'P0',
    'Vừa kéo thả xong',
    '1. Reload trang (F5)',
    '',
    'Thứ tự giữ nguyên sau reload\nAPI đã lưu sort_order')

write_tc(8, 4,
    'Kéo thả khi chỉ có 1 list',
    'P2',
    'Chỉ có 1 list',
    '1. Cố kéo thả',
    '',
    'Không có gì thay đổi')

write_tc(8, 5,
    'Kéo thả API lỗi → rollback',
    'P1',
    'Giả lập API lỗi',
    '1. Kéo thả (API lỗi)',
    '',
    'UI rollback về vị trí cũ\nToast "Không thể sắp xếp danh sách" (danger)')


# =============================================
# IX. CHI TIẾT DANH SÁCH
# =============================================
write_section_row('IX. CHI TIẾT DANH SÁCH')

write_tc(9, 1,
    'Click list → mở chi tiết',
    'P0',
    'Có ≥1 list',
    '1. Click tên list trên sidebar',
    '',
    'Cột trái chuyển sang TodoListDetail\nHeader: tên + mô tả + Sửa/Xóa\nDanh sách todos + input thêm mới')

write_tc(9, 2,
    'Click lại list → quay về chính (toggle)',
    'P0',
    'Đang ở chi tiết list',
    '1. Click lại tên list đó',
    '',
    'Quay về trang chính')

write_tc(9, 3,
    'Header chi tiết hiển thị đúng',
    'P0',
    'List có 3 todo chưa xong, 2 đã xong',
    '1. Mở chi tiết list',
    '',
    '"My To Do > [Tên list]"\n"Tổng: 5 việc | Hoàn thành: 2"\nNút Sửa + Xóa')

write_tc(9, 4,
    'Thêm todo inline (Enter)',
    'P0',
    'Đang ở chi tiết list',
    '1. Gõ "Việc mới" vào input cuối\n2. Enter',
    'title = "Việc mới"',
    'Todo mới xuất hiện cuối\nInput xóa trắng sẵn sàng nhập tiếp')

write_tc(9, 5,
    'Thêm sub-task inline',
    'P0',
    'Có todo trong list',
    '1. Click "Thêm bước con"\n2. Gõ title → Enter',
    '',
    'Sub-task xuất hiện dưới todo cha')

write_tc(9, 6,
    'Toggle todo trong chi tiết',
    'P0',
    '',
    '1. Click checkbox 1 todo',
    '',
    'Toggle trạng thái\nStats header cập nhật ("Hoàn thành: Y+1")')

write_tc(9, 7,
    'Xóa todo trong chi tiết',
    'P0',
    '',
    '1. Hover todo → click icon xóa',
    '',
    'Todo bị xóa\nStats cập nhật')

write_tc(9, 8,
    'Thêm inline title rỗng → không tạo',
    'P1',
    '',
    '1. Enter ở input trống',
    'title = ""',
    'Không tạo todo\nKhông gọi API')


# =============================================
# X. TƯƠNG TÁC ITEM HỆ THỐNG
# =============================================
write_section_row('X. TƯƠNG TÁC ITEM HỆ THỐNG (TASK/ISSUE)')

write_tc(10, 1,
    'Hover Task → hiện action buttons',
    'P0',
    'Có item Task',
    '1. Hover vào item Task',
    '',
    'Hiện buttons tùy quyền: Xem, Sửa (can_edit), Nhập KQ (can_import_result), Duyệt (can_approve), Lịch sử')

write_tc(10, 2,
    'Click "Xem" task',
    'P0',
    'Task status 1-3, 10',
    '1. Click nút Xem',
    '',
    'Mở CreateTaskModal chế độ view (readonly)')

write_tc(10, 3,
    'Click "Sửa" task',
    'P0',
    'Task can_edit = true',
    '1. Click nút Sửa',
    '',
    'Mở CreateTaskModal chế độ edit')

write_tc(10, 4,
    'Click "Nhập KQ" task',
    'P0',
    'Task can_import_result = true',
    '1. Click nút Nhập KQ',
    '',
    'Mở ImportResultModal')

write_tc(10, 5,
    'Click "Duyệt" task',
    'P1',
    'Task can_approve = true',
    '1. Click nút Duyệt',
    '',
    'Mở ImportResultModal chế độ duyệt')

write_tc(10, 6,
    'Click "Lịch sử" task',
    'P1',
    '',
    '1. Click nút Lịch sử',
    '',
    'Mở TaskHistoryModal')

write_tc(10, 7,
    'Click "Xem" issue',
    'P0',
    'Có item Issue',
    '1. Click nút Xem',
    '',
    'Mở CreateIssueModal chế độ view')

write_tc(10, 8,
    'Click "Xử lý" issue',
    'P0',
    'Issue can_handle = true',
    '1. Click nút Xử lý',
    '',
    'Mở CreateIssueModal chế độ xử lý')

write_tc(10, 9,
    'Click "Lịch sử" issue',
    'P1',
    '',
    '1. Click nút Lịch sử',
    '',
    'Mở IssueHistoryModal')

write_tc(10, 10,
    'Click AssignJob → mở tab mới',
    'P0',
    'Có item Phiếu giao việc',
    '1. Click vào item',
    '',
    'window.open(item.url, "_blank")\nMở tab mới tới trang chi tiết phiếu')

write_tc(10, 11,
    'Click Meeting → mở tab mới',
    'P0',
    'Có item Meeting',
    '1. Click vào item',
    '',
    'Mở tab mới tới trang chi tiết cuộc họp')


# =============================================
# XI. DỮ LIỆU & BẢO MẬT
# =============================================
write_section_row('XI. DỮ LIỆU & BẢO MẬT')

write_tc(11, 1,
    'User chỉ thấy dữ liệu của mình (BR-05)',
    'P0',
    'User A tạo 2 todos',
    '1. Đăng nhập User B\n2. Mở My To Do',
    '2 accounts',
    'User B KHÔNG thấy todos của User A')

write_tc(11, 2,
    'Deduplicate items (BR-06)',
    'P0',
    'User vừa là assignee vừa là watcher 1 task',
    '1. Mở My To Do',
    '',
    'Task chỉ hiện 1 lần\nSource = "assigned" (ưu tiên cao nhất)')

write_tc(11, 3,
    'Stats đếm đúng cấp 1 (BR-09)',
    'P0',
    'List có 3 todo cấp 1 (1 xong), 2 sub-task (1 xong)',
    '1. Xem sidebar list',
    '',
    'Count hiện "1/3" (chỉ đếm cấp 1)\nKhông đếm sub-tasks')

write_tc(11, 4,
    'API trả 401 khi chưa đăng nhập',
    'P0',
    'Chưa login',
    '1. Gọi GET /assign/my-todo không có token',
    '',
    'HTTP 401 Unauthorized')

write_tc(11, 5,
    'Không thao tác todo/list của user khác qua API',
    'P0',
    'User A tạo todo id=1',
    '1. Đăng nhập User B\n2. Gọi DELETE /assign/my-todo/todos/1',
    '',
    'HTTP 404 hoặc 403\nTodo không bị xóa')


# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# === SAVE ===
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, 'my-todo-testcases.xlsx')
wb.save(output_path)
print(f'Saved: {output_path}')
# Count actual test cases (rows - section rows)
section_count = 11  # number of write_section_row calls
tc_count = current_row - 7 - section_count
print(f'Total testcases: {tc_count}')
