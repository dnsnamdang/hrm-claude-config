"""Testcase Excel — Quyết toán HĐ: hạch toán tiền vận chuyển + bổ sung TH3 (HĐ hãng & dịch vụ)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/quyet-toan-hach-toan-van-chuyen-th3/testcase.xlsx"
SHEET_NAME   = "HachToanVanChuyenTH3"
FEATURE_NAME = "Quyết toán hợp đồng — hạch toán tiền vận chuyển (bổ sung TH3)"
MODULE_NAME  = "Quyết toán HĐ - Vận chuyển"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Bước '8. Hạch toán tiền vận chuyển' (deliveryAccounting) trong quyết toán Hợp đồng HÃNG (FirmSettlementContractAccountingService) và Hợp đồng DỊCH VỤ (WrSettlementContractAccountingService). Bổ sung TH3: khi không có số tiền vận chuyển trên TK 148, hệ thống xét TK 230 và hạch toán phần chênh lệch tiền vận chuyển."),
    ("2. Đối tượng được tính / hiển thị",
     "► Số tiền vận chuyển lấy từ getAmountDelivery(contractable, TK) + cộng dồn các phụ lục bổ sung (addition_annexes).\n"
     "► TH1: amount(TK148) > 0 và giảm-giá-HĐ ≤ hạn mức giảm giá → hạch toán = before_vat_delivery_cost.\n"
     "► TH2: amount(TK148) > 0 và giảm-giá-HĐ > hạn mức → hạch toán = min(amount148, before_vat_delivery_cost).\n"
     "► TH3 (MỚI): amount(TK148) = 0 → xét amount(TK230); nếu amount230 ≠ 0 và (discount_product+discount_repair_service+discount_delivery_cost) ≤ hạn mức và before_vat_delivery_cost > amount230 → hạch toán = before_vat_delivery_cost − amount230.\n"
     "► Bút toán: 3351↔35241 (số TK148, work TVC); 5211↔3351 (phần ghi nhận giảm trừ, work TVC)."),
    ("3. Đối tượng bị ẩn / không tính",
     "► amount(TK148)=0 và amount(TK230)=0 → không hạch toán vận chuyển.\n"
     "► TH3 không chạy nếu before_vat_delivery_cost ≤ amount230, hoặc giảm giá > hạn mức."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng (bút toán phát sinh khi tính/lưu quyết toán)."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "settlement_contract → contractable (firm_contract / wr_service_contract) + addition_annexes. account_details (bút toán: identify_number 3351/35241/5211, work TVC, type DEPT/HAS, group). settlement_data: before_vat_delivery_cost, discount_product, discount_repair_service, discount_other_cost, discount_delivery_cost, discount_rate_total (firm) / quota_discount_total (wr)."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "amount = getAmountDelivery(HĐ, TK) + Σ getAmountDelivery(phụ lục bổ sung, TK). Mỗi lần tính lại quyết toán xóa & dựng lại bút toán. group tăng dần sau mỗi bước hạch toán."),
    ("7. Phân quyền cấp",
     "• Theo quyền tính/lưu quyết toán hợp đồng (hãng/dịch vụ). Không có permission code riêng cho bước vận chuyển."),
    ("8. Cách tính các ô thống kê",
     "► amount148 = getAmountDelivery(contractable_id, 148) + Σ annex(148).\n"
     "► giam_gia_hd = discount_product + discount_repair_service + discount_other_cost.\n"
     "► TH1: ghi 5211↔3351 = before_vat_delivery_cost.\n"
     "► TH2: ghi 5211↔3351 = min(amount148, before_vat_delivery_cost).\n"
     "► TH3: amount230 = getAmountDelivery(.,230)+Σannex; ghi 5211↔3351 = before_vat_delivery_cost − amount230 (khi điều kiện TH3 thỏa)."),
    ("9. Ghi chú đọc bảng",
     "Note bút toán = '{group}. Hạch toán tiền vận chuyển'. Áp cho cả 2 service (Firm/Wr) — logic tương đương, khác tên field hạn mức (discount_rate_total vs quota_discount_total)."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "TỔNG QUAN BƯỚC HẠCH TOÁN VẬN CHUYỂN", [
        ("001", "Bước 8 'Hạch toán tiền vận chuyển' chạy khi tính quyết toán", "P0",
         "HĐ hãng đủ điều kiện quyết toán, có tiền vận chuyển",
         "1. Tính/lưu quyết toán HĐ\n2. Xem bút toán hạch toán",
         "—",
         "- Phát sinh nhóm bút toán note '{group}. Hạch toán tiền vận chuyển'",
         "deliveryAccounting() bước 8"),
        ("002", "Bút toán 3351 ↔ 35241 khi có số TK 148", "P0",
         "amount(TK148) = 2.000.000",
         "1. Tính quyết toán",
         "amount148=2tr",
         "- Ghi 3351 (Nợ) / 35241 (Có) = 2.000.000, work = TVC",
         "createDataSaveDept 3351/35241"),
        ("003", "Không hạch toán vận chuyển khi không có tiền (148 & 230 = 0)", "P0",
         "amount148=0, amount230=0",
         "1. Tính quyết toán",
         "0/0",
         "- Không phát sinh bút toán vận chuyển",
         "else không vào TH3 khi amount230=0"),
        ("004", "HĐ DỊCH VỤ cũng chạy bước 8 hạch toán vận chuyển", "P0",
         "HĐ dịch vụ đủ điều kiện quyết toán, có tiền vận chuyển",
         "1. Tính quyết toán HĐ dịch vụ\n2. Xem bút toán",
         "Wr",
         "- WrSettlementContractAccountingService::deliveryAccounting chạy, note '{group}. Hạch toán tiền vận chuyển'",
         "Wr bước 8"),
    ]),
    ("V", "CÁC TRƯỜNG HỢP HẠCH TOÁN (TH1/TH2/TH3)", [
        ("001", "[TH1] TK148 > 0, giảm giá HĐ ≤ hạn mức", "P0",
         "amount148=2tr; giam_gia_hd ≤ discount_rate_total; before_vat_delivery_cost=3tr",
         "1. Tính quyết toán HĐ hãng",
         "148=2tr, bvdc=3tr, giảm≤hạn",
         "- 5211 (Nợ) / 3351 (Có) = before_vat_delivery_cost = 3.000.000",
         "TH1: giam_gia_hd <= discount_rate_total"),
        ("002", "[TH2] TK148 > 0, giảm giá HĐ > hạn mức, bvdc ≥ amount148", "P0",
         "amount148=2tr; giam_gia_hd > hạn mức; before_vat_delivery_cost=3tr",
         "1. Tính quyết toán",
         "148=2tr, bvdc=3tr, giảm>hạn",
         "- 5211/3351 = amount148 = 2.000.000 (vì bvdc ≥ amount148, giữ amount148)",
         "TH2: bvdc < amount → amount=bvdc; else giữ amount148"),
        ("003", "[TH2] TK148 > 0, giảm giá > hạn mức, bvdc < amount148", "P1",
         "amount148=3tr; giam_gia_hd > hạn mức; before_vat_delivery_cost=2tr",
         "1. Tính quyết toán",
         "148=3tr, bvdc=2tr",
         "- amount = before_vat_delivery_cost = 2.000.000 (bvdc < amount148 nên lấy bvdc)\n- 5211/3351 = 2.000.000",
         "if bvdc < amount → amount = bvdc"),
        ("004", "[TH3 - MỚI] TK148 = 0, TK230 > 0, giảm giá ≤ hạn mức, bvdc > amount230", "P0",
         "amount148=0; amount230=1tr; (discount_product+discount_repair_service+discount_delivery_cost) ≤ hạn mức; before_vat_delivery_cost=3tr",
         "1. Tính quyết toán HĐ hãng",
         "148=0, 230=1tr, bvdc=3tr",
         "- Vào nhánh else (TH3): hạch toán 5211/3351 = before_vat_delivery_cost − amount230 = 3tr − 1tr = 2.000.000",
         "TH3: amount = bvdc - amount230"),
        ("005", "[TH3] Áp dụng cho HĐ DỊCH VỤ (Wr)", "P0",
         "HĐ dịch vụ; amount148=0, amount230=1tr; giảm giá ≤ quota_discount_total; bvdc=3tr",
         "1. Tính quyết toán HĐ dịch vụ",
         "Wr, 148=0,230=1tr,bvdc=3tr",
         "- TH3 chạy tương tự HĐ hãng: 5211/3351 = bvdc − amount230 = 2.000.000",
         "WrSettlementContractAccountingService TH3 (quota_discount_total)"),
        ("006", "[TH1] Áp dụng cho HĐ DỊCH VỤ (Wr)", "P1",
         "HĐ dịch vụ; amount148=2tr; giảm ≤ quota_discount_total; bvdc=3tr",
         "1. Tính quyết toán HĐ dịch vụ",
         "Wr, 148=2tr",
         "- 5211/3351 = before_vat_delivery_cost = 3.000.000",
         "Wr TH1"),
        ("007", "[TH2] Áp dụng cho HĐ DỊCH VỤ (Wr)", "P1",
         "HĐ dịch vụ; amount148=2tr; giảm > quota_discount_total; bvdc=3tr",
         "1. Tính quyết toán HĐ dịch vụ",
         "Wr, 148=2tr, giảm>hạn",
         "- 5211/3351 = min(amount148, bvdc) = 2.000.000",
         "Wr TH2"),
        ("008", "[TH3] discount_delivery_cost được tính vào điều kiện so hạn mức", "P1",
         "amount148=0; amount230=1tr; discount_delivery_cost làm tổng giảm ≤ hạn mức; bvdc=3tr",
         "1. Tính quyết toán",
         "có discount_delivery_cost",
         "- Điều kiện TH3 dùng (discount_product+discount_repair_service+discount_delivery_cost) ≤ hạn mức → thỏa → ghi chênh = 2.000.000",
         "TH3 gồm discount_delivery_cost"),
    ]),
    ("VI", "EDGE CASES — điều kiện không hạch toán / biên", [
        ("001", "[TH3] before_vat_delivery_cost ≤ amount230 → KHÔNG hạch toán", "P0",
         "amount148=0; amount230=3tr; before_vat_delivery_cost=2tr",
         "1. Tính quyết toán",
         "230=3tr, bvdc=2tr",
         "- Điều kiện 'bvdc > amount230' SAI → không ghi bút toán phần chênh (TH3 bỏ qua)",
         "TH3 cần before_vat_delivery_cost > amount230"),
        ("002", "[TH3] giảm giá > hạn mức → KHÔNG hạch toán phần chênh", "P0",
         "amount148=0; amount230=1tr; (discount_product+repair_service+delivery_cost) > hạn mức; bvdc=3tr",
         "1. Tính quyết toán",
         "giảm > hạn",
         "- Điều kiện giảm giá ≤ hạn mức SAI → TH3 không ghi 5211/3351",
         "TH3 cần discount ≤ discount_rate_total"),
        ("003", "[TH3] amount230 = 0 → không hạch toán", "P1",
         "amount148=0; amount230=0",
         "1. Tính quyết toán",
         "0/0",
         "- amount=0 → không vào bút toán TH3",
         "amount != 0 mới chạy"),
        ("004", "before_vat_delivery_cost = 0 (HĐ không có vận chuyển)", "P1",
         "before_vat_delivery_cost=0",
         "1. Tính quyết toán",
         "bvdc=0",
         "- Không phát sinh bút toán giảm trừ vận chuyển (số 0)",
         "amount = bvdc → 0"),
        ("008", "[TH1] amount148>0 nhưng bvdc=0 → chỉ bút toán 3351/35241", "P2",
         "amount148=2tr; before_vat_delivery_cost=0; giảm ≤ hạn",
         "1. Tính quyết toán",
         "148=2tr, bvdc=0",
         "- Vẫn ghi 3351/35241=2tr; phần 5211/3351 = bvdc = 0 (không phát sinh tiền giảm trừ)",
         "TH1 amount=bvdc=0"),
        ("005", "Biên giảm giá == hạn mức (≤)", "P2",
         "giam_gia_hd == discount_rate_total (TH1 dùng dấu ≤)",
         "1. Tính quyết toán TH1",
         "giảm = hạn",
         "- Vẫn vào TH1 (điều kiện ≤), hạch toán before_vat_delivery_cost",
         "<= bao gồm bằng"),
        ("006", "[TH3] bvdc = amount230 + 1 (biên > )", "P2",
         "amount148=0; amount230=1.000.000; bvdc=1.000.001; giảm ≤ hạn",
         "1. Tính quyết toán",
         "bvdc vừa lớn hơn 230",
         "- bvdc > amount230 đúng → ghi chênh = 1 đồng",
         "biên điều kiện > (không bao gồm bằng)"),
        ("007", "[TH3] bvdc == amount230 → KHÔNG hạch toán", "P1",
         "amount148=0; amount230=1tr; bvdc=1tr",
         "1. Tính quyết toán",
         "bvdc = 230",
         "- bvdc > amount230 SAI (bằng nhau) → không ghi",
         "cần strictly >"),
    ]),
    ("VII", "PHỤ LỤC BỔ SUNG (cộng dồn amount)", [
        ("001", "amount cộng dồn từ phụ lục bổ sung (TK148)", "P0",
         "HĐ amount148=1tr; 2 phụ lục bổ sung amount(148)=500k mỗi cái",
         "1. Tính quyết toán",
         "HĐ 1tr + 2×500k",
         "- amount148 = 1tr + 500k + 500k = 2.000.000; bút toán 3351/35241 = 2.000.000",
         "Σ getAmountDelivery(annex,148)"),
        ("002", "[TH3] amount230 cộng dồn phụ lục", "P1",
         "amount148=0; HĐ amount230=500k + phụ lục 148=500k (theo code annex dùng 148)",
         "1. Tính quyết toán",
         "—",
         "- amount230 = 500k + Σ annex(148) = 1.000.000; TH3 dùng amount230 này",
         "else: amount230 + Σ annex(148)"),
        ("003", "HĐ KHÔNG có phụ lục bổ sung", "P2",
         "HĐ amount148=2tr, không có addition_annexes",
         "1. Tính quyết toán",
         "không phụ lục",
         "- amount148 = 2tr (chỉ từ HĐ gốc); vòng lặp phụ lục không cộng thêm",
         "addition_annexes rỗng"),
    ]),
    ("VIII", "ĐỐI CHIẾU BÚT TOÁN & CÔ LẬP", [
        ("001", "Bút toán dùng đúng work TVC + phòng + nhân viên", "P1",
         "Quyết toán có hạch toán vận chuyển",
         "1. Kiểm bút toán",
         "—",
         "- Các dòng có Work=TVC, employee_department_id = department phòng quyết toán, Employee = employee_id",
         "createDataSaveDept(... Work TVC ...)"),
        ("002", "group tăng sau bước vận chuyển", "P2",
         "Bước 8 hạch toán vận chuyển",
         "1. Kiểm số group các bước",
         "—",
         "- Sau deliveryAccounting, group += 1 (bước kế tiếp dùng group mới)",
         "$group += 1"),
        ("003", "Tính lại quyết toán không nhân đôi bút toán vận chuyển", "P0",
         "Đã tính quyết toán 1 lần",
         "1. Tính lại quyết toán",
         "—",
         "- Bút toán vận chuyển được dựng lại, không bị nhân đôi",
         "recalculate xóa & dựng lại"),
        ("004", "TH3 không ảnh hưởng TH1/TH2 (loại trừ nhánh)", "P1",
         "amount148 > 0",
         "1. Tính quyết toán",
         "148>0",
         "- Khi amount148 > 0 chỉ chạy TH1/TH2; KHÔNG vào nhánh else (TH3)",
         "if(amount){TH1/TH2} else {TH3}"),
        ("005", "Cặp bút toán cân Nợ = Có", "P0",
         "Có hạch toán vận chuyển (bất kỳ TH)",
         "1. Tổng hợp account_details của nhóm vận chuyển",
         "—",
         "- Mỗi cặp cân: 3351(Nợ)=35241(Có); 5211(Nợ)=3351(Có); tổng Nợ = tổng Có",
         "TYPE_DEPT vs TYPE_HAS đối ứng"),
        ("006", "Note bút toán đúng định dạng", "P2",
         "Có hạch toán vận chuyển ở nhóm thứ k",
         "1. Xem note các dòng",
         "—",
         "- Note = '{k}. Hạch toán tiền vận chuyển'",
         "$note = group.'. Hạch toán tiền vận chuyển'"),
        ("007", "Type Nợ/Có đúng cho từng tài khoản", "P1",
         "TH1: 5211 Nợ / 3351 Có",
         "1. Kiểm type từng dòng",
         "—",
         "- 5211 = TYPE_DEPT (Nợ); 3351 (đối ứng 5211) = TYPE_HAS (Có); 3351→35241: 3351 Nợ, 35241 Có",
         "createDataSaveDept type"),
    ]),
]

# ===================== STYLES + BUILD =====================
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True); DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF"); TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79"); SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':24,'B':30,'C':16,'D':46,'E':10,'F':36,'G':46,'H':24,'I':62,'J':36,'K':18,'L':16,'M':16,'N':16,'O':22}
wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)
t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1); ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center"); ws.row_dimensions[11].height = 28
for r, label, formula in [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8
for i, h in enumerate(["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
        "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế","trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"], start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36
current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2): ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data, expected, business_note, "",
              "Not Executed","Not Executed","Not Executed",""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(220, longest // 3)); current_row += 1; data_row_idx += 1
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp, note, group=title)
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE} | data rows 18-{current_row-1}")
