"""Generate testcase Excel cho feature: Cấu hình chặn luồng quá hạn — Tab 2 'Chặn trưởng phòng duyệt'."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/cau-hinh-chan-tp-duyet/testcase.xlsx"
SHEET_NAME   = "CauHinhChanTP"
FEATURE_NAME = "Cấu hình chặn luồng quá hạn — Chặn trưởng phòng duyệt"
MODULE_NAME  = "Cấu hình chặn quá hạn"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Màn admin/due_configs/edit › Tab 2 'Chặn trưởng phòng duyệt'. Cấu hình BẬT/TẮT việc chặn trưởng phòng (TP) duyệt từng loại phiếu, THEO TỪNG CÔNG TY, khi phòng ban do user quản lý có nhân viên QUÁ HẠN hàng mượn / hàng giữ / hàng NXT. Tick = bật chặn; bỏ tích = tắt."),
    ("2. Đối tượng được tính / hiển thị",
     "► 31 mục duyệt (due_configs tab=2, group 'Chặn trưởng phòng duyệt', id 22–52): vd Duyệt yêu cầu xuất giữ/xuất hàng/xuất bán hàng mượn, gia hạn/điều chuyển hàng giữ, gia hạn hàng mượn, nhập hàng, HĐ bán/mua, yêu cầu đặt hàng, đề nghị thanh toán, lắp đặt bàn giao, mua dịch vụ, quyết toán HĐ, các phiếu chấm công/giao việc/HRM...\n"
     "► Ma trận: sub-tab theo TỪNG CÔNG TY × lưới checkbox 31 mục.\n"
     "► Checkbox tick ⇔ có dòng company_due_configs (company_id, due_config_id)."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Tab 1 'Chặn nhân viên' (due_configs tab=1) — KHÔNG thuộc phạm vi testcase này.\n"
     "► Mục due_configs tab ≠ 2 không hiển thị ở Tab 2."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng (màn cấu hình, không có bộ lọc thời gian)."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "due_configs (tab=2) ──< company_due_configs (pivot: company_id ↔ due_config_id) >── companies.\n"
     "Lịch sử: due_config_histories (content = diff dạng '[Công ty] Tích/Bỏ tích: <group> - <name>', updated_by)."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Mỗi (company_id, due_config_id) là 1 dòng pivot duy nhất. Khi Lưu: so sánh tick mới vs cũ theo từng công ty → INSERT mục mới tích, DELETE mục bỏ tích. Mỗi thay đổi ghi 1 dòng vào history; company_id = 'undefined' bị bỏ qua."),
    ("7. Phân quyền cấp",
     "• Route due_configs.edit / due_configs.update KHÔNG gắn middleware checkPermission → truy cập qua menu 'Cấu hình chặn luồng hàng quá hạn' (gating ở cấp menu/role, không có permission code riêng cho màn này)."),
    ("8. Cách tính các ô thống kê",
     "— Không có ô thống kê (màn cấu hình)."),
    ("9. Ghi chú đọc bảng",
     "Tab 2 chia sub-tab theo công ty (tab công ty đầu active mặc định); mỗi công ty 1 lưới checkbox 2 cột + link 'Tích tất cả' / 'Bỏ tích tất cả'. Nút Lưu → POST JSON {config_companies:{company_id:[due_config_id...]}} → thành công thì reload sau ~1s. Bảng 'Lịch sử chỉnh sửa' hiện 50 bản ghi gần nhất."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Truy cập màn cấu hình qua menu", "P0",
         "User đăng nhập, có quyền vào menu cấu hình quá hạn",
         "1. Menu › 'Cấu hình chặn luồng hàng quá hạn'\n2. Quan sát URL + layout",
         "URL = /admin/due_configs/edit",
         "- Mở màn admin/due_configs/edit\n- Tiêu đề 'Cấu hình chặn luồng hàng quá hạn'\n- Có 2 main tab: '1. Chặn nhân viên', '2. Chặn trưởng phòng duyệt'\n- Có nút Lưu / Hủy + bảng 'Lịch sử chỉnh sửa'",
         "BR — route due_configs.edit không checkPermission; vào qua menu"),
        ("002", "Chuyển sang Tab 2 'Chặn trưởng phòng duyệt'", "P0",
         "Đang ở màn cấu hình (mặc định Tab 1)",
         "1. Click main tab '2. Chặn trưởng phòng duyệt'\n2. Quan sát",
         "—",
         "- Hiện dòng mô tả: 'Chặn trưởng phòng duyệt các phiếu dưới đây nếu phòng ban do user quản lý có nhân viên quá hạn hàng mượn / hàng giữ / hàng NXT'\n- Hiện sub-tab theo từng công ty\n- Tab công ty đầu tiên active",
         "activeMainTab = 2"),
        ("003", "Hiển thị đủ 31 mục duyệt trong lưới", "P0",
         "Tab 2, đã chọn 1 công ty; due_configs tab=2 có 31 mục (id 22–52)",
         "1. Quan sát lưới checkbox của 1 công ty",
         "—",
         "- Lưới 2 cột, mỗi mục 1 checkbox + nhãn = display_name || name\n- Đủ các mục: Duyệt yêu cầu xuất giữ, ...xuất hàng, ...xuất bán hàng mượn, ...nhập hàng, HĐ bán, HĐ mua, yêu cầu đặt hàng, đề nghị thanh toán, lắp đặt bàn giao, quyết toán HĐ, ...",
         "Nguồn: due_configs WHERE tab=2"),
        ("004", "Checkbox phản ánh đúng cấu hình hiện tại", "P0",
         "Công ty A đã có company_due_configs cho mục 'Duyệt yêu cầu xuất hàng' (id 26); chưa có 'Duyệt hợp đồng mua' (id 30)",
         "1. Mở Tab 2 › công ty A\n2. Quan sát trạng thái 2 checkbox",
         "Cty A: tick sẵn id 26, bỏ tick id 30",
         "- 'Duyệt yêu cầu xuất hàng' = ĐÃ tick\n- 'Duyệt hợp đồng mua' = CHƯA tick",
         "ng-checked = config_companies[c.id].includes(role.id)"),
        ("005", "Hiển thị bảng Lịch sử chỉnh sửa", "P1",
         "Đã từng có thao tác lưu cấu hình",
         "1. Cuộn xuống panel 'Lịch sử chỉnh sửa'",
         "—",
         "- Bảng cột: STT, Thời gian (d/m/Y H:i), Nội dung chỉnh sửa, Người cập nhật\n- ≤ 50 dòng gần nhất, mới nhất trên cùng\n- Nếu chưa có: hiện 'Chưa có lịch sử'",
         "limit(50) orderBy created_at desc"),
        ("006", "Sub-tab nhiều công ty độc lập", "P1",
         "Hệ thống có ≥ 2 công ty",
         "1. Tab 2 › click lần lượt sub-tab công ty A, B\n2. Quan sát trạng thái tick mỗi công ty",
         "Cty A, Cty B cấu hình khác nhau",
         "- Mỗi công ty có lưới checkbox riêng, trạng thái tick theo company_due_configs của chính công ty đó",
         "config_companies[c.id] riêng từng công ty"),
        ("007", "Nhãn mục hiển thị display_name (fallback name)", "P2",
         "due_configs tab=2 có mục có display_name và mục không có",
         "1. Quan sát nhãn các checkbox",
         "—",
         "- Nhãn = display_name nếu có; nếu trống thì hiển thị name",
         "role.display_name || role.name"),
        ("008", "Nút Hủy rời màn cấu hình", "P2",
         "Đang ở màn cấu hình",
         "1. Click nút 'Hủy'",
         "—",
         "- Điều hướng về trang chủ (href='/'), không lưu thay đổi đang chọn dở",
         "Nút Hủy = link '/'"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH (tick / lưu cấu hình)", [
        ("001", "Tick 1 mục rồi Lưu → tạo cấu hình + ghi history", "P0",
         "Cty A CHƯA tick 'Duyệt yêu cầu xuất hàng' (id 26)",
         "1. Tab 2 › Cty A\n2. Tick 'Duyệt yêu cầu xuất hàng'\n3. Lưu",
         "Cty A, tick id 26",
         "- Toast 'Thao tác thành công!', reload\n- company_due_configs THÊM dòng (company_id=A, due_config_id=26)\n- Lịch sử thêm dòng '[Cty A] Tích: Chặn trưởng phòng duyệt - Duyệt yêu cầu xuất hàng'",
         "update(): new_configs = diff(new, old) → insert + history"),
        ("002", "Bỏ tích 1 mục rồi Lưu → xóa cấu hình + ghi history", "P0",
         "Cty A ĐANG tick 'Duyệt hợp đồng mua' (id 30)",
         "1. Tab 2 › Cty A\n2. Bỏ tích 'Duyệt hợp đồng mua'\n3. Lưu",
         "Cty A, bỏ tick id 30",
         "- company_due_configs XÓA dòng (A, 30)\n- Lịch sử thêm '[Cty A] Bỏ tích: Chặn trưởng phòng duyệt - Duyệt hợp đồng mua'",
         "update(): delete_configs = diff(old, new) → delete + history"),
        ("003", "'Tích tất cả' cho 1 công ty rồi Lưu", "P0",
         "Cty A đang tick 1 phần",
         "1. Tab 2 › Cty A › click 'Tích tất cả'\n2. Lưu",
         "Cty A",
         "- Toàn bộ 31 checkbox được tick\n- Sau lưu: company_due_configs(A) có đủ 31 due_config_id (22–52)\n- Lịch sử ghi các mục mới được tích",
         "toggleAllTab2(c.id, true)"),
        ("004", "'Bỏ tích tất cả' cho 1 công ty rồi Lưu", "P0",
         "Cty A đang tick nhiều mục",
         "1. Tab 2 › Cty A › click 'Bỏ tích tất cả'\n2. Lưu",
         "Cty A",
         "- Toàn bộ checkbox bỏ tick\n- Sau lưu: company_due_configs(A) tab2 rỗng\n- Lịch sử ghi các mục bị bỏ tích",
         "toggleAllTab2(c.id, false)"),
        ("005", "Tick nhiều mục ở nhiều công ty trong 1 lần Lưu", "P1",
         "Cty A, Cty B đều có thay đổi",
         "1. Cty A: tick 'Duyệt yêu cầu nhập hàng'\n2. Cty B: tick 'Duyệt quyết toán HĐ'\n3. Lưu",
         "A: id 28; B: id 36",
         "- 1 lần lưu cập nhật cả 2 công ty\n- pivot thêm (A,28) và (B,36)\n- Lịch sử ghi 2 dòng tương ứng từng công ty",
         "update() lặp theo từng company_id"),
        ("006", "Trạng thái sau Lưu giữ đúng khi reload", "P0",
         "Vừa tick 'Duyệt yêu cầu xuất giữ' (id 22) cho Cty A và Lưu",
         "1. Sau khi màn reload\n2. Mở lại Tab 2 › Cty A",
         "Cty A, id 22",
         "- 'Duyệt yêu cầu xuất giữ' vẫn hiện ĐÃ tick (đọc lại từ company_due_configs)",
         "edit() nạp lại config_companies từ pivot"),
        ("007", "Toggle qua lại 1 mục rồi Lưu = không đổi", "P2",
         "Cty A đang tick id 26",
         "1. Bỏ tích id 26 rồi tích lại id 26\n2. Lưu",
         "Cty A, id 26 (net không đổi)",
         "- Trạng thái cuối = đang tick → so với cũ không khác → KHÔNG insert/delete\n- Không ghi history (changes rỗng)",
         "diff old/new = ∅"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Lưu khi KHÔNG thay đổi gì → success, không ghi history", "P0",
         "Cty A,B cấu hình giữ nguyên",
         "1. Mở Tab 2, không tick/bỏ tích gì\n2. Lưu",
         "—",
         "- Toast 'Thao tác thành công!'\n- KHÔNG có dòng history mới (changes rỗng)\n- pivot không đổi",
         "if(!empty(changes)) mới ghi history"),
        ("002", "config_companies rỗng / không gửi", "P1",
         "—",
         "1. Gửi update với config_companies = null/[]",
         "config_companies = null",
         "- Trả success = true, message 'Thao tác thành công!'\n- Không lỗi, không đổi dữ liệu",
         "if($request->config_companies) bao toàn bộ xử lý"),
        ("003", "Bỏ qua company_id = 'undefined'", "P1",
         "Payload lẫn key 'undefined'",
         "1. Gửi config_companies có key 'undefined'",
         "company_id='undefined'",
         "- Bỏ qua key 'undefined' (continue), các công ty hợp lệ vẫn xử lý bình thường",
         "if($company_id=='undefined') continue;"),
        ("004", "Lưu lần 2 liên tiếp sau khi đã lưu", "P2",
         "Vừa lưu xong 1 thay đổi",
         "1. Lưu lại ngay (không đổi gì)",
         "—",
         "- Lần 2: success, không ghi history (đã đồng bộ)",
         "Idempotent khi không đổi"),
        ("005", "Thiếu/Sai CSRF token", "P1",
         "—",
         "1. POST update không kèm X-CSRF-TOKEN hợp lệ",
         "—",
         "- Bị chặn (419/error), toast 'Đã có lỗi xảy ra'\n- Không thay đổi dữ liệu",
         "Laravel CSRF"),
        ("006", "Mục cấu hình tab=1 không lẫn vào Tab 2", "P1",
         "due_configs có cả tab=1 và tab=2",
         "1. Quan sát lưới Tab 2",
         "—",
         "- Chỉ hiện mục tab=2 (configsTab2 = @json configs_tab2 WHERE tab=2); không hiện mục tab=1",
         "Cô lập tab"),
        ("007", "Reload màn → mặc định main tab = Tab 1", "P2",
         "Vừa thao tác ở Tab 2 rồi Lưu (màn reload)",
         "1. Sau reload, quan sát main tab đang active",
         "—",
         "- Main tab mặc định về '1. Chặn nhân viên' (activeMainTab khởi tạo = 1); cần click lại để vào Tab 2",
         "$scope.activeMainTab = 1 khi init"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Cấu hình công ty A KHÔNG ảnh hưởng công ty B", "P0",
         "Cty A tick id 26; Cty B chưa tick id 26",
         "1. Tick thêm vài mục cho Cty A, Lưu\n2. Mở Cty B",
         "A đổi, B giữ nguyên",
         "- company_due_configs của B không bị thêm/xóa do thao tác trên A\n- Lưới Cty B giữ nguyên trạng thái",
         "update() xử lý riêng theo từng company_id"),
        ("002", "Insert pivot đúng company_id đang thao tác", "P1",
         "Tick id 31 cho Cty A",
         "1. Tick 'Duyệt yêu cầu đặt hàng' Cty A, Lưu\n2. Kiểm DB",
         "A, id 31",
         "- Chỉ tạo dòng (company_id=A, due_config_id=31); không tạo cho công ty khác",
         "rows[] = {due_config_id, company_id} theo đúng vòng lặp"),
        ("003", "Lịch sử ghi đúng người + công ty thao tác", "P2",
         "User TP-X (phòng KD) thực hiện lưu",
         "1. Lưu 1 thay đổi\n2. Xem panel lịch sử",
         "User: TP-X",
         "- Dòng history: updated_by = TP-X; cột Người cập nhật hiện '<mã phòng> - <họ tên>'\n- Nội dung chứa tên công ty bị đổi",
         "DueConfigHistory.updated_by = auth user"),
    ]),
    ("VIII", "E2E — Cấu hình tác động tới chặn duyệt", [
        ("001", "Bật chặn 1 phiếu → TP có NV quá hạn bị chặn duyệt phiếu đó", "P0",
         "Cty X: tick 'Duyệt yêu cầu xuất hàng' (id 26). User TP quản lý phòng có NV đang quá hạn hàng giữ/mượn/NXT.",
         "1. Lưu cấu hình bật id 26 cho Cty X\n2. TP mở duyệt 'Yêu cầu xuất hàng'",
         "Cty X, id 26 bật; phòng TP có NV quá hạn",
         "- TP BỊ CHẶN duyệt yêu cầu xuất hàng (modal/thông báo NV quá hạn)\n- Liên kết: company_due_configs(tab2) → DueConfigBlockService::isManagerBlocked",
         "Config tab2 drive middleware checkDueConfigsManager"),
        ("002", "Tắt chặn (bỏ tích) → TP duyệt được lại", "P0",
         "Cty X đang bật id 26; phòng TP vẫn có NV quá hạn",
         "1. Bỏ tích 'Duyệt yêu cầu xuất hàng' Cty X, Lưu\n2. TP duyệt yêu cầu xuất hàng",
         "Cty X, bỏ id 26",
         "- TP KHÔNG còn bị chặn ở phiếu này (không còn pivot)",
         "Bỏ pivot → isManagerBlocked = false cho action đó"),
        ("003", "Chỉ chặn đúng phiếu đã bật, phiếu chưa bật không chặn", "P1",
         "Cty X bật 'Duyệt yêu cầu xuất hàng' (26), KHÔNG bật 'Duyệt hợp đồng mua' (30); TP có NV quá hạn",
         "1. TP duyệt yêu cầu xuất hàng\n2. TP duyệt hợp đồng mua",
         "Cty X: 26 bật, 30 tắt",
         "- Duyệt xuất hàng: BỊ chặn\n- Duyệt hợp đồng mua: KHÔNG bị chặn",
         "Chặn theo từng due_config được tick"),
        ("004", "TP không quản lý phòng có NV quá hạn → không bị chặn", "P1",
         "Cty X bật id 26; nhưng phòng do TP quản lý KHÔNG có NV quá hạn",
         "1. TP duyệt yêu cầu xuất hàng",
         "Phòng TP: không có NV quá hạn",
         "- KHÔNG bị chặn dù cấu hình bật (điều kiện cần: có NV quá hạn)",
         "isManagerBlocked cần getOverdueEmployees > 0"),
        ("005", "Cấu hình theo công ty của TP", "P2",
         "Cty X bật id 26; Cty Y KHÔNG bật. TP thuộc Cty Y.",
         "1. TP (Cty Y) duyệt yêu cầu xuất hàng",
         "TP ∈ Cty Y; chỉ Cty X bật",
         "- TP Cty Y KHÔNG bị chặn (cấu hình áp theo công ty của user)",
         "company_due_configs lọc theo company của user"),
        ("006", "Cả 3 loại quá hạn (giữ / mượn / NXT) đều trigger chặn", "P1",
         "Cty X bật 'Duyệt yêu cầu xuất hàng' (26). Lần lượt: phòng TP có NV quá hạn (a) hàng giữ, (b) hàng mượn, (c) hàng NXT.",
         "1. Với từng loại quá hạn, TP duyệt yêu cầu xuất hàng",
         "3 ca: giữ / mượn / NXT",
         "- Cả 3 loại quá hạn đều khiến TP bị chặn (chỉ cần có ≥1 NV quá hạn 1 trong 3 loại)",
         "DueConfigBlockService gộp 3 loại quá hạn"),
    ]),
]

# ===================== STYLES =====================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':22,'C':16,'D':42,'E':10,'F':32,'G':55,'H':22,'I':65,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1); ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center"); ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8

HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
           "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế","trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2): ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data, expected, business_note, "",
              "Not Executed","Not Executed","Not Executed",""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, longest // 3))
    current_row += 1; data_row_idx += 1

ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE}")
print(f"Rows: 1-10 desc, 11-15 summary, 17 header, 18-{current_row-1} data")
