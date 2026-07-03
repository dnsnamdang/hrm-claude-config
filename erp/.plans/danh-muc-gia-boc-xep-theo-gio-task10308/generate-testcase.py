"""Testcase Excel — Danh mục giá bốc xếp: thêm đơn giá công ty & thuê ngoài theo GIỜ (task_10308)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/danh-muc-gia-boc-xep-theo-gio-task10308/testcase.xlsx"
SHEET_NAME   = "DanhMucGiaBocXep"
FEATURE_NAME = "Danh mục giá bốc xếp — đơn giá công ty & thuê ngoài theo giờ (task_10308)"
MODULE_NAME  = "Danh mục giá bốc xếp"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Màn Danh mục giá bốc xếp (ArrangeGood — common.delivery_arrange). Bổ sung cấu hình đơn giá THEO GIỜ: 'Đơn giá công ty / Giờ' (company_price_hour) và 'Đơn giá thuê ngoài / Giờ' (price_hour), bên cạnh đơn giá theo TẤN cũ ('Đơn giá công ty / Tấn' = company_price, 'Đơn giá thuê ngoài / Tấn' = price). Mỗi định mức chọn loại tính: Theo trọng lượng (Tấn) và/hoặc Theo thời gian (Giờ)."),
    ("2. Đối tượng được tính / hiển thị",
     "► Modal thêm/sửa: 2 checkbox is_weight (Theo Tấn) / is_time (Theo Giờ).\n"
     "  • Khối is_weight: 'Đơn giá công ty / Tấn' (company_price) + 'Đơn giá thuê ngoài / Tấn' (price).\n"
     "  • Khối is_time (MỚI): 'Đơn giá công ty / Giờ' (company_price_hour) + 'Đơn giá thuê ngoài / Giờ' (price_hour).\n"
     "► Danh sách: cột giá hiển thị dạng 'tấn/giờ' — price = '<price>/<price_hour>', company_price = '<company_price>/<company_price_hour>'.\n"
     "► Phải chọn ≥ 1 loại tính giá."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Khi không bật is_weight: khối đơn giá / Tấn ẩn; lưu price=0, company_price=0.\n"
     "► Khi không bật is_time: khối đơn giá / Giờ ẩn; lưu price_hour=0, company_price_hour=0."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng (màn danh mục). Có báo cáo bốc xếp theo NCC/nhân viên riêng (ngoài phạm vi)."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "arrange_goods: type_name, is_weight, is_time, price (thuê ngoài/tấn), company_price (công ty/tấn), price_hour (thuê ngoài/giờ), company_price_hour (công ty/giờ), status, created_by/updated_by."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Mỗi định mức là 1 bản ghi độc lập. Lưu: chỉ giữ đơn giá của (các) loại đã bật; loại không bật set 0. Không cộng dồn."),
    ("7. Phân quyền cấp",
     "• Theo quyền tạo/sửa danh mục giá bốc xếp (canCreate ở index). Route arrangeGood.* (index/store/edit/delete + báo cáo)."),
    ("8. Cách tính các ô thống kê",
     "— Không có ô thống kê. Cột giá list = ghép 'tấn/giờ'."),
    ("9. Ghi chú đọc bảng",
     "Angular <% %>. Modal: khối Tấn ng-show is_weight, khối Giờ ng-show is_time. Validation BE: ArrangeDeliveryRequest (price/company_price required_if is_weight + not_in:0; price_hour/company_price_hour required_if is_time; ≥1 loại). Lỗi hiển thị inline từng field."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & MODAL", [
        ("001", "Truy cập Danh mục giá bốc xếp", "P0",
         "User có quyền; đã đăng nhập",
         "1. Vào menu Danh mục giá bốc xếp (arrangeGood.index)",
         "—",
         "- Hiện danh sách định mức + nút Thêm (nếu canCreate)",
         "ArrangeGoodController@index"),
        ("002", "Mở modal Thêm định mức", "P0",
         "Đang ở danh sách",
         "1. Click Thêm",
         "—",
         "- Modal hiện: 2 checkbox 'Theo trọng lượng'/'Theo thời gian'; khối đơn giá ẩn/hiện theo loại",
         "modalPopup is_weight/is_time"),
        ("003", "Khối 'Theo Tấn' hiện khi bật is_weight", "P0",
         "Modal mở",
         "1. Bật 'Theo trọng lượng'",
         "—",
         "- Hiện 'Đơn giá công ty / Tấn' + 'Đơn giá thuê ngoài / Tấn'",
         "ng-show form.is_weight"),
        ("004", "Khối 'Theo Giờ' hiện khi bật is_time (MỚI)", "P0",
         "Modal mở",
         "1. Bật 'Theo thời gian'",
         "—",
         "- Hiện 'Đơn giá công ty / Giờ' (company_price_hour) + 'Đơn giá thuê ngoài / Giờ' (price_hour)",
         "ng-show form.is_time"),
        ("005", "Nhập tên/loại bốc xếp (type_name)", "P1",
         "Modal mở",
         "1. Nhập Loại bốc xếp\n2. Quan sát",
         "—",
         "- Có ô nhập type_name (loại/tên định mức bốc xếp) lưu cùng đơn giá",
         "store: type_name"),
    ]),
    ("IV", "DANH SÁCH ĐỊNH MỨC", [
        ("001", "Cột giá hiển thị dạng 'tấn/giờ'", "P0",
         "Định mức: price=100k, price_hour=80k; company_price=90k, company_price_hour=70k",
         "1. Xem danh sách",
         "—",
         "- Cột 'Đơn giá thuê ngoài' = '100.000/80.000'\n- Cột 'Đơn giá công ty' = '90.000/70.000'",
         "editColumn: tan.'/'.gio (number_format)"),
        ("002", "Định mức chỉ Tấn hiển thị giờ = 0", "P1",
         "Định mức is_weight=1, is_time=0: price=100k, price_hour=0",
         "1. Xem danh sách",
         "—",
         "- Cột giá = '100.000/0'",
         "price_hour=0 khi !is_time"),
        ("003", "Định mức chỉ Giờ hiển thị tấn = 0", "P1",
         "is_time=1, is_weight=0: price_hour=80k, price=0",
         "1. Xem danh sách",
         "—",
         "- Cột giá = '0/80.000'",
         "price=0 khi !is_weight"),
        ("004", "Cột người tạo/sửa + thời gian + trạng thái", "P2",
         "Định mức đã tạo",
         "1. Xem danh sách",
         "—",
         "- Hiện created_by/updated_by, created_at/updated_at (định dạng ngày), status, cột thao tác (sửa/xóa)",
         "editColumn created/updated/status/action"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH — CRUD định mức", [
        ("001", "Thêm định mức theo Tấn (đơn giá công ty + thuê ngoài / Tấn)", "P0",
         "Modal mở",
         "1. Bật 'Theo trọng lượng'\n2. Đơn giá công ty/Tấn=90k, thuê ngoài/Tấn=100k\n3. Lưu",
         "company_price=90k, price=100k",
         "- Lưu thành công; arrange_goods: is_weight=1, company_price=90k, price=100k; price_hour/company_price_hour=0",
         "store: !is_time → giờ=0"),
        ("002", "Thêm định mức theo Giờ (đơn giá công ty + thuê ngoài / Giờ) — MỚI", "P0",
         "Modal mở",
         "1. Bật 'Theo thời gian'\n2. Đơn giá công ty/Giờ=70k, thuê ngoài/Giờ=80k\n3. Lưu",
         "company_price_hour=70k, price_hour=80k",
         "- Lưu thành công; arrange_goods: is_time=1, company_price_hour=70k, price_hour=80k; price/company_price=0",
         "store: !is_weight → tấn=0"),
        ("003", "Thêm định mức theo CẢ Tấn và Giờ", "P0",
         "Modal mở",
         "1. Bật cả 'Theo trọng lượng' + 'Theo thời gian'\n2. Nhập đủ 4 đơn giá\n3. Lưu",
         "4 đơn giá",
         "- Lưu đủ 4: price, company_price, price_hour, company_price_hour; is_weight=is_time=1",
         "store giữ cả 2 khối"),
        ("004", "Sửa định mức (đổi đơn giá theo giờ)", "P1",
         "Có định mức theo Giờ price_hour=80k",
         "1. Mở sửa\n2. Đổi thuê ngoài/Giờ=90k\n3. Lưu",
         "price_hour 80k→90k",
         "- Cập nhật price_hour=90k",
         "arrangeGood.edit"),
        ("005", "Đổi loại Tấn → Giờ khi sửa", "P1",
         "Định mức đang theo Tấn",
         "1. Tắt 'Theo trọng lượng', bật 'Theo thời gian'\n2. Nhập đơn giá/Giờ\n3. Lưu",
         "—",
         "- Sau lưu: price/company_price=0, price_hour/company_price_hour có giá trị",
         "store set 0 khi tắt loại"),
        ("006", "Xóa định mức", "P1",
         "Có định mức",
         "1. Click Xóa\n2. Xác nhận",
         "—",
         "- Định mức bị xóa khỏi danh sách",
         "arrangeGood.delete"),
        ("007", "Sửa định mức 2 loại → bỏ 1 loại → đơn giá loại bỏ về 0", "P1",
         "Định mức đang bật cả Tấn + Giờ (đủ 4 đơn giá)",
         "1. Tắt 'Theo thời gian'\n2. Lưu",
         "tắt is_time",
         "- price_hour=0, company_price_hour=0; đơn giá Tấn giữ nguyên",
         "store: !is_time → giờ=0"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Không chọn loại tính nào", "P0",
         "Modal: bỏ cả is_weight và is_time",
         "1. Lưu",
         "is_weight=0,is_time=0",
         "- Lỗi 'Phải chọn ít nhất 1 loại tính giá' (gắn field is_weight); không lưu",
         "withValidator !is_weight && !is_time"),
        ("002", "[Tấn] Thiếu đơn giá công ty / Tấn", "P0",
         "is_weight=1, company_price trống",
         "1. Lưu",
         "company_price rỗng",
         "- Lỗi: company_price required_if is_weight=1 (+ not_in:0)",
         "company_price required_if:is_weight,1|not_in:0"),
        ("003", "[Tấn] Thiếu đơn giá thuê ngoài / Tấn", "P0",
         "is_weight=1, price trống/0",
         "1. Lưu",
         "price rỗng/0",
         "- Lỗi: price required_if is_weight=1 + not_in:0 (phải > 0)",
         "price required_if:is_weight,1|not_in:0"),
        ("004", "[Giờ] Thiếu đơn giá công ty / Giờ", "P0",
         "is_time=1, company_price_hour trống",
         "1. Lưu",
         "company_price_hour rỗng",
         "- Lỗi: company_price_hour required_if is_time=1",
         "company_price_hour required_if:is_time,1|numeric|min:0"),
        ("005", "[Giờ] Thiếu đơn giá thuê ngoài / Giờ", "P0",
         "is_time=1, price_hour trống",
         "1. Lưu",
         "price_hour rỗng",
         "- Lỗi: price_hour required_if is_time=1",
         "price_hour required_if:is_time,1|numeric|min:0"),
        ("006", "Đơn giá âm / không phải số", "P1",
         "—",
         "1. company_price_hour = -5 hoặc 'abc'\n2. Lưu",
         "-5 / abc",
         "- Lỗi numeric/min:0",
         "numeric|min:0"),
        ("007", "Chỉ nhập đơn giá của loại KHÔNG bật → bị bỏ (set 0)", "P2",
         "is_weight=1 (chỉ Tấn) nhưng có nhập tạm price_hour",
         "1. Lưu",
         "—",
         "- price_hour/company_price_hour lưu = 0 (controller ép 0 khi !is_time), không lỗi",
         "if(!is_time) giờ=0"),
    ]),
    ("VII", "STORE LOGIC & CÔ LẬP", [
        ("001", "Chỉ Tấn → đơn giá Giờ = 0", "P0",
         "is_weight=1, is_time=0, nhập đủ Tấn",
         "1. Lưu\n2. Kiểm DB",
         "—",
         "- price_hour=0, company_price_hour=0",
         "store !is_time"),
        ("002", "Chỉ Giờ → đơn giá Tấn = 0", "P0",
         "is_time=1, is_weight=0, nhập đủ Giờ",
         "1. Lưu\n2. Kiểm DB",
         "—",
         "- price=0, company_price=0",
         "store !is_weight"),
        ("003", "Mở lại sửa giữ đúng loại + đơn giá", "P1",
         "Định mức theo Giờ company_price_hour=70k, price_hour=80k",
         "1. Mở sửa",
         "—",
         "- Modal bật 'Theo thời gian', khối Giờ hiện đúng 70k/80k; khối Tấn ẩn",
         "khôi phục is_weight/is_time"),
    ]),
    ("VIII", "E2E — Định mức dùng ở phiếu", [
        ("001", "Định mức theo Giờ dùng được ở Tab bốc xếp phiếu", "P0",
         "Có định mức theo Giờ price_hour=80k",
         "1. Tạo Phiếu xuất kho/xuất hàng › Tab bốc xếp\n2. Chọn loại bốc xếp = định mức này, theo Giờ",
         "—",
         "- Đơn giá/giờ tự lấy price_hour=80k; thành tiền = 80k × số giờ",
         "arrange_type → set price theo loại"),
        ("002", "Định mức công ty vs thuê ngoài tách biệt", "P1",
         "Định mức company_price_hour=70k, price_hour=80k",
         "1. Dùng định mức ở phiếu, chọn rent_type (công ty/thuê ngoài)",
         "—",
         "- Theo rent_type áp đúng đơn giá: công ty=70k/giờ hoặc thuê ngoài=80k/giờ",
         "company_price_hour vs price_hour theo rent_type"),
        ("003", "Báo cáo bốc xếp theo NCC phản ánh đơn giá", "P2",
         "Có phiếu dùng bốc xếp thuê ngoài",
         "1. Mở báo cáo bốc xếp theo NCC",
         "—",
         "- Hiển thị chi phí bốc xếp theo NCC đúng đơn giá đã cấu hình",
         "reportArrangeGoodBySupplier"),
        ("004", "Đổi đơn giá định mức → phiếu lập MỚI dùng giá mới", "P2",
         "Định mức price_hour=80k; đã có phiếu cũ dùng giá 80k",
         "1. Sửa định mức price_hour=90k\n2. Lập phiếu mới chọn định mức theo Giờ",
         "80k → 90k",
         "- Phiếu mới lấy 90k/giờ; phiếu cũ giữ 80k (đã lưu tại thời điểm lập)",
         "đơn giá snapshot vào phiếu"),
    ]),
]

# ===================== STYLES + BUILD =====================
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True); DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF"); TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79"); SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':26,'C':16,'D':46,'E':10,'F':32,'G':52,'H':22,'I':62,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}
wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)
t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1); ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center"); ws.row_dimensions[11].height = 28
for r, label, formula in [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8
for i, h in enumerate(["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
        "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế","trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"], start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36
current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2): ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data, expected, business_note, "",
              "Not Executed","Not Executed","Not Executed",""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(210, longest // 3)); current_row += 1; data_row_idx += 1
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp, note, group=title)
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE} | data rows 18-{current_row-1}")
