"""Testcase Excel — Tab bốc xếp: tính chi phí vận chuyển theo thời gian (task_10308). Phiếu xuất kho + xuất hàng."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/boc-xep-chi-phi-theo-gio-task10308/testcase.xlsx"
SHEET_NAME   = "BocXepTheoGio"
FEATURE_NAME = "Tab bốc xếp — chi phí vận chuyển theo thời gian (task_10308)"
MODULE_NAME  = "Bốc xếp (xuất kho/xuất hàng)"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Cập nhật logic Tab BỐC XẾP trên Phiếu xuất kho (warehouse_exports) và Phiếu xuất hàng (product_exports): bổ sung cách tính chi phí bốc xếp/vận chuyển THEO THỜI GIAN (số giờ × đơn giá/giờ), bên cạnh cách cũ theo TRỌNG LƯỢNG (số tấn × đơn giá/tấn). Mỗi dòng bốc xếp chọn 1 trong 2 loại tính (loại trừ nhau)."),
    ("2. Đối tượng được tính / hiển thị",
     "► Tab bốc xếp có 2 checkbox loại trừ: is_weight (theo trọng lượng) / is_time (theo thời gian). Bật cái này tự tắt cái kia; phải chọn ≥ 1.\n"
     "► is_weight: nhãn 'Khối lượng hàng (Tấn)', đơn giá price (đ/tấn), company_price.\n"
     "► is_time: nhãn 'Số giờ', đơn giá price_hour (đ/giờ), company_price_hour.\n"
     "► Thành tiền (total_money) = đơn giá áp dụng × số lượng (tấn hoặc giờ). + VAT (vat_price = total × vat%), thành tiền sau VAT.\n"
     "► Định mức bốc xếp (ArrangeGood): cấu hình is_weight/is_time + price/price_hour + company_price/company_price_hour; cột giá hiển thị dạng 'tấn/giờ'."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Dòng không chọn loại tính nào (is_weight=0 & is_time=0) → bị chặn lưu.\n"
     "► Khi is_weight: bỏ qua price_hour/company_price_hour (=0). Khi is_time: bỏ qua price/company_price (=0)."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng (đây là tính chi phí trên phiếu, không phải bộ lọc báo cáo)."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "arrange_goods (định mức): + is_weight, is_time, company_price_hour, price_hour (ngoài price, company_price).\n"
     "arrange_deliveries (dòng bốc xếp trên phiếu): + is_weight; lưu is_weight/is_time, weight(=tấn hoặc giờ), price, total_money, vat...\n"
     "Gắn phiếu: ProductImportExportArrangeDelivery / WarehouseImportExportArrangeDelivery."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Mỗi dòng bốc xếp tính độc lập: total_money = đơn giá × số lượng. Tab bốc xếp có thể nhiều dòng; tổng = cộng các dòng. is_weight/is_time loại trừ trên từng dòng."),
    ("7. Phân quyền cấp",
     "• Theo quyền lập/sửa Phiếu xuất kho / Phiếu xuất hàng tương ứng (không có permission code riêng cho tab bốc xếp)."),
    ("8. Cách tính các ô thống kê",
     "► total_money = is_weight ? (price × tấn) : (price_hour × giờ) [dùng chung field số lượng].\n"
     "► vat_price = total_money × vat / 100.\n"
     "► after_total_money = total_money + vat_price."),
    ("9. Ghi chú đọc bảng",
     "Angular <% %>. Toggle is_weight/is_time bằng 2 checkbox loại trừ; đổi loại sẽ gọi change() set lại đơn giá theo định mức. Validation BE: ArrangeDeliveryRequest (required_if + ≥1 loại). Áp dụng cho cả phiếu xuất kho lẫn phiếu xuất hàng (cùng component bốc xếp)."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TAB BỐC XẾP", [
        ("001", "Tab bốc xếp hiển thị trên Phiếu xuất kho", "P0",
         "Mở tạo/sửa Phiếu xuất kho (warehouse_exports)",
         "1. Vào Tab bốc xếp",
         "—",
         "- Hiện 2 checkbox: 'Theo trọng lượng' (is_weight), 'Theo thời gian' (is_time)\n- Có chọn loại thuê (rent_type), loại bốc xếp (arrange_type)",
         "task_10308: + is_time/price_hour"),
        ("002", "Tab bốc xếp hiển thị trên Phiếu xuất hàng", "P0",
         "Mở tạo/sửa Phiếu xuất hàng (product_exports)",
         "1. Vào Tab bốc xếp",
         "—",
         "- Tab bốc xếp giống phiếu xuất kho: 2 checkbox is_weight/is_time + đơn giá/số lượng tương ứng",
         "cùng component arrange_delivery"),
        ("003", "Nhãn số lượng đổi theo loại tính", "P0",
         "Tab bốc xếp",
         "1. Bật 'Theo trọng lượng' → xem nhãn\n2. Bật 'Theo thời gian' → xem nhãn",
         "—",
         "- is_weight: nhãn 'Khối lượng hàng (Tấn)'\n- is_time: nhãn 'Số giờ'",
         "ng-if is_weight ? Tấn : Số giờ"),
        ("004", "Mặc định loại tính khi mở mới", "P1",
         "Tạo phiếu mới",
         "1. Mở Tab bốc xếp",
         "—",
         "- Mặc định is_weight = 1 (theo trọng lượng), is_time = 0",
         "default is_weight=1"),
    ]),
    ("V", "CHỨC NĂNG — TÍNH CHI PHÍ BỐC XẾP", [
        ("001", "[Trọng lượng] Tính chi phí theo Tấn", "P0",
         "Tab bốc xếp; chọn 'Theo trọng lượng'; đơn giá price = 100.000 đ/tấn",
         "1. Bật is_weight\n2. Nhập Khối lượng = 5 (tấn)\n3. Quan sát thành tiền",
         "price=100.000, tấn=5",
         "- Thành tiền (total_money) = 100.000 × 5 = 500.000",
         "total_money = price × weight"),
        ("002", "[Thời gian] Tính chi phí theo Giờ", "P0",
         "Tab bốc xếp; chọn 'Theo thời gian'; đơn giá/giờ price_hour = 80.000",
         "1. Bật is_time\n2. Nhập Số giờ = 3\n3. Quan sát thành tiền",
         "price_hour=80.000, giờ=3",
         "- Thành tiền = 80.000 × 3 = 240.000 (tính theo thời gian)",
         "total_money = price_hour × số giờ"),
        ("003", "Toggle Theo trọng lượng ↔ Theo thời gian (loại trừ)", "P0",
         "Đang bật is_weight",
         "1. Bật 'Theo thời gian'",
         "—",
         "- is_time=1 thì is_weight tự về 0 (và ngược lại); chỉ 1 loại active",
         "is_weight ? is_time=0 (ng-change)"),
        ("004", "Chọn định mức bốc xếp tự set đơn giá", "P1",
         "Có định mức ArrangeGood: price=100k/tấn, price_hour=80k/giờ",
         "1. Chọn arrange_type (định mức)\n2. Bật theo giờ\n3. Quan sát đơn giá",
         "—",
         "- Đơn giá tự set theo loại: is_weight→price(100k), is_time→price_hour(80k); change() cập nhật",
         "change(arrange_type, rent_type)"),
        ("005", "VAT trên chi phí bốc xếp", "P1",
         "total_money=500.000; VAT=8%",
         "1. Nhập VAT=8\n2. Quan sát tiền VAT + sau VAT",
         "VAT=8",
         "- vat_price = 500.000 × 8/100 = 40.000\n- Thành tiền sau VAT = 540.000",
         "vat_price = total × vat/100; after = total + vat"),
        ("006", "Nhiều dòng bốc xếp (trọng lượng + thời gian)", "P1",
         "Tab bốc xếp cho thêm nhiều dòng",
         "1. Dòng1 theo Tấn=5×100k\n2. Dòng2 theo Giờ=3×80k\n3. Quan sát tổng",
         "2 dòng",
         "- Mỗi dòng tính độc lập; tổng chi phí = 500.000 + 240.000 = 740.000",
         "tổng = Σ total_money"),
        ("007", "[Xuất kho] Lưu phiếu với bốc xếp theo giờ", "P0",
         "Phiếu xuất kho; bốc xếp theo giờ 3×80k",
         "1. Nhập bốc xếp theo giờ\n2. Lưu phiếu",
         "giờ=3, price_hour=80k",
         "- Lưu thành công; arrange_deliveries lưu is_time=1, weight(số giờ)=3, price=80k, total_money=240.000",
         "submit_data is_weight/is_time"),
        ("008", "[Xuất hàng] Lưu phiếu với bốc xếp theo giờ", "P0",
         "Phiếu xuất hàng; bốc xếp theo giờ",
         "1. Nhập bốc xếp theo giờ\n2. Lưu phiếu",
         "—",
         "- Lưu thành công; dòng bốc xếp theo thời gian được lưu đúng",
         "product_exports arrange_delivery"),
        ("009", "Mở lại phiếu giữ đúng loại tính + số liệu", "P1",
         "Phiếu đã lưu bốc xếp theo giờ 3×80k",
         "1. Mở lại phiếu (show/edit)",
         "—",
         "- Tab bốc xếp hiện đúng: is_time bật, Số giờ=3, đơn giá/giờ=80k, thành tiền=240.000",
         "khôi phục is_weight/is_time"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Không chọn loại tính nào (is_weight=0 & is_time=0)", "P0",
         "Bỏ tích cả 2 loại (qua thao tác/payload)",
         "1. Lưu dòng bốc xếp khi cả is_weight=0 và is_time=0",
         "is_weight=0,is_time=0",
         "- Báo lỗi 'Phải chọn ít nhất 1 loại tính giá'; không lưu",
         "withValidator: !is_weight && !is_time"),
        ("002", "[Trọng lượng] Thiếu đơn giá price khi is_weight=1", "P0",
         "is_weight=1, price trống/0",
         "1. Lưu",
         "price rỗng/0",
         "- Lỗi: price required_if is_weight=1 + not_in:0 (đơn giá phải > 0)",
         "price required_if:is_weight,1|not_in:0"),
        ("003", "[Trọng lượng] company_price thiếu khi is_weight=1", "P1",
         "is_weight=1, company_price trống",
         "1. Lưu",
         "company_price rỗng",
         "- Lỗi: company_price required_if is_weight=1 + not_in:0",
         "company_price required_if:is_weight,1"),
        ("004", "[Thời gian] Thiếu price_hour khi is_time=1", "P0",
         "is_time=1, price_hour trống",
         "1. Lưu",
         "price_hour rỗng",
         "- Lỗi: price_hour required_if is_time=1",
         "price_hour required_if:is_time,1|numeric|min:0"),
        ("005", "[Thời gian] company_price_hour thiếu khi is_time=1", "P1",
         "is_time=1, company_price_hour trống",
         "1. Lưu",
         "company_price_hour rỗng",
         "- Lỗi: company_price_hour required_if is_time=1",
         "company_price_hour required_if:is_time,1"),
        ("006", "Đơn giá âm / không phải số", "P1",
         "—",
         "1. price = -100 hoặc 'abc'\n2. Lưu",
         "-100 / abc",
         "- Lỗi numeric/min:0",
         "numeric|min:0"),
        ("007", "Số lượng (tấn/giờ) = 0 → thành tiền 0", "P2",
         "is_weight=1, price=100k, tấn=0",
         "1. Quan sát thành tiền",
         "tấn=0",
         "- total_money = 0 (weight=0 → 0)",
         "total_money = weight ? price×weight : 0"),
    ]),
    ("VII", "ĐỊNH MỨC BỐC XẾP (ArrangeGood config)", [
        ("001", "Cấu hình định mức tính theo cả Tấn và Giờ", "P0",
         "Tạo/sửa định mức bốc xếp",
         "1. Bật is_weight + nhập price/company_price\n2. Bật is_time + nhập price_hour/company_price_hour\n3. Lưu",
         "price=100k, price_hour=80k",
         "- Lưu định mức với đủ 4 đơn giá; cột giá hiển thị dạng 'tấn/giờ' (price/price_hour)",
         "arrange_goods is_weight/is_time + price/price_hour"),
        ("002", "Định mức chỉ theo Tấn (is_weight=1, is_time=0)", "P1",
         "Định mức bật is_weight, tắt is_time",
         "1. Lưu",
         "is_time=0",
         "- price_hour, company_price_hour = 0 (controller set 0 khi !is_time)",
         "if(!is_time) price_hour=0"),
        ("003", "Định mức chỉ theo Giờ (is_time=1, is_weight=0)", "P1",
         "Định mức bật is_time, tắt is_weight",
         "1. Lưu",
         "is_weight=0",
         "- price, company_price = 0 (controller set 0 khi !is_weight)",
         "if(!is_weight) price=0"),
        ("004", "Định mức không chọn loại nào → chặn", "P0",
         "is_weight=0 & is_time=0",
         "1. Lưu định mức",
         "—",
         "- Báo lỗi 'Phải chọn ít nhất 1 loại tính giá'",
         "withValidator ArrangeDeliveryRequest"),
        ("005", "Hiển thị cột giá định mức 'tấn/giờ'", "P2",
         "Định mức price=100k, price_hour=80k",
         "1. Xem danh sách định mức",
         "—",
         "- Cột đơn giá = '100.000/80.000' (price/price_hour); company tương tự",
         "editColumn price: tan.'/'.gio"),
    ]),
    ("VIII", "E2E & CÔ LẬP", [
        ("001", "[Xuất kho] Tổng chi phí bốc xếp theo giờ vào tổng phiếu", "P0",
         "Phiếu xuất kho có bốc xếp theo giờ 3×80k + VAT 8%",
         "1. Lưu phiếu\n2. Xem tổng chi phí vận chuyển/bốc xếp + công nợ",
         "—",
         "- Chi phí bốc xếp = 240.000, sau VAT = 259.200 vào tổng phiếu/công nợ đúng",
         "total_money + vat cộng vào phiếu"),
        ("002", "Đổi số giờ → thành tiền cập nhật realtime", "P1",
         "is_time, price_hour=80k, giờ=3",
         "1. Đổi giờ = 5",
         "giờ 3→5",
         "- total_money cập nhật = 80k × 5 = 400.000",
         "binding tự tính"),
        ("003", "Đổi từ Giờ sang Tấn cập nhật đơn giá + thành tiền", "P1",
         "Đang theo giờ (price_hour=80k); định mức có price=100k",
         "1. Bật 'Theo trọng lượng'\n2. Nhập tấn=5",
         "—",
         "- Đơn giá đổi sang price=100k, thành tiền = 100k×5=500.000",
         "change() set lại price"),
        ("004", "Phiếu xuất hàng & xuất kho dùng chung logic", "P1",
         "Cùng định mức",
         "1. Tạo bốc xếp theo giờ trên cả 2 loại phiếu",
         "—",
         "- Cùng công thức + validation; kết quả nhất quán",
         "cùng component"),
        ("005", "Bốc xếp theo giờ không ảnh hưởng giá hàng hóa của phiếu", "P2",
         "Phiếu có hàng hóa + bốc xếp theo giờ",
         "1. Lưu phiếu",
         "—",
         "- Chi phí bốc xếp tách riêng, không cộng vào giá bán hàng hóa",
         "arrange_delivery độc lập tab hàng"),
    ]),
]

# ===================== STYLES + BUILD =====================
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True); DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF"); TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79"); SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':24,'B':26,'C':16,'D':46,'E':10,'F':32,'G':52,'H':22,'I':62,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)
t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1); ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center"); ws.row_dimensions[11].height = 28
for r, label, formula in [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8
for i, h in enumerate(["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
        "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế","trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"], start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36
current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2): ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data, expected, business_note, "",
              "Not Executed","Not Executed","Not Executed",""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(210, longest // 3)); current_row += 1; data_row_idx += 1
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp, note, group=title)
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE} | data rows 18-{current_row-1}")
