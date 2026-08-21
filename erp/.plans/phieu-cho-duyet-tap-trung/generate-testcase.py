"""Generate testcase Excel — Màn Phiếu chờ duyệt tập trung (danh sách)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = "testcase.xlsx"  # chạy trong thư mục .plans/phieu-cho-duyet-tap-trung/
SHEET_NAME   = "PhieuChoDuyetTapTrung"
FEATURE_NAME = "Phiếu chờ duyệt tập trung (màn danh sách)"
MODULE_NAME  = "Phiếu chờ duyệt"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Gom toàn bộ phiếu đang CHỜ DUYỆT của mọi luồng nghiệp vụ (49 loại) về 1 màn hình tập trung dạng "
     "\"hộp duyệt cá nhân\": mỗi user chỉ thấy phiếu MÀ MÌNH CÓ QUYỀN DUYỆT ở cấp hiện tại. "
     "Hỗ trợ lọc, tìm kiếm, phân trang, deep-link mở thẳng phiếu để duyệt, và xuất Excel theo bộ lọc."),
    ("2. Đối tượng được tính / hiển thị",
     "► Chỉ phiếu status = CHỜ DUYỆT (approval_inbox.status = 1 = PENDING).\n"
     "► Chỉ phiếu user hiện tại được duyệt, xác định bởi ApprovalInbox::pendingBuilderForUser:\n"
     "   - (a) Phiếu chỉ định người duyệt đích danh (approver_id NOT NULL) → CHỈ đúng người đó (approver_id = user.id). "
     "Ví dụ: BGĐ duyệt hợp đồng/quyết toán (chuyển duyệt gán valid_approver_id).\n"
     "   - (b) Phiếu CHƯA chỉ định (approver_id NULL) + user CÓ quyền = required_permission + khớp PHẠM VI của quyền đó:\n"
     "        · Quyền scope 'company' (mặc định — Ban kiểm soát, Kế toán, Kế toán trưởng, Thủ quỹ, BGĐ…): "
     "company_id NULL HOẶC = công ty user (KHÔNG check phòng ban).\n"
     "        · Quyền scope 'department' (16 quyền 'Trưởng phòng…' + 'Duyệt hợp đồng' + 3 quyền duyệt KH/chỉ tiêu phòng): "
     "department_id ∈ phòng user QUẢN LÝ (bảng employee_manage_departments).\n"
     "        · Quyền scope 'part': part_id ∈ bộ phận user quản lý (employee_manage_parts) — hiện chưa quyền nào dùng."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Phiếu đã duyệt/từ chối/huỷ (status 2=APPROVED, 3=REJECTED, 4=CANCELED) — không hiện (giữ trong DB cho lịch sử/KPI).\n"
     "► Phiếu thuộc phòng ban user KHÔNG quản lý (với quyền scope department).\n"
     "► Phiếu chỉ định đích danh cho người khác (approver_id = người khác).\n"
     "► Phiếu công ty khác (với quyền scope company, company_id ≠ công ty user và ≠ NULL).\n"
     "► Các luồng KHÔNG phải ký duyệt (đã loại khỏi registry): chuyến xe chờ hạch toán, YC sửa chữa chờ xử lý, "
     "cảnh báo/nhắc/hàng đợi (tài khoản chưa phân quyền, hàng mượn/giữ hết hạn…)."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "Bộ lọc 'Từ ngày' / 'Đến ngày' áp trên cột submitted_at (Thời gian gửi duyệt), so theo NGÀY (whereDate >= from, <= to). "
     "submitted_at = thời điểm phiếu gửi duyệt lần đầu (cố định, không đổi khi chuyển cấp) — live hook = thời điểm bấm gửi thật; "
     "phiếu backfill = created_at của phiếu nguồn."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "Registry: bảng approval_inbox (1 dòng = 1 phiếu đang ở 1 cấp duyệt) + approval_inbox_logs (log từng lượt duyệt/chuyển cấp cho KPI). "
     "Mỗi dòng lưu: doc_type, group_code, source_table/source_id (trỏ phiếu gốc), code, partner, requester, submitted_at, "
     "required_permission (quyền cấp hiện tại), company_id/department_id/part_id, approver_id, current_level, approve_route + approve_params (deep-link). "
     "Config single-source: config/approval_inbox.php (types: label/group/route/permissions; permission_scopes: quyền→scope)."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Mỗi phiếu nguồn = đúng 1 dòng registry (unique source_system+source_table+source_id). "
     "Luồng nhiều cấp: KHÔNG tạo dòng mới mỗi cấp — dòng cũ được advance() đổi required_permission+level (giữ nguyên submitted_at). "
     "Phiếu tái mở chu kỳ mới (reactivate) → round++. Đếm badge/‘Tất cả loại (N)’ = số dòng PENDING trong hộp duyệt user."),
    ("7. Phân quyền cấp",
     "• 'Xem phiếu chờ duyệt tập trung' — quyền TRUY CẬP màn danh sách + endpoint search-data/employees/export (gắn middleware checkPermission).\n"
     "• 'Xem báo cáo phê duyệt' — quyền vào màn Báo cáo phê duyệt (màn RIÊNG, sẽ test sau).\n"
     "• Phạm vi phiếu hiển thị KHÔNG do 1 quyền tổng — mà theo TỪNG required_permission của phiếu + scope trong permission_scopes "
     "(company/department/part) — xem mục 2. Quyền 'department' verify theo bảng employee_manage_departments; 'part' theo employee_manage_parts.\n"
     "• KHÔNG special-case Super Admin ở tầng màn (Super Admin đi qua Gate::before ở AuthServiceProvider → user()->can() trả true cho mọi quyền → thấy phiếu toàn công ty)."),
    ("8. Cách tính các ô thống kê",
     "► Badge menu 'Phiếu chờ duyệt (N)' = ApprovalInbox::pendingCountForUser(user) = tổng dòng PENDING trong hộp duyệt user.\n"
     "► Option 'Loại phiếu' → '— Tất cả loại (N) —' với N = total; mỗi loại 'Nhãn (n)' với n = số phiếu loại đó trong hộp duyệt user (typeCounts group by doc_type).\n"
     "► footInfo phân trang = 'Hiển thị {from}–{to} / {N} phiếu' (N = recordsFiltered theo bộ lọc)."),
    ("9. Ghi chú đọc bảng",
     "► 8 cột: STT | Loại phiếu (badge màu theo nhóm) | Mã phiếu (link ↗ mở TAB MỚI tới phiếu gốc) | Đối tác (tag KH/NCC + tên, "
     "dài quá cắt '…', hover title xem full) | Người yêu cầu (avatar + tên + PHÒNG BAN subtext) | Thời gian gửi (dd/mm/yyyy HH:mm) | "
     "Quyền áp dụng (tên quyền cấp hiện tại) | Hành động (nút Duyệt = deep-link mở phiếu).\n"
     "► Phân trang server-side: 20/50/100/200 (mặc định 50), STT liên tục giữa các trang, đổi bộ lọc → về trang 1.\n"
     "► Bộ lọc: Phòng ban → Bộ phận (cascade) → Người yêu cầu (select search, data theo phòng/bộ phận) → Loại phiếu → Từ/Đến ngày → Tìm kiếm (mã phiếu). "
     "ĐÃ BỎ: lọc Công ty (chỉ 1 cty user), lọc Quyền áp dụng. Danh mục Phòng/Bộ phận/Người yêu cầu GIỚI HẠN theo hộp duyệt user.\n"
     "► Màn full-width (không còn sidebar 'Lọc theo nhóm'). Menu cũ 'Chờ duyệt' đã comment lại, thay bằng menu 'Phiếu chờ duyệt'."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("00", "Truy cập màn khi CÓ quyền xem", "P0",
     "User đã đăng nhập; user CÓ quyền 'Xem phiếu chờ duyệt tập trung'",
     "1. Vào menu 'Phiếu chờ duyệt' > 'Danh sách phiếu chờ duyệt'\n2. Quan sát",
     "User: có quyền xem",
     "- Màn /admin/approval-inbox load thành công (HTTP 200)\n- Hiển thị bộ lọc + bảng danh sách + phân trang",
     "Permission: 'Xem phiếu chờ duyệt tập trung' (middleware checkPermission trên route index)"),
    ("01", "Chặn truy cập khi KHÔNG có quyền xem", "P0",
     "User đã đăng nhập; user KHÔNG có quyền 'Xem phiếu chờ duyệt tập trung'",
     "1. Gõ thẳng URL /admin/approval-inbox\n2. Quan sát",
     "User: không có quyền xem",
     "- Bị chặn bởi middleware checkPermission (403/redirect)\n- Menu 'Phiếu chờ duyệt' KHÔNG hiển thị (@can bọc menu)",
     "BR — Route index/search-data/employees/export đều gắn checkPermission:Xem phiếu chờ duyệt tập trung"),
    ("02", "Quyền scope COMPANY thấy phiếu toàn công ty", "P0",
     "User có quyền company-scope (vd 'Kế toán duyệt quyết toán hợp đồng'); công ty user có phiếu quyết toán ở nhiều phòng",
     "1. Vào màn\n2. Lọc loại 'Quyết toán HĐ' (hoặc để tất cả)\n3. Quan sát phòng ban của các phiếu",
     "User: KT trưởng công ty 1",
     "- Thấy MỌI phiếu required_permission='Kế toán duyệt quyết toán hợp đồng' của công ty (company_id = cty user hoặc NULL)\n"
     "- KHÔNG bị giới hạn theo phòng ban user quản lý",
     "BR — scope 'company' (mặc định trong permission_scopes): company_id NULL OR = công ty user, không check department"),
    ("03", "Quyền scope DEPARTMENT chỉ thấy phiếu phòng mình QUẢN LÝ", "P0",
     "User là Trưởng phòng quản lý phòng X (có bản ghi employee_manage_departments); có phiếu required_permission="
     "'Trưởng phòng duyệt quyết toán hợp đồng' ở phòng X (2 phiếu) và phòng Y không quản lý (3 phiếu)",
     "1. Vào màn\n2. Quan sát danh sách",
     "User: TP quản lý phòng X",
     "- CHỈ thấy 2 phiếu phòng X\n- KHÔNG thấy 3 phiếu phòng Y",
     "BR — scope 'department': department_id ∈ employee_manage_departments của user"),
    ("04", "Quyền department nhưng KHÔNG quản lý phòng nào → không thấy phiếu department", "P1",
     "User có quyền 'Trưởng phòng duyệt quyết toán hợp đồng' nhưng KHÔNG có bản ghi employee_manage_departments nào",
     "1. Vào màn\n2. Quan sát",
     "User: có quyền TP, quản lý 0 phòng",
     "- KHÔNG thấy phiếu nào thuộc quyền department đó (tập phòng quản lý rỗng → whereIn([-1]))\n"
     "- Vẫn thấy phiếu thuộc các quyền company-scope khác (nếu có)",
     "BR — myDeptIds rỗng → không match phiếu department"),
    ("05", "Phiếu chỉ định đích danh chỉ hiện với người được gán", "P0",
     "1 phiếu firm_contract đã chuyển duyệt BGĐ (approver_id = userA.id). Có userA (được gán) và userB (cũng có quyền BGĐ nhưng không được gán)",
     "1. Đăng nhập userA, vào màn → tìm phiếu\n2. Đăng nhập userB, vào màn → tìm phiếu",
     "userA = approver_id; userB có quyền BGĐ",
     "- userA THẤY phiếu (nhánh approver_id NOT NULL AND = user.id)\n- userB KHÔNG thấy phiếu",
     "BR — nhánh (a) approver_id đích danh ưu tiên, chặn người khác cùng quyền"),
    ("06", "User có NHIỀU loại quyền → hộp duyệt là HỢP các phạm vi", "P1",
     "User vừa có quyền company-scope (Kế toán) vừa department-scope (Trưởng phòng, quản lý phòng X)",
     "1. Vào màn\n2. Quan sát các loại phiếu",
     "User: đa quyền",
     "- Thấy phiếu Kế toán toàn công ty + phiếu TP phòng X\n- Tổng = union, không trùng lặp dòng",
     "BR — builder OR nhiều nhánh (company OR department OR part OR approver_id)"),
]

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Load trang danh sách đầy đủ layout", "P0",
         "User có quyền xem, hộp duyệt có phiếu",
         "1. Vào /admin/approval-inbox\n2. Quan sát toàn bộ layout",
         "User: DNS Admin",
         "- Tiêu đề trang 'Phiếu chờ duyệt'\n- Vùng bộ lọc (7 filter + nút Tìm kiếm/Làm mới/Xuất Excel)\n"
         "- Bảng 8 cột\n- Phân trang dưới bảng\n- KHÔNG có sidebar 'Lọc theo nhóm' (đã bỏ, full-width)",
         "—"),
        ("002", "Badge menu hiển thị đúng số phiếu chờ", "P1",
         "Hộp duyệt user có đúng N phiếu PENDING",
         "1. Quan sát menu 'Phiếu chờ duyệt (N)' trên thanh menu",
         "N = pendingCountForUser",
         "- Badge hiển thị số N = tổng phiếu PENDING trong hộp duyệt user\n- N=0 thì không hiện badge",
         "Ô badge = ApprovalInbox::pendingCountForUser(user)"),
        ("003", "Menu cũ 'Chờ duyệt' đã bị gỡ", "P2",
         "User đăng nhập",
         "1. Quan sát thanh menu top",
         "—",
         "- KHÔNG còn menu mega 'Chờ duyệt' cũ\n- Chỉ còn menu 'Phiếu chờ duyệt' (Danh sách + Báo cáo phê duyệt)",
         "BR — menu cũ đã comment {{-- --}} trong topmenubar.blade.php"),
        ("004", "Menu con Báo cáo chỉ hiện khi có quyền", "P2",
         "User có 'Xem phiếu chờ duyệt tập trung' nhưng KHÔNG có 'Xem báo cáo phê duyệt'",
         "1. Mở menu 'Phiếu chờ duyệt'\n2. Quan sát submenu",
         "User: chỉ quyền xem danh sách",
         "- Có 'Danh sách phiếu chờ duyệt'\n- KHÔNG có 'Báo cáo phê duyệt'",
         "@can('Xem báo cáo phê duyệt') bọc submenu báo cáo"),
    ]),
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Lọc theo Loại phiếu (có kèm số lượng)", "P0",
         "Hộp duyệt có nhiều loại; loại 'YC xuất giữ' có n phiếu",
         "1. Mở select 'Loại phiếu'\n2. Chọn 'YC xuất giữ (n)'\n3. Quan sát danh sách + footInfo",
         "Chọn 1 loại có n phiếu",
         "- Option hiển thị 'Nhãn (n)' với n = số phiếu loại đó\n- Danh sách chỉ còn phiếu loại đã chọn\n"
         "- footInfo = 'Hiển thị 1–.. / n phiếu'\n- Về trang 1",
         "Ô option = typeCounts[doc_type]; filter doc_type"),
        ("002", "Option 'Tất cả loại (N)' hiển thị tổng", "P1",
         "Hộp duyệt có tổng N phiếu",
         "1. Mở select 'Loại phiếu'\n2. Quan sát option đầu",
         "N = total",
         "- Option đầu = '— Tất cả loại (N) —'\n- Chọn nó → hiện tất cả loại",
         "N = total (pendingCountForUser)"),
        ("003", "Cascade Phòng ban → Bộ phận", "P0",
         "Hộp duyệt có phiếu ở phòng X (có bộ phận X1, X2)",
         "1. Chọn Phòng ban = X\n2. Mở select Bộ phận",
         "Phòng X có bộ phận X1, X2",
         "- Select Bộ phận chỉ hiện bộ phận thuộc phòng X (X1, X2)\n- Trước khi chọn phòng: Bộ phận = '— Chọn phòng ban trước —'",
         "FE cascade client-side AI_PARTS.filter(part.department_id == dept)"),
        ("004", "Danh mục Phòng ban GIỚI HẠN theo hộp duyệt user", "P0",
         "User department-scope quản lý phòng X (có phiếu). Hệ thống có nhiều phòng khác không phiếu",
         "1. Mở select 'Phòng ban'\n2. Quan sát danh sách phòng",
         "User quản lý phòng X",
         "- Chỉ hiện các phòng CÓ PHIẾU trong hộp duyệt user (không hiện toàn bộ bảng phòng ban)\n"
         "- Không có option 'chết' (chọn ra 0 phiếu)",
         "BR — scopeDeptIds = distinct department_id của pendingBuilderForUser"),
        ("005", "Người yêu cầu — select search, data theo phòng/bộ phận", "P0",
         "Chọn Phòng ban = X",
         "1. Chọn Phòng ban X\n2. Mở 'Người yêu cầu', gõ tên NV\n3. Quan sát gợi ý",
         "Gõ 1 phần tên NV phòng X",
         "- Gợi ý chỉ gồm người GỬI phiếu trong hộp duyệt user + thuộc phòng X (nếu đã chọn)\n"
         "- KHÔNG trả toàn bộ nhân viên hệ thống",
         "BR — employeeOptions giới hạn requesterIds ∈ hộp duyệt user + lọc theo company/dept/part đang chọn"),
        ("006", "Lọc Từ ngày / Đến ngày theo Thời gian gửi", "P0",
         "Phiếu A submitted_at=05/07/2026, phiếu B submitted_at=20/07/2026",
         "1. Từ ngày = 10/07/2026, Đến ngày = 31/07/2026\n2. Tìm kiếm",
         "From=10/07, To=31/07",
         "- Hiện phiếu B (20/07)\n- KHÔNG hiện phiếu A (05/07)",
         "Range: whereDate(submitted_at) >= from AND <= to"),
        ("007", "Tìm kiếm theo Mã phiếu", "P0",
         "Có phiếu mã 'PYCXH-33920'",
         "1. Nhập 'PYCXH-33920' (hoặc '33920') ô Tìm kiếm\n2. Enter / bấm Tìm kiếm",
         "Từ khoá = 33920",
         "- Danh sách lọc còn phiếu có code chứa '33920'\n- Tìm kiếm dạng LIKE %kw%",
         "filter code LIKE %kw%"),
        ("008", "Nút Làm mới reset toàn bộ bộ lọc", "P1",
         "Đang có nhiều filter đang chọn (loại, phòng, ngày, tìm kiếm)",
         "1. Bấm 'Làm mới'\n2. Quan sát",
         "—",
         "- Mọi filter về mặc định (loại='Tất cả', phòng/bộ phận/người yêu cầu trống, ngày trống, tìm kiếm trống)\n"
         "- Danh sách tải lại toàn bộ, về trang 1",
         "resetFilters()"),
        ("009", "ĐÃ BỎ lọc Công ty", "P2",
         "User đăng nhập",
         "1. Quan sát vùng bộ lọc",
         "—",
         "- KHÔNG có dropdown 'Công ty' (vì chỉ 1 công ty của user)\n- Filter đầu tiên là 'Phòng ban'",
         "BR — bỏ f-company; scope đã theo hộp duyệt user"),
        ("010", "ĐÃ BỎ lọc Quyền áp dụng", "P2",
         "User đăng nhập",
         "1. Quan sát vùng bộ lọc",
         "—",
         "- KHÔNG có dropdown 'Quyền áp dụng' trong bộ lọc\n- (Cột 'Quyền áp dụng' trong BẢNG vẫn còn)",
         "BR — bỏ filter f-level, giữ cột hiển thị"),
        ("011", "Thứ tự bộ lọc đúng yêu cầu", "P2",
         "User đăng nhập",
         "1. Quan sát thứ tự các filter trái→phải",
         "—",
         "- Thứ tự: Phòng ban → Bộ phận → Người yêu cầu → Loại phiếu → Từ ngày → Đến ngày → Tìm kiếm",
         "—"),
        ("012", "Select filter là dạng search (select2)", "P1",
         "Có nhiều option loại/phòng",
         "1. Click select 'Loại phiếu'\n2. Gõ 1 phần nhãn để lọc option",
         "Gõ 'xuất'",
         "- Select mở dropdown có ô tìm kiếm\n- Gõ lọc được option theo tên\n- Mở dropdown KHÔNG làm vỡ layout / tràn ngang màn hình",
         "BR — select2 (bỏ dropdownAutoWidth để không vỡ layout)"),
    ]),
    ("III", "THỐNG KÊ ĐẦU TRANG / ĐẾM", [
        ("001", "footInfo phản ánh đúng tổng theo bộ lọc", "P0",
         "Lọc ra M phiếu",
         "1. Áp 1 bộ lọc\n2. Quan sát footInfo + listCount",
         "Kết quả M phiếu",
         "- footInfo = 'Hiển thị {from}–{to} / M phiếu'\n- listCount góc phải = 'M phiếu'",
         "recordsFiltered = M"),
        ("002", "Số trong option loại = số thực tế", "P1",
         "Loại 'HĐ trong nước' có k phiếu trong hộp duyệt",
         "1. Mở select loại\n2. Đọc số trong '(k)'\n3. Chọn loại đó, đối chiếu footInfo",
         "k = typeCounts",
         "- Số '(k)' khớp số phiếu khi lọc đúng loại đó",
         "typeCounts group by doc_type"),
    ]),
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Cột STT liên tục qua các trang", "P0",
         "Có > 50 phiếu, page-size = 50",
         "1. Xem STT trang 1 (1..50)\n2. Sang trang 2\n3. Xem STT dòng đầu",
         "Trang 2",
         "- Trang 1: STT 1..50\n- Trang 2: STT bắt đầu 51 (liên tục, không reset về 1)",
         "STT = offset + i + 1"),
        ("002", "Cột Loại phiếu hiển thị badge màu theo nhóm", "P2",
         "Phiếu thuộc nhóm QUAN_LY_HOP_DONG / KE_TOAN_KHO…",
         "1. Quan sát cột Loại phiếu",
         "—",
         "- Nhãn loại hiển thị dạng badge bo tròn, màu theo group_code (GB map)",
         "—"),
        ("003", "Cột Mã phiếu là deep-link mở TAB MỚI", "P0",
         "Phiếu có approve_route hợp lệ",
         "1. Click mã phiếu (có icon ↗)\n2. Quan sát",
         "—",
         "- Mở TAB MỚI (target=_blank) tới trang chi tiết phiếu gốc = route(approve_route, source_id)\n"
         "- Ví dụ YC nhập hàng → /admin/warehouse/product_import_requests/{id}/show",
         "deep-link = route(approve_route, approve_params.id)"),
        ("004", "Cột Đối tác hiển thị tag KH/NCC + tên", "P1",
         "Phiếu có partner KH; phiếu khác partner NCC; phiếu khác không partner",
         "1. Quan sát cột Đối tác của 3 phiếu",
         "—",
         "- KH → tag 'KH' (xanh) + tên khách hàng\n- NCC → tag 'NCC' (xám) + tên nhà cung cấp\n- Không partner → '—'",
         "partner_type kh/ncc"),
        ("005", "Cột Đối tác dài quá bị cắt '…' + hover xem full", "P1",
         "Phiếu có partner tên rất dài (> ~30 ký tự)",
         "1. Quan sát ô đối tác\n2. Rê chuột hover lên tên",
         "Tên NCC dài",
         "- Tên bị cắt ngắn kết thúc bằng '…' (không xuống dòng, không phá bảng)\n- Hover hiện tooltip title = tên đầy đủ",
         "BR — .partner max-width 260px, .pname ellipsis + title"),
        ("006", "Cột Người yêu cầu có phòng ban subtext", "P0",
         "Phiếu do NV 'Phạm Minh Hiếu' phòng 'PHÒNG THIẾT BỊ Ô TÔ 1' gửi",
         "1. Quan sát cột Người yêu cầu",
         "—",
         "- Dòng trên: avatar + tên NV\n- Dòng dưới (subtext nhỏ, xám): tên PHÒNG BAN của phiếu",
         "department_name lấy từ approval_inbox.department_id"),
        ("007", "Cột Thời gian gửi định dạng ngày + giờ", "P0",
         "Phiếu submitted_at = 2026-07-07 14:39",
         "1. Quan sát cột Thời gian gửi",
         "—",
         "- Hiển thị '07/07/2026 14:39' (dd/mm/yyyy HH:mm)",
         "format submitted_at d/m/Y H:i"),
        ("008", "Cột Quyền áp dụng hiển thị đúng quyền cấp hiện tại", "P1",
         "Phiếu YC xuất hàng ở cấp BGĐ",
         "1. Quan sát cột Quyền áp dụng",
         "—",
         "- Hiển thị tên required_permission cấp hiện tại (vd 'Ban giám đốc duyệt xuất hàng vượt hạn mức công nợ')",
         "required_permission"),
        ("009", "Nút Duyệt (cột Hành động) mở phiếu để duyệt", "P0",
         "Phiếu có approve_route",
         "1. Bấm nút 'Duyệt'",
         "—",
         "- Điều hướng tới trang chi tiết phiếu gốc để thực hiện duyệt (cùng đích deep-link mã phiếu)",
         "action link route(approve_route, id)"),
        ("010", "Phân trang: đổi page-size", "P1",
         "Có > 100 phiếu",
         "1. Chọn page-size = 100\n2. Quan sát số dòng + số trang",
         "page-size = 100",
         "- Hiển thị tối đa 100 dòng/trang\n- Số trang tính lại theo total/100\n- Về trang 1",
         "server-side length"),
        ("011", "Phân trang: nút ‹ vô hiệu ở trang 1", "P2",
         "Đang ở trang 1, có nhiều trang",
         "1. Quan sát nút ‹ (prev)",
         "Trang 1",
         "- Nút ‹ disabled\n- Nút › enabled",
         "—"),
        ("012", "Đổi bộ lọc thì quay về trang 1", "P1",
         "Đang ở trang 3",
         "1. Đổi 1 filter bất kỳ (loại phiếu)\n2. Quan sát trang hiện tại",
         "—",
         "- Danh sách tải lại và về trang 1 (page=1)",
         "loadRows(false) reset page"),
        ("013", "Bảng cuộn ngang khi nội dung dài, không xuống dòng", "P2",
         "Màn hình hẹp / nhiều cột dài",
         "1. Thu nhỏ cửa sổ\n2. Quan sát bảng",
         "—",
         "- Ô dữ liệu KHÔNG wrap nhiều dòng\n- Bảng cuộn NGANG (overflow-x) trong khung, không phá layout trang",
         "BR — td nowrap + tablewrap overflow-x:auto"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH — XUẤT EXCEL", [
        ("001", "Xuất Excel theo đúng bộ lọc hiện tại", "P0",
         "Đang lọc loại = 'Quyết toán HĐ'",
         "1. Bấm 'Xuất Excel'\n2. Mở file tải về",
         "Lọc loại Quyết toán HĐ",
         "- File .xlsx tải về (Content-Type spreadsheetml, tên phieu-cho-duyet-YYYYMMDD-HHMMSS.xlsx)\n"
         "- Nội dung CHỈ gồm phiếu khớp bộ lọc đang áp (không phân trang — xuất toàn bộ)",
         "BR — exportExcel dùng chung applyFilters, bỏ start/length"),
        ("002", "File Excel đủ 8 cột đúng thứ tự", "P1",
         "Có phiếu để xuất",
         "1. Xuất Excel\n2. Mở file, xem header",
         "—",
         "- Cột: STT | Loại phiếu | Mã phiếu | Đối tác | Người yêu cầu | Phòng ban | Thời gian gửi | Quyền áp dụng\n"
         "- Header đậm, có border",
         "ApprovalInboxExcel headings"),
        ("003", "Xuất Excel không có phiếu khớp", "P2",
         "Bộ lọc ra 0 phiếu (vd tìm mã không tồn tại)",
         "1. Lọc ra 0 phiếu\n2. Bấm Xuất Excel",
         "0 phiếu",
         "- File xuất ra chỉ có dòng header, không có dòng dữ liệu (không lỗi)",
         "—"),
        ("004", "Cột Đối tác trong Excel có tiền tố KH:/NCC:", "P2",
         "Phiếu có partner",
         "1. Xuất Excel\n2. Xem cột Đối tác",
         "—",
         "- Giá trị dạng 'KH: <tên>' hoặc 'NCC: <tên>'; không partner để trống",
         "—"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Hộp duyệt rỗng", "P1",
         "User không có phiếu nào chờ duyệt",
         "1. Vào màn",
         "User: 0 phiếu",
         "- Bảng hiện dòng 'Không có phiếu nào khớp bộ lọc'\n- footInfo = 'Không có phiếu'\n- Badge menu không hiện",
         "—"),
        ("002", "Phiếu có company_id/department_id = NULL vẫn hiện với quyền company", "P0",
         "Phiếu doc_type='addition_accounting_request' type=7 (company_id NULL); user có quyền company-scope tương ứng",
         "1. Vào màn\n2. Tìm phiếu đó",
         "Phiếu company_id NULL",
         "- Phiếu VẪN hiện (company scope: company_id NULL OR = cty user)\n"
         "- Không đóng góp vào danh mục Phòng ban (department_id NULL) nhưng lọc 'Tất cả phòng ban' vẫn thấy",
         "BR — vài luồng P5 gate cross-company → company_id NULL cố ý"),
        ("003", "Phiếu nộp thẳng cấp cao (bỏ qua cấp gốc) — hiển thị nhờ backfill", "P2",
         "Phiếu YC xuất hàng tạo thẳng status=10 (bỏ qua CHO_DUYET); đã chạy backfill",
         "1. Vào màn\n2. Tìm phiếu",
         "Phiếu nộp thẳng cấp TP",
         "- Phiếu vẫn hiện đúng cấp (backfill bắt được)\n- (Lưu ý known-gap: live hook chưa push khi nộp thẳng — cần backfill định kỳ)",
         "BR — gap push-condition luồng multi-level, backfill là lưới an toàn"),
        ("004", "Chọn trang vượt quá tổng số trang", "P2",
         "Có 3 trang, đang page 3, rồi lọc còn 1 trang",
         "1. Ở trang cao\n2. Áp lọc thu hẹp còn ít phiếu",
         "—",
         "- page tự kẹp về totalPages hợp lệ, không hiện trang trống lỗi",
         "renderPager kẹp page"),
        ("005", "Tên đối tác/tên phiếu chứa ký tự đặc biệt < > & \" '", "P1",
         "Phiếu có partner_name chứa ký tự HTML đặc biệt",
         "1. Xem cột Đối tác",
         "Partner: 'CTY <A&B>'",
         "- Hiển thị đúng nguyên văn ký tự, KHÔNG bị chèn HTML (đã esc())",
         "BR — chống XSS: esc() mọi field render"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Không thấy phiếu ngoài phạm vi quyền của mình", "P0",
         "User A (TP phòng X) và có phiếu phòng Y (không quản lý)",
         "1. Đăng nhập A\n2. Tìm phiếu phòng Y bằng bộ lọc/tìm kiếm",
         "User A không quản lý phòng Y",
         "- Không có cách nào (kể cả tìm theo mã) làm hiện phiếu phòng Y trong hộp duyệt A\n"
         "- searchData luôn xuất phát từ pendingBuilderForUser(user)",
         "BR — mọi query gốc = pendingForCurrentUser, filter chỉ thu hẹp thêm"),
        ("002", "Endpoint search-data/employees/export đều gate quyền", "P0",
         "User không có 'Xem phiếu chờ duyệt tập trung'",
         "1. Gọi thẳng /admin/approval-inbox/search-data (và /employees, /export)",
         "User không quyền",
         "- Bị chặn checkPermission (403/redirect), không trả dữ liệu",
         "BR — middleware trên cả 4 route"),
        ("003", "Danh mục lọc không lộ phòng/người ngoài phạm vi", "P1",
         "User department-scope quản lý phòng X",
         "1. Mở select Phòng ban / Người yêu cầu",
         "User quản lý phòng X",
         "- Select Phòng ban chỉ có phòng trong hộp duyệt (không lộ toàn bộ cơ cấu công ty)\n"
         "- Người yêu cầu chỉ gồm người gửi phiếu trong hộp duyệt",
         "BR — scopeDeptIds/scopePartIds + requesterIds giới hạn theo builder"),
        ("004", "Xuất Excel cũng chỉ trong phạm vi hộp duyệt", "P0",
         "User A hộp duyệt có 20 phiếu",
         "1. Đăng nhập A\n2. Xuất Excel (không lọc gì)",
         "User A",
         "- File chỉ chứa ≤ 20 phiếu của A, không rò phiếu người/phòng khác",
         "exportExcel dùng pendingForCurrentUser"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Lọc phòng → chọn người yêu cầu → mở phiếu duyệt", "P0",
         "User có phiếu ở phòng X do NV Nguyễn Văn A gửi",
         "1. Chọn Phòng ban X\n2. Chọn Người yêu cầu = Nguyễn Văn A\n3. Bấm mã phiếu / nút Duyệt\n4. Duyệt trên trang chi tiết",
         "Phòng X, NV A",
         "- Danh sách lọc đúng phiếu phòng X của A\n- Deep-link mở đúng phiếu gốc\n"
         "- Sau khi duyệt trên màn gốc, phiếu rời hộp duyệt (sau đồng bộ live hook/backfill)",
         "E2E: lọc → deep-link → duyệt"),
        ("002", "Lọc theo thời gian gửi → xuất Excel", "P0",
         "Có phiếu nhiều mốc submitted_at",
         "1. Đặt Từ ngày/Đến ngày\n2. Kiểm danh sách\n3. Bấm Xuất Excel\n4. Đối chiếu file",
         "From/To cụ thể",
         "- Danh sách và file Excel cùng tập phiếu (khớp bộ lọc thời gian)\n- Số dòng file = tổng theo lọc",
         "E2E: filter thời gian → export"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':22,'C':16,'D':42,'E':10,'F':32,'G':55,'H':22,'I':65,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 15 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL; fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",           '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",      '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",  '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",     '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",               '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8

HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
           "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
           "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18; data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data,
              expected, business_note, "", "Not Executed", "Not Executed", "Not Executed", ""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, longest // 3))
    current_row += 1; data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note, group="Phân quyền & truy cập")

ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Data rows 18-{current_row-1} (tổng {data_row_idx} TC)")
