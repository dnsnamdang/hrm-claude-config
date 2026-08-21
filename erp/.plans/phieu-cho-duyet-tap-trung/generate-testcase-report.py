"""Generate testcase Excel cho màn Báo cáo phê duyệt (Phiếu chờ duyệt tập trung)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = os.path.join(os.path.dirname(__file__), "testcase-report.xlsx")
SHEET_NAME   = "BaoCaoPheDuyet"
FEATURE_NAME = "Báo cáo phê duyệt"
MODULE_NAME  = "Phiếu chờ duyệt tập trung"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
        "Thống kê tình hình phê duyệt theo thời gian và theo người duyệt, dựa trên log bước duyệt "
        "(approval_inbox_logs). Mỗi dòng log = 1 LƯỢT duyệt/từ chối ở 1 cấp. Route: approvalInbox.report "
        "(màn), approvalInbox.reportData (JSON số liệu), approvalInbox.reportExport (Excel)."),
    ("2. Đối tượng được tính / hiển thị",
        "► Các bản ghi approval_inbox_logs có action = 'approve' HOẶC 'reject' (đã xử lý ở 1 cấp).\n"
        "► Đếm THEO LƯỢT: phiếu nhiều vòng (từ chối rồi duyệt lại) tính nhiều lượt (mỗi log 1 lượt).\n"
        "► Lọc theo ended_at (ngày hoàn tất bước) cho khoảng kỳ báo cáo.\n"
        "► KPI 'Đang chờ' lấy từ approval_inbox status = pending (KHÔNG từ log)."),
    ("3. Đối tượng bị ẩn / không tính",
        "► Log action = 'cancel' (hủy) — KHÔNG vào số liệu approve/reject.\n"
        "► Phiếu đang pending chưa có log approve/reject — chỉ tính vào KPI 'Đang chờ', không vào 'Tổng đã xử lý'.\n"
        "► Bảng 'Hiệu suất theo người duyệt' bị ẩn khi lọc đúng 1 người duyệt."),
    ("4. Bộ lọc thời gian áp dụng cho",
        "Cột approval_inbox_logs.ended_at. Kỳ: Tháng này / Tháng trước / Quý này / Năm nay / Tùy chọn (Từ–Đến). "
        "Range gửi lên BE là from/to (whereDate ended_at >= from, <= to). KPI 'Đang chờ'/'Quá hạn' KHÔNG phụ thuộc kỳ "
        "(lấy trạng thái pending hiện tại theo filter org/loại)."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
        "approval_inbox_logs JOIN approval_inbox (inbox_id). Filter org (company_id/department_id/part_id) lấy từ "
        "bảng approval_inbox; filter actor_id/action lấy từ log. Người duyệt = actor_id (người thực sự thao tác bước đó). "
        "TG duyệt 1 lượt = ended_at − started_at của chính log đó (theo TỪNG CẤP)."),
    ("6. Quy tắc cộng dồn / deduplicate",
        "► KHÔNG dedupe theo phiếu — đếm theo lượt (log rows).\n"
        "► TG duyệt TB = trung bình (ended_at − started_at) trên tất cả log trong kỳ.\n"
        "► Tỉ lệ đúng hạn = số lượt có TG ≤ 24h / tổng lượt.\n"
        "► reportData (màn) và reportExport (Excel) DÙNG CHUNG hàm reportRows() → số liệu phải khớp nhau."),
    ("7. Phân quyền cấp",
        "• Xem báo cáo phê duyệt — quyền DUY NHẤT gate cả 3 route (report, reportData, reportExport) qua "
        "middleware checkPermission. Không có quyền → 403 (không vào được màn/không lấy được data/không tải Excel).\n"
        "• Lưu ý: bộ lọc Công ty/Phòng/Bộ phận là lọc DỮ LIỆU, KHÔNG phải phân quyền phạm vi (mọi user có quyền đều "
        "xem được toàn bộ log theo filter họ chọn)."),
    ("8. Cách tính các ô thống kê",
        "► Ô 'Tổng đã xử lý' = COUNT(log action IN ['approve','reject']) trong kỳ + filter.\n"
        "► Ô 'Đã duyệt' = COUNT(action='approve'); tỉ lệ = approved/total.\n"
        "► Ô 'Từ chối' = COUNT(action='reject'); tỉ lệ = rejected/total.\n"
        "► Ô 'Đang chờ duyệt' = COUNT(approval_inbox.status=pending) theo filter org/loại (KHÔNG theo actor/kết quả/kỳ).\n"
        "► Ô 'TG duyệt trung bình' = AVG(ended_at − started_at) giờ, theo từng cấp.\n"
        "► Ô 'Quá hạn (>3 ngày)' = COUNT(pending có level_started_at < now − 3 ngày).\n"
        "► Ô 'Tỉ lệ đúng hạn' = COUNT(TG lượt ≤ 24h) / total × 100."),
    ("9. Ghi chú đọc bảng",
        "► Bảng 'Chi tiết phiếu đã xử lý' hiển thị tối đa 100 dòng gần nhất (sort ended_at desc); Excel xuất TOÀN BỘ theo filter.\n"
        "► Mã phiếu là link mở thẳng phiếu gốc ở TAB MỚI (route <luồng>.show); route rỗng/lỗi → hiện text thường không link.\n"
        "► Dòng 'Từ chối' hiện kèm lý do (approval_inbox_logs.note) dưới badge + cột riêng khi xuất Excel.\n"
        "► Biểu đồ cột: xanh = đã duyệt, đỏ = từ chối, gộp theo tuần ISO của ended_at."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("01", "Truy cập màn báo cáo khi CÓ quyền 'Xem báo cáo phê duyệt'", "P0",
        "User đăng nhập, đã gán quyền 'Xem báo cáo phê duyệt'",
        "1. Vào menu Phiếu chờ duyệt > Báo cáo phê duyệt\n2. Quan sát màn hình",
        "User: có quyền 'Xem báo cáo phê duyệt'",
        "- Màn /admin/approval-inbox/report load thành công\n- Hiện đủ bộ lọc, 7 KPI, 2 biểu đồ, bảng người duyệt, bảng chi tiết",
        "Permission: Xem báo cáo phê duyệt (route approvalInbox.report)"),
    ("02", "Truy cập màn báo cáo khi KHÔNG có quyền", "P0",
        "User đăng nhập, KHÔNG gán quyền 'Xem báo cáo phê duyệt'",
        "1. Gõ trực tiếp URL /admin/approval-inbox/report",
        "User: không có quyền",
        "- Bị chặn 403 (checkPermission) — không vào được màn báo cáo",
        "BR — fail-closed: không quyền thì chặn ở BE, không dựa FE ẩn"),
    ("03", "Gọi endpoint số liệu report-data khi KHÔNG có quyền", "P0",
        "User đăng nhập, KHÔNG có quyền 'Xem báo cáo phê duyệt'",
        "1. Gọi trực tiếp GET /admin/approval-inbox/report-data",
        "User: không có quyền",
        "- Trả 403, không lộ số liệu log duyệt",
        "Permission: reportData cũng gate cùng quyền (defense-in-depth)"),
    ("04", "Tải Excel report-export khi KHÔNG có quyền", "P0",
        "User đăng nhập, KHÔNG có quyền 'Xem báo cáo phê duyệt'",
        "1. Gọi trực tiếp GET /admin/approval-inbox/report-export",
        "User: không có quyền",
        "- Trả 403, không tải được file Excel",
        "Permission: reportExport cũng gate cùng quyền"),
]

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Load màn báo cáo với kỳ mặc định = Tháng này", "P0",
            "User có quyền; có ≥1 log approve/reject trong tháng hiện tại",
            "1. Vào màn Báo cáo phê duyệt\n2. Quan sát kỳ mặc định + số liệu tự đổ",
            "Kỳ mặc định: Tháng này",
            "- Bộ lọc Kỳ = 'Tháng này'\n- KPI/biểu đồ/bảng tự load (gọi reportData ngay khi mở, không cần bấm Xem)\n- Không hiện ô Từ ngày/Đến ngày",
            "Màn auto loadReport() khi mở"),
        ("002", "Bố cục đủ khối theo mockup", "P1",
            "User có quyền; có dữ liệu",
            "1. Quan sát toàn trang từ trên xuống",
            "—",
            "- Thứ tự: Bộ lọc → 7 KPI → 2 biểu đồ (thời gian + nhóm) → bảng Hiệu suất người duyệt → bảng Chi tiết",
            "FE port từ mockup-report.html"),
        ("003", "Nút Xem báo cáo tải lại số liệu theo lọc", "P1",
            "User có quyền; đã đổi vài bộ lọc",
            "1. Đổi Kỳ/Loại phiếu\n2. Bấm '🔍 Xem báo cáo'",
            "—",
            "- Toàn bộ KPI/biểu đồ/bảng cập nhật theo bộ lọc mới",
            "—"),
        ("004", "Nút Làm mới đưa bộ lọc về mặc định", "P1",
            "User có quyền; đã đổi nhiều bộ lọc",
            "1. Đổi Kỳ, Người duyệt, Loại, Kết quả\n2. Bấm '↻ Làm mới'",
            "—",
            "- Kỳ về 'Tháng này', ẩn Từ/Đến ngày\n- Người duyệt/Loại/Kết quả về 'Tất cả'\n- Số liệu load lại",
            "resetFilters()"),
    ]),
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Lọc kỳ 'Tháng trước'", "P0",
            "Có log tháng trước và tháng này khác nhau",
            "1. Chọn Kỳ = 'Tháng trước'",
            "—",
            "- from/to = đầu–cuối tháng trước\n- Số liệu chỉ gồm log có ended_at thuộc tháng trước",
            "Range: month-1"),
        ("002", "Lọc kỳ 'Quý này'", "P1",
            "Có log rải rác trong quý hiện tại",
            "1. Chọn Kỳ = 'Quý này'",
            "—",
            "- from/to = đầu–cuối quý hiện tại\n- KPI tổng gộp cả 3 tháng của quý",
            "Range: quarter hiện tại"),
        ("003", "Lọc kỳ 'Năm nay'", "P1",
            "Có log trong năm",
            "1. Chọn Kỳ = 'Năm nay'",
            "—",
            "- from/to = 01/01 → 31/12 năm hiện tại",
            "Range: year"),
        ("004", "Lọc kỳ 'Tùy chọn' hiện ô Từ/Đến ngày", "P0",
            "User có quyền",
            "1. Chọn Kỳ = 'Tùy chọn…'\n2. Quan sát",
            "—",
            "- Hiện 2 ô 'Từ ngày' và 'Đến ngày'\n- Nhập khoảng ngày → bấm Xem → lọc theo khoảng đó",
            "period=custom → dùng from/to nhập tay"),
        ("005", "Cascade Công ty → Phòng ban", "P0",
            "Có ≥2 công ty, mỗi công ty có phòng ban riêng",
            "1. Chọn Công ty A\n2. Mở dropdown Phòng ban",
            "Công ty A",
            "- Dropdown Phòng ban chỉ còn phòng thuộc Công ty A\n- Bộ phận reset về '— Chọn phòng ban trước —'",
            "Cascade theo company_id"),
        ("006", "Cascade Phòng ban → Bộ phận", "P0",
            "Phòng ban đã chọn có ≥1 bộ phận",
            "1. Chọn Công ty\n2. Chọn Phòng ban\n3. Mở dropdown Bộ phận",
            "—",
            "- Dropdown Bộ phận chỉ còn bộ phận thuộc phòng đã chọn",
            "Cascade theo department_id"),
        ("007", "Lọc theo Người duyệt", "P0",
            "Có ≥2 người duyệt từng thao tác trong kỳ",
            "1. Chọn 1 Người duyệt ở dropdown",
            "Người duyệt: NV X",
            "- Số liệu chỉ gồm log có actor_id = NV X\n- Bảng 'Hiệu suất theo người duyệt' bị ẩn (xem TC IV)",
            "Filter actor_id"),
        ("008", "Lọc theo Loại phiếu", "P1",
            "Có log của ≥2 loại phiếu (doc_type)",
            "1. Chọn 1 Loại phiếu",
            "Loại: Hợp đồng bán hàng",
            "- Số liệu chỉ gồm log của phiếu đúng doc_type đó",
            "Filter doc_type"),
        ("009", "Lọc theo Kết quả = Đã duyệt", "P1",
            "Có cả log approve và reject",
            "1. Chọn Kết quả = 'Đã duyệt'",
            "result=approve",
            "- Chỉ còn log action='approve'\n- KPI 'Từ chối' = 0, bảng chi tiết chỉ badge xanh",
            "Filter action=approve"),
        ("010", "Lọc theo Kết quả = Từ chối", "P0",
            "Có ≥1 log reject có lý do (note)",
            "1. Chọn Kết quả = 'Từ chối'",
            "result=reject",
            "- Chỉ còn log action='reject'\n- Bảng chi tiết hiện badge đỏ + dòng lý do từ chối bên dưới",
            "Filter action=reject; lý do = note"),
        ("011", "Kết hợp nhiều bộ lọc", "P1",
            "Có dữ liệu đủ đa dạng",
            "1. Chọn Công ty + Phòng ban + Loại + Kết quả=Đã duyệt\n2. Bấm Xem",
            "—",
            "- Số liệu là giao của tất cả điều kiện (AND)",
            "Các filter cộng dồn AND"),
    ]),
    ("III", "KPI / THỐNG KÊ ĐẦU TRANG", [
        ("001", "KPI 'Tổng đã xử lý' đúng số lượt", "P0",
            "Trong kỳ có đúng N log approve/reject (đối chiếu tinker)",
            "1. Xem ô 'Tổng đã xử lý'",
            "N = số log approve+reject trong kỳ",
            "- Giá trị = N (đếm theo lượt, không dedupe phiếu)",
            "Ô 'Tổng đã xử lý' = COUNT(action IN approve,reject)"),
        ("002", "KPI 'Đã duyệt' + tỉ lệ", "P0",
            "Có A log approve trong tổng N",
            "1. Xem ô 'Đã duyệt'",
            "A duyệt / N tổng",
            "- Số = A; tỉ lệ = round(A/N×100)%",
            "approved/total"),
        ("003", "KPI 'Từ chối' + tỉ lệ", "P1",
            "Có R log reject trong tổng N",
            "1. Xem ô 'Từ chối'",
            "R từ chối / N tổng",
            "- Số = R; tỉ lệ = round(R/N×100)%",
            "rejected/total"),
        ("004", "KPI 'Đang chờ duyệt' theo trạng thái hiện tại", "P0",
            "Có P phiếu approval_inbox đang pending theo filter org/loại",
            "1. Xem ô 'Đang chờ duyệt'",
            "P phiếu pending",
            "- Số = P (từ approval_inbox status=pending)\n- KHÔNG phụ thuộc kỳ/kết quả/người duyệt",
            "Ô 'Đang chờ' = COUNT(status=pending) theo org/doc_type"),
        ("005", "KPI 'TG duyệt trung bình' theo từng cấp", "P0",
            "Các log có started_at/ended_at hợp lệ",
            "1. Xem ô 'TG duyệt trung bình'",
            "—",
            "- Giá trị = AVG(ended_at − started_at) tính bằng giờ, 1 chữ số thập phân\n- Theo từng cấp (mỗi log 1 lượt)",
            "Chốt design: TG theo TỪNG CẤP, không phải toàn phiếu"),
        ("006", "KPI 'Quá hạn (>3 ngày)'", "P1",
            "Có phiếu pending với level_started_at < now−3 ngày",
            "1. Xem ô 'Quá hạn (>3 ngày)'",
            "—",
            "- Đếm phiếu pending có level_started_at < (hôm nay − 3 ngày)",
            "Ô 'Quá hạn' = COUNT(pending & level_started_at < now-3d)"),
        ("007", "KPI 'Tỉ lệ đúng hạn' (≤24h)", "P2",
            "Có log TG xử lý cả ≤24h lẫn >24h",
            "1. Xem ô 'Tỉ lệ đúng hạn'",
            "—",
            "- = round(COUNT(TG lượt ≤ 24h) / total × 100)%",
            "ontime_rate"),
        ("008", "KPI đổi theo bộ lọc", "P1",
            "Dữ liệu đa dạng theo loại/kỳ",
            "1. Đổi Kỳ hoặc Loại phiếu\n2. Bấm Xem",
            "—",
            "- Tất cả 7 KPI tính lại theo bộ lọc mới",
            "—"),
    ]),
    ("IV", "BIỂU ĐỒ & BẢNG", [
        ("001", "Biểu đồ cột theo tuần: đã duyệt vs từ chối", "P1",
            "Log rải ở nhiều tuần trong kỳ",
            "1. Quan sát biểu đồ 'Phiếu phê duyệt theo thời gian'",
            "—",
            "- Mỗi cột = 1 tuần ISO của ended_at\n- Đoạn xanh = đã duyệt, đỏ = từ chối; số tổng trên đầu cột\n- Nhãn 'Tuần NN'",
            "Group theo format o-\\WW của ended_at"),
        ("002", "Biểu đồ phân bố theo nhóm nghiệp vụ", "P1",
            "Log thuộc ≥2 group_code",
            "1. Quan sát biểu đồ 'Phân bố theo nhóm nghiệp vụ'",
            "—",
            "- Mỗi thanh ngang = 1 nhóm (group_code → label config)\n- Sắp giảm dần theo count",
            "groupBy group_code"),
        ("003", "Bảng hiệu suất người duyệt (nhiều người)", "P0",
            "Có ≥2 người duyệt trong kỳ, KHÔNG lọc 1 người",
            "1. Không chọn Người duyệt cụ thể\n2. Xem bảng 'Hiệu suất theo người duyệt'",
            "—",
            "- Mỗi dòng 1 người: Đã duyệt, Từ chối, Tỉ lệ duyệt, TG duyệt TB, Phiếu lâu nhất\n- Sắp giảm dần theo (đã duyệt+từ chối)\n- Dòng Tổng cộng khớp tổng cột",
            "groupBy actor_id"),
        ("004", "Ẩn bảng hiệu suất khi lọc đúng 1 người duyệt", "P0",
            "Có dữ liệu của NV X",
            "1. Chọn Người duyệt = NV X\n2. Bấm Xem",
            "Người duyệt: NV X",
            "- Bảng 'Hiệu suất theo người duyệt' bị ẩn\n- Hiện infobar: 'Đang lọc theo 1 người duyệt: NV X — đã ẩn bảng…'",
            "single actor → ẩn bảng, hiện infobar"),
        ("005", "Bảng chi tiết phiếu đã xử lý", "P0",
            "Có ≥1 log trong kỳ",
            "1. Xem bảng 'Chi tiết phiếu đã xử lý'",
            "—",
            "- Cột: STT, Loại phiếu, Mã phiếu, Người yêu cầu, Người duyệt, Cấp duyệt, Ngày gửi, Ngày duyệt, TG duyệt, Kết quả\n- Tối đa 100 dòng gần nhất (sort ended_at desc)",
            "detail take(100)"),
        ("006", "Dòng từ chối hiện lý do", "P0",
            "Có log reject với note = 'Thiếu chứng từ'",
            "1. Lọc Kết quả=Từ chối\n2. Xem dòng chi tiết",
            "note='Thiếu chứng từ'",
            "- Badge đỏ 'Từ chối'\n- Dưới badge hiện '⤷ Thiếu chứng từ' (từ note)",
            "reason = approval_inbox_logs.note"),
        ("007", "TG duyệt hiển thị dạng giờ/phút", "P2",
            "Log có TG lẻ (vd 6.2 giờ)",
            "1. Xem cột 'TG duyệt' bảng chi tiết",
            "—",
            "- Định dạng '6h 12m' (giờ + phút)",
            "durFmt"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH (deep-link + Excel)", [
        ("001", "Deep-link mã phiếu mở đúng phiếu ở tab mới", "P0",
            "Có ≥1 log của phiếu có approve_route hợp lệ (vd firm_contract)",
            "1. Ở bảng chi tiết, click mã phiếu (vd HĐ-2026-00012)",
            "—",
            "- Mở TAB MỚI đúng phiếu gốc (route <luồng>.show)\n- Màn báo cáo giữ nguyên ở tab cũ",
            "R1: url build server-side, target=_blank"),
        ("002", "Mã phiếu không có route → text thường", "P1",
            "Có log của loại phiếu chưa map approve_route",
            "1. Xem mã phiếu loại đó ở bảng chi tiết",
            "—",
            "- Mã hiện dạng text thường, KHÔNG phải link, không lỗi JS",
            "url='' khi route rỗng/lỗi (try/catch)"),
        ("003", "Xuất Excel theo bộ lọc hiện tại", "P0",
            "Đã chọn 1 bộ lọc (vd Kỳ=Tháng này, Loại=HĐ bán hàng)",
            "1. Bấm '⬇ Xuất Excel'",
            "—",
            "- Tải file bao-cao-phe-duyet-<timestamp>.xlsx\n- Nội dung khớp bộ lọc đang áp (dùng chung reportRows)",
            "R2: reportExport gửi cùng query như reportData"),
        ("004", "File Excel đủ 11 cột + cột Lý do từ chối", "P0",
            "Bộ lọc có cả log approve và reject",
            "1. Mở file Excel vừa tải",
            "—",
            "- Header: STT, Loại phiếu, Mã phiếu, Người yêu cầu, Người duyệt, Quyền áp dụng, Ngày gửi, Ngày duyệt, TG duyệt (giờ), Kết quả, Lý do từ chối\n- Cột 'Lý do từ chối' có giá trị ở dòng Từ chối, rỗng ở dòng Đã duyệt",
            "ApprovalReportExport 11 cột A–K"),
        ("005", "Excel xuất TOÀN BỘ (không cắt 100)", "P1",
            "Bộ lọc cho ra >100 log",
            "1. Bấm Xuất Excel\n2. Đếm số dòng file",
            ">100 log khớp filter",
            "- File có đủ tất cả dòng khớp filter (màn hình chỉ hiện 100, Excel không cắt)",
            "reportExport không take(100)"),
        ("006", "Số liệu Excel khớp màn hình", "P0",
            "Bộ lọc cố định",
            "1. Ghi lại KPI 'Tổng đã xử lý' trên màn\n2. Xuất Excel, đếm số dòng data",
            "—",
            "- Số dòng data Excel = KPI 'Tổng đã xử lý' (cùng nguồn reportRows)",
            "Chống lệch số liệu màn ↔ Excel"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Không có log trong kỳ", "P1",
            "Chọn kỳ không có dữ liệu (vd năm ngoái)",
            "1. Kỳ=Tùy chọn, chọn khoảng không có log\n2. Bấm Xem",
            "Khoảng ngày rỗng",
            "- KPI = 0\n- Biểu đồ hiện 'Không có dữ liệu'\n- Bảng chi tiết hiện 'Không có phiếu nào khớp bộ lọc'",
            "Empty state"),
        ("002", "Kỳ Tùy chọn nhưng bỏ trống Từ/Đến", "P2",
            "User có quyền",
            "1. Kỳ=Tùy chọn, để trống Từ và Đến\n2. Bấm Xem",
            "from/to rỗng",
            "- BE bỏ qua điều kiện ngày (không lọc kỳ) → trả toàn bộ log theo filter còn lại, không lỗi",
            "filled() check ở BE"),
        ("003", "Phiếu nhiều vòng (từ chối rồi duyệt lại)", "P1",
            "1 phiếu: vòng 1 bị từ chối, vòng 2 được duyệt (2 log)",
            "1. Xem KPI + bảng chi tiết",
            "phiếu có round 1 reject + round 2 approve",
            "- Tính 2 LƯỢT: +1 Từ chối (người vòng 1), +1 Đã duyệt (người vòng 2)\n- 2 dòng riêng trong chi tiết",
            "Đếm theo lượt, không distinct phiếu"),
        ("004", "Nhiều cấp duyệt cùng 1 phiếu", "P2",
            "Phiếu qua 2 cấp (TP → BGĐ), mỗi cấp 1 log approve",
            "1. Xem bảng chi tiết + TG duyệt",
            "phiếu 2 cấp",
            "- 2 dòng, mỗi dòng TG = thời gian giữ phiếu của cấp đó (không cộng dồn toàn phiếu)",
            "TG theo từng cấp"),
        ("005", "Lý do từ chối rỗng", "P2",
            "Có log reject không nhập note",
            "1. Lọc Từ chối, xem dòng đó",
            "note=null",
            "- Badge đỏ 'Từ chối', không hiện dòng lý do (không hiện '⤷ null')",
            "reason rỗng → không render dòng rsn"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Lọc công ty chỉ ra log của công ty đó", "P0",
            "Có log của Công ty A và Công ty B",
            "1. Chọn Công ty A\n2. Xem số liệu",
            "Công ty A",
            "- Chỉ log của phiếu thuộc Công ty A (approval_inbox.company_id = A)\n- Không lẫn log Công ty B",
            "Filter company_id trên inbox"),
        ("002", "reportData không nhận filter ngoài whitelist", "P2",
            "User có quyền",
            "1. Thêm query lạ (vd ?status=hacked) vào report-data",
            "—",
            "- BE bỏ qua param không hỗ trợ, không lỗi, không lộ data ngoài phạm vi",
            "Chỉ đọc filled() các key định trước"),
        ("003", "Deep-link không cho bypass quyền phiếu gốc", "P1",
            "Phiếu gốc có phân quyền riêng ở màn .show",
            "1. Click mã phiếu mở phiếu gốc",
            "—",
            "- Màn .show gốc tự áp quyền của nó (báo cáo chỉ điều hướng, không nới quyền)",
            "Deep-link = điều hướng, quyền do màn gốc quản"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đối chiếu số liệu end-to-end", "P0",
            "Dev có data log; biết trước số liệu tinker",
            "1. Tinker đếm log approve/reject theo 1 filter\n2. Mở màn, áp đúng filter\n3. So KPI với tinker\n4. Xuất Excel, đếm dòng",
            "1 bộ filter cố định",
            "- KPI 'Tổng đã xử lý' = count tinker\n- Số dòng Excel = KPI\n- Bảng chi tiết + biểu đồ nhất quán",
            "Kiểm chứng chéo màn ↔ tinker ↔ Excel"),
        ("002", "Luồng từ chối → hiện ở báo cáo kèm lý do", "P1",
            "Ở màn gốc: từ chối 1 phiếu với lý do 'Sai thông tin KH'",
            "1. Từ chối phiếu ở màn duyệt gốc (nhập lý do)\n2. Backfill/log ghi nhận\n3. Mở báo cáo, lọc Từ chối",
            "Lý do: Sai thông tin KH",
            "- Dòng phiếu đó hiện badge Từ chối + '⤷ Sai thông tin KH'\n- Xuất Excel: cột Lý do từ chối = 'Sai thông tin KH'",
            "resolve(reject, reason) → note"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL      = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL  = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':22,'C':16,'D':42,'E':10,'F':32,'G':55,'H':22,'I':65,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}")
t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",              '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",         '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",     '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",        '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                  '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8

HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
           "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
           "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL; cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data,
              expected, business_note, "", "Not Executed", "Not Executed", "Not Executed", ""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(180, longest // 4))
    current_row += 1
    data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note, group="Phân quyền & truy cập")

ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}")
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)

# Thống kê nhanh
total_tc = data_row_idx
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Tổng TC: {total_tc}")
