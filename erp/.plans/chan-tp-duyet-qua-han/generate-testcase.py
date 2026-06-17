"""Generate testcase Excel — Chặn trưởng phòng duyệt khi NV có hàng quá hạn (ERP)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = "/Users/nguyentrancu/DEV/code/ERP-HRM/ERP/.plans/chan-tp-duyet-qua-han/testcase.xlsx"
SHEET_NAME   = "ChanTPDuyetQuaHan"
FEATURE_NAME = "Chặn trưởng phòng duyệt khi NV có hàng quá hạn (ERP)"
MODULE_NAME  = "Cấu hình đến hạn (Due Configs)"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
        "Chặn trưởng phòng (TP) thực hiện thao tác DUYỆT phiếu khi phòng ban do TP quản lý còn nhân viên "
        "có hàng QUÁ HẠN (hàng giữ / hàng mượn / hàng nhập thẳng). Mục tiêu buộc xử lý trả hàng quá hạn "
        "trước khi được duyệt các phiếu mới.\n"
        "Cơ chế: middleware checkDueConfigsManager:<Tên action> → DueConfigBlockService::isManagerBlocked()."),
    ("2. Đối tượng được tính / hiển thị (BỊ CHẶN khi đủ TẤT CẢ)",
        "► User đăng nhập KHÔNG phải role 'Super Admin'.\n"
        "► Action duyệt có bản ghi trong due_configs với tab = 2.\n"
        "► Action đó ĐƯỢC công ty của TP cấu hình bật trong company_due_configs (company_id + due_config_id).\n"
        "► TP có quản lý ≥1 phòng ban (employee_manage_departments: employee_id = TP, company_id = cty TP).\n"
        "► Tồn tại ≥1 nhân viên thuộc phòng quản lý (employee_infos.department_id ∈ phòng quản lý, cùng company_id) "
        "đang có ≥1 trong 3 loại quá hạn (xem mục 4)."),
    ("3. Đối tượng bị ẩn / KHÔNG CHẶN (cho qua, message nội bộ)",
        "► role 'Super Admin' → luôn cho qua (cả middleware lẫn service).\n"
        "► Action KHÔNG có trong due_configs tab=2 → cho qua (message '1').\n"
        "► Action có tab=2 nhưng company của TP CHƯA tick trong company_due_configs → cho qua (message '2').\n"
        "► TP không quản lý phòng nào (employee_manage_departments rỗng) → cho qua (message '3').\n"
        "► Phòng quản lý không có nhân viên nào → cho qua (message '4').\n"
        "► Có nhân viên nhưng KHÔNG ai quá hạn → cho qua (message '5').\n"
        "► Nhân viên quá hạn nhưng thuộc phòng/công ty KHÁC (không do TP này quản lý) → không tính."),
    ("4. Điều kiện 'quá hạn' (3 loại — bộ lọc thời gian)",
        "► Hàng GIỮ: prepick_details.qty > 0 VÀ expire_date < hôm nay, theo employee_id ∈ NV phòng quản lý.\n"
        "► Hàng MƯỢN: product_export_requests type = XUAT_MUON, borrow_status = DA_MUON, return_date < hôm nay, "
        "VÀ chi tiết base_exported_qty > borrow_returned_qty (còn nợ), theo per.created_by ∈ NV phòng quản lý.\n"
        "► Hàng NHẬP THẲNG (NXT): chỉ check khi company.overdue_date_export_direct > 0; "
        "product_import_direct_details.qty > 0 VÀ created_at < (now − overdue_date_export_direct ngày), "
        "theo company_id của TP + employee_id ∈ NV phòng quản lý.\n"
        "Mốc so sánh dùng date('Y-m-d') / Carbon::now()->subDays(N) tại thời điểm request."),
    ("5. Cấu trúc dữ liệu / quan hệ",
        "due_configs (tab=2, name=action) ↔ company_due_configs (company_id, due_config_id) ↔ "
        "employee_manage_departments (TP ↔ department_id) ↔ employee_infos (department_id, company_id) ↔ employees ↔ "
        "{ prepick_details | product_export_requests + product_export_request_details | product_import_direct_details }."),
    ("6. Quy tắc cộng dồn / quyết định",
        "blocked = true khi tồn tại ÍT NHẤT 1 nhân viên có ÍT NHẤT 1 loại quá hạn. "
        "Các loại quá hạn phát hiện được gộp vào message theo thứ tự: hàng giữ, hàng mượn, hàng nhập thẳng "
        "(phân tách bằng dấu phẩy). Chỉ cần 1 loại là đã chặn."),
    ("7. Phân quyền cấp",
        "Không dùng permission spatie theo cấp cho việc CHẶN. Việc chặn phụ thuộc cấu hình due_configs (tab=2) + "
        "company_due_configs + quan hệ quản lý phòng (employee_manage_departments).\n"
        "• Role 'Super Admin' — BYPASS hoàn toàn (không bao giờ bị chặn).\n"
        "• Một số route vẫn gắn kèm checkPermission riêng (vd 'Trưởng phòng duyệt yêu cầu đặt hàng') — "
        "middleware chặn chạy ĐỘC LẬP với quyền đó."),
    ("8. Cách tính / công thức quyết định",
        "► blocked = (due_configs tab=2 tồn tại) AND (company_due_configs đã tick) AND (TP quản lý ≥1 phòng) "
        "AND (phòng có ≥1 NV) AND (≥1 NV có ≥1 loại quá hạn).\n"
        "► hàng giữ quá hạn = EXISTS(prepick_details: qty>0 AND expire_date < today AND employee_id ∈ NV).\n"
        "► hàng mượn quá hạn = EXISTS(per type=XUAT_MUON, borrow_status=DA_MUON, return_date<today, "
        "base_exported_qty>borrow_returned_qty, created_by ∈ NV).\n"
        "► hàng NXT quá hạn = (overdue_date_export_direct>0) AND EXISTS(import_direct: qty>0, "
        "created_at < now−N, company_id, employee_id ∈ NV)."),
    ("9. Ghi chú đọc bảng",
        "► Khi bị chặn: AJAX trả {success:false, message, show_overdue_modal:true}; non-AJAX redirect back + flash + "
        "modal (layouts/app.blade.php:205).\n"
        "► Modal danh sách NV quá hạn lấy từ DueConfigBlockService::getOverdueEmployees (qua Api/DueConfigController), "
        "kèm nút xuất Excel (OverdueEmployeesExcel).\n"
        "► Cùng service được HRM gọi qua API /api/v1/due-configs/check-manager-block (xem testcase HRM riêng).\n"
        "► message chuẩn: 'Phòng ban bạn quản lý có nhân viên {các loại}. Không thể thực hiện thao tác này.'"),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "ĐIỀU KIỆN KÍCH HOẠT CHẶN (cấu hình due_configs)", [
        ("001", "Action có cấu hình + có NV quá hạn → CHẶN duyệt", "P0",
            "due_configs: name='Duyệt yêu cầu xuất hàng', tab=2 tồn tại.\n"
            "company_due_configs: company TP đã tick action này.\n"
            "TP quản lý phòng X; phòng X có NV A đang giữ 1 hàng expire_date = hôm qua, qty=2.",
            "1. Đăng nhập bằng TP (không phải Super Admin)\n"
            "2. Mở phiếu yêu cầu xuất hàng cần duyệt\n"
            "3. Bấm 'Duyệt' (managerApprove)",
            "TP: trưởng phòng X; Hôm nay = 11/06/2026",
            "- Thao tác bị chặn, KHÔNG duyệt được phiếu\n"
            "- Hiện thông báo cảnh báo + modal danh sách NV quá hạn\n"
            "- Message: 'Phòng ban bạn quản lý có nhân viên hàng giữ quá hạn. Không thể thực hiện thao tác này.'",
            "blocked=true khi đủ: config tab=2 + company tick + có NV quá hạn"),
        ("002", "Action KHÔNG có trong due_configs tab=2 → KHÔNG chặn", "P0",
            "Không có bản ghi due_configs name='Duyệt yêu cầu xuất hàng' với tab=2.\n"
            "Phòng TP vẫn có NV quá hạn.",
            "1. Đăng nhập TP\n2. Bấm 'Duyệt' phiếu",
            "TP: trưởng phòng X",
            "- Duyệt thành công (không chặn)\n"
            "- isManagerBlocked trả blocked=false (message nội bộ '1')",
            "Không có due_configs tab=2 → bỏ qua chặn"),
        ("003", "Company chưa tick company_due_configs → KHÔNG chặn", "P0",
            "due_configs tab=2 tồn tại nhưng company của TP CHƯA có bản ghi trong company_due_configs cho action này.\n"
            "Phòng TP vẫn có NV quá hạn.",
            "1. Đăng nhập TP\n2. Bấm 'Duyệt' phiếu",
            "TP công ty B (chưa tick)",
            "- Duyệt thành công (không chặn)\n"
            "- blocked=false (message nội bộ '2')",
            "Chưa cấu hình company_due_configs → không chặn"),
        ("004", "Cô lập theo công ty: chỉ company đã tick mới bị chặn", "P1",
            "Action tab=2; company A đã tick, company B chưa tick.\n"
            "TP-A (cty A) và TP-B (cty B) đều có NV quá hạn.",
            "1. TP-A bấm Duyệt → quan sát\n2. TP-B bấm Duyệt → quan sát",
            "TP-A: cty A; TP-B: cty B",
            "- TP-A: bị CHẶN\n- TP-B: duyệt THÀNH CÔNG (cty B chưa tick)",
            "Cấu hình chặn theo từng company_id"),
        ("005", "TP không quản lý phòng nào → KHÔNG chặn", "P1",
            "Config đầy đủ (tab=2 + company tick).\n"
            "employee_manage_departments của TP: rỗng (không quản lý phòng).",
            "1. Đăng nhập user không quản lý phòng\n2. Bấm 'Duyệt'",
            "User: nhân viên thường",
            "- Duyệt thành công\n- blocked=false (message nội bộ '3')",
            "Không quản lý phòng → không có phạm vi check"),
    ]),
    ("II", "3 LOẠI QUÁ HẠN", [
        ("001", "Hàng GIỮ quá hạn → chặn + message 'hàng giữ quá hạn'", "P0",
            "Config đủ. Phòng TP có NV A: prepick_details qty=3, expire_date = hôm qua.",
            "1. TP bấm 'Duyệt'\n2. Đọc message",
            "Hôm nay = 11/06/2026; expire_date = 10/06/2026",
            "- Bị chặn\n- Message chứa 'hàng giữ quá hạn'",
            "prepick_details: qty>0 AND expire_date < today"),
        ("002", "Hàng GIỮ đã trả hết (qty=0) → KHÔNG tính", "P0",
            "Phòng TP có NV A: prepick_details expire_date = hôm qua nhưng qty = 0 (đã trả hết). "
            "Không có loại quá hạn nào khác.",
            "1. TP bấm 'Duyệt'",
            "qty = 0",
            "- Duyệt thành công (không chặn)",
            "Điều kiện qty>0 — qty=0 bị loại"),
        ("003", "Hàng GIỮ expire_date = HÔM NAY → KHÔNG tính (biên)", "P1",
            "Phòng TP có NV A: prepick_details qty=2, expire_date = hôm nay (đúng bằng today).",
            "1. TP bấm 'Duyệt'",
            "expire_date = 11/06/2026 = hôm nay",
            "- Duyệt thành công (không chặn)\n- Vì điều kiện là '< today', không phải '<='",
            "Biên: expire_date < today (strictly less)"),
        ("004", "Hàng MƯỢN quá hạn còn nợ → chặn + 'hàng mượn quá hạn'", "P0",
            "Phòng TP có NV B tạo phiếu: per type=XUAT_MUON, borrow_status=DA_MUON, return_date = hôm qua; "
            "chi tiết base_exported_qty=10, borrow_returned_qty=4 (còn nợ 6).",
            "1. TP bấm 'Duyệt'\n2. Đọc message",
            "return_date = 10/06/2026; còn nợ 6",
            "- Bị chặn\n- Message chứa 'hàng mượn quá hạn'",
            "per: XUAT_MUON + DA_MUON + return_date<today + base_exported_qty>borrow_returned_qty"),
        ("005", "Hàng MƯỢN đã trả đủ → KHÔNG tính", "P0",
            "Phòng TP có NV B: per XUAT_MUON/DA_MUON, return_date = hôm qua nhưng "
            "base_exported_qty=10, borrow_returned_qty=10 (trả đủ). Không loại nào khác quá hạn.",
            "1. TP bấm 'Duyệt'",
            "base_exported_qty = borrow_returned_qty = 10",
            "- Duyệt thành công (không chặn)",
            "base_exported_qty > borrow_returned_qty là FALSE → không tính"),
        ("006", "Hàng MƯỢN return_date = hôm nay → KHÔNG tính (biên)", "P1",
            "Phòng TP có NV B: per XUAT_MUON/DA_MUON, return_date = hôm nay, còn nợ.",
            "1. TP bấm 'Duyệt'",
            "return_date = 11/06/2026 = hôm nay",
            "- Duyệt thành công (không chặn)",
            "Biên: return_date < today (strictly less)"),
        ("007", "Hàng MƯỢN sai trạng thái (chưa DA_MUON) → KHÔNG tính", "P2",
            "Phòng TP có NV B: per XUAT_MUON, return_date = hôm qua, còn nợ NHƯNG borrow_status ≠ DA_MUON.",
            "1. TP bấm 'Duyệt'",
            "borrow_status ≠ DA_MUON",
            "- Duyệt thành công (không chặn)",
            "Chỉ tính phiếu borrow_status = DA_MUON"),
        ("008", "Hàng NHẬP THẲNG quá hạn (config>0) → chặn 'hàng nhập thẳng quá hạn'", "P0",
            "company.overdue_date_export_direct = 30. Phòng TP có NV C: "
            "product_import_direct_details qty=5, created_at = 60 ngày trước, cùng company_id.",
            "1. TP bấm 'Duyệt'\n2. Đọc message",
            "overdue_date_export_direct = 30; created_at = 12/04/2026",
            "- Bị chặn\n- Message chứa 'hàng nhập thẳng quá hạn'",
            "import_direct: qty>0 AND created_at < now−30 ngày"),
        ("009", "overdue_date_export_direct = 0 → BỎ QUA check NXT", "P1",
            "company.overdue_date_export_direct = 0. Phòng TP có NV C: import_direct qty=5, created_at = 1 năm trước. "
            "Không loại nào khác quá hạn.",
            "1. TP bấm 'Duyệt'",
            "overdue_date_export_direct = 0",
            "- Duyệt thành công (không chặn)\n- Khối check NXT bị bỏ qua hoàn toàn",
            "Chỉ check NXT khi overdue_date_export_direct > 0"),
        ("010", "NV có CẢ 3 loại quá hạn → message gộp đủ 3", "P0",
            "Phòng TP có NV đang quá hạn cả: hàng giữ, hàng mượn, và hàng nhập thẳng (config>0).",
            "1. TP bấm 'Duyệt'\n2. Đọc message",
            "Đủ 3 loại quá hạn",
            "- Bị chặn\n- Message: 'Phòng ban bạn quản lý có nhân viên hàng giữ quá hạn, hàng mượn quá hạn, "
            "hàng nhập thẳng quá hạn. Không thể thực hiện thao tác này.'",
            "Gộp các loại theo thứ tự giữ, mượn, NXT"),
    ]),
    ("III", "PHẠM VI PHIẾU ÁP DỤNG (routes gắn middleware)", [
        ("001", "Duyệt yêu cầu xuất hàng (managerApprove / switch-board)", "P0",
            "Config đủ + có NV quá hạn.",
            "1. TP duyệt phiếu yêu cầu xuất hàng\n2. Thử cả switch-board-of-manager",
            "action='Duyệt yêu cầu xuất hàng'",
            "- Cả 2 route bị chặn (productExportRequest.managerApprove, switchBoardOfManager)",
            "Route web.php:1034-1035"),
        ("002", "Duyệt yêu cầu xuất giữ / điều chuyển / gia hạn hàng giữ", "P1",
            "Config tick các action 'Duyệt yêu cầu xuất giữ', 'Duyệt điều chuyển hàng giữ', "
            "'Duyệt gia hạn hàng giữ' + có NV quá hạn.",
            "1. TP duyệt từng phiếu tương ứng",
            "3 action hàng giữ",
            "- Cả 3 đều bị chặn khi có NV quá hạn",
            "web.php:1129, 1384, 1996"),
        ("003", "Duyệt yêu cầu xuất bán hàng mượn / gia hạn hàng mượn", "P1",
            "Config tick 'Duyệt yêu cầu xuất bán hàng mượn', 'Duyệt gia hạn hàng mượn' + có NV quá hạn.",
            "1. TP duyệt 2 phiếu tương ứng",
            "2 action hàng mượn",
            "- Cả 2 bị chặn",
            "web.php:1141, 2012"),
        ("004", "Duyệt yêu cầu nhập hàng (control/board/department approve)", "P1",
            "Config tick 'Duyệt yêu cầu nhập hàng' + có NV quá hạn.",
            "1. TP duyệt qua các route nhập hàng (controlBoardApprove, boardOfManagerApprove, "
            "switchBoardManager, controlDepartmentManagerApprove)",
            "action='Duyệt yêu cầu nhập hàng'",
            "- Tất cả route nhập hàng bị chặn",
            "web.php:1413-1420"),
        ("005", "Duyệt yêu cầu đặt hàng (RootOrderRequest approve/departmentApprove)", "P0",
            "Config tick 'Duyệt yêu cầu đặt hàng' + có NV quá hạn.",
            "1. TP duyệt RootOrderRequest (approve)\n2. departmentApprove (kèm checkPermission)",
            "action='Duyệt yêu cầu đặt hàng'",
            "- Cả 2 bị chặn; departmentApprove vẫn chặn dù có thêm checkPermission",
            "web.php:2594-2595, 2476"),
        ("006", "Duyệt hợp đồng mua / bán / quyết toán", "P1",
            "Config tick 'Duyệt hợp đồng mua', 'Duyệt hợp đồng bán', 'Duyệt quyết toán hợp đồng' + NV quá hạn.",
            "1. TP duyệt HĐ mua, HĐ bán (firmContract, annex), quyết toán",
            "3 nhóm action hợp đồng",
            "- Các phiếu duyệt HĐ bị chặn khi có NV quá hạn",
            "web.php:2762/3117/3883/3961/3988/4012/4564"),
        ("007", "Duyệt đề nghị thanh toán / giao việc / kết quả CV / lắp đặt", "P1",
            "Config tick các action tương ứng + NV quá hạn.",
            "1. TP duyệt đề nghị thanh toán, tạo phiếu giao việc, duyệt kết quả CV, duyệt yêu cầu lắp đặt",
            "Nhiều action Customercare/IncomeExpenditure",
            "- Các phiếu bị chặn khi có NV quá hạn",
            "web.php:6152/6184/6189/6245/6606/6607"),
    ]),
    ("IV", "TRƯỜNG HỢP KHÔNG CHẶN", [
        ("001", "Super Admin → luôn cho qua", "P0",
            "User có role 'Super Admin'. Phòng có NV quá hạn, config đủ.",
            "1. Đăng nhập Super Admin\n2. Bấm 'Duyệt' phiếu",
            "Role = Super Admin",
            "- Duyệt thành công (không chặn) — bypass cả ở middleware lẫn service",
            "hasRole('Super Admin') → return next / blocked=false"),
        ("002", "Phòng quản lý không có NV nào → KHÔNG chặn", "P2",
            "Config đủ; TP quản lý phòng X nhưng phòng X hiện không có nhân viên (cùng company_id).",
            "1. TP bấm 'Duyệt'",
            "Phòng X rỗng NV",
            "- Duyệt thành công\n- blocked=false (message nội bộ '4')",
            "employeeIds rỗng → không thể quá hạn"),
        ("003", "Có NV nhưng KHÔNG ai quá hạn → KHÔNG chặn", "P0",
            "Config đủ; phòng có NV nhưng tất cả hàng giữ/mượn/NXT đều trong hạn (hoặc đã trả hết).",
            "1. TP bấm 'Duyệt'",
            "Không NV nào quá hạn",
            "- Duyệt thành công\n- blocked=false (message nội bộ '5')",
            "overdueMessages rỗng → không chặn"),
        ("004", "NV quá hạn thuộc phòng KHÁC (không do TP quản lý) → KHÔNG chặn", "P0",
            "Config đủ; TP quản lý phòng X. NV quá hạn thuộc phòng Y (TP không quản lý).",
            "1. TP (phòng X) bấm 'Duyệt'",
            "NV quá hạn ở phòng Y",
            "- Duyệt thành công (NV phòng Y không nằm trong phạm vi)",
            "Chỉ tính NV thuộc managedDepartmentIds"),
        ("005", "NV quá hạn khác COMPANY với TP → KHÔNG tính", "P2",
            "TP công ty A quản lý phòng X. Có NV trùng department_id nhưng company_id = B đang quá hạn.",
            "1. TP-A bấm 'Duyệt'",
            "NV quá hạn company_id = B",
            "- Duyệt thành công\n- Vì lọc employee_infos.company_id = company TP",
            "Lấy NV theo department_id ∈ phòng AND company_id = cty TP"),
    ]),
    ("V", "THÔNG BÁO & MODAL NV QUÁ HẠN", [
        ("001", "Response AJAX khi bị chặn", "P0",
            "TP bị chặn (đủ điều kiện). Request gửi dạng AJAX/JSON.",
            "1. TP bấm 'Duyệt' (AJAX)\n2. Quan sát response",
            "Header: X-Requested-With / expectsJson",
            "- HTTP 200 body: {success:false, message:'...', show_overdue_modal:true}\n"
            "- FE hiện toastr warning + mở modal NV quá hạn",
            "CheckDueConfigsManager: expectsJson/ajax → json"),
        ("002", "Response non-AJAX khi bị chặn", "P1",
            "TP bị chặn. Request thường (không AJAX).",
            "1. TP submit duyệt (form thường)\n2. Quan sát",
            "Request thường",
            "- redirect back + flash message (alert-type=warning) + show_overdue_modal\n"
            "- layouts/app.blade.php:205 bật modal sau redirect",
            "redirect()->back()->with(show_overdue_modal)"),
        ("003", "Modal hiển thị danh sách NV quá hạn", "P0",
            "TP bị chặn; phòng có NV A (hàng giữ), NV B (hàng mượn), NV C (hàng NXT).",
            "1. Modal bật\n2. Quan sát danh sách",
            "3 NV quá hạn các loại khác nhau",
            "- Bảng liệt kê: mã NV (employee_code), tên NV (employee_name), loại quá hạn (overdue_types)\n"
            "- NV A: 'Hàng giữ'; NV B: 'Hàng mượn'; NV C: 'Hàng NXT'\n"
            "- NV có nhiều loại → gộp 'Hàng mượn, Hàng NXT, Hàng giữ'",
            "getOverdueEmployees() trả mã/tên/loại"),
        ("004", "Xuất Excel danh sách NV quá hạn", "P1",
            "Modal NV quá hạn đang hiển thị danh sách.",
            "1. Bấm 'Xuất Excel' (export=1)",
            "API overdue-employees?export=1",
            "- Tải file 'danh_sach_nhan_vien_qua_han.xlsx' đúng nội dung danh sách",
            "OverdueEmployeesExcel"),
        ("005", "Định dạng message chuẩn", "P1",
            "TP bị chặn do hàng mượn + hàng giữ.",
            "1. Đọc message trả về",
            "2 loại quá hạn",
            "- Đúng cú pháp: 'Phòng ban bạn quản lý có nhân viên hàng giữ quá hạn, hàng mượn quá hạn. "
            "Không thể thực hiện thao tác này.'",
            "implode(', ', overdueMessages)"),
    ]),
    ("VI", "EDGE CASES & BẢO MẬT", [
        ("001", "1 NV được 2 TP quản lý → cả 2 TP bị chặn", "P2",
            "NV A quá hạn; thuộc phòng X. Phòng X được cả TP1 và TP2 quản lý (2 bản ghi employee_manage_departments).",
            "1. TP1 bấm Duyệt\n2. TP2 bấm Duyệt",
            "NV A quá hạn, 2 TP quản lý",
            "- Cả TP1 và TP2 đều bị chặn",
            "Quản lý nhiều-nhiều qua employee_manage_departments"),
        ("002", "TP tự có hàng quá hạn và thuộc phòng mình quản lý", "P2",
            "TP vừa là người quản lý phòng X vừa có bản ghi NV trong phòng X đang quá hạn hàng mượn.",
            "1. TP bấm Duyệt",
            "TP cũng là NV phòng X, quá hạn",
            "- Bị chặn (TP nằm trong employeeIds của phòng mình quản lý)",
            "employeeIds gồm mọi NV phòng, kể cả TP nếu thuộc phòng"),
        ("003", "Action đúng tab nhưng sai tên → không khớp config", "P2",
            "Truyền actionName không trùng due_configs.name (sai chính tả) → không tìm thấy due_configs tab=2.",
            "1. (giả lập) route gắn action sai tên\n2. TP bấm Duyệt",
            "actionName lệch tên",
            "- Không chặn (blocked=false, message '1')",
            "Khớp chính xác name + tab=2"),
        ("004", "Đổi expire_date/return_date sang tương lai → hết chặn ngay", "P1",
            "Đang bị chặn do hàng giữ quá hạn. Gia hạn expire_date sang ngày mai (hoặc trả hết hàng).",
            "1. Gia hạn/trả hàng\n2. TP bấm Duyệt lại",
            "expire_date = ngày mai",
            "- Hết chặn, duyệt thành công ngay lần thử kế tiếp",
            "Check realtime theo today tại thời điểm request"),
    ]),
]

# ============================ BUILD (chuẩn skill) ============================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TL = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TC = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUM_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
EVEN_FILL = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':24,'C':14,'D':42,'E':9,'F':40,'G':46,'H':26,'I':62,'J':34,'K':16,'L':14,'M':14,'N':14,'O':20}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w

ws.cell(1,1,"MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22

for idx,(label,body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx,1,label); a.font=DESC_LABEL_FONT; a.fill=DESC_LABEL_FILL; a.alignment=WRAP_TL; a.border=BORDER
    b = ws.cell(idx,2,body); b.font=DESC_BODY_FONT; b.alignment=WRAP_TL; b.border=BORDER
    ws.merge_cells(start_row=idx,start_column=2,end_row=idx,end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n")*16 + 30)

t = ws.cell(11,1,f"Testcase _ {FEATURE_NAME}"); t.font=TITLE_FONT; t.fill=TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11,6,"TEST SUMMARY"); fs.font=Font(name="Calibri",size=12,bold=True,color="FFFFFF")
fs.fill=TITLE_FILL; fs.alignment=Alignment(vertical="center",horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11,"Số trường hợp kiểm thử đạt (P):",'=COUNTIF(L18:N500,"Passed")'),
    (12,"Số trường hợp kiểm thử không đạt (F):",'=COUNTIF(L18:N500,"Failed")'),
    (13,"Số trường hợp kiểm thử đang xem xét:",'=COUNTIF(L18:N500,"Pending")'),
    (14,"Số trường hợp kiểm thử chưa thực hiện:",'=COUNTIF(L18:N500,"Not Executed")'),
    (15,"Tổng số trường hợp kiểm thử:",'=COUNTIF(L18:N500,"<>")'),
]
for r,label,formula in summary_rows:
    lc = ws.cell(r,9,label); lc.font=Font(name="Calibri",size=11,bold=True); lc.fill=SUM_LABEL_FILL
    lc.alignment=Alignment(vertical="center",horizontal="right"); lc.border=BORDER
    ws.merge_cells(start_row=r,start_column=9,end_row=r,end_column=11)
    vc = ws.cell(r,12,formula); vc.font=Font(name="Calibri",size=11,bold=True); vc.fill=SUM_LABEL_FILL
    vc.alignment=Alignment(horizontal="center",vertical="center"); vc.border=BORDER
    ws.merge_cells(start_row=r,start_column=12,end_row=r,end_column=15)
    if r>11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8

HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện",
    "Test Data","Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
    "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i,h in enumerate(HEADERS, start=1):
    c = ws.cell(17,i,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=HEADER_ALIGN; c.border=BORDER
ws.row_dimensions[17].height = 36

current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row,3,title); cell.font=SECTION_FONT; cell.fill=SECTION_FILL
    cell.alignment=SECTION_ALIGN; cell.border=BORDER
    ws.merge_cells(start_row=current_row,start_column=3,end_row=current_row,end_column=15)
    for col in (1,2):
        ws.cell(current_row,col).fill=SECTION_FILL; ws.cell(current_row,col).border=BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id,function,priority,pre,steps,td,expected,note,group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME,group,tc_id,function,priority,pre,steps,td,expected,note,"",
        "Not Executed","Not Executed","Not Executed",""]
    fill = EVEN_FILL if data_row_idx % 2 == 1 else None
    for i,v in enumerate(values, start=1):
        c = ws.cell(current_row,i,v); c.font=Font(name="Calibri",size=11)
        c.alignment = WRAP_TC if i==5 else WRAP_TL; c.border=BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(40, min(220, longest//3))
    current_row += 1; data_row_idx += 1

roman_list = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman,title,tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = roman_list.index(roman) + 1
    for tc_num,func,prio,pre,steps,td,exp,note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id,func,prio,pre,steps,td,exp,note,group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row+100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"OK: {OUTPUT_FILE} | data rows 18-{current_row-1} | TCs={data_row_idx}")
