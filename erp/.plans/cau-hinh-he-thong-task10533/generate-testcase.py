"""Testcase Excel — Cấu hình hệ thống (task_10533): tái cấu trúc form 7 tab + lưu từng tab + VAT vận chuyển + ẩn loại hàng + lịch sử + xem chi tiết."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/cau-hinh-he-thong-task10533/testcase.xlsx"
SHEET_NAME   = "CauHinhHeThong"
FEATURE_NAME = "Cấu hình hệ thống (task_10533) — form nhiều tab, VAT vận chuyển, ẩn loại hàng, lịch sử"
MODULE_NAME  = "Cấu hình hệ thống"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Tái cấu trúc màn Cấu hình hệ thống (admin/configs/edit, task_10533): chia thành NHIỀU TAB, mỗi tab LƯU ĐỘC LẬP qua updateTab('{tab}'), ghi lịch sử per-tab (config_histories). Gồm 5 phần: (1) Update cấu hình VAT vận chuyển; (2) Update loại hàng hóa không hiển thị trên các phiếu; (3) Chỉnh form cấu hình theo file (chia tab); (4) Lịch sử chỉnh sửa cấu hình; (5) Xem chi tiết lịch sử."),
    ("2. Đối tượng được tính / hiển thị",
     "► 7 tab + field chính:\n"
     "  • general: logo, title, description\n"
     "  • category (Danh mục): max_borrow_date, max_prepick_date, consignment_holding_time, max_prepick_date_project_contract, warning_day, product_types, hidden_product_types[]\n"
     "  • business (Nghiệp vụ): customer_is_following, customer_taken_care, quotation_valid_days, project_quotation_valid_days, customer_register_expiry, customer_groups[], department_groups[], ràng buộc HĐ (is_equipment/is_repair/is_project/is_principle/is_dental_principle)\n"
     "  • warehouse (Kho): vat_delivery_trip_percent (Thuế vận tải)\n"
     "  • xnk: environment_tax\n"
     "  • accounting: coefficient_ecommerce_price, debt_calculation_date\n"
     "  • cskh: serial_product_types[], contract_rows[] (product_group × product_nature × quantity)\n"
     "► hidden_product_types: bảng dòng product_types[] × ticket_types[] (8 loại phiếu).\n"
     "► Lịch sử: theo từng tab, old_value → new_value, người sửa, thời gian."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Tab không hợp lệ (ngoài 7 VALID_TABS) → updateTab/histories trả lỗi 'Tab không hợp lệ'.\n"
     "► Loại phiếu ngoài 8 TICKET_TYPES không có trong hidden_product_types."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng (màn cấu hình). Riêng lịch sử sắp xếp theo created_at desc."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "configs (1 bản ghi) — các cột field theo tab. config_hidden_product_types (config_id, product_types[json], ticket_types[json]). contract_rows (config_id, product_groups[json], product_natures[json], quantity). config_histories (config_id, tab, field_name, old_value, new_value, updated_by, created_at)."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Lưu theo TAB (updateTab). Field thường: so cũ/mới, ghi config_histories mỗi field đổi. Tab category: config_hidden_product_types REPLACE (xóa hết của config rồi tạo lại từ rows). Tab cskh: contract_rows REPLACE. Không cộng dồn."),
    ("7. Phân quyền cấp",
     "• Truy cập màn cấu hình theo quyền vào menu cấu hình hệ thống; updateTab/histories dùng chung quyền màn (không có permission code riêng từng tab)."),
    ("8. Cách tính các ô thống kê",
     "— Không có ô thống kê."),
    ("9. Ghi chú đọc bảng",
     "Angular interpolation <% %>. Lưu từng tab độc lập (không lưu chung). Validation inline (is-invalid + invalid-feedback); hidden_product_types/contract_rows lỗi theo từng dòng theo $index. Lịch sử + xem chi tiết qua API configs/histories/{tab}."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP (form nhiều tab)", [
        ("001", "Truy cập màn Cấu hình hệ thống dạng nhiều tab", "P0",
         "User có quyền cấu hình; đã đăng nhập",
         "1. Vào menu Cấu hình hệ thống (admin/configs/edit)\n2. Quan sát danh sách tab",
         "—",
         "- Hiện đủ 7 tab: Chung, Danh mục, Nghiệp vụ, Kho, XNK, Kế toán, CSKH\n- Mỗi tab có nút Lưu riêng",
         "task_10533: form chia tab + updateTab/{tab}"),
        ("002", "Chuyển giữa các tab giữ dữ liệu mỗi tab", "P1",
         "Đang ở màn cấu hình",
         "1. Click lần lượt các tab\n2. Quan sát field hiển thị",
         "—",
         "- Mỗi tab hiển thị đúng nhóm field của tab đó; chuyển tab không mất dữ liệu đang xem",
         "TAB_FIELDS"),
        ("003", "Tab Chung (general) hiển thị logo/tiêu đề/mô tả", "P1",
         "configs có logo, title, description",
         "1. Mở tab Chung",
         "—",
         "- Hiện ô Logo, Tiêu đề, Mô tả với giá trị hiện tại",
         "fields general"),
        ("004", "Tab Danh mục hiển thị ngày mượn/giữ + tính chất hàng + bảng ẩn hàng", "P0",
         "—",
         "1. Mở tab Danh mục",
         "—",
         "- Hiện: Số ngày mượn/giữ tối đa, giữ hàng gửi, giữ HĐDA, ngày cảnh báo, Tính chất hàng hóa (product_types), bảng 'Loại hàng hóa không hiển thị trên các phiếu'",
         "fields category"),
        ("005", "Tab Kho hiển thị Thuế vận tải (VAT vận chuyển)", "P0",
         "vat_delivery_trip_percent = 8",
         "1. Mở tab Kho",
         "8",
         "- Ô 'Thuế vận tải' kèm %, giá trị = 8",
         "vat_delivery_trip_percent"),
        ("006", "Tab Nghiệp vụ / XNK / Kế toán / CSKH hiển thị đúng field", "P1",
         "—",
         "1. Mở lần lượt Nghiệp vụ, XNK, Kế toán, CSKH",
         "—",
         "- Nghiệp vụ: % theo dõi/chăm sóc, ngày hiệu lực báo giá, ràng buộc HĐ, nhóm KH/phòng ban\n- XNK: Hệ số thuế BVMT\n- Kế toán: hệ số giá TMĐT, ngày khai báo công nợ đầu kỳ\n- CSKH: hàng không bắt buộc serial, contract_rows",
         "TAB_FIELDS business/xnk/accounting/cskh"),
        ("007", "Bảng ẩn hàng hiển thị đúng dữ liệu đã lưu", "P0",
         "config_hidden_product_types có 1 dòng [Vật tư]×[Báo giá dự án]",
         "1. Tab Danh mục › quan sát bảng",
         "1 dòng sẵn",
         "- Dòng tick đúng product_types='Vật tư', ticket_types='Báo giá dự án'",
         "arrayInclude(...)"),
        ("008", "Nút xem Lịch sử chỉnh sửa theo tab", "P1",
         "Đã từng lưu thay đổi",
         "1. Mở lịch sử của 1 tab",
         "—",
         "- Gọi API configs/histories/{tab}; hiện danh sách thay đổi của tab đó",
         "histories(tab)"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH — Lưu cấu hình theo từng tab", [
        ("001", "[general] Lưu Logo/Tiêu đề/Mô tả", "P1",
         "—",
         "1. Tab Chung › sửa Tiêu đề\n2. Lưu",
         "Tiêu đề mới",
         "- updateTab('general') thành công; configs cập nhật; config_histories ghi field title old→new",
         "updateTab('general')"),
        ("002", "[category] Lưu ngày mượn/giữ + tính chất hàng", "P0",
         "—",
         "1. Tab Danh mục › Số ngày mượn tối đa=30, giữ=15\n2. Chọn product_types\n3. Lưu",
         "mượn=30, giữ=15",
         "- Lưu thành công; configs.max_borrow_date=30, max_prepick_date=15, product_types cập nhật; history ghi từng field đổi",
         "updateTab('category')"),
        ("003", "[VAT] Nhập % hợp lệ rồi Lưu tab Kho", "P0",
         "vat_delivery_trip_percent = 8",
         "1. Tab Kho › Thuế vận tải=10\n2. Lưu",
         "10",
         "- configs.vat_delivery_trip_percent=10; history tab=warehouse, old=8, new=10",
         "updateTab('warehouse')"),
        ("004", "[ẨN HÀNG] Thêm 1 dòng rồi Lưu tab Danh mục", "P0",
         "Bảng ẩn hàng trống",
         "1. Thêm dòng: Loại hàng=[Vật tư], Loại phiếu=[Báo giá dự án]\n2. Lưu",
         "[Vật tư]×[bao_gia_du_an]",
         "- Tạo 1 config_hidden_product_types; history ghi thay đổi category",
         "delete + create ConfigHiddenProductType"),
        ("005", "[ẨN HÀNG] Thêm nhiều dòng / nhiều loại hàng × nhiều phiếu", "P1",
         "—",
         "1. Dòng1: Vật tư×[Báo giá dự án]\n2. Dòng2: [Vật tư,Hàng hóa]×[Yêu cầu đặt hàng, Hỏi giá]\n3. Lưu",
         "2 dòng",
         "- 2 dòng config_hidden_product_types đúng mảng product_types/ticket_types",
         "casts array"),
        ("006", "[ẨN HÀNG] Xóa dòng rồi Lưu (replace)", "P0",
         "Bảng có 2 dòng",
         "1. Xóa dòng 2\n2. Lưu",
         "còn 1 dòng",
         "- config_hidden_product_types chỉ còn 1 dòng (xóa hết + tạo lại)",
         "replace"),
        ("007", "[business] Lưu % theo dõi/chăm sóc + ngày hiệu lực báo giá", "P1",
         "—",
         "1. Tab Nghiệp vụ › customer_is_following=80, quotation_valid_days=30\n2. Lưu",
         "80 / 30",
         "- Lưu thành công; history ghi field đổi",
         "updateTab('business')"),
        ("008", "[business] Lưu nhóm KH / phòng ban không ràng buộc + cờ ràng buộc HĐ", "P2",
         "—",
         "1. Chọn customer_groups, department_groups; bật is_equipment\n2. Lưu",
         "—",
         "- Lưu thành công; cấu hình ràng buộc HĐ cập nhật",
         "business fields"),
        ("009", "[xnk] Lưu Hệ số thuế BVMT", "P1",
         "—",
         "1. Tab XNK › environment_tax=1000\n2. Lưu",
         "1000",
         "- configs.environment_tax=1000; history",
         "updateTab('xnk')"),
        ("010", "[accounting] Lưu hệ số giá TMĐT + ngày công nợ đầu kỳ", "P1",
         "—",
         "1. Tab Kế toán › coefficient_ecommerce_price=1.1, debt_calculation_date=01/01/2026\n2. Lưu",
         "1.1 / 01/01/2026",
         "- Lưu thành công; history ghi 2 field",
         "updateTab('accounting')"),
        ("011", "[cskh] Lưu hàng không bắt buộc serial", "P2",
         "—",
         "1. Tab CSKH › chọn serial_product_types\n2. Lưu",
         "—",
         "- Lưu thành công",
         "updateTab('cskh')"),
        ("012", "[cskh] Lưu contract_rows (ràng buộc số lượng)", "P1",
         "—",
         "1. Thêm dòng contract_rows: product_group=[Nhóm A], quantity=5\n2. Lưu",
         "quantity=5",
         "- contract_rows tạo lại (replace) đúng dòng; history",
         "contract_rows replace"),
    ]),
    ("VI", "EDGE CASES & VALIDATION (theo từng tab)", [
        ("001", "[general] Bỏ trống Tiêu đề / Mô tả / Logo", "P0",
         "—",
         "1. Xóa Tiêu đề\n2. Lưu tab Chung",
         "title rỗng",
         "- Lỗi inline 'Bắt buộc nhập' (title); logo trống → 'Bắt buộc chọn'; description trống → 'Bắt buộc nhập'",
         "rules general required"),
        ("002", "[general] Tiêu đề > 255 ký tự", "P2",
         "—",
         "1. Nhập title 256 ký tự\n2. Lưu",
         "256 ký tự",
         "- Lỗi 'Nhập tối đa 255 ký tự'",
         "title max:255"),
        ("003", "[category] Ngày mượn/giữ bỏ trống / <1 / >999", "P0",
         "—",
         "1. max_borrow_date rỗng; max_prepick_date=0; consignment_holding_time=1000\n2. Lưu",
         "rỗng / 0 / 1000",
         "- max_borrow_date: 'Bắt buộc nhập'\n- max_prepick_date<1: 'Bắt buộc nhập' (min)\n- consignment_holding_time: integer, max 999 lỗi",
         "category numeric/integer min1 max999"),
        ("004", "[category] warning_day ngoài 0–100", "P1",
         "—",
         "1. warning_day=150\n2. Lưu",
         "150",
         "- Lỗi (max=100)",
         "warning_day min0 max100"),
        ("005", "[category] product_types để trống", "P0",
         "—",
         "1. Bỏ chọn hết Tính chất hàng hóa\n2. Lưu",
         "product_types rỗng",
         "- Lỗi 'Bắt buộc phải chọn'",
         "product_types required|array|min:1"),
        ("006", "[ẨN HÀNG] Dòng thiếu loại hàng hóa", "P0",
         "Thêm dòng chỉ chọn loại phiếu",
         "1. Lưu",
         "product_types rỗng dòng đó",
         "- Lỗi inline 'Bắt buộc chọn loại hàng hóa' đúng dòng",
         "hidden_product_types.*.product_types required|min:1"),
        ("007", "[ẨN HÀNG] Dòng thiếu loại phiếu", "P0",
         "Thêm dòng chỉ chọn loại hàng",
         "1. Lưu",
         "ticket_types rỗng dòng đó",
         "- Lỗi inline 'Bắt buộc chọn loại phiếu' đúng dòng",
         "hidden_product_types.*.ticket_types required|min:1"),
        ("008", "[ẨN HÀNG] Lỗi hiện đúng dòng khi nhiều dòng", "P1",
         "Dòng1 hợp lệ, dòng2 thiếu loại phiếu",
         "1. Lưu",
         "dòng2 lỗi",
         "- Chỉ dòng2 lỗi (key hidden_product_types.1.ticket_types)",
         "errors theo $index"),
        ("009", "[VAT] required / numeric / min0 / max100", "P0",
         "—",
         "1. Lần lượt: rỗng; 'abc'; -5; 150\n2. Lưu",
         "rỗng/abc/-5/150",
         "- rỗng→'Bắt buộc nhập'; 'abc'→'Không hợp lệ'; -5→min lỗi; 150→max lỗi",
         "vat_delivery_trip_percent required|numeric|min0|max100"),
        ("010", "[business] customer_register_expiry < 1 / không số", "P1",
         "—",
         "1. customer_register_expiry=0 rồi 'abc'\n2. Lưu",
         "0 / abc",
         "- 0→'Không hợp lệ' (min1); abc→'Không hợp lệ' (numeric)",
         "customer_register_expiry required|numeric|min:1"),
        ("011", "[business] quotation_valid_days / project... bỏ trống", "P1",
         "—",
         "1. Xóa trống quotation_valid_days\n2. Lưu",
         "rỗng",
         "- Lỗi 'Bắt buộc nhập' (required, min1)",
         "quotation_valid_days required|min:1"),
        ("012", "[xnk] environment_tax required/numeric", "P1",
         "—",
         "1. Rỗng rồi 'abc'\n2. Lưu",
         "rỗng/abc",
         "- 'Bắt buộc nhập' / 'Không hợp lệ'",
         "environment_tax required|numeric|min0"),
        ("013", "[accounting] debt_calculation_date sai định dạng", "P1",
         "—",
         "1. Nhập ngày không hợp lệ\n2. Lưu",
         "ngày rác",
         "- 'Không hợp lệ' (date); coefficient_ecommerce_price rỗng → 'Bắt buộc nhập'",
         "debt_calculation_date required|date; coefficient required|numeric"),
        ("014", "[cskh] contract_rows quantity < 1 / rỗng", "P1",
         "—",
         "1. Dòng contract_rows quantity=0\n2. Lưu",
         "quantity=0",
         "- Lỗi 'Phải lớn hơn 0' (min:1); rỗng → 'Bắt buộc nhập'",
         "contract_rows.*.quantity required|integer|min:1"),
        ("015", "[cskh] contract_rows dùng CẢ nhóm hàng + tính chất hàng", "P0",
         "Có dòng dùng product_group và dòng khác dùng product_nature",
         "1. Dòng1 chọn Nhóm hàng; Dòng2 chọn Tính chất hàng\n2. Lưu",
         "mix nhóm + tính chất",
         "- Lỗi 'Không được sử dụng cả Nhóm hàng hóa và Tính chất hàng hóa trong các dòng khác nhau'",
         "closure validate contract_rows"),
    ]),
    ("VII", "LỊCH SỬ CHỈNH SỬA & XEM CHI TIẾT (mục #4, #5)", [
        ("001", "Lịch sử ghi khi lưu (old → new)", "P0",
         "vat_delivery_trip_percent = 8",
         "1. Đổi VAT thành 12, Lưu\n2. Xem lịch sử tab Kho",
         "8 → 12",
         "- config_histories thêm dòng: tab=warehouse, field_name='vat_delivery_trip_percent', old_value=8, new_value=12, updated_by=user",
         "ConfigHistory::create"),
        ("002", "Xem chi tiết lịch sử theo tab (API histories/{tab})", "P0",
         "Tab category đã có thay đổi",
         "1. Gọi xem lịch sử tab Danh mục",
         "tab=category",
         "- Trả danh sách: updater (mã_họtên), field_name, old_value, new_value, created_at (d/m/Y H:i), sắp xếp mới nhất trước",
         "histories('category') responseSuccess"),
        ("003", "Lịch sử chỉ hiện thay đổi của TAB đang xem", "P1",
         "Đã đổi cả tab Kho và tab Danh mục",
         "1. Xem lịch sử tab Kho\n2. Xem lịch sử tab Danh mục",
         "—",
         "- Lịch sử tab Kho chỉ có field của warehouse; tab Danh mục chỉ có field category — không lẫn",
         "where tab = {tab}"),
        ("004", "Tên người sửa = mã + họ tên", "P2",
         "User TP-X (mã NV) lưu thay đổi",
         "1. Xem lịch sử",
         "User TP-X",
         "- updater hiển thị '<info.code>_<fullname>'",
         "updater.info.code _ fullname"),
        ("005", "Tab không hợp lệ khi xem lịch sử", "P1",
         "—",
         "1. Gọi histories với tab='abc'",
         "tab='abc'",
         "- Trả lỗi 'Tab không hợp lệ'",
         "in_array(tab, VALID_TABS)"),
        ("006", "Không đổi gì → không phát sinh lịch sử", "P1",
         "—",
         "1. Mở 1 tab, Lưu mà không đổi field nào",
         "—",
         "- Không thêm dòng config_histories cho lần lưu này",
         "chỉ ghi khi field thay đổi"),
    ]),
    ("VIII", "CÔ LẬP TAB & E2E (cấu hình tác động nghiệp vụ)", [
        ("001", "Lưu tab này KHÔNG ảnh hưởng tab khác", "P0",
         "Các tab đều có dữ liệu",
         "1. Đổi+Lưu tab Kho\n2. Kiểm tab Danh mục/Nghiệp vụ",
         "—",
         "- updateTab chỉ cập nhật field của tab gửi lên; các tab khác giữ nguyên (kể cả hidden_product_types/contract_rows)",
         "updateTab cô lập theo tab"),
        ("002", "Tab không hợp lệ khi lưu", "P1",
         "—",
         "1. Gọi updateTab('abc')",
         "tab='abc'",
         "- Trả lỗi 'Tab không hợp lệ', không lưu",
         "VALID_TABS"),
        ("003", "[E2E ẨN HÀNG] Ẩn loại hàng X trên 'Báo giá dự án'", "P0",
         "Cấu hình [Vật tư]×[Báo giá dự án]; có hàng 'Vật tư'",
         "1. Lưu cấu hình\n2. Mở Báo giá dự án › chọn hàng",
         "Hàng 'Vật tư'",
         "- Hàng 'Vật tư' KHÔNG hiển thị/chọn được trên Báo giá dự án; phiếu khác/loại hàng khác vẫn hiện",
         "config_hidden_product_types lọc theo (product_type, ticket_type)"),
        ("004", "[E2E VAT] VAT vận chuyển áp vào cước", "P0",
         "vat_delivery_trip_percent = 10",
         "1. Lưu VAT=10\n2. Tính cước chuyến giao hàng",
         "10%",
         "- Tiền VAT vận chuyển = 10%",
         "vat_delivery_trip_percent dùng khi tính cước"),
        ("005", "[E2E] Ngày mượn/giữ tối đa áp khi tạo phiếu mượn/giữ", "P1",
         "max_borrow_date=30, max_prepick_date=15",
         "1. Lưu\n2. Tạo yêu cầu mượn/giữ hàng",
         "30 / 15",
         "- Hạn mượn ≤ 30 ngày, hạn giữ ≤ 15 ngày theo cấu hình",
         "config áp xuống luồng mượn/giữ"),
        ("006", "[E2E] Ngày hiệu lực báo giá áp khi lập báo giá", "P1",
         "quotation_valid_days=30",
         "1. Lưu\n2. Lập báo giá mới",
         "30",
         "- Hiệu lực báo giá mặc định = ngày lập + 30",
         "quotation_valid_days"),
        ("007", "[E2E] Ràng buộc HĐ theo cờ business", "P2",
         "Bật is_equipment (ràng buộc HĐ bán hàng)",
         "1. Lưu\n2. Thao tác cần HĐ bán",
         "—",
         "- Hệ thống áp ràng buộc HĐ tương ứng cờ đã bật",
         "is_equipment/is_repair/..."),
    ]),
]

# ===================== STYLES + BUILD =====================
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True); DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF"); TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True); SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79"); SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':26,'C':16,'D':46,'E':10,'F':32,'G':55,'H':24,'I':65,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)
t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1); ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF"); fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center"); ws.row_dimensions[11].height = 28
for r, label, formula in [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8
for i, h in enumerate(["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
        "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế","trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"], start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36
current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2): ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data, expected, business_note, "",
              "Not Executed","Not Executed","Not Executed",""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(220, longest // 3)); current_row += 1; data_row_idx += 1
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp, note, group=title)
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE} | data rows 18-{current_row-1}")
