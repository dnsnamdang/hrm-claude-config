"""Helper dựng file Excel testcase chuẩn ERP TPE (dùng chung cho nhiều feature)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22,
}

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]


def build(output_file, sheet_name, feature_name, module_name, description_block, sections):
    """sections = list[(roman, title, [tc,...])]; tc = (tc_num, func, prio, pre, steps, td, exp, note)"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
    ws.merge_cells("B1:O1")
    ws.row_dimensions[1].height = 22

    for idx, (label, body) in enumerate(description_block, start=2):
        a = ws.cell(idx, 1, label)
        a.font = DESC_LABEL_FONT
        a.fill = DESC_LABEL_FILL
        a.alignment = WRAP_TOP_LEFT
        a.border = BORDER
        b = ws.cell(idx, 2, body)
        b.font = DESC_BODY_FONT
        b.alignment = WRAP_TOP_LEFT
        b.border = BORDER
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
        ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

    t = ws.cell(11, 1, f"Testcase _ {feature_name}")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("B11:E11")
    ws.merge_cells("F11:H11")
    fs = ws.cell(11, 6, "TEST SUMMARY")
    fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    fs.fill = TITLE_FILL
    fs.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[11].height = 28

    summary_rows = [
        (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
        (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
        (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
        (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
        (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
    ]
    for r, label, formula in summary_rows:
        lc = ws.cell(r, 9, label)
        lc.font = SUMMARY_LABEL_FONT
        lc.fill = SUMMARY_LABEL_FILL
        lc.alignment = Alignment(vertical="center", horizontal="right")
        lc.border = BORDER
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        vc = ws.cell(r, 12, formula)
        vc.font = SUMMARY_VALUE_FONT
        vc.fill = SUMMARY_LABEL_FILL
        vc.alignment = SUMMARY_VALUE_ALIGN
        vc.border = BORDER
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
        if r > 11:
            ws.row_dimensions[r].height = 22

    ws.row_dimensions[16].height = 8

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(17, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[17].height = 36

    state = {"row": 18, "idx": 0}

    def write_section(title):
        r = state["row"]
        cell = ws.cell(r, 3, title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = SECTION_ALIGN
        cell.border = BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=15)
        for col in (1, 2):
            ws.cell(r, col).fill = SECTION_FILL
            ws.cell(r, col).border = BORDER
        ws.row_dimensions[r].height = 26
        state["row"] += 1

    def write_tc(tc_id, func, prio, pre, steps, td, exp, note, group):
        r = state["row"]
        values = [module_name, group, tc_id, func, prio, pre, steps, td,
                  exp, note, "", "Not Executed", "Not Executed", "Not Executed", ""]
        fill = DATA_FONT_FILL_EVEN if state["idx"] % 2 == 1 else None
        for i, v in enumerate(values, start=1):
            c = ws.cell(r, i, v)
            c.font = Font(name="Calibri", size=11)
            c.alignment = WRAP_TOP_CENTER if i == 5 else WRAP_TOP_LEFT
            c.border = BORDER
            if fill:
                c.fill = fill
        lines = max(str(v).count("\n") + 1 for v in values)
        ws.row_dimensions[r].height = max(30, min(200, lines * 16 + 14))
        state["row"] += 1
        state["idx"] += 1

    for roman, title, tcs in sections:
        write_section(f"{roman}. {title}")
        sec_idx = ROMAN.index(roman) + 1
        for tc_num, func, prio, pre, steps, td, exp, note in tcs:
            tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
            write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

    dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                        allow_blank=True, showDropDown=False)
    dv.add(f"L18:N{state['row'] + 100}")
    ws.add_data_validation(dv)

    wb.save(output_file)
    print(f"✅ {output_file} — data rows 18-{state['row']-1}")
