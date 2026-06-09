"""Builder dùng chung cho testcase phân quyền xem danh sách (Task / Issue)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL      = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL  = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A': 18, 'B': 24, 'C': 14, 'D': 42, 'E': 9,
              'F': 40, 'G': 38, 'H': 24, 'I': 55, 'J': 40,
              'K': 18, 'L': 14, 'M': 14, 'N': 14, 'O': 18}
HEADERS = ["Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
           "Tiền điều kiện", "Bước thực hiện", "Test Data",
           "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
           "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3", "Ghi chú"]
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]


def build_testcase(output_file, sheet_name, feature_name, module_name,
                   description_block, role_tcs, sections):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
    ws.merge_cells("B1:O1")
    ws.row_dimensions[1].height = 22

    for idx, (label, body) in enumerate(description_block, start=2):
        a = ws.cell(idx, 1, label)
        a.font, a.fill, a.alignment, a.border = DESC_LABEL_FONT, DESC_LABEL_FILL, WRAP_TOP_LEFT, BORDER
        b = ws.cell(idx, 2, body)
        b.font, b.alignment, b.border = DESC_BODY_FONT, WRAP_TOP_LEFT, BORDER
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
        ws.row_dimensions[idx].height = max(40, body.count("\n") * 15 + 30)

    t = ws.cell(11, 1, f"Testcase _ {feature_name}")
    t.font, t.fill = TITLE_FONT, TITLE_FILL
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("B11:E11")
    ws.merge_cells("F11:H11")
    fs = ws.cell(11, 6, "TEST SUMMARY")
    fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    fs.fill = TITLE_FILL
    fs.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[11].height = 28

    summary_rows = [
        (11, "Số trường hợp kiểm thử đạt (P):",          '=COUNTIF(L18:N500,"Passed")'),
        (12, "Số trường hợp kiểm thử không đạt (F):",     '=COUNTIF(L18:N500,"Failed")'),
        (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
        (14, "Số trường hợp kiểm thử chưa thực hiện:",    '=COUNTIF(L18:N500,"Not Executed")'),
        (15, "Tổng số trường hợp kiểm thử:",              '=COUNTIF(L18:N500,"<>")'),
    ]
    for r, label, formula in summary_rows:
        lc = ws.cell(r, 9, label)
        lc.font, lc.fill, lc.border = SUMMARY_LABEL_FONT, SUMMARY_LABEL_FILL, BORDER
        lc.alignment = Alignment(vertical="center", horizontal="right")
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        vc = ws.cell(r, 12, formula)
        vc.font, vc.fill, vc.alignment, vc.border = SUMMARY_VALUE_FONT, SUMMARY_LABEL_FILL, SUMMARY_VALUE_ALIGN, BORDER
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
        if r > 11:
            ws.row_dimensions[r].height = 22

    ws.row_dimensions[16].height = 8
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(17, i, h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, BORDER
    ws.row_dimensions[17].height = 36

    state = {"row": 18, "idx": 0}

    def write_section_row(title):
        r = state["row"]
        cell = ws.cell(r, 3, title)
        cell.font, cell.fill, cell.alignment, cell.border = SECTION_FONT, SECTION_FILL, SECTION_ALIGN, BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=15)
        for col in (1, 2):
            ws.cell(r, col).fill = SECTION_FILL
            ws.cell(r, col).border = BORDER
        ws.row_dimensions[r].height = 26
        state["row"] += 1

    def write_tc(tc_id, function, priority, pre, steps, td, exp, note, group=""):
        r = state["row"]
        values = [module_name, group, tc_id, function, priority,
                  pre, steps, td, exp, note, "",
                  "Not Executed", "Not Executed", "Not Executed", ""]
        fill = DATA_FONT_FILL_EVEN if state["idx"] % 2 == 1 else None
        for i, v in enumerate(values, start=1):
            c = ws.cell(r, i, v)
            c.font = Font(name="Calibri", size=11)
            c.alignment = WRAP_TOP_CENTER if i == 5 else WRAP_TOP_LEFT
            c.border = BORDER
            if fill:
                c.fill = fill
        longest = max(len(str(v)) for v in values)
        ws.row_dimensions[r].height = max(40, min(200, longest // 3))
        state["row"] += 1
        state["idx"] += 1

    if role_tcs:
        write_section_row("Phân quyền & truy cập")
        for suffix, func, prio, pre, steps, td, exp, note in role_tcs:
            write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note, group="Phân quyền & truy cập")

    for roman, title, tcs in sections:
        write_section_row(f"{roman}. {title}")
        sec_idx = ROMAN.index(roman) + 1
        for tc_num, func, prio, pre, steps, td, exp, note in tcs:
            write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp, note, group=title)

    dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                        allow_blank=True, showDropDown=False)
    dv.add(f"L18:N{state['row'] + 100}")
    ws.add_data_validation(dv)

    wb.save(output_file)
    print(f"✅ {output_file} | Tổng TC: {state['idx']} | rows 18-{state['row']-1}")
