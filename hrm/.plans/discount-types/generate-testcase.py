"""Generate testcase Excel cho feature Danh mục Loại chiết khấu (module Giao việc / Assign)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = ".plans/discount-types/testcase.xlsx"
SHEET_NAME   = "LoaiChietKhau"
FEATURE_NAME = "Danh mục Loại chiết khấu (Giao việc)"
MODULE_NAME  = "Loại chiết khấu"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý danh mục các Loại chiết khấu (VD: 'CK khách hàng thân thiết') trong module Giao việc (Assign). "
     "Danh mục này được dùng trong Báo giá (Quotation): bảng quotation_discounts liên kết báo giá với một "
     "loại chiết khấu (discount_type_id) kèm phần trăm/số tiền. Chức năng: xem danh sách, lọc/tìm kiếm, "
     "tạo, sửa, khóa/mở khóa, duyệt, xóa (đơn + nhiều)."),
    ("2. Đối tượng được tính / hiển thị",
     "► Danh sách hiển thị tất cả loại chiết khấu (mọi trạng thái) theo bộ lọc.\n"
     "► 3 trạng thái: 1=Hoạt động (STATUS_ACTIVE), 2=Khóa (STATUS_INACTIVE), 3=Chờ duyệt (STATUS_PENDING).\n"
     "► Loại CK tạo từ màn danh mục → mặc định Hoạt động (1).\n"
     "► Loại CK tạo nhanh từ màn Báo giá (quickCreate) → Chờ duyệt (3), gắn quotation_id, phải duyệt mới dùng chung."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Không ẩn theo trạng thái — danh sách hiển thị cả Hoạt động/Khóa/Chờ duyệt (lọc tùy chọn).\n"
     "► Lưu ý: dropdown chọn loại CK trong màn Báo giá (getAll) chỉ lấy status=Hoạt động (+ Chờ duyệt của đúng "
     "báo giá đó) — KHÁC với danh sách màn danh mục này."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "► Lọc theo 'Ngày tạo từ' (created_from) và 'Ngày tạo đến' (created_to), so trên cột created_at (whereDate).\n"
     "► Không có bộ lọc thời gian nào khác."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Bảng discount_types: id, code (string 30, UNIQUE, tự sinh), name (NOT NULL, KHÔNG unique), "
     "status (tinyint default 1), quotation_id (nullable — chỉ có khi tạo nhanh từ báo giá), created_by, updated_by, timestamps.\n"
     "► Quan hệ: hasMany quotation_discounts qua discount_type_id (FK KHÔNG cascade). Đây là căn cứ quyết định có cho xóa hay không.\n"
     "► Không có cây phân cấp."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Không áp dụng (danh mục phẳng, không thống kê). Tên loại CK ĐƯỢC PHÉP trùng (không ràng buộc unique trên name)."),
    ("7. Phân quyền cấp",
     "► KHÔNG phân quyền theo cấp tổ chức.\n"
     "► Một quyền duy nhất: 'Quản lý danh mục loại chiết khấu' (PermissionsTableSeeder id=1090, group='Danh mục').\n"
     "► Quyền này gắn middleware checkPermission cho TẤT CẢ thao tác: index (xem list), updateOrCreate (tạo/sửa), "
     "lock, unlock, approve, destroy, delete_by_ids.\n"
     "► Các route getAll, quick-create, show KHÔNG gắn checkPermission (dùng nội bộ cho màn Báo giá)."),
    ("8. Cách tính các ô thống kê",
     "Không áp dụng — màn này KHÔNG có ô thống kê / stats đầu trang."),
    ("9. Ghi chú đọc bảng",
     "► Cột bảng: checkbox | STT | Mã | Tên loại chiết khấu (+ Người tạo/Ngày tạo) | Cập nhật (+ bởi ai) | Trạng thái (+ nút khóa/mở khóa) | Hành động.\n"
     "► Mã tự sinh dạng CK-{năm}-{maxId+1 pad 5}, vd CK-2026-00001 (không nhập tay, sửa không đổi mã).\n"
     "► Nút Sửa disable khi status=Khóa (2) hoặc !is_can_edit; nút/checkbox Xóa disable khi !is_can_delete (đã dùng trong báo giá); "
     "nút Khóa/Mở khóa ẩn khi status=Chờ duyệt (3); nút Duyệt chỉ hiện khi is_can_approve (status=3).\n"
     "► Phân trang server-side mặc định 10/trang; sort cột code/name/updatedAt, mặc định id giảm dần.\n"
     "► Tạo/Sửa qua MODAL (id modal-discount-type), không phải trang riêng; tạo & sửa dùng chung 1 endpoint POST (phân biệt qua field id).\n"
     "► Filter lưu localStorage key 'assign_discount_types', hết hạn 10 phút.\n"
     "► Validation toàn bộ từ BE (422); FE chỉ hiện lỗi inline trường name + toast."),
]

# Không phân quyền theo cấp → bỏ section TC-ROLE, đưa TC quyền vào section I & VII
HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Load trang danh sách thành công", "P0",
         "User có quyền 'Quản lý danh mục loại chiết khấu'; có ≥1 loại CK",
         "1. Truy cập /assign/discount-types\n2. Quan sát layout",
         "User có quyền quản lý",
         "- Trang tiêu đề 'Danh mục loại chiết khấu' load với panel bộ lọc + bảng danh sách\n"
         "- KHÔNG có khối stats đầu trang",
         "Permission: Quản lý danh mục loại chiết khấu (GET index)"),
        ("002", "Hiển thị đủ các cột bảng", "P0",
         "Danh sách có ≥1 loại CK",
         "1. Truy cập danh sách\n2. Quan sát tiêu đề cột",
         "—",
         "- Cột: checkbox | STT | Mã | Tên loại chiết khấu | Cập nhật | Trạng thái | Hành động\n"
         "- Cột Tên có sub-info 'Người tạo: … - Ngày tạo: …'\n- Cột Cập nhật hiển thị thời gian + 'bởi {người sửa}'",
         "—"),
        ("003", "Nút 'Tạo mới' hiển thị khi có quyền", "P0",
         "User có quyền 'Quản lý danh mục loại chiết khấu'",
         "1. Truy cập danh sách\n2. Quan sát góc trên",
         "User có quyền",
         "- Nút 'Tạo mới' hiển thị, bấm mở modal Thêm loại chiết khấu",
         "Permission: hasAPermission('Quản lý danh mục loại chiết khấu')"),
        ("004", "Ẩn 'Tạo mới' khi không có quyền", "P1",
         "User KHÔNG có quyền 'Quản lý danh mục loại chiết khấu'",
         "1. Truy cập danh sách (nếu vào được)\n2. Quan sát",
         "User không có quyền",
         "- Nút 'Tạo mới' KHÔNG hiển thị",
         "Permission: ẩn nút theo hasAPermission"),
        ("005", "Truy cập khi không có quyền — chặn API index", "P0",
         "User KHÔNG có quyền 'Quản lý danh mục loại chiết khấu'",
         "1. Truy cập trực tiếp /assign/discount-types",
         "User không có quyền",
         "- API GET index bị chặn (403) → không thấy danh sách",
         "Permission: middleware checkPermission GET /assign/discount_types"),
        ("006", "Trạng thái hiển thị đúng nhãn + màu", "P1",
         "Có loại CK ở cả 3 trạng thái",
         "1. Quan sát cột Trạng thái",
         "CK A=1, CK B=2, CK C=3",
         "- A: 'Hoạt động'; B: 'Khóa'; C: 'Chờ duyệt' (đúng nhãn theo status 1/2/3)",
         "BR — status 1=Hoạt động, 2=Khóa, 3=Chờ duyệt"),
    ]),
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm theo mã loại CK", "P0",
         "Có loại CK mã 'CK-2026-00001'",
         "1. Nhập 'CK-2026-00001' vào ô tìm kiếm\n2. Bấm Tìm kiếm",
         "keyword = CK-2026-00001",
         "- Danh sách còn loại CK có mã khớp",
         "BR — keyword LIKE trên code/name (escapeLikeKeyword)"),
        ("002", "Tìm theo tên loại CK", "P0",
         "Có loại CK tên 'CK khách hàng thân thiết'",
         "1. Nhập 'thân thiết' vào ô tìm kiếm\n2. Bấm Tìm",
         "keyword = thân thiết",
         "- Danh sách còn loại CK có tên chứa 'thân thiết'",
         "BR — keyword LIKE trên name"),
        ("003", "Lọc theo Trạng thái", "P1",
         "Có loại CK đủ 3 trạng thái",
         "1. Chọn Trạng thái = Chờ duyệt\n2. Quan sát",
         "status = 3",
         "- Danh sách chỉ còn loại CK Chờ duyệt",
         "Options: 1=Hoạt động, 2=Khóa, 3=Chờ duyệt"),
        ("004", "Lọc theo Người tạo", "P1",
         "Có loại CK tạo bởi nhiều người",
         "1. Chọn Người tạo = NV A\n2. Quan sát",
         "created_by = NV A",
         "- Danh sách còn loại CK do NV A tạo",
         "—"),
        ("005", "Lọc theo khoảng Ngày tạo", "P1",
         "Có loại CK tạo các ngày khác nhau",
         "1. Chọn Ngày tạo từ = 01/06/2026, đến = 15/06/2026\n2. Tìm",
         "created_from=01/06/2026, created_to=15/06/2026",
         "- Danh sách còn loại CK có created_at trong khoảng (bao gồm 2 đầu mút, whereDate)",
         "BR — whereDate created_from/created_to"),
        ("006", "Kết hợp nhiều bộ lọc", "P1",
         "Dữ liệu đa dạng",
         "1. keyword='CK' + Trạng thái=Hoạt động + Người tạo=NV A\n2. Tìm",
         "keyword=CK, status=1, created_by=NV A",
         "- Danh sách thỏa ĐỒNG THỜI cả 3 điều kiện",
         "—"),
        ("007", "Đặt lại bộ lọc", "P2",
         "Đang áp dụng vài filter",
         "1. Bấm 'Đặt lại'\n2. Quan sát",
         "—",
         "- Các filter về initialState, danh sách trở lại toàn bộ",
         "BR — reset về initialStateForm"),
        ("008", "Filter được lưu localStorage", "P2",
         "Đã áp filter, chưa quá 10 phút",
         "1. Áp filter\n2. Reload trang (F5)\n3. Quan sát",
         "key = assign_discount_types",
         "- Sau reload, filter vẫn được khôi phục (hết hạn sau 10 phút)",
         "BR — filterStateMixin localStorage TTL 10'"),
        ("009", "Tìm kiếm không có kết quả", "P1",
         "Danh sách có dữ liệu",
         "1. Nhập keyword 'ZZZ999'\n2. Tìm",
         "keyword=ZZZ999",
         "- Bảng rỗng, không lỗi",
         "—"),
    ]),
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "—",
         "Màn danh mục loại chiết khấu không có khối thống kê đầu trang",
         "—",
         "- Không có ô stats nào để kiểm thử",
         "Không áp dụng"),
    ]),
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Phân trang chuyển trang", "P0",
         "Có > 10 loại CK (vd 23)",
         "1. Xem trang 1\n2. Sang trang 2, 3",
         "Tổng 23, per_page=10",
         "- Trang 1/2: 10 dòng; trang 3: 3 dòng; meta total=23, last_page=3",
         "BR — phân trang server-side mặc định 10/trang"),
        ("002", "Sắp xếp mặc định id giảm dần", "P1",
         "Có nhiều loại CK",
         "1. Truy cập lần đầu",
         "mặc định orderBy id desc",
         "- Loại CK mới tạo (id lớn) ở trên cùng",
         "BR — mặc định id desc"),
        ("003", "Sắp xếp theo cột Mã", "P2",
         "Có nhiều loại CK",
         "1. Bấm tiêu đề cột 'Mã'\n2. Quan sát",
         "sort_field=code",
         "- Dữ liệu sắp theo mã (asc/desc)",
         "Sortable: code"),
        ("004", "Sắp xếp theo cột Tên", "P2",
         "Có nhiều loại CK",
         "1. Bấm tiêu đề cột 'Tên loại chiết khấu'",
         "sort_field=name",
         "- Dữ liệu sắp theo tên (asc/desc)",
         "Sortable: name"),
        ("005", "Sắp xếp theo cột Cập nhật", "P2",
         "Có nhiều loại CK",
         "1. Bấm tiêu đề cột 'Cập nhật'",
         "sort_field=updatedAt → updated_at",
         "- Dữ liệu sắp theo thời gian cập nhật",
         "Sortable whitelist: updatedAt→updated_at"),
        ("006", "Chọn nhiều dòng bằng checkbox", "P1",
         "Có ≥3 loại CK xóa được (is_can_delete=true)",
         "1. Tick checkbox 3 dòng\n2. Quan sát toolbar",
         "—",
         "- Toolbar trái hiện nút 'Xóa' (bulk) + 'Bỏ chọn'",
         "—"),
        ("007", "Checkbox disable với dòng không xóa được", "P1",
         "Có loại CK đã dùng trong báo giá (is_can_delete=false)",
         "1. Quan sát checkbox dòng đó",
         "CK đã dùng",
         "- Checkbox bị disable, không chọn được dòng đó",
         "BR — checkbox disable khi !is_can_delete"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH (CRUD / LOCK / DUYỆT)", [
        ("001", "Tạo loại CK hợp lệ", "P0",
         "User có quyền; modal Thêm đang mở",
         "1. Bấm 'Tạo mới'\n2. Nhập Tên = 'CK khách hàng thân thiết'\n3. Bấm Lưu",
         "name='CK khách hàng thân thiết'",
         "- Lưu thành công, toast thành công, modal đóng\n"
         "- Bản ghi mới có mã tự sinh CK-2026-NNNNN, trạng thái Hoạt động, xuất hiện danh sách",
         "BR — tạo từ danh mục → status=Hoạt động, code=getNextCode()"),
        ("002", "Form tạo chỉ có trường Tên, không có ô Mã", "P1",
         "Modal Thêm đang mở",
         "1. Quan sát các trường trong modal",
         "—",
         "- Chỉ có 1 ô nhập: Tên loại CK (bắt buộc)\n- KHÔNG có ô Mã (mã chỉ hiện & disabled khi Sửa)",
         "BR — code v-if=id và luôn disabled"),
        ("003", "Lưu & Tiếp tục khi tạo mới", "P2",
         "Modal Thêm đang mở",
         "1. Nhập Tên hợp lệ\n2. Bấm 'Lưu & Tiếp tục'",
         "name='CK mùa hè'",
         "- Lưu thành công, modal KHÔNG đóng, form reset để nhập tiếp loại CK mới",
         "BR — nút 'Lưu & Tiếp tục' chỉ có khi tạo mới"),
        ("004", "Sửa tên loại CK (status Hoạt động)", "P0",
         "Có loại CK status=Hoạt động, is_can_edit=true",
         "1. Bấm 'Sửa'\n2. Đổi Tên\n3. Lưu",
         "name cũ → 'CK VIP'",
         "- Lưu thành công, tên cập nhật, MÃ GIỮ NGUYÊN (ô mã disabled)\n- Người/ngày cập nhật đổi",
         "BR — sửa chỉ update name; code không đổi"),
        ("005", "Không sửa được loại CK đang Khóa", "P0",
         "Có loại CK status=Khóa (2)",
         "1. Quan sát nút 'Sửa' của dòng đó\n2. (thử) gọi POST update",
         "status=2",
         "- Nút 'Sửa' bị disable trên FE\n- BE: nếu cố update → trả 404 'Dữ liệu đã thay đổi, vui lòng tải lại'",
         "BR — isCanEdit: status != INACTIVE; FE chặn status===2"),
        ("006", "Xem chi tiết loại CK (readonly)", "P1",
         "Có loại CK bất kỳ",
         "1. Bấm 'Xem'\n2. Quan sát modal",
         "—",
         "- Modal mở tiêu đề 'Xem loại chiết khấu', ô Tên disabled, không có nút Lưu",
         "—"),
        ("007", "Khóa loại CK (Hoạt động → Khóa)", "P0",
         "Có loại CK status=Hoạt động",
         "1. Bấm nút Khóa tại cột trạng thái\n2. Xác nhận trong modal",
         "status 1 → 2",
         "- Sau xác nhận: status chuyển Khóa, nút đổi icon thành Mở khóa",
         "BR — lock set status=INACTIVE; isCanLockUpdate: status != PENDING"),
        ("008", "Mở khóa loại CK (Khóa → Hoạt động)", "P0",
         "Có loại CK status=Khóa",
         "1. Bấm nút Mở khóa\n2. Xác nhận",
         "status 2 → 1",
         "- status chuyển Hoạt động, quotation_id reset = null",
         "BR — unlock set status=ACTIVE, quotation_id=null"),
        ("009", "Không hiện nút Khóa khi đang Chờ duyệt", "P1",
         "Có loại CK status=Chờ duyệt (3)",
         "1. Quan sát cột trạng thái dòng đó",
         "status=3",
         "- Nút Khóa/Mở khóa bị ẩn (không cho khóa khi Chờ duyệt)\n- BE: gọi lock → 400",
         "BR — nút ẩn khi status===3; isCanLockUpdate=false khi PENDING"),
        ("010", "Duyệt loại CK Chờ duyệt", "P0",
         "Có loại CK status=Chờ duyệt (tạo nhanh từ báo giá), is_can_approve=true",
         "1. Bấm 'Duyệt'\n2. Xác nhận",
         "status 3 → 1",
         "- status chuyển Hoạt động, quotation_id=null → áp dụng cho mọi báo giá",
         "BR — approve: chỉ PENDING; set ACTIVE + quotation_id=null"),
        ("011", "Nút Duyệt chỉ hiện với loại CK Chờ duyệt", "P1",
         "Có loại CK Hoạt động và Chờ duyệt",
         "1. Quan sát cột hành động 2 dòng",
         "CK A=1, CK C=3",
         "- Chỉ dòng Chờ duyệt (C) có nút 'Duyệt'; dòng Hoạt động (A) không có",
         "BR — nút hiện khi is_can_approve (status=3)"),
        ("012", "Duyệt loại CK không phải Chờ duyệt — BE chặn", "P0",
         "Loại CK status=Hoạt động",
         "1. Gọi trực tiếp GET .../approve cho CK Hoạt động",
         "status=1",
         "- BE trả 400 'Chỉ có thể duyệt loại chiết khấu đang Chờ duyệt'",
         "BR — approve chặn khi !isCanApprove"),
        ("013", "Xóa đơn loại CK chưa dùng", "P0",
         "Loại CK chưa được dùng trong báo giá (is_can_delete=true)",
         "1. Bấm nút Xóa (đỏ)\n2. Xác nhận xóa",
         "CK chưa dùng",
         "- Xóa cứng thành công, biến mất khỏi danh sách",
         "BR — destroy hard delete khi isCanDelete"),
        ("014", "Không xóa được loại CK đã dùng trong báo giá", "P0",
         "Loại CK đã có ≥1 bản ghi quotation_discounts (is_can_delete=false)",
         "1. Quan sát nút Xóa dòng đó\n2. (thử) gọi DELETE trực tiếp",
         "CK đã dùng, status=1",
         "- Nút Xóa disable, tooltip 'Loại CK đã được sử dụng, không thể xoá'\n- BE: DELETE → 400 cùng message",
         "BR — isCanDelete: false khi tồn tại quotationDiscounts (trừ khi PENDING)"),
        ("015", "Loại CK Chờ duyệt luôn xóa được", "P1",
         "Loại CK status=Chờ duyệt dù có quotation_id",
         "1. Bấm Xóa\n2. Xác nhận",
         "status=3",
         "- Xóa thành công (PENDING luôn cho xóa)",
         "BR — isCanDelete: PENDING → true"),
        ("016", "Xóa nhiều (bulk) — tất cả xóa được", "P1",
         "Chọn 3 loại CK đều is_can_delete=true",
         "1. Tick 3 dòng\n2. Bấm 'Xóa' bulk\n3. Xác nhận",
         "3 CK chưa dùng",
         "- Cả 3 bị xóa, danh sách cập nhật",
         "BR — delete_by_ids"),
        ("017", "Xóa nhiều — có dòng không xóa được", "P0",
         "Chọn 3 loại CK, trong đó 1 đã dùng trong báo giá",
         "1. Tick 3 dòng (nếu chọn được)\n2. Xóa bulk\n3. Quan sát",
         "2 xóa được + 1 đã dùng",
         "- Bỏ qua bản không xóa được, trả lỗi liệt kê tên 'Không thể xoá: … (đã được sử dụng)'\n"
         "- 2 bản còn lại bị xóa",
         "BR — deleteByIds bỏ qua bản !isCanDelete + liệt kê"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Bỏ trống Tên loại CK", "P0",
         "Modal Thêm đang mở",
         "1. Để trống Tên\n2. Bấm Lưu",
         "name=trống",
         "- BE 422, lỗi inline tại Tên 'Tên loại chiết khấu là bắt buộc' (V2BaseError) + toast",
         "Rule: name required"),
        ("002", "Tên vượt quá 255 ký tự", "P1",
         "Modal Thêm đang mở",
         "1. Nhập Tên 256 ký tự\n2. Lưu",
         "name = chuỗi 256 ký tự",
         "- BE 422, lỗi 'Tên loại chiết khấu tối đa 255 ký tự'",
         "Rule: name max:255"),
        ("003", "Cho phép trùng tên loại CK", "P0",
         "Đã có loại CK tên 'CK VIP'",
         "1. Tạo mới Tên = 'CK VIP'\n2. Lưu",
         "name='CK VIP' (đã tồn tại)",
         "- Lưu THÀNH CÔNG (không ràng buộc unique trên name), tạo bản ghi mới mã khác",
         "BR — name KHÔNG unique (cả DB lẫn FormRequest)"),
        ("004", "Tên chỉ chứa khoảng trắng", "P1",
         "Modal Thêm",
         "1. Nhập Tên = '   ' (toàn space)\n2. Lưu",
         "name='   '",
         "- BE xử lý: nếu trim rỗng → 422 lỗi bắt buộc (kiểm chứng hành vi)",
         "Rule: name required|string"),
        ("005", "Mã tự sinh tăng tuần tự", "P2",
         "Loại CK lớn nhất hiện có id=5 (mã CK-2026-00005)",
         "1. Tạo mới loại CK\n2. Kiểm tra mã",
         "maxId=5",
         "- Mã mới = CK-2026-00006 (CK-{năm}-{maxId+1 pad 5})",
         "BR — getNextCode()"),
        ("006", "Sửa loại CK đã bị người khác khóa (race condition)", "P2",
         "User mở modal Sửa khi CK còn Hoạt động; người khác khóa nó trước khi user Lưu",
         "1. Mở Sửa\n2. (CK bị khóa ở nơi khác)\n3. Bấm Lưu",
         "status đổi 1→2 giữa chừng",
         "- BE trả 404 'Dữ liệu đã thay đổi, vui lòng tải lại'",
         "BR — updateOrCreate trả 404 khi !isCanEdit"),
        ("007", "Lưu thất bại hiển thị lỗi inline + toast", "P1",
         "Modal Thêm, để trống Tên",
         "1. Bấm Lưu",
         "name=trống",
         "- Toast 'Bạn chưa nhập đầy đủ thông tin'\n- Ô Tên viền lỗi + message inline",
         "BR — FE gán error từ response.data.errors (422)"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Gọi API tạo/sửa khi không có quyền", "P0",
         "User KHÔNG có quyền 'Quản lý danh mục loại chiết khấu'",
         "1. Gọi trực tiếp POST /assign/discount_types với payload hợp lệ",
         "—",
         "- BE chặn 403 (middleware checkPermission)",
         "Permission: checkPermission POST /assign/discount_types"),
        ("002", "Gọi API khóa/mở khóa khi không có quyền", "P0",
         "User không có quyền quản lý",
         "1. Gọi GET .../{id}/lock hoặc /unlock",
         "—",
         "- BE chặn 403",
         "Permission: checkPermission lock/unlock"),
        ("003", "Gọi API duyệt/xóa khi không có quyền", "P1",
         "User không có quyền quản lý",
         "1. Gọi GET .../{id}/approve, DELETE .../{id}, POST .../delete_by_ids",
         "—",
         "- BE chặn 403 tất cả",
         "Permission: checkPermission approve/destroy/delete_by_ids"),
        ("004", "delete_by_ids với id không tồn tại", "P2",
         "User có quyền",
         "1. Gọi delete_by_ids với ids chứa id không tồn tại",
         "ids=[999999]",
         "- BE 422 validate 'mỗi id phải exists' (hoặc bỏ qua tùy rule)",
         "Rule: ids array min:1, exists"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Luồng tạo → sửa → khóa → mở khóa → xóa", "P0",
         "User có quyền; loại CK chưa dùng trong báo giá",
         "1. Tạo loại CK mới\n2. Sửa tên\n3. Khóa\n4. Mở khóa\n5. Xóa",
         "name='CK test E2E'",
         "- Tạo OK (Hoạt động) → sửa tên OK (mã giữ nguyên) → khóa OK → mở khóa OK → xóa OK (biến mất)",
         "E2E — vòng đời đầy đủ"),
        ("002", "Luồng tạo nhanh từ báo giá → duyệt → áp dụng", "P0",
         "Loại CK tạo nhanh từ màn Báo giá (status Chờ duyệt, gắn quotation_id)",
         "1. Vào danh mục, lọc Trạng thái=Chờ duyệt\n2. Bấm Duyệt loại CK đó\n3. Kiểm tra trạng thái",
         "status 3 → 1",
         "- Sau duyệt: status Hoạt động, quotation_id=null → hiển thị trong dropdown chọn CK ở mọi báo giá (getAll)",
         "E2E — quickCreate → approve"),
        ("003", "Luồng không xóa được khi đã gắn báo giá", "P1",
         "Loại CK Hoạt động đã được 1 báo giá sử dụng",
         "1. Vào danh mục\n2. Thử xóa loại CK đó",
         "CK đã dùng",
         "- Không xóa được (nút disable + BE 400). Phải gỡ khỏi báo giá trước mới xóa được",
         "E2E — ràng buộc usage"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT  = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}")
t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL; fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",              '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",         '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",     '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",        '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                  '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [
        MODULE_NAME, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed", "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    nlines = max(str(v).count("\n") for v in values)
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, nlines * 16 + 30 if nlines else longest // 4))
    current_row += 1
    data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                    allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}")
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Rows: 1-10 description, 11-15 summary, 17 header, 18-{current_row-1} data")
print(f"   Tổng data rows: {data_row_idx}")
