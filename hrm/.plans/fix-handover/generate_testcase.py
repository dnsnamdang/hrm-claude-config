import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ban giao cong viec"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Bàn giao công việc (Assign/Handover)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'Bàn giao công việc'
PREFIX = 'HDV'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. DANH SÁCH PHIẾU (/assign/handover)
# =============================================
write_section_row('I. DANH SÁCH PHIẾU (/assign/handover)')

write_tc(1, 1,
    'Hiển thị trang khi truy cập',
    'P0',
    'User đã đăng nhập, có quyền xem bàn giao',
    '1. Vào menu Giao việc > Bàn giao',
    '',
    'Tiêu đề: "Danh sách bàn giao công việc"\nBộ lọc + bảng hiển thị đầy đủ cột: STT, Mã phiếu, NV bàn giao, Phòng ban, Ngày BG, Lý do, Số item, Trạng thái, Người tạo, Hành động')

write_tc(1, 2,
    'Hiển thị badge trạng thái đúng màu',
    'P0',
    'Có phiếu ở đủ 5 trạng thái',
    '1. Quan sát cột Trạng thái',
    '5 phiếu status 1–5',
    'Nháp = xám (#64748B)\nChờ duyệt = cam (#D97706)\nĐã duyệt = xanh (#2563EB)\nTừ chối = đỏ (#B91C1C)\nHoàn tất = xanh lá (#16A34A)')

write_tc(1, 3,
    'Filter theo Công ty',
    'P0',
    '>=2 công ty có phiếu bàn giao',
    '1. Chọn 1 công ty trong dropdown\n2. Nhấn Tìm kiếm',
    'Công ty A',
    'Chỉ hiển thị phiếu thuộc Công ty A\nPhòng ban chỉ hiện các PB của Công ty A')

write_tc(1, 4,
    'Filter theo Phòng ban',
    'P0',
    'Công ty có nhiều phòng ban có phiếu',
    '1. Chọn công ty\n2. Chọn phòng ban\n3. Tìm kiếm',
    'Phòng Dự án',
    'Chỉ hiển thị phiếu thuộc Phòng Dự án')

write_tc(1, 5,
    'Filter theo Nhân viên bàn giao',
    'P1',
    'Hệ thống có nhiều phiếu của nhiều NV',
    '1. Chọn NV cụ thể\n2. Tìm kiếm',
    'NV: Nguyễn Văn A',
    'Chỉ hiển thị phiếu có employee_id = A')

write_tc(1, 6,
    'Filter theo Trạng thái',
    'P0',
    'Có đủ phiếu ở 5 trạng thái',
    '1. Chọn "Chờ duyệt"\n2. Tìm kiếm',
    'status = 2',
    'Chỉ hiển thị phiếu Chờ duyệt')

write_tc(1, 7,
    'Filter theo Lý do',
    'P1',
    'Có phiếu với nhiều lý do khác nhau',
    '1. Chọn lý do "Nghỉ việc"\n2. Tìm kiếm',
    'reason = 1',
    'Chỉ hiển thị phiếu lý do Nghỉ việc')

write_tc(1, 8,
    'Filter khoảng ngày bàn giao',
    'P1',
    'Có phiếu rải trong nhiều ngày',
    '1. Nhập Từ ngày / Đến ngày\n2. Tìm kiếm',
    'Từ 01/04/2026, Đến 15/04/2026',
    'Chỉ hiển thị phiếu có handover_date nằm trong khoảng')

write_tc(1, 9,
    'Search theo mã phiếu',
    'P1',
    'Phiếu BG.2026.0001 tồn tại',
    '1. Nhập mã vào ô tìm kiếm\n2. Enter',
    'Keyword: BG.2026.0001',
    'Hiển thị đúng phiếu có code khớp')

write_tc(1, 10,
    'Phân trang - Chuyển trang',
    'P1',
    'Có > 10 phiếu',
    '1. Click trang 2 trong pagination',
    '',
    'Bảng load dữ liệu trang 2, pagination cập nhật')

write_tc(1, 11,
    'Phân trang - Đổi số dòng/trang',
    'P1',
    'Có nhiều phiếu',
    '1. Chọn 20 trong dropdown số dòng/trang',
    '',
    'Bảng hiển thị tối đa 20 phiếu/trang')

write_tc(1, 12,
    'Sort theo ngày bàn giao',
    'P1',
    'Bảng có nhiều phiếu',
    '1. Click header cột "Ngày BG" 2 lần',
    '',
    'Lần 1: sort tăng dần; lần 2: sort giảm dần')

write_tc(1, 13,
    'Action - Xem chi tiết',
    'P0',
    'Có phiếu trong danh sách',
    '1. Click icon xem ở dòng phiếu',
    '',
    'Điều hướng tới /assign/handover/{id}, hiển thị chi tiết đúng phiếu')

write_tc(1, 14,
    'Action - Sửa chỉ hiện cho Nháp/Từ chối và chỉ với người tạo',
    'P0',
    'Có phiếu Nháp của current user, phiếu Đã duyệt của user khác',
    '1. Quan sát cột Hành động ở từng dòng',
    'Status 1 (own), Status 3 (others)',
    'Phiếu Nháp của mình: icon Sửa hiện\nPhiếu Đã duyệt hoặc của người khác: icon Sửa ẩn')

write_tc(1, 15,
    'Action - Xoá chỉ hiện cho phiếu Nháp và chỉ với người tạo',
    'P0',
    'Phiếu Nháp của current user',
    '1. Click icon Xoá\n2. Confirm',
    'status = 1, created_by = current',
    'Phiếu bị xoá khỏi DB, biến mất khỏi bảng\nPhiếu status khác: nút Xoá không hiện')


# =============================================
# II. TẠO/SỬA PHIẾU (/assign/handover/add)
# =============================================
write_section_row('II. TẠO/SỬA PHIẾU (/assign/handover/add)')

write_tc(2, 1,
    'Form hiển thị đủ field',
    'P0',
    'User có quyền tạo phiếu',
    '1. Vào /assign/handover/add',
    '',
    'Form có: NV bàn giao, Ngày BG, Lý do (dropdown 5), Ghi chú, Bảng items (tab Task/Issue)\nNút: Lưu nháp, Gửi duyệt')

write_tc(2, 2,
    'Validate bắt buộc NV bàn giao',
    'P0',
    'Form trống',
    '1. Không chọn NV\n2. Click Lưu nháp',
    '',
    'Báo lỗi: "Vui lòng chọn nhân viên bàn giao"\nKhông gọi API')

write_tc(2, 3,
    'Validate bắt buộc Ngày bàn giao',
    'P0',
    'Form đã chọn NV',
    '1. Để trống Ngày BG\n2. Lưu nháp',
    '',
    'Báo lỗi: "Vui lòng chọn ngày bàn giao"')

write_tc(2, 4,
    'Validate bắt buộc Lý do',
    'P0',
    'Đã nhập NV, Ngày',
    '1. Bỏ trống Lý do\n2. Lưu nháp',
    '',
    'Báo lỗi: "Vui lòng chọn lý do"')

write_tc(2, 5,
    'Validate Ghi chú bắt buộc khi lý do "Khác"',
    'P1',
    'Đã chọn NV, ngày',
    '1. Chọn Lý do = "Khác"\n2. Để trống Ghi chú\n3. Lưu',
    'reason = 5',
    'Báo lỗi: "Vui lòng nhập ghi chú khi chọn lý do Khác"')

write_tc(2, 6,
    'Validate phiếu phải có ít nhất 1 item (BR1)',
    'P0',
    'Info phiếu đã nhập đầy đủ',
    '1. Không chọn item nào\n2. Lưu nháp',
    '',
    'Báo lỗi: "Phiếu phải có ít nhất 1 item"')

write_tc(2, 7,
    'Load danh sách task/issue khả dụng theo NV',
    'P0',
    'NV A có 3 task + 2 issue ở trạng thái Đang thực hiện',
    '1. Chọn NV A\n2. Mở bảng items',
    'Task Nháp/Huỷ/Hoàn tất không tính',
    'Bảng hiển thị đúng 3 task + 2 issue\nTask ở trạng thái Nháp/Huỷ/Hoàn tất bị loại')

write_tc(2, 8,
    'Loại items đã nằm trong phiếu active khác (BR3)',
    'P0',
    'Task X đã có trong phiếu khác status 1/2/3',
    '1. Chọn NV\n2. Mở bảng items',
    '',
    'Task X KHÔNG xuất hiện trong list khả dụng')

write_tc(2, 9,
    'Chọn nhiều items + bulk assign receiver',
    'P0',
    'Đã load bảng items',
    '1. Tick nhiều item\n2. Chọn receiver B cho tất cả (nút bulk assign)',
    'receiver = B',
    'Tất cả items được tick đều gán receiver = B')

write_tc(2, 10,
    'Assign receiver khác nhau cho từng item',
    'P1',
    'Đã tick nhiều items',
    '1. Set từng dòng receiver khác nhau',
    'Item 1 → B, Item 2 → C',
    'Mỗi item giữ receiver riêng')

write_tc(2, 11,
    'Nhập ghi chú cho từng item',
    'P2',
    'Đã chọn item',
    '1. Nhập handover_note cho 1 item',
    '',
    'handover_note lưu đúng khi Lưu')

write_tc(2, 12,
    'Lưu nháp thành công',
    'P0',
    'Đã nhập đủ info phiếu và ≥1 item có receiver',
    '1. Click Lưu nháp',
    '',
    'Tạo phiếu status = 1 (Nháp)\nCode format BG.YYYY.NNNN\nĐiều hướng về danh sách hoặc chi tiết')

write_tc(2, 13,
    'Gửi duyệt - tất cả items phải có receiver (BR2)',
    'P0',
    'Có 1 item chưa có receiver',
    '1. Click Gửi duyệt',
    '',
    'Báo lỗi: "Tất cả items phải có receiver trước khi gửi duyệt"\nKhông đổi status')

write_tc(2, 14,
    'Gửi duyệt thành công',
    'P0',
    'Phiếu đủ info, mọi item có receiver',
    '1. Click Gửi duyệt\n2. Confirm',
    '',
    'Phiếu chuyển status = 2 (Chờ duyệt)\nNoti gửi tới TP (section IX.001)\nLog action = "submit"')

write_tc(2, 15,
    'Sửa phiếu Nháp',
    'P0',
    'Phiếu status = 1, user là người tạo',
    '1. Vào Sửa\n2. Đổi ghi chú\n3. Lưu',
    '',
    'Cập nhật thành công\nGiữ nguyên status = 1')

write_tc(2, 16,
    'Sửa phiếu bị từ chối + gửi lại (BR8)',
    'P0',
    'Phiếu status = 4, user là người tạo',
    '1. Vào Sửa\n2. Thấy banner lý do từ chối\n3. Chỉnh items/info\n4. Click Gửi duyệt lại',
    'reject_reason có sẵn',
    'Banner hiển thị đúng reject_reason\nSau khi gửi: status = 2\nLog action = "resubmit"')

write_tc(2, 17,
    'Không cho sửa phiếu Chờ duyệt/Đã duyệt/Hoàn tất',
    'P0',
    'Phiếu status = 2/3/5',
    '1. Thử truy cập /assign/handover/add?id=X với X ở status 2/3/5',
    '',
    'Redirect về chi tiết hoặc báo "Không thể sửa phiếu ở trạng thái này"')

write_tc(2, 18,
    'Không cho người khác sửa phiếu của người khác',
    'P0',
    'Phiếu Nháp do user khác tạo',
    '1. Thử mở Sửa',
    '',
    'Chặn, báo "Không có quyền"')

write_tc(2, 19,
    'Chuyển tab Task ↔ Issue',
    'P1',
    'NV có cả task lẫn issue',
    '1. Click tab Issue\n2. Tick 1 issue\n3. Click lại tab Task',
    '',
    'Bảng chuyển đúng loại items\nTick của tab trước được giữ khi quay lại')

write_tc(2, 20,
    'Group items theo dự án trong bảng chọn',
    'P1',
    'Items thuộc nhiều dự án khác nhau',
    '1. Mở bảng items',
    '',
    'Rows group theo dự án\nMỗi nhóm có header dự án')


# =============================================
# III. CHI TIẾT PHIẾU (/assign/handover/{id})
# =============================================
write_section_row('III. CHI TIẾT PHIẾU (/assign/handover/{id})')

write_tc(3, 1,
    'Layout trang chi tiết',
    'P0',
    'Phiếu tồn tại, user có quyền xem',
    '1. Vào /assign/handover/{id}',
    '',
    'Hiển thị: InfoCard (mã phiếu, NV, PB, ngày BG, lý do, note, status badge, người tạo, người duyệt, thời gian duyệt)\nItemsTable + Timeline + Action buttons')

write_tc(3, 2,
    'Banner lý do từ chối hiển thị khi status = 4',
    'P0',
    'Phiếu Từ chối có reject_reason',
    '1. Quan sát phía trên InfoCard',
    'reject_reason = "Thiếu receiver"',
    'Banner đỏ hiển thị đúng reject_reason')

write_tc(3, 3,
    'Timeline log thao tác',
    'P0',
    'Phiếu đã qua submit + approve',
    '1. Xem HandoverTimeline',
    '',
    'Hiển thị tuần tự: tạo → gửi duyệt → duyệt\nMỗi log có actor + thời gian + nội dung')

write_tc(3, 4,
    'Nút Duyệt chỉ hiện cho user có quyền và status=2',
    'P0',
    'User có "Duyệt bàn giao" + phiếu Chờ duyệt',
    '1. Xem action buttons',
    '',
    'Nút Duyệt hiện với TP có quyền\nUser khác hoặc status khác: nút ẩn')

write_tc(3, 5,
    'Duyệt phiếu thành công',
    'P0',
    'TP có quyền, phiếu Chờ duyệt',
    '1. Click Duyệt\n2. Confirm',
    '',
    'status = 3, set approved_by + approved_at\nNoti gửi tới tất cả receiver\nLog action = "approve"')

write_tc(3, 6,
    'Từ chối phiếu - bắt buộc nhập lý do',
    'P0',
    'TP có quyền, phiếu Chờ duyệt',
    '1. Click Từ chối\n2. Để trống lý do\n3. OK',
    '',
    'Báo lỗi: "Vui lòng nhập lý do từ chối"')

write_tc(3, 7,
    'Từ chối phiếu có lý do',
    'P0',
    'TP có quyền',
    '1. Click Từ chối\n2. Nhập lý do\n3. OK',
    'reject_reason = "Thiếu item quan trọng"',
    'status = 4, lưu reject_reason\nNoti gửi tới người tạo\nLog action = "reject"')

write_tc(3, 8,
    'TP sửa receiver của item trước khi duyệt (BR4)',
    'P0',
    'Phiếu Chờ duyệt, TP có quyền',
    '1. Đổi receiver cell-inline\n2. Duyệt',
    '',
    'Receiver mới được lưu\nDuyệt thành công')

write_tc(3, 9,
    'TP sửa deadline / progress / note của item trước duyệt',
    'P1',
    'Phiếu Chờ duyệt',
    '1. Đổi deadline, progress, note\n2. Lưu',
    '',
    'Cập nhật thành công\nLog action = "edit_item"')

write_tc(3, 10,
    'Không cho sửa item sau khi đã duyệt',
    'P0',
    'Phiếu status = 3',
    '1. Thử sửa receiver/note',
    '',
    'Các field disabled\nKhông gọi API')

write_tc(3, 11,
    'Hiển thị danh sách items với receive_status',
    'P0',
    'Phiếu Đã duyệt, items ở 0/1/2',
    '1. Xem ItemsTable',
    '',
    'Mỗi item: badge "Chờ nhận" / "Đã nhận" / "Từ chối"\nThời gian xử lý\nLý do từ chối (nếu có)')

write_tc(3, 12,
    'User không có quyền xem trả 403 / redirect',
    'P0',
    'User không thuộc phạm vi quyền',
    '1. Mở /assign/handover/{id}',
    '',
    'Không hiển thị chi tiết\nBáo "Không có quyền" hoặc redirect')

write_tc(3, 13,
    'Phiếu không tồn tại trả 404',
    'P1',
    'id không tồn tại',
    '1. Mở /assign/handover/99999',
    'id = 99999',
    'Hiển thị 404 hoặc thông báo "Phiếu không tồn tại"')

write_tc(3, 14,
    'Nút Gửi duyệt hiện cho người tạo khi status = 1 hoặc 4',
    'P0',
    'Phiếu Nháp hoặc Từ chối của current user',
    '1. Xem action buttons',
    '',
    'Nút "Gửi duyệt" hiện ở status 1 hoặc 4\nstatus 3/5: ẩn')

write_tc(3, 15,
    'Badge "Hoàn tất" tự động khi items xử lý hết (BR7) - UI',
    'P0',
    'Phiếu status = 3, còn 1 item chờ',
    '1. Receiver xử lý item cuối cùng\n2. Refresh trang chi tiết',
    '',
    'Status badge chuyển "Hoàn tất" (xanh lá)\nAction buttons cập nhật (không còn nút thao tác)')


# =============================================
# IV. TIẾP NHẬN ITEM (/assign/handover/{id}/receive)
# =============================================
write_section_row('IV. TIẾP NHẬN ITEM (/assign/handover/{id}/receive)')

write_tc(4, 1,
    'Trang chỉ hiện items mà current user là receiver',
    'P0',
    'Phiếu có 5 items, user là receiver của 2',
    '1. Vào /{id}/receive',
    '',
    'Chỉ thấy 2 items user phụ trách\n3 items khác không hiện')

write_tc(4, 2,
    'Accept item → task đổi assignee sang receiver (BR5)',
    'P0',
    'Item chờ (receive_status = 0), task assignee cũ = A',
    '1. Click Nhận',
    'Task T1 assignee_id = A, receiver = B',
    'receive_status = 1\nTask T1 assignee_id = B\nreceived_at set\nNoti tới người tạo (IX.004)\nLog action = "accept_item"')

write_tc(4, 3,
    'Reject item - bắt buộc nhập lý do',
    'P0',
    'Item chờ',
    '1. Click Từ chối\n2. Để trống lý do\n3. OK',
    '',
    'Báo lỗi: "Vui lòng nhập lý do từ chối"')

write_tc(4, 4,
    'Reject item với lý do (BR6)',
    'P0',
    'Item chờ, task assignee = A',
    '1. Click Từ chối\n2. Nhập lý do\n3. OK',
    'reject_reason = "Không đủ thông tin"',
    'receive_status = 2, lưu reject_reason\nTask assignee_id vẫn = A (giữ nguyên)\nNoti tới người tạo')

write_tc(4, 5,
    'Tất cả items accept → phiếu Hoàn tất (BR7)',
    'P0',
    'Phiếu còn 1 item chờ duy nhất',
    '1. Accept item cuối',
    '',
    'Phiếu status = 5\nNoti tới người tạo + TP (IX.006)')

write_tc(4, 6,
    'Tất cả items xử lý mix accept/reject → phiếu Hoàn tất',
    'P0',
    'Phiếu còn 1 item chờ, các item còn lại accepted',
    '1. Reject item cuối',
    '',
    'Phiếu status = 5 (dù có item reject)')

write_tc(4, 7,
    'Không cho thao tác khi phiếu chưa duyệt',
    'P0',
    'Phiếu status = 2',
    '1. Vào /{id}/receive',
    '',
    'Nút Nhận/Từ chối disabled hoặc ẩn\nBáo "Phiếu chưa được duyệt"')

write_tc(4, 8,
    'Không cho thao tác khi phiếu Hoàn tất',
    'P1',
    'Phiếu status = 5',
    '1. Vào /{id}/receive',
    '',
    'Mọi action disabled, chỉ xem')

write_tc(4, 9,
    'Không cho thao tác item không phải của mình',
    'P0',
    'Item có receiver là user khác',
    '1. Gọi API PUT /items/{id}/accept trực tiếp',
    '',
    'API trả 403 "Không có quyền thao tác item này"')

write_tc(4, 10,
    'Idempotent - accept lại item đã accept',
    'P1',
    'Item đã có receive_status = 1',
    '1. Click Nhận lần 2 (hoặc gọi API lại)',
    '',
    'API trả lỗi "Item đã được xử lý"\nKhông tạo log trùng, không đổi DB')

write_tc(4, 11,
    'Hiển thị thông tin task/issue cho receiver xem trước khi quyết định',
    'P1',
    '',
    '1. Mở trang /receive',
    '',
    'Mỗi row hiện: tên task/issue, dự án, deadline, progress, handover_note')

write_tc(4, 12,
    'Filter theo dự án/trạng thái item',
    'P2',
    'Nhiều items thuộc nhiều dự án',
    '1. Chọn filter dự án hoặc status',
    '',
    'Bảng lọc đúng các item thoả filter')


# =============================================
# V. CHỜ DUYỆT (/assign/handover/pending)
# =============================================
write_section_row('V. CHỜ DUYỆT (/assign/handover/pending)')

write_tc(5, 1,
    'User không có quyền "Duyệt bàn giao" không vào được',
    'P0',
    'User thiếu permission',
    '1. Mở /assign/handover/pending',
    '',
    'Redirect hoặc 403, menu Chờ duyệt ẩn')

write_tc(5, 2,
    'Hiển thị chỉ phiếu status = 2 trong phạm vi quyền',
    'P0',
    'Có nhiều phiếu đủ status',
    '1. Vào /pending',
    '',
    'Bảng chỉ show status = 2 thuộc phạm vi user được duyệt')

write_tc(5, 3,
    'Filter công ty/PB/lý do/NV',
    'P1',
    'Có nhiều phiếu',
    '1. Chọn các filter\n2. Tìm kiếm',
    '',
    'Bảng lọc đúng')

write_tc(5, 4,
    'Duyệt nhanh từ danh sách',
    'P0',
    'TP có quyền',
    '1. Click Duyệt trên dòng\n2. Confirm',
    '',
    'Phiếu chuyển status = 3 (hiệu ứng giống màn chi tiết III.005)')

write_tc(5, 5,
    'Từ chối nhanh có lý do',
    'P0',
    'TP có quyền',
    '1. Click Từ chối\n2. Nhập lý do\n3. OK',
    'reject_reason',
    'Phiếu status = 4, lưu reason (giống III.007)')

write_tc(5, 6,
    'Xuất Excel hiển thị "đang phát triển"',
    'P2',
    'Design ghi chưa triển khai',
    '1. Click Xuất Excel',
    '',
    'Thông báo "Chức năng đang phát triển"')


# =============================================
# VI. CHỜ TIẾP NHẬN (/assign/handover/receiving)
# =============================================
write_section_row('VI. CHỜ TIẾP NHẬN (/assign/handover/receiving)')

write_tc(6, 1,
    'Hiển thị chỉ items current user là receiver và receive_status = 0',
    'P0',
    'User là receiver 3 items chờ',
    '1. Mở /receiving',
    '',
    'Bảng 3 items, items khác không hiện')

write_tc(6, 2,
    'Group theo mã phiếu',
    'P1',
    'Items thuộc nhiều phiếu',
    '1. Quan sát bảng',
    '',
    'Rows group theo phiếu\nHeader phiếu hiện: mã phiếu, NV bàn giao, ngày BG')

write_tc(6, 3,
    'Accept từ danh sách',
    'P0',
    'Item chờ',
    '1. Click Nhận trên dòng',
    '',
    'receive_status = 1, task đổi assignee (giống IV.002)')

write_tc(6, 4,
    'Reject từ danh sách có lý do',
    'P0',
    'Item chờ',
    '1. Click Từ chối\n2. Nhập lý do',
    '',
    'receive_status = 2 (giống IV.004)')

write_tc(6, 5,
    'Items phiếu xử lý xong biến mất khỏi DS',
    'P0',
    'Vừa accept xong 1 item',
    '1. Reload /receiving',
    '',
    'Item đã xử lý không còn xuất hiện')

write_tc(6, 6,
    'Xuất Excel hiển thị "đang phát triển"',
    'P2',
    '',
    '1. Click Xuất Excel',
    '',
    'Thông báo "Chức năng đang phát triển"')


# =============================================
# VII. PHÂN QUYỀN THEO CẤP
# =============================================
write_section_row('VII. PHÂN QUYỀN THEO CẤP')

write_tc(7, 1,
    'Quyền "Xem theo tổng công ty" thấy tất cả',
    'P0',
    'User có permission xem theo tổng công ty',
    '1. Vào /assign/handover',
    '',
    'Bảng hiển thị phiếu thuộc mọi công ty')

write_tc(7, 2,
    'Quyền "Xem theo công ty" giới hạn trong công ty',
    'P0',
    'User chỉ có quyền theo công ty',
    '1. Vào /assign/handover',
    '',
    'Chỉ thấy phiếu có company_id = của user')

write_tc(7, 3,
    'Quyền "Xem theo phòng ban" giới hạn trong PB',
    'P0',
    'User chỉ có quyền theo phòng ban',
    '1. Vào /assign/handover',
    '',
    'Chỉ thấy phiếu có department_id = của user')

write_tc(7, 4,
    'User không có quyền xem nào → không vào được menu',
    'P0',
    'User trống quyền liên quan',
    '1. Quan sát menu\n2. Thử mở URL trực tiếp',
    '',
    'Menu ẩn\nMở URL trực tiếp redirect về /dashboard hoặc 403')

write_tc(7, 5,
    'Quyền "Duyệt bàn giao" - chỉ duyệt phiếu trong phạm vi xem',
    'P0',
    'TP PB A, thử mở phiếu PB B qua link',
    '1. Mở /assign/handover/{id} của PB B',
    '',
    'Không hiện nút Duyệt\nAPI /approve trả 403 nếu gọi trực tiếp')

write_tc(7, 6,
    'Người tạo luôn thấy phiếu mình dù không có quyền xem',
    'P0',
    'User không có quyền xem, từng tạo phiếu',
    '1. Mở /assign/handover',
    '',
    'Danh sách vẫn hiện phiếu do mình tạo (override scope)')

write_tc(7, 7,
    'Receiver luôn thấy phiếu mà mình là receiver',
    'P0',
    'User là receiver, không có quyền xem',
    '1. Mở /assign/handover hoặc /receiving',
    '',
    'Thấy phiếu/items tương ứng mà mình là receiver')

write_tc(7, 8,
    'TP chỉ duyệt được phiếu NV trong PB mình quản lý',
    'P1',
    'TP PB A, phiếu NV PB B',
    '1. Gọi API /approve cho phiếu PB B',
    '',
    'API trả 403 "Không thuộc phạm vi quyền duyệt"')


# =============================================
# VIII. BUSINESS RULES & LUỒNG E2E
# =============================================
write_section_row('VIII. BUSINESS RULES & LUỒNG E2E')

write_tc(8, 1,
    'Happy path đầy đủ',
    'P0',
    'Có TP và 2 NV (bàn giao + nhận)',
    '1. NV tạo phiếu với 2 items\n2. Gán receiver\n3. Gửi duyệt\n4. TP duyệt\n5. Receiver accept hết items',
    '',
    'Trải qua status 1→2→3→5\nLog đầy đủ các action\nNoti đúng tại từng bước')

write_tc(8, 2,
    'Luồng từ chối → sửa → gửi lại (BR8)',
    'P0',
    '',
    '1. Tạo phiếu + gửi duyệt\n2. TP từ chối có lý do\n3. Người tạo sửa + gửi lại\n4. TP duyệt',
    '',
    'Trải qua 1→2→4→2→3\nBanner lý do hiện ở lần sửa\nLog: submit → reject → resubmit → approve')

write_tc(8, 3,
    'BR1 - phiếu phải có ≥1 item (double-check BE)',
    'P0',
    '',
    '1. Gọi API POST /handovers với items rỗng',
    '',
    'BE trả 422 "Phiếu phải có ít nhất 1 item"\nFE cũng chặn')

write_tc(8, 4,
    'BR2 - gửi duyệt phải đủ receiver',
    'P0',
    '1 item thiếu receiver',
    '1. Gọi API PUT /handovers/{id}/submit',
    '',
    'BE trả lỗi, phiếu vẫn status = 1')

write_tc(8, 5,
    'BR3 - 1 task không nằm trong 2 phiếu active',
    'P0',
    'Task X đã trong phiếu status = 1 (phiếu A)',
    '1. Tạo phiếu B có task X',
    '',
    'BE trả 422 "Task đã có trong phiếu {code của A}"\nFE không cho chọn task X khi tạo phiếu B')

write_tc(8, 6,
    'BR4 - TP sửa được receiver/note/progress/deadline trước duyệt',
    'P0',
    'Phiếu status = 2',
    '1. TP đổi receiver item\n2. TP đổi deadline + progress\n3. Duyệt',
    '',
    'Giá trị mới được lưu và dùng khi receiver nhận')

write_tc(8, 7,
    'BR5 - Accept đổi assignee task/issue',
    'P0',
    'Task assignee = A, receiver = B',
    '1. B accept item',
    '',
    'Task.assignee_id chuyển thành B\nLịch sử task ghi lại thay đổi assignee')

write_tc(8, 8,
    'BR6 - Reject giữ nguyên assignee',
    'P0',
    'Task assignee = A, receiver = B',
    '1. B reject item',
    '',
    'Task.assignee_id vẫn = A')

write_tc(8, 9,
    'BR7 - Auto Hoàn tất khi items xử lý xong',
    'P0',
    'Phiếu status = 3, còn 1 item chờ',
    '1. Receiver accept hoặc reject item cuối',
    '',
    'status = 5 tự động\nTrigger noti (IX.006)')

write_tc(8, 10,
    'BR8 - Phiếu Từ chối → sửa + gửi lại',
    'P0',
    'Phiếu status = 4',
    '1. Người tạo sửa\n2. Gửi duyệt',
    '',
    'status = 2\nBanner reject_reason biến mất\nLog action = "resubmit" giữ reject_reason cũ trong lịch sử')

write_tc(8, 11,
    'BR9 - Noti khi gửi duyệt (smoke)',
    'P0',
    '',
    '1. Gửi duyệt',
    '',
    'TP nhận noti trong hệ thống (chi tiết ở IX.001)')

write_tc(8, 12,
    'Mã phiếu auto format BG.YYYY.NNNN',
    'P0',
    '',
    '1. Tạo nhiều phiếu liên tiếp',
    '',
    'Code match regex ^BG\\.\\d{4}\\.\\d{4}$\nSố thứ tự tăng dần theo năm, reset đầu năm mới')


# =============================================
# IX. NOTIFICATION
# =============================================
write_section_row('IX. NOTIFICATION')

write_tc(9, 1,
    'Gửi duyệt → noti TP',
    'P0',
    'TP của PB tồn tại, có quyền duyệt',
    '1. Người tạo gửi duyệt',
    '',
    'TP nhận noti trong hệ thống: "Có phiếu bàn giao chờ duyệt" + link tới phiếu')

write_tc(9, 2,
    'Duyệt → noti tất cả receivers',
    'P0',
    'Phiếu có 3 receiver',
    '1. TP duyệt',
    '',
    'Cả 3 receiver nhận noti: "Có phiếu bàn giao cần tiếp nhận" + link /receive')

write_tc(9, 3,
    'Từ chối → noti người tạo',
    'P0',
    '',
    '1. TP từ chối + nhập lý do',
    '',
    'Người tạo nhận noti kèm lý do từ chối + link phiếu')

write_tc(9, 4,
    'Accept item → noti người tạo',
    'P0',
    '',
    '1. Receiver accept item',
    '',
    'Người tạo nhận noti: "Đã nhận bàn giao item {tên task/issue}"')

write_tc(9, 5,
    'Reject item → noti người tạo',
    'P0',
    '',
    '1. Receiver reject item + nhập lý do',
    '',
    'Người tạo nhận noti kèm lý do + link phiếu')

write_tc(9, 6,
    'Hoàn tất → noti người tạo + TP',
    'P0',
    'Phiếu còn 1 item chờ',
    '1. Receiver xử lý item cuối',
    '',
    'Cả người tạo + TP duyệt nhận noti: "Phiếu {code} đã hoàn tất"')


# =============================================
# X. EDGE CASES
# =============================================
write_section_row('X. EDGE CASES')

write_tc(10, 1,
    'Receiver nghỉ việc / bị khoá giữa quá trình',
    'P1',
    'Receiver B đã bị khoá tài khoản',
    '1. Vào /receive với B\n2. Hoặc TP muốn duyệt phiếu có receiver B',
    '',
    'FE cảnh báo "Receiver không còn hoạt động"\nGợi ý TP đổi receiver trước khi duyệt')

write_tc(10, 2,
    'Task bị xoá sau khi phiếu được tạo',
    'P1',
    'Task X bị soft delete',
    '1. Mở chi tiết phiếu chứa task X',
    '',
    'Item hiện trạng thái "Task không còn tồn tại"\nKhông cho accept item này')

write_tc(10, 3,
    '2 TP cùng duyệt phiếu đồng thời',
    'P1',
    '2 TP có quyền, phiếu Chờ duyệt',
    '1. Cả 2 bấm Duyệt cùng lúc',
    '',
    'Người sau nhận lỗi "Phiếu đã được xử lý"\nUI refresh, hiển thị status mới')

write_tc(10, 4,
    'NV tự bàn giao cho chính mình',
    'P1',
    '',
    '1. Tạo phiếu, employee_id = receiver_id của item',
    '',
    'FE/BE chặn: "Người nhận không được trùng người bàn giao"')

write_tc(10, 5,
    'Task đang trong phiếu active → cố thêm vào phiếu khác (BR3 BE)',
    'P0',
    'Task trong phiếu status = 2',
    '1. Tạo phiếu khác có task đó qua API',
    '',
    'BE trả 422 "Task đang trong phiếu active {code}"')

write_tc(10, 6,
    'Concurrent accept - 2 tab cùng click Nhận',
    'P2',
    '',
    '1. Mở 2 tab trên cùng item\n2. Click Nhận gần như cùng lúc',
    '',
    '1 tab thành công\nTab còn lại báo "Item đã được xử lý"')

write_tc(10, 7,
    'Phiếu bàn giao nhưng NV chưa có task/issue nào',
    'P1',
    'NV A trống task/issue',
    '1. Tạo phiếu, chọn NV A',
    '',
    'Bảng items hiển thị "Không có task/issue nào để bàn giao"\nKhông gửi được (vi phạm BR1)')

write_tc(10, 8,
    'Đổi lý do từ "Khác" sang loại khác - ghi chú có bị mất không',
    'P2',
    'Lý do ban đầu = 5, đã nhập note',
    '1. Đổi Lý do = 1 (Nghỉ việc)\n2. Submit',
    '',
    'Note được giữ nguyên, không mất data khi submit')


# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# === SAVE ===
output_path = r'D:\laragon\www\hrm\.plans\fix-handover\Testcase_Ban_Giao_Cong_Viec.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Total testcases: {current_row - 7 - 10}')  # trừ 10 section rows
