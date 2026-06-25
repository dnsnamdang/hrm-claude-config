"""Generate testcase Excel BỔ SUNG cho phần ĐIỂM DANH thành viên trong Meeting."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "testcase-diem-danh.xlsx")
SHEET_NAME   = "DiemDanh"
FEATURE_NAME = "Điểm danh thành viên Meeting"
MODULE_NAME  = "Điểm danh Meeting"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Cho phép thư ký/người phụ trách điểm danh từng thành viên tham dự cuộc họp (nội bộ + khách hàng) "
     "tại tab 'Điểm danh' trong màn chi tiết/sửa Meeting.\n"
     "Là điều kiện bắt buộc trước khi chốt biên bản và hoàn thành cuộc họp (status → Đã hoàn thành)."),
    ("2. Đối tượng được tính / hiển thị",
     "► Thành viên Nội bộ — lấy từ form.company_members (type=1), badge 'Nội bộ'\n"
     "► Thành viên Khách hàng — lấy từ form.customer_members (type=2), badge 'Khách hàng'\n"
     "► Trạng thái điểm danh: 0=Dự kiến tham gia (mặc định), 1=Có mặt, 2=Vắng có lý do, 3=Vắng không lý do\n"
     "► Mỗi dòng: STT, Họ tên, Chức vụ/Phòng ban, Thành phần, Trạng thái điểm danh, Ghi chú/Lý do"),
    ("3. Đối tượng bị ẩn / không tính",
     "► Khi không có thành viên nào (company_members + customer_members rỗng) → hiển thị empty state "
     "'Chưa có thành viên nào để điểm danh', không render bảng\n"
     "► Tab điểm danh KHÔNG ẩn theo loại meeting (luôn hiển thị)"),
    ("4. Bộ lọc thời gian áp dụng cho",
     "— Không áp dụng. Tab điểm danh không có bộ lọc."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Bảng meeting_employees: cột attendance_status (tinyInteger default 0), attendance_note (text nullable)\n"
     "► Lưu qua MeetingService::syncCompanyMembers / syncCustomerMembers (xóa cũ → tạo lại theo type)\n"
     "► Field reactive đảm bảo bởi ensureAttendanceFields (FE) trước khi render"),
    ("6. Quy tắc cộng dồn / deduplicate",
     "► syncCompanyMembers/syncCustomerMembers dùng array_filter loại bỏ giá trị null/empty trước khi fill\n"
     "  → attendance_status = 0 (Dự kiến) và attendance_note rỗng KHÔNG được gửi explicit, "
     "DB sẽ dùng default (status=0, note=null)"),
    ("7. Phân quyền cấp",
     "— Không có permission riêng cho tab điểm danh. Quyền điểm danh đi kèm quyền SỬA meeting.\n"
     "► Điều kiện được phép điểm danh (editable) = đang ở chế độ Sửa (KHÔNG phải màn Xem) VÀ "
     "trạng thái meeting = 2 (Đã chốt lịch).\n"
     "► Các trạng thái khác (0 Nháp / 1 Lên lịch / 3 Hoàn thành / 4 Huỷ) hoặc màn Xem → chỉ đọc (readonly)."),
    ("8. Cách tính các ô thống kê",
     "— Không có ô thống kê dạng số. Logic kiểm tra:\n"
     "► isAllAttended() = mọi member có attendance_status ∈ [1,2,3] (không còn ai = 0)\n"
     "► editable = (!isShow) && Number(form.status) === 2"),
    ("9. Ghi chú đọc bảng",
     "► FE: MeetingAttendance.vue, guard hoàn thành tại MeetingForm.vue (handleComplete/isAllAttended)\n"
     "► BE: MeetingController::update (guard status=HOAN_THANH), MeetingService (sync), MeetingEmployee\n"
     "► Test ở cả 2 lớp: FE chặn trước (toast + chuyển tab), BE chặn lại (422) nếu bypass FE"),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    ("I", "HIỂN THỊ TAB ĐIỂM DANH & DANH SÁCH THÀNH VIÊN", [
        ("001", "Mở tab 'Điểm danh' từ màn chi tiết meeting", "P0",
         "Meeting đã có ≥1 thành viên nội bộ + ≥1 khách hàng. Đang ở màn chi tiết (/assign/meeting/{id}/show)",
         "1. Mở chi tiết meeting\n2. Click tab 'Điểm danh' (icon ri-user-follow-line)",
         "Meeting có 2 thành viên nội bộ + 1 khách hàng",
         "- Hiển thị tiêu đề 'Điểm danh thành viên'\n- Bảng liệt kê đủ 3 dòng theo đúng thứ tự: nội bộ trước, khách hàng sau\n- Cột STT đánh số 1,2,3",
         "attendanceRows = company_members (type=1) + customer_members (type=2)"),
        ("002", "Hiển thị đúng 6 cột của bảng điểm danh", "P1",
         "Meeting có thành viên",
         "1. Mở tab 'Điểm danh'\n2. Quan sát header bảng",
         "—",
         "- Header gồm: STT | Họ và tên | Chức vụ / Phòng ban | Thành phần | Trạng thái điểm danh | Ghi chú / Lý do",
         "Cấu trúc cột cố định trong MeetingAttendance.vue"),
        ("003", "Hiển thị badge Thành phần đúng loại", "P1",
         "Meeting có 1 nội bộ + 1 khách hàng",
         "1. Mở tab 'Điểm danh'\n2. Quan sát cột 'Thành phần'",
         "—",
         "- Dòng nội bộ: badge 'Nội bộ' (badge-soft-brand, xanh)\n- Dòng khách hàng: badge 'Khách hàng' (badge-soft-customer, xanh dương)",
         "type===1 → Nội bộ; ngược lại Khách hàng"),
        ("004", "Empty state khi meeting không có thành viên", "P1",
         "Meeting có company_members = [] và customer_members = []",
         "1. Mở tab 'Điểm danh'",
         "Meeting không thành viên",
         "- Không render bảng\n- Hiển thị 'Chưa có thành viên nào để điểm danh' (icon ri-team-line)",
         "v-if attendanceRows.length, ngược lại empty state"),
        ("005", "Hiển thị tên & chức vụ, fallback '-' khi thiếu", "P2",
         "Meeting có 1 thành viên thiếu role (chức vụ rỗng)",
         "1. Mở tab 'Điểm danh'\n2. Quan sát dòng thiếu chức vụ",
         "member.name='Nguyễn A', member.role=null",
         "- Cột Họ tên: 'Nguyễn A'\n- Cột Chức vụ / Phòng ban: '-'",
         "row.member.name || '-'; row.member.role || '-'"),
    ]),
    ("II", "ĐIỀU KIỆN ĐƯỢC PHÉP ĐIỂM DANH (editable)", [
        ("001", "Cho phép điểm danh khi Sửa meeting trạng thái 'Đã chốt lịch'", "P0",
         "Meeting status=2 (Đã chốt lịch). Mở màn Sửa (/assign/meeting/{id}/edit)",
         "1. Mở màn Sửa meeting\n2. Vào tab 'Điểm danh'\n3. Quan sát các điều khiển",
         "status=2, isShow=false",
         "- Hiển thị cụm chip chọn trạng thái cho từng dòng (Có mặt / Vắng có lý do / Vắng không lý do)\n- Ô Ghi chú cho nhập\n- Có nút 'Điểm danh nhanh: Tất cả có mặt'",
         "editable = !isShow && Number(status)===2"),
        ("002", "Readonly khi ở màn Xem chi tiết (isShow=true)", "P0",
         "Meeting status=2. Ở màn Xem chi tiết (show)",
         "1. Mở chi tiết meeting (show)\n2. Vào tab 'Điểm danh'",
         "status=2, isShow=true",
         "- KHÔNG hiển thị chip\n- Hiển thị label trạng thái dạng text (Có mặt/Vắng.../Dự kiến tham gia)\n- Ô Ghi chú bị disabled\n- KHÔNG có nút 'Điểm danh nhanh'",
         "isShow=true → editable=false dù status=2"),
        ("003", "Readonly khi Sửa meeting trạng thái 'Lên lịch hẹn' (status=1)", "P0",
         "Meeting status=1. Mở màn Sửa",
         "1. Mở Sửa meeting\n2. Vào tab 'Điểm danh'",
         "status=1, isShow=false",
         "- Hiển thị readonly (label text, không chip, không nút điểm danh nhanh)",
         "editable yêu cầu status===2; status=1 → false"),
        ("004", "Readonly khi meeting trạng thái 'Lưu nháp' (status=0)", "P1",
         "Meeting status=0. Mở Sửa",
         "1. Mở Sửa meeting\n2. Vào tab 'Điểm danh'",
         "status=0",
         "- Hiển thị readonly, không điểm danh được",
         "status=0 ≠ 2 → editable=false"),
        ("005", "Readonly khi meeting đã 'Hoàn thành' (status=3)", "P0",
         "Meeting status=3 (đã hoàn thành)",
         "1. Mở chi tiết meeting\n2. Vào tab 'Điểm danh'",
         "status=3",
         "- Hiển thị readonly: label trạng thái đã điểm danh trước đó\n- Không sửa được",
         "BR — meeting hoàn thành không cho sửa điểm danh"),
        ("006", "Readonly khi meeting đã 'Huỷ' (status=4)", "P1",
         "Meeting status=4 (đã huỷ)",
         "1. Mở chi tiết meeting\n2. Vào tab 'Điểm danh'",
         "status=4",
         "- Hiển thị readonly, không sửa được",
         "status=4 ≠ 2 → editable=false"),
    ]),
    ("III", "ĐIỂM DANH TỪNG THÀNH VIÊN (chip trạng thái + ghi chú)", [
        ("001", "Chọn trạng thái 'Có mặt' cho 1 thành viên", "P0",
         "Đang Sửa meeting status=2, tab Điểm danh, member đang ở trạng thái Dự kiến (0)",
         "1. Tại dòng thành viên, click chip 'Có mặt'",
         "member.attendance_status: 0 → click 'Có mặt'",
         "- Chip 'Có mặt' active (nền xanh #dcfce7, viền xanh, dot sáng)\n- attendance_status = 1\n- Các chip khác trở về trạng thái thường",
         "$set(row.member,'attendance_status',1)"),
        ("002", "Chọn trạng thái 'Vắng có lý do'", "P0",
         "Đang Sửa meeting status=2, tab Điểm danh",
         "1. Click chip 'Vắng có lý do' tại 1 dòng",
         "click 'Vắng có lý do'",
         "- Chip 'Vắng có lý do' active (nền vàng #fef3c7)\n- attendance_status = 2",
         "value=2, cls=excused"),
        ("003", "Chọn trạng thái 'Vắng không lý do'", "P0",
         "Đang Sửa meeting status=2, tab Điểm danh",
         "1. Click chip 'Vắng không lý do' tại 1 dòng",
         "click 'Vắng không lý do'",
         "- Chip 'Vắng không lý do' active (nền đỏ #fee2e2)\n- attendance_status = 3",
         "value=3, cls=absent"),
        ("004", "Đổi trạng thái qua lại giữa các chip", "P1",
         "Đang Sửa meeting status=2",
         "1. Click 'Có mặt'\n2. Click 'Vắng không lý do'\n3. Click lại 'Có mặt'",
         "—",
         "- Tại mỗi bước chỉ 1 chip active đúng trạng thái vừa chọn\n- Giá trị cuối = 1 (Có mặt)",
         "Mỗi click ghi đè attendance_status"),
        ("005", "Nhập Ghi chú / Lý do cho thành viên", "P1",
         "Đang Sửa meeting status=2",
         "1. Tại ô Ghi chú của 1 dòng, nhập 'Đi công tác đột xuất'",
         "attendance_note='Đi công tác đột xuất'",
         "- Ô nhận text, attendance_note lưu đúng giá trị",
         "v-model row.member.attendance_note"),
        ("006", "Ô Ghi chú bị khoá ở chế độ readonly", "P1",
         "Màn Xem chi tiết, status bất kỳ",
         "1. Vào tab Điểm danh ở màn Xem\n2. Thử click vào ô Ghi chú",
         "isShow=true",
         "- Ô Ghi chú disabled, không nhập được",
         ":disabled='!editable'"),
        ("007", "Điểm danh độc lập cho từng dòng", "P1",
         "Meeting có 3 thành viên, đang Sửa status=2",
         "1. Dòng 1 → 'Có mặt'\n2. Dòng 2 → 'Vắng có lý do'\n3. Dòng 3 → giữ Dự kiến",
         "—",
         "- Dòng 1 status=1, dòng 2 status=2, dòng 3 status=0\n- Việc chọn dòng này không ảnh hưởng dòng khác",
         "Mỗi row gắn member riêng"),
    ]),
    ("IV", "ĐIỂM DANH NHANH — TẤT CẢ CÓ MẶT", [
        ("001", "Nút 'Điểm danh nhanh' set tất cả thành Có mặt", "P0",
         "Đang Sửa meeting status=2, có 3 thành viên ở các trạng thái khác nhau (0,2,3)",
         "1. Click nút 'Điểm danh nhanh: Tất cả có mặt'",
         "3 thành viên: status 0/2/3",
         "- Tất cả dòng chip 'Có mặt' active\n- attendance_status mọi member = 1",
         "markAllPresent() set tất cả về 1"),
        ("002", "Nút 'Điểm danh nhanh' chỉ hiện khi editable", "P1",
         "So sánh màn Sửa(status=2) và màn Xem",
         "1. Màn Sửa status=2 → quan sát\n2. Màn Xem → quan sát",
         "—",
         "- Màn Sửa status=2: hiện nút\n- Màn Xem hoặc status≠2: ẩn nút",
         "v-if='editable'"),
        ("003", "Điểm danh nhanh không tác động khi không có thành viên", "P2",
         "Meeting không thành viên (empty state)",
         "1. Quan sát màn (Sửa status=2 nhưng members rỗng)",
         "members rỗng",
         "- Không có bảng, nút điểm danh nhanh không gây lỗi (attendanceRows rỗng)",
         "forEach trên mảng rỗng → no-op"),
    ]),
    ("V", "CHẾ ĐỘ XEM (readonly) — HIỂN THỊ LABEL TRẠNG THÁI", [
        ("001", "Hiển thị label 'Có mặt' màu xanh ở màn Xem", "P1",
         "Màn Xem, member attendance_status=1",
         "1. Vào tab Điểm danh màn Xem",
         "status member=1",
         "- Text 'Có mặt' (class status-present, màu xanh #16a34a)",
         "statusLabel(1)='Có mặt'; statusClass(1)=status-present"),
        ("002", "Hiển thị label 'Vắng có lý do'/'Vắng không lý do' màu đỏ", "P1",
         "Màn Xem, có member status=2 và member status=3",
         "1. Vào tab Điểm danh màn Xem",
         "status 2 và 3",
         "- 'Vắng có lý do' và 'Vắng không lý do' đều class status-absent (đỏ #ef4444)",
         "statusClass(2/3)=status-absent"),
        ("003", "Hiển thị 'Dự kiến tham gia' cho member chưa điểm danh", "P1",
         "Màn Xem, member attendance_status=0",
         "1. Vào tab Điểm danh màn Xem",
         "status member=0",
         "- Text 'Dự kiến tham gia' (class status-pending, xám italic)",
         "statusLabel(0)='Dự kiến tham gia'"),
    ]),
    ("VI", "GUARD HOÀN THÀNH & VALIDATION (FE + BE)", [
        ("001", "Chặn Hoàn thành khi còn thành viên chưa điểm danh (FE)", "P0",
         "Đang Sửa meeting status=2, đã có biên bản, còn ≥1 member status=0",
         "1. Click nút 'Hoàn thành'",
         "1 member status=0, đã có reports",
         "- Toast lỗi 'Vui lòng hoàn thành điểm danh cho tất cả thành viên trước khi chốt biên bản và hoàn thành cuộc họp.'\n- Tự chuyển sang tab 'Điểm danh'\n- KHÔNG gọi API hoàn thành",
         "isAllAttended()=false → chặn; mọi member phải ∈[1,2,3]"),
        ("002", "Cho phép Hoàn thành khi đã điểm danh đủ + có biên bản", "P0",
         "Đang Sửa meeting status=2, đã có biên bản, mọi member status ∈[1,2,3]",
         "1. Click nút 'Hoàn thành'",
         "mọi member đã điểm danh, có reports",
         "- Không có lỗi điểm danh\n- Gọi save(3), meeting chuyển 'Đã hoàn thành'",
         "isAllAttended()=true → save(3)"),
        ("003", "Chặn Hoàn thành khi chưa có biên bản (ưu tiên trước điểm danh)", "P1",
         "Đang Sửa meeting status=2, CHƯA có biên bản",
         "1. Click 'Hoàn thành'",
         "reports=[]",
         "- Toast 'Vui lòng thêm biên bản cuộc họp trước khi hoàn thành!'\n- Chuyển sang tab 'Biên bản'\n- Không kiểm tra điểm danh ở bước này",
         "hasMeetingReport() check trước isAllAttended()"),
        ("004", "BE chặn hoàn thành khi member chưa điểm danh (bypass FE)", "P0",
         "Gọi trực tiếp API update với status=3 (HOAN_THANH), payload còn member attendance_status=0",
         "1. POST update meeting status=3 với 1 member status=0",
         "status=3, member status=0",
         "- BE trả lỗi (422) 'Vui lòng hoàn thành điểm danh cho tất cả thành viên trước khi chốt biên bản và hoàn thành cuộc họp.'\n- Không cập nhật trạng thái",
         "MeetingController guard: status===HOAN_THANH && attendance_status ∉[1,2,3]"),
        ("005", "BE cho hoàn thành khi mọi member đã điểm danh", "P0",
         "Gọi API update status=3, mọi member status ∈[1,2,3]",
         "1. POST update status=3 với members đã điểm danh đủ",
         "mọi member ∈[1,2,3]",
         "- BE qua guard, cập nhật meeting thành công status=3",
         "guard pass"),
    ]),
    ("VII", "LƯU TRỮ DỮ LIỆU & EDGE CASES", [
        ("001", "Lưu meeting (status=2) giữ lại trạng thái điểm danh đã chọn", "P0",
         "Đang Sửa meeting status=2, điểm danh: A=Có mặt, B=Vắng có lý do + ghi chú",
         "1. Điểm danh như trên\n2. Click 'Lưu' (giữ status=2)\n3. Reload / mở lại chi tiết",
         "A status=1; B status=2, note='Bận'",
         "- Sau reload: A hiển thị 'Có mặt', B 'Vắng có lý do' + ghi chú 'Bận'\n- Dữ liệu lưu đúng vào meeting_employees",
         "syncCompanyMembers/syncCustomerMembers ghi attendance_status, attendance_note"),
        ("002", "attendance_status=0 (Dự kiến) lưu về default DB", "P0",
         "Đang Sửa meeting status=2, 1 member giữ Dự kiến (0), không ghi chú",
         "1. Lưu meeting\n2. Kiểm tra DB / reload",
         "member status=0, note rỗng",
         "- Do array_filter loại bỏ giá trị 0/rỗng, status không gửi explicit → DB lưu default 0 (Dự kiến), note=null\n- Reload hiển thị 'Dự kiến tham gia'",
         "BR — array_filter bỏ null/empty; cột default 0"),
        ("003", "Sync xoá-tạo-lại: bỏ thành viên rồi lưu", "P1",
         "Meeting status=2 có 3 thành viên đã điểm danh, xoá 1 thành viên ở tab Thông tin",
         "1. Xoá 1 thành viên\n2. Lưu meeting\n3. Vào lại tab Điểm danh",
         "còn 2 thành viên",
         "- meeting_employees cũ bị xóa, tạo lại 2 bản ghi\n- Điểm danh của 2 thành viên còn lại giữ nguyên",
         "sync: delete() rồi tạo lại theo type"),
        ("004", "ensureAttendanceFields chuẩn hoá status dạng string", "P2",
         "API trả member.attendance_status='2' (chuỗi)",
         "1. Mở tab Điểm danh",
         "attendance_status='2' (string)",
         "- Chip 'Vắng có lý do' active đúng (so sánh Number)\n- Không lỗi do kiểu dữ liệu",
         "ensureAttendanceFields: string → Number"),
        ("005", "Thành viên nội bộ và khách hàng lưu đúng type", "P1",
         "Meeting status=2, điểm danh cả nội bộ + khách hàng rồi lưu",
         "1. Điểm danh A(nội bộ)=Có mặt, K(khách hàng)=Vắng\n2. Lưu",
         "—",
         "- A lưu type=1, K lưu type=2 trong meeting_employees\n- Trạng thái điểm danh đúng từng người",
         "syncCompanyMembers type=1, syncCustomerMembers type=2"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL  = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 28, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 26, 'I': 65, 'J': 38,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT
    a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT
    a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT
    b.alignment = WRAP_TOP_LEFT
    b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}")
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",              '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",         '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",     '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",        '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                  '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT
    lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right")
    lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT
    vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN
    vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [
        MODULE_NAME, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed",
        "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(180, longest // 4))
    current_row += 1
    data_row_idx += 1

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
    showDropDown=False,
)
dv.add(f"L18:N{current_row + 100}")
ws.add_data_validation(dv)

os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE}")
print(f"Rows: 1-10 description, 11-15 summary, 17 header, 18-{current_row-1} data")
