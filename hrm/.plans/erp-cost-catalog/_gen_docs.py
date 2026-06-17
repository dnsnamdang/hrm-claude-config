# -*- coding: utf-8 -*-
import html, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTDIR = "/Users/dnsnamdang/Documents/DNSMEDIA/websites/hrm/.plans/erp-cost-catalog"
TODAY = "2026-06-04"

# ============================================================
# TEST CASES — BOM -> Quotation (erp-cost-catalog Phase P5)
# Mỗi TC: (id, nhom, chuc_nang, tien_dieu_kien, cac_buoc, du_lieu, ket_qua, uu_tien)
# ============================================================
TC = [
 # A. Danh mục dịch vụ / chi phí (ERP costs) + tạo nhanh
 ("TC-01","A. Danh mục DV/CP","Tìm dịch vụ từ danh mục costs ERP",
  "Đăng nhập, mở form tạo BOM hoặc Báo giá, đã chọn loại tiền tệ",
  "1. Mở popup Thêm mới phần Dịch vụ & Chi phí\n2. Gõ từ khoá vào ô tìm kiếm\n3. Lọc theo 'Dịch vụ có tính DT' / 'Chi phí khác' / 'Tất cả'",
  "Từ khoá: tên dịch vụ có thật",
  "Danh sách costs (status=1, kind_of=2) hiển thị đúng; badge phân biệt Dịch vụ (xanh) / Chi phí khác (cam); lọc đúng theo revenue_calculation",
  "Cao"),
 ("TC-02","A. Danh mục DV/CP","Tạo nhanh dịch vụ trên BÁO GIÁ (đủ giá)",
  "Đang ở popup Dịch vụ trên Báo giá",
  "1. Bấm 'Thêm mới' trong panel danh mục\n2. Chọn Loại = Dịch vụ có tính DT\n3. Nhập Tên, Giá vốn, Giá bán, VAT\n4. Lưu",
  "Tên: 'DV Test 1'; Giá vốn 100.000; Giá bán 150.000; VAT 8",
  "Tạo cost mới trong costs ERP (kind_of=2, type=null, revenue_calculation=1, rate_value_capital≈66.67); tự chọn vào dòng; giá vốn/giá bán/VAT điền sẵn vào dòng dịch vụ",
  "Cao"),
 ("TC-03","A. Danh mục DV/CP","Tạo nhanh dịch vụ trên BOM (ẩn giá)",
  "Đang ở popup Dịch vụ trên form BOM",
  "1. Bấm 'Thêm mới'\n2. Quan sát các field\n3. Nhập Loại + Tên\n4. Lưu",
  "Loại = Chi phí khác; Tên: 'CP Test 1'",
  "Popup BOM CHỈ có Loại + Tên (KHÔNG có Giá vốn/Giá bán/VAT), kèm dòng 'Giá & VAT sẽ nhập ở bước Báo giá'; tạo cost với vat=0, rate=0",
  "Cao"),
 ("TC-04","A. Danh mục DV/CP","Validate tạo nhanh thiếu thông tin",
  "Đang ở popup tạo nhanh (Báo giá)",
  "1. Để trống Loại / Tên / Giá\n2. Bấm Lưu",
  "-",
  "Hiện lỗi inline đỏ tại field bắt buộc + toast; không gọi API; không đóng popup",
  "TB"),

 # B. BOM - Phần A Hàng hoá
 ("TC-05","B. BOM Hàng hoá","Form BOM luôn hiển thị 2 phần A & B",
  "Mở form tạo BOM",
  "1. Quan sát bảng Chi tiết BOM LIST",
  "-",
  "Luôn thấy 2 dải header xanh: 'A — Hàng hoá' và 'B — Dịch vụ & Chi phí khác' (kể cả khi chưa có dữ liệu)",
  "Cao"),
 ("TC-06","B. BOM Hàng hoá","Thêm hàng hoá ERP vào phần A",
  "Form BOM, loại = thành phần (không phải gộp)",
  "1. Bấm 'Thêm mới' ở header A (popup chỉ tab Hàng hoá)\n2. Chọn 1+ hàng ERP\n3. Áp dụng",
  "Chọn 2 hàng hoá",
  "Popup chỉ có nội dung Hàng hoá (không có tab Dịch vụ); hàng được thêm vào phần A; SL nhập được, căn giữa",
  "Cao"),
 ("TC-07","B. BOM Hàng hoá","Thêm hàng tạm (thủ công) không có ô Giá nhập",
  "Form BOM, popup Hàng hoá đang mở",
  "1. Bấm 'Thêm hàng tạm'\n2. Quan sát form",
  "-",
  "Form Thêm hàng tạm KHÔNG còn ô 'Giá nhập'; chỉ có Tên/Mã/Model/Thương hiệu/Xuất xứ/ĐVT/SL/Ghi chú",
  "Cao"),
 ("TC-08","B. BOM Hàng hoá","BOM không còn cột giá",
  "Form BOM có hàng hoá",
  "1. Quan sát các cột bảng + nút cấu hình cột",
  "-",
  "KHÔNG còn cột 'Giá nhập', 'Thành tiền nhập', 'Tỷ suất LN', 'Giá bán', 'Thành tiền bán'; nút cấu hình cột không liệt kê các cột giá",
  "Cao"),
 ("TC-09","B. BOM Hàng hoá","Kéo-thả xếp lại thứ tự hàng hoá",
  "Form BOM có ≥2 hàng cha",
  "1. Kéo icon Move ở cột Thao tác để đổi thứ tự\n2. Lưu, reload",
  "-",
  "Thứ tự thay đổi theo thao tác; sau lưu/reload giữ đúng thứ tự (sort_order)",
  "TB"),

 # C. BOM - Phần B Dịch vụ & Chi phí
 ("TC-10","C. BOM Dịch vụ","Thêm dịch vụ từ danh mục vào phần B",
  "Form BOM",
  "1. Bấm 'Thêm mới' ở header B (popup chỉ tab Dịch vụ)\n2. Chọn 1 dịch vụ\n3. Quan sát dòng",
  "Chọn 'DV Test 1'",
  "Popup chỉ Dịch vụ (không tab Hàng hoá, KHÔNG có ô chọn nhóm hàng hoá); dòng dịch vụ vào phần B; SL=1 căn giữa; icon phân biệt DV/CP đúng",
  "Cao"),
 ("TC-11","C. BOM Dịch vụ","Section B nằm TRONG bảng Chi tiết (không tách rời)",
  "Form BOM có dịch vụ",
  "1. Quan sát vị trí phần B",
  "-",
  "Phần B là 1 nhóm nằm trong bảng Chi tiết BOM LIST (sau hàng hoá), không phải card độc lập tách rời",
  "Cao"),
 ("TC-12","C. BOM Dịch vụ","Kéo-thả + xoá dòng dịch vụ",
  "Form BOM có ≥2 dịch vụ",
  "1. Kéo Move đổi thứ tự dòng dịch vụ\n2. Xoá 1 dòng\n3. Lưu, reload",
  "-",
  "Đổi thứ tự + xoá hoạt động; lưu/reload giữ đúng",
  "TB"),
 ("TC-13","C. BOM Dịch vụ","Icon phân biệt Dịch vụ vs Chi phí khác",
  "Form BOM có 1 dịch vụ (rev=1) + 1 chi phí (rev=0)",
  "1. Quan sát icon đầu dòng",
  "-",
  "Dịch vụ: icon dịch vụ màu xanh; Chi phí khác: icon tiền màu cam",
  "TB"),
 ("TC-14","C. BOM Dịch vụ","Lưu & reload dịch vụ BOM",
  "Form BOM có hàng hoá + dịch vụ",
  "1. Lưu BOM\n2. Reload trang",
  "-",
  "Dịch vụ lưu vào bảng bom_list_service_items; sau reload hiển thị đúng tên/SL/icon, đúng thứ tự",
  "Cao"),

 # D. BOM - Gộp BOM
 ("TC-15","D. Gộp BOM","Gộp nhiều BOM con — hàng hoá",
  "Có ≥2 BOM con cùng loại tiền tệ, cùng cấu trúc nhóm",
  "1. Tạo BOM loại 'gộp'\n2. Chọn các BOM con\n3. Áp dụng",
  "2 BOM con có hàng hoá",
  "Hàng hoá của các BOM con được gộp vào BOM tổng hợp; nhóm trùng tên được gộp",
  "Cao"),
 ("TC-16","D. Gộp BOM","Gộp BOM con CÓ dịch vụ (bug đã fix)",
  "≥2 BOM con, mỗi BOM con có dịch vụ/chi phí khác",
  "1. Tạo BOM gộp\n2. Chọn các BOM con có dịch vụ\n3. Áp dụng\n4. Lưu, reload",
  "BOM con #1: 2 dịch vụ; BOM con #2: 1 dịch vụ",
  "BOM tổng hợp hiển thị ĐỦ 3 dịch vụ của cả 2 BOM con (trước đây bị mất); lưu/reload giữ nguyên",
  "Cao"),
 ("TC-17","D. Gộp BOM","Gộp BOM khác loại tiền tệ / khác cấu trúc",
  "BOM con khác currency hoặc 1 có nhóm 1 không",
  "1. Chọn các BOM con không đồng nhất\n2. Áp dụng",
  "-",
  "Hiện cảnh báo validate, không gộp",
  "TB"),

 # E. BOM - Khác
 ("TC-18","E. BOM Khác","Đổi loại BOM clear cả hàng hoá + dịch vụ",
  "Form BOM có hàng hoá + dịch vụ",
  "1. Đổi 'Loại BOM'\n2. Xác nhận",
  "-",
  "Popup xác nhận hiện kể cả khi chỉ có dịch vụ; sau xác nhận: cả hàng hoá VÀ dịch vụ bị xoá sạch",
  "Cao"),
 ("TC-19","E. BOM Khác","Màn XEM BOM hiển thị 2 phần A/B",
  "BOM đã lưu có hàng hoá + dịch vụ",
  "1. Mở màn xem BOM",
  "-",
  "Hiển thị 2 phần A/B (read-only) giống tạo/sửa; không có nút thêm/sửa/xoá; không có cột giá",
  "TB"),
 ("TC-20","E. BOM Khác","Cột Số lượng căn giữa đồng nhất",
  "BOM có hàng hoá + dịch vụ (tạo/sửa/xem)",
  "1. Quan sát cột Số lượng",
  "-",
  "SL của hàng hoá và dịch vụ đều căn giữa (đồng nhất)",
  "Thấp"),
 ("TC-21","E. BOM Khác","Nền row item",
  "BOM có dữ liệu",
  "1. Quan sát màu nền các dòng",
  "-",
  "Dòng hàng hoá/dịch vụ nền trắng; chỉ dòng tiêu đề nhóm + header A/B có nền màu",
  "Thấp"),

 # F. BOM -> Báo giá (luồng chảy)
 ("TC-22","F. BOM→Báo giá","Tạo báo giá từ BOM — hàng hoá chảy sang",
  "BOM đã duyệt có hàng hoá",
  "1. Tạo báo giá từ BOM (hoặc chọn dự án có BOM)\n2. Quan sát phần A",
  "-",
  "Toàn bộ hàng hoá BOM chảy sang phần A báo giá; giữ Model/Thương hiệu/Xuất xứ/TSKT; giá ERP load tại báo giá",
  "Cao"),
 ("TC-23","F. BOM→Báo giá","Dịch vụ BOM chảy sang Báo giá",
  "BOM có dịch vụ/chi phí khác",
  "1. Tạo báo giá từ BOM\n2. Quan sát phần B",
  "-",
  "Dịch vụ/chi phí BOM được copy sang quotation_service_items (cost_id, tên, vat); hiển thị phần B; icon DV/CP đúng",
  "Cao"),
 ("TC-24","F. BOM→Báo giá","Nhập giá tại Báo giá",
  "Báo giá tạo từ BOM (giá BOM = 0)",
  "1. Nhập Giá vốn/Giá bán cho hàng hoá + dịch vụ\n2. Quan sát tổng",
  "-",
  "Nhập giá được; tổng theo cột ở header A/B + dòng TỔNG cuối bảng cập nhật đúng (A+B = TỔNG)",
  "Cao"),
 ("TC-25","F. BOM→Báo giá","Gộp BOM → Báo giá đủ dịch vụ",
  "BOM gộp (TC-16) có dịch vụ từ nhiều BOM con",
  "1. Tạo báo giá từ BOM gộp",
  "-",
  "Báo giá nhận đủ dịch vụ đã gộp từ các BOM con",
  "Cao"),

 # G. Báo giá tự tạo (direct)
 ("TC-26","G. Báo giá direct","Báo giá tự tạo có cột Thao tác đầu",
  "Tạo báo giá trực tiếp (không từ BOM)",
  "1. Quan sát bảng chi tiết",
  "-",
  "Cột 'Thao tác' nằm ĐẦU bảng (chứa Move + Xoá); không còn cột thao tác ở cuối",
  "Cao"),
 ("TC-27","G. Báo giá direct","Thêm mới tách 2 popup Hàng hoá / Dịch vụ",
  "Báo giá direct",
  "1. Bấm 'Thêm mới' ở header A\n2. Bấm 'Thêm mới' ở header B",
  "-",
  "A mở popup Hàng hoá; B mở popup Dịch vụ (đúng tab tương ứng)",
  "Cao"),
 ("TC-28","G. Báo giá direct","Kéo-thả hàng hoá + dịch vụ",
  "Báo giá direct có ≥2 hàng + ≥2 dịch vụ",
  "1. Kéo Move đổi thứ tự cha/con + dịch vụ\n2. Lưu, reload",
  "-",
  "Đổi thứ tự cha (kéo cả con theo), con trong cha, và dịch vụ; lưu/reload giữ đúng (sort_order)",
  "Cao"),
 ("TC-29","G. Báo giá direct","Thêm con vào hàng cha",
  "Báo giá direct có hàng cha tự tạo",
  "1. Bấm 'Thêm con' dưới tên hàng cha",
  "-",
  "Mở popup Hàng hoá (đúng tab); thêm con vào đúng cha",
  "TB"),

 # H. Báo giá - cột chi tiết / tổng / icon
 ("TC-30","H. Báo giá cột","4 cột Model/Thương hiệu/Xuất xứ/TSKT",
  "Báo giá (tạo/sửa/xem) có hàng hoá",
  "1. Quan sát sau cột Tên hàng",
  "-",
  "Có 4 cột Model, Thương hiệu, Xuất xứ, Thông số kỹ thuật (mặc định HIỆN); TSKT render HTML",
  "Cao"),
 ("TC-31","H. Báo giá cột","Nút Ẩn/Hiện cột chi tiết",
  "Báo giá có hàng hoá",
  "1. Bấm 'Ẩn cột chi tiết'\n2. Bấm lại 'Hiện cột chi tiết'",
  "-",
  "4 cột ẩn/hiện đúng; bảng + dòng tổng + header A/B vẫn thẳng cột (colspan đúng) ở mọi chế độ chiết khấu",
  "Cao"),
 ("TC-32","H. Báo giá cột","Tổng theo cột trên header A/B",
  "Báo giá có hàng hoá + dịch vụ có giá",
  "1. Quan sát số tổng trên header A và B",
  "-",
  "Tổng (Giá vốn nếu xem được, Thành tiền sau VAT...) căn đúng cột, thẳng hàng với dòng TỔNG cuối; A+B khớp TỔNG",
  "Cao"),
 ("TC-33","H. Báo giá cột","Icon DV/CP trên Báo giá",
  "Báo giá có dịch vụ (rev=1) + chi phí (rev=0)",
  "1. Quan sát dòng dịch vụ (sửa) + badge (xem)",
  "-",
  "Sửa: icon xanh (DV) / cam (CP); Xem: badge 'DV' xanh / 'CP' cam",
  "TB"),

 # I. Báo giá - lưu / xem
 ("TC-34","I. Báo giá lưu","Lưu & reload báo giá",
  "Báo giá có hàng hoá + dịch vụ + giá",
  "1. Lưu\n2. Reload",
  "-",
  "Dữ liệu (hàng, dịch vụ, giá, VAT, thứ tự, 4 cột chi tiết) giữ nguyên sau reload",
  "Cao"),
 ("TC-35","I. Báo giá lưu","Màn XEM báo giá",
  "Báo giá đã lưu",
  "1. Mở màn xem",
  "-",
  "Hiển thị 2 phần A/B + 4 cột chi tiết + nút Ẩn/Hiện + tổng theo cột; icon/badge DV-CP đúng",
  "TB"),
 ("TC-36","I. Báo giá lưu","Xuất Excel / In",
  "Báo giá đã lưu có dịch vụ",
  "1. Xuất Excel\n2. Xem trước in",
  "-",
  "Nhóm dịch vụ hiển thị nhãn 'Dịch vụ & Chi phí khác' (đồng nhất); dữ liệu đúng",
  "TB"),

 # J. Edge / Regression
 ("TC-37","J. Edge","Popup không tự đóng khi focus input number",
  "Popup tạo nhanh dịch vụ / thêm hàng tạm có ô số (SL, VAT)",
  "1. Focus/bấm spinner ô số\n2. Dùng phím mũi tên tăng/giảm",
  "-",
  "Popup KHÔNG tự đóng; giá trị tăng/giảm bình thường",
  "Cao"),
 ("TC-38","J. Edge","Backward-compat dòng cũ (cost_id null / product_type=2)",
  "BOM/Báo giá cũ có dịch vụ kiểu cũ",
  "1. Mở BOM/Báo giá cũ",
  "-",
  "Vẫn hiển thị + lưu bình thường; icon mặc định Dịch vụ nếu không xác định loại",
  "TB"),
 ("TC-39","J. Edge","SL dịch vụ luôn = 1",
  "Thêm dịch vụ từ danh mục",
  "1. Quan sát + thử sửa SL dịch vụ",
  "-",
  "SL dịch vụ mặc định 1, không sửa; payload gửi qty=1",
  "TB"),
 ("TC-40","J. Edge","Tạo nhanh ghi đúng danh mục ERP",
  "Tạo nhanh dịch vụ mới",
  "1. Tạo nhanh\n2. Mở lại popup tìm kiếm danh mục",
  "-",
  "Dịch vụ vừa tạo xuất hiện trong danh mục costs ERP (dùng lại được lần sau)",
  "TB"),
]

# ---------- Build testcase.xlsx ----------
wb = Workbook()
ws = wb.active
ws.title = "BOM-Quotation"
headers = ["ID","Nhóm","Chức năng","Tiền điều kiện","Các bước thực hiện","Dữ liệu test","Kết quả mong đợi","Ưu tiên"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F6F54")
hfont = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="CFCFCF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
widths=[10,18,30,28,40,22,46,10]
from openpyxl.utils import get_column_letter
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w
prio_fill={"Cao":PatternFill("solid",fgColor="FDE2E1"),"TB":PatternFill("solid",fgColor="FFF3CD"),"Thấp":PatternFill("solid",fgColor="E2F0D9")}
for row in TC:
    ws.append(list(row))
    r=ws.max_row
    for ci,c in enumerate(ws[r],1):
        c.alignment=Alignment(vertical="top",wrap_text=True); c.border=border; c.font=Font(size=10)
    ws.cell(r,8).fill=prio_fill.get(row[7],PatternFill())
    ws.cell(r,8).alignment=Alignment(horizontal="center",vertical="center")
ws.freeze_panes="A2"
ws.auto_filter.ref=f"A1:H{ws.max_row}"
wb.save(f"{OUTDIR}/testcase.xlsx")
print("Saved testcase.xlsx with", len(TC), "cases")

# ---------- Build changes-summary HTML ----------
def esc(s): return html.escape(str(s)).replace("\n","<br>")

tc_rows = "\n".join(
 f"<tr><td class=id>{esc(t[0])}</td><td>{esc(t[1])}</td><td><b>{esc(t[2])}</b></td>"
 f"<td>{esc(t[4])}</td><td>{esc(t[6])}</td><td class=p p-{ 'cao' if t[7]=='Cao' else ('tb' if t[7]=='TB' else 'thap') }>{esc(t[7])}</td></tr>"
 for t in TC)

CSS = """
@page { size: A4; margin: 16mm 14mm; }
* { box-sizing: border-box; }
body { font-family: 'Segoe UI','Helvetica Neue',Arial,sans-serif; color:#1f2937; font-size:11px; line-height:1.5; }
h1 { font-size:21px; color:#0f5132; margin:0 0 2px; }
h2 { font-size:15px; color:#0f6f54; border-bottom:2px solid #9bdcc6; padding-bottom:4px; margin:18px 0 8px; }
h3 { font-size:12.5px; color:#155e44; margin:12px 0 4px; }
.sub { color:#6b7280; font-size:10.5px; margin-bottom:10px; }
ul { margin:4px 0 8px 0; padding-left:18px; }
li { margin:2px 0; }
.tag { display:inline-block; background:#e2f5ee; color:#0f8a63; border:1px solid #9bdcc6; border-radius:10px; padding:0 7px; font-size:9.5px; font-weight:600; margin-right:4px;}
table { width:100%; border-collapse:collapse; margin:6px 0 12px; }
th,td { border:1px solid #cfd8d3; padding:4px 6px; vertical-align:top; text-align:left; }
th { background:#1f6f54; color:#fff; font-size:10px; }
td { font-size:9.7px; }
td.id { font-weight:700; color:#0f5132; white-space:nowrap; }
td.p { text-align:center; font-weight:700; white-space:nowrap; }
.p-cao{ background:#fde2e1;} .p-tb{ background:#fff3cd;} .p-thap{ background:#e2f0d9;}
.box { background:#f6fbf9; border:1px solid #cfeae0; border-radius:6px; padding:8px 12px; margin:6px 0; }
.flow { background:#eef6ff; border:1px solid #cfe0f5; border-radius:6px; padding:8px 12px; font-size:10.5px;}
code { background:#eef2f0; padding:0 4px; border-radius:3px; font-size:9.7px; }
.files td { font-size:9.5px; }
.small { font-size:9.5px; color:#6b7280; }
"""

HTML = f"""<!doctype html><html lang="vi"><head><meta charset="utf-8"><style>{CSS}</style></head><body>
<h1>Tổng hợp thay đổi luồng BOM → Báo giá</h1>
<div class="sub">Feature: <b>erp-cost-catalog — Phase P5</b> · Ngày: {TODAY} · Phụ trách: @dnsnamdang · Module: Assign (BOM List + Quotation)</div>

<h2>1. Mục tiêu</h2>
<ul>
<li>Đồng nhất cách xây dựng nội dung (hàng hoá & dịch vụ) giữa <b>BOM</b> và <b>Báo giá tự tạo</b>.</li>
<li>Dịch vụ &amp; Chi phí khác lấy từ <b>danh mục costs ERP</b> (mysql2), tạo nhanh ghi thẳng danh mục.</li>
<li>Chuyển toàn bộ việc <b>nhập giá</b> sang bước Báo giá; BOM chỉ quản lý cấu trúc.</li>
</ul>

<h2>2. Thay đổi trên BOM</h2>
<h3>Cấu trúc &amp; hiển thị</h3>
<ul>
<li>Bảng Chi tiết BOM LIST luôn có <b>2 phần</b>: <span class="tag">A — Hàng hoá</span> <span class="tag">B — Dịch vụ &amp; Chi phí khác</span> (header xanh, nằm trong cùng bảng).</li>
<li>Dịch vụ/Chi phí lưu ở <b>bảng riêng</b> <code>bom_list_service_items</code> (không còn dùng <code>product_type=2</code>).</li>
<li>Mỗi dòng dịch vụ có <b>icon phân biệt</b>: Dịch vụ (xanh) / Chi phí khác (cam) theo <code>revenue_calculation</code>.</li>
<li>Cột <b>Số lượng căn giữa</b> đồng nhất; <b>bỏ nền</b> các dòng item (chỉ giữ nền nhóm + header A/B).</li>
</ul>
<h3>Bỏ toàn bộ cột giá trên BOM</h3>
<ul>
<li>Ẩn hết cột <b>Giá nhập / Thành tiền nhập / Tỷ suất LN / Giá bán / Thành tiền bán</b>; bỏ khỏi cấu hình cột.</li>
<li>Form "Thêm hàng tạm" + form sửa dòng: <b>bỏ ô Giá nhập</b>.</li>
<li>Popup tạo nhanh dịch vụ ở BOM: chỉ còn <b>Loại + Tên</b> (giá &amp; VAT nhập ở Báo giá).</li>
</ul>
<h3>Tương tác &amp; sửa lỗi</h3>
<ul>
<li><b>Kéo-thả</b> xếp lại thứ tự hàng hoá (cha/con) và dịch vụ.</li>
<li>Popup "Thêm mới" tách riêng theo phần: A → popup Hàng hoá, B → popup Dịch vụ (ẩn tab không liên quan); bỏ ô chọn nhóm ở tab Dịch vụ.</li>
<li><b>Đổi loại BOM</b> nay clear cả hàng hoá <b>và dịch vụ</b> (trước chỉ clear hàng hoá).</li>
<li><b>Gộp BOM</b> nay gộp cả <b>service_items</b> của các BOM con (trước bị mất khi gộp).</li>
<li>Popup không còn <b>tự đóng</b> khi focus/bấm input số (chặn bubble sự kiện).</li>
</ul>

<h2>3. Thay đổi trên Báo giá (Quotation)</h2>
<ul>
<li>2 phần <span class="tag">A — Hàng hoá</span> <span class="tag">B — Dịch vụ &amp; Chi phí khác</span> (header xanh) ở cả tạo/sửa/xem; <b>tổng cộng căn theo cột</b> trên header (thẳng với dòng TỔNG cuối).</li>
<li>Báo giá tự tạo: thêm <b>cột "Thao tác" ở đầu</b> (Move kéo-thả + Xoá), bỏ cột thao tác cuối; <b>kéo-thả</b> hàng hoá (cha/con) + dịch vụ.</li>
<li>Thêm <b>4 cột</b>: Model · Thương hiệu · Xuất xứ · Thông số kỹ thuật (sau cột Tên hàng) + nút <b>Ẩn/Hiện cột chi tiết</b> (mặc định hiện) — cả tạo/sửa/xem.</li>
<li>Popup tạo nhanh dịch vụ ở Báo giá: <b>đủ Giá vốn + Giá bán + VAT</b>; icon phân biệt DV/CP; SL dịch vụ = 1 (khóa).</li>
</ul>

<h2>4. Luồng dữ liệu BOM → Báo giá</h2>
<div class="flow">
<b>BOM</b> (cấu trúc, không giá)<br>
&nbsp;&nbsp;• Hàng hoá → <code>bom_list_products</code> &nbsp;|&nbsp; Dịch vụ/Chi phí → <code>bom_list_service_items</code><br>
&nbsp;&nbsp;↓ Tạo Báo giá từ BOM (<code>create</code> / <code>createFromBom</code>)<br>
<b>Báo giá</b>: hàng hoá → <code>quotation_product_prices</code> (giá ERP load tại đây) &nbsp;|&nbsp; dịch vụ → <code>quotation_service_items</code> (copy cost_id/tên/VAT)<br>
&nbsp;&nbsp;↓ Người dùng <b>nhập giá vốn / giá bán / VAT / chiết khấu</b> tại Báo giá<br>
&nbsp;&nbsp;↓ Tổng theo cột (A + B = TỔNG) + xuất Excel / trình duyệt.
</div>

<h2>5. Thay đổi Backend</h2>
<ul>
<li><code>DetailBomListResource</code> + <code>DetailQuotationResource</code>: trả thêm <code>revenue_calculation</code> cho service_items (tra <code>TpCost</code> theo cost_id) để FE hiện icon DV/CP.</li>
<li><code>DetailQuotationResource</code> (báo giá direct): thêm <code>orderBy('sort_order')</code> cho danh sách hàng hoá → giữ đúng thứ tự kéo-thả sau reload.</li>
<li>Quick-create ghi <code>costs</code> ERP (kind_of=2, revenue_calculation 1/0, type=null); BOM gửi vat=0/rate=0.</li>
<li>Sync dịch vụ qua <code>syncServiceItems</code> (BOM) / service_items (Báo giá) theo thứ tự mảng FE gửi.</li>
</ul>

<h2>6. File thay đổi chính</h2>
<table class="files">
<tr><th>Lớp</th><th>File</th><th>Nội dung</th></tr>
<tr><td>FE</td><td>BomBuilderTableCard.vue</td><td>2 phần A/B, bỏ cột giá, icon DV/CP, kéo-thả dịch vụ, SL căn giữa, bỏ nền row, tổng theo cột</td></tr>
<tr><td>FE</td><td>BomBuilderEditor.vue</td><td>serviceItems bảng riêng, ẩn cột giá, đổi loại clear dịch vụ, gộp BOM gộp dịch vụ, reorder</td></tr>
<tr><td>FE</td><td>BomBuilderAddProductModal.vue</td><td>tách popup goodsOnly/serviceOnly, bỏ ô Giá nhập + chọn nhóm tab DV, chặn bubble</td></tr>
<tr><td>FE</td><td>CostCatalogPanel / CostQuickCreateModal.vue</td><td>tạo nhanh, hidePricing (BOM), icon, chống tự đóng popup</td></tr>
<tr><td>FE</td><td>quotations/_id/edit.vue</td><td>cột Thao tác đầu, 4 cột chi tiết + Ẩn/Hiện, kéo-thả, 2 phần A/B + tổng cột, icon</td></tr>
<tr><td>FE</td><td>quotations/_id/index.vue</td><td>màn xem: 2 phần A/B, 4 cột chi tiết + Ẩn/Hiện, badge DV/CP, tổng cột</td></tr>
<tr><td>BE</td><td>DetailBomListResource / DetailQuotationResource</td><td>revenue_calculation, orderBy sort_order</td></tr>
<tr><td>BE</td><td>BomListService / QuotationService</td><td>sync service_items, copy BOM→Báo giá</td></tr>
</table>

<h2>7. Test case đầy đủ: BOM → Báo giá ({len(TC)} ca)</h2>
<div class="small">Ưu tiên: <span class="p p-cao" style="padding:0 6px">Cao</span> <span class="p p-tb" style="padding:0 6px">TB</span> <span class="p p-thap" style="padding:0 6px">Thấp</span> — chi tiết đầy đủ (tiền điều kiện, dữ liệu) xem file <b>testcase.xlsx</b>.</div>
<table>
<tr><th style="width:42px">ID</th><th style="width:90px">Nhóm</th><th style="width:150px">Chức năng</th><th>Các bước</th><th>Kết quả mong đợi</th><th style="width:34px">UT</th></tr>
{tc_rows}
</table>

<div class="small" style="margin-top:14px">— Hết — Tài liệu sinh tự động cho feature erp-cost-catalog (Phase P5).</div>
</body></html>"""

with open(f"{OUTDIR}/changes-summary.html","w",encoding="utf-8") as f:
    f.write(HTML)
print("Saved changes-summary.html")
