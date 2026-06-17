"""Generate testcase Excel cho feature Phân quyền hàng loạt (Bulk Permission)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = "/Users/manhcuong/Desktop/dns/HRM/.plans/bulk-permission/testcase.xlsx"
SHEET_NAME   = "PhanQuyenHangLoat"
FEATURE_NAME = "Phân quyền hàng loạt"
MODULE_NAME  = "Phân quyền hàng loạt"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
        "Cho phép admin CẤP (grant) hoặc THU HỒI (revoke) một tập permission cho NHIỀU nhân viên cùng lúc "
        "thông qua popup mở từ màn /timesheet/setting/roles (nút 'Phân quyền hàng loạt').\n"
        "► Lọc đối tượng theo: Khối / Phòng ban / Bộ phận / Chức vụ / Chức danh + NV bổ sung, có thể loại trừ thủ công.\n"
        "► Thao tác trên permission gán TRỰC TIẾP cho NV (model_has_permissions), KHÔNG đụng Role/RBAC.\n"
        "► Quyền cuối của NV = quyền từ Role ∪ quyền gán trực tiếp (cộng dồn)."),
    ("2. Đối tượng được tính / hiển thị",
        "► NV đang làm việc thuộc current_company của admin đăng nhập: employees.status = 1 VÀ employee_infos.status = 1 (đang làm việc).\n"
        "► NV khớp bộ lọc đối tượng (Khối/PB/BP/CV/CD) theo quy tắc kết hợp (xem mục 6).\n"
        "► NV trong danh sách 'NV bổ sung' (OR ngoài cùng — kể cả không khớp bộ lọc chính).\n"
        "► Tab 'Đã chọn' = NV có tick (không nằm trong excluded). Tab 'Loại trừ' = NV bị bỏ tick (excluded)."),
    ("3. Đối tượng bị ẩn / không tính",
        "► NV nghỉ việc: employee_infos.status ≠ 1 (đã bị loại khỏi mọi danh sách — fix 8b).\n"
        "► NV thuộc company khác current_company.\n"
        "► NV bị bỏ tick (excluded) → không nằm trong 'Tổng SL NV sẽ cập nhật', không bị apply.\n"
        "► Ô 'NV bổ sung' KHÔNG hiển thị NV đã thuộc bộ lọc chính (chỉ hiện NV ngoại lệ)."),
    ("4. Bộ lọc thời gian áp dụng cho",
        "— Không áp dụng. Feature không có bộ lọc thời gian; chỉ lọc theo cơ cấu tổ chức + danh sách NV."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
        "► Khối (groups) → Phòng ban (departments.group_id) → Bộ phận (parts).\n"
        "► Chức vụ = employee_infos.employee_work_position_id (bảng working_positions).\n"
        "► Chức danh = employee_infos.employee_role_id (bảng titles — KHÁC bảng roles của RBAC).\n"
        "► Permission lấy từ bảng permissions (spatie). Cấp/thu hồi ghi vào model_has_permissions (global, chưa có cột company_id)."),
    ("6. Quy tắc cộng dồn / deduplicate",
        "► Khối/Phòng ban/Bộ phận: UNION drill-down (fix 7) — Khối không drill → lấy hết NV thuộc khối; "
        "khối có chọn phòng (thuộc khối) → chỉ phòng đó; phòng có chọn bộ phận → chỉ bộ phận đó. Các nhánh OR với nhau.\n"
        "► Chức vụ / Chức danh: AND với cụm tổ chức; OR trong cùng trường.\n"
        "► NV bổ sung: OR ngoài cùng (cộng thêm vào tập kết quả).\n"
        "► 'Tất cả' của 1 trường (all_*): chọn mọi NV CÓ gán trường đó (whereNotNull), loại NV chưa gán.\n"
        "► State checkbox lưu dạng excluded list (Set NV bị bỏ tick), không lưu 'đã chọn'."),
    ("7. Phân quyền cấp",
        "• 'Quản lý phân quyền' — quyền DUY NHẤT gác toàn bộ tính năng:\n"
        "   - Ẩn/hiện nút 'Phân quyền hàng loạt' trên màn /timesheet/setting/roles (FE check hasAPermission).\n"
        "   - Gắn middleware checkPermission cho 3 route: preview-employees, apply, export-employees, "
        "eligible-additional-employees → không có quyền gọi API trả 403.\n"
        "(Không thêm permission mới; không phân quyền theo cấp tổ chức cho chính tính năng này.)"),
    ("8. Cách tính các ô thống kê",
        "► Tab 'Tất cả (X)' = tổng NV match bộ lọc chính ∪ NV bổ sung (đang làm việc, đúng company).\n"
        "► Tab 'Đã chọn (Y)' = X − số NV bị loại trừ (excluded).\n"
        "► Tab 'Loại trừ (Z)' = số phần tử trong excludedEmployeeIds.\n"
        "► Ô 'Tổng SL NV sẽ cập nhật' = totalAffected − excludedEmployeeIds.size (realtime, = Y)."),
    ("9. Ghi chú đọc bảng",
        "► Cột bảng NV: ☑ checkbox | STT | Tên NV (mã - tên) | Phòng ban/Bộ phận | Chức vụ | Chức danh.\n"
        "► Pagination 10/25/50/100. Header checkbox chỉ tick/untick trang HIỆN TẠI; mix → indeterminate.\n"
        "► Excluded persist qua paging, đổi sub-filter, đổi bộ lọc chính (kèm banner cảnh báo).\n"
        "► Export Excel có cột 'Nguồn': Theo bộ lọc / Bổ sung / Loại trừ.\n"
        "► Testcase viết theo góc nhìn người dùng cuối (UI), không cover riêng API/BE."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("01", "Tài khoản CÓ quyền 'Quản lý phân quyền' thấy & dùng được tính năng", "P0",
        "Đăng nhập tài khoản admin company A đã được gán quyền 'Quản lý phân quyền'",
        "1. Vào màn /timesheet/setting/roles\n2. Quan sát khu vực nút bên cạnh 'Thêm mới'\n3. Click 'Phân quyền hàng loạt'",
        "User: admin có quyền 'Quản lý phân quyền'",
        "- Nút 'Phân quyền hàng loạt' hiển thị\n- Click mở được popup Phân quyền hàng loạt\n- Thao tác preview/apply/export hoạt động (không bị 403)",
        "Permission: 'Quản lý phân quyền' — gác toàn bộ tính năng"),
    ("02", "Tài khoản KHÔNG có quyền 'Quản lý phân quyền' không thấy nút", "P0",
        "Đăng nhập tài khoản KHÔNG được gán quyền 'Quản lý phân quyền'",
        "1. Vào màn /timesheet/setting/roles\n2. Quan sát khu vực nút",
        "User: không có quyền 'Quản lý phân quyền'",
        "- KHÔNG hiển thị nút 'Phân quyền hàng loạt'\n- Không có lối vào popup từ UI",
        "FE check hasAPermission('Quản lý phân quyền') để render nút"),
    ("03", "Gọi trực tiếp API khi không có quyền → 403", "P1",
        "Tài khoản đăng nhập hợp lệ nhưng KHÔNG có quyền 'Quản lý phân quyền'",
        "1. Gọi POST /timesheet/bulk-permissions/preview-employees (hoặc apply / export-employees)\n2. Quan sát response",
        "Token user thiếu quyền",
        "- API trả HTTP 403 (middleware checkPermission chặn)\n- Không trả dữ liệu NV, không thực thi cấp/thu hồi",
        "Middleware checkPermission:Quản lý phân quyền trên cả 4 route"),
    ("04", "Phạm vi dữ liệu giới hạn theo current_company của admin", "P0",
        "Admin thuộc company A. DB có NV ở company A và company B",
        "1. Mở popup, không chọn bộ lọc hoặc chọn 'Tất cả'\n2. Quan sát danh sách NV bị ảnh hưởng",
        "Admin: current_company = A",
        "- Chỉ NV thuộc company A hiển thị\n- NV company B KHÔNG xuất hiện trong danh sách và không bị apply",
        "Scope: employee_infos.company_id = current_company"),
]

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Mở popup Phân quyền hàng loạt từ màn Vai trò", "P0",
            "Đăng nhập admin có quyền 'Quản lý phân quyền', đang ở /timesheet/setting/roles",
            "1. Click nút 'Phân quyền hàng loạt'\n2. Quan sát popup",
            "—",
            "- Popup (modal kích thước xl) mở ra\n- Có khối 'Đối tượng phân quyền' (5 trường + NV bổ sung + Quyền)\n- Có khối 'DS NV bị ảnh hưởng' với 3 tab + bảng + phân trang\n- Footer có 3 nút: Hủy / Áp dụng và làm tiếp / Áp dụng",
            "—"),
        ("002", "Trạng thái mặc định khi mở popup", "P0",
            "Vừa mở popup, chưa thao tác gì",
            "1. Quan sát giá trị mặc định các trường",
            "—",
            "- Hành động mặc định = 'Cấp quyền' (grant) được chọn\n- Các trường bộ lọc và Quyền đang trống\n- Bảng NV chưa có dữ liệu (chưa chọn bộ lọc) hoặc rỗng\n- 'Tổng SL NV sẽ cập nhật' = 0",
            "Default action = grant"),
        ("003", "Đổi hành động Cấp ↔ Thu hồi", "P1",
            "Popup đang mở, đã chọn 1 bộ lọc cho ra ≥1 NV",
            "1. Đang ở 'Cấp quyền', tick chọn radio 'Thu hồi quyền'\n2. Quan sát danh sách NV",
            "—",
            "- Radio chuyển sang 'Thu hồi quyền'\n- Danh sách NV được fetch lại (auto-search theo action)\n- Không mất bộ lọc đã chọn",
            "Deep watcher trên action + objectFilter → fetch lại"),
        ("004", "Đóng popup bằng nút X hoặc Hủy", "P1",
            "Popup đang mở, đã nhập một số bộ lọc",
            "1. Click [X] góc trên (hoặc nút 'Hủy')\n2. Quan sát",
            "—",
            "- Popup đóng lại\n- KHÔNG thực hiện cấp/thu hồi quyền nào\n- Quay về màn /timesheet/setting/roles giữ nguyên",
            "Hủy = không apply"),
    ]),
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Lọc chỉ theo Khối", "P0",
            "Company A có Khối K1 chứa các phòng/NV; tổng NV thuộc K1 = 12 (đang làm việc)",
            "1. Chọn Khối = K1\n2. Quan sát danh sách + tab 'Tất cả'",
            "Khối: K1 (12 NV)",
            "- Danh sách hiển thị đúng 12 NV thuộc các phòng ban trong Khối K1\n- Tab 'Tất cả (12)'",
            "Khối lọc qua departments.group_id ∈ chọn"),
        ("002", "Lọc chỉ theo Phòng ban", "P0",
            "Phòng ban PB-KD1 có 5 NV đang làm việc",
            "1. Chọn Phòng ban = PB-KD1\n2. Quan sát",
            "Phòng ban: PB-KD1 (5 NV)",
            "- Hiển thị đúng 5 NV thuộc PB-KD1\n- Tab 'Tất cả (5)'",
            "employee_infos.department_id ∈ chọn"),
        ("003", "Cascade: chọn Khối → option Phòng ban/Bộ phận thu hẹp theo Khối", "P0",
            "Khối K1 gồm PB-KD1, PB-KD2; Khối K2 gồm PB-NS1",
            "1. Chọn Khối = K1\n2. Mở dropdown Phòng ban\n3. Mở dropdown Bộ phận",
            "Khối: K1",
            "- Dropdown Phòng ban CHỈ hiện PB thuộc K1 (PB-KD1, PB-KD2), không hiện PB-NS1\n- Dropdown Bộ phận chỉ hiện BP thuộc scope K1\n- Vẫn chọn được Phòng ban (không bị disable rỗng — fix 7)",
            "filteredDepartmentOptions theo group_id; fix bug chọn Khối xong không chọn được PB"),
        ("004", "Drill-down UNION: Khối A + Khối B + chọn Phòng X thuộc A", "P0",
            "Khối A có PB-X (3 NV) + PB-Y (4 NV); Khối B có 6 NV",
            "1. Chọn Khối = A, B\n2. Chọn thêm Phòng ban = PB-X (thuộc A)\n3. Quan sát",
            "Khối: A,B; Phòng ban: PB-X",
            "- Kết quả = chỉ PB-X của khối A (3) ∪ toàn bộ khối B (6) = 9 NV\n- KHÔNG lấy PB-Y vì khối A đã bị drill xuống PB-X",
            "BR fix 7 — UNION drill-down: khối có drill chỉ lấy nhánh được drill"),
        ("005", "Kết hợp AND: (Phòng ban) AND (Chức vụ)", "P0",
            "PB-KD1 có 5 NV, trong đó 2 NV có Chức vụ 'Nhân viên'",
            "1. Chọn Phòng ban = PB-KD1\n2. Chọn Chức vụ = Nhân viên\n3. Quan sát",
            "Phòng ban: PB-KD1; Chức vụ: Nhân viên",
            "- Hiển thị đúng 2 NV (giao của PB-KD1 và chức vụ Nhân viên)",
            "Chức vụ AND với cụm tổ chức"),
        ("006", "OR trong cùng trường: chọn 2 Chức danh", "P1",
            "Chức danh CD1 có 3 NV, CD2 có 2 NV (không trùng người)",
            "1. Chọn Chức danh = CD1, CD2\n2. Quan sát",
            "Chức danh: CD1, CD2",
            "- Hiển thị 5 NV (CD1 ∪ CD2)",
            "OR trong cùng 1 trường"),
        ("007", "Tùy chọn 'Tất cả' cho một trường", "P1",
            "Có NV chưa gán Bộ phận (part_id null) và NV đã gán Bộ phận",
            "1. Tick 'Tất cả' ở trường Bộ phận\n2. Quan sát",
            "Bộ phận: Tất cả",
            "- Chỉ hiện 1 chip 'Tất cả' (không liệt kê từng bộ phận)\n- Danh sách lấy mọi NV CÓ gán Bộ phận (whereNotNull), loại NV chưa gán bộ phận",
            "BR fix 2 — all_part = whereNotNull part_id"),
        ("008", "Không cho chọn trùng option đã chọn", "P2",
            "Đang chọn Phòng ban = PB-KD1",
            "1. Mở lại dropdown Phòng ban\n2. Quan sát option PB-KD1",
            "—",
            "- Option PB-KD1 hiển thị trạng thái đã chọn (bôi đen/disable), không thể chọn lại tạo trùng",
            "—"),
        ("009", "Xóa hết lựa chọn của một trường", "P2",
            "Đang chọn Khối = K1, K2",
            "1. Bấm nút xóa của trường Khối\n2. Quan sát",
            "—",
            "- Trường Khối được clear về rỗng\n- Danh sách NV fetch lại theo các trường còn lại (hoặc rỗng nếu không còn trường nào)",
            "—"),
        ("010", "NV bổ sung chỉ hiện NV KHÔNG thuộc bộ lọc chính", "P0",
            "Bộ lọc chính = PB-KD1 (gồm NV An, Bình). NV Cường thuộc PB khác",
            "1. Chọn Phòng ban = PB-KD1\n2. Mở dropdown 'NV bổ sung'\n3. Tìm 'An' và 'Cường'",
            "Phòng ban: PB-KD1",
            "- 'NV bổ sung' KHÔNG hiện An, Bình (đã thuộc bộ lọc)\n- Hiện Cường (không thuộc bộ lọc) → chọn được\n- Sau khi chọn Cường, danh sách NV = PB-KD1 ∪ Cường",
            "BR fix 6 — eligible-additional-employees: whereNotIn matchedIds"),
        ("011", "NV bổ sung được prune khi đổi bộ lọc khiến NV đó vào nhóm chính", "P2",
            "Đã chọn NV bổ sung = An (lúc đó An chưa thuộc bộ lọc)",
            "1. Sau đó chọn Phòng ban = phòng của An\n2. Quan sát chip NV bổ sung",
            "—",
            "- An tự bị gỡ khỏi 'NV bổ sung' vì giờ đã thuộc bộ lọc chính (tránh đếm trùng)",
            "FE prune NV đã chọn nay đã thuộc filter"),
        ("012", "Chọn quyền (permission) — required min 1", "P0",
            "Popup đang mở",
            "1. Mở dropdown 'Quyền'\n2. Tìm và chọn 2-3 permission",
            "Quyền: 'Xem chấm công', 'Duyệt đơn'",
            "- Dropdown có searchbox, hiển thị DS permission hoạt động\n- Chọn được nhiều permission, hiển thị dạng chip\n- Option đã chọn không chọn lại được",
            "permissions table; required min 1"),
        ("013", "Sub-filter: tìm theo tên/mã NV trong danh sách", "P1",
            "Danh sách hiện 30 NV, có NV mã 'NV001 - Nguyễn An'",
            "1. Gõ 'An' vào ô tìm NV của sub-filter (hoặc Enter)\n2. Quan sát",
            "Keyword: An",
            "- Bảng lọc còn các NV khớp 'An'\n- Bộ lọc chính + excluded KHÔNG bị ảnh hưởng (chỉ lọc hiển thị)",
            "Sub-filter lọc trong tập đã match"),
        ("014", "Sub-filter theo Phòng ban/Bộ phận/Chức vụ/Chức danh", "P2",
            "Danh sách gồm NV nhiều phòng ban",
            "1. Chọn sub-filter Phòng ban = PB-KD1\n2. Quan sát",
            "Sub Phòng ban: PB-KD1",
            "- Bảng chỉ hiển thị NV thuộc PB-KD1 trong tập kết quả hiện tại\n- Có nút xóa từng trường sub-filter",
            "—"),
        ("015", "Nút refresh (⟳) clear sub-filter, giữ excluded", "P2",
            "Đã nhập sub-filter keyword + đang loại trừ 2 NV (excluded)",
            "1. Bấm nút refresh (⟳)\n2. Quan sát",
            "—",
            "- Sub-filter được clear, trang về 1\n- Danh sách loại trừ (excluded) VẪN giữ nguyên 2 NV",
            "Refresh chỉ reset sub-filter + page"),
    ]),
    ("III", "TAB & THỐNG KÊ SỐ LƯỢNG", [
        ("001", "3 tab hiển thị đúng số đếm", "P0",
            "Bộ lọc cho ra 20 NV, đã bỏ tick 3 NV (excluded)",
            "1. Quan sát 3 tab phía trên bảng",
            "—",
            "- Tab 'Tất cả (20)'\n- Tab 'Đã chọn (17)'\n- Tab 'Loại trừ (3)'",
            "BR — Đã chọn = Tất cả − Loại trừ"),
        ("002", "Tab 'Đã chọn' chỉ hiện NV có tick", "P1",
            "20 NV, excluded 3 NV cụ thể",
            "1. Bấm tab 'Đã chọn'\n2. Quan sát danh sách",
            "—",
            "- Bảng chỉ hiển thị 17 NV đang được tick\n- 3 NV bị loại trừ không xuất hiện",
            "view = selected"),
        ("003", "Tab 'Loại trừ' chỉ hiện NV bị bỏ tick", "P1",
            "20 NV, excluded 3 NV cụ thể",
            "1. Bấm tab 'Loại trừ'\n2. Quan sát",
            "—",
            "- Bảng chỉ hiển thị đúng 3 NV bị bỏ tick",
            "view = excluded"),
        ("004", "Ô 'Tổng SL NV sẽ cập nhật' realtime", "P0",
            "Bộ lọc cho ra 20 NV, chưa loại trừ ai",
            "1. Xem footer 'Tổng SL NV sẽ cập nhật' = 20\n2. Bỏ tick 2 NV\n3. Xem lại footer",
            "—",
            "- Ban đầu = 20\n- Sau khi bỏ tick 2 NV → cập nhật ngay còn 18\n- = totalAffected − excluded.size",
            "Realtime = totalAffected − excludedEmployeeIds.size"),
    ]),
    ("IV", "DANH SÁCH NV & LOẠI TRỪ", [
        ("001", "Cột bảng NV hiển thị đúng", "P1",
            "Bộ lọc cho ra ≥1 NV có đầy đủ thông tin phòng/bộ phận/chức vụ/chức danh",
            "1. Quan sát các cột bảng",
            "—",
            "- Cột: ☑ checkbox | STT | Tên NV (mã - tên) | Phòng ban/Bộ phận | Chức vụ | Chức danh\n- Dữ liệu khớp với NV",
            "—"),
        ("002", "Mặc định tất cả NV đều được tick", "P0",
            "Vừa chọn bộ lọc cho ra 20 NV, chưa thao tác checkbox",
            "1. Quan sát checkbox các dòng + header",
            "—",
            "- Tất cả dòng đều tick sẵn\n- Header checkbox tick đầy\n- Tab 'Đã chọn (20)', 'Loại trừ (0)'",
            "BR — mặc định tick hết"),
        ("003", "Bỏ tick 1 NV → vào danh sách loại trừ", "P0",
            "20 NV đang tick hết",
            "1. Bỏ tick NV 'Nguyễn An'\n2. Quan sát tab + footer",
            "NV: Nguyễn An",
            "- Tab 'Loại trừ (1)', 'Đã chọn (19)'\n- 'Tổng SL NV sẽ cập nhật' = 19\n- Checkbox flip tức thì (UX không reload toàn bảng — fix 9)",
            "State lưu excluded list"),
        ("004", "Header checkbox tick/untick chỉ trang hiện tại", "P0",
            "50 NV, pagination 10/trang (5 trang), đang ở trang 1",
            "1. Bỏ tick header (untick trang 1)\n2. Sang trang 2\n3. Quan sát header trang 2",
            "perPage = 10",
            "- Untick header chỉ bỏ tick 10 NV trang 1 (thêm vào excluded)\n- Trang 2 các NV vẫn tick\n- Tab 'Loại trừ (10)'",
            "BR — header chỉ phạm vi trang hiện tại"),
        ("005", "Header indeterminate khi trang có mix tick/untick", "P1",
            "Trang 1 có 10 NV, bỏ tick 3 NV bất kỳ",
            "1. Bỏ tick 3/10 NV trang 1\n2. Quan sát ô header checkbox",
            "—",
            "- Header checkbox ở trạng thái indeterminate (gạch ngang, không phải tick đầy/trống)",
            "—"),
        ("006", "Excluded persist khi chuyển trang qua lại", "P0",
            "50 NV, perPage 10. Trang 1 bỏ tick NV 'An'",
            "1. Bỏ tick 'An' (trang 1)\n2. Sang trang 2 → quay lại trang 1\n3. Quan sát 'An'",
            "—",
            "- 'An' VẪN ở trạng thái bỏ tick khi quay lại trang 1\n- Tab 'Loại trừ' vẫn đếm 'An'",
            "Excluded persist qua paging"),
        ("007", "Excluded persist khi đổi sub-filter", "P1",
            "20 NV, đã bỏ tick 'An'",
            "1. Nhập sub-filter keyword 'B' (ẩn 'An' khỏi view)\n2. Xóa sub-filter\n3. Quan sát 'An'",
            "—",
            "- Sau khi xóa sub-filter, 'An' vẫn ở trạng thái loại trừ",
            "Excluded persist khi đổi sub-filter"),
        ("008", "Banner cảnh báo loại trừ + nút 'Xóa loại trừ'", "P0",
            "Đang loại trừ 4 NV",
            "1. Quan sát banner phía trên bảng\n2. Bấm 'Xóa loại trừ'\n3. Quan sát",
            "—",
            "- Banner cảnh báo: 'Đang loại trừ 4 nhân viên — [Xóa loại trừ]'\n- Sau khi bấm → excluded clear về 0, tất cả NV tick lại, banner ẩn\n- Tab 'Loại trừ (0)'",
            "Banner hiện khi excluded.size > 0"),
        ("009", "Excluded persist khi đổi bộ lọc chính + banner hiện", "P1",
            "Bộ lọc PB-KD1 (5 NV), đã bỏ tick 2 NV",
            "1. Thêm Phòng ban = PB-KD2 vào bộ lọc\n2. Quan sát banner + danh sách loại trừ",
            "—",
            "- Đổi bộ lọc KHÔNG tự clear excluded\n- Banner vẫn báo 'Đang loại trừ 2 nhân viên'\n- User chủ động bấm 'Xóa loại trừ' nếu muốn reset",
            "BR — đổi filter giữ excluded, hiện banner"),
        ("010", "Phân trang đổi số dòng/trang", "P2",
            "50 NV",
            "1. Đổi perPage từ 10 → 25\n2. Quan sát",
            "perPage: 10 → 25",
            "- Bảng hiển thị 25 dòng/trang\n- Excluded + tổng số đếm không thay đổi (chỉ đổi cách phân trang)",
            "perPage 10/25/50/100"),
    ]),
    ("V", "ÁP DỤNG / THU HỒI / EXPORT", [
        ("001", "Áp dụng CẤP quyền (grant) cộng dồn với Role", "P0",
            "NV X đang có quyền từ Role = A, B, C (không có D, E trực tiếp). Chọn NV X, action Cấp, quyền D, E",
            "1. Action = Cấp quyền\n2. Bộ lọc cho ra NV X\n3. Chọn Quyền = D, E\n4. Bấm 'Áp dụng'\n5. Xác nhận",
            "Action: grant; Quyền: D, E; NV: X",
            "- Toast thành công, popup đóng\n- Quyền cuối của NV X = A, B, C (từ Role) + D, E (trực tiếp) = A,B,C,D,E\n- Role của NV X không bị thay đổi",
            "BR — cộng dồn: Role ∪ permission trực tiếp"),
        ("002", "Cấp quyền idempotent (đã có thì bỏ qua)", "P1",
            "NV X đã có quyền trực tiếp D. Apply grant lại quyền D",
            "1. Action Cấp, NV X, Quyền = D\n2. Bấm 'Áp dụng'",
            "Action: grant; Quyền: D (NV đã có D)",
            "- Toast thành công, không lỗi trùng\n- NV X vẫn chỉ có 1 bản ghi quyền D (không nhân đôi)",
            "BR — grant skip nếu đã có"),
        ("003", "Thu hồi (revoke) quyền gán trực tiếp", "P0",
            "NV X có quyền trực tiếp D, E. Action Thu hồi, quyền D",
            "1. Action = Thu hồi quyền\n2. Bộ lọc cho ra NV X\n3. Chọn Quyền = D\n4. Bấm 'Áp dụng'",
            "Action: revoke; Quyền: D; NV: X",
            "- Toast thành công\n- NV X còn lại quyền trực tiếp E (D bị gỡ)",
            "BR — revoke xóa khỏi model_has_permissions"),
        ("004", "Thu hồi quyền đến từ Role → NV vẫn còn quyền", "P0",
            "NV X có quyền A do Role cấp (không gán trực tiếp). Action Thu hồi, quyền A",
            "1. Action = Thu hồi quyền\n2. NV X\n3. Chọn Quyền = A\n4. Bấm 'Áp dụng'",
            "Action: revoke; Quyền: A (đến từ Role)",
            "- Toast thành công\n- NV X VẪN còn quyền A (vì quyền đến từ Role, revoke chỉ đụng layer trực tiếp)\n- Role không bị thay đổi",
            "BR — revoke không đụng Role (hành vi mong muốn)"),
        ("005", "Apply chỉ tác động NV được tick (loại excluded)", "P0",
            "Bộ lọc 20 NV, bỏ tick 5 NV. Apply grant quyền D",
            "1. Bỏ tick 5 NV\n2. Footer 'Tổng SL NV sẽ cập nhật' = 15\n3. Bấm 'Áp dụng'",
            "15 NV được chọn, 5 NV loại trừ",
            "- Chỉ 15 NV được cấp quyền D\n- 5 NV bị loại trừ KHÔNG bị thay đổi quyền",
            "Apply trên tập đã chọn = totalAffected − excluded"),
        ("006", "Nút 'Áp dụng và làm tiếp' reset form, giữ popup mở", "P1",
            "Đã cấu hình action + bộ lọc + quyền cho ra ≥1 NV",
            "1. Bấm 'Áp dụng và làm tiếp'\n2. Quan sát sau khi thành công",
            "—",
            "- Toast thành công\n- Popup VẪN mở\n- Toàn bộ form reset về mặc định (action=grant, bộ lọc trống, quyền trống, excluded trống)",
            "continueAfter = reset form"),
        ("007", "Export Excel danh sách NV bị ảnh hưởng", "P1",
            "Bộ lọc cho ra NV gồm: NV theo bộ lọc, 1 NV bổ sung, 2 NV loại trừ",
            "1. Bấm 'Export Excel'\n2. Mở file tải về",
            "—",
            "- File Excel tải về (có Authorization Bearer — không 401, fix 3)\n- Có title 'DANH SÁCH NHÂN VIÊN ÁP DỤNG PHÂN QUYỀN HÀNG LOẠT', header tô màu, border đầy đủ\n- Cột 'Nguồn' đúng: 'Theo bộ lọc' / 'Bổ sung' / 'Loại trừ'",
            "BR fix 5 — khôi phục cột Nguồn"),
        ("008", "Export tôn trọng sub-filter hiện tại", "P2",
            "Danh sách 30 NV, đang sub-filter keyword 'An' còn 4 NV",
            "1. Bấm 'Export Excel'\n2. Mở file",
            "Sub keyword: An",
            "- File chỉ chứa các NV khớp filter chính + sub-filter hiện tại",
            "Export dùng cùng schema preview"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Validate thiếu đối tượng (chưa chọn trường nào)", "P0",
            "Popup mới mở, chưa chọn bộ lọc và chưa chọn NV bổ sung",
            "1. Chọn Quyền = D\n2. Bấm 'Áp dụng'",
            "Đối tượng: rỗng; Quyền: D",
            "- Báo lỗi inline tại khối 'Đối tượng phân quyền' (viền đỏ is-invalid + text invalid-feedback)\n- Lỗi chỉ hiện sau submit đầu (touched)\n- KHÔNG gọi apply / không thực thi",
            "BR validate — phải có ≥1 trong 6 trường (5 cấp + NV bổ sung)"),
        ("002", "Validate thiếu quyền (chưa chọn permission)", "P0",
            "Đã chọn bộ lọc cho ra ≥1 NV nhưng chưa chọn Quyền",
            "1. Bấm 'Áp dụng'",
            "Đối tượng: PB-KD1; Quyền: rỗng",
            "- Báo lỗi inline tại trường 'Quyền' (viền đỏ + text lỗi)\n- KHÔNG gọi apply",
            "BR validate — permission required min 1"),
        ("003", "Chặn submit khi loại trừ hết NV (0 NV được chọn)", "P0",
            "Bộ lọc cho ra 3 NV, bỏ tick cả 3 (Tổng SL = 0)",
            "1. Bỏ tick toàn bộ\n2. Chọn Quyền = D\n3. Bấm 'Áp dụng'",
            "Tổng SL NV sẽ cập nhật = 0",
            "- Toast lỗi: 'Phải có ít nhất 1 nhân viên được chọn để áp dụng quyền' (fix 8)\n- KHÔNG gọi apply\n- BE cũng phòng thủ trả 422 nếu vẫn lọt (defense-in-depth)",
            "BR fix 8 — totalToApply <= 0 bị chặn cả FE + BE"),
        ("004", "Bộ lọc cho ra 0 NV", "P1",
            "Chọn tổ hợp bộ lọc không có NV nào khớp (vd Khối K1 AND Chức vụ không tồn tại trong khối)",
            "1. Chọn tổ hợp bộ lọc trống kết quả\n2. Quan sát",
            "—",
            "- Bảng hiển thị trạng thái rỗng (không có NV)\n- Tab 'Tất cả (0)', Tổng SL = 0\n- Không apply được (rơi vào validate 0 NV)",
            "—"),
        ("005", "NV nghỉ việc không xuất hiện trong danh sách", "P0",
            "PB-KD1 có 5 NV đang làm + 2 NV đã nghỉ việc (employee_infos.status ≠ 1)",
            "1. Chọn Phòng ban = PB-KD1\n2. Quan sát danh sách + tab",
            "PB-KD1: 5 active + 2 nghỉ",
            "- Chỉ 5 NV đang làm việc hiển thị\n- 2 NV nghỉ việc KHÔNG xuất hiện và không bị apply",
            "BR fix 8b — lọc employee_infos.status = 1"),
        ("006", "NV nghỉ việc không hiện trong ô 'NV bổ sung'", "P2",
            "Có NV đã nghỉ việc thuộc company A",
            "1. Mở dropdown 'NV bổ sung', tìm NV đã nghỉ",
            "—",
            "- NV đã nghỉ việc KHÔNG xuất hiện trong option NV bổ sung",
            "BR fix 8b — eligibleAdditionalEmployees lọc status active"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Không apply cho NV company khác", "P0",
            "Admin company A. NV Y thuộc company B",
            "1. Mở popup, chọn 'Tất cả' các trường\n2. Tìm NV Y trong danh sách",
            "Admin: company A; NV Y: company B",
            "- NV Y không xuất hiện → không thể chọn → không bị apply\n- Nếu cố gửi NV Y qua payload → BE loại bỏ/không tác động (scope company)",
            "Scope employee_infos.company_id = current_company"),
        ("002", "Không có quyền 'Quản lý phân quyền' → mọi route trả 403", "P1",
            "Tài khoản thiếu quyền 'Quản lý phân quyền'",
            "1. Gọi lần lượt preview-employees / apply / export-employees / eligible-additional-employees",
            "Token thiếu quyền",
            "- Tất cả trả HTTP 403\n- Không lộ dữ liệu NV, không cấp/thu hồi quyền",
            "Middleware checkPermission trên 4 route"),
        ("003", "Bulk-permission KHÔNG đụng Role/RBAC", "P0",
            "NV X gắn Role 'Trưởng phòng'. Thực hiện grant + revoke quyền trực tiếp cho X",
            "1. Grant quyền D cho X\n2. Revoke quyền E cho X\n3. Kiểm tra Role của X",
            "—",
            "- Role 'Trưởng phòng' của X giữ nguyên (không thêm/bớt)\n- Chỉ layer permission trực tiếp thay đổi",
            "BR — chỉ thao tác model_has_permissions"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Luồng cấp quyền hàng loạt đầy đủ", "P0",
            "Admin company A có quyền 'Quản lý phân quyền'. PB-KD1 có 8 NV active",
            "1. Vào /timesheet/setting/roles → 'Phân quyền hàng loạt'\n2. Action = Cấp quyền\n3. Chọn Phòng ban = PB-KD1\n4. Bỏ tick 1 NV\n5. Chọn Quyền = D, E\n6. Bấm 'Áp dụng' → xác nhận",
            "PB-KD1 (8 NV), loại trừ 1, Quyền D,E",
            "- Footer 'Tổng SL NV sẽ cập nhật' = 7\n- Toast thành công, popup đóng\n- 7 NV được gán quyền D, E (cộng dồn với quyền cũ); NV bị loại trừ không đổi",
            "E2E grant"),
        ("002", "Luồng thu hồi rồi làm tiếp", "P1",
            "Admin có quyền. NV nhóm A có quyền trực tiếp D",
            "1. Action = Thu hồi, lọc nhóm A, Quyền = D\n2. Bấm 'Áp dụng và làm tiếp'\n3. Quan sát form\n4. Cấu hình tiếp nhóm B, grant quyền F → 'Áp dụng'",
            "—",
            "- Lần 1: nhóm A bị gỡ quyền D, form reset, popup vẫn mở\n- Lần 2: nhóm B được cấp quyền F, popup đóng\n- Hai lần thao tác độc lập, không dính state cũ",
            "E2E revoke + continueAfter"),
        ("003", "Luồng export đối chiếu trước khi apply", "P2",
            "Bộ lọc phức tạp (drill-down + NV bổ sung + loại trừ vài NV)",
            "1. Cấu hình bộ lọc + NV bổ sung + loại trừ\n2. Export Excel\n3. Đối chiếu cột Nguồn\n4. Quay lại popup → 'Áp dụng'",
            "—",
            "- File Excel đúng NV + cột Nguồn (Theo bộ lọc/Bổ sung/Loại trừ)\n- Apply tác động đúng tập NV đã đối chiếu (trừ loại trừ)",
            "E2E export → apply"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT  = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}")
t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",            '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",       '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",   '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",      '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL; cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [
        MODULE_NAME, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed", "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_CENTER if i == 5 else WRAP_TOP_LEFT
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    nlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, max(longest // 4, nlines * 15 + 20)))
    current_row += 1
    data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note, group="Phân quyền & truy cập")

for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                    allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}")
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Data rows: 18-{current_row-1}")
