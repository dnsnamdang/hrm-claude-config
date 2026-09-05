# -*- coding: utf-8 -*-
"""Sinh file Excel mo ta luong 'Phieu yeu cau xuat hang muon' ben ERP."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "luong_phieu_yeu_cau_xuat_hang_muon.xlsx")

C_TITLE = "1F3864"
C_HEAD = "2E75B6"
C_SUB = "DDEBF7"
C_WARN = "FFF2CC"
C_OK = "E2EFDA"
C_BAD = "FCE4E4"

thin = Side(style="thin", color="9DC3E6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def sheet(name, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def title(ws, text, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def header(ws, row, cols):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def rows(ws, start, data, fills=None, valign="top"):
    r = start
    for idx, line in enumerate(data):
        for i, v in enumerate(line, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(vertical=valign, wrap_text=True)
            c.border = BORDER
            if fills and fills[idx]:
                c.fill = PatternFill("solid", fgColor=fills[idx])
        r += 1
    return r


def section(ws, row, text, ncol):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="1F3864")
    c.fill = PatternFill("solid", fgColor=C_SUB)
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
    return row + 1


# ---------------------------------------------------------------- 1. TONG QUAN
ws = sheet("1. Tổng quan", [26, 100])
title(ws, "PHIẾU YÊU CẦU XUẤT HÀNG MƯỢN (ERP) — TỔNG QUAN NGHIỆP VỤ", 2)
r = 3
header(ws, r, ["Hạng mục", "Nội dung"])
r += 1
data = [
    ["Tên chứng từ", "Phiếu yêu cầu xuất hàng mượn (mã phiếu: PYCXHM-xxxxx)"],
    ["Chứng từ duyệt kèm theo", "Phiếu xuất hàng mượn (mã phiếu: PXHM-xxxxx) — do Kế toán kho lập khi duyệt"],
    ["Hệ thống", "ERP TanPhat (Laravel 8 + AngularJS/Blade) — module Kho / Quản lý hàng mượn"],
    ["Mục đích nghiệp vụ",
     "Khách hàng đã MƯỢN hàng qua phiếu \"Yêu cầu xuất hàng\" loại XUẤT MƯỢN. Khi khách KHÔNG trả hàng về kho mà "
     "chốt luôn số hàng đang mượn (đã có phiếu/đường xuất khác), NVKD lập phiếu này để đề nghị TẤT TOÁN số hàng "
     "đang mượn đó. Kế toán kho duyệt → hệ thống ghi nhận số hàng mượn đã được xử lý bằng đường khác "
     "(returned_by_other), phiếu xuất mượn gốc chuyển sang trạng thái ĐÃ TRẢ."],
    ["Điểm mấu chốt cần nhớ",
     "Phiếu này KHÔNG trừ tồn kho và KHÔNG nhập hàng về kho. Hàng đã ra khỏi kho từ phiếu xuất mượn gốc. "
     "Phiếu chỉ TẤT TOÁN công nợ hàng mượn trên phiếu xuất mượn gốc."],
    ["Đối tượng tham gia",
     "1) Người lập: NVKD — chính là người đã tạo phiếu xuất mượn gốc\n"
     "2) Người duyệt: nhân sự có quyền \"Kế toán kho\"\n"
     "(Luồng này KHÔNG có bước Trưởng phòng / Ban giám đốc duyệt)"],
    ["Điều kiện đầu vào (tiền đề)",
     "Tồn tại phiếu Yêu cầu xuất hàng (product_export_requests) thoả:\n"
     "• Loại xuất = Xuất mượn (export_type = 3)\n"
     "• Trạng thái = Đã hạch toán (status = 5)\n"
     "• Trạng thái mượn = Đã mượn (borrow_status = 2)\n"
     "• Người đăng nhập chính là người tạo phiếu xuất mượn đó (accessor can_borrow_export)"],
    ["Kết quả đầu ra",
     "• Phiếu yêu cầu chuyển trạng thái Đã duyệt / Không duyệt\n"
     "• Sinh phiếu xuất hàng mượn (PXHM) khi được duyệt\n"
     "• Cộng dồn borrow_returned_qty + returned_by_other trên dòng hàng của phiếu xuất mượn gốc\n"
     "• Khi tất toán hết: phiếu xuất mượn gốc → borrow_status = 3 (Đã trả)"],
    ["Vị trí trên menu (người lập)", "Kho → Quản lý hàng mượn → \"Phiếu Yêu cầu xuất hàng mượn\""],
    ["Vị trí trên menu (kế toán kho)",
     "Kho → \"Yêu cầu xuất hàng mượn\" (màn forAccounting), \"Phiếu xuất hàng mượn\",\n"
     "và mục Duyệt → \"Phiếu yêu cầu xuất hàng mượn chờ duyệt\""],
    ["File nguồn chính (BE)",
     "app/Http/Controllers/Warehouse/BorrowExportRequestController.php\n"
     "app/Http/Controllers/Warehouse/BorrowExportController.php\n"
     "app/Model/Warehouse/BorrowExportRequest.php · app/Model/Warehouse/BorrowExport.php"],
    ["File nguồn chính (FE)",
     "resources/views/warehouse/borrow_export_requests/{index,create,form,formJs,show,forAccounting}.blade.php\n"
     "resources/views/warehouse/borrow_exports/{index,create,form,formJs,show}.blade.php"],
]
rows(ws, r, data)
ws.freeze_panes = "A4"

# ------------------------------------------------------------- 2. SO DO LUONG
ws = sheet("2. Sơ đồ luồng", [110])
title(ws, "SƠ ĐỒ LUỒNG ĐI CỦA PHIẾU", 1)
diagram = """
                    [ NVKD ]                                              [ KẾ TOÁN KHO ]
                        |                                                        |
  (Tiền đề) Phiếu Yêu cầu xuất hàng — loại XUẤT MƯỢN                             |
      status = 5 (Đã hạch toán) · borrow_status = 2 (Đã mượn)                    |
                        |                                                        |
                        v                                                        |
  B1. Vào menu Phiếu Yêu cầu xuất hàng mượn → Thêm mới                           |
                        |                                                        |
                        v                                                        |
  B2. Chọn 1..n phiếu xuất mượn (popup lọc sẵn: đã hạch toán,                     |
      đã mượn, loại xuất mượn, do chính mình tạo)                                |
                        |                                                        |
                        v                                                        |
  B3. Hệ thống nạp hàng hoá đang mượn + số lượng còn được xuất                    |
                        |                                                        |
                        v                                                        |
  B4. Nhập SL xuất / ĐVT · Ghi chú (bắt buộc) · Đính kèm PDF (bắt buộc)          |
                        |                                                        |
                        v                                                        |
  B5. Lưu → sinh mã PYCXHM-xxxxx, trạng thái = CHỜ DUYỆT (2) ─────► thông báo ───┤
                                                                                 |
                                                                                 v
                                                              B6. Mở danh sách chờ duyệt → xem chi tiết
                                                                                 |
                                              ┌──────────────────────────────────┴──────────────────────────────┐
                                              |                                                                 |
                                              v                                                                 v
                          B7a. TỪ CHỐI (nhập lý do bắt buộc)                          B7b. DUYỆT = bấm "Tạo phiếu xuất hàng mượn"
                                              |                                                                 |
                                              v                                                                 v
                          status = 4 (KHÔNG DUYỆT) ──► thông báo về NVKD              B8. Nhập SL duyệt từng dòng → Lưu
                                              |                                                                 |
                                              v                                                                 v
                                         KẾT THÚC                                     • Sinh phiếu PXHM-xxxxx (Đã duyệt)
                                                                                      • Cộng borrow_returned_qty + returned_by_other
                                                                                        trên dòng hàng phiếu xuất mượn gốc
                                                                                      • Nếu tất toán hết → phiếu xuất mượn gốc
                                                                                        chuyển borrow_status = 3 (ĐÃ TRẢ)
                                                                                      • Phiếu yêu cầu → status = 1 (ĐÃ DUYỆT)
                                                                                                                |
                                                                                                                v
                                                                                            thông báo về NVKD ─► KẾT THÚC

  Ghi chú: KHÔNG có bước sửa / xoá phiếu. Không có bước duyệt của Trưởng phòng hay Ban giám đốc.
"""
r = 3
for line in diagram.strip("\n").split("\n"):
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name="Consolas", size=10)
    c.alignment = Alignment(vertical="center")
    r += 1

# ------------------------------------------------------------ 3. LUONG CHI TIET
ws = sheet("3. Luồng chi tiết theo bước", [7, 16, 30, 34, 34, 40, 18, 26])
title(ws, "LUỒNG ĐI CHI TIẾT THEO TỪNG BƯỚC", 8)
r = 3
header(ws, r, ["Bước", "Ai làm", "Màn hình / Route", "Hành động của người dùng",
               "Điều kiện để làm được", "Hệ thống xử lý gì", "Trạng thái phiếu sau bước",
               "Thông báo / Chứng từ sinh ra"])
r += 1
steps = [
    ["B0", "NVKD", "Yêu cầu xuất hàng — loại Xuất mượn\n(product_export_requests)",
     "Đã tạo và hoàn tất phiếu xuất mượn trước đó; hàng đã giao cho khách mượn",
     "Phiếu ở trạng thái Đã hạch toán (5) và Đã mượn (borrow_status = 2)",
     "Ghi nhận số lượng đã xuất mượn trên từng dòng hàng (base_exported_qty)",
     "(chưa có phiếu yêu cầu)", "—"],
    ["B1", "NVKD", "Danh sách phiếu\nGET borrowExportRequest.index?type=all\n→ nút Thêm mới (borrowExportRequest.create)",
     "Vào menu Kho → Quản lý hàng mượn → Phiếu Yêu cầu xuất hàng mượn → Thêm mới",
     "Đăng nhập; không cần quyền riêng để tạo",
     "Mở màn tạo mới (form AngularJS)",
     "—", "—"],
    ["B2", "NVKD", "Màn tạo mới — popup \"Phiếu yêu cầu xuất hàng\"\n(productExportRequest.searchData)",
     "Bấm dấu + ở ô \"Phiếu xuất mượn\", chọn 1 hoặc nhiều phiếu xuất mượn",
     "Popup chỉ hiện phiếu: status = 5, borrow_status = 2, export_type = 3\nvà do chính người đang đăng nhập tạo",
     "Gọi productExportRequest.getDataForBorrowExport để lấy dữ liệu phiếu mượn",
     "—", "—"],
    ["B3", "Hệ thống", "Màn tạo mới — bảng Chi tiết",
     "(tự động)",
     "Có ít nhất 1 phiếu xuất mượn được chọn",
     "Nạp danh sách hàng hoá đang mượn của các phiếu đã chọn; mỗi dòng tính\n"
     "SL còn được xuất = (base_exported_qty − borrow_returned_qty − SL đang chờ xử lý) / hệ số ĐVT",
     "—", "—"],
    ["B4", "NVKD", "Màn tạo mới",
     "Nhập SL xuất cho từng dòng hàng theo từng phiếu mượn; chọn ĐVT; nhập Ghi chú; đính kèm file PDF",
     "Ghi chú bắt buộc (≤255 ký tự); tối thiểu 1 file đính kèm và phải là PDF;\nphải có ít nhất 1 dòng SL > 0",
     "Kiểm tra hợp lệ tại client (dòng vượt SL đang mượn bị bôi đỏ, nút Lưu bị khoá)",
     "—", "—"],
    ["B5", "NVKD", "POST borrowExportRequest.store",
     "Bấm Lưu",
     "Qua toàn bộ validate ở cột Quy tắc (sheet 6)",
     "Trong 1 transaction: tạo phiếu, sinh mã PYCXHM-xxxxx, upload file lên S3 (thư mục borrow_export_requests),\n"
     "gán company_id/department_id/part_id theo người tạo, lưu liên kết các phiếu xuất mượn,\n"
     "lưu dòng hàng + dòng chi tiết theo từng phiếu mượn, tính đơn giá xuất bình quân",
     "2 — Chờ duyệt",
     "Thông báo tới TẤT CẢ nhân sự có quyền \"Kế toán kho\":\n\"<Người lập> vừa tạo yêu cầu xuất hàng mượn: <mã phiếu>\""],
    ["B6", "Kế toán kho", "borrowExportRequest.forAccounting\nhoặc borrowExportRequest.index?type=for-approve\n→ borrowExportRequest.show",
     "Mở danh sách yêu cầu chờ duyệt, mở chi tiết phiếu, xem file đính kèm",
     "Có quyền \"Kế toán kho\"; phiếu đang ở trạng thái Chờ duyệt (2)",
     "Hiển thị thông tin chung, danh sách phiếu xuất mượn liên quan, chi tiết hàng hoá, file đính kèm",
     "2 — Chờ duyệt", "—"],
    ["B7a", "Kế toán kho", "Màn chi tiết → nút Từ chối\nPOST borrowExportRequest.deny",
     "Bấm \"Từ chối\", nhập lý do trong popup, xác nhận",
     "Có quyền \"Kế toán kho\"; phiếu đang Chờ duyệt (2); lý do bắt buộc (≤255 ký tự)",
     "Ghi comment (lý do), approver_id, approved_time; đổi trạng thái phiếu",
     "4 — Không duyệt (kết thúc luồng)",
     "Thông báo về NGƯỜI LẬP:\n\"<Người duyệt> vừa từ chối yêu cầu xuất hàng mượn: <mã phiếu>\""],
    ["B7b", "Kế toán kho", "Màn chi tiết → nút \"Tạo phiếu xuất hàng mượn\"\n→ borrowExport.create?borrow_export_request_id=...",
     "Bấm \"Tạo phiếu xuất hàng mượn\" — đây CHÍNH LÀ thao tác duyệt",
     "Có quyền \"Kế toán kho\"; phiếu đang Chờ duyệt (2)",
     "Mở màn lập phiếu xuất hàng mượn, nạp lại dữ liệu từ phiếu yêu cầu (getDataForBorrowExport)",
     "2 — Chờ duyệt", "—"],
    ["B8", "Kế toán kho", "POST borrowExport.store",
     "Nhập số lượng duyệt cho từng dòng (có thể ít hơn số đề nghị), nhập ghi chú, bấm Lưu",
     "SL duyệt từng dòng ≤ SL đang mượn còn lại của dòng đó; ít nhất 1 dòng > 0",
     "Trong 1 transaction:\n"
     "① tạo phiếu xuất hàng mượn PXHM-xxxxx (trạng thái Đã duyệt) + dòng hàng/chi tiết\n"
     "② updateWarehouse(): ghi approved_qty về phiếu yêu cầu; cộng borrow_returned_qty và returned_by_other\n"
     "   trên dòng hàng của phiếu xuất mượn gốc (quy đổi về ĐVT cơ bản)\n"
     "③ nếu mọi dòng của phiếu xuất mượn gốc đã tất toán hết → borrow_status = 3 (Đã trả)\n"
     "④ approve(): phiếu yêu cầu → Đã duyệt, ghi người duyệt + thời gian duyệt",
     "1 — Đã duyệt (kết thúc luồng)",
     "Sinh phiếu PXHM-xxxxx.\nThông báo về NGƯỜI LẬP:\n\"<Người duyệt> vừa duyệt yêu cầu xuất hàng mượn: <mã phiếu>\""],
    ["B9", "NVKD / Kế toán kho", "borrowExportRequest.print · borrowExportRequest.exportList\n· warehouseInfo.borrowIndex",
     "In phiếu, xuất Excel danh sách phiếu, tra cứu Danh sách hàng mượn để đối chiếu",
     "Người lập hoặc người có quyền \"Kế toán kho\" (phiếu không ở trạng thái Đang tạo)",
     "In theo mẫu báo cáo \"Yêu cầu xuất hàng mượn\"; xuất file danh_sach_yeu_cau_xuat_hang_muon.xlsx",
     "không đổi", "—"],
]
fills = [None, None, None, None, None, C_SUB, None, C_BAD, C_OK, C_OK, None]
rows(ws, r, steps, fills)
ws.freeze_panes = "A4"

# ------------------------------------------------------------- 4. TRANG THAI
ws = sheet("4. Trạng thái phiếu", [10, 24, 12, 46, 44, 40])
title(ws, "TRẠNG THÁI PHIẾU & CHUYỂN TRẠNG THÁI", 6)
r = 3
r = section(ws, r, "A. Trạng thái phiếu yêu cầu xuất hàng mượn (borrow_export_requests.status)", 6)
header(ws, r, ["Giá trị", "Tên hiển thị", "Màu nhãn", "Ý nghĩa", "Ai nhìn thấy", "Hành động cho phép"])
r += 1
st = [
    [3, "Đang tạo", "Đỏ", "Trạng thái nháp có trong code nhưng luồng hiện tại KHÔNG dùng "
     "(khi lưu là vào thẳng Chờ duyệt)", "Chỉ người tạo", "Không (đã bỏ chức năng sửa/xoá)"],
    [2, "Chờ duyệt", "Đỏ", "Đã gửi, đang chờ Kế toán kho xử lý", "Người tạo + Kế toán kho + người có quyền xem theo cấp",
     "Kế toán kho: Từ chối hoặc Tạo phiếu xuất hàng mượn (duyệt)"],
    [1, "Đã duyệt", "Xanh", "Đã duyệt và đã sinh phiếu xuất hàng mượn PXHM", "Như trên", "Xem, in, xuất Excel"],
    [4, "Không duyệt", "Đỏ", "Bị từ chối, có ghi lý do ở trường comment", "Như trên", "Xem, in, xuất Excel"],
]
r = rows(ws, r, st, [None, C_WARN, C_OK, C_BAD])
r += 1
r = section(ws, r, "B. Trạng thái phiếu xuất hàng mượn (borrow_exports.status) — luôn = 1 (Đã duyệt) ngay khi lập", 6)
r += 1
r = section(ws, r, "C. Trạng thái mượn trên phiếu xuất mượn gốc (product_export_requests.borrow_status)", 6)
header(ws, r, ["Giá trị", "Tên hiển thị", "Màu nhãn", "Ý nghĩa", "Ai nhìn thấy", "Hành động cho phép"])
r += 1
st2 = [
    [1, "Chờ mượn", "—", "Đã lập phiếu xuất mượn nhưng chưa xuất kho xong", "—", "—"],
    [2, "Đã mượn", "—", "Khách đang giữ hàng — điều kiện để lập phiếu yêu cầu xuất hàng mượn", "—",
     "Trả hàng mượn / Gia hạn / Xuất bán hàng mượn / Xuất hàng mượn"],
    [3, "Đã trả", "—", "Toàn bộ hàng mượn đã được tất toán (trả về kho hoặc xử lý bằng đường khác)", "—",
     "Không còn thao tác hàng mượn"],
]
rows(ws, r, st2)

# --------------------------------------------------------- 5. PHAN QUYEN
ws = sheet("5. Phân quyền & tầm nhìn", [38, 58, 58])
title(ws, "PHÂN QUYỀN & PHẠM VI DỮ LIỆU NHÌN THẤY", 3)
r = 3
header(ws, r, ["Quyền", "Được làm gì trong luồng này", "Ghi chú kỹ thuật"])
r += 1
perms = [
    ["(không cần quyền)", "Tạo phiếu yêu cầu xuất hàng mượn cho chính phiếu xuất mượn mình đã tạo; "
     "xem lại phiếu do mình tạo", "Điều kiện nằm ở accessor can_borrow_export: bắt buộc created_by = người đăng nhập"],
    ["Kế toán kho", "Xem mọi phiếu (trừ trạng thái Đang tạo); Từ chối; Duyệt bằng cách lập phiếu xuất hàng mượn; "
     "vào màn Yêu cầu xuất hàng mượn (forAccounting) và Phiếu xuất hàng mượn",
     "canView(), canApprove(); toàn bộ nhóm route borrow_exports gắn middleware checkPermission:Kế toán kho"],
    ["Xem phiếu hàng mượn theo tổng công ty", "Ở màn danh sách type=all: xem phiếu của mọi công ty",
     "Thực tế bị vô hiệu — xem sheet 9, mục rủi ro"],
    ["Xem phiếu hàng mượn theo công ty", "Ở màn danh sách type=all: xem phiếu cùng công ty với mình + phiếu do mình tạo",
     "searchByFilter(): where company_id = công ty của mình OR created_by = mình"],
    ["Xem phiếu hàng mượn theo phòng ban", "Xem phiếu thuộc các phòng ban mình quản lý + phiếu do mình tạo",
     "Lấy danh sách phòng ban từ bảng employee_manage_departments"],
    ["(không có quyền xem nào)", "Chỉ xem phiếu do chính mình tạo", "where created_by = mình"],
]
r = rows(ws, r, perms)
r += 1
r = section(ws, r, "Các chế độ danh sách (tham số type trên URL)", 3)
header(ws, r, ["type", "Màn dùng ở đâu", "Lọc dữ liệu"])
r += 1
types = [
    ["all", "Menu \"Phiếu Yêu cầu xuất hàng mượn\"", "Theo quyền xem ở trên; ẩn phiếu Đang tạo của người khác"],
    ["accounting", "Màn Yêu cầu xuất hàng mượn của kế toán", "Bắt buộc có quyền Kế toán kho; lấy mọi phiếu trừ status 0 và 3"],
    ["for-approve", "Menu Duyệt → phiếu chờ duyệt", "status = 2 và cùng công ty với người đăng nhập"],
    ["return", "Phục vụ luồng trả hàng", "status = 1 và do chính mình tạo"],
    ["(không truyền)", "Mặc định", "Chỉ phiếu do chính mình tạo"],
]
rows(ws, r, types)
ws.freeze_panes = "A4"

# ------------------------------------------------------- 6. QUY TAC & VALIDATE
ws = sheet("6. Quy tắc & validate", [8, 30, 60, 46])
title(ws, "QUY TẮC NGHIỆP VỤ & KIỂM TRA DỮ LIỆU", 4)
r = 3
r = section(ws, r, "A. Khi NVKD lưu phiếu yêu cầu (POST borrowExportRequest.store)", 4)
header(ws, r, ["STT", "Trường / Đối tượng", "Quy tắc", "Thông báo lỗi hiển thị"])
r += 1
rules = [
    [1, "Phiếu xuất mượn (product_export_request_ids)", "Bắt buộc chọn ít nhất 1; phải tồn tại trong bảng product_export_requests", "Bắt buộc phải chọn"],
    [2, "Phiếu xuất mượn — kiểm tra lại ở BE", "Từng phiếu phải còn hợp lệ: do chính mình tạo và đang ở trạng thái mượn hợp lệ (can_borrow_export)", "Có yêu cầu xuất mượn không hợp lệ!"],
    [3, "Danh sách hàng hoá (products)", "Bắt buộc có ít nhất 1 dòng", "Bắt buộc phải chọn"],
    [4, "Số lượng xuất từng dòng", "Bắt buộc nhập, là số, ≥ 0, tối đa 6 chữ số (≤ 999.999)", "Bắt buộc nhập / Không hợp lệ / Phải lớn hơn 0 / Không được vượt quá 6 chữ số"],
    [5, "Số lượng xuất từng dòng — trần nghiệp vụ",
     "SL xuất ≤ (SL đã xuất mượn − SL đã tất toán − SL đang chờ xử lý) ÷ hệ số ĐVT.\n"
     "SL đang chờ xử lý gồm: phiếu trả hàng chưa hoàn tất + yêu cầu xuất bán hàng mượn đang chờ duyệt + "
     "yêu cầu xuất hàng mượn khác đang chờ duyệt", "Số lượng không hợp lệ"],
    [6, "Dòng chi tiết phải thuộc phiếu đã chọn", "product_export_request_id của từng dòng phải nằm trong danh sách phiếu mượn đã chọn", "Dữ liệu không hợp lệ"],
    [7, "Tổng thể", "Phải có ít nhất 1 dòng có số lượng > 0", "Không có thay đổi"],
    [8, "Ghi chú (note)", "BẮT BUỘC nhập, tối đa 255 ký tự", "Bắt buộc nhập / Không được vượt quá 255 ký tự"],
    [9, "File đính kèm (attachments)", "BẮT BUỘC ít nhất 1 file, chỉ chấp nhận định dạng PDF", "Bắt buộc phải chọn / Không hợp lệ"],
]
r = rows(ws, r, rules)
r += 1
r = section(ws, r, "B. Khi Kế toán kho từ chối (POST borrowExportRequest.deny)", 4)
header(ws, r, ["STT", "Trường / Đối tượng", "Quy tắc", "Thông báo lỗi hiển thị"])
r += 1
rules2 = [
    [1, "Quyền", "Phải có quyền Kế toán kho và phiếu đang ở trạng thái Chờ duyệt", "Không đủ quyền!"],
    [2, "Lý do từ chối (comment)", "Bắt buộc nhập, tối đa 255 ký tự", "Bắt buộc phải nhập / Không được vượt quá 255 ký tự"],
]
r = rows(ws, r, rules2)
r += 1
r = section(ws, r, "C. Khi Kế toán kho duyệt = lập phiếu xuất hàng mượn (POST borrowExport.store)", 4)
header(ws, r, ["STT", "Trường / Đối tượng", "Quy tắc", "Thông báo lỗi hiển thị"])
r += 1
rules3 = [
    [1, "Phiếu yêu cầu (borrow_export_request_id)", "Bắt buộc, phải tồn tại và còn duyệt được (Chờ duyệt + người duyệt có quyền Kế toán kho)", "Bắt buộc phải chọn / Không thể duyệt yêu cầu này!"],
    [2, "Số lượng duyệt từng dòng", "Bắt buộc nhập, là số, ≥ 0, tối đa 6 chữ số", "Bắt buộc nhập / Không hợp lệ / Không được vượt quá 6 chữ số"],
    [3, "Số lượng duyệt — trần nghiệp vụ", "≤ SL đang mượn còn lại của dòng đó (loại trừ chính phiếu yêu cầu đang duyệt khỏi phần đang chờ xử lý)", "Số lượng không hợp lệ"],
    [4, "Tổng thể", "Phải có ít nhất 1 dòng > 0", "Không có thay đổi"],
    [5, "Ghi chú", "Không bắt buộc, tối đa 255 ký tự", "Không được vượt quá 255 ký tự"],
]
r = rows(ws, r, rules3)
r += 1
r = section(ws, r, "D. Quy tắc chung khác", 4)
header(ws, r, ["STT", "Trường / Đối tượng", "Quy tắc", "Thông báo lỗi hiển thị"])
r += 1
rules4 = [
    [1, "Mã phiếu", "Sinh tự động sau khi lưu: PYCXHM-<5 chữ số theo id> (phiếu xuất hàng mượn: PXHM-<5 chữ số>)", "—"],
    [2, "Sửa / Xoá phiếu", "KHÔNG có — route edit/update/delete đã bị tắt trong mã nguồn", "—"],
    [3, "Đơn giá", "Đơn giá lấy theo hàng hoá; đơn giá xuất (export_price) tính bình quân theo số lượng của các dòng chi tiết", "—"],
    [4, "Quy đổi đơn vị tính", "Mọi so sánh số lượng đều quy về đơn vị cơ bản qua unit_coefficient", "—"],
    [5, "Tồn kho", "Phiếu KHÔNG tác động tồn kho — chỉ tất toán số hàng đang mượn", "—"],
]
rows(ws, r, rules4)
ws.freeze_panes = "A4"

# ------------------------------------------------------------ 7. BANG DU LIEU
ws = sheet("7. Bảng dữ liệu", [40, 34, 70])
title(ws, "BẢNG DỮ LIỆU & TRƯỜNG QUAN TRỌNG", 3)
r = 3
header(ws, r, ["Bảng", "Vai trò", "Trường đáng chú ý"])
r += 1
tables = [
    ["borrow_export_requests", "Phiếu yêu cầu xuất hàng mượn (bảng chính)",
     "code (PYCXHM-...), status (1/2/3/4), note, comment (lý do từ chối), attachments (danh sách URL S3, ngăn bởi dấu phẩy),\n"
     "approver_id, approved_time, created_by, updated_by, company_id, department_id, part_id, timestamps"],
    ["borrow_export_request_has_export_requests", "Liên kết nhiều-nhiều với phiếu xuất mượn gốc",
     "borrow_export_request_id, product_export_request_id"],
    ["borrow_export_request_products", "Dòng hàng hoá của phiếu yêu cầu",
     "parent_id, product_id, product_name, code, unit_id/unit_name, unit_coefficient, brand_*, model_*,\n"
     "qty (SL đề nghị), approved_qty (SL được duyệt), returned_qty, price, export_price"],
    ["borrow_export_request_product_details", "Tách dòng hàng theo từng phiếu xuất mượn gốc",
     "parent_id (dòng hàng), request_id (phiếu yêu cầu), product_export_request_id,\n"
     "product_export_request_detail_id, product_id, unit_id, qty, approved_qty"],
    ["borrow_exports", "Phiếu xuất hàng mượn — chứng từ duyệt",
     "code (PXHM-...), status (=1), borrow_export_request_id, note, created_by"],
    ["borrow_export_products / borrow_export_product_details", "Dòng hàng & chi tiết của phiếu xuất hàng mượn",
     "Cấu trúc tương tự bên phiếu yêu cầu; details giữ liên kết ngược về dòng của phiếu xuất mượn gốc"],
    ["product_export_requests", "Phiếu xuất mượn gốc (nguồn của luồng)",
     "type/export_type = 3 (Xuất mượn), status = 5 (Đã hạch toán),\nborrow_status: 1 Chờ mượn · 2 Đã mượn · 3 Đã trả"],
    ["product_export_request_details", "Dòng hàng của phiếu xuất mượn gốc — nơi bị cập nhật khi duyệt",
     "base_exported_qty (SL đã xuất mượn, ĐVT cơ bản),\nborrow_returned_qty (SL đã tất toán),\n"
     "returned_by_other (phần tất toán bằng đường khác — do chính luồng này cộng vào), export_price"],
    ["notifications", "Thông báo trong hệ thống", "url, content, receiver_id, status, created_by (đẩy realtime qua Redis)"],
]
rows(ws, r, tables)
ws.freeze_panes = "A4"

# ------------------------------------------------------ 8. LUONG LIEN QUAN
ws = sheet("8. Luồng liên quan", [34, 20, 26, 46, 40])
title(ws, "4 LUỒNG XỬ LÝ HÀNG MƯỢN — ĐỂ PHÂN BIỆT", 5)
r = 3
header(ws, r, ["Luồng", "Mã phiếu", "Người duyệt", "Dùng khi nào", "Tác động tới hàng mượn"])
r += 1
related = [
    ["Yêu cầu xuất hàng mượn\n(luồng đang mô tả)", "PYCXHM → PXHM", "Kế toán kho",
     "Chốt/tất toán số hàng khách đang mượn mà không nhập lại kho",
     "Cộng borrow_returned_qty + returned_by_other; tất toán hết → Đã trả"],
    ["Trả hàng mượn", "Phiếu yêu cầu nhập hàng\n(product_import_requests)", "Theo luồng nhập kho",
     "Khách trả hàng thật về kho",
     "Nhập hàng về kho, cộng borrow_returned_qty; tất toán hết → Đã trả"],
    ["Yêu cầu xuất bán hàng mượn", "PYCXBHM → PXBHM\n(borrow_sell_requests / borrow_sells)",
     "Trưởng phòng / Ban giám đốc (có cấu hình hạn duyệt) rồi Kế toán kho",
     "Khách quyết định MUA luôn số hàng đang mượn",
     "Chuyển hàng mượn thành hàng bán, ghi nhận doanh thu/công nợ"],
    ["Yêu cầu gia hạn hàng mượn", "borrow_extend_requests", "Người duyệt gia hạn (có cấu hình hạn duyệt)",
     "Khách cần mượn thêm thời gian",
     "Chỉ kéo dài hạn mượn, không đổi số lượng"],
]
r = rows(ws, r, related)
r += 1
r = section(ws, r, "Lưu ý: cả 4 luồng đều trừ chung vào \"số lượng đang mượn\" của phiếu xuất mượn gốc. "
                   "Số lượng của phiếu/yêu cầu đang chờ xử lý ở luồng này sẽ bị khoá lại, không cho luồng khác dùng trùng.", 5)
ws.freeze_panes = "A4"

# ------------------------------------------------------------- 9. GHI CHU
ws = sheet("9. Ghi chú & rủi ro", [8, 34, 56, 46])
title(ws, "GHI CHÚ TRIỂN KHAI & ĐIỂM CẦN LƯU Ý", 4)
r = 3
header(ws, r, ["STT", "Vấn đề", "Mô tả", "Ảnh hưởng / Đề xuất"])
r += 1
notes = [
    [1, "Quyền \"Xem phiếu hàng mượn theo tổng công ty\" đang bị vô hiệu",
     "Cuối hàm searchByFilter() luôn thêm điều kiện cứng company_id = công ty của người đăng nhập, "
     "áp cho mọi nhánh quyền — kể cả nhánh tổng công ty vừa mới bỏ qua điều kiện công ty ở phía trên.",
     "Người có quyền tổng công ty vẫn chỉ thấy phiếu của công ty mình. Cần rà lại nếu port sang HRM."],
    [2, "Không có chức năng Sửa / Xoá",
     "Các route edit / update / delete đã bị comment trong controller và file route. Trạng thái 3 (Đang tạo) "
     "vì vậy không bao giờ phát sinh.",
     "Sai sót chỉ xử lý được bằng cách Từ chối rồi lập phiếu mới."],
    [3, "Duyệt không có nút \"Duyệt\" riêng",
     "Thao tác duyệt chính là bấm \"Tạo phiếu xuất hàng mượn\" và lưu phiếu đó. "
     "Phiếu yêu cầu chỉ chuyển sang Đã duyệt sau khi phiếu xuất hàng mượn lưu thành công.",
     "Cần giữ nguyên ngữ nghĩa này khi viết tài liệu HDSD hoặc port sang hệ thống mới."],
    [4, "Kiểm tra người tạo rất chặt",
     "Chỉ chính người tạo phiếu xuất mượn mới lập được yêu cầu xuất hàng mượn cho phiếu đó "
     "(điều kiện created_by trong can_borrow_export).",
     "Nhân sự nghỉ việc / bàn giao khách sẽ không lập được phiếu — cần cơ chế bàn giao."],
    [5, "Điều kiện trạng thái ở can_borrow_export viết lỏng",
     "Biểu thức status != 5 && borrow_status != 2 khiến chỉ cần thoả MỘT trong hai điều kiện (đã hạch toán "
     "HOẶC đang ở trạng thái Đã mượn) là qua được.",
     "Trên giao diện popup đã lọc đủ cả 3 điều kiện nên ít lộ ra; nhưng API gọi trực tiếp thì lọt."],
    [6, "Thông báo tạo phiếu gửi cho toàn bộ Kế toán kho trong công ty",
     "Danh sách người nhận lấy theo quyền \"Kế toán kho\" và lọc theo công ty của người lập phiếu; "
     "KHÔNG lọc theo kho hay theo phòng ban (đoạn lọc theo thủ kho của kho đã bị comment).",
     "Kế toán kho không phụ trách kho liên quan vẫn nhận thông báo. Kế toán kho công ty khác thì không nhận."],
    [7, "Logo/letterhead khi in và xuất Excel",
     "Bản in lấy logo theo công ty của NGƯỜI TẠO phiếu; file Excel danh sách lấy header theo công ty của "
     "NGƯỜI ĐANG ĐĂNG NHẬP.",
     "Nếu port sang HRM phải sửa lại: lấy theo company_id ghi trên chứng từ (xem quy ước trong CLAUDE.md)."],
    [8, "File đính kèm",
     "Bắt buộc, chỉ nhận PDF, lưu trên S3 ở thư mục borrow_export_requests, ghép thành 1 chuỗi ngăn bởi dấu phẩy.",
     "Không có chức năng xoá/thay file sau khi đã lưu."],
]
rows(ws, r, notes)
ws.freeze_panes = "A4"

del wb["Sheet"]
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("saved:", OUT)
