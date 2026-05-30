import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Phan cong quan ly GP"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Phân công quản lý giải pháp (Assign/Solutions)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'Phân công quản lý GP'
PREFIX = 'PCQL'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(vertical='center')
    for col_idx in range(1, 14):
        ws.cell(row=current_row, column=col_idx).fill = section_fill
        ws.cell(row=current_row, column=col_idx).border = thin_border
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority, precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# === SECTION 1: GIAO DIỆN ===
write_section_row('1. GIAO DIỆN — Tab Nhân sự')

write_tc(1, 1, 'Hi���n thị button Phân công', 'P0',
    'User là Trưởng phòng GP (created_by) hoặc PM; GP status ≠ Đóng',
    '1. Mở giải pháp → Tab Nhân sự\n2. Quan sát action bar',
    'GP status = Đang triển khai',
    'Button "Phân công" (icon ri-user-settings-line) hiện bên trái button "Thêm nhân sự"')

write_tc(1, 2, 'Ẩn button Phân công khi GP Đóng', 'P0',
    'GP status = 2 (Đóng)',
    '1. Mở GP đ�� đóng → Tab Nhân sự\n2. Quan sát action bar',
    'GP status = 2',
    'KHÔNG hiện button "Phân công"')

write_tc(1, 3, 'Ẩn button Phân công khi user không có quyền', 'P0',
    'User không phải created_by và không phải pm_id',
    '1. Login bằng user là thành viên (không phải TP GP / PM)\n2. Mở GP → Tab Nhân sự',
    'User l�� member thường',
    'KHÔNG hiện button "Phân công"')

write_tc(1, 4, 'Hiển thị button Xem lịch sử phân công', 'P1',
    'User có quyền xem tab Nhân sự',
    '1. Mở GP → Tab Nhân sự\n2. Quan sát action bar',
    '',
    'Button "Xem lịch sử phân công" (icon ri-history-line) luôn hiện')

write_tc(1, 5, 'Popup Phân công — layout', 'P1',
    'User có quyền phân công',
    '1. Click button "Phân công"\n2. Quan sát popup',
    '',
    'Popup size lg, header icon + title "Phân công quản lý", footer có button "Xác nhận phân công" + "Đóng"')

write_tc(1, 6, 'Popup Phân công — radio loại', 'P0',
    'User mở popup phân công',
    '1. Quan sát radio buttons',
    '',
    'Có 2 radio: "Phân công PM" và "Phân công Leader". Radio disabled nếu user không có quyền tương ứng')

write_tc(1, 7, 'Popup Lịch sử — layout', 'P1',
    'User click "Xem lịch sử phân công"',
    '1. Click button "Xem lịch sử"\n2. Quan sát popup',
    '',
    'Popup size xl, bảng 8 cột: Ngày | Loại | Hạng mục | Người cũ | Người mới | Vai trò mới | Ghi chú | Người thực hiện')

# === SECTION 2: PHÂN CÔNG PM ===
write_section_row('2. CHỨC NĂNG — Phân công PM')

write_tc(2, 1, 'Phân công PM — happy path (GP không module)', 'P0',
    'User là TP GP; GP không có module; GP status ≠ Đóng',
    '1. Click "Phân công"\n2. Chọn "Phân công PM"\n3. Chọn vai trò mới cho PM cũ: "Chuyên viên"\n4. Chọn PM mới\n5. Nh���p ghi chú ≥50 ký tự\n6. Click "Xác nhận" → OK confirm',
    'PM mới: NV cùng phòng',
    'solutions.pm_id cập nhật = PM mới.\nPM cũ được thêm vào solution_members v��i vai trò "Chuyên viên", start_date = today.\nLog ghi nhận.\nNotification gửi cho PM cũ + PM mới.')

write_tc(2, 2, 'Phân công PM — happy path (GP có module)', 'P0',
    'User là TP GP; GP CÓ module; GP status ≠ Đóng',
    '1. Click "Phân công"\n2. Chọn "Phân công PM"\n3. Chọn vai trò mới\n4. Select "Module cho PM cũ" hiện → chọn module\n5. Chọn PM mới\n6. Ghi chú ≥50 ký tự\n7. Xác nhận',
    'GP có 2 module HM01, HM02',
    'solutions.pm_id cập nhật.\nPM cũ thêm vào solution_module_members c��a module đã chọn.\nLog ghi nhận old_manager_module_id.')

write_tc(2, 3, 'Phân công PM — không chọn vai trò mới', 'P0',
    'User là TP GP; GP status ≠ Đóng',
    '1. Click "Phân công"\n2. Chọn "Phân công PM"\n3. KHÔNG chọn vai trò mới\n4. Chọn PM mới\n5. Ghi chú ≥50 ký tự\n6. Xác nhận',
    '',
    'solutions.pm_id cập nhật.\nPM cũ KHÔNG được thêm vào bất kỳ danh sách nhân sự nào (rời dự án).\nLog ghi old_manager_new_role_id = null.')

write_tc(2, 4, 'Phân công PM — user không phải TP GP', 'P0',
    'User là PM nhưng KHÔNG phải created_by',
    '1. Radio "Phân công PM" bị disabled\nHOẶC: Gọi API trực tiếp',
    '',
    'Radio disabled (FE).\nN���u gọi API: trả 403 "Bạn không có quyền phân công PM"')

write_tc(2, 5, 'Phân công PM — notification nội dung', 'P1',
    'Phân công PM thành công',
    '1. Hoàn thành phân công PM\n2. Kiểm tra notification PM cũ + PM mới',
    '',
    'PM cũ nhận: "Bạn đã được thay thế vị trí PM giải pháp [code - name]. PM mới: [tên]. Vai trò mới: [role]"\nPM mới nhận: "Bạn được phân công làm PM giải pháp [code - name]"')

# === SECTION 3: PHÂN CÔNG LEADER ===
write_section_row('3. CHỨC NĂNG — Phân công Leader')

write_tc(3, 1, 'Phân công Leader — happy path', 'P0',
    'User là TP GP hoặc PM; GP có module',
    '1. Click "Phân công"\n2. Chọn "Phân công Leader"\n3. Chọn hạng mục → hiện Leader hiện tại\n4. Chọn vai trò mới cho Leader cũ\n5. Chọn Leader mới\n6. Ghi chú ≥50 ký tự\n7. Xác nhận',
    'Hạng mục HM01 có Leader A',
    'solution_modules.leader_id cập nhật = Leader mới.\nLeader cũ thêm vào solution_module_members với vai trò đã chọn.\nLog type=leader, solution_module_id = HM01.')

write_tc(3, 2, 'Phân công Leader — không chọn vai trò', 'P0',
    'User có quyền; GP có module',
    '1. Chọn "Phân công Leader"\n2. Chọn hạng mục\n3. KHÔNG chọn vai trò mới\n4. Chọn Leader mới\n5. Ghi chú\n6. Xác nhận',
    '',
    'Leader cũ rời hạng mục (không thêm vào member).\nLeader mới cập nhật.')

write_tc(3, 3, 'Phân công Leader — Leader mới đang là member', 'P1',
    'Leader mới đang là member trong hạng mục đó',
    '1. Chọn "Phân công Leader"\n2. Chọn hạng mục\n3. Chọn Leader mới = 1 member hiện có\n4. Xác nhận',
    'NV B là member HM01',
    'solution_modules.leader_id = NV B.\nRecord member của NV B trong solution_module_members KHÔNG bị xóa (giữ nguyên).')

write_tc(3, 4, 'Phân công Leader — PM th��c hiện', 'P0',
    'User là PM (pm_id) nhưng KHÔNG phải TP GP',
    '1. Click "Phân công"\n2. Radio "Phân công PM" disabled, "Phân công Leader" enabled\n3. Chọn Leader\n4. Hoàn thành flow',
    '',
    'PM được phép phân công Leader thành công.\nLog created_by = pm_id.')

# === SECTION 4: VALIDATION ===
write_section_row('4. VALIDATION & EDGE CASES')

write_tc(4, 1, 'Validate ghi chú < 50 ký tự', 'P0',
    'User điền form phân công',
    '1. Nhập ghi chú chỉ 30 ký tự\n2. Click Xác nhận',
    'Ghi chú: "Thay đổi PM" (10 chars)',
    'Hiện lỗi: "Ghi chú phải có ít nhất 50 ký tự (hiện 10)".\nKhông submit.')

write_tc(4, 2, 'Validate không chọn PM/Leader mới', 'P0',
    'User điền form nhưng bỏ trống PM/Leader mới',
    '1. Chọn loại phân công\n2. KHÔNG chọn người mới\n3. Click Xác nhận',
    '',
    'Hiện lỗi: "Vui lòng chọn người mới".\nKhông submit.')

write_tc(4, 3, 'Validate không chọn loại phân công', 'P0',
    'User mở popup nhưng không chọn radio',
    '1. Mở popup\n2. Không chọn loại\n3. Click Xác nhận',
    '',
    'Hiện lỗi: "Vui lòng chọn loại phân công".\nKhông submit.')

write_tc(4, 4, 'Validate module cho PM cũ (khi GP có module + chọn vai trò)', 'P0',
    'GP có module; User chọn type=pm + chọn vai trò mới',
    '1. Chọn PM, chọn vai trò mới\n2. Select "Module cho PM cũ" hiện\n3. KHÔNG chọn module\n4. Click Xác nhận',
    'GP có module',
    'Hiện lỗi: "Vui lòng chọn hạng mục cho PM cũ".\nKhông submit.')

write_tc(4, 5, 'Validate hạng mục khi chọn Leader', 'P0',
    'User chọn type=leader',
    '1. Chọn "Phân công Leader"\n2. KHÔNG chọn hạng mục\n3. Click Xác nhận',
    '',
    'Hiện lỗi: "Vui lòng chọn hạng mục".\nKhông submit.')

write_tc(4, 6, 'Confirm popup trước submit', 'P0',
    'User điền form hợp lệ',
    '1. Điền đầy đủ form\n2. Click "Xác nhận phân công"\n3. Confirm popup hiện\n4. Click "Huỷ"',
    '',
    'Không gọi API.\nModal phân công vẫn mở, dữ liệu form giữ nguyên.')

write_tc(4, 7, 'DS chọn chỉ hiện NV c��ng phòng', 'P1',
    'GP thuộc phòng A; C�� NV phòng A + phòng B',
    '1. Mở popup phân công\n2. Xem dropdown PM/Leader mới',
    'GP.department_id = phòng A',
    'Chỉ hiện NV thuộc phòng A.\nKhông hiện NV phòng B.\nKhông hiện PM/Leader hiện tại trong dropdown.')

write_tc(4, 8, 'GP tr���ng thái Đóng — API reject', 'P1',
    'GP status = 2; Gọi API trực tiếp (bypass FE)',
    '1. POST /assign-manager v��i solution_id GP đã đóng',
    'GP status=2',
    'Response 422: "Giải pháp đã đóng, không thể phân công."')

# === SECTION 5: LỊCH SỬ ===
write_section_row('5. LỊCH SỬ PHÂN CÔNG')

write_tc(5, 1, 'Xem lịch sử — có dữ liệu', 'P1',
    'GP đã có ít nhất 1 lần phân công',
    '1. Click "Xem lịch sử phân công"\n2. Quan sát bảng',
    '',
    'Bảng hiện đầy đủ 8 cột.\nDữ liệu sort mới nhất lên trên (created_at DESC).\nCột "Người cũ", "Người mới", "Người thực hiện" hiện tên đầy đủ.')

write_tc(5, 2, 'Xem lịch sử — không có dữ liệu', 'P2',
    'GP chưa t���ng phân công',
    '1. Click "Xem lịch sử"\n2. Quan sát popup',
    '',
    'Hiện text: "Chưa có lịch sử phân công."')

write_tc(5, 3, 'Lịch sử — phân biệt PM vs Leader', 'P1',
    'GP có cả log phân công PM và Leader',
    '1. Xem lịch sử\n2. Quan sát cột "Loại"',
    '',
    'Log PM: badge "Phân công PM" (xanh info).\nLog Leader: badge "Phân công Leader" (vàng warning).\nLog Leader có cột "Hạng mục" hiện tên, log PM cột này trống.')

write_tc(5, 4, 'Lịch sử — vai trò mới hiển thị', 'P1',
    'Có log có vai trò mới + log không có',
    '1. Xem lịch sử\n2. Quan sát cột "Vai trò mới (người cũ)"',
    '',
    'Nếu có vai trò: hiện tên role.\nNếu không có: hiện "Rời dự án".')

# === DATA VALIDATION ===
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
dv.error = "Chọn giá trị hợp lệ"
dv.errorTitle = "Giá trị không hợp lệ"
ws.add_data_validation(dv)
dv.add(f'L7:L{current_row}')

# === SAVE ===
output_path = '/Users/dnsnamdang/Documents/DNSMEDIA/websites/hrm/.plans/solution-manager-assignment/testcase.xlsx'
wb.save(output_path)
print(f'Testcase Excel saved: {output_path}')
print(f'Total test cases: {current_row - 7 - 5}')  # subtract section rows
