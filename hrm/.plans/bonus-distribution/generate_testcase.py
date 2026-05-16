import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Testcase Bonus Distribution"

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
data_font = Font(size=11)
data_alignment = Alignment(vertical='top', wrap_text=True)
summary_font = Font(size=11)

col_widths = {'A': 14, 'B': 18, 'C': 16, 'D': 40, 'E': 10, 'F': 30, 'G': 50, 'H': 20, 'I': 12, 'J': 55, 'K': 15, 'L': 14, 'M': 15}
for c, w in col_widths.items():
    ws.column_dimensions[c].width = w

ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Bảng chia thưởng cuối năm (Bonus Distribution)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

for i, (label, formula) in enumerate([
    ('Passed:', '=COUNTIF(L8:L500,"Passed")'), ('Failed:', '=COUNTIF(L8:L500,"Failed")'),
    ('Pending:', '=COUNTIF(L8:L500,"Pending")'), ('Not Executed:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng:', '=COUNTA(L8:L500)')]):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority', 'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data', 'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border

M = 'Payroll'
G = 'Bảng chia thưởng'
P = 'BD'
r = 7

def sec(row, title):
    ws.merge_cells(f'C{row}:M{row}')
    c = ws.cell(row=row, column=3, value=title); c.font = section_font; c.fill = section_fill
    for col in range(1, 14): ws.cell(row=row, column=col).border = thin_border

def tc(row, tc_id, func, pri, pre='', steps='', td='', exp=''):
    vals = [M, G, tc_id, func, pri, pre, steps, td, '', exp, '', 'Not Executed', '']
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v); c.font = data_font; c.alignment = data_alignment; c.border = thin_border

# I. DANH SÁCH
sec(r, 'I. TRANG DANH SÁCH BẢNG CHIA THƯỞNG (/payroll/bonus-distribution)'); r += 1
tc(r, f'{P}_001.001', 'Hiển thị trang danh sách', 'P0', 'Đăng nhập, có quyền Payroll', '1. Vào menu Tính lương > Bảng chia thưởng cuối năm', '', 'Hiển thị: tiêu đề, nút "Tạo bảng chia thưởng", bộ lọc (Công ty, PB, Năm, Trạng thái, Người tạo, Người duyệt, Từ ngày, Đến ngày), bảng danh sách'); r += 1
tc(r, f'{P}_001.002', 'Hiển thị dữ liệu đúng', 'P0', 'Có dữ liệu bảng chia thưởng', '1. Mở trang danh sách', '', 'Cột: STT, Mã phiếu (link), Công ty, Phòng ban, Năm, Tổng quỹ (format dấu phẩy), Người lập, Ngày lập, Người duyệt, Ngày duyệt, Trạng thái (badge màu), Thao tác. Sắp xếp mới nhất trước'); r += 1
tc(r, f'{P}_001.003', 'Lọc theo công ty', 'P0', '', '1. Chọn công ty trong bộ lọc\n2. Nhấn tìm kiếm', 'Company: Công ty A', 'Chỉ hiện bảng thuộc công ty đã chọn'); r += 1
tc(r, f'{P}_001.004', 'Lọc theo phòng ban', 'P0', '', '1. Chọn phòng ban\n2. Nhấn tìm kiếm', '', 'Chỉ hiện bảng thuộc phòng ban đã chọn'); r += 1
tc(r, f'{P}_001.005', 'Lọc theo năm', 'P0', '', '1. Chọn năm\n2. Nhấn tìm kiếm', 'Year: 2025', 'Chỉ hiện bảng của năm 2025'); r += 1
tc(r, f'{P}_001.006', 'Lọc theo trạng thái', 'P0', '', '1. Chọn trạng thái "Đã duyệt"\n2. Nhấn tìm kiếm', 'Status: Đã duyệt', 'Chỉ hiện bảng trạng thái Đã duyệt'); r += 1
tc(r, f'{P}_001.007', 'Lọc theo khoảng thời gian', 'P1', '', '1. Chọn Từ ngày, Đến ngày\n2. Nhấn tìm kiếm', '', 'Chỉ hiện bảng có ngày duyệt trong khoảng'); r += 1
tc(r, f'{P}_001.008', 'Phân trang', 'P1', 'Có > 10 bảng', '1. Đổi số bản ghi/trang\n2. Chuyển trang', '', 'Phân trang hoạt động đúng, tổng số chính xác'); r += 1
tc(r, f'{P}_001.009', 'Nút Xem/Sửa/Xoá theo trạng thái', 'P0', 'Có bảng ở các trạng thái khác nhau', '1. Kiểm tra cột Thao tác', '', 'Nháp/Không duyệt: hiện Xem + Sửa + Xoá. Chờ duyệt/Đã duyệt: chỉ hiện Xem'); r += 1
tc(r, f'{P}_001.010', 'Click mã phiếu → chi tiết', 'P0', '', '1. Click vào mã phiếu', '', 'Chuyển sang trang chi tiết bảng chia thưởng'); r += 1

# II. TẠO MỚI
sec(r, 'II. TẠO MỚI BẢNG CHIA THƯỞNG'); r += 1
tc(r, f'{P}_002.001', 'Mở form tạo mới', 'P0', '', '1. Nhấn "Tạo bảng chia thưởng"', '', 'Chuyển sang form tạo mới. Tiêu đề "Tạo bảng chia thưởng". Select phòng ban + năm enabled. Có nút HỦY, LƯU NHÁP, LƯU & GỬI DUYỆT'); r += 1
tc(r, f'{P}_002.002', 'Chọn phòng ban + năm → lấy quỹ thưởng từ ERP', 'P0', 'ERP có quyết toán đã duyệt cho PB + năm', '1. Chọn phòng ban\n2. Chọn năm', 'PB: Phòng KD, Năm: 2025', 'Ô "Tổng quỹ thưởng phòng" tự động hiển thị số tiền từ ERP (format dấu phẩy)'); r += 1
tc(r, f'{P}_002.003', 'Chọn PB + năm — ERP chưa có quyết toán', 'P1', 'ERP chưa có quyết toán đã duyệt', '1. Chọn PB + năm', '', 'Hiển thị lỗi "Chưa có quyết toán đã duyệt cho phòng ban này". Quỹ = 0'); r += 1
tc(r, f'{P}_002.004', 'Thêm thành phần tính thưởng', 'P0', '', '1. Nhấn "Thêm thành phần"\n2. Chọn thành phần từ dropdown\n3. Nhập % quỹ', '', 'Dòng mới hiển thị: dropdown TP, Loại (badge), Công thức, % quỹ (input), Quỹ TP (auto), Checkbox điều chỉnh, Nút xoá'); r += 1
tc(r, f'{P}_002.005', 'Không cho chọn trùng thành phần', 'P0', 'Đã chọn TP001 ở dòng 1', '1. Thêm dòng mới\n2. Mở dropdown', '', 'Dropdown không hiển thị TP001 (đã chọn). Chỉ hiện các TP chưa chọn'); r += 1
tc(r, f'{P}_002.006', 'Validate % quỹ tổng không vượt 100', 'P0', 'Dòng 1: 60%', '1. Dòng 2 nhập 50%', '', 'Tự động giảm về 40% (max allowed). Hiện toast "Tổng % quỹ không được vượt quá 100%. Tối đa: 40%"'); r += 1
tc(r, f'{P}_002.007', 'Xoá thành phần trong Select2', 'P1', 'Đã chọn TP001', '1. Nhấn X xoá trong Select2', '', 'Thông tin Loại, Công thức bị reset về trống'); r += 1
tc(r, f'{P}_002.008', 'Xoá dòng thành phần', 'P0', 'Có 2 dòng TP', '1. Nhấn icon xoá dòng 2', '', 'Dòng 2 biến mất. Tổng % cập nhật'); r += 1
tc(r, f'{P}_002.009', 'Lưu nháp thành công', 'P0', 'Đã thiết lập TP, có quỹ', '1. Nhấn "LƯU NHÁP"', '', 'Toast "Lưu thành công". Mã phiếu tự sinh. Trạng thái = Nháp'); r += 1
tc(r, f'{P}_002.010', 'Lưu & Gửi duyệt thành công', 'P0', 'Tổng % = 100%, có quỹ', '1. Nhấn "LƯU & GỬI DUYỆT"', '', 'Toast "Gửi duyệt thành công". Trạng thái → Chờ duyệt. Form chuyển readonly'); r += 1
tc(r, f'{P}_002.011', 'Validate trùng PB + năm', 'P0', 'Đã có bảng PB KD năm 2025', '1. Tạo mới, chọn PB KD + 2025\n2. Lưu', '', 'Báo lỗi "Phòng ban này đã có bảng chia thưởng cho năm 2025"'); r += 1

# III. TÍNH BẢNG THƯỞNG
sec(r, 'III. TÍNH BẢNG THƯỞNG (PREVIEW CALCULATE)'); r += 1
tc(r, f'{P}_003.001', 'Nút Tính bảng thưởng disabled khi % ≠ 100', 'P0', 'Tổng % = 80%', '1. Kiểm tra nút "Tính bảng thưởng"', '', 'Nút bị disabled (xám). Không click được'); r += 1
tc(r, f'{P}_003.002', 'Nút Tính bảng thưởng disabled khi không có quỹ', 'P0', 'Tổng % = 100%, quỹ = 0', '1. Kiểm tra nút', '', 'Nút bị disabled'); r += 1
tc(r, f'{P}_003.003', 'Tính bảng thưởng — loading', 'P0', 'Tổng % = 100%, có quỹ', '1. Nhấn "Tính bảng thưởng"', '', 'Hiển thị spinner "Đang tính toán bảng thưởng...". Sau khi xong: hiện bảng phân bổ + toast thành công'); r += 1
tc(r, f'{P}_003.004', 'TP công thức — hiển thị căn cứ + số tiền', 'P0', 'TP công thức: luong_p1 + luong_p2', '1. Tính bảng thưởng', '', 'Cột "Căn cứ tính toán": hiển thị base_value (format dấu phẩy, không thập phân). Cột "Số tiền": hiển thị amount tính từ tỷ lệ base_value/totalBase * quỹ TP'); r += 1
tc(r, f'{P}_003.005', 'TP thủ công — hiển thị input % chia + số tiền', 'P0', 'TP mode manual', '1. Tính bảng thưởng', '', 'Cột "% chia": input nhập tay. Cột "Số tiền": input nhập tay. Mặc định 0'); r += 1
tc(r, f'{P}_003.006', 'Dòng tổng quỹ phòng (sticky)', 'P0', 'Đã tính xong', '1. Scroll dọc bảng phân bổ', '', 'Dòng TỔNG QUỸ PHÒNG cố định ở trên. Hiện: % phân bổ, quỹ TP, "Đã chia: xxx" (xanh nếu khớp, đỏ nếu chênh)'); r += 1
tc(r, f'{P}_003.007', 'Tổng thưởng mỗi NV', 'P0', 'Có nhiều TP', '1. Kiểm tra cột Tổng thưởng', '', 'Tổng = sum amount các TP. Hiển thị (x% quỹ) bên dưới'); r += 1
tc(r, f'{P}_003.008', 'Preview không lưu DB', 'P0', '', '1. Tính bảng thưởng\n2. Nhấn HỦY (không lưu)\n3. Quay lại trang danh sách', '', 'Bảng chia thưởng KHÔNG được tạo trong DB (vì chưa lưu)'); r += 1

# IV. ĐIỀU CHỈNH THỦ CÔNG
sec(r, 'IV. ĐIỀU CHỈNH THỦ CÔNG'); r += 1
tc(r, f'{P}_004.001', 'Nhập % chia → auto tính số tiền', 'P0', 'TP thủ công, đã tính bảng', '1. Nhập % chia = 10 cho NV A', '% chia: 10', 'Số tiền NV A = quỹ TP * 10%. Cập nhật realtime'); r += 1
tc(r, f'{P}_004.002', 'Nhập số tiền → auto tính % chia', 'P0', 'TP thủ công', '1. Nhập số tiền = 5000000 cho NV A', 'Amount: 5,000,000', '% chia = 5000000 / quỹ TP * 100. Cập nhật realtime'); r += 1
tc(r, f'{P}_004.003', 'Chia đều thành phần thủ công', 'P0', 'Có TP thủ công, 10 NV', '1. Nhấn "Chia đều thành phần thủ công"', '', 'Mỗi NV được chia đều: % = 100/10 = 10%, amount = quỹ TP / 10'); r += 1
tc(r, f'{P}_004.004', 'TP công thức có allow_adjustment', 'P1', 'TP công thức, checkbox điều chỉnh = true', '1. Tính bảng\n2. Sửa số tiền NV A', '', 'Input số tiền enabled, có thể sửa tay'); r += 1

# V. SỬA
sec(r, 'V. SỬA BẢNG CHIA THƯỞNG'); r += 1
tc(r, f'{P}_005.001', 'Mở form sửa — load data đúng', 'P0', 'Bảng trạng thái Nháp', '1. Nhấn icon sửa', '', 'Load đúng: PB (disabled), Năm (disabled), Quỹ, Ghi chú, Thành phần, Bảng phân bổ (nếu đã tính)'); r += 1
tc(r, f'{P}_005.002', 'Sửa thành phần + tính lại', 'P0', 'Đang sửa', '1. Thêm/bớt TP\n2. Sửa %\n3. Nhấn "Tính bảng thưởng"\n4. Lưu', '', 'Tính lại đúng với thành phần mới. Lưu thành công'); r += 1
tc(r, f'{P}_005.003', 'Không cho sửa khi trạng thái Chờ duyệt', 'P0', 'Bảng Chờ duyệt', '1. Mở chi tiết', '', 'Tất cả input disabled. Không có nút LƯU. Chỉ có nút HỦY'); r += 1
tc(r, f'{P}_005.004', 'Không cho sửa khi trạng thái Đã duyệt', 'P0', 'Bảng Đã duyệt', '1. Mở chi tiết', '', 'Tất cả input disabled. Không có nút LƯU'); r += 1
tc(r, f'{P}_005.005', 'Cho sửa khi trạng thái Không duyệt', 'P0', 'Bảng Không duyệt', '1. Mở chi tiết', '', 'Form editable. Có nút LƯU NHÁP, LƯU & GỬI DUYỆT'); r += 1

# VI. XOÁ
sec(r, 'VI. XOÁ BẢNG CHIA THƯỞNG'); r += 1
tc(r, f'{P}_006.001', 'Xoá bảng Nháp — thành công', 'P0', 'Bảng Nháp', '1. Nhấn icon xoá\n2. Xác nhận', '', 'Toast "Xóa thành công". Bảng biến mất khỏi danh sách'); r += 1
tc(r, f'{P}_006.002', 'Xoá bảng Không duyệt — thành công', 'P0', 'Bảng Không duyệt', '1. Nhấn xoá\n2. Xác nhận', '', 'Xoá thành công'); r += 1
tc(r, f'{P}_006.003', 'Không xoá được bảng Chờ duyệt', 'P0', 'Bảng Chờ duyệt', '1. Kiểm tra cột Thao tác', '', 'Không hiển thị nút xoá'); r += 1
tc(r, f'{P}_006.004', 'Huỷ xoá', 'P1', '', '1. Nhấn xoá\n2. Nhấn "Huỷ"', '', 'Modal đóng, bảng vẫn còn'); r += 1

# VII. DUYỆT
sec(r, 'VII. DUYỆT / TỪ CHỐI BẢNG CHIA THƯỞNG'); r += 1
tc(r, f'{P}_007.001', 'Hiện nút DUYỆT khi có quyền + trạng thái Chờ duyệt', 'P0', 'User có quyền "BGĐ duyệt" (id 1054), bảng Chờ duyệt', '1. Mở chi tiết bảng', '', 'Hiện nút "DUYỆT" và "KHÔNG DUYỆT"'); r += 1
tc(r, f'{P}_007.002', 'Không hiện nút DUYỆT khi không có quyền', 'P0', 'User KHÔNG có quyền 1054, bảng Chờ duyệt', '1. Mở chi tiết bảng', '', 'KHÔNG hiện nút DUYỆT / KHÔNG DUYỆT. can_approve = false từ API'); r += 1
tc(r, f'{P}_007.003', 'Duyệt thành công', 'P0', 'Có quyền, bảng Chờ duyệt', '1. Nhấn "DUYỆT"', '', 'Toast "Duyệt thành công". Trạng thái → Đã duyệt. Ghi nhận người duyệt + ngày duyệt. Form chuyển readonly'); r += 1
tc(r, f'{P}_007.004', 'Từ chối duyệt thành công', 'P0', 'Có quyền, bảng Chờ duyệt', '1. Nhấn "KHÔNG DUYỆT"', '', 'Toast "Đã từ chối duyệt". Trạng thái → Không duyệt. Người tạo có thể sửa lại'); r += 1
tc(r, f'{P}_007.005', 'API trả 403 khi không có quyền gọi approve', 'P0', 'Gọi API trực tiếp không có quyền', '1. POST /approve không có quyền', '', 'HTTP 403: "Bạn không có quyền duyệt"'); r += 1

# VIII. VALIDATION
sec(r, 'VIII. VALIDATION'); r += 1
tc(r, f'{P}_008.001', 'Thiếu phòng ban khi tạo', 'P0', '', '1. Không chọn PB\n2. Lưu', '', 'Lỗi 422: "Bắt buộc chọn phòng ban"'); r += 1
tc(r, f'{P}_008.002', 'Thiếu năm khi tạo', 'P0', '', '1. Không chọn năm\n2. Lưu', '', 'Lỗi 422: "Bắt buộc chọn năm"'); r += 1
tc(r, f'{P}_008.003', 'Thiếu thành phần', 'P0', '', '1. Không thêm TP\n2. Lưu', '', 'Lỗi 422: "Bắt buộc chọn ít nhất 1 thành phần"'); r += 1
tc(r, f'{P}_008.004', 'Gửi duyệt khi tổng % ≠ 100', 'P0', 'Tổng % = 80%', '1. Nhấn "LƯU & GỬI DUYỆT"', '', 'Lỗi: "Tổng % thành phần phải bằng 100%"'); r += 1
tc(r, f'{P}_008.005', 'Năm ngoài khoảng 2000-2100', 'P1', '', '1. Gọi API với year = 1999', '', 'Lỗi 422 validation'); r += 1

# DATA VALIDATION
dv = DataValidation(type='list', formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
dv.error = 'Chọn 1 trong 4 giá trị'; dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv); dv.add('L8:L500')

output = '/Users/nguyentrancu/DEV/code/HRM/hrm-claude-config/.plans/bonus-distribution/Testcase_Bonus_Distribution.xlsx'
wb.save(output)
print(f'Saved: {output}')
