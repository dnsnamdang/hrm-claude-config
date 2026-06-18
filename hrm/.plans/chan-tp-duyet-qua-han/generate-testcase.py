"""Generate testcase Excel — Chặn trưởng phòng duyệt khi NV có hàng quá hạn (HRM, gọi API ERP)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = "/Users/nguyentrancu/DEV/code/ERP-HRM/HRM/.plans/chan-tp-duyet-qua-han/testcase.xlsx"
SHEET_NAME   = "ChanTPDuyetQuaHan_HRM"
FEATURE_NAME = "Chặn trưởng phòng duyệt khi NV có hàng quá hạn (HRM)"
MODULE_NAME  = "HRM - Chặn duyệt (Due Configs Manager)"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
        "Chặn trưởng phòng (TP) DUYỆT/ TẠO một số phiếu trên HRM khi phòng do TP quản lý còn nhân viên có hàng "
        "QUÁ HẠN (giữ/mượn/nhập thẳng). HRM KHÔNG tự tính quá hạn mà GỌI API sang ERP để kiểm tra.\n"
        "Cơ chế: middleware checkDueConfigsManager:<Tên action> → ErpApiService->get('/api/v1/due-configs/check-manager-block')."),
    ("2. Đối tượng được tính / hiển thị (BỊ CHẶN khi đủ TẤT CẢ)",
        "► Setting use_erp đang BẬT (MasterSetting category='use_erp', content truthy).\n"
        "► User đã đăng nhập (auth()->user() khác null).\n"
        "► Map được NV: HRM user.info.id → TpEmployee.employee_info_id → erp_employeeId.\n"
        "► API ERP /api/v1/due-configs/check-manager-block trả về result.blocked = true.\n"
        "(ERP quyết định blocked dựa trên: action có due_configs tab=2 + company tick + phòng TP có NV quá hạn — "
        "xem testcase ERP riêng)."),
    ("3. Đối tượng bị ẩn / KHÔNG CHẶN (cho qua)",
        "► use_erp TẮT / null → middleware bỏ qua, cho thực hiện ngay (không gọi ERP).\n"
        "► Chưa đăng nhập (auth null) → cho qua.\n"
        "► ERP trả blocked = false → cho thực hiện.\n"
        "► Gọi ERP lỗi / không kết nối được (Exception) → FAIL-OPEN: cho qua + ghi Log::warning.\n"
        "► erp_employeeId = null (NV chưa map ERP) → API ERP nhận employee_id null → ERP trả blocked=false → cho qua."),
    ("4. Bộ lọc thời gian / điều kiện quá hạn",
        "Do ERP xác định (HRM không tính). 3 loại trên ERP: hàng giữ (expire_date<today), "
        "hàng mượn (return_date<today, còn nợ), hàng nhập thẳng (created_at < now − overdue_date_export_direct). "
        "Chi tiết xem testcase ERP."),
    ("5. Cấu trúc dữ liệu / luồng tích hợp",
        "HRM: MasterSetting(use_erp) → auth user → TpEmployee(employee_info_id→id) → ErpApiService(HTTP GET) → "
        "ERP DueConfigController@checkManagerBlock → DueConfigBlockService::isManagerBlocked → trả {blocked, message}."),
    ("6. Quy tắc quyết định",
        "block HRM = use_erp BẬT AND có user AND ERP.blocked = true. "
        "Mọi trường hợp lỗi/thiếu dữ liệu/tắt cờ đều nghiêng về CHO QUA (fail-open) để không chặn nhầm nghiệp vụ HRM."),
    ("7. Phân quyền cấp",
        "Không dùng permission spatie cho việc chặn. Middleware HRM KHÔNG tự bypass Super Admin "
        "(khác middleware ERP) — việc bỏ qua Super Admin do phía ERP (isManagerBlocked) xử lý theo NV ERP tương ứng.\n"
        "Action name truyền vào middleware phải trùng due_configs.name (tab=2) trên ERP."),
    ("8. Cách tính / công thức quyết định",
        "► useErp = MasterSetting where category='use_erp' value('content'); nếu falsy → next().\n"
        "► erp_employeeId = TpEmployee where employee_info_id = user.info.id → value('id').\n"
        "► result = ERP GET check-manager-block(employee_id, action).\n"
        "► if (!empty(result['blocked'])) → response 403 {status:'fail', message}.\n"
        "► catch Exception → Log::warning + next() (fail-open)."),
    ("9. Ghi chú đọc bảng",
        "► Khi bị chặn: HTTP 403 body {status:'fail', message}. FE axios bắt 403 + message chứa 'quá hạn' → "
        "mở modal OverdueEmployeesModal (hrm-client), có nút Xuất Excel.\n"
        "► Danh sách NV quá hạn lấy qua API /api/v1/due-configs/overdue-employees (HRM proxy gọi ERP).\n"
        "► message mặc định khi ERP không trả message: 'Phòng ban bạn quản lý có nhân viên quá hạn. "
        "Không thể thực hiện thao tác này.'\n"
        "► 20 route áp dụng thuộc 2 module: Timesheet (chấm công) và Assign (giao việc)."),
]

SECTIONS = [
    ("I", "ĐIỀU KIỆN GỌI CHECK (use_erp / auth)", [
        ("001", "use_erp TẮT → KHÔNG gọi ERP, cho qua", "P0",
            "MasterSetting category='use_erp' content = null/0/false. NV phòng TP đang quá hạn trên ERP.",
            "1. Đảm bảo use_erp tắt\n2. TP duyệt 1 phiếu (vd Duyệt đơn xin nghỉ)",
            "use_erp = OFF",
            "- Thực hiện thành công, KHÔNG gọi ERP\n- Middleware return next() ngay",
            "if(!useErp) return next()"),
        ("002", "use_erp BẬT → gọi API ERP check", "P0",
            "use_erp BẬT. TP đăng nhập, map được erp_employeeId.",
            "1. TP duyệt 1 phiếu có gắn middleware",
            "use_erp = ON",
            "- Có request HTTP GET tới ERP /api/v1/due-configs/check-manager-block kèm employee_id + action\n"
            "- Kết quả block/allow theo phản hồi ERP",
            "Gọi ErpApiService->get(...)"),
        ("003", "Chưa đăng nhập → cho qua", "P2",
            "use_erp BẬT nhưng auth()->user() = null (token thiếu/hết hạn — tình huống lý thuyết).",
            "1. Gọi route khi chưa auth",
            "auth = null",
            "- Middleware return next() (không chặn ở tầng này)",
            "if(!user) return next()"),
    ]),
    ("II", "MAPPING NHÂN VIÊN & GỌI API ERP", [
        ("001", "Map đúng HRM→ERP employee và truyền action", "P0",
            "use_erp BẬT. HRM user.info.id = 500; TpEmployee có employee_info_id=500 → id(ERP)=88.",
            "1. TP (info.id=500) duyệt 'Duyệt đơn xin nghỉ'\n2. Theo dõi request sang ERP",
            "user.info.id=500 → erp_employeeId=88",
            "- API ERP nhận employee_id=88 và action='Duyệt đơn xin nghỉ'",
            "TpEmployee where employee_info_id=user.info.id → id"),
        ("002", "NV chưa map ERP (erp_employeeId null) → cho qua", "P0",
            "use_erp BẬT. user.info.id không có bản ghi TpEmployee → erp_employeeId = null.",
            "1. TP chưa map duyệt phiếu",
            "erp_employeeId = null",
            "- API ERP nhận employee_id=null → ERP trả blocked=false (thiếu employee_id)\n- HRM cho thực hiện",
            "ERP checkManagerBlock thiếu employee_id → responseErrors/blocked=false"),
        ("003", "ERP trả blocked=true → HRM trả 403", "P0",
            "use_erp BẬT, map OK. ERP phản hồi {blocked:true, message:'Phòng ban bạn quản lý có nhân viên hàng mượn quá hạn...'}.",
            "1. TP duyệt phiếu\n2. Quan sát HTTP status + body",
            "ERP.blocked = true",
            "- HTTP 403 Forbidden\n- Body: {status:'fail', message: <message từ ERP>}",
            "if(!empty(result['blocked'])) → 403"),
        ("004", "ERP trả blocked=false → cho thực hiện", "P0",
            "use_erp BẬT, map OK. ERP phản hồi {blocked:false}.",
            "1. TP duyệt phiếu",
            "ERP.blocked = false",
            "- Thực hiện thành công (next())",
            "blocked falsy → next()"),
        ("005", "ERP lỗi/không kết nối → FAIL-OPEN cho qua + log", "P0",
            "use_erp BẬT nhưng ErpApiService ném Exception (ERP sập / timeout / 500).",
            "1. (giả lập) ERP không phản hồi\n2. TP duyệt phiếu\n3. Kiểm tra log HRM",
            "ERP unreachable",
            "- HRM KHÔNG chặn (thực hiện thành công)\n- Ghi Log::warning('CheckDueConfigsManager: không thể check ERP', ...)",
            "catch Exception → log + next() (fail-open)"),
        ("006", "ERP trả blocked=true nhưng KHÔNG có message → dùng message mặc định", "P1",
            "ERP phản hồi {blocked:true} không kèm message.",
            "1. TP duyệt phiếu\n2. Đọc message",
            "result.message = null",
            "- 403 + message mặc định: 'Phòng ban bạn quản lý có nhân viên quá hạn. Không thể thực hiện thao tác này.'",
            "result['message'] ?? message mặc định"),
    ]),
    ("III", "PHẠM VI PHIẾU HRM — TIMESHEET (chấm công)", [
        ("001", "Duyệt đơn xin nghỉ (storeApprove + nshc-approve)", "P0",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt đơn xin nghỉ (approve)\n2. Thử nshc-approve",
            "action='Duyệt đơn xin nghỉ'",
            "- Cả 2 route bị chặn 403",
            "AttendanceController storeApprove/nshcApprove"),
        ("002", "Duyệt đề nghị tra soát công", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt đề nghị tra soát công",
            "action='Duyệt đề nghị tra soát công'",
            "- Bị chặn 403",
            "RequestUpdateTimeSheetController storeApprove"),
        ("003", "Tạo/Duyệt phiếu giao việc (phân hệ chấm công)", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP tạo phiếu giao việc\n2. TP duyệt phiếu giao việc",
            "action='Tạo/Duyệt phiếu giao việc (phân hệ chấm công)'",
            "- Cả store và approve bị chặn 403",
            "JobAssignmentNoteController store/storeApprove"),
        ("004", "Duyệt đơn đi muộn về sớm", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt đơn đi muộn về sớm",
            "action='Duyệt đơn đi muộn về sớm'",
            "- Bị chặn 403",
            "LateEarlyOutController storeApprove"),
        ("005", "Duyệt phiếu yêu cầu làm thêm / đăng ký làm thêm", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt yêu cầu làm thêm\n2. TP duyệt đăng ký làm thêm",
            "2 action OT",
            "- Cả 2 bị chặn 403",
            "OvertimeRequirement/OvertimeAssignment storeApprove"),
        ("006", "Tạo/Duyệt phiếu đi công tác + kết quả công tác (chấm công)", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP tạo phiếu đi công tác\n2. duyệt\n3. duyệt kết quả công tác (storeResult)",
            "action công tác chấm công",
            "- store, approve, storeResult đều bị chặn 403",
            "BusinessTripAssignController store/storeApprove/storeResult"),
    ]),
    ("IV", "PHẠM VI PHIẾU HRM — ASSIGN (giao việc)", [
        ("001", "Tạo/Update phiếu công tác (giao việc)", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP tạo phiếu công tác (assign-business)\n2. update {id}",
            "action='Tạo phiếu công tác (phân hệ giao việc)' / 'Duyệt kết quả công tác trên HRM'",
            "- store và update bị chặn 403",
            "AssignBusinessController store/update"),
        ("002", "Duyệt kết quả công tác (assign-job approveResult)", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt kết quả công tác",
            "action='Duyệt kết quả công tác trên HRM'",
            "- Bị chặn 403",
            "AssignJobController approveResult"),
        ("003", "Duyệt quyết toán công theo HĐ", "P0",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt quyết toán công theo HĐ",
            "action='Duyệt quyết toán công theo HĐ'",
            "- Bị chặn 403",
            "SettlementContractController approve"),
        ("004", "Duyệt đề nghị thanh toán HRM", "P0",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt đề nghị thanh toán",
            "action='Duyệt đề nghị thanh toán HRM'",
            "- Bị chặn 403",
            "PaymentBusinessRequestController approve"),
        ("005", "Duyệt hồ sơ thanh toán", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt hồ sơ thanh toán",
            "action='Duyệt hồ sơ thanh toán'",
            "- Bị chặn 403",
            "PaymentProfileController approve"),
        ("006", "Chặn gia hạn & kết thúc phiếu công tác (giao việc)", "P1",
            "use_erp BẬT, ERP.blocked=true.",
            "1. TP duyệt gia hạn/kết thúc sớm phiếu công tác",
            "action='Chặn gia hạn và kết thúc phiếu công tác (phân hệ giao việc)'",
            "- Bị chặn 403",
            "ExtendEndSoonRequestController approve"),
    ]),
    ("V", "FE — MODAL NV QUÁ HẠN (hrm-client)", [
        ("001", "Nhận 403 + message 'quá hạn' → mở modal", "P0",
            "TP duyệt phiếu, backend trả 403 {status:'fail', message chứa 'quá hạn'}.",
            "1. TP bấm Duyệt trên UI\n2. Quan sát",
            "403 + message có 'quá hạn'",
            "- axios bắt lỗi → set state showOverdueModal=true → mở OverdueEmployeesModal\n- Không điều hướng/đăng xuất",
            "plugins/axios.js: 403 + 'quá hạn' → modal"),
        ("002", "Modal load danh sách NV quá hạn", "P1",
            "Modal NV quá hạn vừa mở.",
            "1. Quan sát danh sách trong modal",
            "API overdue-employees (HRM→ERP)",
            "- Hiển thị mã NV, tên NV, loại quá hạn (Hàng giữ/Hàng mượn/Hàng NXT)",
            "GET /api/v1/due-configs/overdue-employees"),
        ("003", "Nút Xuất Excel trong modal", "P2",
            "Modal đang hiển thị danh sách NV quá hạn.",
            "1. Bấm 'Xuất Excel'",
            "export=1",
            "- Tải file danh sách NV quá hạn (.xlsx)",
            "OverdueEmployeesExcel (ERP) qua proxy HRM"),
        ("004", "403 KHÔNG chứa 'quá hạn' → không mở modal quá hạn", "P2",
            "Backend trả 403 vì lý do khác (không phải due config).",
            "1. Trigger 403 khác\n2. Quan sát",
            "403 message thường",
            "- Không mở OverdueEmployeesModal (chỉ mở khi message chứa 'quá hạn')",
            "Điều kiện FE lọc theo nội dung message"),
    ]),
    ("VI", "EDGE CASES & TÍCH HỢP", [
        ("001", "Action name HRM không khớp due_configs ERP → cho qua", "P0",
            "use_erp BẬT, action truyền vào không trùng due_configs.name tab=2 trên ERP.",
            "1. TP duyệt phiếu có action lệch tên",
            "actionName lệch",
            "- ERP trả blocked=false → HRM cho thực hiện",
            "Khớp tên action ERP; lệch → không chặn"),
        ("002", "Hết quá hạn bên ERP → HRM hết chặn ngay lần sau", "P1",
            "Đang bị chặn. NV trả hết hàng/gia hạn trên ERP.",
            "1. Xử lý quá hạn trên ERP\n2. TP duyệt lại trên HRM",
            "ERP.blocked chuyển false",
            "- Lần duyệt kế tiếp thành công (HRM gọi ERP realtime mỗi request)",
            "Mỗi request gọi ERP mới, không cache"),
        ("003", "Đồng nhất quyết định với ERP", "P2",
            "Cùng TP, cùng thời điểm: thao tác phiếu tương ứng trên ERP cũng bị chặn.",
            "1. So sánh kết quả chặn giữa ERP và HRM cho cùng action+TP",
            "Cùng employee + action",
            "- Kết quả block/allow của HRM TRÙNG với ERP (vì dùng chung isManagerBlocked)",
            "HRM ủy quyền quyết định cho ERP"),
        ("004", "use_erp BẬT, ERP trả định dạng bất ngờ (không có key blocked)", "P2",
            "ERP trả body thiếu key 'blocked'.",
            "1. TP duyệt phiếu",
            "result không có 'blocked'",
            "- empty(result['blocked']) = true → coi như không chặn → cho qua",
            "!empty(result['blocked']) là điều kiện chặn"),
    ]),
]

# ============================ BUILD ============================
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri",size=11,bold=True); DESC_LABEL_FILL = PatternFill("solid",fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri",size=11)
WRAP_TL = Alignment(wrap_text=True,vertical="top",horizontal="left")
WRAP_TC = Alignment(wrap_text=True,vertical="top",horizontal="center")
TITLE_FONT = Font(name="Calibri",size=14,bold=True,color="FFFFFF"); TITLE_FILL = PatternFill("solid",fgColor="4472C4")
SUM_LABEL_FILL = PatternFill("solid",fgColor="D9E1F2")
HEADER_FONT = Font(name="Calibri",size=11,bold=True,color="FFFFFF"); HEADER_FILL = PatternFill("solid",fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True,vertical="center",horizontal="center")
SECTION_FONT = Font(name="Calibri",size=12,bold=True,color="1F4E79"); SECTION_FILL = PatternFill("solid",fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True,vertical="center",horizontal="left",indent=1)
EVEN_FILL = PatternFill("solid",fgColor="F2F2F2")
COL_WIDTHS = {'A':24,'B':24,'C':14,'D':42,'E':9,'F':40,'G':44,'H':24,'I':60,'J':36,'K':16,'L':14,'M':14,'N':14,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col,w in COL_WIDTHS.items(): ws.column_dimensions[col].width = w
ws.cell(1,1,"MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True,size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
for idx,(label,body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx,1,label); a.font=DESC_LABEL_FONT; a.fill=DESC_LABEL_FILL; a.alignment=WRAP_TL; a.border=BORDER
    b = ws.cell(idx,2,body); b.font=DESC_BODY_FONT; b.alignment=WRAP_TL; b.border=BORDER
    ws.merge_cells(start_row=idx,start_column=2,end_row=idx,end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n")*16 + 30)
t = ws.cell(11,1,f"Testcase _ {FEATURE_NAME}"); t.font=TITLE_FONT; t.fill=TITLE_FILL
t.alignment = Alignment(vertical="center",horizontal="left",indent=1)
ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11,6,"TEST SUMMARY"); fs.font=Font(name="Calibri",size=12,bold=True,color="FFFFFF")
fs.fill=TITLE_FILL; fs.alignment=Alignment(vertical="center",horizontal="center")
ws.row_dimensions[11].height = 28
for r,label,formula in [
    (11,"Số trường hợp kiểm thử đạt (P):",'=COUNTIF(L18:N500,"Passed")'),
    (12,"Số trường hợp kiểm thử không đạt (F):",'=COUNTIF(L18:N500,"Failed")'),
    (13,"Số trường hợp kiểm thử đang xem xét:",'=COUNTIF(L18:N500,"Pending")'),
    (14,"Số trường hợp kiểm thử chưa thực hiện:",'=COUNTIF(L18:N500,"Not Executed")'),
    (15,"Tổng số trường hợp kiểm thử:",'=COUNTIF(L18:N500,"<>")')]:
    lc = ws.cell(r,9,label); lc.font=Font(name="Calibri",size=11,bold=True); lc.fill=SUM_LABEL_FILL
    lc.alignment=Alignment(vertical="center",horizontal="right"); lc.border=BORDER
    ws.merge_cells(start_row=r,start_column=9,end_row=r,end_column=11)
    vc = ws.cell(r,12,formula); vc.font=Font(name="Calibri",size=11,bold=True); vc.fill=SUM_LABEL_FILL
    vc.alignment=Alignment(horizontal="center",vertical="center"); vc.border=BORDER
    ws.merge_cells(start_row=r,start_column=12,end_row=r,end_column=15)
    if r>11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8
HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện",
    "Test Data","Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
    "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i,h in enumerate(HEADERS, start=1):
    c = ws.cell(17,i,h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=HEADER_ALIGN; c.border=BORDER
ws.row_dimensions[17].height = 36
current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row,3,title); cell.font=SECTION_FONT; cell.fill=SECTION_FILL
    cell.alignment=SECTION_ALIGN; cell.border=BORDER
    ws.merge_cells(start_row=current_row,start_column=3,end_row=current_row,end_column=15)
    for col in (1,2):
        ws.cell(current_row,col).fill=SECTION_FILL; ws.cell(current_row,col).border=BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1
def write_tc(tc_id,function,priority,pre,steps,td,expected,note,group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME,group,tc_id,function,priority,pre,steps,td,expected,note,"",
        "Not Executed","Not Executed","Not Executed",""]
    fill = EVEN_FILL if data_row_idx % 2 == 1 else None
    for i,v in enumerate(values, start=1):
        c = ws.cell(current_row,i,v); c.font=Font(name="Calibri",size=11)
        c.alignment = WRAP_TC if i==5 else WRAP_TL; c.border=BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(40, min(220, longest//3))
    current_row += 1; data_row_idx += 1
roman_list = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman,title,tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = roman_list.index(roman) + 1
    for tc_num,func,prio,pre,steps,td,exp,note in tcs:
        write_tc(f"TC_{sec_idx:02d}.{int(tc_num):03d}",func,prio,pre,steps,td,exp,note,group=title)
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row+100}"); ws.add_data_validation(dv)
wb.save(OUTPUT_FILE)
print(f"OK: {OUTPUT_FILE} | data rows 18-{current_row-1} | TCs={data_row_idx}")
