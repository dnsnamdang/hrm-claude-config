"""Generate testcase Excel cho feature Danh mục Khách hàng (module Giao việc / Assign)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE  = ".plans/customer-catalog/testcase.xlsx"
SHEET_NAME   = "DanhMucKhachHang"
FEATURE_NAME = "Danh mục Khách hàng (Giao việc)"
MODULE_NAME  = "Khách hàng"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý danh mục Khách hàng trong module Giao việc (Assign): xem danh sách, lọc/tìm kiếm, "
     "tạo mới và chỉnh sửa khách hàng theo form đầy đủ. Khi Lưu, dữ liệu được ghi đồng thời vào "
     "DB ERP TanPhat (connection mysql2, bảng customers) và đồng bộ sang DB HRM. "
     "Danh sách dùng làm nguồn KH cho Dự án tiền khả thi, Báo giá, Họp, giải pháp..."),
    ("2. Đối tượng được tính / hiển thị",
     "► Chỉ KH có is_customer = 1 trên ERP.\n"
     "► Phạm vi hiển thị theo phân quyền cấp của user (xem mục 7):\n"
     "   - 'Xem tất cả khách hàng' → thấy mọi KH.\n"
     "   - 'Xem tất cả khách hàng của công ty' → KH thuộc company_id của user.\n"
     "   - 'Xem tất cả khách hàng của phòng ban' → KH thuộc department_id.\n"
     "   - 'Xem tất cả khách hàng của bộ phận' → KH thuộc part_id.\n"
     "   - Không có cấp nào → chỉ KH do chính user tạo (created_by = erp employee id).\n"
     "► Ngoại lệ: KH do mình tạo luôn thấy; tìm khớp ĐÚNG số điện thoại (phone_exact_bypass) bỏ qua giới hạn cấp."),
    ("3. Đối tượng bị ẩn / không tính",
     "► KH có is_customer != 1 (nhà cung cấp thuần, đối tượng khác trên ERP).\n"
     "► KH nằm ngoài phạm vi quyền cấp của user (vd: KH công ty khác khi user chỉ có quyền cấp công ty).\n"
     "► KH chưa có báo giá gắn company/department/part → chỉ người tạo mới thấy (cấp dựa trên báo giá liên kết)."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "Không áp dụng — màn này KHÔNG có bộ lọc theo khoảng thời gian/ngày."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Nguồn chính: bảng ERP customers (connection mysql2). HRM customers chỉ là bản đồng bộ rút gọn, "
     "liên kết bằng cột code. Ghi ERP trước (transaction) → best-effort sync HRM (lỗi sync KHÔNG làm fail Lưu).\n"
     "► id trên route /{id} chính là ERP customers.id.\n"
     "► 5 loại KH (customer_type): 1=Cá nhân; 2=Doanh nghiệp tư nhân; 3=Doanh nghiệp nước ngoài; "
     "4=Tổ chức phi chính phủ; 5=Cơ quan nhà nước. Loại 2-5 = 'tổ chức' (form hiển thị section khác cá nhân).\n"
     "► Bảng phụ ERP: customer_deputies (người đại diện), customer_contacts (người liên hệ), "
     "customer_has_groups/customer_customer_scopes (nhóm + lĩnh vực), customer_has_vehicle_manufacts (hãng xe), "
     "bank_accounts, delivery_places (địa chỉ giao hàng)."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Không áp dụng cộng dồn (đây là danh sách thuần, không thống kê). "
     "Khi Sửa, các bảng phụ (deputies/contacts/scopes/bank accounts) được sync theo kiểu delete-then-insert."),
    ("7. Phân quyền cấp",
     "Quyền ĐỌC TỪ ERP (không nằm trong PermissionsTableSeeder HRM), định nghĩa tại ErpPermissionHelper:\n"
     "• Xem khách hàng — vào màn danh sách (GET /assign/customers).\n"
     "• Thêm khách hàng — nút 'Tạo mới' + POST /assign/customers.\n"
     "• Sửa khách hàng — nút 'Sửa' + POST /assign/customers/{id}.\n"
     "• Xóa khách hàng — định nghĩa nhưng KHÔNG có route dùng (không có chức năng xóa).\n"
     "Phân quyền XEM theo cấp tổ chức:\n"
     "• Xem tất cả khách hàng (is_all_company)\n"
     "• Xem tất cả khách hàng của công ty (is_company)\n"
     "• Xem tất cả khách hàng của phòng ban (is_department)\n"
     "• Xem tất cả khách hàng của bộ phận (is_part)\n"
     "Account HRM phải map được sang ERP employee; nếu không map → coi như KHÔNG có quyền nào."),
    ("8. Cách tính các ô thống kê",
     "Không áp dụng — màn này KHÔNG có ô thống kê / stats đầu trang."),
    ("9. Ghi chú đọc bảng",
     "► Cột 'SĐT' và 'Nhóm KH' ở danh sách luôn hiển thị '-' / '—' do resource hard-code "
     "(CustomerListResource) — KHÔNG phản ánh dữ liệu thật, đừng test giá trị thật ở 2 cột này.\n"
     "► Cột Trạng thái: pill 'Hoạt động' (xanh) khi status=1, 'Khóa' khi khác 1. Tạo mới luôn set status=1; "
     "màn này KHÔNG có UI khóa/mở khóa.\n"
     "► Phân trang server-side, mặc định 10 dòng/trang; sắp xếp mặc định id giảm dần.\n"
     "► Mã KH không nhập tay, tự sinh khi Lưu (theo biển số tỉnh + viết tắt tỉnh/phường, trùng thêm -01,-02). Sửa KHÔNG đổi mã.\n"
     "► Toàn bộ validation hiển thị inline đến từ BE (HTTP 422); FE không chặn submit trước."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("00", "Truy cập mặc định khi không có quyền cấp nào", "P0",
     "User đã đăng nhập, account ĐÃ map ERP employee; KHÔNG có bất kỳ quyền 'Xem tất cả...' cấp nào; "
     "ERP có 5 KH do user tạo + 20 KH do người khác tạo",
     "1. Truy cập /assign/customers\n2. Quan sát danh sách",
     "User: chỉ có quyền 'Xem khách hàng'",
     "- Chỉ hiển thị đúng 5 KH do chính user tạo (created_by = erp employee id)\n"
     "- KHÔNG thấy 20 KH của người khác",
     "BR — Không có cấp nào → chỉ thấy KH do mình tạo"),
    ("01", "Truy cập với quyền 'Xem khách hàng'", "P0",
     "User có quyền 'Xem khách hàng'; account map ERP employee",
     "1. Truy cập /assign/customers\n2. Quan sát màn hình load",
     "User: quyền 'Xem khách hàng'",
     "- Vào được màn danh sách, bảng + bộ lọc hiển thị bình thường",
     "Permission: Xem khách hàng (erpPermission GET /assign/customers)"),
    ("02", "Không có quyền 'Xem khách hàng'", "P0",
     "User KHÔNG có quyền 'Xem khách hàng'",
     "1. Truy cập trực tiếp URL /assign/customers",
     "User: không có quyền xem",
     "- Bị chặn truy cập (403/redirect), KHÔNG thấy danh sách KH",
     "Permission: middleware erpPermission:Xem khách hàng"),
    ("03", "Quyền 'Xem tất cả khách hàng' (toàn hệ thống)", "P0",
     "User có quyền 'Xem tất cả khách hàng'; ERP có KH thuộc nhiều công ty khác nhau",
     "1. Truy cập danh sách\n2. Quan sát phạm vi dữ liệu",
     "User: quyền tất cả KH",
     "- Hiển thị TOÀN BỘ KH mọi công ty, không giới hạn cấp",
     "Permission: Xem tất cả khách hàng (is_all_company)"),
    ("04", "Quyền 'Xem tất cả khách hàng của công ty'", "P0",
     "User thuộc Công ty A; có quyền 'Xem tất cả khách hàng của công ty'; ERP có KH cty A và cty B",
     "1. Truy cập danh sách\n2. Quan sát phạm vi dữ liệu",
     "User: quyền cấp công ty, company A",
     "- Chỉ hiển thị KH gắn company_id = A\n- KHÔNG thấy KH công ty B",
     "Permission: Xem tất cả khách hàng của công ty (is_company)"),
    ("05", "Quyền 'Xem tất cả khách hàng của phòng ban'", "P1",
     "User thuộc Phòng ban X; có quyền cấp phòng ban; ERP có KH phòng X và phòng Y",
     "1. Truy cập danh sách\n2. Quan sát phạm vi",
     "User: quyền cấp phòng ban, department X",
     "- Chỉ hiển thị KH gắn department_id = X",
     "Permission: Xem tất cả khách hàng của phòng ban (is_department)"),
    ("06", "Quyền 'Xem tất cả khách hàng của bộ phận'", "P1",
     "User thuộc Bộ phận P; có quyền cấp bộ phận; ERP có KH bộ phận P và bộ phận Q",
     "1. Truy cập danh sách\n2. Quan sát phạm vi",
     "User: quyền cấp bộ phận, part P",
     "- Chỉ hiển thị KH gắn part_id = P",
     "Permission: Xem tất cả khách hàng của bộ phận (is_part)"),
    ("07", "Quyền 'Thêm khách hàng' — nút Tạo mới hiển thị", "P0",
     "User có quyền 'Thêm khách hàng'",
     "1. Truy cập danh sách\n2. Quan sát góc phải trên",
     "User: quyền thêm",
     "- Nút 'Tạo mới' hiển thị\n- Bấm vào điều hướng /assign/customers/add",
     "Permission: Thêm khách hàng (canCreate)"),
    ("08", "Không có quyền 'Thêm khách hàng' — ẩn nút Tạo mới", "P0",
     "User KHÔNG có quyền 'Thêm khách hàng'",
     "1. Truy cập danh sách\n2. Quan sát\n3. Thử POST /assign/customers trực tiếp",
     "User: không có quyền thêm",
     "- Nút 'Tạo mới' KHÔNG hiển thị\n- Gọi API store bị chặn (403)",
     "Permission: middleware erpPermission:Thêm khách hàng"),
    ("09", "Quyền 'Sửa khách hàng' — nút Sửa hiển thị", "P0",
     "User có quyền 'Sửa khách hàng'; danh sách có ≥1 KH",
     "1. Truy cập danh sách\n2. Quan sát action trên từng dòng",
     "User: quyền sửa",
     "- Nút 'Sửa' hiển thị trên mỗi dòng (cùng nút 'Xem')\n- Bấm Sửa mở form chỉnh sửa",
     "Permission: Sửa khách hàng (canEdit)"),
    ("10", "Không có quyền 'Sửa khách hàng' — chỉ Xem", "P1",
     "User KHÔNG có quyền 'Sửa khách hàng' nhưng có 'Xem khách hàng'",
     "1. Truy cập danh sách\n2. Quan sát action từng dòng",
     "User: chỉ quyền xem",
     "- Chỉ có nút 'Xem' trên mỗi dòng, KHÔNG có nút 'Sửa'\n- POST /assign/customers/{id} bị chặn (403)",
     "Permission: middleware erpPermission:Sửa khách hàng"),
    ("11", "Account HRM chưa map được ERP employee", "P0",
     "User đã đăng nhập HRM nhưng account KHÔNG gắn employee_info_id hoặc không map được sang ERP employee",
     "1. Truy cập danh sách\n2. Thử tạo mới KH",
     "User: chưa map ERP",
     "- Coi như không có quyền nào: danh sách trống / không tạo được\n"
     "- Khi Lưu tạo/sửa → ValidationException báo lỗi gắn nhân sự",
     "BR — Tạo/Sửa yêu cầu map được ERP employee (CustomerService::save)"),
    ("12", "Tìm khớp đúng SĐT bỏ qua giới hạn cấp (phone_exact_bypass)", "P2",
     "User chỉ có quyền cấp công ty A; tồn tại KH cty B có SĐT = 0912345678",
     "1. Nhập đúng SĐT 0912345678 vào ô tìm kiếm SĐT\n2. Tìm",
     "SĐT tìm: 0912345678 (khớp tuyệt đối)",
     "- KH cty B xuất hiện dù ngoài phạm vi quyền cấp của user",
     "BR — phone_exact_bypass: khớp đúng SĐT bỏ qua giới hạn quyền"),
]

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Load trang danh sách thành công", "P0",
         "User có quyền 'Xem khách hàng'; ERP có ≥1 KH trong phạm vi quyền",
         "1. Truy cập /assign/customers\n2. Quan sát layout",
         "User bất kỳ có quyền xem",
         "- Trang load với 2 khối: panel Bộ lọc (mặc định thu gọn) + bảng danh sách\n"
         "- KHÔNG có khối stats/thống kê đầu trang",
         "BR — Layout: V2BaseFilterPanel + V2BaseDataTable, không có stats"),
        ("002", "Hiển thị đủ các cột bảng danh sách", "P0",
         "Danh sách có ≥1 KH",
         "1. Truy cập danh sách\n2. Quan sát tiêu đề cột",
         "—",
         "- Cột theo thứ tự: STT | Mã KH - Tên khách hàng | Loại | MST | SĐT | Email | "
         "Nhóm KH | Địa chỉ | Tỉnh/TP | Trạng thái\n"
         "- Cell đầu có 2 nút inline: Xem, Sửa (Sửa chỉ khi có quyền)",
         "—"),
        ("003", "Bộ lọc mặc định thu gọn", "P2",
         "Truy cập lần đầu",
         "1. Vào màn hình\n2. Quan sát panel bộ lọc",
         "—",
         "- Panel bộ lọc nâng cao ở trạng thái thu gọn (filterCollapsed=true)\n"
         "- Vẫn hiển thị ô tìm kiếm nhanh (keyword)",
         "—"),
        ("004", "Hiển thị cột SĐT và Nhóm KH là gạch ngang", "P1",
         "KH trong danh sách thực tế CÓ số điện thoại và nhóm KH",
         "1. Xem 1 dòng KH đã có SĐT/nhóm\n2. Quan sát cột SĐT, Nhóm KH",
         "KH có mobile=0901xxx, nhóm='KH lẻ'",
         "- Cột 'SĐT' hiển thị '-' và 'Nhóm KH' hiển thị '—' (hard-code ở resource)\n"
         "- Đây là hành vi đúng theo hiện trạng, không phải bug",
         "Lưu ý: CustomerListResource trả mobile='-', groupNames='—' cứng"),
        ("005", "Pill trạng thái Hoạt động / Khóa", "P1",
         "ERP có KH status=1 và KH status≠1",
         "1. Quan sát cột Trạng thái",
         "KH A status=1; KH B status=0",
         "- KH A: pill 'Hoạt động' (xanh)\n- KH B: pill 'Khóa'",
         "BR — status=1 → Hoạt động, khác → Khóa"),
        ("006", "Danh sách rỗng trong phạm vi quyền", "P1",
         "User có quyền cấp công ty nhưng công ty chưa có KH nào",
         "1. Truy cập danh sách",
         "User cty chưa có KH",
         "- Bảng hiển thị trạng thái rỗng (không có dòng dữ liệu), không lỗi",
         "—"),
    ]),
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm nhanh theo mã KH", "P0",
         "ERP có KH mã '29HNHK' trong phạm vi quyền",
         "1. Nhập '29HNHK' vào ô tìm nhanh\n2. Nhấn Enter (hoặc bấm Tìm)",
         "keyword = 29HNHK",
         "- Danh sách lọc còn KH có mã chứa '29HNHK'",
         "BR — keyword tìm theo mã / tên / tên viết tắt"),
        ("002", "Tìm nhanh theo tên khách hàng", "P0",
         "ERP có KH tên 'CÔNG TY ABC'",
         "1. Nhập 'ABC' vào ô tìm nhanh\n2. Nhấn Enter",
         "keyword = ABC",
         "- Danh sách lọc còn KH có tên chứa 'ABC'",
         "BR — keyword tìm theo tên"),
        ("003", "Ô text chỉ search khi Enter/bấm Tìm", "P1",
         "Danh sách có nhiều KH",
         "1. Gõ ký tự vào ô keyword/Mã KH/Tên KH KHÔNG nhấn Enter\n2. Quan sát",
         "Gõ 'AB' rồi dừng",
         "- Danh sách CHƯA lọc (chưa gọi API) cho tới khi nhấn Enter hoặc bấm Tìm",
         "BR — ignoredFields: keyword, tax_code, name, gara_name, code, identity_card_number chỉ search khi submit"),
        ("004", "Select/dropdown search ngay khi đổi", "P1",
         "Panel bộ lọc nâng cao mở; danh sách nhiều KH",
         "1. Chọn 1 giá trị ở dropdown (vd Trạng thái, Tỉnh/TP)\n2. Quan sát",
         "Trạng thái = Hoạt động",
         "- Danh sách tự động lọc ngay khi đổi dropdown (deep watch), không cần bấm Tìm",
         "BR — filter select auto-search (deep watcher)"),
        ("005", "Lọc theo Công ty / Phòng ban / Nhân viên", "P1",
         "ERP có KH tạo bởi nhiều nhân viên thuộc nhiều công ty/phòng ban",
         "1. Mở bộ lọc nâng cao\n2. Chọn Công ty → Phòng ban → Nhân viên",
         "Công ty A, Phòng ban X",
         "- Danh sách lọc theo công ty/phòng ban/nhân viên đã chọn (part bị disable)",
         "BR — V2BaseCompanyDepartmentFilter, disable_part=true"),
        ("006", "Lọc theo Tỉnh/TP", "P1",
         "ERP có KH ở Hà Nội và TP.HCM",
         "1. Chọn Tỉnh/TP = Hà Nội",
         "province = Hà Nội",
         "- Danh sách còn KH thuộc Hà Nội",
         "—"),
        ("007", "Lọc theo Trạng thái", "P1",
         "ERP có KH Hoạt động và Khóa",
         "1. Chọn Trạng thái = Khóa",
         "status = 0",
         "- Danh sách còn KH có status ≠ 1",
         "—"),
        ("008", "Lọc theo MST / SĐT (ô gộp)", "P1",
         "ERP có KH MST '0101234567'",
         "1. Nhập '0101234567' vào ô MST/SĐT\n2. Nhấn Enter",
         "tax/phone = 0101234567",
         "- Danh sách lọc theo MST hoặc SĐT khớp",
         "—"),
        ("009", "Lọc Nhóm lĩnh vực KH (multi-select) + Lĩnh vực phụ thuộc", "P1",
         "ERP có nhóm lĩnh vực và lĩnh vực con",
         "1. Chọn ≥1 Nhóm lĩnh vực KH\n2. Quan sát dropdown Lĩnh vực KH",
         "Nhóm = 'Ô tô'",
         "- Dropdown Lĩnh vực KH chỉ còn lĩnh vực thuộc nhóm đã chọn (top-down)\n- Danh sách lọc tương ứng",
         "BR — Lĩnh vực lọc theo Nhóm đã chọn"),
        ("010", "Lọc 'Khách hàng hãng' + Hãng xe + Cấp đại lý", "P2",
         "ERP có KH is_manufacturer=1 gắn hãng xe",
         "1. Bật lọc 'Khách hàng hãng'\n2. Chọn Hãng xe, Cấp đại lý",
         "is_manufacturer=1, hãng='Toyota'",
         "- Danh sách còn KH là khách hãng khớp hãng xe/cấp đại lý",
         "—"),
        ("011", "Lọc theo Người sửa gần nhất (editor)", "P2",
         "ERP có KH sửa bởi nhiều người",
         "1. Chọn Người sửa gần nhất = NV B",
         "editor_id = NV B",
         "- Danh sách còn KH có người sửa gần nhất là NV B",
         "—"),
        ("012", "Kết hợp nhiều filter", "P1",
         "ERP có dữ liệu đa dạng",
         "1. Chọn Tỉnh = Hà Nội + Trạng thái = Hoạt động + keyword 'CÔNG TY'\n2. Tìm",
         "province=HN, status=1, keyword=CÔNG TY",
         "- Danh sách thỏa ĐỒNG THỜI cả 3 điều kiện",
         "—"),
        ("013", "Xóa/Reset bộ lọc", "P2",
         "Đang áp dụng vài filter",
         "1. Bấm nút xóa lọc / reset\n2. Quan sát",
         "—",
         "- Các filter về mặc định, danh sách trở lại toàn bộ KH trong phạm vi quyền",
         "—"),
        ("014", "Tìm kiếm không có kết quả", "P1",
         "Danh sách có KH",
         "1. Nhập keyword không tồn tại 'ZZZZZ999'\n2. Tìm",
         "keyword = ZZZZZ999",
         "- Bảng hiển thị rỗng, không lỗi",
         "—"),
    ]),
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "—",
         "Màn Danh mục Khách hàng không có khối thống kê đầu trang",
         "—",
         "- Không có ô stats nào để kiểm thử",
         "Không áp dụng"),
    ]),
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Phân trang chuyển trang", "P0",
         "ERP có > 10 KH trong phạm vi quyền (vd 25 KH)",
         "1. Quan sát trang 1 (10 dòng)\n2. Bấm sang trang 2\n3. Bấm trang 3",
         "Tổng 25 KH, per_page=10",
         "- Trang 1: 10 dòng, trang 2: 10 dòng, trang 3: 5 dòng\n- meta total=25, last_page=3",
         "BR — phân trang server-side, mặc định 10/trang"),
        ("002", "Đổi số dòng mỗi trang", "P1",
         "ERP có > 20 KH",
         "1. Đổi per_page (vd 20)\n2. Quan sát",
         "per_page = 20",
         "- Bảng hiển thị tối đa 20 dòng/trang, phân trang tính lại",
         "—"),
        ("003", "Sắp xếp mặc định id giảm dần", "P1",
         "ERP có nhiều KH",
         "1. Truy cập danh sách lần đầu",
         "sort_by=id, sort_desc=true",
         "- KH mới tạo (id lớn) hiển thị trên cùng",
         "BR — mặc định sort_by=id desc"),
        ("004", "Sắp xếp theo cột", "P2",
         "Danh sách có nhiều KH",
         "1. Bấm tiêu đề cột cho phép sort\n2. Quan sát",
         "—",
         "- Dữ liệu sắp xếp theo cột đã chọn (asc/desc), gửi sort_by/sort_desc lên server",
         "—"),
        ("005", "Cột địa chỉ ghép đúng định dạng", "P2",
         "KH có apartment_number, ward_name, province_name đầy đủ",
         "1. Quan sát cột Địa chỉ",
         "Số nhà=12, phường='Dịch Vọng', tỉnh='Hà Nội'",
         "- Hiển thị '12, Dịch Vọng, Hà Nội' (ghép theo thứ tự số nhà, phường, tỉnh)",
         "—"),
        ("006", "Cột Loại map đúng tên", "P1",
         "ERP có KH các customer_type 1-5",
         "1. Quan sát cột Loại",
         "KH type=1,2,3,4,5",
         "- Hiển thị đúng: Cá nhân / DN tư nhân / DN nước ngoài / Tổ chức phi chính phủ / Cơ quan nhà nước",
         "BR — renderCustomerType map 1-5"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH (TẠO / SỬA / XEM)", [
        ("001", "Tạo KH Cá nhân hợp lệ", "P0",
         "User có quyền 'Thêm khách hàng', account map ERP employee",
         "1. Bấm 'Tạo mới'\n2. Loại = Cá nhân\n3. Nhập Tên KH, Loại hình tổ chức, Nhóm + Lĩnh vực KH, "
         "Tên đơn vị, ≥1 SĐT đúng định dạng, địa chỉ (Quốc gia/Tỉnh/Phường)\n4. Bấm Lưu",
         "Tên='Nguyễn Văn A'; SĐT='0901234567'; Tỉnh='Hà Nội', Phường='Dịch Vọng'",
         "- Lưu thành công, ghi ERP customers (is_customer=1, status=1) + sync HRM\n"
         "- Mã KH tự sinh theo biển số tỉnh + viết tắt (vd 29...)\n- Quay về danh sách, KH mới xuất hiện",
         "BR — Cá nhân cần ≥1 SĐT; mã tự sinh; status=1"),
        ("002", "Tạo KH Tổ chức (có MST) hợp lệ", "P0",
         "User có quyền thêm, map ERP employee",
         "1. Tạo mới, Loại = Doanh nghiệp tư nhân\n2. Nhập Tên KH, Loại hình, Nhóm+Lĩnh vực, "
         "MST hợp lệ, Địa chỉ xuất hóa đơn, ≥1 Người đại diện, ≥1 Người liên hệ (có ≥1 SĐT)\n3. Lưu",
         "Tên='CÔNG TY ABC'; MST='0101234567'; đại diện='Trần B' (GĐ); liên hệ='Lê C' (Kế toán, 0912345678)",
         "- Lưu thành công, ghi ERP + sync HRM\n- Tên KH được uppercase\n- Mã tự sinh, quay về danh sách",
         "BR — Tổ chức cần đại diện ≥1, liên hệ ≥1, địa chỉ xuất HĐ, MST khi không có công ty mẹ"),
        ("003", "Tạo KH Tổ chức có Công ty mẹ (MST không bắt buộc)", "P1",
         "User có quyền thêm; tồn tại 1 KH tổ chức làm công ty mẹ",
         "1. Tạo mới Loại tổ chức\n2. Chọn Công ty mẹ\n3. Bỏ trống MST\n4. Điền các trường bắt buộc còn lại\n5. Lưu",
         "Công ty mẹ='TẬP ĐOÀN X'; MST=trống",
         "- Lưu thành công, KHÔNG báo lỗi MST",
         "BR — có parent_id → tax_code nullable"),
        ("004", "Form đổi section theo Loại KH", "P1",
         "Đang ở form tạo",
         "1. Chọn Loại = Cá nhân → quan sát\n2. Đổi Loại = Doanh nghiệp tư nhân → quan sát",
         "—",
         "- Cá nhân: hiện section Thông tin cá nhân (CCCD, SĐT, ...), KHÔNG có người đại diện/liên hệ\n"
         "- Tổ chức: hiện Thông tin tổ chức + Người liên hệ (MST, đại diện, liên hệ)",
         "BR — isOrganization (type 2-5) đổi layout form"),
        ("005", "Tick 'Là khách hãng' bắt buộc chọn Hãng xe", "P1",
         "Đang ở form tạo",
         "1. Tick 'Là khách hãng'\n2. Bỏ trống Hãng xe\n3. Điền trường khác hợp lệ\n4. Lưu",
         "is_manufacturer=1, Hãng xe=trống",
         "- BE trả 422, lỗi inline tại Hãng xe (bắt buộc khi là khách hãng)",
         "BR — vehicle_manufacts required_if is_manufacturer=1"),
        ("006", "Nhóm và Lĩnh vực KH đồng bộ 2 chiều", "P2",
         "Form tạo, Nhóm có nhiều lĩnh vực con",
         "1. Chọn Nhóm lĩnh vực → quan sát Lĩnh vực\n2. Chọn lĩnh vực thuộc nhóm khác",
         "Nhóm='Ô tô'",
         "- Lĩnh vực chỉ cho chọn trong nhóm đã chọn; chọn lĩnh vực khác nhóm thì nhóm cập nhật theo (watch 2 chiều)",
         "BR — customer_scope_ids phải thuộc nhóm đã chọn (validate BE)"),
        ("007", "Sửa KH — load đúng dữ liệu", "P0",
         "User có quyền sửa; tồn tại KH tổ chức đầy đủ thông tin",
         "1. Bấm 'Sửa' 1 KH\n2. Quan sát form",
         "KH 'CÔNG TY ABC'",
         "- Form load đủ: thông tin chung, MST, đại diện, liên hệ, lĩnh vực, địa chỉ, "
         "thêm section Địa chỉ giao hàng + Nhân viên phụ trách + Cấp đại lý (chỉ có khi sửa)",
         "—"),
        ("008", "Sửa KH — lưu giữ nguyên mã code", "P0",
         "User có quyền sửa; KH có mã '29HNHK'",
         "1. Sửa KH, đổi Tên KH\n2. Lưu\n3. Kiểm tra mã",
         "Mã cũ='29HNHK', tên mới='CÔNG TY ABC 2'",
         "- Lưu thành công, mã KH GIỮ NGUYÊN '29HNHK' (không sinh lại)\n- Tên cập nhật",
         "BR — Update giữ nguyên code"),
        ("009", "Sửa KH — thêm/xóa Người liên hệ (sync delete-then-insert)", "P1",
         "KH tổ chức có 2 người liên hệ",
         "1. Sửa KH, xóa 1 liên hệ, thêm 1 liên hệ mới\n2. Lưu\n3. Mở lại form",
         "Trước: 2 liên hệ; sau: 1 cũ + 1 mới",
         "- Sau lưu còn đúng 2 liên hệ (1 cũ giữ lại + 1 mới), không trùng",
         "BR — sync sub-entity kiểu delete-then-insert"),
        ("010", "Xem chi tiết KH (readonly)", "P1",
         "Tồn tại KH; user có quyền xem",
         "1. Bấm 'Xem' 1 KH\n2. Quan sát form",
         "KH bất kỳ",
         "- Form mở ở chế độ readonly (/assign/customers/{id}), các field bị disable, không có nút Lưu",
         "BR — page _id/index render CustomerForm readonly"),
        ("011", "Mã KH không nhập tay", "P2",
         "Đang ở form tạo",
         "1. Quan sát khu vực mã KH",
         "—",
         "- Không có ô nhập mã KH (đã ẩn); mã sinh tự động khi Lưu",
         "BR — mã tự sinh ở thời điểm create"),
        ("012", "Thêm địa chỉ giao hàng chỉ khi Sửa", "P2",
         "Form tạo và form sửa",
         "1. Mở form tạo → tìm section Địa chỉ giao hàng\n2. Mở form sửa → tìm section đó",
         "—",
         "- Form tạo: KHÔNG có section Địa chỉ giao hàng\n- Form sửa: CÓ section Địa chỉ giao hàng",
         "BR — delivery_places chỉ thao tác khi isEdit"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Bỏ trống Tên KH", "P0",
         "Form tạo",
         "1. Bỏ trống Tên KH, điền trường khác\n2. Lưu",
         "fullname = trống",
         "- BE 422, lỗi inline 'Bắt buộc phải nhập' tại Tên KH (viền đỏ)",
         "Rule: fullname required|max:255"),
        ("002", "Cá nhân — bỏ trống SĐT", "P0",
         "Form tạo, Loại = Cá nhân",
         "1. Không nhập SĐT nào\n2. Điền trường khác\n3. Lưu",
         "phones = []",
         "- BE 422, lỗi inline tại SĐT (cần ≥1 SĐT)",
         "Rule: phones required|array|min:1 khi type=1"),
        ("003", "Cá nhân — SĐT sai định dạng", "P0",
         "Form tạo, Loại = Cá nhân",
         "1. Nhập SĐT '123' (sai)\n2. Lưu",
         "phones[0] = '123'",
         "- BE 422, lỗi 'Số điện thoại không đúng định dạng' tại SĐT",
         "Rule: phones.* regex ^(0)[0-9]{9,11}$"),
        ("004", "Tổ chức — bỏ trống MST khi không có công ty mẹ", "P0",
         "Form tạo, Loại tổ chức, không chọn công ty mẹ",
         "1. Bỏ trống MST\n2. Lưu",
         "tax_code=trống, parent_id=null",
         "- BE 422, lỗi 'Bắt buộc phải nhập' tại MST",
         "Rule: tax_code required khi không parent_id"),
        ("005", "Tổ chức — MST sai định dạng", "P1",
         "Form tạo, Loại tổ chức",
         "1. Nhập MST 'ABC@123' (có ký tự lạ)\n2. Lưu",
         "tax_code='ABC@123'",
         "- BE 422, lỗi định dạng MST",
         "Rule: tax_code regex ^[\\d\\-]{1,14}$"),
        ("006", "Tổ chức — không có Người đại diện", "P0",
         "Form tạo, Loại tổ chức",
         "1. Xóa hết người đại diện\n2. Lưu",
         "deputies = []",
         "- BE 422, lỗi tại Người đại diện (cần ≥1)",
         "Rule: deputies required|array|min:1"),
        ("007", "Tổ chức — Người đại diện thiếu Tên hoặc Chức vụ", "P1",
         "Form tạo, Loại tổ chức",
         "1. Thêm đại diện chỉ điền Tên, bỏ Chức vụ\n2. Lưu",
         "deputies.0.role=trống",
         "- BE 422, lỗi inline tại deputies.0.role",
         "Rule: deputies.*.name & deputies.*.role required"),
        ("008", "Tổ chức — không có Người liên hệ", "P0",
         "Form tạo, Loại tổ chức",
         "1. Xóa hết người liên hệ\n2. Lưu",
         "contacts = []",
         "- BE 422, lỗi tại Người liên hệ (cần ≥1)",
         "Rule: contacts required|array|min:1"),
        ("009", "Tổ chức — Người liên hệ thiếu SĐT", "P1",
         "Form tạo, Loại tổ chức",
         "1. Thêm liên hệ có Họ tên + Chức vụ nhưng không SĐT\n2. Lưu",
         "contacts.0.phones=[]",
         "- BE 422, lỗi inline tại contacts.0.phones",
         "Rule: contacts.*.phones required|array|min:1"),
        ("010", "Tổ chức — bỏ trống Địa chỉ xuất hóa đơn", "P1",
         "Form tạo, Loại tổ chức",
         "1. Bỏ trống Địa chỉ xuất hóa đơn\n2. Lưu",
         "invoice_export_address=trống",
         "- BE 422, lỗi tại Địa chỉ xuất hóa đơn",
         "Rule: invoice_export_address required"),
        ("011", "Bỏ trống địa chỉ bắt buộc (Quốc gia/Tỉnh/Phường)", "P0",
         "Form tạo",
         "1. Bỏ trống Tỉnh/TP (hoặc Phường/Quốc gia)\n2. Lưu",
         "province_id=null",
         "- BE 422, lỗi tại trường địa chỉ tương ứng",
         "Rule: nation_id/province_id/ward_id required|exists"),
        ("012", "Bỏ trống Nhóm hoặc Lĩnh vực KH", "P0",
         "Form tạo",
         "1. Bỏ trống Nhóm lĩnh vực (hoặc không chọn lĩnh vực nào)\n2. Lưu",
         "customer_scope_group_id=null / customer_scope_ids=[]",
         "- BE 422, lỗi tại Nhóm / Lĩnh vực KH",
         "Rule: customer_scope_group_id required; customer_scope_ids required|array|min:1"),
        ("013", "Lĩnh vực không thuộc Nhóm đã chọn", "P2",
         "Form tạo (giả lập payload lĩnh vực sai nhóm)",
         "1. Gửi customer_scope_ids chứa lĩnh vực thuộc nhóm khác\n2. Lưu",
         "scope_ids không thuộc group",
         "- BE 422, lỗi lĩnh vực không hợp lệ",
         "Rule: closure check scope thuộc group"),
        ("014", "Email trùng (unique ERP)", "P0",
         "ERP đã có KH email 'a@abc.com'",
         "1. Tạo KH mới nhập email 'a@abc.com'\n2. Lưu",
         "email='a@abc.com' (đã tồn tại)",
         "- BE 422, lỗi 'Email đã tồn tại'",
         "Rule: email unique trên ERP (closure)"),
        ("015", "MST trùng (unique ERP)", "P0",
         "ERP đã có KH MST '0101234567'",
         "1. Tạo KH tổ chức nhập MST '0101234567'\n2. Lưu",
         "tax_code='0101234567' (đã tồn tại)",
         "- BE 422, lỗi 'Mã số thuế đã tồn tại'",
         "Rule: tax_code unique trên ERP"),
        ("016", "CCCD trùng (unique ERP)", "P1",
         "ERP đã có KH CCCD '012345678901'",
         "1. Tạo KH cá nhân nhập CCCD '012345678901'\n2. Lưu",
         "identity_card_number trùng",
         "- BE 422, lỗi 'Số CMND/CCCD đã tồn tại'",
         "Rule: identity_card_number unique trên ERP"),
        ("017", "Email sai định dạng", "P1",
         "Form tạo",
         "1. Nhập email 'abc@' (sai)\n2. Lưu",
         "email='abc@'",
         "- BE 422, lỗi định dạng email",
         "Rule: email nullable|email|max:255"),
        ("018", "Sửa KH — unique loại trừ chính bản ghi", "P1",
         "KH 'CÔNG TY ABC' có email 'a@abc.com'",
         "1. Sửa chính KH này, giữ nguyên email 'a@abc.com'\n2. Lưu",
         "email không đổi",
         "- Lưu thành công, KHÔNG báo trùng email với chính nó",
         "Rule: Update loại trừ id hiện tại khi check unique"),
        ("019", "Lưu thất bại — hiển thị toast + lỗi inline", "P1",
         "Form tạo thiếu nhiều trường bắt buộc",
         "1. Để trống nhiều trường\n2. Lưu",
         "—",
         "- Toast 'Bạn chưa nhập đầy đủ thông tin'\n- Mọi field lỗi có viền đỏ + text lỗi inline (icon warning)",
         "BR — FE gán formError từ response.data.errors (422)"),
        ("020", "Loại KH ngoài 1-5", "P2",
         "Giả lập payload customer_type=9",
         "1. Gửi store với customer_type=9\n2. Lưu",
         "customer_type=9",
         "- BE 422, lỗi customer_type không hợp lệ",
         "Rule: customer_type required|in:1,2,3,4,5"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "User cấp công ty không thấy KH công ty khác", "P0",
         "User cty A (quyền cấp công ty); ERP có KH cty B",
         "1. Truy cập danh sách\n2. Tìm KH của cty B",
         "User cty A",
         "- KH cty B KHÔNG xuất hiện trong danh sách lẫn tìm kiếm (trừ khớp đúng SĐT)",
         "BR — applyErpVisibilityScope theo company_id"),
        ("002", "Truy cập trực tiếp /{id} KH ngoài phạm vi quyền", "P1",
         "User cty A; KH X thuộc cty B (ngoài quyền)",
         "1. Truy cập trực tiếp /assign/customers/{idKHX}",
         "id = KH cty B",
         "- Không xem được dữ liệu KH ngoài phạm vi (404/chặn) — kiểm chứng cô lập",
         "BR — phân quyền cấp áp dụng cả show"),
        ("003", "Gọi API store khi không có quyền thêm", "P0",
         "User KHÔNG có quyền 'Thêm khách hàng'",
         "1. Gọi trực tiếp POST /assign/customers với payload hợp lệ",
         "—",
         "- BE chặn 403 (middleware erpPermission)",
         "BR — middleware erpPermission:Thêm khách hàng"),
        ("004", "Gọi API update khi không có quyền sửa", "P0",
         "User KHÔNG có quyền 'Sửa khách hàng'",
         "1. Gọi trực tiếp POST /assign/customers/{id}",
         "—",
         "- BE chặn 403",
         "BR — middleware erpPermission:Sửa khách hàng"),
        ("005", "Lỗi sync HRM không làm fail Lưu", "P2",
         "Tạo KH thành công trên ERP nhưng bước sync HRM lỗi (giả lập)",
         "1. Tạo KH hợp lệ\n2. Quan sát kết quả",
         "—",
         "- KH vẫn lưu thành công trên ERP, không rollback; sync HRM best-effort",
         "BR — sync HRM best-effort sau commit ERP"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Luồng tạo KH Cá nhân đầy đủ → xuất hiện danh sách", "P0",
         "User có quyền thêm, map ERP employee",
         "1. Vào danh sách → Tạo mới\n2. Loại Cá nhân, điền đủ thông tin hợp lệ\n3. Lưu\n"
         "4. Quay về danh sách, tìm theo tên vừa tạo",
         "Tên='Nguyễn Văn A', SĐT='0901234567'",
         "- Lưu thành công → về danh sách → tìm thấy KH mới với mã tự sinh, trạng thái Hoạt động",
         "E2E — create cá nhân"),
        ("002", "Luồng tạo KH Tổ chức → sửa → xem", "P0",
         "User có quyền thêm + sửa",
         "1. Tạo KH tổ chức đầy đủ → Lưu\n2. Mở Sửa, đổi tên + thêm liên hệ → Lưu\n3. Mở Xem chi tiết",
         "Tên='CÔNG TY ABC' → 'CÔNG TY ABC 2'",
         "- Tạo OK → sửa OK (mã giữ nguyên) → xem readonly thể hiện đúng dữ liệu mới",
         "E2E — create → update → view"),
        ("003", "Luồng lọc → mở chi tiết → quay lại giữ filter", "P2",
         "Danh sách nhiều KH",
         "1. Áp filter Tỉnh=Hà Nội\n2. Mở Xem 1 KH\n3. Quay lại danh sách",
         "province=Hà Nội",
         "- Sau khi quay lại, danh sách vẫn giữ kết quả lọc (nếu hỗ trợ)/hoặc về mặc định nhất quán",
         "E2E — filter + navigation"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT  = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL; a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}")
t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL; fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",              '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",         '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",     '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",        '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                  '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0

def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, business_note, group=""):
    global current_row, data_row_idx
    values = [
        MODULE_NAME, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed", "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    nlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, longest // 4, nlines * 16 + 30) if nlines else min(180, longest // 4))
    current_row += 1
    data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                    allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}")
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Rows: 1-10 description, 11-15 summary, 17 header, 18-{current_row-1} data")
print(f"   Tổng data rows: {data_row_idx}")
