# -*- coding: utf-8 -*-
"""Builder dùng chung cho testcase.xlsx (6 màn elearning quản lý).
Mỗi màn import build() và truyền config. Nội dung TC viết theo ngôn ngữ UI thật cho tester.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(THIN, THIN, THIN, THIN)
WRAP_TL = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TC = Alignment(wrap_text=True, vertical="top", horizontal="center")
COL_WIDTHS = {'A':16,'B':26,'C':14,'D':40,'E':9,'F':32,'G':46,'H':26,'I':56,'J':34,'K':16,'L':14,'M':14,'N':14,'O':16}
HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
           "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
           "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
ROMAN = {"I":1,"II":2,"III":3,"IV":4,"V":5,"VI":6,"VII":7,"VIII":8,"IX":9,"X":10}


def build(output_file, sheet_name, feature_name, module_name, description_block,
          sections, role_tcs=None):
    wb = Workbook(); ws = wb.active; ws.title = sheet_name
    for c, w in COL_WIDTHS.items():
        ws.column_dimensions[c].width = w

    ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
    ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22
    for idx, (label, body) in enumerate(description_block, start=2):
        a = ws.cell(idx, 1, label); a.font = Font(bold=True, size=11); a.fill = PatternFill("solid", fgColor="FFF2CC"); a.alignment = WRAP_TL; a.border = BORDER
        b = ws.cell(idx, 2, body); b.font = Font(size=11); b.alignment = WRAP_TL; b.border = BORDER
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
        ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

    t = ws.cell(11, 1, f"Testcase _ {feature_name}"); t.font = Font(bold=True, size=14, color="FFFFFF"); t.fill = PatternFill("solid", fgColor="4472C4"); t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
    fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(bold=True, size=12, color="FFFFFF"); fs.fill = PatternFill("solid", fgColor="4472C4"); fs.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[11].height = 28
    for r, label, f in [(11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
                        (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
                        (13, "Số trường hợp kiểm thử đang xem xét:", '=COUNTIF(L18:N500,"Pending")'),
                        (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
                        (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")')]:
        lc = ws.cell(r, 9, label); lc.font = Font(bold=True, size=11); lc.fill = PatternFill("solid", fgColor="D9E1F2"); lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        vc = ws.cell(r, 12, f); vc.font = Font(bold=True, size=11); vc.fill = PatternFill("solid", fgColor="D9E1F2"); vc.alignment = Alignment(vertical="center", horizontal="center"); vc.border = BORDER
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
        if r > 11:
            ws.row_dimensions[r].height = 22
    ws.row_dimensions[16].height = 8

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(17, i, h); c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="4472C4"); c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER
    ws.row_dimensions[17].height = 36

    state = {"cur": 18, "ridx": 0}

    def section(title):
        cur = state["cur"]
        c = ws.cell(cur, 3, title); c.font = Font(bold=True, size=12, color="1F4E79"); c.fill = PatternFill("solid", fgColor="D6E4F0"); c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1); c.border = BORDER
        ws.merge_cells(start_row=cur, start_column=3, end_row=cur, end_column=15)
        for col in (1, 2):
            ws.cell(cur, col).fill = PatternFill("solid", fgColor="D6E4F0"); ws.cell(cur, col).border = BORDER
        ws.row_dimensions[cur].height = 26
        state["cur"] += 1

    def tc(tcid, func, prio, pre, steps, td, exp, note, group):
        cur = state["cur"]
        vals = [module_name, group, tcid, func, prio, pre, steps, td, exp, note, "",
                "Not Executed", "Not Executed", "Not Executed", ""]
        fill = PatternFill("solid", fgColor="F2F2F2") if state["ridx"] % 2 == 1 else None
        for i, v in enumerate(vals, start=1):
            c = ws.cell(cur, i, v); c.font = Font(size=11); c.alignment = WRAP_TC if i == 5 else WRAP_TL; c.border = BORDER
            if fill:
                c.fill = fill
        longest = max(len(str(v)) for v in vals)
        ws.row_dimensions[cur].height = max(30, min(230, longest // 3))
        state["cur"] += 1; state["ridx"] += 1

    if role_tcs:
        section("Phân quyền & truy cập")
        for suf, func, prio, pre, steps, td, exp, note in role_tcs:
            tc(f"TC-ROLE-{suf}", func, prio, pre, steps, td, exp, note, "Phân quyền & truy cập")

    for roman, title, tcs in sections:
        base = roman.rstrip("b")  # cho phép "VIIb" gộp cùng số VII
        section(f"{base}. {title}")
        for num, func, prio, pre, steps, td, exp, note in tcs:
            tcid = f"TC_{ROMAN[base]:02d}.{int(num):03d}"
            tc(tcid, func, prio, pre, steps, td, exp, note, title)

    dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
    dv.add(f"L18:N{state['cur'] + 100}")
    ws.add_data_validation(dv)
    wb.save(output_file)
    print(f"OK: {output_file} | data rows 18-{state['cur'] - 1}")
