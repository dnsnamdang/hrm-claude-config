import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Testcase Bonus Component"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 18, 'C': 16, 'D': 35, 'E': 10,
    'F': 30, 'G': 45, 'H': 20, 'I': 12, 'J': 50,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Thành phần tính thưởng (Bonus Component)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

MODULE = 'Payroll'
GROUP = 'Thành phần tính thưởng'
PREFIX = 'BC'

current_row = 7

def write_section_row(row, title):
    ws.merge_cells(f'C{row}:M{row}')
    cell = ws.cell(row=row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    for col in range(1, 14):
        ws.cell(row=row, column=col).border = thin_border

def write_tc_row(row, tc_id, func, priority, precondition='', steps='', test_data='', expected='', status='Not Executed'):
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', status, '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# ========================================
# I. TRANG DANH SÁCH
# ========================================
write_section_row(current_row, 'I. TRANG DANH SÁCH THÀNH PHẦN TÍNH THƯỞNG (/payroll/bonus-component)')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.001', 'Hiển thị trang danh sách', 'P0',
    'Đã đăng nhập, có quyền Payroll',
    '1. Vào menu Tính lương > Thành phần tính thưởng',
    '',
    'Hiển thị đúng layout: tiêu đề "Thành phần tính thưởng", nút "Tạo thành phần", bảng danh sách với các cột: STT, Mã TP, Tên thành phần, Loại, Công thức, Mô tả, Thao tác')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.002', 'Hiển thị dữ liệu đúng', 'P0',
    'Có ít nhất 1 thành phần trong DB',
    '1. Mở trang danh sách',
    '',
    'Các cột hiển thị đúng: Mã TP, Tên, Loại (Công thức/Thủ công dạng badge), Công thức (dạng code), Mô tả. Sắp xếp theo thời gian cập nhật mới nhất')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.003', 'Tìm kiếm theo mã thành phần', 'P0',
    'Có thành phần mã TP001',
    '1. Nhập "TP001" vào ô tìm kiếm\n2. Chờ 300ms (debounce)',
    'Keyword: TP001',
    'Danh sách chỉ hiển thị thành phần có mã chứa "TP001"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.004', 'Tìm kiếm theo tên thành phần', 'P0',
    'Có thành phần tên "Theo lương"',
    '1. Nhập "Theo lương" vào ô tìm kiếm\n2. Chờ 300ms',
    'Keyword: Theo lương',
    'Danh sách chỉ hiển thị thành phần có tên chứa "Theo lương"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.005', 'Tìm kiếm không có kết quả', 'P1',
    '',
    '1. Nhập "XYZZZZ" vào ô tìm kiếm',
    'Keyword: XYZZZZ',
    'Hiển thị "Không có dữ liệu"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.006', 'Phân trang', 'P1',
    'Có > 10 thành phần',
    '1. Đổi số bản ghi/trang sang 10\n2. Kiểm tra phân trang hiển thị',
    '',
    'Hiển thị đúng 10 bản ghi, pagination controls hoạt động đúng, tổng số bản ghi chính xác')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_001.007', 'Nút sửa hiển thị', 'P0',
    'Có ít nhất 1 thành phần',
    '1. Kiểm tra cột Thao tác',
    '',
    'Hiển thị nút sửa (icon bút) và nút xoá (icon thùng rác) cho mỗi dòng')
current_row += 1

# ========================================
# II. TẠO MỚI
# ========================================
write_section_row(current_row, 'II. TẠO MỚI THÀNH PHẦN TÍNH THƯỞNG')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.001', 'Mở modal tạo mới', 'P0',
    'Đang ở trang danh sách',
    '1. Nhấn nút "Tạo thành phần"',
    '',
    'Modal hiển thị với tiêu đề "Tạo thành phần tính thưởng". Mã TP tự sinh (format TPxxx). Cách tính mặc định "Theo công thức"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.002', 'Tạo thành phần mode Công thức - thành công', 'P0',
    'Modal tạo mới đang mở',
    '1. Nhập mã: TP001\n2. Nhập tên: Theo lương P1+P2\n3. Chọn cách tính: Theo công thức\n4. Nhập công thức: luong_p1 + luong_p2\n5. Nhấn "Lưu"',
    'Code: TP001, Name: Theo lương P1+P2, Formula: luong_p1 + luong_p2',
    'Toast "Lưu thành công". Modal đóng. Thành phần mới hiển thị đầu danh sách')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.003', 'Tạo thành phần mode Thủ công - thành công', 'P0',
    'Modal tạo mới đang mở',
    '1. Nhập mã: TP002\n2. Nhập tên: Điều chỉnh thủ công\n3. Chọn cách tính: Nhập thủ công\n4. Nhấn "Lưu"',
    'Code: TP002, Name: Điều chỉnh thủ công, Mode: manual',
    'Toast "Lưu thành công". Modal đóng. Thành phần mới hiển thị với loại "Thủ công"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.004', 'Validate - thiếu mã thành phần', 'P0',
    'Modal tạo mới đang mở',
    '1. Để trống mã TP\n2. Nhập tên\n3. Nhấn "Lưu"',
    '',
    'Hiển thị lỗi "Bắt buộc nhập mã thành phần". Không tạo được')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.005', 'Validate - thiếu tên thành phần', 'P0',
    'Modal tạo mới đang mở',
    '1. Nhập mã TP\n2. Để trống tên\n3. Nhấn "Lưu"',
    '',
    'Hiển thị lỗi "Bắt buộc nhập tên thành phần". Không tạo được')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.006', 'Validate - mã trùng', 'P0',
    'Đã có thành phần mã TP001',
    '1. Nhập mã: TP001 (trùng)\n2. Nhập tên\n3. Nhấn "Lưu"',
    'Code: TP001 (đã tồn tại)',
    'Hiển thị lỗi "Mã thành phần đã tồn tại". Không tạo được')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.007', 'Validate - mã quá 50 ký tự', 'P1',
    'Modal tạo mới đang mở',
    '1. Nhập mã dài > 50 ký tự\n2. Nhấn "Lưu"',
    'Code: 51 ký tự',
    'Hiển thị lỗi "Mã thành phần tối đa 50 ký tự"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_002.008', 'Validate - tên quá 255 ký tự', 'P1',
    'Modal tạo mới đang mở',
    '1. Nhập tên dài > 255 ký tự\n2. Nhấn "Lưu"',
    'Name: 256 ký tự',
    'Hiển thị lỗi "Tên thành phần tối đa 255 ký tự"')
current_row += 1

# ========================================
# III. SOẠN CÔNG THỨC
# ========================================
write_section_row(current_row, 'III. SOẠN CÔNG THỨC')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.001', 'Autocomplete biến hệ thống', 'P0',
    'Modal tạo/sửa, mode = Công thức',
    '1. Gõ "luong" vào ô công thức',
    'Input: luong',
    'Hiển thị gợi ý: luong_p1, luong_p2, luong_p3. Mỗi gợi ý có code + tên mô tả')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.002', 'Chọn gợi ý autocomplete', 'P0',
    'Gợi ý đang hiển thị',
    '1. Click vào gợi ý "luong_p1"\nHoặc nhấn Tab để chọn gợi ý đầu tiên',
    '',
    'Biến "luong_p1" được chèn vào vị trí con trỏ trong ô công thức. Gợi ý đóng')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.003', 'Chèn biến từ danh sách nhanh', 'P0',
    'Modal tạo/sửa, mode = Công thức',
    '1. Click vào biến "ds_net_nv" trong phần "Biến hệ thống"',
    '',
    'Biến "ds_net_nv" được chèn vào cuối ô công thức')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.004', 'Chèn hàm từ danh sách nhanh', 'P0',
    'Modal tạo/sửa, mode = Công thức',
    '1. Click vào hàm "IF" trong phần "Hàm tính toán"',
    '',
    'Template "IF(condition, value_if_true, value_if_false)" được chèn vào ô công thức')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.005', 'Tham chiếu thành phần khác', 'P0',
    'Có thành phần TP001 (mode formula)',
    '1. Gõ "TP" vào ô công thức',
    '',
    'Hiển thị gợi ý thành phần khác (TP001...). Click để chèn mã TP vào công thức')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.006', 'Preview công thức realtime', 'P1',
    'Modal tạo/sửa, mode = Công thức',
    '1. Nhập công thức: luong_p1 + luong_p2 * 2',
    '',
    'Preview hiển thị công thức với syntax highlighting: biến (xanh), số (tím), toán tử (xám)')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.007', 'Validate công thức - thành công', 'P0',
    'Công thức hợp lệ đã nhập',
    '1. Nhập: luong_p1 + luong_p2\n2. Nhấn "Kiểm tra công thức"',
    'Formula: luong_p1 + luong_p2',
    'Hiển thị thông báo "Công thức hợp lệ" (màu xanh)')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.008', 'Validate công thức - biến không tồn tại', 'P0',
    'Modal tạo/sửa',
    '1. Nhập: abc_xyz + luong_p1\n2. Nhấn "Kiểm tra công thức"',
    'Formula: abc_xyz + luong_p1',
    'Hiển thị lỗi: biến "abc_xyz" không tồn tại trong hệ thống')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.009', 'Validate công thức - ngoặc không khớp', 'P1',
    'Modal tạo/sửa',
    '1. Nhập: MAX(luong_p1, luong_p2\n2. Nhấn "Kiểm tra công thức"',
    'Formula: MAX(luong_p1, luong_p2',
    'Hiển thị lỗi: ngoặc mở/đóng không khớp')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_003.010', 'Đổi mode từ Công thức sang Thủ công', 'P1',
    'Modal tạo/sửa, mode = Công thức, đã nhập công thức',
    '1. Đổi cách tính sang "Nhập thủ công"',
    '',
    'Ẩn phần soạn công thức. Hiển thị thông báo "Thành phần thủ công sẽ nhập % chia và số tiền trực tiếp..."')
current_row += 1

# ========================================
# IV. SỬA THÀNH PHẦN
# ========================================
write_section_row(current_row, 'IV. SỬA THÀNH PHẦN TÍNH THƯỞNG')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_004.001', 'Mở modal sửa', 'P0',
    'Có thành phần TP001 trong danh sách',
    '1. Click icon sửa trên dòng TP001',
    '',
    'Modal mở với tiêu đề "Sửa thành phần tính thưởng". Load đúng dữ liệu: mã, tên, cách tính, công thức, ghi chú')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_004.002', 'Sửa tên - thành công', 'P0',
    'Modal sửa đang mở',
    '1. Sửa tên thành "Theo lương P1+P2+P3"\n2. Nhấn "Lưu"',
    '',
    'Toast "Lưu thành công". Modal đóng. Tên mới hiển thị trong danh sách')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_004.003', 'Sửa mã - trùng với thành phần khác', 'P0',
    'Có TP001 và TP002. Đang sửa TP001',
    '1. Đổi mã thành TP002\n2. Nhấn "Lưu"',
    'Code: TP002 (trùng với thành phần khác)',
    'Hiển thị lỗi "Mã thành phần đã tồn tại"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_004.004', 'Sửa mã - giữ nguyên mã cũ', 'P1',
    'Đang sửa TP001',
    '1. Giữ nguyên mã TP001\n2. Sửa tên\n3. Nhấn "Lưu"',
    '',
    'Lưu thành công (unique validation loại trừ chính nó)')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_004.005', 'Sửa công thức', 'P0',
    'Thành phần mode Công thức đang mở',
    '1. Sửa công thức thành: luong_p1 + luong_p2 + luong_p3\n2. Nhấn "Lưu"',
    '',
    'Lưu thành công. Công thức mới hiển thị đúng trong danh sách')
current_row += 1

# ========================================
# V. XOÁ THÀNH PHẦN
# ========================================
write_section_row(current_row, 'V. XOÁ THÀNH PHẦN TÍNH THƯỞNG')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_005.001', 'Xoá thành phần - xác nhận', 'P0',
    'Có thành phần TP003 trong danh sách',
    '1. Click icon xoá trên dòng TP003',
    '',
    'Hiển thị modal xác nhận "Bạn có chắc chắn muốn xoá thành phần TP003 — [tên]?"')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_005.002', 'Xoá thành phần - thành công', 'P0',
    'Modal xác nhận đang hiển thị',
    '1. Nhấn "Xoá"',
    '',
    'Toast "Xoá thành công". Thành phần biến mất khỏi danh sách. Tổng số bản ghi giảm 1')
current_row += 1

write_tc_row(current_row, f'{PREFIX}_005.003', 'Xoá thành phần - huỷ', 'P1',
    'Modal xác nhận đang hiển thị',
    '1. Nhấn "Huỷ"',
    '',
    'Modal đóng. Thành phần vẫn còn trong danh sách')
current_row += 1

# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add(f'L8:L500')

# === SAVE ===
output_path = '/Users/nguyentrancu/DEV/code/HRM/hrm-claude-config/.plans/bonus-component/Testcase_Bonus_Component.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
