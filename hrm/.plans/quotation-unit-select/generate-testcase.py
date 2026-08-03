"""Generate testcase Excel cho feature Chọn ĐVT khi tạo/sửa báo giá (hàng ERP)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE  = ".plans/quotation-unit-select/testcase.xlsx"
SHEET_NAME   = "QuotationUnitSelect"
FEATURE_NAME = "Chọn ĐVT khi tạo/sửa báo giá (hàng ERP)"
MODULE_NAME  = "Báo giá / Giao việc"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Cho phép chọn Đơn vị tính (ĐVT) cho dòng hàng ERP ở màn Tạo/Sửa báo giá. Khi đổi ĐVT, phần mềm LẤY LẠI đơn giá bán + đơn giá vốn theo đúng đơn vị được chọn (giá lưu theo từng đơn vị bên ERP). Màn Xem chi tiết vẫn hiển thị ĐVT dạng text."),
    ("2. Đối tượng được tính / hiển thị (cột ĐVT = SELECT)",
     "► Báo giá type=2 (tự lập / SELF_BUILT), màn Tạo/Sửa.\n"
     "► Dòng hàng ERP ĐƠN: erp_product_id != null, KHÔNG phải cha-có-con, KHÔNG phải con (parent_id null).\n"
     "► Dropdown = TẤT CẢ đơn vị của sản phẩm (product_units), kể cả đơn vị thiếu giá bán lẻ (→ giá 0)."),
    ("3. Đối tượng bị ẩn / không áp (cột ĐVT = TEXT)",
     "► Hàng tạm (tự tạo), dòng dịch vụ.\n"
     "► Combo ERP: dòng cha có con recipe + các dòng con của nó.\n"
     "► Báo giá type=1 (từ BOM) — kế thừa nguyên vẹn từ BOM.\n"
     "► Màn Xem chi tiết báo giá — luôn text."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "Không áp dụng — feature thao tác trên dòng hàng, không có bộ lọc thời gian."),
    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "ERP (connection mysql2): product_units(product_id, unit_id, is_base, unit_coefficient, cost_price) + "
     "product_unit_prices(product_unit_id, price_type_id=1 bán lẻ, price) + units(id, name). "
     "Mỗi sản phẩm có N đơn vị; giá vốn ở product_units.cost_price, giá bán ở product_unit_prices.price theo từng đơn vị."),
    ("6. Quy tắc cộng dồn / tính toán",
     "► Thành tiền bán dòng = quoted_price × qty_needed; Thành tiền nhập = estimated_price × qty_needed.\n"
     "► TSLN dòng = (Thành tiền bán sau CK − Thành tiền nhập) / Thành tiền nhập × 100%.\n"
     "► Tổng báo giá = Σ dòng (computed reactive, tự cập nhật khi đổi giá).\n"
     "► Đổi ĐVT CHỈ đổi đơn giá; SỐ LƯỢNG giữ nguyên (user tự sửa, KHÔNG auto quy đổi theo hệ số)."),
    ("7. Phân quyền cấp",
     "• 'Xem giá vốn hàng hoá' (permission id 1045, guard api) — gate hiển thị Giá nhập / Thành tiền nhập / TSLN của hàng ERP. "
     "Endpoint POST /assign/quotations/erp-product-units trả cost_price=NULL nếu KHÔNG có quyền (chặn tại API, không dựa FE). "
     "canEdit (sửa dòng) = báo giá status=1 (Đang tạo) AND người đăng nhập là người tạo."),
    ("8. Cách tính đơn giá theo ĐVT",
     "► Giá bán = product_unit_prices.price (price_type_id=1) của đơn vị được chọn, ÷ tỷ giá nếu rate>1, làm tròn 2 số lẻ.\n"
     "► Giá vốn = product_units.cost_price của đơn vị được chọn, cùng cách quy đổi.\n"
     "► LƯU Ý: dùng GIÁ GỐC, KHÔNG nhân hệ số công ty (product_company_coefficients) — khác popup 'Thêm hàng hoá' (proxy ERP). Nhất quán hành vi lưu cũ.\n"
     "► BE re-derive giá theo (erp_product_id, unit_id) khi LƯU (nguồn chân lý, không tin giá FE) — cả tạo lẫn cập nhật."),
    ("9. Ghi chú đọc bảng",
     "Ô Giá nhập hàng ERP luôn disabled (khoá, không gõ tay). Đơn vị bị xoá bên ERP → fallback đơn vị cơ bản. "
     "Đổi ĐVT có kèm khép rủi ro chiết khấu (xem section VI)."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("00", "Mở màn Sửa báo giá type=2 có hàng ERP",
     "P0", "User là người tạo báo giá; báo giá type=2, status=1 (Đang tạo); có ≥1 dòng hàng ERP đơn có nhiều đơn vị (vd product 3920: Cái/Hộp/Thùng)",
     "1. Vào Sửa báo giá\n2. Quan sát cột ĐVT của dòng hàng ERP",
     "Báo giá type=2, dòng ERP 3920",
     "- Cột ĐVT dòng hàng ERP đơn là <select> đủ các đơn vị của sản phẩm (Cái, Hộp, Thùng)\n- Đơn vị đang lưu (Cái) được chọn sẵn",
     "canEdit = status=1 AND creator = current user"),
    ("01", "CÓ quyền 'Xem giá vốn hàng hoá' → đổi ĐVT thấy giá vốn",
     "P0", "User có quyền 'Xem giá vốn hàng hoá'; đang Sửa báo giá type=2 có hàng ERP 3920",
     "1. Đổi ĐVT dòng ERP từ 'Cái' → 'Hộp'\n2. Quan sát cột Giá nhập, Thành tiền nhập, TSLN",
     "User: có quyền cost; Hộp giá bán 90.000, giá vốn 50.849",
     "- Giá bán dòng = 90.000 (theo Hộp)\n- Giá nhập (giá vốn) = 50.849 hiển thị (ô disabled)\n- Thành tiền nhập + TSLN cập nhật theo",
     "Permission: 'Xem giá vốn hàng hoá' (1045) → BE trả cost_price thật"),
    ("02", "KHÔNG quyền 'Xem giá vốn' → đổi ĐVT giá bán đổi, giá vốn ẩn",
     "P0", "User KHÔNG có quyền 'Xem giá vốn hàng hoá'; đang Sửa báo giá type=2 của chính họ có hàng ERP 3920",
     "1. Đổi ĐVT dòng ERP từ 'Cái' → 'Hộp'\n2. Quan sát Giá bán + Giá nhập + TSLN\n3. Kiểm response API erp-product-units (DevTools)",
     "User: KHÔNG quyền cost",
     "- Giá bán = 90.000 (đổi bình thường — không thuộc gate)\n- Giá nhập / Thành tiền nhập / TSLN hiển thị '—'\n- Response API: cost_price=null cho MỌI đơn vị (không leak 50.849)",
     "BR bảo mật: gate tại API (cost=null), không dựa FE ẩn"),
]

SECTIONS = [
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Cột ĐVT hàng ERP đơn là select", "P0",
         "Sửa báo giá type=2, dòng hàng ERP đơn (product 3920)",
         "1. Mở Sửa báo giá\n2. Quan sát cột ĐVT dòng ERP",
         "product 3920 (3 đơn vị)",
         "- Cột ĐVT là <select>\n- Options = Cái, Hộp, Thùng (đủ đơn vị của SP)\n- Giá trị chọn sẵn = đơn vị đang lưu",
         "BR — chỉ hàng ERP đơn type=2 mới có select"),
        ("002", "Hàng tạm / dịch vụ giữ ĐVT dạng text", "P1",
         "Sửa báo giá type=2 có 1 dòng hàng tạm (tự tạo) + 1 dòng dịch vụ",
         "1. Mở Sửa báo giá\n2. Quan sát cột ĐVT dòng hàng tạm và dòng dịch vụ",
         "1 hàng tạm + 1 dịch vụ",
         "- Cột ĐVT dòng hàng tạm hiển thị TEXT (không select)\n- Dòng dịch vụ ĐVT text",
         "BR — hàng tạm/dịch vụ ngoài phạm vi"),
        ("003", "Combo ERP (cha có con) giữ ĐVT text", "P1",
         "Sửa báo giá type=2 có 1 hàng ERP là combo (cha có con recipe)",
         "1. Mở Sửa báo giá\n2. Quan sát ĐVT dòng cha combo và dòng con",
         "1 combo ERP cha-con",
         "- ĐVT dòng cha combo: TEXT (không select)\n- ĐVT dòng con: TEXT",
         "BR — combo ngoài phạm vi (tránh vỡ roll-up)"),
        ("004", "Báo giá type=1 (từ BOM) giữ ĐVT text", "P1",
         "Sửa báo giá type=1 (tạo từ BOM) có hàng ERP",
         "1. Mở Sửa báo giá type=1\n2. Quan sát cột ĐVT",
         "Báo giá type=1",
         "- Cột ĐVT hiển thị TEXT toàn bộ (không select)",
         "BR — type=1 kế thừa nguyên vẹn từ BOM"),
        ("005", "Màn Xem chi tiết báo giá — ĐVT text", "P1",
         "Báo giá type=2 có hàng ERP; mở màn Xem chi tiết (không phải Sửa)",
         "1. Mở Xem chi tiết báo giá\n2. Quan sát cột ĐVT",
         "Báo giá type=2",
         "- Cột ĐVT hiển thị TEXT (không có select)",
         "BR — chỉ màn Tạo/Sửa mới cho chọn ĐVT"),
        ("006", "Nhiều dòng hàng ERP — mỗi dòng select độc lập", "P1",
         "Sửa báo giá type=2 có 2 dòng hàng ERP khác nhau (SP A và SP B, mỗi SP nhiều đơn vị)",
         "1. Mở Sửa\n2. Đổi ĐVT dòng A\n3. Quan sát dòng B",
         "2 dòng ERP A, B",
         "- Mỗi dòng có select ĐVT riêng theo đơn vị của SP đó\n- Đổi ĐVT dòng A KHÔNG ảnh hưởng giá/ĐVT dòng B",
         "BR — options theo từng erp_product_id"),
    ]),
    ("V", "CHỨC NĂNG CHÍNH — ĐỔI ĐVT & TÍNH LẠI GIÁ / TỔNG / TSLN", [
        ("001", "Đổi ĐVT → lấy lại giá bán + giá vốn theo đơn vị", "P0",
         "Sửa báo giá type=2, user có quyền cost; dòng ERP 3920 đang ĐVT 'Cái'",
         "1. Đổi ĐVT 'Cái' → 'Hộp'\n2. Quan sát Giá bán, Giá nhập",
         "Cái: giá 0 ; Hộp: bán 90.000 / vốn 50.849",
         "- Giá bán dòng đổi 0 → 90.000\n- Giá vốn dòng đổi → 50.849\n- Cả 2 lấy cùng đơn vị 'Hộp'",
         "BR chính của feature"),
        ("002", "Thành tiền + Tổng giá trị báo giá cập nhật khi đổi ĐVT", "P0",
         "Như trên, SL dòng = 1, không CK",
         "1. Ghi lại Tổng trước\n2. Đổi ĐVT 'Cái' → 'Hộp'\n3. Quan sát Thành tiền bán, bảng Tổng hợp, Tổng cộng",
         "SL=1",
         "- Thành tiền bán dòng = 90.000 × 1 = 90.000\n- Bảng Tổng hợp giá trị + Tổng cộng cập nhật đúng theo giá mới",
         "Công thức: Thành tiền = đơn giá × SL; Tổng = Σ dòng"),
        ("003", "TSLN cập nhật đúng khi đổi ĐVT", "P0",
         "User có quyền cost; dòng ERP đổi sang 'Hộp' (bán 90.000, vốn 50.849), SL=1, không CK",
         "1. Đổi ĐVT sang 'Hộp'\n2. Quan sát cột TSLN",
         "bán 90.000 / vốn 50.849",
         "- TSLN = (90.000 − 50.849)/50.849 × 100 ≈ 77%\n- Cả giá bán và giá vốn cùng đơn vị → TSLN nhất quán",
         "Công thức: TSLN=(sale sau CK − import)/import"),
        ("004", "Số lượng GIỮ NGUYÊN khi đổi ĐVT", "P0",
         "Dòng ERP có SL = 5 (Cái)",
         "1. Đổi ĐVT 'Cái' → 'Hộp'\n2. Quan sát cột Số lượng",
         "SL=5",
         "- Số lượng vẫn = 5 (không tự quy đổi theo hệ số)\n- Thành tiền = 90.000 × 5 = 450.000",
         "BR đã chốt: SL user tự sửa, không auto convert"),
        ("005", "Lưu → mở lại: giá đúng theo ĐVT đã chọn (BE re-derive)", "P0",
         "Sửa báo giá, đổi ĐVT sang 'Hộp', bấm Lưu",
         "1. Đổi ĐVT sang 'Hộp'\n2. Lưu báo giá\n3. Mở lại màn Sửa/Xem",
         "Hộp",
         "- Sau lưu, dòng giữ ĐVT 'Hộp'\n- Giá bán/giá vốn = giá của 'Hộp' (BE tự lấy lại theo unit_id, không tin giá FE)",
         "BR — BE là nguồn chân lý (saveDirectProduct re-derive create+update)"),
        ("006", "Đổi qua lại nhiều đơn vị", "P1",
         "Dòng ERP 3920 (Cái/Hộp/Thùng)",
         "1. Đổi 'Cái' → 'Hộp' (giá 90.000)\n2. Đổi 'Hộp' → 'Thùng' (giá 0)\n3. Đổi 'Thùng' → 'Cái'",
         "Cái 0 / Hộp 90.000 / Thùng 0",
         "- Mỗi lần đổi, giá bán + giá vốn cập nhật đúng theo đơn vị hiện chọn\n- Không giữ giá của đơn vị trước",
         "BR — luôn lấy giá theo đơn vị hiện tại"),
        ("007", "Đổi ĐVT KHÔNG đổi VAT của dòng", "P1",
         "Dòng ERP có VAT (vat_percent) = 8%",
         "1. Ghi lại VAT dòng\n2. Đổi ĐVT sang đơn vị khác\n3. Quan sát cột VAT",
         "VAT=8%",
         "- VAT dòng vẫn = 8% (VAT theo sản phẩm ERP, không theo đơn vị)\n- Chỉ đơn giá đổi",
         "BR — onChangeUnit không đụng vat_percent"),
        ("008", "Báo giá ngoại tệ (rate>1) — giá quy đổi theo tỷ giá khi đổi ĐVT", "P1",
         "Báo giá tiền tệ USD, exchange_rate > 1; dòng ERP nhiều đơn vị",
         "1. Đổi ĐVT sang đơn vị có giá VND = P\n2. Quan sát đơn giá hiển thị",
         "USD, rate=R>1",
         "- Đơn giá hiển thị = round(P / R, 2) (quy đổi theo tỷ giá)\n- Nhất quán với cách thêm hàng ERP",
         "Công thức: rate>1 ? round(price/rate,2) : price"),
    ]),
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Đơn vị thiếu giá bán lẻ → giá bán = 0", "P1",
         "Dòng ERP có đơn vị 'Thùng' không khai giá bán lẻ (product_unit_prices trống)",
         "1. Đổi ĐVT sang 'Thùng'\n2. Quan sát Giá bán",
         "Thùng: retail null",
         "- Giá bán dòng = 0 (hiển thị rõ để user xử lý)\n- Không lỗi/không crash",
         "LEFT JOIN + coalesce 0"),
        ("002", "CK theo % → tính lại discount_amount theo giá mới", "P0",
         "Dòng ERP ĐVT 'Cái' (giá X), nhập chiết khấu theo % (mode=percent), vd 10%",
         "1. Nhập CK 10% (theo %)\n2. Đổi ĐVT sang 'Hộp' (giá 90.000)\n3. Quan sát CK tiền + Thành tiền sau CK + TSLN",
         "CK 10%; Hộp 90.000",
         "- discount_amount tính lại = 90.000 × 10% = 9.000\n- Thành tiền sau CK + TSLN đúng theo giá mới (không dùng CK cũ stale)",
         "Fix: onChangeUnit gọi onDiscountPercentInput khi mode=percent"),
        ("003", "CK theo tiền > giá mới → validation báo đỏ", "P0",
         "Dòng ERP ĐVT 'Hộp' (90.000), nhập CK theo tiền = 50.000; đổi sang đơn vị rẻ hơn (giá < 50.000)",
         "1. Nhập CK tiền = 50.000\n2. Đổi ĐVT sang đơn vị có giá < 50.000\n3. Quan sát ô CK",
         "CK 50.000",
         "- Ô CK bị đánh dấu is-invalid (viền đỏ) + báo 'GG không được lớn hơn đơn giá bán'\n- Chặn gửi duyệt cho tới khi user sửa",
         "isItemDiscountInvalid: discount_amount > quoted_price"),
        ("004", "Phương pháp 2 (phân bổ) → cảnh báo phân bổ lại", "P0",
         "Báo giá dùng CK phương pháp 2, đã phân bổ giảm giá; dòng ERP có allocated_discount_amount > 0",
         "1. Đổi ĐVT dòng ERP\n2. Quan sát thông báo",
         "PP2, đã phân bổ",
         "- Hiện toast 'Giá bán đã thay đổi, vui lòng phân bổ lại giảm giá'\n- allocationStale = true (nhắc user phân bổ lại)",
         "Fix: onChangeUnit gọi onSalePriceChange"),
        ("005", "Đơn vị bị xoá bên ERP → fallback đơn vị cơ bản khi lưu", "P2",
         "Dòng đã lưu unit_id là đơn vị sau đó bị xoá bên ERP",
         "1. Mở Sửa báo giá\n2. Lưu lại",
         "unit_id không còn trong product_units",
         "- BE getUnitPrice fallback đơn vị cơ bản (is_base=1)\n- Không lỗi; giá lấy theo đơn vị cơ bản",
         "getUnitPrice fallback base khi unit không khớp"),
        ("006", "ERP offline khi nạp đơn vị → dòng giữ text, không vỡ màn", "P2",
         "mysql2/ERP tạm không truy vấn được khi load màn Sửa",
         "1. Mở Sửa báo giá khi nạp unit lỗi",
         "erp-product-units lỗi",
         "- loadUnitOptions catch im lặng → dòng ERP hiển thị ĐVT text (fallback)\n- Màn không crash",
         "try/catch im lặng ở loadUnitOptions"),
        ("007", "Sản phẩm ERP chỉ có 1 đơn vị", "P2",
         "Dòng hàng ERP mà sản phẩm chỉ khai 1 đơn vị (chỉ base)",
         "1. Mở Sửa\n2. Quan sát select ĐVT dòng đó",
         "SP 1 đơn vị",
         "- Select ĐVT hiển thị 1 option (đơn vị cơ bản)\n- Không lỗi khi mở/chọn",
         "Edge — dropdown 1 phần tử vẫn hợp lệ"),
        ("008", "Đổi ĐVT rồi Hủy (không lưu) → DB không đổi", "P1",
         "Dòng ERP đang ĐVT 'Cái'; user đổi sang 'Hộp' nhưng bấm Hủy/không lưu",
         "1. Đổi ĐVT sang 'Hộp'\n2. Rời màn / bấm Hủy (không Lưu)\n3. Mở lại báo giá",
         "Không bấm Lưu",
         "- Báo giá giữ ĐVT 'Cái' như cũ (thay đổi chỉ ở FE, chưa persist)\n- Không ảnh hưởng dữ liệu",
         "BR — chỉ Lưu mới ghi DB"),
    ]),
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Không leak giá vốn khi thiếu quyền (API null)", "P0",
         "User KHÔNG có quyền 'Xem giá vốn hàng hoá'",
         "1. Mở Sửa báo giá có hàng ERP\n2. Xem response API POST erp-product-units (DevTools)",
         "User không quyền cost",
         "- Response: mỗi đơn vị có cost_price = null (mọi sản phẩm, mọi đơn vị)\n- FE không có dữ liệu giá vốn để hiện",
         "BR bảo mật: gate tại API, defense-in-depth"),
        ("002", "BE re-derive giá khi lưu — không tin giá FE", "P0",
         "User dùng DevTools sửa giá FE của dòng ERP rồi Lưu",
         "1. Sửa payload quoted_price/estimated_price của dòng ERP\n2. Lưu\n3. Mở lại kiểm giá",
         "Payload giá FE bị sửa",
         "- Sau lưu, giá dòng ERP = giá lấy từ ERP theo unit_id (KHÔNG theo giá FE gửi)",
         "BR — saveDirectProduct re-derive theo (erp_product_id, unit_id)"),
        ("003", "Không sửa được dòng của báo giá không phải mình / đã duyệt", "P1",
         "Báo giá status != 1 (đã duyệt) hoặc người khác tạo",
         "1. Mở màn Sửa (nếu vào được)\n2. Quan sát select ĐVT",
         "status != 1 hoặc khác creator",
         "- canEdit=false → select ĐVT disabled (không đổi được)",
         "canEdit = status=1 AND creator = current user"),
    ]),
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đầy đủ: đổi ĐVT → lưu → mở lại (có quyền)", "P0",
         "User có quyền cost; báo giá type=2 có hàng ERP 3920",
         "1. Mở Sửa\n2. Đổi ĐVT 'Cái' → 'Hộp'\n3. Kiểm giá bán 90.000 + giá vốn 50.849 + tổng/TSLN\n4. Lưu\n5. Mở lại",
         "product 3920",
         "- Bước 3: giá bán 90.000, giá vốn 50.849, tổng + TSLN cập nhật\n- Sau lưu: ĐVT 'Hộp', giá đúng theo ĐVT (BE re-derive)",
         "E2E chính (đã tự động hoá Playwright)"),
        ("002", "Luồng không quyền: đổi ĐVT giá bán đổi, giá vốn ẩn", "P0",
         "User KHÔNG quyền cost; báo giá type=2 của họ có hàng ERP 3920",
         "1. Mở Sửa\n2. Đổi ĐVT 'Cái' → 'Hộp'\n3. Kiểm giá bán + giá vốn",
         "product 3920",
         "- Giá bán 90.000 (đổi); Giá vốn/Thành tiền nhập/TSLN '—' (không lộ 50.849)",
         "E2E case permission (đã tự động hoá Playwright)"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT  = Font(name="Calibri", size=11)
WRAP_TOP_LEFT   = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
TITLE_FONT      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL      = PatternFill("solid", fgColor="4472C4")
SUMMARY_LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL  = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT  = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
COL_WIDTHS = {'A':22,'B':22,'C':16,'D':42,'E':10,'F':32,'G':55,'H':22,'I':65,'J':35,'K':18,'L':16,'M':16,'N':16,'O':22}

wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1"); ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label); a.font = DESC_LABEL_FONT; a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT; a.border = BORDER
    b = ws.cell(idx, 2, body); b.font = DESC_BODY_FONT; b.alignment = WRAP_TOP_LEFT; b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, f"Testcase _ {FEATURE_NAME}"); t.font = TITLE_FONT; t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11"); ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY"); fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL; fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):",              '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):",         '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):",     '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:",        '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:",                  '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label); lc.font = SUMMARY_LABEL_FONT; lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right"); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula); vc.font = SUMMARY_VALUE_FONT; vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN; vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11: ws.row_dimensions[r].height = 22
ws.row_dimensions[16].height = 8

HEADERS = ["Module","Nhóm chức năng","TC ID","Chức năng","Priority","Tiền điều kiện","Bước thực hiện","Test Data",
           "Expected Result (chi tiết)","Giải thích nghiệp vụ","KQ thực tế",
           "trạng thái check lần 1","trạng thái check lần 2","trạng thái check lần 3","Ghi chú"]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = HEADER_ALIGN; c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18; data_row_idx = 0
def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN; cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL; ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26; current_row += 1

def write_tc(tc_id, function, priority, precondition, steps, test_data, expected, note, group=""):
    global current_row, data_row_idx
    values = [MODULE_NAME, group, tc_id, function, priority, precondition, steps, test_data,
              expected, note, "", "Not Executed", "Not Executed", "Not Executed", ""]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v); c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER; c.border = BORDER
        if fill: c.fill = fill
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, longest // 3))
    current_row += 1; data_row_idx += 1

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc(f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp, note, group="Phân quyền & truy cập")

ROMANS = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]
for roman, title, tcs in SECTIONS:
    write_section_row(f"{roman}. {title}")
    sec_idx = ROMANS.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = f"TC_{sec_idx:02d}.{int(tc_num):03d}"
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)

dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True, showDropDown=False)
dv.add(f"L18:N{current_row + 100}"); ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"✅ Generated: {OUTPUT_FILE}")
print(f"   Rows: 1-10 description, 11-15 summary, 17 header, 18-{current_row-1} data")
