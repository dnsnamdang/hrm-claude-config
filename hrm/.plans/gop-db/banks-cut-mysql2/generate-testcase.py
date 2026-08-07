# -*- coding: utf-8 -*-
"""Generate testcase Excel cho man Danh muc ngan hang (Modules/MasterData - pages/master-data/banks)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE = ".plans/gop-db/banks-cut-mysql2/testcase.xlsx"
SHEET_NAME = "DanhMucNganHang"
FEATURE_NAME = "Danh mục ngân hàng"
MODULE_NAME = "Danh mục ngân hàng"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý danh mục ngân hàng dùng chung toàn hệ thống HRM (Danh mục chung → Danh mục ngân hàng, URL /master-data/banks).\n"
     "Cho phép: xem danh sách + lọc, tạo mới / sửa / xem chi tiết / xoá ngân hàng, khoá – mở khoá ngân hàng, "
     "tra cứu nhanh thông tin ngân hàng từ dịch vụ ngoài (VietQR) để điền tự động, và quản lý danh sách chi nhánh của từng ngân hàng.\n"
     "Dữ liệu này là nguồn cho các màn: Hồ sơ nhân viên (tài khoản ngân hàng / tài khoản uỷ quyền), "
     "Danh mục tài khoản ngân hàng công ty (company_accounts), và đồng bộ sang CRM khi bật cấu hình use_crm."),

    ("2. Đối tượng được tính / hiển thị",
     "► Toàn bộ bản ghi bảng `banks` — KHÔNG lọc theo công ty / phòng ban / người tạo (danh mục dùng chung, không phân cấp dữ liệu).\n"
     "► Bao gồm CẢ ngân hàng status = 1 (Hoạt động) và status = 2 (Khoá) — bản ghi khoá vẫn hiển thị, chỉ bị chặn thao tác Sửa/Xoá.\n"
     "► Cột 'Chi nhánh' đếm toàn bộ bản ghi `bank_branches` có bank_id = id ngân hàng (eager load `with('branches')`), không phụ thuộc bộ lọc.\n"
     "► Modal 'Chi nhánh ngân hàng': hiển thị bản ghi `bank_branches` của đúng bank_id, INNER JOIN `provinces` để lấy tên tỉnh/TP; mặc định limit = 100.\n"
     "► Modal 'Tra cứu' (Thông tin ngân hàng): dữ liệu lấy trực tiếp từ API ngoài https://api.vietqr.io/v2/banks (không phải dữ liệu HRM)."),

    ("3. Đối tượng bị ẩn / không tính",
     "► Bản ghi đã xoá cứng (màn không dùng SoftDeletes — xoá là xoá vĩnh viễn khỏi `banks`, kèm xoá toàn bộ `bank_branches` con).\n"
     "► Trong modal chi nhánh: chi nhánh có `province_id` NULL hoặc trỏ tới tỉnh/TP không tồn tại sẽ KHÔNG hiển thị (do INNER JOIN provinces).\n"
     "► Ở các màn tiêu thụ danh mục (Hồ sơ nhân viên): dropdown chỉ nhận ngân hàng status = 1, ngân hàng đã Khoá bị loại khỏi danh sách chọn.\n"
     "► Không có khái niệm 'ẩn theo quyền' — mọi user đã đăng nhập đều thấy đủ danh sách."),

    ("4. Bộ lọc thời gian áp dụng cho",
     "Không áp dụng — màn danh mục không có bộ lọc theo khoảng thời gian.\n"
     "Cột ngày duy nhất là 'Cập nhật' (`banks.updated_at`, format qua Helper::formatDate) và chỉ dùng để hiển thị + sắp xếp, không dùng để lọc."),

    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► `banks` (1) — (N) `bank_branches` qua `bank_branches.bank_id`. Quan hệ 2 cấp, không có cây sâu hơn.\n"
     "► `bank_branches.province_id` → `provinces.id` (lấy tên tỉnh/TP hiển thị cột Tỉnh/TP).\n"
     "► Trường của `banks`: code, name, short_name, international_business_name, business_address, logo (URL S3), bin, status, updated_by.\n"
     "► Trường của `bank_branches`: bank_id, name, province_id.\n"
     "► Ngoài HRM: khi MasterSetting category = 'use_crm' có content → mọi create/update Bank & BankBranch đồng bộ sang CRM (mate.bank / mate.bank.branch) và ghi bảng ModuleMapping."),

    ("6. Quy tắc cộng dồn / deduplicate",
     "► Không cộng dồn số liệu. Mỗi dòng = 1 bản ghi `banks`, phân trang server-side (paginate, mặc định limit = 10).\n"
     "► Chống trùng: `code` unique toàn bảng `banks`, `name` unique toàn bảng `banks` (bỏ qua chính bản ghi đang sửa).\n"
     "► Chống trùng chi nhánh: `bank_branches.name` unique TRONG PHẠM VI 1 bank_id (2 ngân hàng khác nhau được phép có chi nhánh cùng tên).\n"
     "► STT hiển thị = (page - 1) × limit + index + 1, tính lại theo từng trang."),

    ("7. Phân quyền cấp",
     "Không áp dụng — màn KHÔNG phân quyền theo cấp tổ chức và KHÔNG gắn middleware checkPermission.\n"
     "• Toàn bộ route nhóm `/v1/master-data/banks/*` chỉ gắn middleware `auth:api` (JWT).\n"
     "• Menu 'Danh mục ngân hàng' (components/subsystem-menu/master-data.js) khai báo không kèm permission key → hiện với mọi user đăng nhập.\n"
     "• Hệ quả: mọi user đăng nhập đều xem/tạo/sửa/xoá/khoá được. Test bảo mật chỉ kiểm tra tầng xác thực (401 khi thiếu/hết hạn token) — xem mục VII."),

    ("8. Cách tính các ô thống kê",
     "Không có khối thống kê (stats card) trên màn này. Các ô số/đếm duy nhất:\n"
     "► Cột 'Chi nhánh' = COUNT(bank_branches WHERE bank_id = banks.id) — lấy từ item.branches.length, hiển thị 0 nếu rỗng.\n"
     "► Cột 'STT' = (currentPage - 1) × pageSize + index + 1.\n"
     "► Dòng tổng dưới bảng = tổng số bản ghi khớp bộ lọc (`total` từ paginate), kèm khoảng from–to của trang hiện tại."),

    ("9. Ghi chú đọc bảng",
     "► Phân trang server-side; đổi số dòng/trang reset về trang 1.\n"
     "► Cột 'Cập nhật': dòng trên là thời điểm updated_at, dòng dưới là 'bởi <tên NV>' (banks.updated_by, tự gán theo user đăng nhập khi lưu); nếu chưa có thì không hiện dòng 'bởi'.\n"
     "► Cột 'Mã ngân hàng - Tên ngân hàng' gộp 3 thông tin: code - name (dòng 1), 'Tên viết tắt: …' (dòng 2) và 3 nút hành động Xem / Sửa / Xoá.\n"
     "► Trạng thái hiển thị dạng pill: xanh 'Hoạt động' (status = 1), xám/đỏ 'Khoá' (status = 2); nút ổ khoá bên cạnh để khoá / mở khoá.\n"
     "► Giá trị trống hiển thị dấu '—'.\n"
     "► Nút bị disable luôn kèm tooltip giải thích lý do (đang được sử dụng / đang bị khoá)."),
]

HAS_ROLE_SECTION = False
ROLE_TCS = []

SECTIONS = [
    # ---------------------------------------------------------------- I
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Truy cập màn hình danh mục ngân hàng từ menu", "P0",
         "User đã đăng nhập hệ thống HRM; DB có ít nhất 12 bản ghi trong bảng `banks`",
         "1. Đăng nhập HRM\n2. Vào phân hệ 'Danh mục chung'\n3. Click menu 'Danh mục ngân hàng'\n4. Quan sát toàn bộ layout",
         "User bất kỳ đã đăng nhập",
         "- URL chuyển sang /master-data/banks\n- Tiêu đề tab trình duyệt = 'Danh mục ngân hàng'\n"
         "- Hiển thị panel 'Bộ lọc danh mục ngân hàng' ở trên (đang thu gọn)\n"
         "- Hiển thị bảng 'Danh mục ngân hàng' bên dưới với nút 'Tạo mới' góc phải\n- Không có lỗi console",
         "Route /v1/master-data/banks chỉ yêu cầu auth:api — mọi user đăng nhập đều truy cập được"),

        ("002", "Kiểm tra đủ 8 cột của bảng danh sách", "P0",
         "Đang ở màn /master-data/banks, có ≥ 1 ngân hàng",
         "1. Quan sát dòng header của bảng\n2. Đối chiếu tên và thứ tự từng cột",
         "—",
         "Header hiển thị đúng thứ tự 8 cột:\n1. STT\n2. Logo\n3. Mã ngân hàng - Tên ngân hàng\n4. Tên giao dịch quốc tế\n"
         "5. Địa chỉ giao dịch\n6. Chi nhánh\n7. Cập nhật\n8. Trạng thái\n- Chỉ cột 'Cập nhật' có icon sắp xếp",
         "tableColumns: chỉ key updatedAt có sortable = true"),

        ("003", "Truy cập bằng URL cũ /human/banks", "P1",
         "User đã đăng nhập",
         "1. Gõ trực tiếp URL <host>/human/banks lên thanh địa chỉ\n2. Enter",
         "URL: /human/banks",
         "- Hệ thống tự chuyển hướng sang /master-data/banks\n- Màn danh mục ngân hàng load bình thường, dữ liệu đầy đủ",
         "config/route-redirects.js: { from: '/human/banks', to: '/master-data/banks' }"),

        ("004", "Trạng thái loading khi tải dữ liệu", "P2",
         "Mạng ở chế độ throttle (Slow 3G) trong DevTools",
         "1. Bật throttle Slow 3G\n2. Reload trang /master-data/banks\n3. Quan sát vùng bảng khi API chưa trả",
         "Network: Slow 3G",
         "- Bảng hiển thị trạng thái đang tải (loading) trong lúc chờ API\n- Sau khi API trả về, loading tắt và dữ liệu hiển thị đầy đủ\n- Không hiện đồng thời loading và dòng 'Không có dữ liệu'",
         "loading = true trước gọi API, finally set false"),

        ("005", "Hiển thị khi danh mục không có bản ghi nào", "P1",
         "Bảng `banks` rỗng HOẶC lọc với từ khoá chắc chắn không khớp",
         "1. Nhập ô tìm nhanh: 'zzzkhongtontai'\n2. Bấm Tìm kiếm\n3. Quan sát vùng bảng",
         "keyword = 'zzzkhongtontai'",
         "- Bảng hiển thị dòng 'Không có dữ liệu phù hợp bộ lọc.'\n- Tổng số bản ghi = 0\n- Không hiển thị dòng dữ liệu rác, không lỗi JS",
         "emptyText = 'Không có dữ liệu phù hợp bộ lọc.'"),

        ("006", "Mở rộng / thu gọn panel bộ lọc nâng cao", "P1",
         "Đang ở màn danh mục ngân hàng, panel lọc đang thu gọn (mặc định)",
         "1. Click nút mở rộng bộ lọc\n2. Quan sát các trường nâng cao\n3. Click lại để thu gọn",
         "—",
         "- Khi mở: hiện 3 trường 'Tên viết tắt', 'Tên giao dịch quốc tế', 'Trạng thái'\n"
         "- Khi thu gọn: 3 trường bị ẩn, ô tìm nhanh vẫn hiển thị\n- Giá trị đã nhập KHÔNG bị mất khi thu gọn rồi mở lại",
         "filterCollapsed mặc định = true; v-show nên DOM giữ nguyên giá trị"),
    ]),

    # ---------------------------------------------------------------- II
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm nhanh theo TÊN ngân hàng (khớp một phần)", "P0",
         "DB có: 'Ngân hàng TMCP Ngoại thương Việt Nam' (code VCB), 'Ngân hàng TMCP Công thương Việt Nam' (code CTG), 'Ngân hàng TMCP Á Châu' (code ACB)",
         "1. Nhập ô tìm nhanh: 'Ngoại thương'\n2. Bấm Tìm kiếm\n3. Quan sát kết quả",
         "keyword = 'Ngoại thương'",
         "- Chỉ hiển thị 'Ngân hàng TMCP Ngoại thương Việt Nam'\n- Tổng số bản ghi = 1\n- Trang hiện tại reset về 1",
         "banks.name LIKE %keyword%"),

        ("002", "Tìm nhanh theo MÃ ngân hàng", "P0",
         "DB có ngân hàng code = 'VCB', name = 'Ngân hàng TMCP Ngoại thương Việt Nam'",
         "1. Nhập ô tìm nhanh: 'VCB'\n2. Bấm Tìm kiếm",
         "keyword = 'VCB'",
         "- Hiển thị đúng ngân hàng có mã VCB\n- Cột 'Mã ngân hàng - Tên ngân hàng' hiện 'VCB - Ngân hàng TMCP Ngoại thương Việt Nam'",
         "Điều kiện OR: banks.code LIKE %keyword%"),

        ("003", "Tìm nhanh KHÔNG áp dụng cho tên viết tắt", "P1",
         "DB có ngân hàng: code = 'VCB', name = 'Ngân hàng TMCP Ngoại thương Việt Nam', short_name = 'Vietcombank'",
         "1. Nhập ô tìm nhanh: 'Vietcombank'\n2. Bấm Tìm kiếm\n3. Quan sát kết quả",
         "keyword = 'Vietcombank'",
         "- Không trả về bản ghi nào (vì tìm nhanh chỉ quét name và code)\n- Muốn tìm theo tên viết tắt phải dùng ô 'Tên viết tắt' ở bộ lọc nâng cao",
         "getBanks(): keyword chỉ so name OR code, KHÔNG so short_name"),

        ("004", "Lọc theo Tên viết tắt (bộ lọc nâng cao)", "P0",
         "DB có 3 ngân hàng short_name lần lượt: 'Vietcombank', 'Vietinbank', 'ACB'",
         "1. Mở bộ lọc nâng cao\n2. Nhập 'Viet' vào ô 'Tên viết tắt'\n3. Bấm Tìm kiếm",
         "short_name = 'Viet'",
         "- Trả về đúng 2 bản ghi: Vietcombank và Vietinbank\n- Không trả về ACB\n- Cột 'Tên viết tắt' của mỗi dòng chứa chuỗi 'Viet'",
         "banks.short_name LIKE %short_name%"),

        ("005", "Lọc theo Tên giao dịch quốc tế", "P1",
         "DB có ngân hàng international_business_name = 'Joint Stock Commercial Bank for Foreign Trade of Vietnam'",
         "1. Mở bộ lọc nâng cao\n2. Nhập 'Foreign Trade' vào 'Tên giao dịch quốc tế'\n3. Bấm Tìm kiếm",
         "international_business_name = 'Foreign Trade'",
         "- Chỉ hiển thị các ngân hàng có tên giao dịch quốc tế chứa 'Foreign Trade'\n- Cột 'Tên giao dịch quốc tế' hiển thị đúng giá trị đầy đủ",
         "banks.international_business_name LIKE %…%"),

        ("006", "Lọc theo trạng thái = Hoạt động", "P0",
         "DB có 5 ngân hàng status = 1 và 2 ngân hàng status = 2",
         "1. Mở bộ lọc nâng cao\n2. Chọn Trạng thái = 'Hoạt động'\n3. Bấm Tìm kiếm",
         "status = 1",
         "- Trả về đúng 5 bản ghi\n- Mọi dòng đều hiện pill xanh 'Hoạt động'\n- Không có dòng nào hiện 'Khoá'",
         "banks.status = 1 (Bank::STATUS_ACTIVE)"),

        ("007", "Lọc theo trạng thái = Khoá", "P0",
         "DB có 5 ngân hàng status = 1 và 2 ngân hàng status = 2",
         "1. Mở bộ lọc nâng cao\n2. Chọn Trạng thái = 'Khoá'\n3. Bấm Tìm kiếm",
         "status = 2",
         "- Trả về đúng 2 bản ghi\n- Mọi dòng hiện pill 'Khoá' + icon ổ khoá\n- Nút Sửa và Xoá của các dòng này bị disable",
         "banks.status = 2 (Bank::STATUS_INACTIVE)"),

        ("008", "Kết hợp Tên viết tắt + Trạng thái", "P0",
         "DB có: 'Vietcombank' status 1, 'Vietinbank' status 2",
         "1. Mở bộ lọc nâng cao\n2. Nhập Tên viết tắt = 'Viet'\n3. Chọn Trạng thái = 'Hoạt động'\n4. Bấm Tìm kiếm",
         "short_name = 'Viet', status = 1",
         "- Chỉ trả về Vietcombank (thoả CẢ 2 điều kiện)\n- Không trả về Vietinbank",
         "Các điều kiện lọc kết hợp bằng AND"),

        ("009", "Kết hợp tìm nhanh (keyword) + Trạng thái", "P0",
         "DB có: 'Ngân hàng TMCP Á Châu' (ACB) status = 2; 'Ngân hàng TMCP Ngoại thương' (VCB) status = 1",
         "1. Nhập ô tìm nhanh: 'Ngân hàng'\n2. Mở bộ lọc nâng cao, chọn Trạng thái = 'Hoạt động'\n3. Bấm Tìm kiếm\n4. Đối chiếu từng dòng trả về",
         "keyword = 'Ngân hàng', status = 1",
         "- Mọi dòng trả về phải thoả ĐỒNG THỜI: (name hoặc code chứa 'Ngân hàng') VÀ status = 1\n"
         "- Ngân hàng ACB (status = 2) KHÔNG được xuất hiện\n- Tổng số bản ghi khớp đúng số đếm thủ công trong DB",
         "CHÚ Ý: query dùng where(name like)->orWhere(code like) KHÔNG bọc nhóm — nếu kết quả lọt bản ghi status = 2 thì là lỗi gộp điều kiện, ghi nhận Failed"),

        ("010", "Kết hợp tìm nhanh + Tên viết tắt", "P1",
         "DB có: code 'VCB' / short_name 'Vietcombank'; code 'CTG' / short_name 'Vietinbank'",
         "1. Nhập ô tìm nhanh: 'VCB'\n2. Nhập Tên viết tắt: 'Vietinbank'\n3. Bấm Tìm kiếm",
         "keyword = 'VCB', short_name = 'Vietinbank'",
         "- Không có bản ghi nào thoả cả 2 điều kiện → bảng hiện 'Không có dữ liệu phù hợp bộ lọc.'\n- Tổng số bản ghi = 0",
         "AND giữa keyword và short_name — kiểm tra điều kiện OR của keyword không phá nhóm AND"),

        ("011", "Nút Làm mới (reset) bộ lọc", "P0",
         "Đã nhập: tìm nhanh 'VCB', Tên viết tắt 'Viet', Trạng thái 'Khoá'; đang ở trang 3",
         "1. Bấm nút Làm mới / Đặt lại trên panel lọc\n2. Quan sát các ô lọc và bảng dữ liệu",
         "—",
         "- Cả 3 ô lọc nâng cao + ô tìm nhanh về rỗng\n- Trạng thái về placeholder 'Chọn trạng thái'\n"
         "- Danh sách tải lại toàn bộ dữ liệu\n- Về trang 1, sắp xếp về mặc định (id giảm dần)",
         "handleReset(): filters = {...initialStateForm}, currentPage = 1"),

        ("012", "Chọn rồi xoá (clear) giá trị Trạng thái", "P1",
         "Đang lọc Trạng thái = 'Khoá', kết quả 2 bản ghi",
         "1. Bấm dấu x xoá lựa chọn ở ô Trạng thái\n2. Quan sát danh sách",
         "status: 2 → rỗng",
         "- Danh sách tự tải lại, hiển thị cả ngân hàng Hoạt động và Khoá\n- Tổng số bản ghi tăng đúng bằng tổng toàn danh mục",
         "allowClear = true; watch filters deep → tự loadData"),

        ("013", "Bộ lọc reset về trang 1 khi đang ở trang sau", "P1",
         "Danh mục có ≥ 25 ngân hàng; đang đứng ở trang 3",
         "1. Đang ở trang 3\n2. Nhập Tên viết tắt = 'Viet'\n3. Bấm Tìm kiếm",
         "short_name = 'Viet', trang trước đó = 3",
         "- Kết quả hiển thị từ trang 1\n- Chỉ báo phân trang hiển thị trang 1\n- Không xảy ra trường hợp bảng trắng do trang vượt quá số trang mới",
         "handleSearch(): currentPage = 1 trước khi loadData"),

        ("014", "Tìm kiếm chuỗi có khoảng trắng đầu/cuối", "P2",
         "DB có ngân hàng name = 'Ngân hàng TMCP Á Châu'",
         "1. Nhập ô tìm nhanh: '  Á Châu  ' (có dấu cách 2 đầu)\n2. Bấm Tìm kiếm",
         "keyword = '  Á Châu  '",
         "- Ghi nhận thực tế: nếu trả về đúng 'Ngân hàng TMCP Á Châu' → Passed\n- Nếu trả 0 bản ghi do không trim → ghi Failed kèm mô tả (cần trim keyword trước khi LIKE)",
         "Backend LIKE %keyword% không trim — cần xác nhận hành vi mong muốn"),

        ("015", "Tìm kiếm ký tự đặc biệt % và _", "P2",
         "DB có ≥ 10 ngân hàng, không ngân hàng nào chứa ký tự '%' trong tên",
         "1. Nhập ô tìm nhanh: '%'\n2. Bấm Tìm kiếm\n3. Quan sát kết quả",
         "keyword = '%'",
         "- Kết quả KHÔNG được trả về toàn bộ danh mục (ký tự wildcard phải được escape)\n- Kỳ vọng: 0 bản ghi\n- Không lỗi 500",
         "Bảo vệ LIKE wildcard injection"),
    ]),

    # ---------------------------------------------------------------- III
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "Đang ở màn /master-data/banks",
         "1. Quan sát vùng phía trên bảng dữ liệu",
         "—",
         "- Màn KHÔNG có khối thẻ thống kê (stats card)\n- Chỉ có panel bộ lọc và bảng dữ liệu\n"
         "- Số liệu duy nhất là tổng số bản ghi + khoảng from–to hiển thị ở vùng phân trang",
         "Section giữ lại theo chuẩn tài liệu — feature không có stats"),
    ]),

    # ---------------------------------------------------------------- IV
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Sắp xếp mặc định khi vào màn", "P0",
         "DB có ≥ 12 ngân hàng, id tăng dần theo thứ tự tạo",
         "1. Vào /master-data/banks lần đầu (không đụng vào cột sắp xếp)\n2. Đối chiếu thứ tự bản ghi với id trong DB",
         "sortBy = id, sortDesc = desc",
         "- Bản ghi có id lớn nhất (mới tạo nhất) đứng đầu danh sách\n- Thứ tự giảm dần theo id",
         "initialStateForm: sort_by = 'id', sort_desc = true"),

        ("002", "Sắp xếp theo cột Cập nhật — tăng dần", "P0",
         "DB có ≥ 5 ngân hàng với updated_at khác nhau",
         "1. Click vào tiêu đề cột 'Cập nhật'\n2. Quan sát mũi tên sắp xếp và thứ tự dòng",
         "sortBy = updated_at, sortDesc = asc",
         "- Danh sách sắp xếp tăng dần theo thời điểm cập nhật (cũ nhất trên cùng)\n- Icon mũi tên trên cột 'Cập nhật' đổi trạng thái\n- Về trang 1",
         "handleSort map field 'updatedAt' → 'updated_at'"),

        ("003", "Sắp xếp theo cột Cập nhật — giảm dần", "P0",
         "Đang sắp xếp cột 'Cập nhật' tăng dần",
         "1. Click lần 2 vào tiêu đề cột 'Cập nhật'\n2. Quan sát thứ tự dòng",
         "sortBy = updated_at, sortDesc = desc",
         "- Bản ghi cập nhật gần nhất đứng đầu\n- Thứ tự đảo ngược hoàn toàn so với lần click trước",
         "sort_desc = (direction === 'desc')"),

        ("004", "Các cột khác không cho sắp xếp", "P2",
         "Đang ở màn danh mục ngân hàng",
         "1. Click lần lượt vào tiêu đề: STT, Logo, Mã ngân hàng - Tên ngân hàng, Chi nhánh, Trạng thái",
         "—",
         "- Không cột nào trong số đó đổi thứ tự dữ liệu\n- Không hiển thị icon sắp xếp trên các cột này\n- Không phát sinh request API thừa",
         "Chỉ cột updatedAt khai báo sortable"),

        ("005", "Phân trang — chuyển sang trang kế tiếp", "P0",
         "DB có 25 ngân hàng; số dòng/trang = 10",
         "1. Ở trang 1, ghi lại tên ngân hàng dòng đầu\n2. Bấm sang trang 2\n3. Quan sát STT và dữ liệu",
         "total = 25, pageSize = 10",
         "- STT dòng đầu trang 2 = 11\n- Dữ liệu khác hoàn toàn trang 1\n- Vùng phân trang hiển thị đúng trang hiện tại = 2 và tổng 3 trang",
         "STT = (currentPage - 1) × pageSize + index + 1"),

        ("006", "Phân trang — trang cuối hiển thị đúng số dòng lẻ", "P1",
         "DB có đúng 25 ngân hàng; số dòng/trang = 10",
         "1. Chuyển tới trang 3 (trang cuối)\n2. Đếm số dòng hiển thị",
         "total = 25, page = 3",
         "- Trang 3 hiển thị đúng 5 dòng\n- STT chạy từ 21 đến 25\n- Nút sang trang kế bị vô hiệu",
         "lastPage = ceil(25/10) = 3"),

        ("007", "Đổi số dòng trên mỗi trang", "P0",
         "DB có ≥ 60 ngân hàng; đang ở trang 3, 10 dòng/trang",
         "1. Đổi số dòng/trang sang 50\n2. Quan sát danh sách và chỉ số trang",
         "pageSize: 10 → 50",
         "- Danh sách hiển thị tối đa 50 dòng\n- Tự động quay về trang 1\n- STT dòng đầu = 1\n- Tổng số bản ghi không đổi",
         "handlePageSizeChange(): currentPage = 1"),

        ("008", "Hiển thị logo ngân hàng", "P1",
         "Có ngân hàng A đã upload logo, ngân hàng B chưa có logo (logo = null)",
         "1. Tìm ngân hàng A và B trên danh sách\n2. Quan sát cột Logo",
         "A: có logo; B: logo = null",
         "- Dòng A hiển thị ảnh logo, chiều cao 32px, không méo hình (object-fit contain)\n- Dòng B hiển thị dấu '—' màu xám\n- Ảnh lỗi không làm vỡ layout dòng",
         "cell-logo: v-if item.logo, ngược lại hiện '—'"),

        ("009", "Hiển thị các trường rỗng bằng dấu '—'", "P1",
         "Có ngân hàng với international_business_name = null, business_address = null, short_name = null",
         "1. Tìm ngân hàng đó trên danh sách\n2. Quan sát các cột tương ứng",
         "3 trường đều null",
         "- Cột 'Tên giao dịch quốc tế' hiển thị '—'\n- Cột 'Địa chỉ giao dịch' hiển thị '—'\n"
         "- Dòng phụ 'Tên viết tắt:' hiển thị '—'\n- Không hiển thị chữ 'null' hay 'undefined'",
         "FE fallback: item.field || '—'"),

        ("010", "Hiển thị số lượng chi nhánh", "P0",
         "Ngân hàng A có 3 chi nhánh; ngân hàng B có 0 chi nhánh",
         "1. Quan sát cột 'Chi nhánh' của A và B",
         "A: 3 branches; B: 0 branch",
         "- Dòng A hiển thị số 3, dạng link màu xanh, có thể click\n- Dòng B hiển thị số 0\n- Tooltip khi rê chuột: 'Xem danh sách chi nhánh'",
         "Số = item.branches.length (eager load with('branches'))"),

        ("011", "Hiển thị người và thời điểm cập nhật", "P1",
         "Ngân hàng A vừa được user 'Nguyễn Văn A' sửa lúc 10:30 hôm nay",
         "1. Tìm ngân hàng A\n2. Quan sát cột 'Cập nhật'",
         "updated_by = id của Nguyễn Văn A",
         "- Dòng trên: thời điểm cập nhật đúng định dạng hệ thống\n- Dòng dưới: 'bởi Nguyễn Văn A' (tên in đậm)\n"
         "- Với bản ghi chưa có người cập nhật: chỉ hiện thời gian, không hiện dòng 'bởi'",
         "BaseModel tự gán updated_by = Auth::user()->id khi saving"),

        ("012", "Hiển thị pill trạng thái Hoạt động / Khoá", "P0",
         "Ngân hàng A status = 1, ngân hàng B status = 2",
         "1. Quan sát cột 'Trạng thái' của A và B",
         "A: status 1; B: status 2",
         "- A: pill xanh, icon dấu tích, chữ 'Hoạt động', nút bên cạnh là icon ổ khoá (tooltip 'Khoá ngân hàng')\n"
         "- B: pill khoá, icon ổ khoá, chữ 'Khoá', nút bên cạnh là icon mở khoá (tooltip 'Mở khoá ngân hàng')",
         "renderBankStatus(): status = 2 → 'Khoá', còn lại 'Hoạt động'"),

        ("013", "Trạng thái nút hành động của ngân hàng đang bị Khoá", "P0",
         "Ngân hàng B: status = 2, can_delete = true",
         "1. Tìm ngân hàng B\n2. Rê chuột lần lượt vào 3 nút Xem / Sửa / Xoá",
         "B: status = 2",
         "- Nút Xem: bật, click được\n- Nút Sửa: bị mờ, không click được\n"
         "- Nút Xoá: bị mờ, tooltip 'Không thể xóa ngân hàng đang bị khoá'",
         "Điều kiện FE: :disabled = item.status === 2"),

        ("014", "Trạng thái nút Xoá khi ngân hàng đang được sử dụng", "P0",
         "Ngân hàng C: status = 1 và đang được gán ở tài khoản ngân hàng của 1 nhân viên đang làm việc",
         "1. Tìm ngân hàng C\n2. Rê chuột vào nút Xoá",
         "C: can_delete = false",
         "- Nút Xoá bị mờ, không click được\n- Tooltip: 'Không thể xóa bản ghi, ngân hàng đang được sử dụng trên hệ thống'\n- Nút Sửa vẫn bật",
         "can_delete = Bank::canDelete() trả về từ API"),
    ]),

    # ---------------------------------------------------------------- V
    ("V", "CHỨC NĂNG CHÍNH (CRUD / ACTION)", [
        # --- Tạo mới
        ("001", "Mở modal Tạo mới ngân hàng", "P0",
         "Đang ở màn /master-data/banks",
         "1. Bấm nút 'Tạo mới' góc phải bảng\n2. Quan sát modal",
         "—",
         "- Mở modal tiêu đề 'Tạo mới ngân hàng'\n- Toàn bộ ô nhập rỗng\n"
         "- Hiện khối 'Gợi ý' + nút 'Tra cứu'\n- Footer có 3 nút: Lưu, Lưu & Tiếp tục, Đóng\n- Không hiện trường Trạng thái",
         "id = null → tiêu đề 'Tạo mới'; isShow = false → hiện khối Gợi ý"),

        ("002", "Tạo mới ngân hàng với đủ trường bắt buộc", "P0",
         "Chưa tồn tại ngân hàng có code = 'TEST01' và name = 'Ngân hàng Kiểm Thử 01'",
         "1. Bấm 'Tạo mới'\n2. Nhập Mã ngân hàng = 'TEST01'\n3. Nhập Tên ngân hàng = 'Ngân hàng Kiểm Thử 01'\n"
         "4. Nhập Tên viết tắt = 'NHKT01'\n5. Bấm Lưu",
         "code = TEST01, name = Ngân hàng Kiểm Thử 01, short_name = NHKT01",
         "- Toast xanh 'Đã lưu thành công!'\n- Modal tự đóng\n- Danh sách tải lại, bản ghi mới đứng đầu (sắp xếp id desc)\n"
         "- Trạng thái mặc định 'Hoạt động'\n- Cột Chi nhánh = 0",
         "POST /v1/master-data/banks; validate name, code, short_name required"),

        ("003", "Tạo mới với đầy đủ mọi trường (kể cả không bắt buộc)", "P1",
         "Chưa tồn tại code = 'TEST02'",
         "1. Bấm 'Tạo mới'\n2. Nhập Mã = 'TEST02', Tên = 'Ngân hàng Kiểm Thử 02', Tên viết tắt = 'NHKT02'\n"
         "3. Nhập Tên giao dịch quốc tế = 'Test Bank 02 JSC'\n4. Nhập Địa chỉ giao dịch = '123 Trần Duy Hưng, Hà Nội'\n5. Bấm Lưu",
         "5 trường như bước thực hiện",
         "- Lưu thành công\n- Trên danh sách: cột 'Tên giao dịch quốc tế' = 'Test Bank 02 JSC', 'Địa chỉ giao dịch' = '123 Trần Duy Hưng, Hà Nội'\n"
         "- Mở lại modal Xem thấy đúng toàn bộ giá trị đã nhập",
         "international_business_name và business_address là trường tuỳ chọn"),

        ("004", "Nút 'Lưu & Tiếp tục' khi tạo mới", "P0",
         "Chưa tồn tại code = 'TEST03' và 'TEST04'",
         "1. Bấm 'Tạo mới'\n2. Nhập Mã 'TEST03', Tên 'NH Kiểm Thử 03', Tên viết tắt 'NHKT03'\n"
         "3. Bấm 'Lưu & Tiếp tục'\n4. Quan sát modal\n5. Nhập tiếp Mã 'TEST04', Tên 'NH Kiểm Thử 04', Tên viết tắt 'NHKT04' rồi bấm Lưu",
         "2 bản ghi liên tiếp",
         "- Sau bước 3: toast 'Đã lưu thành công!', modal KHÔNG đóng, toàn bộ ô nhập được xoá trắng\n"
         "- Sau bước 5: modal đóng\n- Danh sách có đủ cả TEST03 và TEST04",
         "isContinue = true → resetModal() thay vì hide()"),

        ("005", "Nút 'Lưu & Tiếp tục' không xuất hiện khi Sửa", "P1",
         "Có ít nhất 1 ngân hàng đang Hoạt động",
         "1. Bấm nút Sửa trên 1 dòng bất kỳ\n2. Quan sát footer modal",
         "id != null",
         "- Footer chỉ có 2 nút: 'Lưu' và 'Đóng'\n- KHÔNG có nút 'Lưu & Tiếp tục'",
         "v-if = '!id && !isShow'"),

        # --- Sửa
        ("006", "Mở modal Sửa và kiểm tra dữ liệu nạp sẵn", "P0",
         "Ngân hàng 'TEST01' (code TEST01, short_name NHKT01, có logo) đang Hoạt động",
         "1. Bấm nút Sửa trên dòng TEST01\n2. Quan sát toàn bộ trường trong modal",
         "id = id của TEST01",
         "- Tiêu đề modal = 'Sửa ngân hàng'\n- Mã, Tên, Tên viết tắt, Tên giao dịch quốc tế, Địa chỉ điền sẵn đúng dữ liệu hiện tại\n"
         "- Vùng logo hiển thị ảnh hiện tại\n- Các ô đều cho phép chỉnh sửa",
         "resetModal(): GET /v1/master-data/banks/{id}"),

        ("007", "Sửa và lưu thành công", "P0",
         "Ngân hàng 'TEST01' đang có tên viết tắt = 'NHKT01'",
         "1. Bấm Sửa dòng TEST01\n2. Đổi Tên viết tắt thành 'NHKT01-NEW'\n3. Bấm Lưu\n4. Quan sát dòng trên danh sách",
         "short_name: NHKT01 → NHKT01-NEW",
         "- Toast 'Đã lưu thành công!', modal đóng\n- Dòng TEST01 hiển thị 'Tên viết tắt: NHKT01-NEW'\n"
         "- Cột 'Cập nhật' đổi thành thời điểm hiện tại, dòng 'bởi <user đang đăng nhập>'\n- Không tạo thêm bản ghi mới",
         "createBank(): có id → find + fill + save; BaseModel gán updated_by"),

        ("008", "Sửa giữ nguyên mã/tên của chính nó (không báo trùng)", "P0",
         "Ngân hàng 'TEST01' đã tồn tại",
         "1. Bấm Sửa dòng TEST01\n2. Giữ nguyên Mã và Tên\n3. Chỉ đổi Địa chỉ giao dịch\n4. Bấm Lưu",
         "code, name giữ nguyên; business_address đổi",
         "- Lưu thành công, KHÔNG báo lỗi 'Mã ngân hàng này đã tồn tại' hay 'Tên ngân hàng này đã tồn tại'\n- Địa chỉ mới hiển thị trên danh sách",
         "Rule unique có ignore chính id đang sửa"),

        # --- Xem
        ("009", "Xem chi tiết — mọi trường ở chế độ chỉ đọc", "P0",
         "Ngân hàng 'TEST02' có đầy đủ thông tin và logo",
         "1. Bấm nút Xem (icon con mắt) trên dòng TEST02\n2. Thử click và gõ vào từng ô\n3. Quan sát footer",
         "isShow = true",
         "- Tiêu đề 'Xem chi tiết ngân hàng'\n- Tất cả ô ở trạng thái disabled, không gõ được\n"
         "- Hiện thêm ô 'Trạng thái' với giá trị 'Hoạt động'\n- Ẩn khối 'Gợi ý' + nút Tra cứu, ẩn nút Tải ảnh lên\n"
         "- Footer chỉ còn nút 'Đóng'",
         "isShow = true → :disabled, ẩn nút Lưu và khối tra cứu"),

        ("010", "Xem chi tiết ngân hàng đang Khoá", "P1",
         "Ngân hàng 'TEST05' đang ở trạng thái Khoá (status = 2)",
         "1. Bấm nút Xem trên dòng TEST05\n2. Quan sát ô Trạng thái",
         "status = 2",
         "- Ô 'Trạng thái' hiển thị 'Khoá'\n- Vẫn xem được đầy đủ thông tin\n- Không có nút Lưu",
         "data.status === 2 ? 'Khoá' : 'Hoạt động'"),

        ("011", "Đóng modal không lưu thay đổi", "P1",
         "Ngân hàng 'TEST02' có Tên viết tắt = 'NHKT02'",
         "1. Bấm Sửa dòng TEST02\n2. Đổi Tên viết tắt thành 'ABC-XYZ'\n3. Bấm nút 'Đóng'\n4. Quan sát danh sách\n5. Mở lại modal Sửa",
         "Thay đổi không bấm Lưu",
         "- Modal đóng, không có toast thành công\n- Dòng TEST02 vẫn hiển thị 'NHKT02'\n"
         "- Mở lại modal Sửa: giá trị nạp lại đúng 'NHKT02', không còn 'ABC-XYZ'",
         "resetModal() gọi lại API mỗi lần @show"),

        # --- Tra cứu VietQR
        ("012", "Tra cứu thông tin ngân hàng từ nguồn ngoài", "P1",
         "Đang mở modal 'Tạo mới ngân hàng'; máy có kết nối Internet ra api.vietqr.io",
         "1. Nhập ô 'Gợi ý' = 'Vietcombank'\n2. Bấm nút 'Tra cứu'\n3. Quan sát bảng kết quả",
         "keyword tra cứu = 'Vietcombank'",
         "- Mở modal 'Thông tin ngân hàng' với dòng phụ 'Click vào một dòng để điền nhanh thông tin'\n"
         "- Bảng 5 cột: STT, Logo, Tên ngân hàng, Mã ngân hàng, Tên viết tắt\n"
         "- Chỉ hiển thị các ngân hàng khớp 'Vietcombank'",
         "GET https://api.vietqr.io/v2/banks + lọc client-side"),

        ("013", "Tra cứu không dấu vẫn khớp tên có dấu", "P1",
         "Đang mở modal Tra cứu, kết nối Internet bình thường",
         "1. Ở modal Tạo mới, nhập Gợi ý = 'ngoai thuong' (không dấu)\n2. Bấm 'Tra cứu'",
         "keyword = 'ngoai thuong'",
         "- Bảng vẫn trả về 'Ngân hàng TMCP Ngoại Thương Việt Nam'\n- Không phân biệt hoa/thường và dấu tiếng Việt",
         "normalizeString(): NFD + bỏ dấu + đ→d + lowercase"),

        ("014", "Chọn 1 dòng tra cứu để điền tự động", "P0",
         "Modal 'Thông tin ngân hàng' đang mở với ≥ 1 kết quả",
         "1. Click vào dòng 'Ngân hàng TMCP Ngoại Thương Việt Nam'\n2. Quan sát modal Tạo mới phía sau",
         "Dòng chọn: VCB / Vietcombank",
         "- Modal tra cứu tự đóng\n- Ô Mã ngân hàng, Tên ngân hàng, Tên viết tắt được điền tự động theo dòng đã chọn\n"
         "- Vùng logo hiển thị logo lấy từ nguồn tra cứu\n- Các trường khác giữ nguyên giá trị đang nhập",
         "selectBankInfo(): gán name, code, logo, short_name"),

        ("015", "Tra cứu khi không gọi được dịch vụ ngoài", "P2",
         "Chặn domain api.vietqr.io (tắt mạng hoặc chặn qua DevTools)",
         "1. Mở modal Tạo mới\n2. Bấm 'Tra cứu'\n3. Quan sát bảng kết quả",
         "Network: block api.vietqr.io",
         "- Bảng hiển thị 'Chưa có dữ liệu', không treo loading vĩnh viễn\n"
         "- Modal Tạo mới vẫn nhập tay và lưu bình thường\n- Không văng lỗi trắng màn hình",
         "catch lỗi axios, finally tắt isLoading"),

        # --- Logo
        ("016", "Upload logo hợp lệ", "P0",
         "File ảnh logo.png dung lượng 200KB",
         "1. Mở modal Tạo mới\n2. Bấm 'Tải ảnh lên', chọn logo.png\n3. Quan sát vùng xem trước\n"
         "4. Nhập đủ Mã / Tên / Tên viết tắt\n5. Bấm Lưu",
         "File: logo.png (200KB)",
         "- Ngay sau chọn file: vùng xem trước hiển thị ảnh, hiện thêm nút 'Xóa ảnh'\n"
         "- Sau khi Lưu: cột Logo trên danh sách hiển thị đúng ảnh vừa tải\n- Không có thông báo lỗi",
         "IMAGE = jpg/jpeg/png; upload S3 rồi gán data.logo"),

        ("017", "Upload file sai định dạng", "P0",
         "File tailieu.pdf",
         "1. Mở modal Tạo mới\n2. Bấm 'Tải ảnh lên', chọn tailieu.pdf (đổi bộ lọc file thành All files)\n3. Quan sát",
         "File: tailieu.pdf",
         "- Hiển thị lỗi đỏ dưới vùng logo: 'File không hợp lệ'\n- Vùng xem trước quay về 'Chưa có logo'\n- Không gửi request upload",
         "Chỉ chấp nhận jpg, jpeg, png"),

        ("018", "Upload ảnh vượt quá 5MB", "P0",
         "File anh_lon.jpg dung lượng 8MB",
         "1. Mở modal Tạo mới\n2. Bấm 'Tải ảnh lên', chọn anh_lon.jpg\n3. Quan sát",
         "File: anh_lon.jpg (8MB)",
         "- Hiển thị lỗi đỏ: 'Dung lượng tối đa: 5MB'\n- Vùng xem trước không hiển thị ảnh\n- Không gửi request upload lên S3",
         "Điều kiện: file.size / 1024 / 1024 <= 5"),

        ("019", "Xoá ảnh logo đã chọn", "P1",
         "Đang mở modal Tạo mới, đã chọn 1 ảnh logo hợp lệ",
         "1. Bấm nút 'Xóa ảnh'\n2. Quan sát vùng logo",
         "—",
         "- Ảnh xem trước biến mất, hiện chữ 'Chưa có logo'\n- Nút 'Xóa ảnh' ẩn đi\n"
         "- Modal KHÔNG bị đóng và form KHÔNG bị submit\n- Bấm Lưu → bản ghi lưu không có logo",
         "clearLogo() có preventDefault chống submit ngoài ý muốn"),

        ("020", "Tooltip hướng dẫn định dạng logo", "P2",
         "Đang mở modal Tạo mới",
         "1. Rê chuột vào icon (i) cạnh nhãn 'Logo'",
         "—",
         "- Hiện tooltip: 'Các loại ảnh có thể tải lên (tối đa 5Mb): .jpg, .jpeg, .png'",
         "v-b-tooltip.hover.rightbottom"),

        # --- Khoá / Mở khoá
        ("021", "Khoá một ngân hàng đang hoạt động", "P0",
         "Ngân hàng 'TEST01' đang status = 1",
         "1. Bấm nút ổ khoá ở cột Trạng thái dòng TEST01\n2. Đọc nội dung hộp xác nhận\n3. Bấm 'Khoá'",
         "id = TEST01, status 1 → 2",
         "- Hộp xác nhận tiêu đề 'Xác nhận khoá', nội dung \"Bạn có chắc muốn khoá ngân hàng 'TEST01'?\", nút xác nhận ghi 'Khoá'\n"
         "- Sau xác nhận: toast 'Khoá thành công'\n- Dòng đổi sang pill 'Khoá'\n- Nút Sửa và Xoá của dòng chuyển sang mờ",
         "PUT /v1/master-data/banks/{id}/lock → status = 2"),

        ("022", "Mở khoá một ngân hàng đang bị khoá", "P0",
         "Ngân hàng 'TEST01' đang status = 2",
         "1. Bấm nút mở khoá dòng TEST01\n2. Đọc hộp xác nhận\n3. Bấm 'Mở khoá'",
         "id = TEST01, status 2 → 1",
         "- Hộp xác nhận tiêu đề 'Xác nhận mở khoá', nút xác nhận ghi 'Mở khoá'\n"
         "- Toast 'Mở khoá thành công'\n- Dòng trở về pill xanh 'Hoạt động'\n- Nút Sửa hoạt động trở lại",
         "PUT /v1/master-data/banks/{id}/unlock → status = 1"),

        ("023", "Huỷ thao tác khoá", "P1",
         "Ngân hàng 'TEST02' đang status = 1",
         "1. Bấm nút ổ khoá dòng TEST02\n2. Bấm 'Hủy' trên hộp xác nhận\n3. Quan sát dòng dữ liệu",
         "—",
         "- Hộp xác nhận đóng\n- Trạng thái vẫn là 'Hoạt động'\n- Không có toast\n- Không phát sinh request PUT",
         "Chỉ gọi API khi bấm nút xác nhận"),

        ("024", "Ngân hàng bị khoá không xuất hiện ở nơi chọn ngân hàng", "P0",
         "Ngân hàng 'TEST01' vừa bị Khoá; có 1 hồ sơ nhân viên đang mở form cập nhật thông tin",
         "1. Khoá ngân hàng TEST01\n2. Vào Hồ sơ nhân viên → mục Tài khoản ngân hàng\n3. Mở dropdown chọn ngân hàng\n4. Tìm 'TEST01'",
         "TEST01: status = 2",
         "- Danh sách chọn KHÔNG còn ngân hàng TEST01\n- Các ngân hàng status = 1 vẫn hiển thị đầy đủ\n"
         "- Hồ sơ đã gán TEST01 từ trước vẫn giữ nguyên dữ liệu cũ, không bị xoá",
         "FE tiêu thụ danh mục lọc bank.status == 1"),

        # --- Xoá
        ("025", "Xoá ngân hàng chưa được sử dụng", "P0",
         "Ngân hàng 'TEST04': status = 1, không chi nhánh, không nhân viên nào dùng, không có trong company_accounts",
         "1. Bấm nút Xoá (thùng rác) dòng TEST04\n2. Đọc hộp xác nhận\n3. Bấm 'Xóa'",
         "id = TEST04, can_delete = true",
         "- Hộp xác nhận: \"Bạn có chắc muốn xóa ngân hàng 'TEST04'?\"\n- Toast 'Xoá ngân hàng thành công'\n"
         "- Dòng biến mất khỏi danh sách, tổng số bản ghi giảm 1\n- Tìm lại theo mã TEST04 → không còn kết quả",
         "DELETE /v1/master-data/banks/{id} sau khi qua canDelete()"),

        ("026", "Xoá ngân hàng kéo theo xoá toàn bộ chi nhánh", "P0",
         "Ngân hàng 'TEST03' có 3 chi nhánh, không nhân viên nào dùng ngân hàng và chi nhánh này",
         "1. Ghi nhận cột Chi nhánh của TEST03 = 3\n2. Bấm Xoá, xác nhận 'Xóa'\n3. Kiểm tra bảng bank_branches theo bank_id của TEST03",
         "TEST03 + 3 bank_branches",
         "- Xoá thành công, dòng TEST03 biến mất\n- 3 bản ghi chi nhánh cũng bị xoá khỏi bank_branches (không còn bản ghi mồ côi)",
         "deleteBank(): $bank->branches()->delete() trước khi xoá bank"),

        ("027", "Huỷ thao tác xoá", "P1",
         "Ngân hàng 'TEST04' đang hiển thị trên danh sách",
         "1. Bấm nút Xoá dòng TEST04\n2. Bấm 'Hủy'",
         "—",
         "- Hộp xác nhận đóng\n- Dòng TEST04 vẫn còn trên danh sách\n- Không phát sinh request DELETE",
         "BaseConfirmModal chỉ emit event khi bấm nút xác nhận"),

        # --- Chi nhánh
        ("028", "Mở danh sách chi nhánh từ cột 'Chi nhánh'", "P0",
         "Ngân hàng 'TEST02' có 2 chi nhánh: 'CN Hà Nội' (Hà Nội), 'CN Đà Nẵng' (Đà Nẵng)",
         "1. Click vào con số ở cột 'Chi nhánh' dòng TEST02\n2. Quan sát modal",
         "bank_id = TEST02",
         "- Mở modal 'Chi nhánh ngân hàng'\n- Bảng 4 cột: STT, Tên chi nhánh, Tỉnh/TP, Hành động\n"
         "- Hiển thị đúng 2 dòng CN Hà Nội và CN Đà Nẵng kèm đúng tỉnh/TP\n- Có ô lọc Tỉnh/TP, Tên chi nhánh và nút 'Thêm chi nhánh'",
         "GET /v1/master-data/banks/{id}/branches"),

        ("029", "Thêm chi nhánh mới", "P0",
         "Đang mở modal chi nhánh của 'TEST02'; chưa có chi nhánh tên 'CN Cầu Giấy'",
         "1. Bấm 'Thêm chi nhánh'\n2. Nhập Tên chi nhánh = 'CN Cầu Giấy'\n3. Chọn Tỉnh/TP = 'Hà Nội'\n4. Bấm Lưu",
         "name = CN Cầu Giấy, province = Hà Nội",
         "- Toast 'Đã lưu thành công!'\n- Modal thêm đóng, bảng chi nhánh tải lại và có dòng 'CN Cầu Giấy - Hà Nội'\n"
         "- Sau khi đóng modal: cột 'Chi nhánh' của TEST02 tăng từ 2 lên 3",
         "POST /v1/master-data/banks/add-bank-branches; emit event cập nhật lại số đếm ở danh sách cha"),

        ("030", "Thêm chi nhánh bằng 'Lưu & Tiếp tục'", "P1",
         "Đang mở modal thêm chi nhánh của 'TEST02'",
         "1. Nhập 'CN Long Biên', chọn 'Hà Nội'\n2. Bấm 'Lưu & Tiếp tục'\n3. Quan sát form\n"
         "4. Nhập tiếp 'CN Hoàng Mai', chọn 'Hà Nội', bấm Lưu",
         "2 chi nhánh liên tiếp",
         "- Sau bước 2: toast thành công, form trống lại, modal vẫn mở\n"
         "- Sau bước 4: modal đóng, bảng chi nhánh có đủ cả 'CN Long Biên' và 'CN Hoàng Mai'",
         "isContinue = true → resetModal()"),

        ("031", "Sửa tên chi nhánh", "P0",
         "Chi nhánh 'CN Cầu Giấy' thuộc TEST02, tỉnh Hà Nội",
         "1. Trong modal chi nhánh, bấm nút Sửa dòng 'CN Cầu Giấy'\n2. Kiểm tra dữ liệu nạp sẵn\n"
         "3. Đổi tên thành 'CN Cầu Giấy 1'\n4. Bấm Lưu",
         "name: CN Cầu Giấy → CN Cầu Giấy 1",
         "- Modal 'Sửa chi nhánh ngân hàng' nạp sẵn tên và tỉnh/TP hiện tại\n- Sau Lưu: toast thành công, bảng hiển thị 'CN Cầu Giấy 1'\n"
         "- Số lượng chi nhánh KHÔNG tăng (vẫn là sửa, không tạo mới)\n- Modal Sửa không có nút 'Lưu & Tiếp tục'",
         "addBankBranches(): có id → find + fill + save"),

        ("032", "Đổi tỉnh/thành phố của chi nhánh", "P1",
         "Chi nhánh 'CN Cầu Giấy 1' đang gắn tỉnh Hà Nội",
         "1. Bấm Sửa dòng 'CN Cầu Giấy 1'\n2. Đổi Tỉnh/TP sang 'Hải Phòng'\n3. Bấm Lưu",
         "province: Hà Nội → Hải Phòng",
         "- Lưu thành công\n- Cột Tỉnh/TP của dòng đó đổi thành 'Hải Phòng'\n"
         "- Lọc chi nhánh theo Tỉnh/TP = 'Hà Nội' không còn trả về chi nhánh này",
         "bank_branches.province_id cập nhật + join provinces lấy tên mới"),

        ("033", "Lọc chi nhánh theo Tỉnh/TP", "P0",
         "TEST02 có: 'CN Hà Nội' (Hà Nội), 'CN Long Biên' (Hà Nội), 'CN Đà Nẵng' (Đà Nẵng)",
         "1. Mở modal chi nhánh của TEST02\n2. Chọn Tỉnh/TP = 'Hà Nội'\n3. Bấm 'Tìm kiếm'",
         "province_id = Hà Nội",
         "- Bảng chỉ còn 2 dòng: CN Hà Nội, CN Long Biên\n- Không hiển thị CN Đà Nẵng\n- STT đánh lại từ 1",
         "bank_branches.province_id = province_id"),

        ("034", "Lọc chi nhánh theo tên (khớp một phần)", "P0",
         "TEST02 có các chi nhánh: 'CN Hà Nội', 'CN Long Biên', 'CN Đà Nẵng'",
         "1. Nhập ô 'Tên chi nhánh' = 'Long'\n2. Bấm 'Tìm kiếm'",
         "name = 'Long'",
         "- Bảng chỉ còn 'CN Long Biên'\n- Có thể nhấn Enter tại ô nhập để tìm thay cho bấm nút",
         "bank_branches.name LIKE %name%; @keyup.enter → getData"),

        ("035", "Làm mới bộ lọc chi nhánh", "P1",
         "Đang lọc chi nhánh theo Tỉnh/TP = 'Hà Nội' và Tên = 'Long'",
         "1. Bấm nút 'Làm mới'\n2. Quan sát ô lọc và bảng",
         "—",
         "- Cả 2 ô lọc về rỗng\n- Bảng hiển thị lại toàn bộ chi nhánh của ngân hàng đang mở",
         "reset(): formFilter về rỗng rồi getData()"),

        ("036", "Xoá chi nhánh chưa được sử dụng", "P0",
         "Chi nhánh 'CN Hoàng Mai' của TEST02 chưa được nhân viên nào sử dụng (can_delete = true)",
         "1. Bấm nút Xoá dòng 'CN Hoàng Mai'\n2. Đọc hộp xác nhận\n3. Bấm 'Xóa'",
         "id chi nhánh CN Hoàng Mai",
         "- Hộp xác nhận: \"Bạn có chắc muốn xóa chi nhánh 'CN Hoàng Mai'?\"\n- Toast 'Xoá chi nhánh ngân hàng thành công'\n"
         "- Dòng biến mất khỏi bảng\n- Đóng modal: cột 'Chi nhánh' ngoài danh sách giảm 1",
         "DELETE /v1/master-data/banks/bank-branches/{id}"),

        ("037", "Nút Xoá chi nhánh bị chặn khi đang được sử dụng", "P0",
         "Chi nhánh 'CN Hà Nội' của TEST02 đang được gán trong tài khoản ngân hàng của 1 nhân viên đang làm việc (status 1)",
         "1. Mở modal chi nhánh của TEST02\n2. Rê chuột vào nút Xoá dòng 'CN Hà Nội'",
         "can_delete = false",
         "- Nút Xoá bị mờ, không click được\n- Tooltip: 'Không thể xóa bản ghi, chi nhánh đang được sử dụng trên hệ thống'\n- Nút Sửa vẫn bật",
         "BankBranch::canDelete() kiểm tra employee_bank_accounts / employee_authorized_bank_accounts theo bank_branch_id"),

        ("038", "Đóng modal chi nhánh", "P2",
         "Đang mở modal 'Chi nhánh ngân hàng'",
         "1. Bấm nút 'Đóng' ở footer\n2. Quan sát màn danh sách",
         "—",
         "- Modal đóng, quay lại màn danh mục ngân hàng\n- Số lượng chi nhánh trên dòng tương ứng đã đồng bộ với thao tác vừa thực hiện",
         "eventHandler type = 'bank-branches' cập nhật item.branches"),
    ]),

    # ---------------------------------------------------------------- VI
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Bỏ trống toàn bộ trường bắt buộc khi tạo mới", "P0",
         "Đang mở modal 'Tạo mới ngân hàng', chưa nhập gì",
         "1. Bấm Lưu ngay",
         "Tất cả trường rỗng",
         "- Toast đỏ 'Vui lòng kiểm tra lại thông tin'\n"
         "- Hiện lỗi đỏ 'Bắt buộc phải nhập' dưới cả 3 ô: Mã ngân hàng, Tên ngân hàng, Tên viết tắt\n"
         "- Không gửi request POST, modal không đóng",
         "FE chặn trước: name, code, short_name required"),

        ("002", "Bỏ trống riêng ô Mã ngân hàng", "P0",
         "Đang mở modal Tạo mới",
         "1. Nhập Tên ngân hàng = 'NH Kiểm Thử X', Tên viết tắt = 'NHKTX'\n2. Để trống Mã ngân hàng\n3. Bấm Lưu",
         "code = rỗng",
         "- Lỗi 'Bắt buộc phải nhập' chỉ hiện dưới ô Mã ngân hàng\n- 2 ô còn lại không báo lỗi\n- Không gửi request",
         "Validate từng trường độc lập"),

        ("003", "Bỏ trống riêng ô Tên viết tắt", "P0",
         "Đang mở modal Tạo mới",
         "1. Nhập Mã = 'TESTX', Tên = 'NH Kiểm Thử X'\n2. Để trống Tên viết tắt\n3. Bấm Lưu",
         "short_name = rỗng",
         "- Lỗi 'Bắt buộc phải nhập' hiện dưới ô Tên viết tắt\n- Không lưu bản ghi",
         "Rule BE: short_name required"),

        ("004", "Tạo mới trùng MÃ ngân hàng đã tồn tại", "P0",
         "Đã tồn tại ngân hàng code = 'TEST01'",
         "1. Bấm Tạo mới\n2. Nhập Mã = 'TEST01', Tên = 'Ngân hàng khác hoàn toàn', Tên viết tắt = 'NHK'\n3. Bấm Lưu",
         "code trùng = TEST01",
         "- API trả 422\n- Lỗi đỏ dưới ô Mã ngân hàng: 'Mã ngân hàng này đã tồn tại'\n"
         "- Toast đỏ 'Vui lòng kiểm tra lại thông tin'\n- Modal không đóng, KHÔNG tạo thêm bản ghi",
         "Rule: code unique:banks,code (bỏ qua id đang sửa)"),

        ("005", "Tạo mới trùng TÊN ngân hàng đã tồn tại", "P0",
         "Đã tồn tại ngân hàng name = 'Ngân hàng Kiểm Thử 01'",
         "1. Bấm Tạo mới\n2. Nhập Mã = 'TESTY', Tên = 'Ngân hàng Kiểm Thử 01', Tên viết tắt = 'NHY'\n3. Bấm Lưu",
         "name trùng",
         "- Lỗi đỏ dưới ô Tên ngân hàng: 'Tên ngân hàng này đã tồn tại'\n- Không tạo bản ghi mới",
         "Rule: name unique:banks,name"),

        ("006", "Sửa thành mã trùng với ngân hàng khác", "P0",
         "Có 2 ngân hàng: code 'TEST01' và code 'TEST02'",
         "1. Bấm Sửa dòng TEST02\n2. Đổi Mã ngân hàng thành 'TEST01'\n3. Bấm Lưu",
         "code TEST02 → TEST01",
         "- Lỗi 'Mã ngân hàng này đã tồn tại' dưới ô Mã ngân hàng\n- Bản ghi TEST02 giữ nguyên mã cũ trên danh sách",
         "unique ignore chỉ bỏ qua chính id đang sửa"),

        ("007", "Lỗi validate biến mất sau khi sửa đúng", "P1",
         "Đang hiển thị lỗi 'Bắt buộc phải nhập' ở cả 3 ô",
         "1. Nhập đầy đủ Mã, Tên, Tên viết tắt hợp lệ\n2. Bấm Lưu",
         "Dữ liệu hợp lệ",
         "- Toàn bộ dòng lỗi đỏ biến mất\n- Lưu thành công, toast xanh 'Đã lưu thành công!'",
         "submitSave() reset this.error = {} ở đầu mỗi lần submit"),

        ("008", "Chống double-click nút Lưu", "P0",
         "Đang mở modal Tạo mới với dữ liệu hợp lệ (Mã 'TESTZ')",
         "1. Bấm Lưu liên tiếp 3 lần thật nhanh\n2. Kiểm tra danh sách và bảng banks",
         "3 lần click < 1 giây",
         "- Nút Lưu bị vô hiệu ngay sau lần bấm đầu tiên cho tới khi API trả về\n"
         "- Chỉ tạo ra ĐÚNG 1 bản ghi mã TESTZ\n- Không có bản ghi trùng lặp",
         "isSubmitSave = true → :disabled trên cả 3 nút footer"),

        ("009", "Bỏ trống trường bắt buộc khi thêm chi nhánh", "P0",
         "Đang mở modal 'Thêm chi nhánh ngân hàng'",
         "1. Bấm Lưu ngay khi chưa nhập gì",
         "name và province_id rỗng",
         "- Toast đỏ 'Vui lòng kiểm tra lại thông tin'\n- Lỗi 'Bắt buộc phải nhập' dưới CẢ ô Tên chi nhánh và ô Tỉnh/Thành phố\n- Không gửi request",
         "Rule: name required, province_id required"),

        ("010", "Thêm chi nhánh trùng tên trong CÙNG ngân hàng", "P0",
         "Ngân hàng TEST02 đã có chi nhánh 'CN Hà Nội'",
         "1. Mở modal chi nhánh TEST02\n2. Bấm 'Thêm chi nhánh'\n3. Nhập tên = 'CN Hà Nội', chọn tỉnh bất kỳ\n4. Bấm Lưu",
         "name = 'CN Hà Nội', bank_id = TEST02",
         "- API trả 422\n- Lỗi đỏ dưới ô Tên chi nhánh: 'Tên chi nhánh ngân hàng này đã tồn tại'\n- Không tạo bản ghi mới",
         "Rule unique scope theo bank_id"),

        ("011", "Thêm chi nhánh trùng tên nhưng KHÁC ngân hàng", "P0",
         "Ngân hàng TEST02 có 'CN Hà Nội'; ngân hàng TEST03 chưa có chi nhánh nào",
         "1. Mở modal chi nhánh của TEST03\n2. Thêm chi nhánh tên 'CN Hà Nội', tỉnh Hà Nội\n3. Bấm Lưu",
         "name = 'CN Hà Nội', bank_id = TEST03",
         "- Lưu THÀNH CÔNG (không báo trùng)\n- TEST03 có 1 chi nhánh 'CN Hà Nội'\n- Chi nhánh của TEST02 không bị ảnh hưởng",
         "unique scope where bank_id → cho phép trùng tên giữa 2 ngân hàng khác nhau"),

        ("012", "Xoá ngân hàng qua API khi đang được nhân viên sử dụng", "P0",
         "Ngân hàng 'TEST06' đang được gán cho tài khoản ngân hàng của 1 nhân viên có status = 1; FE đã disable nút Xoá",
         "1. Lấy id của TEST06\n2. Gọi trực tiếp DELETE /v1/master-data/banks/{id} bằng Postman với token hợp lệ\n3. Quan sát response và DB",
         "id = TEST06",
         "- API trả HTTP 400 với message: 'Không thể xóa bản ghi, ngân hàng đang được sử dụng trên hệ thống'\n"
         "- Bản ghi TEST06 vẫn còn trong bảng banks\n- Các chi nhánh của nó không bị xoá",
         "Guard server-side trong deleteBank(): throw khi canDelete() = false"),

        ("013", "Xoá ngân hàng đang được dùng ở danh mục tài khoản ngân hàng công ty", "P0",
         "Ngân hàng 'TEST07' đang được tham chiếu bởi 1 bản ghi trong company_accounts (Tài chính → Danh mục tài khoản ngân hàng)",
         "1. Mở màn danh mục ngân hàng, tìm TEST07\n2. Rê chuột vào nút Xoá\n3. Thử gọi DELETE qua API",
         "TEST07 có trong company_accounts",
         "- Nút Xoá mờ, tooltip 'Không thể xóa bản ghi, ngân hàng đang được sử dụng trên hệ thống'\n"
         "- Gọi API trực tiếp trả HTTP 400, bản ghi không bị xoá",
         "canDelete() kiểm tra thêm company_accounts.bank_id"),

        ("014", "Ngân hàng chỉ được nhân viên đã nghỉ việc sử dụng vẫn xoá được", "P0",
         "Ngân hàng 'TEST08' chỉ được gán cho nhân viên có status ngoài [1, 3] (đã nghỉ việc); không nằm trong company_accounts",
         "1. Tìm TEST08 trên danh sách\n2. Quan sát nút Xoá\n3. Bấm Xoá và xác nhận",
         "Nhân viên liên quan: status ≠ 1 và ≠ 3",
         "- Nút Xoá KHÔNG bị mờ\n- Xoá thành công, toast 'Xoá ngân hàng thành công'\n- Bản ghi biến mất khỏi danh sách",
         "canDelete() chỉ tính nhân viên có employee_infos.status thuộc [1, 3]"),

        ("015", "Xoá bản ghi đã bị người khác xoá trước đó", "P1",
         "Ngân hàng 'TEST09' đang hiển thị trên màn của user A; user B vừa xoá bản ghi này",
         "1. Trên màn user A (chưa reload), bấm Xoá dòng TEST09\n2. Xác nhận 'Xóa'\n3. Quan sát",
         "Bản ghi không còn trong DB",
         "- Hệ thống không văng lỗi trắng màn hình\n- Danh sách tải lại và dòng TEST09 không còn\n- Nếu API trả lỗi, hiện toast đỏ với thông báo rõ ràng",
         "deleteBank(): find trả null → trả về null, controller phản hồi 404"),

        ("016", "Nhập giá trị rất dài cho tên ngân hàng", "P2",
         "Đang mở modal Tạo mới",
         "1. Nhập Tên ngân hàng chuỗi 300 ký tự\n2. Nhập Mã 'TESTLONG', Tên viết tắt 'LONG'\n3. Bấm Lưu\n4. Quan sát danh sách",
         "name: chuỗi 300 ký tự",
         "- Nếu vượt giới hạn cột DB: hiển thị lỗi rõ ràng, KHÔNG trả 500\n"
         "- Nếu lưu được: danh sách xuống dòng đúng (text-wrap), không vỡ layout bảng, không tràn ngang",
         "Cột bankInfo có cellClass 'text-wrap'"),

        ("017", "Nhập ký tự đặc biệt và tiếng Việt có dấu", "P1",
         "Đang mở modal Tạo mới",
         "1. Nhập Tên = 'Ngân hàng TMCP Đông Á & Cộng sự (Chi nhánh #1)'\n2. Mã = 'TEST-Đ.1'\n3. Tên viết tắt = 'ĐA&CS'\n4. Bấm Lưu",
         "Chuỗi có dấu tiếng Việt, &, #, (), dấu chấm",
         "- Lưu thành công\n- Danh sách hiển thị đúng nguyên văn, không bị mã hoá HTML (không thấy &amp;)\n- Mở lại modal Sửa thấy đúng chuỗi đã nhập",
         "Kiểm tra encode UTF-8 và escape HTML"),

        ("018", "Nhập thẻ script vào ô nhập (XSS)", "P0",
         "Đang mở modal Tạo mới",
         "1. Nhập Tên ngân hàng = '<script>alert(1)</script>'\n2. Nhập Mã = 'TESTXSS', Tên viết tắt = 'XSS'\n3. Bấm Lưu\n4. Quan sát danh sách",
         "name = '<script>alert(1)</script>'",
         "- KHÔNG xuất hiện hộp thoại alert\n- Chuỗi hiển thị nguyên văn dưới dạng text trên danh sách\n- Không lỗi JS console",
         "Vue nội suy {{ }} tự escape; kiểm tra cả cột dùng v-html"),

        ("019", "Chi nhánh có tỉnh/TP không hợp lệ", "P2",
         "Chi nhánh 'CN Lỗi Data' của TEST02 có province_id = NULL hoặc trỏ tới tỉnh không tồn tại (tạo bằng SQL trực tiếp)",
         "1. Mở modal chi nhánh của TEST02\n2. Tìm 'CN Lỗi Data' trong bảng\n3. Đối chiếu với cột 'Chi nhánh' ngoài danh sách",
         "province_id = NULL",
         "- Ghi nhận thực tế: chi nhánh này KHÔNG hiển thị trong modal (do INNER JOIN provinces)\n"
         "- Nhưng cột 'Chi nhánh' ngoài danh sách VẪN đếm nó → số đếm lệch so với số dòng trong modal\n"
         "- Ghi Failed nếu nghiệp vụ yêu cầu 2 số phải khớp",
         "getBankBranches() dùng join (INNER); cột đếm dùng with('branches') không join"),

        ("020", "Mất kết nối mạng khi đang lưu", "P2",
         "Đang mở modal Tạo mới với dữ liệu hợp lệ",
         "1. Ngắt mạng (DevTools → Offline)\n2. Bấm Lưu\n3. Quan sát",
         "Network: Offline",
         "- Hiện toast đỏ thông báo lỗi\n- Loading bar tắt (không quay mãi)\n- Nút Lưu được bật lại để thử lại\n- Modal không đóng, dữ liệu đã nhập còn nguyên",
         "finally: $loading.finish() + isSubmitSave = false"),
    ]),

    # ---------------------------------------------------------------- VII
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Gọi API danh sách khi chưa đăng nhập", "P0",
         "Chưa có token hoặc đã đăng xuất",
         "1. Gọi GET /api/v1/master-data/banks bằng Postman, KHÔNG gắn header Authorization\n2. Quan sát response",
         "Không có Bearer token",
         "- API trả HTTP 401 Unauthorized\n- Không trả về bất kỳ dữ liệu ngân hàng nào trong body",
         "Route group middleware auth:api"),

        ("002", "Truy cập URL màn hình khi chưa đăng nhập", "P0",
         "Trình duyệt ở trạng thái đã đăng xuất (xoá token)",
         "1. Gõ URL /master-data/banks\n2. Quan sát",
         "—",
         "- Bị chuyển về màn đăng nhập\n- Không hiển thị dữ liệu danh mục ngân hàng dù chỉ trong tích tắc",
         "Middleware auth phía FE"),

        ("003", "Token hết hạn khi đang thao tác", "P1",
         "Đang mở màn danh mục ngân hàng, để token hết hạn (hoặc xoá token trong storage)",
         "1. Xoá/làm hỏng token\n2. Bấm sang trang 2\n3. Quan sát",
         "Token không hợp lệ",
         "- Hệ thống xử lý lỗi 401 (chuyển về đăng nhập hoặc hiện thông báo), KHÔNG hiện dữ liệu cũ như thể vẫn hợp lệ\n- Không lỗi trắng màn hình",
         "Xử lý 401 tập trung ở tầng axios"),

        ("004", "Không phân quyền theo cấp — 2 user khác công ty thấy dữ liệu giống nhau", "P0",
         "User A thuộc Công ty 1, User B thuộc Công ty 2; danh mục có 20 ngân hàng",
         "1. Đăng nhập user A, ghi lại tổng số bản ghi và 5 mã đầu tiên\n2. Đăng xuất, đăng nhập user B\n3. Vào cùng màn, so sánh",
         "A: công ty 1; B: công ty 2",
         "- Cả 2 user thấy CÙNG tổng số bản ghi = 20 và cùng danh sách\n"
         "- Đây là hành vi đúng: danh mục dùng chung, KHÔNG lọc theo công ty/phòng ban",
         "getBanks() không có điều kiện company_id/department_id; route không gắn checkPermission"),

        ("005", "User thường thực hiện được thao tác ghi", "P1",
         "User C là nhân viên thường, không có vai trò quản trị",
         "1. Đăng nhập user C\n2. Vào /master-data/banks\n3. Thử Tạo mới, Sửa, Khoá, Xoá 1 bản ghi thử nghiệm",
         "User: nhân viên thường",
         "- Mọi thao tác đều thực hiện được (không có màn 403)\n"
         "- Ghi nhận đây là hiện trạng do route chưa gắn checkPermission — nếu nghiệp vụ yêu cầu giới hạn quyền thì báo là điểm cần bổ sung",
         "Toàn bộ route banks chỉ có auth:api, KHÔNG có middleware checkPermission"),

        ("006", "Xử lý khi API trả 403", "P1",
         "Giả lập API /master-data/banks trả HTTP 403 (dùng DevTools override response)",
         "1. Giả lập 403\n2. Reload màn danh mục ngân hàng",
         "Response: 403",
         "- Hệ thống chuyển sang trang /pages/extras/403\n- Không hiện toast lỗi tải dữ liệu chồng lên\n- Không lỗi trắng màn hình",
         "loadData catch: status 403 → router.push('/pages/extras/403')"),

        ("007", "Xử lý khi API trả lỗi 500", "P1",
         "Giả lập API /master-data/banks trả HTTP 500",
         "1. Giả lập 500\n2. Reload màn",
         "Response: 500",
         "- Hiện toast đỏ 'Lỗi khi tải dữ liệu'\n- Bảng ở trạng thái rỗng, loading đã tắt\n- Vẫn thao tác được panel lọc để thử lại",
         "catch chung: toast 'Lỗi khi tải dữ liệu'"),
    ]),

    # ---------------------------------------------------------------- VIII
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đầy đủ: tạo → thêm chi nhánh → sửa → khoá → mở khoá → xoá", "P0",
         "User đã đăng nhập; chưa tồn tại ngân hàng code = 'E2E01'",
         "1. Tạo mới ngân hàng: Mã 'E2E01', Tên 'Ngân hàng E2E 01', Tên viết tắt 'E2E01', upload logo hợp lệ\n"
         "2. Click cột 'Chi nhánh' của E2E01 → thêm 2 chi nhánh 'CN A' (Hà Nội), 'CN B' (Đà Nẵng) → đóng modal\n"
         "3. Bấm Sửa E2E01, đổi Địa chỉ giao dịch thành '99 Nguyễn Trãi' → Lưu\n"
         "4. Bấm nút ổ khoá → xác nhận 'Khoá'\n5. Bấm nút mở khoá → xác nhận 'Mở khoá'\n"
         "6. Xoá 2 chi nhánh trong modal chi nhánh\n7. Bấm Xoá E2E01 → xác nhận",
         "code = E2E01",
         "- B1: bản ghi mới đứng đầu danh sách, logo hiển thị, Chi nhánh = 0\n"
         "- B2: cột Chi nhánh = 2\n- B3: cột 'Địa chỉ giao dịch' = '99 Nguyễn Trãi', cột Cập nhật đổi thành thời điểm hiện tại\n"
         "- B4: pill 'Khoá', nút Sửa/Xoá mờ\n- B5: pill 'Hoạt động', nút Sửa/Xoá bật lại\n"
         "- B6: cột Chi nhánh về 0\n- B7: dòng biến mất, tìm 'E2E01' không còn kết quả",
         "Bao trùm toàn bộ CRUD + lock/unlock + quan hệ bank ↔ branch"),

        ("002", "Luồng tạo nhanh bằng Tra cứu rồi dùng ở hồ sơ nhân viên", "P0",
         "Chưa tồn tại ngân hàng lấy từ tra cứu; có 1 nhân viên đang làm việc để gán tài khoản",
         "1. Tạo mới → Tra cứu 'Techcombank' → click dòng kết quả → kiểm tra Mã/Tên/Tên viết tắt/Logo được điền\n"
         "2. Bấm Lưu\n3. Thêm chi nhánh 'CN Hoàn Kiếm' (Hà Nội)\n"
         "4. Vào Hồ sơ nhân viên → Tài khoản ngân hàng → chọn ngân hàng vừa tạo và chi nhánh 'CN Hoàn Kiếm' → Lưu\n"
         "5. Quay lại danh mục ngân hàng, rê chuột vào nút Xoá của ngân hàng đó",
         "Ngân hàng lấy từ nguồn tra cứu",
         "- B1: 4 trường được điền tự động đúng dữ liệu tra cứu\n- B2: lưu thành công\n"
         "- B4: dropdown ngân hàng và chi nhánh có đủ dữ liệu vừa tạo, lưu hồ sơ thành công\n"
         "- B5: nút Xoá đã chuyển sang mờ với tooltip 'Không thể xóa bản ghi, ngân hàng đang được sử dụng trên hệ thống'",
         "Kiểm chứng liên thông danh mục → hồ sơ nhân viên và cơ chế can_delete"),

        ("003", "Luồng lọc — phân trang — sắp xếp giữ đúng ngữ cảnh", "P1",
         "Danh mục có ≥ 30 ngân hàng, trong đó ≥ 15 bản ghi có tên viết tắt chứa 'Bank'",
         "1. Lọc Tên viết tắt = 'Bank'\n2. Đổi số dòng/trang = 10\n3. Sang trang 2\n"
         "4. Click sắp xếp cột 'Cập nhật'\n5. Bấm 'Làm mới'",
         "short_name = 'Bank'",
         "- B3: trang 2 vẫn chỉ chứa bản ghi khớp 'Bank', STT bắt đầu từ 11\n"
         "- B4: sắp xếp áp dụng trên tập đã lọc, quay về trang 1, bộ lọc 'Bank' KHÔNG bị mất\n"
         "- B5: mọi bộ lọc + sắp xếp về mặc định, danh sách hiển thị toàn bộ dữ liệu từ trang 1",
         "Bộ lọc, phân trang và sắp xếp phải cộng dồn đúng trên cùng một truy vấn"),

        ("004", "Luồng đồng bộ CRM khi bật cấu hình use_crm", "P2",
         "MasterSetting category = 'use_crm' đang bật (có content); hệ thống CRM kết nối được",
         "1. Tạo mới ngân hàng 'E2E-CRM'\n2. Kiểm tra bảng module_mappings\n"
         "3. Sửa tên ngân hàng thành 'E2E-CRM đổi tên'\n4. Khoá ngân hàng\n5. Kiểm tra dữ liệu bên CRM",
         "use_crm = bật",
         "- B1: tạo thành công bên HRM\n- B2: có bản ghi module_mappings với hrm_model = Bank, crm_model = 'mate.bank', crm_id ≠ null\n"
         "- B3: tên bên CRM cập nhật theo\n- B4: bản ghi CRM chuyển active = false\n"
         "- Nếu CRM lỗi: giao dịch rollback, KHÔNG tạo bản ghi mồ côi bên HRM",
         "Bank::create/updating đồng bộ CRMBankService trong DB::transaction"),

        ("005", "Luồng khi tắt cấu hình use_crm", "P2",
         "MasterSetting category = 'use_crm' đang TẮT (không có bản ghi hoặc content rỗng)",
         "1. Tạo mới ngân hàng 'E2E-NOCRM'\n2. Sửa tên\n3. Thêm 1 chi nhánh\n4. Xoá chi nhánh vừa thêm",
         "use_crm = tắt",
         "- Toàn bộ thao tác chạy bình thường và nhanh (không chờ gọi API ngoài)\n"
         "- Không tạo bản ghi trong module_mappings\n- Không có lỗi kết nối CRM trong log",
         "Nhánh else: chỉ static::query()->create(), bỏ qua toàn bộ logic CRM"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT
    a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT
    a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT
    b.alignment = WRAP_TOP_LEFT
    b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, "Testcase _ %s" % FEATURE_NAME)
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT
    lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right")
    lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT
    vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN
    vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0


def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1


def write_tc(tc_id, function, priority, precondition, steps, test_data, expected,
             business_note, module=MODULE_NAME, group=""):
    global current_row, data_row_idx
    values = [
        module, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed",
        "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    newlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, max(longest // 4, newlines * 15 + 20)))
    current_row += 1
    data_row_idx += 1


if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc("TC-ROLE-%s" % suffix, func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")

ROMANS = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
total_tc = 0
p0_tc = 0
for roman, title, tcs in SECTIONS:
    write_section_row("%s. %s" % (roman, title))
    sec_idx = ROMANS.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = "TC_%02d.%03d" % (sec_idx, int(tc_num))
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

dv = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
    showDropDown=False,
)
dv.add("L18:N%d" % (current_row + 100))
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print("Generated: %s" % OUTPUT_FILE)
print("Rows: 1-10 description, 11-15 summary, 17 header, 18-%d data" % (current_row - 1))
print("Total TC: %d | P0: %d (%.0f%%)" % (total_tc, p0_tc, p0_tc * 100.0 / total_tc))
