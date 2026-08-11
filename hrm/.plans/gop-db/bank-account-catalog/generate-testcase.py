# -*- coding: utf-8 -*-
"""Generate testcase Excel cho man Danh muc tai khoan ngan hang (Modules/Finance - pages/finance/account-banks)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE = ".plans/gop-db/bank-account-catalog/testcase.xlsx"
SHEET_NAME = "DanhMucTKNganHang"
FEATURE_NAME = "Danh mục tài khoản ngân hàng"
MODULE_NAME = "DM tài khoản ngân hàng"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý danh mục tài khoản ngân hàng CỦA CÔNG TY (phân hệ Tài chính → Danh mục → Danh mục tài khoản ngân hàng, URL /finance/account-banks).\n"
     "Màn port từ ERP (`admin/accounting/account-banks`), chạy song song với ERP trên CÙNG bảng `company_accounts` của DB gộp — KHÔNG đổi schema.\n"
     "Phạm vi chức năng (tối giản đúng như ERP): xem danh sách + lọc, Thêm mới, Sửa, Xem chi tiết, Khóa / Mở khóa.\n"
     "KHÔNG có: Xóa, Export, Import, In, Lịch sử thay đổi, sắp xếp cột."),

    ("2. Đối tượng được tính / hiển thị",
     "► Chỉ bản ghi `company_accounts` có `company_id` = công ty của user đang đăng nhập (lấy qua `auth()->user()->info->company_id`).\n"
     "► Bao gồm CẢ status = 1 (Hoạt động) và status = 0 (Khóa) — bản ghi khóa vẫn hiển thị và vẫn Sửa được, chỉ đổi nhãn trạng thái.\n"
     "► Bản ghi do màn ERP tạo vẫn hiển thị bình thường ở HRM nếu cùng company_id (2 màn dùng chung 1 bảng).\n"
     "► Dropdown Ngân hàng ở form Thêm/Sửa: CHỈ ngân hàng `banks.status = 1`, sắp xếp theo tên.\n"
     "► Dropdown Chi nhánh: API trả TOÀN BỘ `bank_branches`, FE tự lọc theo `bank_id` đang chọn."),

    ("3. Đối tượng bị ẩn / không tính",
     "► Tài khoản của công ty KHÁC — không hiển thị trong danh sách; gọi trực tiếp API show/update/lock/unlock theo id đó trả 404 (không lộ sự tồn tại).\n"
     "► Nếu user đăng nhập KHÔNG gắn hồ sơ nhân viên / không có company_id: danh sách trả RỖNG (query `whereRaw('1 = 0')`), thao tác Thêm bị chặn với thông báo 422.\n"
     "► Ngân hàng đã bị khóa (`banks.status = 2`): không xuất hiện trong dropdown chọn ngân hàng khi Thêm/Sửa; ở chế độ Xem vẫn hiển thị tên (lấy từ `bank_name` denormalized).\n"
     "► Không có bản ghi bị ẩn theo phòng ban / người tạo — trong 1 công ty mọi user có quyền đều thấy như nhau."),

    ("4. Bộ lọc thời gian áp dụng cho",
     "Không áp dụng — màn không có bộ lọc theo khoảng thời gian và không hiển thị cột ngày.\n"
     "`created_at` chỉ dùng để sắp xếp mặc định (giảm dần — bản ghi mới nhất lên đầu), không lộ ra giao diện."),

    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Bảng chính `company_accounts` (dùng chung với ERP). Không có cây phân cấp — danh sách phẳng.\n"
     "► Khóa ngoài: `bank_id` → `banks.id`; `bank_branch_id` → `bank_branches.id` (chi nhánh phải thuộc đúng ngân hàng); `currency_id` → `currencies.id`; `company_id` → `companies.id`.\n"
     "► 2 cột denormalized `bank_name`, `bank_branch`: lưu TÊN tại thời điểm lưu bản ghi (fill từ id) — dùng để hiển thị trên danh sách và lọc.\n"
     "► Entity kế thừa Model THUẦN (không BaseModel) vì bảng không có `created_by` / `updated_by`.\n"
     "► Cột `currency_id` nullable — bản ghi cũ do ERP tạo có thể null."),

    ("6. Quy tắc cộng dồn / deduplicate",
     "► Không cộng dồn số liệu. Mỗi dòng = 1 bản ghi `company_accounts`, phân trang server-side (mặc định `per_page` = 10).\n"
     "► Chống trùng: `account_number` unique TOÀN BẢNG `company_accounts` (KHÔNG scope theo company_id) — 2 công ty khác nhau cũng không được trùng số tài khoản. Khi Sửa thì bỏ qua chính bản ghi đang sửa.\n"
     "► Bảng KHÔNG có unique index ở DB → về lý thuyết vẫn có thể race khi ERP và HRM ghi đồng thời (chấp nhận, giữ như hiện trạng ERP).\n"
     "► FE có cơ chế chống gọi API trùng (DedupeLoadMixin): 2 request cùng tham số trong 800ms chỉ chạy 1.\n"
     "► STT hiển thị = (trang hiện tại − 1) × số dòng/trang + vị trí + 1."),

    ("7. Phân quyền cấp",
     "Màn dùng ĐÚNG 1 quyền (giống ERP), KHÔNG phân quyền theo cấp tổ chức:\n"
     "• `Quản lý danh mục tài khoản ngân hàng` — gắn cho TẤT CẢ route: GET /, GET /options, POST /, PUT /{id}, GET /{id}/lock, GET /{id}/unlock, GET /{id}.\n"
     "  → Không có quyền: menu bị ẩn, mọi API trả 403 (kể cả xem danh sách).\n"
     "  → Có quyền: xem + thêm + sửa + khóa/mở khóa (không có thao tác xóa).\n"
     "• FE dùng `hasAPermission('Quản lý danh mục tài khoản ngân hàng')` (biến `canManage`) để ẩn/hiện nút Thêm tài khoản, nút Sửa và nút khóa/mở khóa.\n"
     "• Phạm vi dữ liệu KHÔNG do permission quyết định mà do `company_id` của user — xem mục 2 và section VII."),

    ("8. Cách tính các ô thống kê",
     "Không có khối thẻ thống kê. Các ô số/đếm duy nhất:\n"
     "► Ô 'STT' = (currentPage − 1) × pageSize + index + 1.\n"
     "► Dòng 'Hiển thị x–y / z' dưới bảng: z = `total` từ API; x = total ? (currentPage − 1) × pageSize + 1 : 0; y = (currentPage − 1) × pageSize + số dòng thực tế của trang.\n"
     "► Ô 'Loại tiền tệ' = mã tiền tệ (`currencies.code`) join theo `currency_id`; nếu null → hiển thị '—'.\n"
     "► Ô 'Trạng thái' = 'Hoạt động' khi status = 1, 'Khóa' khi status = 0."),

    ("9. Ghi chú đọc bảng",
     "► Bảng 7 cột, cột STT và Số tài khoản được ghim (sticky) khi cuộn ngang.\n"
     "► KHÔNG cột nào cho phép sắp xếp — thứ tự luôn là bản ghi tạo mới nhất lên đầu.\n"
     "► Nút Xem / Sửa nằm ngay trong ô 'Số tài khoản'; nút khóa–mở khóa nằm cạnh nhãn trạng thái.\n"
     "► 'Chủ tài khoản' và 'Ngân hàng' luôn hiển thị CHỮ IN HOA (hệ thống tự chuyển khi lưu).\n"
     "► Bộ lọc được ghi nhớ trong 10 phút khi rời trang rồi quay lại (lưu ở trình duyệt, key `finance_account_banks`).\n"
     "► Giá trị trống hiển thị '—'."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("01", "Truy cập màn hình khi CÓ quyền quản lý", "P0",
     "User A được gán quyền 'Quản lý danh mục tài khoản ngân hàng'; công ty của A có 8 tài khoản ngân hàng",
     "1. Đăng nhập bằng user A\n2. Vào phân hệ Tài chính → Danh mục → 'Danh mục tài khoản ngân hàng'\n3. Quan sát danh sách và các nút thao tác",
     "User A: có quyền",
     "- Menu 'Danh mục tài khoản ngân hàng' hiển thị\n- Danh sách load đủ 8 bản ghi của công ty A\n"
     "- Hiển thị nút 'Thêm tài khoản'\n- Mỗi dòng có nút Xem và Sửa; cột Trạng thái có nút khóa/mở khóa",
     "Permission: Quản lý danh mục tài khoản ngân hàng — gắn cho toàn bộ route của màn"),

    ("02", "Menu bị ẩn khi KHÔNG có quyền", "P0",
     "User B KHÔNG được gán quyền 'Quản lý danh mục tài khoản ngân hàng'",
     "1. Đăng nhập bằng user B\n2. Mở phân hệ Tài chính → nhóm menu Danh mục\n3. Tìm mục 'Danh mục tài khoản ngân hàng'",
     "User B: không có quyền",
     "- Mục menu 'Danh mục tài khoản ngân hàng' KHÔNG xuất hiện\n- Các mục danh mục khác mà B có quyền vẫn hiển thị bình thường",
     "finance.js: isShow = ['Quản lý danh mục tài khoản ngân hàng']"),

    ("03", "Gõ thẳng URL khi KHÔNG có quyền", "P0",
     "User B không có quyền, đã đăng nhập",
     "1. Gõ URL /finance/account-banks lên thanh địa chỉ\n2. Enter\n3. Quan sát màn hình và tab Network",
     "User B: không có quyền",
     "- API GET /v1/finance/account-banks trả HTTP 403\n- Bảng KHÔNG hiển thị bất kỳ dữ liệu tài khoản nào\n"
     "- Không hiện toast 'Lỗi khi tải dữ liệu' (403 được xử lý im lặng)\n- Không lỗi trắng màn hình",
     "index cũng gắn checkPermission; FE bỏ qua toast khi status = 403"),

    ("04", "Gọi trực tiếp API ghi khi KHÔNG có quyền", "P0",
     "User B không có quyền; đã có 1 tài khoản id = X thuộc công ty của B",
     "1. Lấy token của user B\n2. Gọi lần lượt bằng Postman: POST /v1/finance/account-banks; PUT /v1/finance/account-banks/X; "
     "GET /v1/finance/account-banks/X/lock; GET /v1/finance/account-banks/X/unlock; GET /v1/finance/account-banks/options",
     "User B: token hợp lệ, thiếu quyền",
     "- Cả 5 request đều trả HTTP 403\n- Không có bản ghi nào được tạo/sửa\n- Trạng thái của bản ghi X không đổi",
     "Toàn bộ route gắn checkPermission:Quản lý danh mục tài khoản ngân hàng"),

    ("05", "User có quyền nhưng CHƯA gắn công ty", "P0",
     "User C có quyền quản lý nhưng hồ sơ nhân viên không có company_id (hoặc chưa có employee_info)",
     "1. Đăng nhập user C\n2. Vào /finance/account-banks\n3. Bấm 'Thêm tài khoản', nhập đủ thông tin hợp lệ và bấm Lưu",
     "User C: company_id = null",
     "- Danh sách hiển thị RỖNG ('Không có dữ liệu phù hợp bộ lọc.'), tổng = 0\n"
     "- KHÔNG lộ tài khoản của công ty khác (kể cả bản ghi có company_id = NULL)\n"
     "- Khi Lưu: API trả 422 với thông báo 'Tài khoản đăng nhập chưa gắn công ty, không thể thao tác' và hiện toast lỗi\n- Không tạo bản ghi",
     "searchByFilter: company null → whereRaw('1 = 0'); store: chặn 422 trước khi ghi"),

    ("06", "Ẩn nút thao tác theo quyền trên giao diện", "P1",
     "So sánh giao diện giữa user A (có quyền) và trạng thái khi biến canManage = false",
     "1. Đăng nhập user A, chụp lại giao diện danh sách\n2. Đối chiếu 3 vị trí: nút 'Thêm tài khoản', nút Sửa trên dòng, nút khóa/mở khóa",
     "canManage = true / false",
     "- Khi có quyền: cả 3 nút hiển thị\n- Khi không có quyền: cả 3 nút bị ẩn, chỉ còn nút Xem\n- Nút Xem luôn hiển thị",
     "FE: v-if=\"canManage\" trên nút Thêm / Sửa / toggle khóa"),
]

SECTIONS = [
    # ---------------------------------------------------------------- I
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Truy cập màn hình từ menu Tài chính", "P0",
         "User có quyền quản lý; công ty có ≥ 12 tài khoản ngân hàng",
         "1. Đăng nhập\n2. Vào phân hệ Tài chính → Danh mục → 'Danh mục tài khoản ngân hàng'\n3. Quan sát toàn bộ layout",
         "User: có quyền",
         "- URL = /finance/account-banks\n- Tiêu đề trang = 'Danh mục tài khoản ngân hàng'\n"
         "- Hiển thị panel 'Bộ lọc danh mục tài khoản ngân hàng' (đang thu gọn)\n"
         "- Hiển thị bảng 'Danh sách tài khoản ngân hàng'\n- Nút 'Thêm tài khoản' nằm ở khu vực dưới bảng\n- Không lỗi console",
         "Route index gắn checkPermission; menu isShow theo cùng quyền"),

        ("002", "Kiểm tra đủ 7 cột của bảng", "P0",
         "Đang ở màn danh mục tài khoản ngân hàng, có ≥ 1 bản ghi",
         "1. Quan sát dòng tiêu đề bảng\n2. Đối chiếu tên và thứ tự cột",
         "—",
         "Hiển thị đúng thứ tự 7 cột:\n1. STT\n2. Số tài khoản\n3. Loại tiền tệ\n4. Chủ tài khoản\n5. Ngân hàng\n6. Chi nhánh\n7. Trạng thái\n"
         "- KHÔNG cột nào có icon sắp xếp",
         "tableColumns không khai báo sortable cho bất kỳ cột nào"),

        ("003", "Cột STT và Số tài khoản được ghim khi cuộn ngang", "P1",
         "Cửa sổ trình duyệt thu hẹp để bảng phải cuộn ngang",
         "1. Thu nhỏ cửa sổ tới khi bảng xuất hiện thanh cuộn ngang\n2. Cuộn ngang sang phải\n3. Quan sát 2 cột đầu",
         "Độ rộng cửa sổ ~1024px",
         "- Cột STT và Số tài khoản luôn dính bên trái, không trôi theo\n- Các cột còn lại cuộn bình thường\n- Nội dung 2 cột ghim không bị đè lên cột khác",
         "tableColumns: sticky = true cho index và account_number"),

        ("004", "Trạng thái loading khi tải dữ liệu", "P2",
         "Bật throttle Slow 3G trong DevTools",
         "1. Bật Slow 3G\n2. Reload /finance/account-banks\n3. Quan sát vùng bảng khi API chưa trả",
         "Network: Slow 3G",
         "- Bảng ở trạng thái đang tải trong lúc chờ API\n- Sau khi API trả về: loading tắt, dữ liệu hiển thị\n- Không hiện đồng thời loading và dòng 'không có dữ liệu'",
         "loading = true trước gọi API, finally = false"),

        ("005", "Hiển thị khi công ty chưa có tài khoản nào", "P1",
         "Công ty của user chưa có bản ghi nào trong company_accounts (hoặc lọc bằng từ khoá chắc chắn không khớp)",
         "1. Nhập ô tìm nhanh 'zzzkhongtontai'\n2. Bấm Tìm kiếm\n3. Quan sát bảng",
         "keyword = 'zzzkhongtontai'",
         "- Hiển thị dòng 'Không có dữ liệu phù hợp bộ lọc.'\n- Tổng số bản ghi = 0, dòng 'Hiển thị 0–0 / 0'\n- Nút 'Thêm tài khoản' vẫn hiển thị và bấm được",
         "emptyText = 'Không có dữ liệu phù hợp bộ lọc.'"),

        ("006", "Mở rộng / thu gọn bộ lọc nâng cao", "P1",
         "Đang ở màn danh mục, panel lọc thu gọn (mặc định)",
         "1. Bấm mở rộng bộ lọc\n2. Quan sát các trường\n3. Bấm thu gọn lại",
         "—",
         "- Khi mở: hiện 5 trường 'Số tài khoản', 'Chủ tài khoản', 'Ngân hàng', 'Chi nhánh', 'Trạng thái'\n"
         "- Khi thu gọn: 5 trường ẩn, ô tìm nhanh vẫn còn\n- Giá trị đã nhập không mất khi thu gọn rồi mở lại",
         "filterCollapsed mặc định true; dùng v-show nên giữ giá trị"),
    ]),

    # ---------------------------------------------------------------- II
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm nhanh theo SỐ TÀI KHOẢN", "P0",
         "Công ty có tài khoản số '19001234567' (chủ TK 'CONG TY ABC', NH 'VIETCOMBANK') và 4 tài khoản khác",
         "1. Nhập ô tìm nhanh: '19001234'\n2. Bấm Tìm kiếm",
         "keyword = '19001234'",
         "- Chỉ hiển thị tài khoản 19001234567\n- Tổng số bản ghi = 1\n- Về trang 1",
         "keyword quét account_number LIKE %kw%"),

        ("002", "Tìm nhanh theo CHỦ TÀI KHOẢN", "P0",
         "Công ty có tài khoản chủ TK = 'CONG TY ABC'",
         "1. Nhập ô tìm nhanh: 'ABC'\n2. Bấm Tìm kiếm",
         "keyword = 'ABC'",
         "- Trả về các tài khoản có chủ tài khoản chứa 'ABC'\n- Cột 'Chủ tài khoản' hiển thị chữ in hoa 'CONG TY ABC'",
         "keyword quét thêm account_name"),

        ("003", "Tìm nhanh theo TÊN NGÂN HÀNG", "P0",
         "Công ty có tài khoản gắn ngân hàng 'VIETCOMBANK' và tài khoản gắn 'BIDV'",
         "1. Nhập ô tìm nhanh: 'VIETCOM'\n2. Bấm Tìm kiếm",
         "keyword = 'VIETCOM'",
         "- Chỉ hiển thị tài khoản của VIETCOMBANK\n- Không hiển thị tài khoản BIDV",
         "keyword quét thêm bank_name (cột denormalized)"),

        ("004", "Tìm nhanh KHÔNG quét tên chi nhánh", "P1",
         "Công ty có tài khoản chi nhánh = 'CN HOAN KIEM', số TK/chủ TK/ngân hàng đều không chứa chữ 'HOAN KIEM'",
         "1. Nhập ô tìm nhanh: 'HOAN KIEM'\n2. Bấm Tìm kiếm",
         "keyword = 'HOAN KIEM'",
         "- Không trả về bản ghi nào (tìm nhanh chỉ quét số TK, chủ TK, ngân hàng)\n- Muốn tìm theo chi nhánh phải dùng ô 'Chi nhánh' ở bộ lọc nâng cao",
         "keyword không bao gồm bank_branch"),

        ("005", "Lọc theo ô Số tài khoản (khớp một phần)", "P0",
         "Công ty có 2 tài khoản: '19001234567' và '19009999999'",
         "1. Mở bộ lọc nâng cao\n2. Nhập Số tài khoản = '1900'\n3. Bấm Tìm kiếm",
         "account_number = '1900'",
         "- Trả về CẢ 2 tài khoản (khớp một phần, không cần nhập đủ số)\n- Tổng = 2",
         "HRM đổi từ so khớp chính xác của ERP sang LIKE %…%"),

        ("006", "Lọc theo ô Chủ tài khoản", "P0",
         "Công ty có chủ TK 'CONG TY ABC' và 'CONG TY XYZ'",
         "1. Mở bộ lọc nâng cao\n2. Nhập Chủ tài khoản = 'XYZ'\n3. Bấm Tìm kiếm",
         "account_name = 'XYZ'",
         "- Chỉ trả về tài khoản của CONG TY XYZ\n- Cột Chủ tài khoản khớp chuỗi tìm kiếm",
         "account_name LIKE %…%"),

        ("007", "Lọc theo ô Ngân hàng", "P0",
         "Công ty có tài khoản ở VIETCOMBANK (2 TK) và BIDV (1 TK)",
         "1. Nhập ô Ngân hàng = 'BIDV'\n2. Bấm Tìm kiếm",
         "bank_name = 'BIDV'",
         "- Trả về đúng 1 tài khoản của BIDV\n- Cột Ngân hàng hiển thị 'BIDV'",
         "bank_name LIKE %…% (cột denormalized, không join banks)"),

        ("008", "Lọc theo ô Chi nhánh", "P0",
         "Công ty có tài khoản chi nhánh 'CN HOAN KIEM' và 'CN CAU GIAY'",
         "1. Nhập ô Chi nhánh = 'CAU GIAY'\n2. Bấm Tìm kiếm",
         "bank_branch = 'CAU GIAY'",
         "- Chỉ trả về tài khoản có chi nhánh CN CAU GIAY",
         "bank_branch LIKE %…%"),

        ("009", "Lọc theo Trạng thái = Hoạt động", "P0",
         "Công ty có 6 tài khoản status = 1 và 2 tài khoản status = 0",
         "1. Mở bộ lọc nâng cao\n2. Chọn Trạng thái = 'Hoạt động'\n3. Bấm Tìm kiếm",
         "status = '1'",
         "- Trả về đúng 6 bản ghi\n- Mọi dòng có nhãn trạng thái 'Hoạt động'",
         "status = 1"),

        ("010", "Lọc theo Trạng thái = Khóa (giá trị '0')", "P0",
         "Công ty có 6 tài khoản status = 1 và 2 tài khoản status = 0",
         "1. Chọn Trạng thái = 'Khóa'\n2. Bấm Tìm kiếm\n3. Quan sát kết quả",
         "status = '0'",
         "- Trả về ĐÚNG 2 bản ghi khóa (giá trị '0' phải được nhận, không bị coi là rỗng)\n- Mọi dòng có nhãn 'Khóa'",
         "BE dùng filled() nên chuỗi '0' vẫn được áp; nếu trả về toàn bộ danh sách → Failed"),

        ("011", "Kết hợp nhiều bộ lọc (AND)", "P0",
         "Công ty có: TK '19001234567' NH VIETCOMBANK status 1; TK '19007777777' NH VIETCOMBANK status 0",
         "1. Nhập Ngân hàng = 'VIETCOMBANK'\n2. Chọn Trạng thái = 'Hoạt động'\n3. Bấm Tìm kiếm",
         "bank_name = 'VIETCOMBANK', status = '1'",
         "- Chỉ trả về TK 19001234567\n- Không trả về TK đang khóa",
         "Các điều kiện lọc nối bằng AND"),

        ("012", "Kết hợp tìm nhanh + bộ lọc nâng cao", "P0",
         "Công ty có TK '19001234567' (VIETCOMBANK, status 1) và TK '19001239999' (BIDV, status 1)",
         "1. Nhập ô tìm nhanh = '1900123'\n2. Nhập ô Ngân hàng = 'BIDV'\n3. Bấm Tìm kiếm",
         "keyword = '1900123', bank_name = 'BIDV'",
         "- Chỉ trả về TK 19001239999\n- Điều kiện OR của tìm nhanh được bọc trong ngoặc, không phá vế AND của bộ lọc ngân hàng",
         "BE bọc nhóm keyword bằng where(function($q){...}) — kiểm chứng nhóm điều kiện đúng"),

        ("013", "Nút Đặt lại (reset) bộ lọc", "P0",
         "Đã nhập tìm nhanh + 4 ô text + trạng thái; đang ở trang 3",
         "1. Bấm nút Đặt lại trên panel lọc\n2. Quan sát các ô và bảng",
         "—",
         "- Toàn bộ ô lọc về rỗng, Trạng thái về placeholder 'Chọn trạng thái'\n"
         "- Danh sách tải lại đầy đủ, về trang 1\n- Chỉ phát sinh 1 request GET",
         "handleReset(): filters về initialStateForm + currentPage = 1 + loadData()"),

        ("014", "Đổi bộ lọc tự động tải lại và về trang 1", "P0",
         "Công ty có ≥ 25 tài khoản; đang đứng ở trang 3",
         "1. Ở trang 3, chọn Trạng thái = 'Hoạt động' (không bấm Tìm kiếm)\n2. Quan sát",
         "status = '1', trang trước = 3",
         "- Danh sách tự tải lại ngay khi đổi lựa chọn\n- Về trang 1\n- Không xảy ra bảng trắng do trang vượt quá tổng số trang mới",
         "watch filters deep → currentPage = 1 + loadData()"),

        ("015", "Gõ ô tìm nhanh KHÔNG tự gọi API cho tới khi bấm Tìm kiếm", "P1",
         "Đang ở màn danh mục, mở tab Network",
         "1. Gõ từng ký tự '1900' vào ô tìm nhanh, KHÔNG bấm Tìm kiếm\n2. Quan sát tab Network\n3. Bấm Tìm kiếm",
         "keyword gõ 4 ký tự",
         "- Trong bước 1: KHÔNG phát sinh request nào (tránh gọi API mỗi ký tự)\n- Bước 3: phát sinh đúng 1 request với keyword = '1900'",
         "keyword nằm trong ignoredFields của deep watcher"),

        ("016", "Chống gọi API trùng lặp", "P1",
         "Đang ở màn danh mục, mở tab Network",
         "1. Bấm nút Tìm kiếm 3 lần liên tiếp thật nhanh (trong ~1 giây), không đổi tham số\n2. Đếm số request GET",
         "3 lần bấm < 800ms",
         "- Chỉ phát sinh 1 request GET /finance/account-banks\n- Danh sách hiển thị đúng, không nhấp nháy nhiều lần",
         "DedupeLoadMixin chặn request trùng tham số trong 800ms"),

        ("017", "Ghi nhớ bộ lọc khi rời trang rồi quay lại", "P1",
         "Đang lọc: Ngân hàng = 'BIDV', Trạng thái = 'Hoạt động'",
         "1. Chuyển sang màn khác trong hệ thống\n2. Quay lại /finance/account-banks trong vòng 10 phút\n3. Quan sát ô lọc và kết quả",
         "Thời gian rời trang < 10 phút",
         "- Các ô lọc được khôi phục đúng: Ngân hàng = 'BIDV', Trạng thái = 'Hoạt động'\n"
         "- Trạng thái mở/thu gọn của panel lọc cũng được khôi phục\n- Danh sách hiển thị đúng kết quả đã lọc",
         "filterStateMixin: key 'finance_account_banks', hạn 10 phút"),

        ("018", "Bộ lọc ghi nhớ hết hạn sau 10 phút", "P2",
         "Đã lọc Ngân hàng = 'BIDV', sau đó rời trang > 10 phút",
         "1. Lọc rồi rời trang\n2. Chờ quá 10 phút (hoặc chỉnh thời gian lưu trong trình duyệt)\n3. Quay lại màn",
         "Thời gian rời trang > 10 phút",
         "- Bộ lọc KHÔNG được khôi phục, các ô về rỗng\n- Danh sách hiển thị toàn bộ tài khoản của công ty",
         "expirationTime = 10 × 60 × 1000 ms"),

        ("019", "Lọc bằng chuỗi có khoảng trắng đầu/cuối", "P2",
         "Công ty có chủ TK = 'CONG TY ABC'",
         "1. Nhập ô Chủ tài khoản = '  ABC  '\n2. Bấm Tìm kiếm",
         "account_name = '  ABC  '",
         "- Ghi nhận thực tế: nếu trả về 'CONG TY ABC' → Passed\n- Nếu trả 0 bản ghi do không cắt khoảng trắng → ghi Failed kèm mô tả",
         "BE ghép LIKE trực tiếp, không trim giá trị lọc"),

        ("020", "Lọc theo chữ thường vẫn khớp dữ liệu in hoa", "P1",
         "Công ty có chủ TK lưu dạng in hoa 'CONG TY ABC'",
         "1. Nhập ô Chủ tài khoản = 'cong ty abc' (chữ thường)\n2. Bấm Tìm kiếm",
         "account_name = 'cong ty abc'",
         "- Vẫn trả về đúng tài khoản 'CONG TY ABC' (MySQL so khớp không phân biệt hoa/thường theo collation mặc định)\n- Tổng ≥ 1",
         "Dữ liệu luôn lưu in hoa do hook saving — người dùng thường gõ chữ thường"),
    ]),

    # ---------------------------------------------------------------- III
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "Đang ở màn /finance/account-banks",
         "1. Quan sát vùng phía trên bảng dữ liệu",
         "—",
         "- Màn KHÔNG có thẻ thống kê nào\n- Chỉ có panel lọc và bảng dữ liệu\n"
         "- Số liệu duy nhất là dòng 'Hiển thị x–y / z' ở vùng phân trang",
         "Section giữ lại theo chuẩn tài liệu — feature không có stats"),
    ]),

    # ---------------------------------------------------------------- IV
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Thứ tự mặc định — bản ghi mới nhất lên đầu", "P0",
         "Công ty có ≥ 5 tài khoản với created_at khác nhau",
         "1. Vào màn danh mục\n2. Đối chiếu thứ tự hiển thị với created_at trong DB",
         "—",
         "- Bản ghi được tạo gần nhất đứng đầu danh sách\n- Thứ tự giảm dần theo thời điểm tạo",
         "orderBy('company_accounts.created_at', 'desc')"),

        ("002", "Bản ghi vừa thêm nhảy lên đầu danh sách", "P0",
         "Đang ở trang 1 danh sách",
         "1. Thêm mới 1 tài khoản hợp lệ\n2. Quan sát vị trí bản ghi mới sau khi danh sách tải lại",
         "Tài khoản mới tạo",
         "- Bản ghi mới đứng ở dòng đầu tiên (STT = 1)\n- Tổng số bản ghi tăng đúng 1",
         "Sắp xếp created_at desc"),

        ("003", "Không có cột nào cho phép sắp xếp", "P1",
         "Đang ở màn danh mục",
         "1. Click lần lượt vào tiêu đề cả 7 cột\n2. Quan sát thứ tự dữ liệu và tab Network",
         "—",
         "- Không cột nào đổi thứ tự dữ liệu\n- Không hiện icon sắp xếp\n- Không phát sinh request thừa",
         "tableColumns không khai báo sortable"),

        ("004", "Phân trang — chuyển sang trang kế tiếp", "P0",
         "Công ty có 25 tài khoản, số dòng/trang = 10",
         "1. Ở trang 1, ghi lại số TK dòng đầu\n2. Sang trang 2\n3. Quan sát STT và dữ liệu",
         "total = 25, pageSize = 10",
         "- STT dòng đầu trang 2 = 11\n- Dữ liệu khác hoàn toàn trang 1\n- Dòng thông tin hiển thị 'Hiển thị 11–20 / 25'",
         "STT = (currentPage − 1) × pageSize + index + 1"),

        ("005", "Phân trang — trang cuối và dòng 'Hiển thị x–y / z'", "P0",
         "Công ty có đúng 25 tài khoản, 10 dòng/trang",
         "1. Chuyển tới trang 3 (trang cuối)\n2. Đếm số dòng và đọc dòng thông tin phân trang",
         "total = 25, page = 3",
         "- Trang 3 có đúng 5 dòng, STT từ 21 đến 25\n- Dòng thông tin: 'Hiển thị 21–25 / 25'\n- Nút sang trang kế bị vô hiệu",
         "from/to tính ở FE vì API không trả; to = (page−1) × size + số dòng thực tế"),

        ("006", "Đổi số dòng trên mỗi trang", "P0",
         "Công ty có ≥ 60 tài khoản; đang ở trang 3, 10 dòng/trang",
         "1. Đổi số dòng/trang sang 50\n2. Quan sát danh sách và dòng thông tin phân trang",
         "pageSize: 10 → 50",
         "- Hiển thị tối đa 50 dòng, tự về trang 1\n- STT dòng đầu = 1\n"
         "- Dòng thông tin hiển thị đúng 'Hiển thị 1–50 / <tổng>' (không bị lệch do kiểu chuỗi của per_page)",
         "FE ép Number() cho perPage/total/currentPage vì Laravel trả per_page dạng chuỗi"),

        ("007", "Đổi trang KHÔNG bị coi là đổi bộ lọc", "P0",
         "Đang lọc Ngân hàng = 'VIETCOMBANK', kết quả ≥ 15 dòng, mở tab Network",
         "1. Bấm sang trang 2\n2. Đếm số request phát sinh và kiểm tra tham số",
         "page: 1 → 2",
         "- Chỉ 1 request GET với page = 2 và giữ nguyên bank_name = 'VIETCOMBANK'\n"
         "- KHÔNG bị reset về trang 1\n- Bộ lọc không bị xoá",
         "page/per_page để ngoài object filters nên deep watcher không kích hoạt"),

        ("008", "Hiển thị mã loại tiền tệ", "P1",
         "Tài khoản A gắn currency_id của 'VND'; tài khoản B có currency_id = null (bản ghi cũ do ERP tạo)",
         "1. Tìm dòng A và B\n2. Quan sát cột 'Loại tiền tệ'",
         "A: VND; B: null",
         "- Dòng A hiển thị mã tiền tệ 'VND'\n- Dòng B hiển thị '—'\n- Không hiển thị 'null' hay ô trống",
         "currency_text = currencies.code hoặc '—'"),

        ("009", "Chủ tài khoản và Ngân hàng hiển thị chữ IN HOA", "P0",
         "Thêm mới tài khoản với Chủ tài khoản nhập chữ thường 'cong ty test'",
         "1. Thêm tài khoản, nhập Chủ tài khoản = 'cong ty test'\n2. Lưu\n3. Quan sát dòng trên danh sách",
         "account_name nhập = 'cong ty test'",
         "- Cột 'Chủ tài khoản' hiển thị 'CONG TY TEST'\n- Cột 'Ngân hàng' cũng hiển thị in hoa\n"
         "- Mở lại modal Sửa: giá trị nạp lên cũng là chữ in hoa",
         "Hook saving: mb_strtoupper cho account_name và bank_name"),

        ("010", "Nhãn trạng thái Hoạt động / Khóa", "P0",
         "Tài khoản A status = 1, tài khoản B status = 0",
         "1. Quan sát cột Trạng thái của A và B",
         "A: 1; B: 0",
         "- A: nhãn 'Hoạt động' (màu thương hiệu), nút cạnh là icon ổ khóa (tooltip 'Khóa tài khoản')\n"
         "- B: nhãn 'Khóa' (màu cảnh báo), nút cạnh là icon mở khóa (tooltip 'Mở khóa tài khoản')",
         "status_text: 1 → 'Hoạt động', 0 → 'Khóa'"),

        ("011", "Nút Xem và Sửa nằm trong ô Số tài khoản", "P1",
         "Có ≥ 1 tài khoản trên danh sách; user có quyền quản lý",
         "1. Rê chuột vào ô 'Số tài khoản' của 1 dòng\n2. Quan sát các nút",
         "canManage = true",
         "- Có 2 nút: Xem (icon con mắt) và Sửa (icon bút)\n- KHÔNG có nút Xóa (màn không hỗ trợ xóa)\n- Tooltip tương ứng 'Xem' và 'Sửa'",
         "Chốt thiết kế: không có chức năng Xóa, giống ERP"),

        ("012", "Tài khoản đang Khóa vẫn Sửa được", "P0",
         "Tài khoản B đang status = 0 (Khóa)",
         "1. Tìm dòng B\n2. Rê chuột vào nút Sửa\n3. Bấm Sửa",
         "B: status = 0",
         "- Nút Sửa KHÔNG bị mờ\n- Modal 'Sửa tài khoản ngân hàng' mở bình thường, ô Trạng thái hiển thị 'Khóa'\n"
         "- Khác với màn Danh mục ngân hàng (ở đó bản ghi khóa bị chặn sửa)",
         "FE không gắn điều kiện disable theo status ở nút Sửa"),

        ("013", "Nội dung dài xuống dòng không vỡ bảng", "P2",
         "Có tài khoản với Chủ tài khoản dài ~120 ký tự và tên chi nhánh dài",
         "1. Tìm dòng đó trên danh sách\n2. Quan sát cột Chủ tài khoản, Ngân hàng, Chi nhánh",
         "Chuỗi ~120 ký tự",
         "- Nội dung tự xuống dòng trong ô, không tràn ngang\n- Chiều cao dòng giãn ra, không đè lên dòng khác\n- Trang không xuất hiện cuộn ngang bất thường",
         "3 cột này có cellClass = 'text-wrap'"),
    ]),

    # ---------------------------------------------------------------- V
    ("V", "CHỨC NĂNG CHÍNH (CRUD / ACTION)", [
        # --- Thêm mới
        ("001", "Mở modal Thêm tài khoản", "P0",
         "User có quyền, đang ở màn danh mục",
         "1. Bấm nút 'Thêm tài khoản'\n2. Quan sát modal",
         "—",
         "- Mở modal tiêu đề 'Thêm tài khoản ngân hàng'\n"
         "- 6 trường: Số tài khoản*, Loại tiền tệ*, Chủ tài khoản*, Ngân hàng*, Chi nhánh*, Trạng thái\n"
         "- 5 trường bắt buộc có dấu * đỏ\n- Trạng thái mặc định = 'Hoạt động'\n"
         "- Ô Chi nhánh đang bị vô hiệu (chưa chọn ngân hàng)\n- Footer có nút Lưu và Đóng\n- Chưa hiện lỗi đỏ nào",
         "form.status mặc định '1'; Chi nhánh disabled khi !form.bank_id"),

        ("002", "Dropdown Ngân hàng chỉ liệt kê ngân hàng đang hoạt động", "P0",
         "Danh mục ngân hàng có: VIETCOMBANK (hoạt động), BIDV (hoạt động), TESTLOCK (đang Khóa)",
         "1. Mở modal Thêm tài khoản\n2. Mở dropdown Ngân hàng\n3. Tìm 'TESTLOCK'",
         "TESTLOCK: banks.status = 2",
         "- Danh sách có VIETCOMBANK và BIDV\n- KHÔNG có TESTLOCK\n- Danh sách sắp xếp theo tên ngân hàng",
         "options(): banks where status = 1, orderBy name"),

        ("003", "Chi nhánh lọc theo ngân hàng đã chọn", "P0",
         "VIETCOMBANK có 3 chi nhánh; BIDV có 2 chi nhánh khác",
         "1. Mở modal Thêm\n2. Chọn Ngân hàng = 'VIETCOMBANK'\n3. Mở dropdown Chi nhánh\n4. Đổi Ngân hàng sang 'BIDV'\n5. Mở lại dropdown Chi nhánh",
         "VCB: 3 CN; BIDV: 2 CN",
         "- Sau bước 2: ô Chi nhánh được bật\n- Bước 3: chỉ hiện đúng 3 chi nhánh của VIETCOMBANK\n"
         "- Bước 5: chỉ hiện 2 chi nhánh của BIDV, không lẫn chi nhánh VCB",
         "branchOptionsForBank lọc theo bank_id ở client"),

        ("004", "Đổi ngân hàng làm reset chi nhánh đã chọn", "P0",
         "Đang mở modal Thêm, đã chọn VIETCOMBANK + chi nhánh 'CN HOAN KIEM'",
         "1. Đổi Ngân hàng sang 'BIDV'\n2. Quan sát ô Chi nhánh",
         "bank: VCB → BIDV",
         "- Ô Chi nhánh bị xóa trắng (không giữ 'CN HOAN KIEM' của ngân hàng cũ)\n- Bắt buộc chọn lại chi nhánh mới\n- Không lưu được khi chưa chọn lại",
         "watch form.bank_id: reset bank_branch_id nếu chi nhánh không thuộc ngân hàng mới"),

        ("005", "Thêm mới tài khoản đủ trường bắt buộc", "P0",
         "Chưa tồn tại tài khoản số '19008888888' trong toàn bảng company_accounts",
         "1. Bấm 'Thêm tài khoản'\n2. Số tài khoản = '19008888888'\n3. Loại tiền tệ = 'VND'\n"
         "4. Chủ tài khoản = 'cong ty kiem thu'\n5. Ngân hàng = 'VIETCOMBANK', Chi nhánh = 'CN HOAN KIEM'\n"
         "6. Giữ Trạng thái = 'Hoạt động'\n7. Bấm Lưu",
         "6 trường như bước thực hiện",
         "- Toast xanh 'Thêm mới thành công'\n- Modal đóng, danh sách tải lại\n"
         "- Dòng mới đứng đầu: số TK '19008888888', tiền tệ 'VND', chủ TK 'CONG TY KIEM THU' (in hoa), NH 'VIETCOMBANK', CN 'CN HOAN KIEM', trạng thái 'Hoạt động'",
         "POST /v1/finance/account-banks; company_id gán tự động từ user login"),

        ("006", "Thêm mới với trạng thái = Khóa", "P1",
         "Chưa tồn tại tài khoản số '19006666666'",
         "1. Thêm tài khoản, nhập đủ thông tin\n2. Đổi Trạng thái sang 'Khóa'\n3. Bấm Lưu",
         "status = '0'",
         "- Lưu thành công\n- Dòng mới hiển thị nhãn 'Khóa'\n- Nút cạnh trạng thái là icon mở khóa",
         "status nhận 0 hoặc 1"),

        ("007", "Tên ngân hàng và chi nhánh được lưu theo id đã chọn", "P0",
         "Vừa thêm tài khoản với Ngân hàng = 'VIETCOMBANK', Chi nhánh = 'CN HOAN KIEM'",
         "1. Sau khi lưu, kiểm tra bản ghi trong bảng company_accounts\n2. Đối chiếu 4 cột bank_id, bank_name, bank_branch_id, bank_branch",
         "—",
         "- bank_id và bank_branch_id lưu đúng id đã chọn\n"
         "- bank_name lưu tên ngân hàng dạng IN HOA, bank_branch lưu tên chi nhánh tương ứng\n- Không có cột nào để trống",
         "createOrUpdate() fill bank_name/bank_branch từ id trước khi ghi"),

        ("008", "company_id được gán tự động theo user đăng nhập", "P0",
         "User A thuộc công ty 1",
         "1. Đăng nhập user A, thêm 1 tài khoản mới\n2. Kiểm tra cột company_id của bản ghi vừa tạo",
         "User A: company_id = 1",
         "- company_id = 1 (công ty của user A)\n- Form KHÔNG có ô chọn công ty (người dùng không tự chọn được)",
         "createOrUpdate(): company_id = currentCompanyId()"),

        # --- Sửa
        ("009", "Mở modal Sửa và kiểm tra dữ liệu nạp sẵn", "P0",
         "Tài khoản '19008888888' đã tồn tại với đủ thông tin",
         "1. Bấm nút Sửa trên dòng đó\n2. Quan sát toàn bộ trường",
         "id của TK 19008888888",
         "- Tiêu đề 'Sửa tài khoản ngân hàng'\n- 6 trường điền sẵn đúng dữ liệu hiện tại (kể cả Loại tiền tệ và Trạng thái)\n"
         "- Ô Chi nhánh đã bật và hiển thị đúng chi nhánh đang gắn (không bị xóa trắng khi mở)\n- Các ô cho phép chỉnh sửa",
         "loadDetail(); watch bank_id so sánh trước khi reset để không xoá giá trị vừa nạp"),

        ("010", "Sửa và lưu thành công", "P0",
         "Tài khoản '19008888888' có Chủ tài khoản = 'CONG TY KIEM THU'",
         "1. Bấm Sửa\n2. Đổi Chủ tài khoản thành 'cong ty kiem thu moi'\n3. Bấm Lưu\n4. Quan sát danh sách",
         "account_name đổi",
         "- Toast 'Cập nhật thành công', modal đóng\n- Dòng hiển thị 'CONG TY KIEM THU MOI' (in hoa)\n"
         "- Không tạo thêm bản ghi mới, tổng số không đổi",
         "PUT /v1/finance/account-banks/{id}"),

        ("011", "Sửa giữ nguyên số tài khoản không báo trùng", "P0",
         "Tài khoản '19008888888' đã tồn tại",
         "1. Bấm Sửa dòng đó\n2. Giữ nguyên Số tài khoản\n3. Chỉ đổi Chi nhánh sang chi nhánh khác cùng ngân hàng\n4. Bấm Lưu",
         "account_number giữ nguyên",
         "- Lưu thành công, KHÔNG báo 'Số tài khoản đã tồn tại'\n- Cột Chi nhánh cập nhật tên chi nhánh mới",
         "Rule unique có ignore theo id đang sửa"),

        ("012", "Sửa làm mới lại tên ngân hàng đã đổi ở danh mục", "P1",
         "Tài khoản gắn ngân hàng X; sau đó vào Danh mục ngân hàng đổi tên X thành 'NGAN HANG X MOI'",
         "1. Đổi tên ngân hàng X ở màn Danh mục ngân hàng\n2. Quay lại màn tài khoản ngân hàng, quan sát cột Ngân hàng\n"
         "3. Mở Sửa tài khoản đó và bấm Lưu (không đổi gì)\n4. Quan sát lại cột Ngân hàng",
         "Tên ngân hàng đổi ở danh mục gốc",
         "- Bước 2: cột Ngân hàng VẪN hiển thị tên cũ (do lưu tên tại thời điểm ghi)\n"
         "- Bước 4: sau khi Lưu lại, cột Ngân hàng cập nhật thành 'NGAN HANG X MOI'",
         "bank_name là dữ liệu denormalized — chỉ refresh khi lưu lại bản ghi"),

        ("013", "Sửa tài khoản có ngân hàng đã bị khóa", "P0",
         "Tài khoản D đang gắn ngân hàng 'TESTLOCK'; ngân hàng này vừa bị Khóa ở màn Danh mục ngân hàng",
         "1. Bấm Sửa dòng D\n2. Quan sát ô Ngân hàng và thông báo\n3. Thử bấm Lưu ngay",
         "bank của D: status = 2",
         "- Ô Ngân hàng bị xóa trắng, hiện toast đỏ 'Ngân hàng của tài khoản này đã bị khóa, vui lòng chọn ngân hàng khác'\n"
         "- Ô Chi nhánh cũng trống và bị vô hiệu\n- Bấm Lưu: báo lỗi 'Bắt buộc phải nhập' ở Ngân hàng, không lưu được\n"
         "- Chọn ngân hàng khác + chi nhánh rồi Lưu → thành công",
         "options chỉ trả bank status = 1; FE clear bank_id nếu không còn trong options"),

        # --- Xem
        ("014", "Xem chi tiết — mọi trường chỉ đọc", "P0",
         "Tài khoản '19008888888' đầy đủ thông tin",
         "1. Bấm nút Xem trên dòng đó\n2. Thử click và gõ vào từng ô\n3. Quan sát footer",
         "isView = true",
         "- Tiêu đề 'Xem chi tiết tài khoản ngân hàng'\n- Cả 6 ô bị vô hiệu, không gõ / không mở dropdown được\n"
         "- Footer CHỈ có nút 'Đóng', không có nút Lưu\n- Dữ liệu hiển thị đúng bản ghi",
         "isView → :disabled toàn bộ + ẩn nút Lưu"),

        ("015", "Xem tài khoản có ngân hàng đã bị khóa vẫn thấy tên ngân hàng", "P0",
         "Tài khoản D gắn ngân hàng 'TESTLOCK' đang bị Khóa",
         "1. Bấm nút Xem dòng D\n2. Quan sát ô Ngân hàng và ô Chi nhánh",
         "bank: status = 2",
         "- Ô Ngân hàng hiển thị tên 'TESTLOCK' (không bị trống, không hiện '—')\n"
         "- Ô Chi nhánh hiển thị đúng chi nhánh đang gắn\n- Khác với chế độ Sửa (ở đó bị xóa trắng)",
         "bankOptionsForSelect: chế độ Xem append option tạm từ bank_name của detail"),

        ("016", "Đóng modal không lưu thay đổi", "P1",
         "Tài khoản '19008888888' có Chủ tài khoản = 'CONG TY KIEM THU'",
         "1. Bấm Sửa\n2. Đổi Chủ tài khoản thành 'ABC XYZ'\n3. Bấm 'Đóng'\n4. Quan sát danh sách\n5. Mở lại modal Sửa",
         "Thay đổi không bấm Lưu",
         "- Modal đóng, không có toast thành công\n- Dòng vẫn hiển thị 'CONG TY KIEM THU'\n"
         "- Mở lại modal: dữ liệu nạp đúng giá trị cũ, không còn 'ABC XYZ', không còn lỗi đỏ",
         "@hide → reset() xoá form và cờ touched"),

        ("017", "Mở Thêm mới ngay sau khi mở Sửa — form phải sạch", "P0",
         "Vừa mở modal Sửa 1 tài khoản rồi đóng lại",
         "1. Bấm Sửa 1 dòng, quan sát dữ liệu\n2. Bấm 'Đóng'\n3. Bấm 'Thêm tài khoản'\n4. Quan sát các trường",
         "—",
         "- Modal mở với tiêu đề 'Thêm tài khoản ngân hàng'\n- Toàn bộ ô rỗng, không còn dữ liệu của bản ghi vừa xem\n"
         "- Trạng thái về mặc định 'Hoạt động'\n- Ô Chi nhánh bị vô hiệu trở lại",
         "open() gọi reset() trước khi hiển thị"),

        # --- Khóa / Mở khóa
        ("018", "Khóa một tài khoản đang hoạt động", "P0",
         "Tài khoản '19008888888' đang status = 1",
         "1. Bấm nút ổ khóa ở cột Trạng thái\n2. Đọc hộp xác nhận\n3. Bấm 'Khóa'",
         "status 1 → 0",
         "- Hộp xác nhận tiêu đề 'Xác nhận khóa', nội dung \"Bạn có chắc muốn khóa tài khoản '19008888888'?\", nút xác nhận ghi 'Khóa'\n"
         "- Toast 'Khóa thành công'\n- Dòng đổi sang nhãn 'Khóa', nút chuyển thành icon mở khóa\n- Bản ghi vẫn nằm trong danh sách",
         "GET /v1/finance/account-banks/{id}/lock → status = 0"),

        ("019", "Mở khóa một tài khoản đang bị khóa", "P0",
         "Tài khoản '19008888888' đang status = 0",
         "1. Bấm nút mở khóa\n2. Đọc hộp xác nhận\n3. Bấm 'Mở khóa'",
         "status 0 → 1",
         "- Hộp xác nhận tiêu đề 'Xác nhận mở khóa', nút xác nhận ghi 'Mở khóa'\n"
         "- Toast 'Mở khóa thành công'\n- Dòng trở về nhãn 'Hoạt động'",
         "GET /v1/finance/account-banks/{id}/unlock → status = 1"),

        ("020", "Hủy thao tác khóa", "P1",
         "Tài khoản '19008888888' đang status = 1",
         "1. Bấm nút ổ khóa\n2. Bấm 'Hủy'\n3. Quan sát dòng và tab Network",
         "—",
         "- Hộp xác nhận đóng\n- Trạng thái vẫn 'Hoạt động'\n- Không có toast\n- Không phát sinh request lock",
         "Chỉ gọi API khi bấm nút xác nhận"),

        ("021", "Khóa/mở khóa giữ nguyên bộ lọc và trang hiện tại", "P1",
         "Đang lọc Ngân hàng = 'VIETCOMBANK', đang ở trang 2",
         "1. Khóa 1 tài khoản trên trang 2\n2. Quan sát bộ lọc, trang và danh sách sau khi tải lại",
         "bank_name = 'VIETCOMBANK', page = 2",
         "- Bộ lọc vẫn giữ 'VIETCOMBANK'\n- Vẫn ở trang 2\n- Dòng vừa khóa hiển thị nhãn 'Khóa' đã cập nhật",
         "handleConfirmToggleLock: resetLoadDedupe() rồi loadData() với params hiện tại"),

        ("022", "Khóa bản ghi đã bị người khác xóa/đổi", "P1",
         "Tài khoản E đang hiển thị trên màn user A; bản ghi này vừa bị xóa khỏi DB (hoặc đổi company_id)",
         "1. Trên màn user A (chưa reload), bấm khóa dòng E\n2. Xác nhận 'Khóa'\n3. Quan sát",
         "Bản ghi không còn thuộc công ty user A",
         "- Hiện toast đỏ 'Dữ liệu đã thay đổi, vui lòng tải lại'\n- Không lỗi trắng màn hình\n- Danh sách vẫn thao tác được",
         "API trả 404 → FE map sang thông báo riêng"),

        ("023", "Không có chức năng Xóa trên toàn màn", "P1",
         "Đang ở màn danh mục tài khoản ngân hàng",
         "1. Rà toàn bộ nút trên danh sách và trong modal\n2. Thử gọi DELETE /v1/finance/account-banks/{id} bằng Postman",
         "—",
         "- Giao diện KHÔNG có nút Xóa ở bất kỳ đâu\n- API DELETE trả 404/405 (route không tồn tại)\n"
         "- Muốn ngừng dùng 1 tài khoản thì dùng chức năng Khóa",
         "Chốt thiết kế: bỏ Xóa giống ERP → không cần điều kiện is_can_delete"),
    ]),

    # ---------------------------------------------------------------- VI
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Bỏ trống toàn bộ trường bắt buộc", "P0",
         "Đang mở modal 'Thêm tài khoản ngân hàng', chưa nhập gì",
         "1. Bấm Lưu ngay",
         "Tất cả trường rỗng",
         "- Hiện lỗi 'Bắt buộc phải nhập' dưới CẢ 5 ô: Số tài khoản, Loại tiền tệ, Chủ tài khoản, Ngân hàng, Chi nhánh\n"
         "- 5 ô có viền đỏ\n- KHÔNG gửi request lên server\n- Modal không đóng",
         "validateLocal() chặn trước; touched = true nên lỗi mới hiện"),

        ("002", "Lỗi chỉ hiện SAU lần bấm Lưu đầu tiên", "P0",
         "Vừa mở modal Thêm tài khoản",
         "1. Quan sát ngay khi modal vừa mở\n2. Click vào ô Số tài khoản rồi click ra ngoài (không nhập)\n3. Bấm Lưu",
         "—",
         "- Bước 1 và 2: KHÔNG hiện lỗi đỏ nào, không viền đỏ\n- Bước 3: lỗi mới hiện đồng loạt ở các ô còn thiếu",
         "Cờ touched: chỉ hiện lỗi sau lần submit đầu"),

        ("003", "Bỏ trống riêng Số tài khoản", "P0",
         "Đang mở modal Thêm",
         "1. Nhập đủ Loại tiền tệ, Chủ tài khoản, Ngân hàng, Chi nhánh\n2. Để trống Số tài khoản\n3. Bấm Lưu",
         "account_number rỗng",
         "- Chỉ ô Số tài khoản báo 'Bắt buộc phải nhập' và viền đỏ\n- 4 ô còn lại không báo lỗi\n- Không gửi request",
         "Validate từng trường độc lập"),

        ("004", "Nhập toàn dấu cách vào trường bắt buộc", "P0",
         "Đang mở modal Thêm",
         "1. Nhập Số tài khoản = '   ' (3 dấu cách), Chủ tài khoản = '   '\n2. Nhập đủ 3 trường còn lại\n3. Bấm Lưu",
         "Chuỗi chỉ có khoảng trắng",
         "- Báo lỗi 'Bắt buộc phải nhập' ở cả Số tài khoản và Chủ tài khoản\n- Không tạo bản ghi rỗng",
         "validateLocal() dùng trim() trước khi kiểm tra"),

        ("005", "Khoảng trắng đầu/cuối bị cắt khi lưu", "P1",
         "Đang mở modal Thêm; chưa tồn tại số TK '19005555555'",
         "1. Nhập Số tài khoản = '  19005555555  ', Chủ tài khoản = '  cong ty co trim  '\n2. Nhập đủ trường còn lại\n3. Bấm Lưu\n4. Kiểm tra dữ liệu đã lưu",
         "Chuỗi có khoảng trắng 2 đầu",
         "- Lưu thành công\n- Dữ liệu lưu là '19005555555' và 'CONG TY CO TRIM' — không còn khoảng trắng thừa\n"
         "- Tìm kiếm theo '19005555555' ra đúng bản ghi",
         "FE trim trước khi gửi + BE trim lần nữa trong createOrUpdate()"),

        ("006", "Thêm mới trùng SỐ TÀI KHOẢN đã tồn tại", "P0",
         "Đã tồn tại tài khoản số '19008888888' trong công ty hiện tại",
         "1. Bấm Thêm tài khoản\n2. Nhập Số tài khoản = '19008888888', các trường khác hợp lệ\n3. Bấm Lưu",
         "account_number trùng",
         "- API trả 422\n- Lỗi đỏ dưới ô Số tài khoản: 'Số tài khoản đã tồn tại' (KHÔNG phải 'Tên tài khoản đã tồn tại')\n"
         "- Modal không đóng, không tạo bản ghi",
         "Rule unique + message đã sửa lại cho đúng ngữ nghĩa"),

        ("007", "Trùng số tài khoản với công ty KHÁC cũng bị chặn", "P0",
         "Công ty 2 đã có tài khoản số '19007777777'; user đang đăng nhập thuộc công ty 1 và công ty 1 chưa có số này",
         "1. Đăng nhập user công ty 1\n2. Thêm tài khoản số '19007777777', các trường khác hợp lệ\n3. Bấm Lưu",
         "Số TK đã tồn tại ở công ty khác",
         "- API trả 422, báo 'Số tài khoản đã tồn tại'\n- Không tạo bản ghi\n"
         "- Ghi nhận: quy tắc unique áp dụng TOÀN BẢNG, không tách theo công ty — xác nhận lại với nghiệp vụ nếu đây không phải mong muốn",
         "Rule::unique('company_accounts','account_number') không kèm điều kiện company_id"),

        ("008", "Sửa thành số tài khoản trùng bản ghi khác", "P0",
         "Có 2 tài khoản: '19008888888' và '19006666666'",
         "1. Bấm Sửa dòng '19006666666'\n2. Đổi Số tài khoản thành '19008888888'\n3. Bấm Lưu",
         "account_number đổi sang giá trị đã tồn tại",
         "- Lỗi 'Số tài khoản đã tồn tại' dưới ô Số tài khoản\n- Bản ghi giữ nguyên số cũ trên danh sách",
         "unique ignore chỉ bỏ qua chính bản ghi đang sửa"),

        ("009", "Chọn chi nhánh không thuộc ngân hàng (giả mạo qua API)", "P0",
         "Chi nhánh 'CN HOAN KIEM' (id = 10) thuộc VIETCOMBANK (id = 1); BIDV có id = 2",
         "1. Gọi POST /v1/finance/account-banks bằng Postman với bank_id = 2, bank_branch_id = 10, các trường khác hợp lệ\n2. Quan sát response",
         "bank_id = 2, bank_branch_id = 10",
         "- API trả 422 với message 'Chi nhánh không thuộc ngân hàng đã chọn'\n- Không tạo bản ghi\n"
         "- (Ghi chú: giao diện đã lọc sẵn nên chỉ tái hiện được qua API)",
         "Rule::exists('bank_branches','id')->where('bank_id', ...) — HRM bổ sung so với ERP"),

        ("010", "Gửi ngân hàng đang bị khóa qua API", "P0",
         "Ngân hàng 'TESTLOCK' id = 99 đang status = 2",
         "1. Gọi POST /v1/finance/account-banks với bank_id = 99, chi nhánh thuộc ngân hàng đó, trường khác hợp lệ\n2. Quan sát response",
         "bank_id = 99 (đã khóa)",
         "- API trả 422 với message 'Ngân hàng không tồn tại hoặc đã bị khóa'\n- Không tạo bản ghi",
         "Rule::exists('banks','id')->where('status', 1)"),

        ("011", "Thiếu loại tiền tệ khi CẬP NHẬT", "P0",
         "Tài khoản cũ do ERP tạo có currency_id = null",
         "1. Bấm Sửa tài khoản đó\n2. Quan sát ô Loại tiền tệ\n3. Bấm Lưu ngay khi chưa chọn tiền tệ\n4. Chọn 'VND' rồi Lưu lại",
         "currency_id = null",
         "- Bước 2: ô Loại tiền tệ để trống (không tự điền)\n"
         "- Bước 3: báo lỗi 'Bắt buộc phải nhập' ở ô Loại tiền tệ, không lưu được\n"
         "- Bước 4: lưu thành công, cột 'Loại tiền tệ' đổi từ '—' thành 'VND'",
         "HRM áp required currency_id cho CẢ update (ERP thiếu ở nhánh update)"),

        ("012", "Gửi loại tiền tệ không tồn tại qua API", "P1",
         "currency_id = 999999 không có trong bảng currencies",
         "1. Gọi POST /v1/finance/account-banks với currency_id = 999999, trường khác hợp lệ",
         "currency_id = 999999",
         "- API trả 422 với message 'Loại tiền tệ không tồn tại'\n- Không tạo bản ghi",
         "Rule: currency_id required|exists:currencies,id"),

        ("013", "Gửi trạng thái không hợp lệ qua API", "P1",
         "Đã có token hợp lệ và đủ quyền",
         "1. Gọi POST /v1/finance/account-banks với status = 5, các trường khác hợp lệ",
         "status = 5",
         "- API trả 422 với message 'Trạng thái không hợp lệ'\n- Không tạo bản ghi",
         "Rule: status required|in:0,1"),

        ("014", "Chống double-click nút Lưu", "P0",
         "Đang mở modal Thêm với dữ liệu hợp lệ (số TK '19004444444')",
         "1. Bấm Lưu liên tiếp 3 lần thật nhanh\n2. Kiểm tra danh sách và bảng company_accounts",
         "3 lần click < 1 giây",
         "- Nút Lưu bị vô hiệu ngay sau lần bấm đầu cho tới khi API trả về\n"
         "- Chỉ tạo ĐÚNG 1 bản ghi số '19004444444'\n- Không có bản ghi trùng",
         "submitting = true → :disabled trên nút Lưu và Đóng"),

        ("015", "Sửa bản ghi của công ty khác qua API", "P0",
         "Tài khoản id = Y thuộc công ty 2; user đang đăng nhập thuộc công ty 1, có đủ quyền",
         "1. Gọi PUT /v1/finance/account-banks/Y với dữ liệu hợp lệ\n2. Gọi GET /v1/finance/account-banks/Y\n"
         "3. Gọi GET /v1/finance/account-banks/Y/lock",
         "id Y thuộc công ty khác",
         "- Cả 3 request đều trả HTTP 404 với message 'Không tìm thấy tài khoản ngân hàng'\n"
         "- KHÔNG trả 403 (không lộ việc bản ghi có tồn tại)\n- Dữ liệu công ty 2 không bị thay đổi",
         "findForCompany() lọc kèm company_id → không thấy thì 404"),

        ("016", "Nhập ký tự đặc biệt và tiếng Việt có dấu", "P1",
         "Đang mở modal Thêm",
         "1. Chủ tài khoản = 'Công ty TNHH Đông Á & Cộng sự (CN #1)'\n2. Nhập đủ trường còn lại\n3. Bấm Lưu\n4. Quan sát danh sách",
         "Chuỗi có dấu tiếng Việt, &, #, ()",
         "- Lưu thành công\n- Danh sách hiển thị 'CÔNG TY TNHH ĐÔNG Á & CỘNG SỰ (CN #1)' — in hoa đúng cả chữ có dấu\n"
         "- Không bị mã hoá HTML (không thấy &amp;)",
         "mb_strtoupper với bảng mã UTF-8 xử lý đúng chữ tiếng Việt"),

        ("017", "Nhập thẻ script vào ô nhập (XSS)", "P0",
         "Đang mở modal Thêm",
         "1. Chủ tài khoản = '<script>alert(1)</script>'\n2. Nhập đủ trường còn lại\n3. Bấm Lưu\n4. Quan sát danh sách",
         "account_name = '<script>alert(1)</script>'",
         "- KHÔNG xuất hiện hộp thoại alert\n- Chuỗi hiển thị nguyên văn dạng text trên danh sách\n- Không lỗi JS console",
         "Vue nội suy tự escape — kiểm tra cả ô lọc và ô hiển thị"),

        ("018", "Số tài khoản rất dài", "P2",
         "Đang mở modal Thêm",
         "1. Nhập Số tài khoản 100 ký tự số\n2. Nhập đủ trường còn lại\n3. Bấm Lưu",
         "account_number: 100 ký tự",
         "- Nếu vượt giới hạn cột DB: báo lỗi rõ ràng, KHÔNG trả 500\n"
         "- Nếu lưu được: cột Số tài khoản hiển thị không vỡ bảng, cột ghim vẫn hoạt động",
         "Kiểm tra ràng buộc độ dài — Request hiện chưa khai max"),

        ("019", "Mất kết nối mạng khi đang lưu", "P2",
         "Đang mở modal Thêm với dữ liệu hợp lệ",
         "1. Ngắt mạng (DevTools → Offline)\n2. Bấm Lưu\n3. Quan sát",
         "Network: Offline",
         "- Hiện toast đỏ thông báo lỗi\n- Thanh loading tắt (không quay mãi)\n"
         "- Nút Lưu bật lại để thử lại\n- Modal không đóng, dữ liệu đã nhập còn nguyên",
         "finally: submitting = false + $loading.finish()"),

        ("020", "Lỗi 422 không kèm chi tiết trường", "P2",
         "User có quyền nhưng chưa gắn công ty (tái hiện lại tình huống ở TC-ROLE-05)",
         "1. Nhập đủ thông tin hợp lệ\n2. Bấm Lưu\n3. Quan sát cách hiển thị lỗi",
         "Response 422 không có mảng errors",
         "- Hiện TOAST với nội dung 'Tài khoản đăng nhập chưa gắn công ty, không thể thao tác'\n"
         "- Không có ô nào bị đánh dấu đỏ vô cớ\n- Modal vẫn mở",
         "FE: 422 không có errors → hiển thị message dạng toast"),

        ("021", "Hai người sửa cùng bản ghi", "P2",
         "User A và user B cùng công ty, cùng mở modal Sửa tài khoản '19008888888'",
         "1. User A đổi Chủ tài khoản thành 'AAA', bấm Lưu\n2. User B (chưa reload) đổi Chi nhánh, bấm Lưu\n3. Kiểm tra dữ liệu cuối cùng",
         "2 phiên sửa song song",
         "- Cả 2 lần lưu đều thành công (màn không có khoá phiên bản)\n"
         "- Bản ghi cuối mang dữ liệu của lần lưu SAU, ghi đè thay đổi của user A ở các trường B gửi lên\n"
         "- Ghi nhận hành vi để nghiệp vụ quyết định có cần cảnh báo xung đột không",
         "Không có cơ chế optimistic lock — hiện trạng giống ERP"),
    ]),

    # ---------------------------------------------------------------- VII
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Gọi API khi chưa đăng nhập", "P0",
         "Chưa có token hoặc đã đăng xuất",
         "1. Gọi GET /api/v1/finance/account-banks không gắn header Authorization\n2. Quan sát response",
         "Không có Bearer token",
         "- API trả HTTP 401\n- Body không chứa bất kỳ dữ liệu tài khoản nào",
         "Route group middleware auth"),

        ("002", "Truy cập URL màn hình khi chưa đăng nhập", "P0",
         "Trình duyệt đã đăng xuất (xoá token)",
         "1. Gõ URL /finance/account-banks\n2. Quan sát",
         "—",
         "- Chuyển về màn đăng nhập\n- Không hiển thị dữ liệu tài khoản dù chỉ trong tích tắc",
         "Middleware auth phía FE"),

        ("003", "2 user khác công ty thấy dữ liệu khác nhau", "P0",
         "User A (công ty 1) có 8 tài khoản; user B (công ty 2) có 3 tài khoản; cả 2 đều có quyền quản lý",
         "1. Đăng nhập user A, ghi lại tổng số và danh sách số TK\n2. Đăng xuất, đăng nhập user B\n3. Vào cùng màn, so sánh",
         "A: công ty 1; B: công ty 2",
         "- User A thấy đúng 8 bản ghi của công ty 1\n- User B thấy đúng 3 bản ghi của công ty 2\n"
         "- KHÔNG có số tài khoản nào xuất hiện ở cả 2 danh sách\n- Tổng số bản ghi khác nhau",
         "searchByFilter luôn kèm where company_id = công ty user login"),

        ("004", "2 user CÙNG công ty thấy dữ liệu giống nhau", "P0",
         "User A và user D cùng thuộc công ty 1, cùng có quyền; công ty 1 có 8 tài khoản",
         "1. Đăng nhập user A, ghi lại danh sách\n2. Đăng nhập user D, so sánh",
         "A và D: cùng company_id",
         "- Cả 2 thấy CÙNG 8 bản ghi\n- Không lọc theo phòng ban / người tạo (màn không phân quyền theo cấp)\n"
         "- User D sửa được bản ghi do user A tạo",
         "Phạm vi chỉ theo company_id, không theo department/part/creator"),

        ("005", "Bản ghi có company_id NULL không bị lộ", "P0",
         "DB có 1 bản ghi company_accounts với company_id = NULL (do ERP sinh); user đăng nhập có company_id = 1",
         "1. Đăng nhập user công ty 1\n2. Vào màn danh mục, tìm bản ghi đó\n3. Thử lọc bằng số tài khoản của nó",
         "Bản ghi: company_id = NULL",
         "- Bản ghi KHÔNG xuất hiện trong danh sách\n- Lọc theo số TK của nó trả về 0 kết quả\n"
         "- Gọi GET theo id của nó trả 404",
         "Không đưa null vào where() — company null được chặn bằng nhánh riêng"),

        ("006", "Bản ghi do màn ERP tạo hiển thị đúng ở HRM", "P1",
         "Trên ERP, tạo 1 tài khoản cho công ty 1 với số '19003333333'",
         "1. Tạo bản ghi ở màn ERP\n2. Đăng nhập HRM bằng user công ty 1\n3. Vào /finance/account-banks, tìm '19003333333'",
         "Bản ghi tạo từ ERP",
         "- Bản ghi hiển thị trong danh sách HRM với đầy đủ số TK, chủ TK, ngân hàng, chi nhánh, trạng thái\n"
         "- Sửa được ở HRM; sau khi sửa, ERP cũng thấy dữ liệu mới (chung 1 bảng)",
         "2 màn chạy song song trên cùng bảng company_accounts"),

        ("007", "Xử lý khi API trả lỗi 500", "P1",
         "Giả lập API /finance/account-banks trả HTTP 500 (DevTools override response)",
         "1. Giả lập 500\n2. Reload màn",
         "Response: 500",
         "- Hiện toast đỏ 'Lỗi khi tải dữ liệu'\n- Bảng ở trạng thái rỗng, loading đã tắt\n- Panel lọc vẫn thao tác được để thử lại",
         "catch: chỉ bỏ qua toast khi status = 403"),
    ]),

    # ---------------------------------------------------------------- VIII
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đầy đủ: thêm → sửa → khóa → mở khóa → xem", "P0",
         "User có quyền, thuộc công ty 1; chưa tồn tại số TK 'E2E00011122'",
         "1. Thêm tài khoản: số 'E2E00011122', tiền tệ VND, chủ TK 'cong ty e2e', NH VIETCOMBANK, CN HOAN KIEM, trạng thái Hoạt động\n"
         "2. Bấm Sửa: đổi chủ TK thành 'cong ty e2e sua', đổi chi nhánh sang chi nhánh khác cùng ngân hàng → Lưu\n"
         "3. Bấm nút ổ khóa → xác nhận 'Khóa'\n4. Bấm nút mở khóa → xác nhận 'Mở khóa'\n"
         "5. Bấm nút Xem, đối chiếu toàn bộ dữ liệu",
         "Số TK = 'E2E00011122'",
         "- B1: bản ghi mới đứng đầu danh sách, chủ TK hiển thị 'CONG TY E2E' in hoa\n"
         "- B2: toast 'Cập nhật thành công', chủ TK đổi thành 'CONG TY E2E SUA', cột Chi nhánh đổi tên mới\n"
         "- B3: nhãn 'Khóa'\n- B4: nhãn 'Hoạt động'\n"
         "- B5: modal Xem hiển thị đúng 6 giá trị mới nhất, mọi ô chỉ đọc, footer chỉ có nút Đóng",
         "Bao trùm toàn bộ chức năng của màn (không có Xóa)"),

        ("002", "Luồng liên thông với Danh mục ngân hàng", "P0",
         "User có quyền cả 2 màn; ngân hàng 'NH E2E' đang hoạt động và có chi nhánh 'CN E2E'",
         "1. Ở Danh mục ngân hàng: xác nhận 'NH E2E' đang Hoạt động và có chi nhánh 'CN E2E'\n"
         "2. Sang màn tài khoản ngân hàng: thêm tài khoản gắn 'NH E2E' + 'CN E2E' → Lưu\n"
         "3. Quay lại Danh mục ngân hàng: rê chuột vào nút Xóa của 'NH E2E'\n"
         "4. Khóa 'NH E2E'\n5. Về màn tài khoản: bấm Xem tài khoản đó, rồi bấm Sửa",
         "Ngân hàng 'NH E2E'",
         "- B2: lưu thành công, cột Ngân hàng hiển thị 'NH E2E' in hoa\n"
         "- B3: nút Xóa của 'NH E2E' bị mờ với tooltip 'Không thể xóa bản ghi, ngân hàng đang được sử dụng trên hệ thống'\n"
         "- B5 (Xem): ô Ngân hàng vẫn hiển thị 'NH E2E'\n"
         "- B5 (Sửa): ô Ngân hàng bị xóa trắng kèm toast 'Ngân hàng của tài khoản này đã bị khóa, vui lòng chọn ngân hàng khác'",
         "canDelete() của Bank kiểm tra company_accounts; options chỉ trả bank đang hoạt động"),

        ("003", "Luồng lọc — phân trang giữ đúng ngữ cảnh", "P1",
         "Công ty có ≥ 30 tài khoản, trong đó ≥ 15 tài khoản của VIETCOMBANK",
         "1. Lọc Ngân hàng = 'VIETCOMBANK'\n2. Đổi số dòng/trang = 10\n3. Sang trang 2\n"
         "4. Khóa 1 tài khoản trên trang 2\n5. Bấm Đặt lại",
         "bank_name = 'VIETCOMBANK'",
         "- B3: trang 2 chỉ chứa tài khoản VIETCOMBANK, STT bắt đầu từ 11\n"
         "- B4: sau khi khóa vẫn ở trang 2, vẫn giữ bộ lọc, dòng vừa khóa hiển thị nhãn 'Khóa'\n"
         "- B5: bộ lọc về rỗng, danh sách hiển thị toàn bộ tài khoản của công ty từ trang 1",
         "page/per_page tách khỏi filters nên đổi trang không reset bộ lọc"),

        ("004", "Luồng phân quyền: gán quyền giữa chừng", "P1",
         "User B ban đầu KHÔNG có quyền 'Quản lý danh mục tài khoản ngân hàng'",
         "1. Đăng nhập user B, xác nhận menu không hiện\n2. Quản trị gán quyền cho vai trò của B\n"
         "3. User B đăng xuất và đăng nhập lại\n4. Vào màn và thử Thêm 1 tài khoản",
         "User B: chưa quyền → có quyền",
         "- B1: menu ẩn, gõ URL trực tiếp thì API trả 403 và bảng rỗng\n"
         "- B3-B4: menu hiển thị, danh sách load được, thêm tài khoản thành công\n"
         "- Bản ghi tạo ra mang company_id của công ty user B",
         "checkPermission đọc quyền theo phiên đăng nhập → cần đăng nhập lại sau khi gán"),

        ("005", "Luồng dữ liệu cũ do ERP tạo (thiếu tiền tệ)", "P1",
         "Có bản ghi cũ do ERP tạo: currency_id = null, thuộc công ty của user",
         "1. Vào màn, tìm bản ghi đó, quan sát cột Loại tiền tệ\n2. Lọc theo số tài khoản của nó\n"
         "3. Bấm Sửa, bấm Lưu ngay\n4. Chọn tiền tệ 'VND', bấm Lưu\n5. Quan sát lại danh sách",
         "currency_id = null",
         "- B1: cột Loại tiền tệ hiển thị '—'\n- B2: lọc ra đúng bản ghi\n"
         "- B3: chặn lưu, báo 'Bắt buộc phải nhập' ở ô Loại tiền tệ\n"
         "- B4: lưu thành công\n- B5: cột Loại tiền tệ đổi thành 'VND'",
         "HRM buộc currency_id ở cả update — dữ liệu cũ được chuẩn hoá dần khi sửa lại"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT
    a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT
    a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT
    b.alignment = WRAP_TOP_LEFT
    b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, "Testcase _ %s" % FEATURE_NAME)
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT
    lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right")
    lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT
    vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN
    vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0


def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1


def write_tc(tc_id, function, priority, precondition, steps, test_data, expected,
             business_note, module=MODULE_NAME, group=""):
    global current_row, data_row_idx
    values = [
        module, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed",
        "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    newlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(200, max(longest // 4, newlines * 15 + 20)))
    current_row += 1
    data_row_idx += 1


total_tc = 0
p0_tc = 0

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc("TC-ROLE-%s" % suffix, func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

ROMANS = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row("%s. %s" % (roman, title))
    sec_idx = ROMANS.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = "TC_%02d.%03d" % (sec_idx, int(tc_num))
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

dv = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
    showDropDown=False,
)
dv.add("L18:N%d" % (current_row + 100))
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print("Generated: %s" % OUTPUT_FILE)
print("Rows: 1-10 description, 11-15 summary, 17 header, 18-%d data" % (current_row - 1))
print("Total TC: %d | P0: %d (%.0f%%)" % (total_tc, p0_tc, p0_tc * 100.0 / total_tc))
