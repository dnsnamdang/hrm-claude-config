# -*- coding: utf-8 -*-
"""Generate testcase Excel cho man Danh muc goi bao duong (Modules/CustomerCare - pages/customer-care/services)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE = ".plans/gop-db/customer-care-services-catalog/testcase.xlsx"
SHEET_NAME = "DanhMucGoiBaoDuong"
FEATURE_NAME = "Danh mục gói bảo dưỡng"
MODULE_NAME = "DM gói bảo dưỡng"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý danh mục gói bảo dưỡng (phân hệ CSKH → Danh mục - Dịch vụ → Danh mục gói bảo dưỡng, URL /customer-care/services).\n"
     "Màn port từ ERP (`admin/sale/services`), chạy song song với ERP trên CÙNG bảng `services` + 5 bảng con của DB gộp — KHÔNG đổi schema.\n"
     "Phạm vi chức năng: xem danh sách + lọc + sắp xếp, Thêm mới, Sửa, Sao chép, Xóa (hoặc tự chuyển Khóa), In phiếu 'Danh mục kiểm tra bảo dưỡng', Xuất Excel.\n"
     "Form thêm/sửa là TRANG RIÊNG (không phải modal) gồm 5 khối: Thông tin chung · Ma trận danh mục kiểm tra bảo dưỡng định kỳ · "
     "Giá vốn theo công ty · Áp dụng cho hàng hóa · File đính kèm PDF."),

    ("2. Đối tượng được tính / hiển thị",
     "► Toàn bộ bản ghi bảng `services` — KHÔNG lọc theo công ty / phòng ban / người tạo (mọi user đăng nhập đều xem được toàn bộ danh mục).\n"
     "► Bao gồm CẢ status = 1 (Hoạt động) và status = 0 (Khóa) — gói đã Khóa vẫn hiển thị và vẫn Sửa được.\n"
     "► Gói do cổng ERP tạo cũng hiển thị ở HRM (chung 1 bảng) và ngược lại.\n"
     "► Dropdown ĐVT: `units` có status = 1. Dropdown Cấp bảo dưỡng: toàn bộ `levels`. Dropdown Ghi chú: toàn bộ `note_maintenances` (hiển thị theo `key_name`). Dropdown Công ty: toàn bộ `companies`.\n"
     "► Popup chọn hàng hóa: `products` có `deleted_at` IS NULL và `status` ≠ 0. Popup chọn nhóm hàng: `groups` có status = 1."),

    ("3. Đối tượng bị ẩn / không tính",
     "► Bản ghi đã xóa cứng khỏi `services` (màn không dùng SoftDeletes).\n"
     "► Popup hàng hóa: KHÔNG hiển thị hàng đã xóa mềm (`deleted_at` khác null) và hàng `status` = 0.\n"
     "► Popup nhóm hàng: KHÔNG hiển thị nhóm `status` ≠ 1.\n"
     "► Dropdown ĐVT: KHÔNG hiển thị đơn vị tính `status` ≠ 1.\n"
     "► Màn Sửa: hàng hóa đã gắn vào gói vẫn hiển thị đủ dù hàng đó đã bị xóa mềm sau khi gắn (chủ đích — tránh lưu lại làm mất pivot).\n"
     "► Không có bản ghi bị ẩn theo quyền — 3 quyền của màn chỉ chặn thao tác ghi, không chặn xem."),

    ("4. Bộ lọc thời gian áp dụng cho",
     "Không áp dụng — màn không có bộ lọc theo khoảng thời gian.\n"
     "Hai cột ngày 'Ngày tạo' (`created_at`) và 'Ngày sửa' (`updated_at`) chỉ dùng để hiển thị (định dạng d/m/Y) và sắp xếp, không dùng để lọc."),

    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Bảng chính `services` (dùng chung với ERP). 5 bảng con/liên kết:\n"
     "   • `service_maintains` — mỗi dòng = 1 HÀNG của ma trận (nội dung kiểm tra, ĐVT, SL)\n"
     "   • `service_maintain_levels` — mỗi dòng = 1 ghi chú của 1 ô (service_maintain_id × level_id × note_maintenance_id, kèm `order` = thứ tự cột)\n"
     "   • `service_levels` — mỗi dòng = 1 CỘT cấp bảo dưỡng của gói (quota_work, benefit_coefficient, base_price, key_word, order)\n"
     "   • `company_service_coefficients` (pivot) — hệ số giá bán theo từng công ty\n"
     "   • `service_has_products` (pivot) — hàng hóa áp dụng, kèm `group_id` của nhóm tại thời điểm gắn\n"
     "► Khóa ngoài khác: `services.company_id` → `companies.id`; `service_levels.level_id` → `levels.id`; "
     "`service_maintain_levels.note_maintenance_id` → `note_maintenances.id`; `service_maintains.unit_id` → `units.id`.\n"
     "► `services.attachments` là CHUỖI các URL S3 nối bằng ', ' (không có bảng đính kèm riêng).\n"
     "► Ma trận là bảng 2 chiều: HÀNG = nội dung kiểm tra, CỘT = cấp bảo dưỡng, Ô = danh sách ghi chú (chọn nhiều)."),

    ("6. Quy tắc cộng dồn / deduplicate",
     "► Không cộng dồn số liệu. Mỗi dòng danh sách = 1 bản ghi `services`, phân trang server-side (mặc định 10 dòng/trang).\n"
     "► Chống trùng: `name` unique toàn bảng `services`; `code` unique toàn bảng `services` (bỏ qua chính bản ghi đang sửa). `code` luôn được lưu IN HOA.\n"
     "► Ma trận: không cho chọn TRÙNG cấp bảo dưỡng giữa 2 cột (FE chặn kèm cảnh báo).\n"
     "► Hàng hóa: 1 hàng hóa chỉ được thêm 1 lần vào gói; thêm lại sẽ bị bỏ qua và báo 'đã được chọn trước đó'.\n"
     "► Khi lưu: toàn bộ `service_maintains` + `service_maintain_levels` cũ bị XÓA rồi tạo lại từ dữ liệu gửi lên; riêng `service_levels` chỉ xóa cấp không còn dùng và giữ nguyên id của cấp còn dùng (để không phá báo giá dịch vụ đang trỏ tới).\n"
     "► STT hiển thị = (trang − 1) × số dòng/trang + vị trí + 1."),

    ("7. Phân quyền cấp",
     "Màn dùng 3 quyền (không phân quyền theo cấp tổ chức, không lọc dữ liệu theo công ty):\n"
     "• `Thêm danh mục gói bảo dưỡng` — POST /v1/customer-care/services. FE dùng để ẩn/hiện nút 'Thêm mới' và nút 'Sao chép'.\n"
     "• `Sửa danh mục gói bảo dưỡng` — POST /v1/customer-care/services/{id}. FE dùng để ẩn/hiện nút 'Sửa'.\n"
     "• `Xóa danh mục gói bảo dưỡng` — DELETE /v1/customer-care/services/{id}. FE dùng để ẩn/hiện nút 'Xóa'.\n"
     "→ KHÔNG gate quyền (chỉ cần đăng nhập): xem danh sách (GET /), xem chi tiết (GET /{id}), dữ liệu dropdown (GET /options-data), "
     "xuất Excel (GET /export), dữ liệu in (GET /{id}/print-data), tìm hàng hóa / nhóm / model / danh mục lọc.\n"
     "→ Menu 'Danh mục gói bảo dưỡng' KHÔNG gate quyền — hiện với mọi user đăng nhập (giống ERP)."),

    ("8. Cách tính các ô thống kê",
     "Màn không có thẻ thống kê. Các ô số/tính toán:\n"
     "► Tooltip 'Giá' khi rê chuột vào tên gói ở danh sách = FLOOR(đơn giá công của công ty quản lý × Định mức công của cấp × Hệ số giá bán gói) — tính cho từng cấp.\n"
     "► Ô 'Giá vốn' trên form = đơn giá công công ty quản lý × Định mức công × Hệ số công nghệ (bỏ trống Hệ số công nghệ = 1).\n"
     "► Ô 'Giá công thức' trên form = Giá vốn × Hệ số giá bán gói bảo dưỡng (bỏ trống = 1).\n"
     "► Ô 'Giá bán cơ sở' = người dùng nhập tay; mặc định được gán = ROUND(Giá công thức) và bị ghi đè lại mỗi khi đổi Định mức công / Hệ số giá bán gói / Hệ số công nghệ.\n"
     "► Dòng 'Giá bán theo công ty' = Giá bán cơ sở × Hệ số giá bán của công ty đó (công ty quản lý luôn hệ số 1).\n"
     "► Khối 'Giá vốn theo công ty': ô giá vốn từng cột = Định mức công × Đơn giá công của công ty ở dòng đó.\n"
     "► Ô 'STT' = (trang − 1) × số dòng/trang + vị trí + 1."),

    ("9. Ghi chú đọc bảng",
     "► Danh sách 9 cột; cột STT và Tên gói bảo dưỡng được ghim khi cuộn ngang. 5 cột cho phép sắp xếp: Tên, Mã, Trạng thái, Ngày tạo, Ngày sửa.\n"
     "► 4 nút hành động (Sao chép · Sửa · In · Xóa) nằm ngay dưới tên gói ở cột 'Tên gói bảo dưỡng'; nút In luôn hiện, mở tab mới.\n"
     "► Nút Xóa LUÔN hiện (không bị mờ): hệ thống tự quyết xóa hẳn hay chuyển sang trạng thái Khóa tùy gói đã được sử dụng hay chưa.\n"
     "► Trạng thái hiển thị dạng pill: xanh 'Hoạt động' (status = 1), khóa 'Khóa' (status = 0).\n"
     "► Số tiền hiển thị phân cách hàng nghìn bằng dấu PHẨY (1,400,000) — khớp định dạng ERP.\n"
     "► Giá trị trống hiển thị '—'.\n"
     "► Bộ lọc được ghi nhớ 10 phút khi rời trang rồi quay lại (lưu ở trình duyệt, key `customer_care_services`).\n"
     "► Ô 'Tên gói bảo dưỡng' ở bộ lọc nâng cao dùng CHUNG một trường với ô tìm nhanh phía trên."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("01", "Xem danh sách KHÔNG cần quyền nào", "P0",
     "User X đã đăng nhập, KHÔNG được gán bất kỳ quyền nào trong 3 quyền của màn; danh mục có 15 gói",
     "1. Đăng nhập user X\n2. Vào CSKH → Danh mục - Dịch vụ → 'Danh mục gói bảo dưỡng'\n3. Quan sát danh sách và các nút",
     "User X: không có quyền nào của màn",
     "- Menu hiển thị bình thường (menu không gate quyền)\n- Danh sách load đủ 15 gói, lọc / sắp xếp / phân trang hoạt động\n"
     "- Nút 'Xuất excel' và nút 'In' trên từng dòng VẪN hiển thị và dùng được\n"
     "- KHÔNG hiển thị: nút 'Thêm mới', nút 'Sao chép', nút 'Sửa', nút 'Xóa'",
     "Route index/show/export/print-data/options-data KHÔNG gắn checkPermission"),

    ("02", "Có quyền 'Thêm danh mục gói bảo dưỡng'", "P0",
     "User A chỉ có quyền `Thêm danh mục gói bảo dưỡng`",
     "1. Đăng nhập user A\n2. Vào màn danh sách, quan sát nút\n3. Bấm 'Thêm mới', nhập đủ thông tin hợp lệ, bấm Lưu\n4. Bấm nút 'Sao chép' trên 1 dòng",
     "User A: chỉ quyền Thêm",
     "- Hiển thị nút 'Thêm mới' và nút 'Sao chép'; KHÔNG hiển thị nút 'Sửa' và 'Xóa'\n"
     "- Bước 3: tạo thành công, toast 'Tạo gói bảo dưỡng thành công'\n- Bước 4: mở được màn tạo bản sao",
     "Permission: Thêm danh mục gói bảo dưỡng (gắn cho POST /); nút Sao chép dùng chung quyền này"),

    ("03", "Có quyền 'Sửa danh mục gói bảo dưỡng'", "P0",
     "User B chỉ có quyền `Sửa danh mục gói bảo dưỡng`; có gói 'GBD-TEST-01'",
     "1. Đăng nhập user B\n2. Quan sát nút trên dòng\n3. Bấm 'Sửa' gói GBD-TEST-01, đổi Ghi chú, bấm Lưu",
     "User B: chỉ quyền Sửa",
     "- Hiển thị nút 'Sửa'; KHÔNG hiển thị 'Thêm mới', 'Sao chép', 'Xóa'\n"
     "- Lưu thành công, toast 'Cập nhật gói bảo dưỡng thành công'\n"
     "- Trong màn sửa: nút 'Sao chép' cuối trang KHÔNG hiển thị (thiếu quyền Thêm)",
     "Permission: Sửa danh mục gói bảo dưỡng (gắn cho POST /{id})"),

    ("04", "Có quyền 'Xóa danh mục gói bảo dưỡng'", "P0",
     "User C chỉ có quyền `Xóa danh mục gói bảo dưỡng`; có gói chưa gắn hàng hóa và chưa dùng ở báo giá",
     "1. Đăng nhập user C\n2. Quan sát nút trên dòng\n3. Bấm 'Xóa' gói đó và xác nhận",
     "User C: chỉ quyền Xóa",
     "- Hiển thị nút 'Xóa'; KHÔNG hiển thị 'Thêm mới', 'Sao chép', 'Sửa'\n"
     "- Xóa thành công, toast 'Xóa gói bảo dưỡng thành công', dòng biến mất",
     "Permission: Xóa danh mục gói bảo dưỡng (gắn cho DELETE /{id})"),

    ("05", "Gọi API ghi khi thiếu quyền tương ứng", "P0",
     "User X không có quyền nào; có gói id = Y",
     "1. Lấy token user X\n2. Gọi POST /v1/customer-care/services (tạo mới)\n"
     "3. Gọi POST /v1/customer-care/services/Y (sửa)\n4. Gọi DELETE /v1/customer-care/services/Y",
     "User X: token hợp lệ, thiếu cả 3 quyền",
     "- Cả 3 request trả HTTP 403\n- Không tạo/sửa/xóa bản ghi nào\n"
     "- Trong khi đó GET /v1/customer-care/services và GET /v1/customer-care/services/Y vẫn trả 200",
     "3 route ghi gắn 3 quyền riêng biệt; route đọc không gate"),

    ("06", "Có quyền Thêm nhưng gọi API Sửa", "P0",
     "User A chỉ có quyền `Thêm danh mục gói bảo dưỡng`; có gói id = Y",
     "1. Lấy token user A\n2. Gọi POST /v1/customer-care/services/Y với dữ liệu hợp lệ\n3. Kiểm tra bản ghi Y trong DB",
     "User A: chỉ quyền Thêm",
     "- API trả HTTP 403\n- Bản ghi Y không thay đổi (kể cả các bảng con)\n- 3 quyền độc lập, không quyền nào bao quyền nào",
     "Mỗi route ghi gắn đúng 1 quyền riêng"),

    ("07", "Truy cập thẳng URL màn tạo/sửa khi thiếu quyền", "P1",
     "User X không có quyền nào; có gói id = Y",
     "1. Gõ URL /customer-care/services/create\n2. Nhập dữ liệu hợp lệ và bấm Lưu\n"
     "3. Gõ URL /customer-care/services/Y/edit, sửa gì đó rồi bấm Lưu",
     "User X: không quyền",
     "- Trang form vẫn mở được (FE không chặn route) nhưng khi Lưu: API trả 403\n"
     "- KHÔNG hiện toast lỗi rác (403 được nuốt im lặng), KHÔNG tạo/sửa dữ liệu\n"
     "- Ghi nhận: nếu nghiệp vụ muốn chặn ngay từ URL thì đây là điểm cần bổ sung",
     "FE chỉ ẩn nút theo quyền; chặn thật nằm ở BE"),
]

SECTIONS = [
    # ---------------------------------------------------------------- I
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Truy cập màn hình từ menu CSKH", "P0",
         "User đã đăng nhập; danh mục có ≥ 12 gói bảo dưỡng",
         "1. Đăng nhập\n2. Vào phân hệ Chăm sóc khách hàng → nhóm 'Danh mục - Dịch vụ'\n"
         "3. Bấm 'Danh mục gói bảo dưỡng'\n4. Quan sát layout",
         "User bất kỳ",
         "- URL = /customer-care/services\n- Tiêu đề trang/tab = 'Danh mục gói bảo dưỡng'\n"
         "- Hiển thị panel 'Bộ lọc gói bảo dưỡng' (đang thu gọn) và bảng 'Danh mục gói bảo dưỡng'\n"
         "- Cuối bảng có nút 'Thêm mới' (nếu có quyền) và 'Xuất excel'\n- Không lỗi console",
         "Menu customer-care.js không gate quyền"),

        ("002", "Kiểm tra đủ 9 cột của bảng danh sách", "P0",
         "Đang ở màn danh mục gói bảo dưỡng, có ≥ 1 bản ghi",
         "1. Quan sát dòng tiêu đề bảng\n2. Đối chiếu tên, thứ tự và icon sắp xếp từng cột",
         "—",
         "Đúng thứ tự 9 cột:\n1. STT\n2. Tên gói bảo dưỡng (sắp xếp được)\n3. Mã (sắp xếp được)\n"
         "4. Công ty quản lý gói bảo dưỡng\n5. Trạng thái (sắp xếp được)\n6. Người tạo\n"
         "7. Ngày tạo (sắp xếp được)\n8. Người sửa\n9. Ngày sửa (sắp xếp được)\n"
         "- Cột 'Công ty quản lý', 'Người tạo', 'Người sửa' KHÔNG có icon sắp xếp",
         "allowedSortFields BE: name, code, status, created_at, updated_at"),

        ("003", "Cột STT và Tên gói được ghim khi cuộn ngang", "P1",
         "Thu hẹp cửa sổ để bảng phải cuộn ngang",
         "1. Thu nhỏ cửa sổ đến khi có thanh cuộn ngang\n2. Cuộn sang phải\n3. Quan sát 2 cột đầu",
         "Độ rộng cửa sổ ~1024px",
         "- Cột STT và 'Tên gói bảo dưỡng' luôn dính bên trái\n- Các cột còn lại cuộn bình thường\n- Nội dung không bị đè lên nhau",
         "tableColumns: sticky = true cho index và name"),

        ("004", "Hiển thị khi không có dữ liệu phù hợp", "P1",
         "Danh mục có dữ liệu nhưng lọc bằng từ khoá chắc chắn không khớp",
         "1. Nhập ô tìm nhanh 'zzzkhongtontai'\n2. Bấm Tìm kiếm\n3. Quan sát bảng",
         "name = 'zzzkhongtontai'",
         "- Hiển thị 'Không có dữ liệu phù hợp bộ lọc.'\n- Tổng số bản ghi = 0\n- Nút 'Thêm mới' và 'Xuất excel' vẫn dùng được",
         "emptyText của V2BaseDataTable"),

        ("005", "Mở rộng / thu gọn bộ lọc nâng cao", "P1",
         "Đang ở màn danh sách, panel lọc thu gọn (mặc định)",
         "1. Bấm mở rộng bộ lọc\n2. Quan sát các trường\n3. Thu gọn lại rồi mở lại",
         "—",
         "- Khi mở: hiện 4 trường 'Tên gói bảo dưỡng', 'Mã', 'Trạng thái', 'Người tạo'\n"
         "- Ô 'Tên gói bảo dưỡng' luôn đồng bộ giá trị với ô tìm nhanh phía trên\n"
         "- Giá trị đã nhập không mất sau khi thu gọn rồi mở lại",
         "Ô tìm nhanh và ô 'Tên gói bảo dưỡng' cùng bind filters.name"),

        ("006", "Trạng thái loading khi tải dữ liệu", "P2",
         "Bật throttle Slow 3G trong DevTools",
         "1. Bật Slow 3G\n2. Reload màn\n3. Quan sát vùng bảng",
         "Network: Slow 3G",
         "- Bảng ở trạng thái đang tải trong lúc chờ API\n- Sau khi API trả về: loading tắt, dữ liệu hiện\n- Không hiện đồng thời loading và dòng 'không có dữ liệu'",
         "loading = true trước gọi API, finally = false"),
    ]),

    # ---------------------------------------------------------------- II
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm nhanh theo tên gói bảo dưỡng", "P0",
         "Danh mục có 'Bảo dưỡng máy xúc cấp 1', 'Bảo dưỡng xe nâng', 'Sửa chữa động cơ'",
         "1. Nhập ô tìm nhanh: 'máy xúc'\n2. Bấm Tìm kiếm",
         "name = 'máy xúc'",
         "- Chỉ hiển thị 'Bảo dưỡng máy xúc cấp 1'\n- Tổng số = 1\n- Về trang 1",
         "services.name LIKE %name%"),

        ("002", "Gõ ô tìm nhanh KHÔNG tự gọi API", "P0",
         "Đang ở màn danh sách, mở tab Network",
         "1. Gõ từng ký tự 'máy' vào ô tìm nhanh, KHÔNG bấm Tìm kiếm\n2. Quan sát Network\n3. Bấm Tìm kiếm",
         "3 ký tự",
         "- Bước 1: KHÔNG phát sinh request nào\n- Bước 3: phát sinh đúng 1 request với name = 'máy'",
         "name nằm trong ignoredFields của deep watcher — chờ bấm Tìm kiếm"),

        ("003", "Lọc theo Mã gói bảo dưỡng", "P0",
         "Danh mục có gói mã 'GBD-001' và 'GBD-002', 'SC-001'",
         "1. Mở bộ lọc nâng cao\n2. Nhập Mã = 'GBD'\n3. Bấm Tìm kiếm",
         "code = 'GBD'",
         "- Trả về 2 gói GBD-001 và GBD-002\n- Không trả về SC-001",
         "services.code LIKE %code%"),

        ("004", "Lọc theo Trạng thái = Hoạt động", "P0",
         "Danh mục có 8 gói status = 1 và 3 gói status = 0",
         "1. Chọn Trạng thái = 'Hoạt động'\n2. Quan sát danh sách",
         "status = 1",
         "- Trả về đúng 8 bản ghi\n- Mọi dòng có pill xanh 'Hoạt động'\n- Danh sách tự tải lại ngay khi chọn (không cần bấm Tìm kiếm)",
         "Bộ lọc khác `name` kích hoạt deep watcher → tự gọi API"),

        ("005", "Lọc theo Trạng thái = Khóa (giá trị 0)", "P0",
         "Danh mục có 8 gói status = 1 và 3 gói status = 0",
         "1. Chọn Trạng thái = 'Khóa'\n2. Quan sát danh sách và tham số request",
         "status = '0'",
         "- Trả về ĐÚNG 3 bản ghi khóa\n- Request gửi lên có status='0' (không bị rớt mất giá trị)\n- Mọi dòng hiện pill 'Khóa'",
         "BE dùng has() + kiểm chuỗi rỗng nên nhận giá trị 0; FE dùng id chuỗi '0' để không bị falsy"),

        ("006", "Lọc theo Người tạo", "P0",
         "User 'NV001 - Nguyễn Văn A' tạo 3 gói; user khác tạo 5 gói",
         "1. Mở bộ lọc nâng cao\n2. Chọn Người tạo = 'NV001 - Nguyễn Văn A'\n3. Quan sát",
         "created_by = id của NV001",
         "- Trả về đúng 3 gói\n- Cột 'Người tạo' của mọi dòng đều là 'NV001 - Nguyễn Văn A'",
         "services.created_by = giá trị chọn"),

        ("007", "Kết hợp nhiều bộ lọc (AND)", "P0",
         "Có gói 'Bảo dưỡng máy xúc' mã GBD-001 status 1; gói 'Bảo dưỡng máy xúc mini' mã GBD-002 status 0",
         "1. Nhập ô tìm nhanh 'máy xúc'\n2. Chọn Trạng thái = 'Hoạt động'\n3. Bấm Tìm kiếm",
         "name = 'máy xúc', status = 1",
         "- Chỉ trả về 'Bảo dưỡng máy xúc' (GBD-001)\n- Không trả về gói đang Khóa",
         "Các điều kiện lọc nối bằng AND"),

        ("008", "Nút Đặt lại bộ lọc", "P0",
         "Đã nhập tìm nhanh + mã + trạng thái + người tạo, đang sắp xếp theo Mã, ở trang 3",
         "1. Bấm nút Đặt lại\n2. Quan sát các ô lọc, thứ tự và trang",
         "—",
         "- Mọi ô lọc về rỗng\n- Sắp xếp về mặc định (mới nhất trước)\n- Danh sách tải lại đầy đủ từ trang 1",
         "handleReset(): filters = initialStateForm, watcher lo gọi API"),

        ("009", "Ghi nhớ bộ lọc khi rời trang rồi quay lại", "P1",
         "Đang lọc Mã = 'GBD', Trạng thái = 'Hoạt động'",
         "1. Chuyển sang màn khác\n2. Quay lại /customer-care/services trong vòng 10 phút\n3. Quan sát ô lọc và kết quả",
         "Rời trang < 10 phút",
         "- Ô lọc khôi phục đúng Mã = 'GBD', Trạng thái = 'Hoạt động'\n- Trạng thái mở/thu gọn của panel cũng khôi phục\n"
         "- Chỉ phát sinh 1 request khi vào lại màn (không gọi thừa)",
         "filterStateMixin key 'customer_care_services'; watcher bỏ qua khi giá trị không đổi"),

        ("010", "Bộ lọc ghi nhớ hết hạn sau 10 phút", "P2",
         "Đã lọc rồi rời trang > 10 phút",
         "1. Lọc Mã = 'GBD', rời trang\n2. Chờ quá 10 phút\n3. Quay lại màn",
         "Rời trang > 10 phút",
         "- Bộ lọc KHÔNG khôi phục, các ô về rỗng\n- Danh sách hiển thị toàn bộ gói",
         "expirationTime = 10 phút"),

        ("011", "Chống gọi API trùng lặp", "P1",
         "Đang ở màn danh sách, mở tab Network",
         "1. Bấm Tìm kiếm 3 lần liên tiếp thật nhanh, không đổi tham số\n2. Đếm request GET",
         "3 lần bấm < 1 giây",
         "- Chỉ phát sinh 1 request\n- Danh sách không nhấp nháy nhiều lần",
         "DedupeLoadMixin chặn request trùng tham số"),

        ("012", "Lọc bằng ký tự đặc biệt của LIKE", "P1",
         "Danh mục có ≥ 10 gói, không gói nào chứa ký tự '%' trong tên",
         "1. Nhập ô tìm nhanh '%'\n2. Bấm Tìm kiếm\n3. Quan sát kết quả",
         "name = '%'",
         "- KHÔNG trả về toàn bộ danh mục (ký tự đại diện phải được vô hiệu hóa)\n- Kỳ vọng 0 bản ghi\n- Không lỗi 500",
         "BE dùng escapeLikeKeyword() để chặn wildcard injection"),
    ]),

    # ---------------------------------------------------------------- III
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "Đang ở màn /customer-care/services",
         "1. Quan sát vùng phía trên bảng dữ liệu",
         "—",
         "- Màn KHÔNG có thẻ thống kê nào\n- Chỉ có panel lọc và bảng dữ liệu\n"
         "- Con số duy nhất ngoài bảng là tổng bản ghi ở vùng phân trang và tooltip giá theo cấp",
         "Section giữ lại theo chuẩn tài liệu — feature không có stats"),
    ]),

    # ---------------------------------------------------------------- IV
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Thứ tự mặc định — gói mới nhất lên đầu", "P0",
         "Danh mục có ≥ 5 gói với thời điểm tạo khác nhau",
         "1. Vào màn (không đụng cột sắp xếp)\n2. Đối chiếu thứ tự với created_at trong DB",
         "—",
         "- Gói tạo gần nhất đứng đầu\n- Thứ tự giảm dần theo ngày tạo",
         "ERP mặc định orderBy created_at desc"),

        ("002", "Sắp xếp theo cột Tên gói bảo dưỡng", "P0",
         "Danh mục có ≥ 5 gói tên khác nhau",
         "1. Click tiêu đề cột 'Tên gói bảo dưỡng'\n2. Quan sát thứ tự\n3. Click lần 2",
         "sort_by = name",
         "- Lần 1: sắp xếp tăng dần theo tên (A→Z)\n- Lần 2: giảm dần (Z→A)\n- Mỗi lần đều về trang 1\n- Bộ lọc đang áp dụng KHÔNG bị mất",
         "sort_by/sort_desc nằm trong filters → watcher reset trang và gọi API"),

        ("003", "Sắp xếp theo Mã / Trạng thái / Ngày tạo / Ngày sửa", "P0",
         "Danh mục có ≥ 10 gói đủ 2 trạng thái, ngày tạo và ngày sửa khác nhau",
         "1. Lần lượt click tiêu đề 4 cột: Mã, Trạng thái, Ngày tạo, Ngày sửa\n2. Mỗi lần đối chiếu thứ tự dữ liệu",
         "sort_by = code / status / created_at / updated_at",
         "- Cả 4 cột đều đổi thứ tự đúng theo giá trị cột đó\n- Icon sắp xếp hiển thị đúng chiều\n- Không phát sinh lỗi API",
         "allowedSortFields BE: name, code, status, created_at, updated_at"),

        ("004", "Cột không cho sắp xếp", "P2",
         "Đang ở màn danh sách",
         "1. Click tiêu đề cột 'Công ty quản lý gói bảo dưỡng', 'Người tạo', 'Người sửa'",
         "—",
         "- Không cột nào đổi thứ tự dữ liệu\n- Không có icon sắp xếp\n- Không phát sinh request thừa",
         "3 cột này không khai sortable; BE cũng chặn bằng allowedSortFields"),

        ("005", "Tooltip giá theo cấp khi rê chuột vào tên gói", "P0",
         "Gói 'BD máy xúc': công ty quản lý có đơn giá công = 100.000; cấp 1 có Định mức công = 2; cấp 2 = 3; Hệ số giá bán gói = 1,5",
         "1. Rê chuột vào tên gói trên danh sách\n2. Đọc nội dung tooltip",
         "work_price = 100000; quota 2 và 3; hệ số 1.5",
         "- Hiện tooltip 'Giá:' liệt kê từng cấp\n- Cấp 1 = 300,000 (floor(100000 × 2 × 1,5))\n"
         "- Cấp 2 = 450,000\n- Số có phân cách hàng nghìn bằng dấu phẩy",
         "priceByLevel = floor(work_price × quota_work × coefficient_cost_price_service)"),

        ("006", "Tooltip giá KHÔNG nhân Hệ số công nghệ", "P1",
         "Gói có Hệ số công nghệ (benefit_coefficient) của cấp 1 = 2, đơn giá công 100.000, định mức 2, hệ số giá bán 1",
         "1. Mở form Sửa gói, ghi lại ô 'Giá công thức' của cấp 1\n2. Quay lại danh sách, rê chuột vào tên gói\n3. So sánh 2 con số",
         "benefit_coefficient = 2",
         "- Form hiển thị Giá công thức = 400.000 (có nhân hệ số công nghệ)\n"
         "- Tooltip danh sách hiển thị 200,000 (KHÔNG nhân hệ số công nghệ)\n"
         "- Ghi nhận chênh lệch này và xác nhận lại với nghiệp vụ xem con số nào đúng",
         "priceByLevel không dùng benefit_coefficient, khác công thức primeCost() ở form"),

        ("007", "Gói không có cấp bảo dưỡng nào", "P2",
         "Gói 'GBD-RONG' chưa khai cấp bảo dưỡng nào (service_levels rỗng)",
         "1. Tìm gói GBD-RONG\n2. Rê chuột vào tên gói",
         "service_levels rỗng",
         "- Không hiện tooltip giá (hoặc hiện tooltip rỗng), không lỗi JS\n- Dòng vẫn hiển thị bình thường",
         "hasPriceByLevel() kiểm tra mảng rỗng"),

        ("008", "Phân trang — chuyển trang", "P0",
         "Danh mục có 25 gói, 10 dòng/trang",
         "1. Ở trang 1, ghi lại tên gói dòng đầu\n2. Sang trang 2\n3. Quan sát STT và dữ liệu",
         "total = 25, pageSize = 10",
         "- STT dòng đầu trang 2 = 11\n- Dữ liệu khác hoàn toàn trang 1\n- Bộ lọc hiện tại được giữ nguyên",
         "Phân trang không đụng filters nên không reset về trang 1"),

        ("009", "Đổi số dòng trên mỗi trang", "P0",
         "Danh mục có ≥ 60 gói; đang ở trang 3, 10 dòng/trang",
         "1. Đổi số dòng/trang sang 50\n2. Quan sát danh sách và vùng phân trang",
         "pageSize: 10 → 50",
         "- Hiển thị tối đa 50 dòng, tự về trang 1, STT bắt đầu từ 1\n"
         "- KHÔNG phát sinh vòng lặp gọi API (chỉ 1 request)\n- Tổng số bản ghi không đổi",
         "FE ép Number(meta.per_page) vì API trả chuỗi"),

        ("010", "Hiển thị pill trạng thái", "P0",
         "Gói A status = 1, gói B status = 0",
         "1. Quan sát cột Trạng thái của A và B",
         "A: 1; B: 0",
         "- A: pill xanh, icon dấu tích, chữ 'Hoạt động'\n- B: pill khóa, icon ổ khóa, chữ 'Khóa'",
         "services: 1 = Hoạt động, 0 = Khóa"),

        ("011", "Hiển thị người tạo / người sửa dạng 'MÃ - Họ tên'", "P0",
         "Gói được tạo bởi nhân viên mã NV001 tên 'Nguyễn Văn A', sửa bởi NV002 'Trần Thị B'",
         "1. Quan sát cột 'Người tạo' và 'Người sửa' của dòng đó",
         "created_by = NV001, updated_by = NV002",
         "- Cột Người tạo hiển thị 'NV001 - Nguyễn Văn A'\n- Cột Người sửa hiển thị 'NV002 - Trần Thị B'\n"
         "- Gói chưa có người sửa hiển thị '—'",
         "ServiceListResource ghép code + fullname từ employee_infos"),

        ("012", "Hiển thị 4 nút hành động dưới tên gói", "P0",
         "User có đủ 3 quyền của màn",
         "1. Quan sát vùng dưới tên gói ở 1 dòng bất kỳ\n2. Rê chuột từng nút",
         "Có đủ quyền",
         "- Đúng 4 nút theo thứ tự: Sao chép · Sửa · In · Xóa\n- Tooltip lần lượt: 'Sao chép', 'Sửa', 'In', 'Xóa'\n"
         "- Nút Xóa màu đỏ và LUÔN bật (không bị mờ dù gói đang được sử dụng)",
         "is_can_delete luôn = true — BE tự quyết xóa hẳn hay chuyển Khóa"),

        ("013", "Định dạng ngày tạo / ngày sửa", "P1",
         "Gói được tạo ngày 05/08/2026",
         "1. Quan sát cột 'Ngày tạo' và 'Ngày sửa'",
         "created_at = 2026-08-05",
         "- Hiển thị '05/08/2026' (dạng ngày/tháng/năm, không kèm giờ)\n- Ô trống hiển thị '—'",
         "Helper::formatDateTime(..., 'd/m/Y')"),
    ]),

    # ---------------------------------------------------------------- V
    ("V", "CHỨC NĂNG CHÍNH (CRUD / ACTION)", [
        # --- Mở form
        ("001", "Mở màn Thêm gói bảo dưỡng", "P0",
         "User có quyền Thêm; đang ở màn danh sách",
         "1. Bấm nút 'Thêm mới'\n2. Quan sát toàn bộ trang",
         "—",
         "- Chuyển sang URL /customer-care/services/create\n- Tiêu đề 'Thêm gói bảo dưỡng'\n"
         "- Hiển thị đủ 5 khối: Thông tin chung, Danh mục kiểm tra bảo dưỡng định kỳ, Giá vốn theo công ty, Áp dụng cho hàng hóa, File đính kèm (PDF)\n"
         "- KHÔNG hiển thị ô 'Trạng thái' (chỉ có ở màn sửa)\n- Cuối trang có 3 nút: Lưu, (Sao chép chỉ ở màn sửa), Hủy",
         "isEdit = false → ẩn ô status"),

        ("002", "Công ty quản lý mặc định khi thêm mới", "P0",
         "User đăng nhập thuộc công ty 'Công ty A'",
         "1. Bấm 'Thêm mới'\n2. Quan sát ô 'Công ty quản lý gói bảo dưỡng'",
         "User: company_id = Công ty A",
         "- Ô công ty đã chọn sẵn 'Công ty A'\n- Ở khối 'Giá vốn theo công ty', dòng Công ty A được in đậm và ô Hệ số giá bán bị vô hiệu (luôn = 1)",
         "applyDefaultCompany(); công ty quản lý luôn hệ số 1"),

        ("003", "Khối Giá vốn theo công ty nạp đủ danh sách công ty", "P0",
         "Hệ thống có 5 công ty với đơn giá công khác nhau",
         "1. Bấm 'Thêm mới'\n2. Cuộn tới khối 'Giá vốn theo công ty'\n3. Đếm số dòng và đối chiếu đơn giá công",
         "5 công ty",
         "- Có đủ 5 dòng công ty\n- Cột 'Đơn giá công' hiển thị đúng giá trị của từng công ty, phân cách hàng nghìn\n"
         "- Cột 'Hệ số giá bán' mặc định = 1 cho mọi công ty\n- Dòng công ty quản lý in đậm, ô hệ số bị vô hiệu",
         "optionsData trả companies kèm work_price; mặc định coefficient = 1"),

        # --- Ma trận
        ("004", "Thêm cột cấp bảo dưỡng", "P0",
         "Đang ở màn Thêm, ma trận chưa có cột nào",
         "1. Bấm nút '+' ở cuối dòng tiêu đề ma trận\n2. Chọn Cấp bảo dưỡng = 'Cấp 1'\n3. Bấm '+' lần nữa, chọn 'Cấp 2'",
         "2 cột: Cấp 1, Cấp 2",
         "- Mỗi lần bấm thêm 1 cột mới với ô chọn cấp\n- Ma trận hiện thêm cột tương ứng ở mọi hàng đã có\n"
         "- Khối 'Giá vốn theo công ty' cũng hiện thêm cột giá vốn tương ứng, tiêu đề là tên cấp",
         "levelCols tách khỏi maintains → thêm cột được cả khi chưa có hàng"),

        ("005", "Chặn chọn TRÙNG cấp bảo dưỡng giữa 2 cột", "P0",
         "Ma trận đã có cột 1 chọn 'Cấp 1'",
         "1. Thêm cột 2\n2. Chọn cũng 'Cấp 1' cho cột 2\n3. Quan sát",
         "2 cột cùng chọn Cấp 1",
         "- Hiện cảnh báo 'Cấp bảo dưỡng này đã được chọn ở cột khác'\n"
         "- Ô chọn của cột 2 tự trả về rỗng, không giữ giá trị trùng\n- Cột 1 không bị ảnh hưởng",
         "onLevelColChange() chặn trùng — BE lưu theo `order` nên trùng cấp sẽ hỏng dữ liệu"),

        ("006", "Xóa cột cấp bảo dưỡng có xác nhận", "P0",
         "Ma trận có 2 cột (Cấp 1, Cấp 2), mỗi ô đã chọn ghi chú, đã nhập định mức và giá",
         "1. Bấm dấu x đỏ trên tiêu đề cột 'Cấp 2'\n2. Đọc hộp xác nhận\n3. Bấm 'Xác nhận'",
         "Xóa cột Cấp 2",
         "- Hộp xác nhận: 'Xóa cột \"Cấp 2\" sẽ xóa toàn bộ ghi chú và giá của cột này. Bạn có chắc chắn?'\n"
         "- Sau xác nhận: cột Cấp 2 biến mất ở CẢ ma trận và khối Giá vốn theo công ty\n"
         "- Dữ liệu cột Cấp 1 giữ nguyên (ghi chú, định mức, giá không bị lệch cột)",
         "removeLevelCol() xóa cả noteIdsByCol tương ứng của từng hàng"),

        ("007", "Hủy xác nhận xóa cột", "P1",
         "Ma trận có 2 cột",
         "1. Bấm dấu x trên cột 2\n2. Bấm 'Hủy'",
         "—",
         "- Hộp xác nhận đóng\n- Cột 2 vẫn còn nguyên cùng toàn bộ dữ liệu",
         "pendingAction chỉ chạy khi bấm nút xác nhận"),

        ("008", "Thêm hàng nội dung kiểm tra bảo dưỡng", "P0",
         "Ma trận đã có 2 cột cấp",
         "1. Bấm 'Thêm danh mục kiểm tra bảo dưỡng'\n2. Nhập Nội dung = 'Kiểm tra dầu máy', ĐVT = 'Lần', SL = 1\n"
         "3. Ở ô cột Cấp 1 chọn 2 ghi chú, ô cột Cấp 2 chọn 1 ghi chú",
         "1 hàng, 2 cột",
         "- Hàng mới xuất hiện với STT = 1 và đủ 2 ô ghi chú tương ứng 2 cột\n"
         "- Ô ghi chú cho phép chọn NHIỀU giá trị, hiển thị theo ký hiệu (key_name)\n"
         "- Sau khi có ≥ 1 hàng, chân bảng hiện các dòng: Định mức công, Hệ số công nghệ, Giá vốn, Giá công thức, Giá bán cơ sở, Gợi ý hàng hoá, Giá bán theo công ty",
         "Chân bảng chỉ render khi maintains.length > 0"),

        ("009", "Xóa hàng nội dung kiểm tra", "P0",
         "Ma trận có 3 hàng",
         "1. Bấm icon thùng rác ở cuối hàng thứ 2\n2. Quan sát",
         "Xóa hàng 2",
         "- Hàng 2 biến mất ngay (không hỏi xác nhận)\n- STT các hàng còn lại đánh lại liên tục 1, 2\n- Dữ liệu 2 hàng còn lại không bị lệch",
         "removeMaintain() xóa trực tiếp"),

        ("010", "Bảng rỗng khi chưa có hàng nào", "P2",
         "Đang ở màn Thêm, chưa thêm hàng nào",
         "1. Quan sát thân bảng ma trận",
         "—",
         "- Hiển thị dòng 'Không có danh mục kiểm tra bảo dưỡng'\n- KHÔNG hiển thị chân bảng (định mức, giá...)\n- Vẫn có nút 'Thêm danh mục kiểm tra bảo dưỡng'",
         "v-if maintains.length"),

        # --- Công thức giá
        ("011", "Tính Giá vốn theo công thức", "P0",
         "Công ty quản lý có đơn giá công = 100.000; nhập Định mức công = 2, Hệ số công nghệ = 1,5",
         "1. Chọn công ty quản lý\n2. Nhập Định mức công = 2\n3. Nhập Hệ số công nghệ = 1,5\n4. Đọc ô 'Giá vốn'",
         "100000 × 2 × 1,5",
         "- Ô 'Giá vốn' = 300,000\n- Ô này chỉ đọc (không nhập tay được)\n- Số phân cách hàng nghìn bằng dấu phẩy",
         "Giá vốn = đơn giá công × định mức công × hệ số công nghệ"),

        ("012", "Hệ số công nghệ bỏ trống được coi bằng 1", "P0",
         "Đơn giá công 100.000, Định mức công = 2, ô Hệ số công nghệ để TRỐNG",
         "1. Nhập Định mức công = 2, không nhập Hệ số công nghệ\n2. Đọc ô Giá vốn",
         "benefit_coefficient = rỗng",
         "- Giá vốn = 200,000 (coi hệ số = 1, KHÔNG phải 0)\n- Giá công thức tính tiếp từ 200,000",
         "primeCost(): hệ số rỗng quy về 1"),

        ("013", "Tính Giá công thức", "P0",
         "Giá vốn = 300.000; nhập Hệ số giá bán gói bảo dưỡng = 2",
         "1. Nhập Hệ số giá bán gói = 2\n2. Đọc ô 'Giá công thức'",
         "300000 × 2",
         "- Ô 'Giá công thức' = 600,000\n- Ô chỉ đọc\n- Hệ số giá bán gói để trống thì coi = 1",
         "Giá công thức = Giá vốn × Hệ số giá bán gói"),

        ("014", "Giá bán cơ sở tự điền theo Giá công thức", "P0",
         "Ma trận có 1 cột, đơn giá công 100.000",
         "1. Nhập Định mức công = 2\n2. Quan sát ô 'Giá bán cơ sở'",
         "quota_work = 2",
         "- Ô 'Giá bán cơ sở' tự điền = làm tròn của Giá công thức (200.000)\n- Ô này CHO PHÉP sửa tay\n- Nhập tiền hiển thị phân cách hàng nghìn",
         "onQuotaWorkInput() ghi đè base_price = ROUND(recipeCost)"),

        ("015", "Đổi Hệ số giá bán gói ghi đè Giá bán cơ sở đã sửa tay", "P0",
         "Đã tự sửa Giá bán cơ sở thành 999.999",
         "1. Sửa Giá bán cơ sở thành 999.999\n2. Đổi Hệ số giá bán gói bảo dưỡng từ 1 sang 2\n3. Quan sát ô Giá bán cơ sở",
         "coefficient: 1 → 2",
         "- Giá bán cơ sở bị GHI ĐÈ bằng giá công thức mới (mất giá trị 999.999)\n"
         "- Đây là hành vi đúng theo ERP — cần nhập lại giá tay sau khi đổi hệ số",
         "onCoefficientInput() luôn set lại base_price khi hệ số thực sự đổi"),

        ("016", "Gõ ký tự chữ vào ô hệ số KHÔNG ghi đè giá đã sửa tay", "P1",
         "Hệ số giá bán gói đang = 2; đã sửa Giá bán cơ sở thành 999.999",
         "1. Click vào ô Hệ số giá bán gói, gõ thêm ký tự 'a'\n2. Quan sát ô hệ số và ô Giá bán cơ sở",
         "Gõ ký tự bị lọc bỏ",
         "- Ô hệ số vẫn là 2 (ký tự chữ bị loại ngay, không hiển thị)\n"
         "- Giá bán cơ sở VẪN là 999.999 (không bị ghi đè vì hệ số không thực sự đổi)",
         "Chỉ tính lại khi giá trị sau khi lọc KHÁC giá trị cũ"),

        ("017", "Đổi công ty quản lý cập nhật lại giá vốn", "P0",
         "Công ty A đơn giá công 100.000; công ty B đơn giá công 150.000; Định mức công = 2",
         "1. Đang chọn công ty A, ghi lại Giá vốn\n2. Đổi Công ty quản lý sang B\n3. Quan sát Giá vốn, Giá công thức và Giá bán cơ sở",
         "work_price: 100000 → 150000",
         "- Giá vốn đổi từ 200,000 thành 300,000\n- Giá công thức tính lại theo giá vốn mới\n"
         "- Giá bán cơ sở GIỮ NGUYÊN (không bị ghi đè)\n- Dòng in đậm ở khối Giá vốn theo công ty chuyển sang công ty B, ô hệ số của B bị vô hiệu",
         "Đổi công ty chỉ đổi đơn giá công, không đụng base_price (đúng ERP)"),

        ("018", "Giá bán theo công ty = Giá bán cơ sở × hệ số công ty", "P0",
         "Giá bán cơ sở của cấp 1 = 200.000; công ty B có Hệ số giá bán = 1,2; công ty quản lý là A",
         "1. Nhập Giá bán cơ sở = 200.000\n2. Ở khối Giá vốn theo công ty, nhập Hệ số giá bán của công ty B = 1,2\n"
         "3. Đọc dòng 'Giá bán theo công ty' trong ma trận",
         "base_price = 200000; hệ số B = 1,2",
         "- Dòng công ty A hiển thị 200,000 (công ty quản lý luôn hệ số 1)\n"
         "- Dòng công ty B hiển thị 240,000\n- Dòng công ty quản lý được in đậm",
         "salePriceByCompany(); công ty quản lý cưỡng bức hệ số 1"),

        ("019", "Hệ số công ty bỏ trống được coi bằng 1", "P1",
         "Giá bán cơ sở = 200.000; xóa trắng ô Hệ số giá bán của công ty C",
         "1. Xóa trắng ô hệ số của công ty C\n2. Đọc dòng giá bán của công ty C",
         "coefficient C = rỗng",
         "- Giá bán của công ty C = 200,000 (hệ số coi = 1, KHÔNG phải 0)\n- Khi lưu, hệ số của C được ghi = 1",
         "Pivot coefficient NOT NULL DEFAULT 1 — trống quy về 1"),

        ("020", "Nhập 'Gợi ý hàng hoá' dạng thẻ", "P1",
         "Ma trận có 1 cột cấp",
         "1. Ở dòng 'Gợi ý hàng hoá', gõ 'Dao tiện' rồi Enter\n2. Gõ tiếp 'Mũi khoan, Bulong' (có dấu phẩy)\n3. Quan sát",
         "3 gợi ý",
         "- Mỗi giá trị hiển thị thành 1 thẻ riêng, dấu phẩy tự tách thành 2 thẻ\n"
         "- Bấm x trên thẻ hoặc phím xóa lùi loại bỏ được thẻ\n- Lưu rồi mở lại: các thẻ vẫn còn đủ",
         "key_word lưu dạng [{'text': '...'}] để cổng ERP đọc được"),

        ("021", "Ô nhập số chỉ nhận ký tự số", "P1",
         "Đang ở màn Thêm",
         "1. Gõ 'abc12.5xyz' vào ô 'Định mức đàm phán giá (%)'\n2. Gõ tương tự vào VAT (%), Hệ số giá bán gói, Định mức công, Hệ số công nghệ, SL",
         "Chuỗi lẫn chữ",
         "- Các ô chỉ giữ lại '12.5' (ô SL chỉ giữ số nguyên '125')\n- Ký tự chữ bị loại ngay khi gõ, không đợi tới lúc Lưu\n"
         "- Tối đa 2 chữ số thập phân",
         "sanitizeNumberEvent(): lọc và ghi đè lại giá trị trên DOM"),

        # --- Hàng hóa
        ("022", "Mở popup chọn hàng hóa", "P0",
         "Đang ở màn Thêm/Sửa",
         "1. Bấm nút 'Hàng hóa' ở khối 'Áp dụng cho hàng hóa'\n2. Quan sát popup",
         "—",
         "- Mở popup 'Bộ lọc hàng hoá' với ô tìm nhanh + bộ lọc nâng cao nhiều tiêu chí (tính chất, loại, thương hiệu, hãng sản xuất, xuất xứ, lĩnh vực, chương, nhóm CV, cụm CV, nhóm hàng, model, hãng/loại/đời xe...)\n"
         "- Bảng kết quả có ô tick chọn, hiển thị ảnh, tên, mã, ĐVT, giá niêm yết, bảo hành\n- Nút 'Thêm N hàng hoá' ở cuối",
         "ProductSearchModal — UX như popup chọn hàng của báo giá"),

        ("023", "Tìm và thêm nhiều hàng hóa cùng lúc", "P0",
         "Có ≥ 5 hàng hóa tên chứa 'bulong'",
         "1. Mở popup Hàng hóa\n2. Nhập tìm nhanh 'bulong', tìm kiếm\n3. Tick 3 dòng\n4. Bấm 'Thêm 3 hàng hoá'\n5. Đóng popup, quan sát bảng ngoài",
         "3 hàng hóa",
         "- Toast 'Đã thêm 3 hàng hóa'\n- Popup KHÔNG tự đóng (còn chọn tiếp được)\n"
         "- Bảng ngoài nhóm hàng theo nhóm hàng hóa, mỗi nhóm có tiêu đề 'Nhóm hàng: <tên>'\n"
         "- Mỗi dòng hiển thị STT, ảnh, tên hàng, mã hàng và nút xóa",
         "onApplyProducts() gom theo group_id"),

        ("024", "Thêm trùng hàng hóa đã có", "P0",
         "Đã thêm hàng 'Bulong M10' vào gói",
         "1. Mở popup Hàng hóa, tick lại đúng 'Bulong M10' và 1 hàng mới\n2. Bấm 'Thêm 2 hàng hoá'\n3. Quan sát bảng ngoài",
         "1 trùng + 1 mới",
         "- Toast 'Đã thêm 1 hàng hóa' và cảnh báo '1 hàng hóa đã được chọn trước đó'\n"
         "- Bảng ngoài chỉ tăng 1 dòng, KHÔNG có dòng trùng lặp",
         "Chống trùng theo product.id trong từng nhóm"),

        ("025", "Thêm cả nhóm hàng hóa", "P0",
         "Nhóm 'Vòng bi' có 12 hàng hóa đang hoạt động",
         "1. Bấm nút 'Nhóm hàng hóa'\n2. Tìm 'Vòng bi', tick nhóm đó\n3. Bấm nút thêm\n4. Đóng popup, quan sát bảng ngoài",
         "1 nhóm, 12 hàng",
         "- Toast 'Đã thêm 12 hàng hóa từ 1 nhóm'\n- Bảng ngoài xuất hiện khối 'Nhóm hàng: Vòng bi' với đủ 12 dòng\n"
         "- Popup hiển thị số lượng hàng của từng nhóm để đối chiếu",
         "onApplyGroups() gọi search-products?group_id=... lấy đủ hàng của nhóm (không bị cắt 50)"),

        ("026", "Thêm nhóm không có hàng hóa nào", "P2",
         "Nhóm 'Nhóm rỗng' đang hoạt động nhưng không có hàng hóa nào status ≠ 0",
         "1. Bấm 'Nhóm hàng hóa', tick 'Nhóm rỗng', bấm thêm\n2. Quan sát",
         "Nhóm 0 hàng",
         "- Cảnh báo '1 nhóm không có hàng hóa nào'\n- Bảng ngoài KHÔNG xuất hiện khối nhóm rỗng\n- Không lỗi JS",
         "emptyGroups được đếm và cảnh báo riêng"),

        ("027", "Xóa 1 hàng hóa khỏi nhóm", "P0",
         "Nhóm 'Vòng bi' trong gói đang có 3 hàng",
         "1. Bấm nút xóa ở 1 dòng hàng hóa\n2. Quan sát",
         "Xóa 1/3 hàng",
         "- Dòng hàng biến mất ngay, không hỏi xác nhận\n- Khối nhóm 'Vòng bi' vẫn còn với 2 dòng\n- STT trong nhóm đánh lại từ 1",
         "removeProduct()"),

        ("028", "Xóa hàng hóa cuối cùng làm mất luôn nhóm", "P0",
         "Nhóm 'Vòng bi' trong gói chỉ còn 1 hàng",
         "1. Xóa hàng hóa cuối cùng của nhóm\n2. Quan sát bảng",
         "Xóa hàng duy nhất",
         "- Cả khối 'Nhóm hàng: Vòng bi' biến mất (không để lại nhóm rỗng)\n- Các nhóm khác không bị ảnh hưởng",
         "removeProduct() tự loại nhóm khi hết hàng (đúng ERP)"),

        ("029", "Xóa cả nhóm hàng có xác nhận", "P0",
         "Nhóm 'Vòng bi' có 12 hàng trong gói",
         "1. Bấm nút xóa trên dòng tiêu đề 'Nhóm hàng: Vòng bi'\n2. Đọc hộp xác nhận\n3. Bấm 'Xác nhận'",
         "Xóa nhóm 12 hàng",
         "- Hộp xác nhận: 'Nếu bạn xóa nhóm hàng này, toàn bộ hàng hóa trong nhóm sẽ bị xóa. Bạn có chắc chắn?'\n"
         "- Sau xác nhận: cả nhóm và 12 dòng biến mất\n- Các nhóm khác giữ nguyên",
         "askRemoveGroup() + BaseConfirmModal"),

        # --- File
        ("030", "Đính kèm file PDF", "P0",
         "Đang ở màn Thêm; có file huongdan.pdf",
         "1. Bấm ô 'Thêm file'\n2. Chọn huongdan.pdf\n3. Quan sát khối File đính kèm\n4. Nhập đủ thông tin bắt buộc và Lưu\n5. Mở lại màn Sửa gói vừa tạo",
         "1 file PDF",
         "- Sau bước 2: hiện thẻ file với icon PDF, tên file và nút x để bỏ chọn\n"
         "- Sau bước 4: lưu thành công\n- Bước 5: file hiển thị ở danh sách file đã lưu, bấm vào mở được file trên tab mới",
         "Upload S3 prefix `services`; attachments lưu chuỗi URL"),

        ("031", "Chọn file không phải PDF", "P0",
         "Có file anh.png",
         "1. Bấm 'Thêm file'\n2. Chọn anh.png (đổi bộ lọc file thành All files)\n3. Quan sát",
         "File .png",
         "- Cảnh báo 'File \"anh.png\" không phải PDF'\n- File KHÔNG được thêm vào danh sách\n"
         "- Nếu vẫn gửi lên được, BE trả 422 với 'Chỉ nhận file PDF'",
         "FE kiểm đuôi .pdf; BE rule mimes:pdf"),

        ("032", "Bỏ file mới chọn trước khi lưu", "P1",
         "Đã chọn 2 file PDF nhưng chưa lưu",
         "1. Bấm nút x trên thẻ file thứ nhất\n2. Bấm Lưu",
         "2 file → bỏ 1",
         "- Thẻ file bị bỏ biến mất khỏi danh sách\n- Sau khi lưu chỉ có 1 file được đính kèm",
         "removeNewFile()"),

        ("033", "File đã lưu KHÔNG có nút xóa", "P1",
         "Gói đã có 2 file PDF đã lưu",
         "1. Mở màn Sửa gói đó\n2. Quan sát 2 thẻ file đã lưu",
         "2 file đã lưu",
         "- 2 thẻ file đã lưu KHÔNG có nút x (chỉ file mới chọn mới bỏ được)\n"
         "- Bấm vào thẻ mở được file\n- Đây là hành vi nguyên trạng ERP: file đính kèm chỉ thêm, không xóa",
         "uploadAttachments() chỉ nối thêm, không xóa"),

        ("034", "Thêm file vào gói đã có file", "P1",
         "Gói đã có 1 file PDF",
         "1. Mở Sửa, thêm 1 file PDF mới\n2. Bấm Lưu\n3. Mở lại màn Sửa",
         "1 file cũ + 1 file mới",
         "- Sau khi lưu, gói có ĐỦ 2 file (file cũ không bị mất, không bị ghi đè)\n- Cả 2 mở được",
         "attachments = chuỗi cũ + ', ' + URL mới"),

        # --- Lưu / Sửa
        ("035", "Tạo gói bảo dưỡng đầy đủ thông tin", "P0",
         "Chưa tồn tại gói tên 'Gói kiểm thử 01' hoặc mã 'GKT-01'",
         "1. Bấm 'Thêm mới'\n2. Tên = 'Gói kiểm thử 01', Mã = 'gkt-01', Định mức đàm phán giá = 10, VAT = 8\n"
         "3. Công ty quản lý = công ty của mình, Ghi chú = 'ghi chú test', Hệ số giá bán gói = 2\n"
         "4. Thêm 2 cột cấp (Cấp 1, Cấp 2), thêm 1 hàng 'Kiểm tra dầu máy' (ĐVT 'Lần', SL 1), chọn ghi chú cho cả 2 ô\n"
         "5. Nhập Định mức công cột 1 = 2, cột 2 = 3\n6. Thêm 2 hàng hóa\n7. Bấm Lưu",
         "Bộ dữ liệu đầy đủ",
         "- Toast 'Tạo gói bảo dưỡng thành công'\n- Quay về màn danh sách, gói mới đứng đầu\n"
         "- Mã hiển thị IN HOA 'GKT-01' (dù nhập chữ thường)\n- Trạng thái mặc định 'Hoạt động'\n"
         "- Rê chuột vào tên: tooltip hiện giá của 2 cấp",
         "Mã tự chuyển in hoa; gói mới luôn STATUS_ACTIVE"),

        ("036", "Mở màn Sửa và kiểm tra dữ liệu nạp lại", "P0",
         "Gói 'Gói kiểm thử 01' đã có 2 cột cấp, 1 hàng, 2 hàng hóa, 1 file PDF",
         "1. Bấm nút 'Sửa' trên dòng đó\n2. Đối chiếu từng khối với dữ liệu đã lưu",
         "id gói kiểm thử 01",
         "- Tiêu đề 'Sửa gói bảo dưỡng'\n- Khối 1 nạp đủ tên, mã, VAT, định mức đàm phán, công ty, ghi chú, hệ số giá bán\n"
         "- Hiện thêm ô 'Trạng thái' với giá trị hiện tại\n"
         "- Ma trận nạp đúng 2 cột (đúng cấp) và 1 hàng, ghi chú từng ô đúng, định mức/hệ số/giá bán cơ sở/gợi ý hàng hoá đúng\n"
         "- Khối hàng hóa hiện đủ 2 hàng theo nhóm; khối file hiện 1 file\n- Cuối trang có thêm nút 'Sao chép'",
         "dataForEdit() dựng lại ma trận từ maintains + service_levels"),

        ("037", "Sửa và lưu thành công", "P0",
         "Gói 'Gói kiểm thử 01' đang có ghi chú 'ghi chú test'",
         "1. Bấm Sửa\n2. Đổi Ghi chú thành 'ghi chú đã sửa', thêm 1 hàng nội dung kiểm tra mới\n3. Bấm Lưu",
         "Đổi note + thêm 1 hàng",
         "- Toast 'Cập nhật gói bảo dưỡng thành công', quay về danh sách\n"
         "- Cột 'Ngày sửa' và 'Người sửa' cập nhật theo user đang đăng nhập\n"
         "- Mở lại màn Sửa: có đủ 2 hàng và ghi chú mới",
         "POST /v1/customer-care/services/{id}"),

        ("038", "Đổi trạng thái sang Khóa ở màn Sửa", "P0",
         "Gói 'Gói kiểm thử 01' đang Hoạt động",
         "1. Bấm Sửa\n2. Đổi Trạng thái sang 'Khóa'\n3. Bấm Lưu\n4. Quan sát danh sách",
         "status: 1 → 0",
         "- Lưu thành công\n- Dòng hiển thị pill 'Khóa'\n- Gói vẫn hiển thị trong danh sách và vẫn mở Sửa được\n"
         "- Lọc Trạng thái = 'Khóa' có gói này",
         "status required in:0,1 khi update"),

        ("039", "Nút Hủy quay về danh sách không lưu", "P1",
         "Đang ở màn Sửa, đã đổi Tên gói",
         "1. Đổi Tên gói thành 'ABC XYZ'\n2. Bấm 'Hủy'\n3. Quan sát danh sách",
         "Thay đổi chưa lưu",
         "- Quay về màn danh sách, không có toast thành công\n- Tên gói trên danh sách vẫn là tên cũ\n"
         "- Mở lại màn Sửa: dữ liệu vẫn là bản cũ",
         "goBack() chỉ điều hướng, không gọi API"),

        # --- Sao chép
        ("040", "Sao chép gói từ màn danh sách", "P0",
         "Gói 'Gói kiểm thử 01' có đủ ma trận, hàng hóa, file",
         "1. Bấm nút 'Sao chép' trên dòng đó\n2. Quan sát màn tạo mới",
         "copy_from = id gói nguồn",
         "- Chuyển sang /customer-care/services/create?copy_from=<id>\n- Tiêu đề 'Thêm gói bảo dưỡng'\n"
         "- Hiện dải nhắc 'Đang sao chép từ gói <tên> — hãy đổi tên/mã trước khi lưu' và toast tương ứng\n"
         "- Toàn bộ ma trận, hàng hóa, hệ số công ty, file đính kèm được sao chép sang\n"
         "- Tên và Mã giữ nguyên của gói nguồn (chờ user đổi); KHÔNG hiện ô Trạng thái",
         "loadService(id, isCopy = true)"),

        ("041", "Lưu bản sao khi chưa đổi tên/mã", "P0",
         "Đang ở màn sao chép, giữ nguyên tên và mã gói nguồn",
         "1. Bấm Lưu ngay",
         "name/code trùng gói nguồn",
         "- API trả 422\n- Ô Tên gói và Mã gói đều báo lỗi 'Đã tồn tại' (viền đỏ)\n"
         "- Toast 'Vui lòng kiểm tra lại dữ liệu nhập', trang tự cuộn tới ô lỗi đầu tiên\n- KHÔNG tạo bản ghi",
         "name và code unique toàn bảng services"),

        ("042", "Lưu bản sao sau khi đổi tên/mã", "P0",
         "Đang ở màn sao chép từ 'Gói kiểm thử 01' (2 cột cấp, 1 hàng, 2 hàng hóa, 1 file)",
         "1. Đổi Tên = 'Gói kiểm thử 01 - bản sao', Mã = 'GKT-01-C'\n2. Bấm Lưu\n"
         "3. Mở màn Sửa gói vừa tạo, đối chiếu từng khối với gói gốc",
         "Tên/mã mới",
         "- Tạo thành công, gói mới đứng đầu danh sách với trạng thái 'Hoạt động'\n"
         "- Ma trận, hàng hóa, hệ số công ty giống hệt gói gốc\n"
         "- File đính kèm của gói gốc ĐƯỢC mang sang gói mới (mở được)\n- Gói gốc không bị thay đổi gì",
         "existing_attachments được gửi kèm để BE ghi vào gói mới"),

        ("043", "Sao chép từ trong màn Sửa", "P1",
         "Đang ở màn Sửa gói 'Gói kiểm thử 01'; user có quyền Thêm",
         "1. Bấm nút 'Sao chép' cuối trang\n2. Quan sát",
         "—",
         "- Mở màn tạo bản sao của chính gói đang sửa\n- Hiển thị dải nhắc đổi tên/mã\n"
         "- Nếu user KHÔNG có quyền Thêm thì nút này bị ẩn",
         "Nút Sao chép ở màn sửa dùng quyền `Thêm danh mục gói bảo dưỡng`"),

        # --- Xóa / Khóa
        ("044", "Xóa gói chưa được sử dụng", "P0",
         "Gói 'GKT-XOA': chưa gắn hàng hóa nào và chưa xuất hiện ở báo giá dịch vụ",
         "1. Bấm nút 'Xóa' trên dòng đó\n2. Đọc hộp xác nhận\n3. Bấm 'Xóa'\n4. Kiểm tra DB",
         "Gói chưa dùng",
         "- Hộp xác nhận có nội dung: 'Gói đang được sử dụng sẽ chuyển sang trạng thái Khóa thay vì xóa. Bạn có chắc chắn muốn xóa <tên>?'\n"
         "- Toast 'Xóa gói bảo dưỡng thành công'\n- Dòng biến mất khỏi danh sách\n"
         "- Các bảng con (`service_maintains`, `service_maintain_levels`, `service_levels`) của gói cũng bị xóa sạch",
         "destroy(): isCanDelete = true → xóa gói + bảng con"),

        ("045", "Xóa gói ĐÃ gắn hàng hóa → chuyển Khóa", "P0",
         "Gói 'GKT-DUNG' đã gắn ≥ 1 hàng hóa (service_has_products)",
         "1. Bấm 'Xóa' trên dòng đó, xác nhận\n2. Quan sát danh sách và DB",
         "Gói đã gắn hàng hóa",
         "- Toast 'Gói bảo dưỡng đang được sử dụng nên đã được chuyển sang trạng thái Khóa'\n"
         "- Dòng VẪN còn trong danh sách nhưng đổi sang pill 'Khóa'\n- Bản ghi và toàn bộ dữ liệu con vẫn nguyên trong DB",
         "isCanDelete kiểm service_has_products + service_quotation_items"),

        ("046", "Xóa gói đã dùng ở báo giá dịch vụ → chuyển Khóa", "P0",
         "Gói 'GKT-BAOGIA' đã xuất hiện trong service_quotation_items (dù chưa gắn hàng hóa)",
         "1. Bấm 'Xóa', xác nhận\n2. Quan sát",
         "Gói có trong báo giá dịch vụ",
         "- Toast báo đã chuyển sang trạng thái Khóa\n- Gói vẫn còn, trạng thái 'Khóa'\n- Dữ liệu báo giá dịch vụ liên quan không bị ảnh hưởng",
         "isCanDelete kiểm service_quotation_items theo service_id"),

        ("047", "Hủy thao tác xóa", "P1",
         "Gói bất kỳ đang hiển thị",
         "1. Bấm 'Xóa'\n2. Bấm 'Hủy'\n3. Quan sát",
         "—",
         "- Hộp xác nhận đóng\n- Gói vẫn còn nguyên trạng thái cũ\n- Không phát sinh request DELETE",
         "Chỉ gọi API khi bấm nút xác nhận"),

        # --- In
        ("048", "Mở màn in phiếu kiểm tra bảo dưỡng", "P0",
         "Gói 'Gói kiểm thử 01' có 2 cột cấp và 3 hàng nội dung kiểm tra",
         "1. Bấm nút 'In' trên dòng\n2. Quan sát tab mới",
         "id gói",
         "- Mở TAB MỚI tới /customer-care/services/<id>/print\n"
         "- Hiển thị bản xem trước: ảnh letterhead công ty của user đăng nhập, tên gói IN HOA, bảng 'Nội dung kiểm tra bảo dưỡng'\n"
         "- Bảng có cột STT, Nội dung kiểm tra bảo dưỡng, SL, nhóm cột 'Cấp bảo dưỡng' (2 cột theo tên cấp), 'Kiểm tra' (Có/Không), Ghi chú\n"
         "- Bên dưới có chú giải các ký hiệu ghi chú, khối 'Nội dung đề xuất' và bảng ký KỸ THUẬT VIÊN / KHÁCH HÀNG",
         "Mẫu in report_templates id 191, dùng chung với ERP"),

        ("049", "Bấm nút In để mở hộp thoại in", "P0",
         "Đang ở màn xem trước bản in",
         "1. Bấm nút 'In' màu xanh\n2. Quan sát hộp thoại in của trình duyệt",
         "—",
         "- Mở hộp thoại in\n- Bản xem trong hộp thoại: KHÔNG có nút 'In', bảng có đầy đủ viền đen, font Times New Roman\n"
         "- Bảng ký KỸ THUẬT VIÊN/KHÁCH HÀNG KHÔNG có viền\n- Nội dung không tràn lề phải",
         "Nút In phải bấm tay (tự mở sẽ bị chặn popup); style truyền qua options.styles"),

        ("050", "In gói không tồn tại", "P2",
         "Không tồn tại gói id = 999999",
         "1. Truy cập /customer-care/services/999999/print",
         "id không tồn tại",
         "- Hiển thị thông báo lỗi đỏ 'Không thể in gói bảo dưỡng này (gói không tồn tại hoặc có lỗi tải dữ liệu)'\n"
         "- Nút In bị vô hiệu\n- Không lỗi trắng màn hình",
         "loadError được hiển thị thay cho nội dung"),

        ("051", "In gói chưa có nội dung kiểm tra nào", "P2",
         "Gói 'GBD-RONG' chưa có hàng nội dung kiểm tra và chưa có cột cấp",
         "1. Bấm In gói đó\n2. Quan sát bản xem trước",
         "maintains rỗng",
         "- Bảng in vẫn hiện với dòng tiêu đề nhưng phần thân rỗng\n- Vẫn có chú giải ghi chú và bảng ký\n- Không lỗi 500, không trắng trang",
         "buildPrintData() vẫn dựng khung bảng khi không có dữ liệu"),

        # --- Export
        ("052", "Xuất Excel danh sách", "P0",
         "Danh mục có 30 gói bảo dưỡng",
         "1. Bấm nút 'Xuất excel'\n2. Chờ tải xong, mở file",
         "30 gói",
         "- Tải về file 'Danh_sach_dich_vu.xlsx'\n- Toast 'Xuất Excel thành công'\n"
         "- File chứa ĐỦ 30 gói (không chỉ trang hiện tại)\n- Có các cột tên, mã, công ty, trạng thái, người tạo/sửa, ngày tạo/sửa và giá theo cấp",
         "export() lấy toàn bộ Service, không phân trang"),

        ("053", "Xuất Excel KHÔNG áp bộ lọc đang chọn", "P0",
         "Danh mục có 30 gói; đang lọc Trạng thái = 'Khóa' (chỉ còn 3 dòng trên màn)",
         "1. Lọc Trạng thái = 'Khóa'\n2. Bấm 'Xuất excel'\n3. Mở file và đếm số dòng",
         "Lọc còn 3 dòng",
         "- File Excel chứa ĐỦ 30 gói, KHÔNG chỉ 3 gói đang lọc\n"
         "- Đây là hành vi giữ nguyên theo ERP — cần xác nhận với nghiệp vụ nếu mong muốn xuất theo bộ lọc",
         "ERP export toàn bộ, không nhận tham số lọc"),

        ("054", "Xuất Excel khi danh mục rỗng", "P2",
         "Bảng services không có bản ghi nào (hoặc môi trường test trắng)",
         "1. Bấm 'Xuất excel'\n2. Mở file",
         "0 bản ghi",
         "- Vẫn tải về file hợp lệ với dòng tiêu đề, phần dữ liệu rỗng\n- Không lỗi 500, không file hỏng",
         "Excel::download vẫn chạy với collection rỗng"),
    ]),

    # ---------------------------------------------------------------- VI
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Bỏ trống toàn bộ trường bắt buộc ở Thông tin chung", "P0",
         "Đang ở màn Thêm, chưa nhập gì",
         "1. Bấm Lưu ngay",
         "Tất cả trống",
         "- Ô 'Tên gói bảo dưỡng' và 'Mã gói bảo dưỡng' báo 'Bắt buộc phải nhập' (viền đỏ)\n"
         "- Ô 'Công ty quản lý' báo lỗi nếu chưa chọn được công ty mặc định\n"
         "- Toast 'Vui lòng kiểm tra lại dữ liệu nhập', trang tự cuộn tới ô lỗi đầu tiên\n- KHÔNG gửi request lên server",
         "validate() chặn ở FE trước khi submit"),

        ("002", "Lỗi chỉ hiện SAU lần bấm Lưu đầu tiên", "P0",
         "Vừa mở màn Thêm",
         "1. Quan sát ngay khi vào form\n2. Click vào ô Tên rồi click ra ngoài (không nhập)\n3. Bấm Lưu",
         "—",
         "- Bước 1 và 2: KHÔNG có ô nào viền đỏ hay hiện lỗi\n- Bước 3: lỗi mới hiện đồng loạt ở các ô còn thiếu",
         "Cờ touched — đúng chuẩn CLAUDE.md"),

        ("003", "Tạo trùng TÊN gói bảo dưỡng", "P0",
         "Đã tồn tại gói tên 'Gói kiểm thử 01'",
         "1. Thêm mới, nhập Tên = 'Gói kiểm thử 01', Mã = 'GKT-KHAC', đủ trường khác\n2. Bấm Lưu",
         "name trùng",
         "- API trả 422\n- Ô Tên gói báo 'Đã tồn tại', viền đỏ\n- Toast 'Vui lòng kiểm tra lại dữ liệu nhập'\n- Không tạo bản ghi",
         "Rule unique services.name"),

        ("004", "Tạo trùng MÃ gói bảo dưỡng (kể cả khác hoa/thường)", "P0",
         "Đã tồn tại gói mã 'GKT-01'",
         "1. Thêm mới, nhập Mã = 'gkt-01' (chữ thường), Tên khác hoàn toàn\n2. Bấm Lưu",
         "code = 'gkt-01'",
         "- Mã tự chuyển thành 'GKT-01' ngay khi gõ\n- API trả 422, ô Mã báo 'Đã tồn tại'\n- Không tạo bản ghi",
         "code luôn in hoa (FE + BE) rồi mới kiểm unique"),

        ("005", "Sửa giữ nguyên tên/mã của chính nó", "P0",
         "Gói 'Gói kiểm thử 01' mã 'GKT-01' đã tồn tại",
         "1. Bấm Sửa gói đó\n2. Giữ nguyên Tên và Mã, chỉ đổi Ghi chú\n3. Bấm Lưu",
         "name/code giữ nguyên",
         "- Lưu thành công, KHÔNG báo 'Đã tồn tại'\n- Ghi chú mới được lưu",
         "Rule unique có ignore theo id đang sửa"),

        ("006", "Tên / Mã vượt 255 ký tự", "P1",
         "Đang ở màn Thêm",
         "1. Nhập Tên = chuỗi 300 ký tự, Mã = chuỗi 300 ký tự\n2. Nhập đủ trường còn lại, bấm Lưu",
         "Chuỗi 300 ký tự",
         "- API trả 422\n- Cả 2 ô báo 'Tối đa 255 ký tự'\n- Không tạo bản ghi, không lỗi 500",
         "Rule max:255 cho name và code"),

        ("007", "Ghi chú vượt 255 ký tự", "P2",
         "Đang ở màn Thêm",
         "1. Nhập Ghi chú = chuỗi 300 ký tự\n2. Nhập đủ trường bắt buộc, bấm Lưu",
         "note 300 ký tự",
         "- API trả 422, ô Ghi chú báo 'Tối đa 255 ký tự'\n- Không tạo bản ghi",
         "Rule note max:255"),

        ("008", "VAT vượt khoảng cho phép", "P0",
         "Đang ở màn Thêm",
         "1. Nhập VAT = 150\n2. Nhập đủ trường bắt buộc, bấm Lưu",
         "vat_percent = 150",
         "- API trả 422, ô VAT báo 'Tối đa 100'\n- Không tạo bản ghi\n"
         "- Nhập VAT = 8 thì lưu bình thường",
         "Rule vat_percent nullable numeric min:0 max:100 (ERP thiếu numeric → HRM đã sửa)"),

        ("009", "VAT để trống được lưu thành 0", "P0",
         "Đang ở màn Thêm",
         "1. Để trống ô VAT\n2. Nhập đủ trường bắt buộc, bấm Lưu\n3. Kiểm tra cột vat_percent trong DB",
         "vat_percent = rỗng",
         "- Lưu thành công (VAT không bắt buộc)\n- Cột vat_percent trong DB = 0, KHÔNG phải null\n"
         "- Mở lại màn Sửa: ô VAT hiển thị 0",
         "BE quy trống về 0 vì cột NOT NULL DEFAULT 0"),

        ("010", "Định mức đàm phán giá vượt 99", "P1",
         "Đang ở màn Thêm",
         "1. Nhập Định mức đàm phán giá = 120\n2. Nhập đủ trường bắt buộc, bấm Lưu",
         "sale_max_percent = 120",
         "- API trả 422, ô đó báo 'Tối đa 99'\n- Không tạo bản ghi",
         "Rule sale_max_percent nullable numeric min:0 max:99"),

        ("011", "Hệ số giá bán gói nhỏ hơn 1", "P0",
         "Đang ở màn Thêm",
         "1. Nhập Hệ số giá bán gói bảo dưỡng = 0.5\n2. Nhập đủ trường bắt buộc, bấm Lưu",
         "coefficient = 0.5",
         "- API trả 422, ô đó báo 'Không được nhỏ hơn 1'\n- Không tạo bản ghi\n"
         "- Ghi nhận: rule này HRM áp cho CẢ tạo mới (ERP chỉ kiểm khi sửa)",
         "Rule coefficient_cost_price_service nullable numeric min:1 max:100"),

        ("012", "Hệ số giá bán gói để trống được lưu = 1", "P1",
         "Đang ở màn Thêm",
         "1. Để trống ô Hệ số giá bán gói\n2. Nhập đủ trường bắt buộc, bấm Lưu\n3. Mở lại màn Sửa",
         "coefficient = rỗng",
         "- Lưu thành công\n- Mở lại màn Sửa thấy hệ số = 1\n- Giá công thức trên form bằng đúng giá vốn",
         "BE: coefficient_cost_price_service ?? 1"),

        ("013", "Dấu phẩy được hiểu là dấu thập phân", "P1",
         "Đang ở màn Thêm",
         "1. Nhập VAT = '8,5' (dùng dấu phẩy)\n2. Nhập đủ trường bắt buộc, bấm Lưu\n3. Kiểm tra giá trị đã lưu",
         "vat_percent = '8,5'",
         "- Lưu thành công với giá trị 8.5 (KHÔNG phải 85)\n- Không báo lỗi 'Phải là số'",
         "prepareForValidation() đổi ',' thành '.' cho 3 trường phần trăm/hệ số"),

        ("014", "Bỏ trống trường bắt buộc trong ma trận", "P0",
         "Ma trận có 1 cột cấp và 1 hàng, chưa nhập gì trong hàng",
         "1. Thêm 1 cột cấp và 1 hàng trống\n2. Nhập đủ Thông tin chung\n3. Bấm Lưu",
         "Hàng trống",
         "- Báo 'Bắt buộc phải nhập' tại: ô Nội dung kiểm tra, ô ĐVT, ô SL, ô ghi chú của cột, ô Định mức công\n"
         "- Nếu cột cấp chưa chọn cấp thì ô chọn cấp cũng báo lỗi\n- Không gửi request",
         "validate() kiểm từng ô của ma trận theo đúng key BE"),

        ("015", "Có hàng nội dung nhưng chưa có cột cấp nào", "P0",
         "Đã thêm 1 hàng nội dung kiểm tra, ma trận chưa có cột cấp nào",
         "1. Thêm 1 hàng, nhập đủ Nội dung / ĐVT / SL\n2. Nhập đủ Thông tin chung\n3. Bấm Lưu",
         "levelCols rỗng, maintains có 1 dòng",
         "- Cảnh báo 'Cần ít nhất 1 cột cấp bảo dưỡng'\n- Không gửi request lên server\n- Trang không điều hướng",
         "validate(): maintains có dòng thì buộc phải có ≥ 1 cột cấp"),

        ("016", "Tạo gói KHÔNG có nội dung kiểm tra nào", "P0",
         "Đang ở màn Thêm; chưa tồn tại mã 'GKT-EMPTY'",
         "1. Nhập đủ Thông tin chung (Tên, Mã 'GKT-EMPTY', Công ty)\n2. KHÔNG thêm cột cấp, KHÔNG thêm hàng nào\n3. Bấm Lưu",
         "maintains rỗng",
         "- Lưu THÀNH CÔNG (ma trận không bắt buộc)\n- Gói mới xuất hiện trong danh sách\n"
         "- Rê chuột vào tên: không có tooltip giá (chưa có cấp nào)",
         "maintains là nullable ở BE"),

        ("017", "Xóa hết hàng ma trận rồi lưu — dữ liệu cũ KHÔNG bị xóa", "P0",
         "Gói 'Gói kiểm thử 01' đang có 3 hàng nội dung kiểm tra",
         "1. Bấm Sửa\n2. Xóa hết cả 3 hàng nội dung kiểm tra\n3. Bấm Lưu\n4. Mở lại màn Sửa và kiểm tra DB",
         "Gửi lên mảng maintains rỗng",
         "- Lưu thành công (không báo lỗi)\n- Ghi nhận thực tế: 3 hàng CŨ VẪN CÒN trong DB và hiện lại khi mở Sửa\n"
         "- Đây là hành vi nguyên trạng ERP (mảng rỗng thì bỏ qua, không xóa) — nếu nghiệp vụ muốn xóa được hết thì ghi Failed",
         "saveServiceMaintain(): mảng rỗng → return ngay, không đụng dữ liệu cũ"),

        ("018", "Xóa hết hàng hóa rồi lưu — dữ liệu CÓ bị xóa", "P0",
         "Gói 'Gói kiểm thử 01' đang gắn 5 hàng hóa",
         "1. Bấm Sửa\n2. Xóa hết các nhóm hàng hóa\n3. Bấm Lưu\n4. Mở lại màn Sửa và kiểm tra pivot service_has_products",
         "product_groups rỗng",
         "- Lưu thành công\n- Toàn bộ pivot hàng hóa của gói bị XÓA sạch\n"
         "- Mở lại màn Sửa: khối hàng hóa hiển thị 'Chưa chọn hàng hóa áp dụng'\n"
         "- Lưu ý sự KHÁC BIỆT với ma trận ở TC_06.017 (ma trận không xóa, hàng hóa thì xóa)",
         "syncProducts() luôn sync kể cả mảng rỗng"),

        ("019", "Bỏ cấp bảo dưỡng đang được dùng ở báo giá dịch vụ", "P0",
         "Gói 'GKT-BAOGIA' có cột 'Cấp 2' đang được tham chiếu bởi service_quotation_items",
         "1. Bấm Sửa gói đó\n2. Xóa cột 'Cấp 2' và xác nhận\n3. Bấm Lưu\n4. Kiểm tra DB",
         "Cấp 2 đang dùng ở báo giá",
         "- API trả lỗi với thông báo 'Không thể xóa cấp dịch vụ đã được sử dụng!'\n- Hiện toast đúng thông báo đó\n"
         "- QUAN TRỌNG: dữ liệu ma trận trong DB KHÔNG bị hỏng — mở lại màn Sửa vẫn thấy đủ cột Cấp 2 và các hàng như trước khi thao tác",
         "Controller bọc DB::transaction → rollback toàn bộ khi ném lỗi ở bước sau"),

        ("020", "Bỏ cấp bảo dưỡng CHƯA được dùng", "P0",
         "Gói có cột 'Cấp 3' chưa xuất hiện ở báo giá dịch vụ nào",
         "1. Bấm Sửa\n2. Xóa cột 'Cấp 3', xác nhận\n3. Bấm Lưu\n4. Mở lại màn Sửa",
         "Cấp 3 chưa dùng",
         "- Lưu thành công\n- Cột Cấp 3 biến mất khỏi ma trận\n- Các cột còn lại giữ nguyên định mức, giá, ghi chú (không bị lệch)",
         "Chỉ xóa service_levels của cấp không còn dùng"),

        ("021", "Giữ nguyên cấp cũ khi lưu lại không phá liên kết báo giá", "P0",
         "Gói có cấp 1 đang được báo giá dịch vụ tham chiếu (service_level_id = X)",
         "1. Ghi lại id bản ghi service_levels của cấp 1\n2. Bấm Sửa gói, đổi Định mức công của cấp 1, bấm Lưu\n"
         "3. Kiểm tra lại id bản ghi service_levels của cấp 1",
         "service_level_id = X",
         "- Lưu thành công, định mức mới được ghi nhận\n- id bản ghi service_levels VẪN là X (không bị tạo mới)\n"
         "- Báo giá dịch vụ đang trỏ tới X không bị mồ côi",
         "firstOrNew theo (service_id, level_id) — giữ id"),

        ("022", "Gửi cấp bảo dưỡng không tồn tại qua API", "P1",
         "level_id = 999999 không có trong bảng levels",
         "1. Gọi POST /v1/customer-care/services với maintains[0].levels[0].level_id = 999999, các trường khác hợp lệ",
         "level_id = 999999",
         "- API trả 422 với thông báo 'Không tồn tại' cho trường cấp bảo dưỡng\n- Không tạo bản ghi",
         "Rule exists:levels,id"),

        ("023", "Gửi ĐVT / ghi chú / hàng hóa không tồn tại qua API", "P1",
         "unit_id = 999999, note_maintenance_id = 999999, product_id = 999999",
         "1. Lần lượt gọi POST với từng giá trị sai ở trên, các trường khác hợp lệ",
         "3 lần gọi với id sai",
         "- Cả 3 lần đều trả 422 với thông báo 'Không tồn tại' ở đúng trường tương ứng\n- Không tạo bản ghi",
         "Rule exists cho unit_id, note_maintenance_ids.*, product_groups.*.product_id"),

        ("024", "Gửi trạng thái không hợp lệ khi cập nhật", "P1",
         "Có gói id = Y",
         "1. Gọi POST /v1/customer-care/services/Y với status = 5, các trường khác hợp lệ",
         "status = 5",
         "- API trả 422 với 'Trạng thái không hợp lệ'\n- Bản ghi không đổi",
         "Rule status required in:0,1 khi update"),

        ("025", "Chống double-click nút Lưu", "P0",
         "Đang ở màn Thêm với dữ liệu hợp lệ (mã 'GKT-DBL')",
         "1. Bấm Lưu liên tiếp 3 lần thật nhanh\n2. Kiểm tra danh sách và bảng services",
         "3 lần bấm < 1 giây",
         "- Nút Lưu bị vô hiệu và chuyển sang biểu tượng đang xử lý sau lần bấm đầu\n"
         "- Chỉ tạo ĐÚNG 1 gói mã GKT-DBL\n- Không có bản ghi trùng, không có bảng con nhân đôi",
         "isSubmitSave chặn gọi lại + :interactable trên nút"),

        ("026", "Nhập ký tự đặc biệt và tiếng Việt có dấu", "P1",
         "Đang ở màn Thêm",
         "1. Tên = 'Bảo dưỡng máy xúc & thiết bị (cấp 1) #A'\n2. Nhập đủ trường bắt buộc, bấm Lưu\n3. Quan sát danh sách và bản in",
         "Chuỗi có dấu, &, #, ()",
         "- Lưu thành công\n- Danh sách hiển thị đúng nguyên văn, không bị mã hoá HTML (không thấy &amp;)\n"
         "- Bản in hiển thị tên IN HOA đúng cả chữ có dấu",
         "mb_strtoupper UTF-8 cho tên khi in"),

        ("027", "Nhập thẻ script vào ô nhập (XSS)", "P0",
         "Đang ở màn Thêm",
         "1. Tên = '<script>alert(1)</script>', Mã = 'GKT-XSS'\n2. Bấm Lưu\n3. Quan sát danh sách\n4. Rê chuột lên tên gói (tooltip giá)",
         "name = '<script>alert(1)</script>'",
         "- KHÔNG xuất hiện hộp thoại alert ở cả danh sách lẫn tooltip\n"
         "- Chuỗi hiển thị nguyên văn dạng text\n- Không lỗi JS console",
         "Tooltip build bằng HTML nên phải escape — kiểm tra kỹ chỗ này"),

        ("028", "Mất kết nối mạng khi đang lưu", "P2",
         "Đang ở màn Thêm với dữ liệu hợp lệ",
         "1. Ngắt mạng (DevTools → Offline)\n2. Bấm Lưu\n3. Quan sát",
         "Network: Offline",
         "- Hiện toast lỗi\n- Thanh loading tắt, nút Lưu bật lại để thử lại\n"
         "- Trang KHÔNG điều hướng, toàn bộ dữ liệu đã nhập (kể cả ma trận, hàng hóa, file) còn nguyên",
         "finally: isSubmitSave = false + $loading.finish(); không goBack khi lỗi"),

        ("029", "Mở màn Sửa gói không tồn tại", "P1",
         "Không tồn tại gói id = 999999",
         "1. Truy cập /customer-care/services/999999/edit",
         "id không tồn tại",
         "- Hiện toast lỗi tải dữ liệu\n- Tự quay về màn danh sách\n- Không lỗi trắng màn hình",
         "loadService() catch lỗi → toast + goBack()"),

        ("030", "Ma trận nhiều cột nhiều hàng", "P1",
         "Đang ở màn Thêm",
         "1. Thêm 5 cột cấp và 20 hàng nội dung kiểm tra, điền đủ dữ liệu\n2. Bấm Lưu\n3. Mở lại màn Sửa và đối chiếu",
         "5 cột × 20 hàng = 100 ô ghi chú",
         "- Lưu thành công trong thời gian chấp nhận được (không timeout)\n"
         "- Mở lại đủ 5 cột và 20 hàng, ghi chú từng ô ĐÚNG vị trí, không lệch cột\n"
         "- Bảng cuộn ngang được, không vỡ layout",
         "service_maintain_levels lưu theo `order` = thứ tự cột"),

        ("031", "Hàng hóa đã bị xóa mềm sau khi gắn vào gói", "P2",
         "Gói đã gắn hàng 'Bulong M10'; sau đó hàng này bị xóa mềm ở danh mục hàng hóa",
         "1. Xóa mềm hàng hóa đó\n2. Mở màn Sửa gói\n3. Quan sát khối 'Áp dụng cho hàng hóa'\n4. Bấm Lưu (không đổi gì)",
         "products.deleted_at khác null",
         "- Hàng 'Bulong M10' VẪN hiển thị trong bảng của gói (chủ đích, để không mất pivot khi lưu lại)\n"
         "- Sau khi Lưu: pivot của hàng đó vẫn còn\n- Popup chọn hàng hóa thì KHÔNG còn hàng này",
         "dataForEdit không lọc deleted_at; searchProducts thì có lọc"),
    ]),

    # ---------------------------------------------------------------- VII
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Gọi API khi chưa đăng nhập", "P0",
         "Chưa có token hoặc đã đăng xuất",
         "1. Gọi GET /api/v1/customer-care/services không gắn Authorization\n2. Gọi tiếp GET /v1/customer-care/services/{id}/print-data",
         "Không có Bearer token",
         "- Cả 2 request trả HTTP 401\n- Body không chứa dữ liệu gói bảo dưỡng nào",
         "Route group middleware auth:api"),

        ("002", "Truy cập URL màn hình khi chưa đăng nhập", "P0",
         "Trình duyệt đã đăng xuất",
         "1. Gõ URL /customer-care/services\n2. Gõ URL /customer-care/services/1/print",
         "—",
         "- Cả 2 đều chuyển về màn đăng nhập\n- Không hiển thị dữ liệu dù chỉ trong tích tắc",
         "Middleware auth phía FE"),

        ("003", "2 user khác công ty thấy CÙNG danh sách", "P0",
         "User A thuộc công ty 1, user B thuộc công ty 2; danh mục có 20 gói",
         "1. Đăng nhập user A, ghi lại tổng số và 5 mã đầu\n2. Đăng nhập user B, so sánh",
         "A: công ty 1; B: công ty 2",
         "- Cả 2 thấy CÙNG 20 gói và cùng thứ tự\n"
         "- Đây là hành vi đúng: danh mục dùng chung, KHÔNG lọc theo công ty (khác màn Danh mục tài khoản ngân hàng)\n"
         "- User B sửa được gói do user A tạo (nếu có quyền Sửa)",
         "ServiceService::index() không có điều kiện company_id"),

        ("004", "Ảnh letterhead khi in theo công ty người in", "P0",
         "User A thuộc công ty 1 (có letterhead), user B thuộc công ty 2 (letterhead khác); cùng in 1 gói",
         "1. User A in gói X, chụp lại phần đầu trang\n2. User B in cùng gói X, so sánh",
         "Cùng gói, khác người in",
         "- Ảnh letterhead khác nhau: mỗi người thấy letterhead của CÔNG TY MÌNH\n"
         "- Phần nội dung bảng kiểm tra bảo dưỡng giống hệt nhau\n"
         "- User không gắn công ty: bản in không có letterhead nhưng vẫn in được nội dung",
         "companyHeader() lấy theo công ty của user đang đăng nhập, không theo công ty quản lý gói"),

        ("005", "Xử lý khi API danh sách trả lỗi 500", "P1",
         "Giả lập GET /customer-care/services trả HTTP 500",
         "1. Giả lập 500\n2. Reload màn",
         "Response: 500",
         "- Hiện toast 'Lỗi khi tải dữ liệu'\n- Bảng rỗng, loading đã tắt\n- Panel lọc vẫn thao tác được để thử lại",
         "catch: chỉ bỏ qua toast khi status = 403"),

        ("006", "Xử lý khi API xuất Excel lỗi", "P2",
         "Giả lập GET /customer-care/services/export trả HTTP 500",
         "1. Giả lập lỗi\n2. Bấm 'Xuất excel'",
         "Response: 500",
         "- Hiện toast 'Lỗi khi xuất Excel'\n- Thanh loading tắt\n- Không tải về file rỗng/hỏng",
         "exportExcel() catch lỗi + finally tắt loading"),

        ("007", "Dữ liệu ghi từ HRM đọc được ở cổng ERP", "P1",
         "Có quyền truy cập cả HRM và ERP",
         "1. Tạo gói mới ở HRM với 2 cấp, 2 hàng, gợi ý hàng hoá\n2. Mở màn tương ứng ở ERP\n3. Đối chiếu ma trận và gợi ý hàng hoá",
         "Gói tạo từ HRM",
         "- ERP hiển thị đủ gói vừa tạo với đúng ma trận và gợi ý hàng hoá (không rỗng)\n"
         "- Sửa ở ERP rồi mở lại HRM cũng thấy dữ liệu mới\n- Không phát sinh bản ghi trùng",
         "2 cổng chạy song song trên cùng bảng; key_word lưu đúng shape [{'text': ...}] của ERP"),
    ]),

    # ---------------------------------------------------------------- VIII
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đầy đủ: tạo → sửa → in → sao chép → xóa", "P0",
         "User có đủ 3 quyền; chưa tồn tại mã 'E2E-GBD-01'",
         "1. Thêm mới: Tên 'Gói E2E 01', Mã 'e2e-gbd-01', Công ty = công ty mình, VAT 8, Hệ số giá bán gói 2; "
         "thêm 2 cột cấp, 2 hàng nội dung kiểm tra (chọn ghi chú đủ 4 ô), nhập định mức công 2 và 3; thêm 3 hàng hóa; đính 1 file PDF → Lưu\n"
         "2. Kiểm tra dòng mới trên danh sách và tooltip giá\n"
         "3. Bấm Sửa: thêm 1 hàng nội dung, đổi Ghi chú → Lưu\n"
         "4. Bấm In, kiểm tra bản xem trước\n"
         "5. Bấm Sao chép, đổi Tên/Mã thành 'Gói E2E 01 - sao chép' / 'E2E-GBD-01C' → Lưu\n"
         "6. Xóa bản sao (chưa gắn hàng hóa) và xóa gói gốc",
         "Mã 'E2E-GBD-01'",
         "- B1: tạo thành công, mã hiển thị IN HOA\n- B2: dòng đứng đầu danh sách, tooltip hiện giá 2 cấp đúng công thức\n"
         "- B3: cập nhật thành công, Người sửa/Ngày sửa cập nhật\n"
         "- B4: bản in đủ letterhead, tên gói in hoa, bảng 2 cột cấp và 3 hàng, có chú giải + bảng ký\n"
         "- B5: bản sao có đủ ma trận, hàng hóa và file của bản gốc\n"
         "- B6: bản sao (nếu chưa gắn hàng hóa) bị xóa hẳn; gói gốc ĐÃ gắn hàng hóa nên chuyển sang trạng thái 'Khóa' kèm thông báo tương ứng",
         "Bao trùm toàn bộ chức năng của màn"),

        ("002", "Luồng công thức giá đầu-cuối", "P0",
         "Công ty A (quản lý) đơn giá công 100.000; công ty B đơn giá công 150.000, Hệ số giá bán 1,2",
         "1. Thêm gói, chọn công ty quản lý = A, Hệ số giá bán gói = 2\n"
         "2. Thêm 1 cột cấp, 1 hàng; nhập Định mức công = 2, Hệ số công nghệ = 1,5\n"
         "3. Đọc lần lượt Giá vốn, Giá công thức, Giá bán cơ sở\n"
         "4. Ở khối Giá vốn theo công ty, nhập hệ số công ty B = 1,2\n"
         "5. Đọc dòng 'Giá bán theo công ty'\n6. Lưu rồi về danh sách rê chuột lên tên gói",
         "work_price A = 100000; quota 2; hệ số CN 1,5; hệ số gói 2; hệ số B 1,2",
         "- B3: Giá vốn = 300,000; Giá công thức = 600,000; Giá bán cơ sở tự điền 600,000\n"
         "- B4: khối Giá vốn theo công ty hiện giá vốn của B = 300,000 (2 × 150.000)\n"
         "- B5: Giá bán công ty A = 600,000; công ty B = 720,000\n"
         "- B6: tooltip danh sách hiện 400,000 cho cấp đó (không nhân hệ số công nghệ) → ghi nhận chênh lệch với Giá công thức để nghiệp vụ xác nhận",
         "Chuỗi công thức: giá vốn → giá công thức → giá bán cơ sở → giá bán theo công ty; tooltip dùng công thức khác"),

        ("003", "Luồng chọn hàng hóa quy mô lớn", "P1",
         "Nhóm 'Vòng bi' có 12 hàng, nhóm 'Bulong' có 8 hàng; có hàng thuộc cả 2 nhóm khác nhau",
         "1. Mở form Thêm, bấm 'Nhóm hàng hóa', tick cả 2 nhóm, bấm thêm\n"
         "2. Bấm 'Hàng hóa', lọc nâng cao theo Tính chất + Thương hiệu, tick thêm 3 hàng (trong đó 1 hàng đã có)\n"
         "3. Xóa 1 hàng bất kỳ, xóa cả nhóm 'Bulong'\n4. Lưu và mở lại màn Sửa",
         "20 hàng từ 2 nhóm + 3 hàng lẻ",
         "- B1: toast 'Đã thêm 20 hàng hóa từ 2 nhóm', bảng có 2 khối nhóm\n"
         "- B2: toast thêm 2 hàng + cảnh báo 1 hàng đã được chọn trước đó\n"
         "- B3: nhóm 'Bulong' và toàn bộ hàng của nó biến mất sau khi xác nhận\n"
         "- B4: mở lại đúng số hàng còn lại, nhóm hiển thị đúng nhóm lúc gắn",
         "Pivot service_has_products lưu group_id tại thời điểm gắn"),

        ("004", "Luồng khóa gói và ảnh hưởng tới nơi sử dụng", "P1",
         "Gói 'Gói E2E 01' đang Hoạt động và đang được dùng ở báo giá dịch vụ",
         "1. Bấm Xóa gói → xác nhận\n2. Quan sát danh sách và thông báo\n"
         "3. Lọc Trạng thái = 'Khóa' để tìm lại gói\n4. Mở Sửa, đổi trạng thái về 'Hoạt động' → Lưu",
         "Gói đang được sử dụng",
         "- B2: thông báo gói đã chuyển sang trạng thái Khóa, gói vẫn còn trong danh sách với pill 'Khóa'\n"
         "- B3: lọc ra được gói\n- B4: gói trở lại 'Hoạt động'\n"
         "- Dữ liệu báo giá dịch vụ liên quan không bị ảnh hưởng ở mọi bước",
         "Khóa là cách 'ngừng dùng' thay cho xóa khi gói đã phát sinh sử dụng"),

        ("005", "Luồng lọc — sắp xếp — phân trang — xuất Excel", "P1",
         "Danh mục có ≥ 30 gói, trong đó ≥ 15 gói tên chứa 'Bảo dưỡng'",
         "1. Nhập tìm nhanh 'Bảo dưỡng', bấm Tìm kiếm\n2. Đổi số dòng/trang = 10, sang trang 2\n"
         "3. Click sắp xếp cột 'Ngày tạo'\n4. Bấm 'Xuất excel'\n5. Bấm Đặt lại",
         "name = 'Bảo dưỡng'",
         "- B2: trang 2 chỉ chứa gói khớp 'Bảo dưỡng', STT bắt đầu từ 11, bộ lọc không bị mất\n"
         "- B3: sắp xếp áp dụng trên tập đã lọc, về trang 1, bộ lọc vẫn giữ\n"
         "- B4: file Excel chứa TOÀN BỘ gói (không theo bộ lọc) — ghi nhận đúng hành vi hiện tại\n"
         "- B5: mọi bộ lọc và sắp xếp về mặc định, danh sách hiển thị lại từ trang 1",
         "Bộ lọc/sắp xếp/phân trang cộng dồn đúng; export cố tình không áp bộ lọc"),

        ("006", "Luồng phối hợp với danh mục Cấp dịch vụ và Ghi chú kiểm tra", "P1",
         "Có quyền quản lý cả 3 màn: Cấp dịch vụ bảo dưỡng, Ghi chú kiểm tra bảo dưỡng, Gói bảo dưỡng",
         "1. Ở màn 'Cấp dịch vụ bảo dưỡng', thêm cấp mới 'Cấp E2E'\n"
         "2. Ở màn 'Danh mục ghi chú kiểm tra bảo dưỡng', thêm ghi chú ký hiệu 'E2E'\n"
         "3. Về màn gói bảo dưỡng, thêm gói mới, chọn cột cấp = 'Cấp E2E', ô ghi chú chọn 'E2E' → Lưu\n"
         "4. In gói vừa tạo\n5. Quay lại màn 'Cấp dịch vụ bảo dưỡng', thử xóa 'Cấp E2E'",
         "Cấp và ghi chú mới tạo",
         "- B3: 2 danh mục mới xuất hiện ngay trong dropdown của form gói bảo dưỡng, lưu thành công\n"
         "- B4: bản in có cột tên 'Cấp E2E' và ô ghi chú hiển thị ký hiệu 'E2E'; phần chú giải có dòng của ghi chú E2E\n"
         "- B5: hệ thống chặn xóa cấp đang được gói bảo dưỡng sử dụng, báo lý do rõ ràng",
         "Liên thông 3 danh mục CSKH; điều kiện chặn xóa của cấp dịch vụ kiểm đủ bảng tham chiếu"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT
    a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT
    a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT
    b.alignment = WRAP_TOP_LEFT
    b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, "Testcase _ %s" % FEATURE_NAME)
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT
    lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right")
    lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT
    vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN
    vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0


def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1


def write_tc(tc_id, function, priority, precondition, steps, test_data, expected,
             business_note, module=MODULE_NAME, group=""):
    global current_row, data_row_idx
    values = [
        module, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed",
        "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    newlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(210, max(longest // 4, newlines * 15 + 20)))
    current_row += 1
    data_row_idx += 1


total_tc = 0
p0_tc = 0

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc("TC-ROLE-%s" % suffix, func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

ROMANS = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row("%s. %s" % (roman, title))
    sec_idx = ROMANS.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = "TC_%02d.%03d" % (sec_idx, int(tc_num))
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

dv = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
    showDropDown=False,
)
dv.add("L18:N%d" % (current_row + 100))
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print("Generated: %s" % OUTPUT_FILE)
print("Rows: 1-10 description, 11-15 summary, 17 header, 18-%d data" % (current_row - 1))
print("Total TC: %d | P0: %d (%.0f%%)" % (total_tc, p0_tc, p0_tc * 100.0 / total_tc))
