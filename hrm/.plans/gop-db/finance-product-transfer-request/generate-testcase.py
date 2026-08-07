# -*- coding: utf-8 -*-
"""Generate testcase Excel cho man Phieu yeu cau chuyen hang (Modules/Finance - pages/finance/product-transfer-requests)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# CONFIG
# =========================================================================
OUTPUT_FILE = ".plans/gop-db/finance-product-transfer-request/testcase.xlsx"
SHEET_NAME = "PhieuYCChuyenHang"
FEATURE_NAME = "Phiếu yêu cầu chuyển hàng"
MODULE_NAME = "Phiếu YC chuyển hàng"

DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Quản lý phiếu yêu cầu chuyển hàng (phân hệ Tài chính → nhóm Xuất hàng → Phiếu yêu cầu chuyển hàng, URL /finance/product-transfer-requests).\n"
     "Màn port từ ERP `admin/warehouse/product_transfer_requests?type=all`, chạy song song với ERP trên CÙNG 3 bảng của DB gộp "
     "(`product_transfer_requests` + `product_transfer_request_products` + `product_transfer_request_product_details`) — KHÔNG đổi schema. HRM là bản thay thế lâu dài.\n"
     "Nghiệp vụ: người lập tạo phiếu (chọn hàng hóa → khai từng khách hàng cần hàng với số lượng + ngày cần) → Lưu nháp hoặc Gửi duyệt → "
     "Kế toán kho tiếp nhận (bấm Tổng hợp để lập phiếu đề nghị xuất hàng bên ERP) hoặc Không duyệt (trả về nháp kèm ghi chú).\n"
     "Phạm vi chức năng: danh sách + lọc, tạo/sửa/xóa nháp, gửi duyệt, xem chi tiết, Không duyệt, Tổng hợp (mở tab ERP), In (mẫu ERP 87), Xuất Excel, đính kèm PDF trên S3."),

    ("2. Đối tượng được tính / hiển thị",
     "► Phạm vi phiếu thấy được XÁC ĐỊNH THEO QUYỀN (port nguyên ERP `searchByFilter` nhánh type=all):\n"
     "   • Có `Xem yêu cầu chuyển hàng theo tổng công ty` HOẶC role Super admin (id 18) → thấy TOÀN BỘ phiếu mọi công ty.\n"
     "   • Có `Xem yêu cầu chuyển hàng theo công ty` → chỉ phiếu cùng `company_id` với mình.\n"
     "   • Có `Xem yêu cầu chuyển hàng theo phòng ban` → phiếu thuộc phòng ban mình quản lý (bảng `employee_manage_departments`) + phòng ban của chính mình, HOẶC phiếu do chính mình tạo.\n"
     "   • Có `Xem yêu cầu chuyển hàng theo bộ phận` → tương tự nhưng theo `employee_manage_parts` + bộ phận của mình, HOẶC phiếu do chính mình tạo.\n"
     "   • Không có quyền nào ở trên → CHỈ phiếu do chính mình tạo.\n"
     "► Ngoài ra LUÔN áp thêm 1 điều kiện cuối: phiếu nháp (status = 3) chỉ hiện với CHÍNH người tạo; mọi trạng thái khác thì hiện theo phạm vi ở trên.\n"
     "► Phiếu do cổng ERP tạo cũng hiển thị ở HRM và ngược lại (chung bảng)."),

    ("3. Đối tượng bị ẩn / không tính",
     "► Phiếu nháp (status = 3 'Đang tạo') của NGƯỜI KHÁC — không bao giờ hiện, kể cả với quyền xem theo tổng công ty.\n"
     "► Với user không gắn hồ sơ nhân viên (không có company_id): nhánh quyền theo công ty trả RỖNG hoàn toàn (không lộ phiếu có company_id NULL).\n"
     "► Màn HRM chỉ port 1 danh sách duy nhất (tương đương `type=all` bên ERP) — KHÔNG port 2 danh sách phụ `forAccounting` (type=approve) và `can_request` của ERP.\n"
     "► Popup chọn hàng hóa: không hiển thị hàng đã xóa mềm và hàng `status` = 0; không có đường 'Thêm hàng tạm' (phiếu chỉ nhận hàng thật trong bảng `products`).\n"
     "► Cột 'Được nhận' (allocated_qty) ở màn chi tiết chỉ hiển thị khi phiếu ở trạng thái 12 'Đã phân bổ'."),

    ("4. Bộ lọc thời gian áp dụng cho",
     "Áp dụng cho cột NGÀY TẠO phiếu (`product_transfer_requests.created_at`), qua 2 ô 'Ngày tạo từ' / 'Ngày tạo đến'.\n"
     "► 'Ngày tạo từ' → created_at >= ngày chọn (00:00 của ngày đó).\n"
     "► 'Ngày tạo đến' → created_at <= (ngày chọn + 1 ngày), tức BAO TRỌN cả ngày được chọn kể cả phiếu tạo lúc 23:59 (port nguyên công thức addDay của ERP).\n"
     "► KHÔNG lọc theo ngày tiếp nhận (approved_time) và ngày cần hàng (date_needed).\n"
     "► Khi xuất Excel có lọc ngày, file có thêm dòng 'Từ ngày ... đến ngày ...' ở đầu."),

    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► 3 bảng lồng nhau (dùng chung với ERP):\n"
     "   • `product_transfer_requests` — phiếu: code (PYCCH-xxxxx), status, note, comment (ghi chú duyệt), attachments (chuỗi URL S3 nối ', '), approver_id, approved_time, company_id/department_id/part_id, created_by.\n"
     "   • `product_transfer_request_products` — DÒNG HÀNG HÓA: product_id + denormalize product_name/code/model_name/brand_name/unit_name/price, qty = TỔNG số lượng các dòng con.\n"
     "   • `product_transfer_request_product_details` — DÒNG CON theo khách hàng: customer_id + customer_name, qty, date_needed, note, allocated_qty.\n"
     "► company_id / department_id / part_id được gán TỰ ĐỘNG theo hồ sơ người lập tại thời điểm tạo — người dùng không chọn.\n"
     "► Mã phiếu sinh sau khi có id: `PYCCH-` + id đệm 5 chữ số (cùng công thức với ERP nên 2 cổng không đụng mã nhau).\n"
     "► Xóa phiếu → xóa cascade dòng hàng rồi dòng con (xóa cứng, không SoftDeletes)."),

    ("6. Quy tắc cộng dồn / deduplicate",
     "► Không cộng dồn giữa các phiếu. Mỗi dòng danh sách = 1 phiếu, phân trang server-side (mặc định 10 dòng/trang).\n"
     "► Trong 1 phiếu: KHÔNG cho phép trùng hàng hóa (cùng `product_id` xuất hiện 2 dòng) — chặn ở cả popup chọn hàng và ở server.\n"
     "► Số lượng của dòng hàng = TỔNG số lượng của toàn bộ dòng con khách hàng thuộc dòng hàng đó (tự tính khi lưu, không nhập tay).\n"
     "► Khi lưu: toàn bộ dòng hàng + dòng con CŨ bị XÓA rồi tạo lại từ dữ liệu gửi lên (không so sánh từng dòng) — port nguyên ERP.\n"
     "► File đính kèm: chỉ nối thêm khi lưu; muốn bỏ file cũ phải bấm nút xóa riêng từng file (gọi API xóa file).\n"
     "► STT hiển thị = (trang − 1) × số dòng/trang + vị trí + 1."),

    ("7. Phân quyền cấp",
     "Route KHÔNG gắn middleware checkPermission (giống ERP) — quyền chặn ở tầng nghiệp vụ:\n"
     "• `Xem yêu cầu chuyển hàng theo tổng công ty` — thấy toàn bộ phiếu; bật bộ lọc Công ty + Phòng ban trên giao diện.\n"
     "• `Xem yêu cầu chuyển hàng theo công ty` — thấy phiếu cùng công ty; bật bộ lọc Phòng ban.\n"
     "• `Xem yêu cầu chuyển hàng theo phòng ban` — thấy phiếu phòng ban mình quản lý + phiếu của mình; bật bộ lọc Phòng ban.\n"
     "• `Xem yêu cầu chuyển hàng theo bộ phận` — thấy phiếu bộ phận mình quản lý + phiếu của mình.\n"
     "• `Kế toán kho` — được XEM mọi phiếu cùng công ty, được 'Không duyệt' và 'Tổng hợp' phiếu Chờ duyệt cùng công ty.\n"
     "• Role `Super admin` (id 18) — xem toàn bộ; riêng 'Không duyệt'/'Tổng hợp' VẪN phải cùng công ty với phiếu (khớp đúng ERP).\n"
     "Điều kiện theo trạng thái/người lập (không phải permission):\n"
     "• Sửa / Xóa: chỉ phiếu status = 3 'Đang tạo' VÀ do CHÍNH người đăng nhập lập.\n"
     "• Không duyệt / Tổng hợp: chỉ phiếu status = 2 'Chờ duyệt' VÀ cùng công ty VÀ có quyền Kế toán kho (hoặc Super admin).\n"
     "• Xem chi tiết / In: người tạo, hoặc Kế toán kho cùng công ty, hoặc Super admin.\n"
     "Ghi chú kỹ thuật: kiểm tra quyền đọc thẳng bảng phân quyền theo TÊN quyền (không qua thư viện) vì trên DB gộp cùng 1 tên quyền tồn tại 2 bản ERP/HRM và role gán từ ERP có kiểu dữ liệu khác."),

    ("8. Cách tính các ô thống kê",
     "Màn không có thẻ thống kê. Các ô số/tính toán:\n"
     "► Ô 'STT' = (trang − 1) × số dòng/trang + vị trí + 1.\n"
     "► Ô 'SL' của dòng hàng (form + chi tiết + bản in) = TỔNG số lượng các dòng con khách hàng của dòng hàng đó.\n"
     "► Ô 'Giá' hiển thị trên form = giá bán lẻ theo ĐVT đã chọn, đã áp hệ số công ty: nếu có hệ số ≠ 1 thì = làm tròn(giá × hệ số ÷ 1000) × 1000.\n"
     "► LƯU Ý: giá LƯU vào DB là giá bán lẻ THÔ theo ĐVT (KHÔNG áp hệ số công ty) — đúng nguyên bản ERP; con số trên form và con số trong DB có thể lệch nhau khi công ty có hệ số ≠ 1.\n"
     "► Cột tồn kho ('Xem tồn theo kho'): 'Tồn kho' = tổng tồn kế toán trong phạm vi kho đang xem TRỪ lượng đang chờ xuất; "
     "'Khả dụng' = max(0, min(khả dụng toàn công ty, tồn trong kho)); còn có 'Giữ hàng' và 'Hàng gửi' của chính nhân viên.\n"
     "► Dòng 'Hiển thị x–y / z' dưới bảng: z = tổng phiếu khớp bộ lọc trong phạm vi quyền."),

    ("9. Ghi chú đọc bảng",
     "► Danh sách 8 cột; cột STT và Mã yêu cầu được ghim khi cuộn ngang. KHÔNG cột nào sắp xếp được — thứ tự luôn là phiếu mới tạo nhất lên đầu.\n"
     "► Mã yêu cầu là link mở màn chi tiết phiếu.\n"
     "► 13 trạng thái: 1 Đã tiếp nhận · 2 Chờ duyệt · 3 Đang tạo · 4 Đang đề nghị · 5 Đang xuất kho · 6 Đã xuất kho · 7 Đang vận chuyển · "
     "8 và 9 Đang nhập kho · 10 Đã nhập kho · 11 Đã nhập hàng · 12 Đã phân bổ · 13 Đã hủy. HRM chỉ ghi được 2 và 3; các trạng thái còn lại do chuỗi nghiệp vụ kho bên ERP đẩy sang (13 hiện chưa nơi nào set).\n"
     "► Cột 'Hành động': nút Sửa và Xóa luôn hiển thị nhưng bị mờ kèm tooltip khi không đủ điều kiện; nút Tổng hợp chỉ hiện khi được duyệt; nút In luôn hiện, mở tab mới.\n"
     "► Giá trị trống hiển thị '—'.\n"
     "► Bộ lọc được ghi nhớ 10 phút khi rời trang rồi quay lại (key `finance_product_transfer_requests`)."),
]

HAS_ROLE_SECTION = True
ROLE_TCS = [
    ("01", "Không có quyền xem theo cấp — chỉ thấy phiếu của mình", "P0",
     "User A không có 4 quyền 'Xem yêu cầu chuyển hàng theo ...', không phải Kế toán kho, không Super admin; "
     "A đã tạo 3 phiếu; công ty của A còn 10 phiếu do người khác tạo",
     "1. Đăng nhập user A\n2. Vào Tài chính → Xuất hàng → 'Phiếu yêu cầu chuyển hàng'\n3. Đối chiếu danh sách với DB",
     "User A: không quyền xem theo cấp",
     "- Chỉ hiển thị đúng 3 phiếu do A tạo, tổng = 3\n- Không thấy phiếu của người khác dù cùng công ty\n"
     "- Bộ lọc nâng cao KHÔNG hiện ô Công ty và Phòng ban",
     "searchByFilter nhánh else: where created_by = mình; meta is_big_boss/is_boss/is_manager = false"),

    ("02", "Quyền 'Xem yêu cầu chuyển hàng theo công ty'", "P0",
     "User B có quyền theo công ty, thuộc công ty 1; công ty 1 có 12 phiếu (không tính nháp người khác); công ty 2 có 7 phiếu",
     "1. Đăng nhập user B\n2. Vào màn danh sách, ghi lại tổng số\n3. Kiểm tra bộ lọc nâng cao",
     "User B: quyền theo công ty, company_id = 1",
     "- Thấy đúng 12 phiếu của công ty 1\n- KHÔNG thấy phiếu nào của công ty 2\n"
     "- Bộ lọc nâng cao hiện ô 'Phòng ban', KHÔNG hiện ô 'Công ty'",
     "where company_id = công ty user; meta is_boss = true"),

    ("03", "Quyền 'Xem yêu cầu chuyển hàng theo tổng công ty'", "P0",
     "User C có quyền theo tổng công ty; hệ thống có 19 phiếu ở 2 công ty khác nhau (không tính nháp người khác)",
     "1. Đăng nhập user C\n2. Vào màn danh sách\n3. Kiểm tra bộ lọc nâng cao",
     "User C: quyền tổng công ty",
     "- Thấy đủ 19 phiếu của CẢ 2 công ty\n- Bộ lọc nâng cao hiện CẢ ô 'Công ty' và ô 'Phòng ban'\n"
     "- Lọc theo Công ty = công ty 2 thì chỉ còn 7 phiếu",
     "Nhánh tổng công ty: không filter thêm; meta is_big_boss = true"),

    ("04", "Quyền 'Xem yêu cầu chuyển hàng theo phòng ban'", "P0",
     "User D có quyền theo phòng ban; D thuộc phòng ban P1 và được gán quản lý thêm phòng ban P2; "
     "P1 có 4 phiếu, P2 có 3 phiếu, phòng ban P3 có 5 phiếu (trong đó 1 phiếu do chính D tạo)",
     "1. Đăng nhập user D\n2. Vào màn danh sách\n3. Đối chiếu từng phiếu với phòng ban",
     "User D: quyền theo phòng ban, quản lý P2",
     "- Thấy 4 phiếu P1 + 3 phiếu P2 + 1 phiếu của chính D ở P3 = 8 phiếu\n"
     "- KHÔNG thấy 4 phiếu còn lại của P3\n- Bộ lọc nâng cao hiện ô 'Phòng ban'",
     "whereIn(department_id, [phòng ban quản lý + phòng ban của mình]) OR created_by = mình"),

    ("05", "Quyền 'Xem yêu cầu chuyển hàng theo bộ phận'", "P0",
     "User E có quyền theo bộ phận; E thuộc bộ phận BP1 và quản lý thêm BP2; BP1 có 3 phiếu, BP2 có 2 phiếu, BP3 có 4 phiếu (1 phiếu do E tạo)",
     "1. Đăng nhập user E\n2. Vào màn danh sách, đối chiếu từng phiếu theo bộ phận",
     "User E: quyền theo bộ phận, quản lý BP2",
     "- Thấy 3 + 2 + 1 = 6 phiếu\n- KHÔNG thấy 3 phiếu còn lại của BP3\n"
     "- Ghi nhận: quyền 'theo bộ phận' có thể chưa được khai trong seeder — nếu không gán được quyền này thì ghi Pending kèm ghi chú",
     "whereIn(part_id, [bộ phận quản lý + bộ phận của mình]) OR created_by = mình"),

    ("06", "Phiếu nháp của người khác luôn bị ẩn", "P0",
     "User C có quyền xem theo TỔNG CÔNG TY; user A có 2 phiếu nháp (status = 3) chưa gửi",
     "1. Đăng nhập user A, xác nhận thấy 2 phiếu nháp của mình\n2. Đăng nhập user C, tìm 2 mã phiếu nháp đó\n"
     "3. Lọc Trạng thái = 'Đang tạo' bằng user C",
     "2 phiếu status = 3 của user A",
     "- User C KHÔNG thấy 2 phiếu nháp của A dù có quyền cao nhất\n- Lọc 'Đang tạo' bằng user C chỉ ra nháp của CHÍNH C\n"
     "- Gọi trực tiếp API chi tiết phiếu nháp của A bằng token C trả 403",
     "Điều kiện cuối luôn append: created_by = mình OR status != 3"),

    ("07", "Quyền 'Kế toán kho' — xem và duyệt phiếu cùng công ty", "P0",
     "User F có quyền `Kế toán kho`, thuộc công ty 1; có phiếu Chờ duyệt (status = 2) của công ty 1 và 1 phiếu Chờ duyệt của công ty 2",
     "1. Đăng nhập user F\n2. Mở chi tiết phiếu Chờ duyệt công ty 1\n3. Mở chi tiết phiếu Chờ duyệt công ty 2",
     "User F: Kế toán kho, company_id = 1",
     "- Phiếu công ty 1: mở được, hiện khối 'Ghi chú duyệt' nhập được, có nút 'Không duyệt' và 'Tổng hợp'\n"
     "- Phiếu công ty 2: KHÔNG mở được (403 'Bạn không có quyền xem phiếu này') — trừ khi F còn có quyền xem theo tổng công ty, khi đó mở được nhưng KHÔNG có nút Không duyệt/Tổng hợp",
     "canView/canApprove đều yêu cầu cùng company_id"),

    ("08", "Super admin — xem hết nhưng duyệt phải cùng công ty", "P0",
     "User G có role Super admin (id 18), thuộc công ty 1; có phiếu Chờ duyệt của công ty 4",
     "1. Đăng nhập user G\n2. Xác nhận thấy phiếu của mọi công ty ở danh sách\n"
     "3. Mở chi tiết phiếu Chờ duyệt công ty 4\n4. Đối chiếu với màn ERP của cùng phiếu đó",
     "Super admin công ty 1, phiếu công ty 4",
     "- Danh sách: thấy phiếu của mọi công ty\n- Mở chi tiết được (canView cho Super admin)\n"
     "- KHÔNG hiện nút 'Không duyệt' và 'Tổng hợp' (vì khác công ty) — GIỐNG HỆT bên ERP\n"
     "- Gọi API reject phiếu đó trả 403",
     "Super admin thay thế vế QUYỀN nhưng vẫn phải cùng company_id — khớp 1-1 ERP"),

    ("09", "User không gắn hồ sơ nhân viên", "P0",
     "User H có quyền 'Xem yêu cầu chuyển hàng theo công ty' nhưng không có hồ sơ nhân viên (company_id null); "
     "DB có phiếu với company_id = NULL",
     "1. Đăng nhập user H\n2. Vào màn danh sách",
     "User H: company_id = null",
     "- Danh sách RỖNG (không lộ phiếu company_id NULL của bất kỳ ai)\n- Không lỗi 500\n"
     "- Nếu H tự tạo phiếu thì sau đó chỉ thấy đúng phiếu của mình",
     "companyId null → whereRaw('1 = 0') thay vì where(company_id, null)"),

    ("10", "Sửa/xóa phiếu của người khác qua API", "P0",
     "Phiếu nháp id = X do user A tạo; user C có quyền xem tổng công ty",
     "1. Lấy token user C\n2. Gọi POST /v1/finance/product-transfer-requests/X (_method=PUT) với dữ liệu hợp lệ\n"
     "3. Gọi DELETE /v1/finance/product-transfer-requests/X",
     "Phiếu nháp của người khác",
     "- Cả 2 request trả 403 với message 'Chỉ sửa được phiếu Đang tạo do chính bạn lập' / 'Chỉ xóa được phiếu Đang tạo do chính bạn lập'\n"
     "- Phiếu X không thay đổi, các bảng con nguyên vẹn",
     "canEdit/canDelete = status 3 AND created_by = mình"),
]

SECTIONS = [
    # ---------------------------------------------------------------- I
    ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
        ("001", "Truy cập màn hình từ menu Tài chính", "P0",
         "User đã đăng nhập, có phạm vi xem ≥ 12 phiếu",
         "1. Đăng nhập\n2. Vào phân hệ Tài chính → nhóm 'Xuất hàng' → 'Phiếu yêu cầu chuyển hàng'\n3. Quan sát layout",
         "User bất kỳ",
         "- URL = /finance/product-transfer-requests\n- Tiêu đề trang = 'Phiếu yêu cầu chuyển hàng'\n"
         "- Hiển thị panel 'Bộ lọc phiếu yêu cầu chuyển hàng' (thu gọn) và bảng 'Phiếu yêu cầu chuyển hàng'\n"
         "- Cuối bảng có nút 'Thêm mới' và 'Xuất Excel'\n- Không lỗi console",
         "Menu không gate quyền — mọi user đăng nhập đều vào được"),

        ("002", "Kiểm tra đủ 8 cột của bảng", "P0",
         "Đang ở màn danh sách, có ≥ 1 phiếu",
         "1. Quan sát dòng tiêu đề bảng\n2. Đối chiếu tên và thứ tự cột với màn ERP",
         "—",
         "Đúng thứ tự 8 cột: STT · Mã yêu cầu · Người tạo · Ngày tạo · Người tiếp nhận · Ngày tiếp nhận · Trạng thái · Hành động\n"
         "- Bộ cột khớp với màn ERP (ERP cũng 7 cột dữ liệu + cột thao tác)\n- KHÔNG cột nào có icon sắp xếp",
         "Đối chiếu ERP all.blade.php — cùng bộ cột"),

        ("003", "Cột STT và Mã yêu cầu ghim khi cuộn ngang", "P1",
         "Thu hẹp cửa sổ để bảng cuộn ngang",
         "1. Thu nhỏ cửa sổ tới khi có thanh cuộn ngang\n2. Cuộn sang phải",
         "Độ rộng ~1024px",
         "- Cột STT và 'Mã yêu cầu' luôn dính bên trái\n- Nội dung không bị đè lên nhau",
         "tableColumns: sticky = true cho index và code"),

        ("004", "Mã yêu cầu là link mở chi tiết", "P0",
         "Có phiếu mã 'PYCCH-00123'",
         "1. Bấm vào mã phiếu trên danh sách\n2. Quan sát",
         "PYCCH-00123",
         "- Chuyển sang /finance/product-transfer-requests/<id>\n- Màn chi tiết hiển thị đúng phiếu vừa bấm\n"
         "- Mã hiển thị màu xanh in đậm ở danh sách",
         "nuxt-link tới màn chi tiết"),

        ("005", "Hiển thị khi không có phiếu nào phù hợp", "P1",
         "Lọc bằng mã chắc chắn không tồn tại",
         "1. Nhập ô tìm nhanh 'PYCCH-99999999'\n2. Bấm Tìm kiếm",
         "code không tồn tại",
         "- Hiển thị 'Không có dữ liệu phù hợp bộ lọc.'\n- Tổng = 0\n- Nút 'Thêm mới' và 'Xuất Excel' vẫn dùng được",
         "emptyText của V2BaseDataTable"),

        ("006", "Mở rộng / thu gọn bộ lọc nâng cao", "P1",
         "Đang ở màn danh sách, panel lọc thu gọn (mặc định)",
         "1. Bấm mở rộng bộ lọc\n2. Đếm và đọc tên các trường\n3. Thu gọn rồi mở lại",
         "—",
         "- Hiện các ô: Trạng thái, Tên/mã hàng hóa, Người tạo, Người tiếp nhận, Ngày tạo từ, Ngày tạo đến\n"
         "- Ô Công ty / Phòng ban chỉ hiện theo phạm vi quyền (xem TC-ROLE-02, 03, 04)\n"
         "- Giá trị đã nhập không mất khi thu gọn rồi mở lại",
         "V2BaseCompanyDepartmentFilter nhận permissions từ meta API"),
    ]),

    # ---------------------------------------------------------------- II
    ("II", "BỘ LỌC & TÌM KIẾM", [
        ("001", "Tìm nhanh theo mã yêu cầu", "P0",
         "Trong phạm vi xem có phiếu 'PYCCH-00123' và 'PYCCH-00456'",
         "1. Nhập ô tìm nhanh: '00123'\n2. Bấm Tìm kiếm",
         "code = '00123'",
         "- Chỉ hiển thị PYCCH-00123\n- Tổng = 1\n- Về trang 1",
         "code LIKE %code%"),

        ("002", "Gõ ô tìm nhanh KHÔNG tự gọi API", "P0",
         "Đang ở màn danh sách, mở tab Network",
         "1. Gõ '00123' vào ô tìm nhanh, không bấm Tìm kiếm\n2. Quan sát Network\n3. Bấm Tìm kiếm",
         "5 ký tự",
         "- Bước 1: KHÔNG phát sinh request\n- Bước 3: đúng 1 request với code = '00123'",
         "code nằm trong ignoredFields của deep watcher"),

        ("003", "Lọc theo Trạng thái", "P0",
         "Trong phạm vi xem có 5 phiếu 'Chờ duyệt', 3 phiếu 'Đã tiếp nhận', 2 phiếu 'Đang tạo' của chính mình",
         "1. Mở bộ lọc nâng cao\n2. Chọn Trạng thái = 'Chờ duyệt'\n3. Quan sát",
         "status = 2",
         "- Trả về đúng 5 phiếu\n- Mọi dòng hiện badge 'Chờ duyệt' (màu cảnh báo)\n"
         "- Danh sách tự tải lại ngay khi chọn (không cần bấm Tìm kiếm)",
         "Danh sách trạng thái lấy từ meta `statuses` của API"),

        ("004", "Dropdown Trạng thái có đủ 13 giá trị", "P1",
         "Đang ở màn danh sách",
         "1. Mở dropdown Trạng thái\n2. Đếm và đối chiếu tên với ERP",
         "—",
         "- Có đủ 13 lựa chọn: Đã tiếp nhận, Chờ duyệt, Đang tạo, Đang đề nghị, Đang xuất kho, Đã xuất kho, "
         "Đang vận chuyển, Đang nhập kho (2 mục trùng tên - id 8 và 9), Đã nhập kho, Đã nhập hàng, Đã phân bổ, Đã hủy\n"
         "- Tên khớp 1-1 với danh sách trạng thái bên ERP",
         "STATUSES port nguyên ERP (id 8 và 9 trùng tên 'Đang nhập kho' là đúng gốc)"),

        ("005", "Lọc theo Tên/mã hàng hóa", "P0",
         "Phiếu P1 có hàng 'Bulong M10' mã 'BL-M10'; phiếu P2 chỉ có 'Vòng bi 6203'",
         "1. Nhập ô 'Tên/mã hàng hóa' = 'Bulong'\n2. Quan sát\n3. Đổi thành 'BL-M10' và quan sát lại",
         "product_name = 'Bulong' rồi 'BL-M10'",
         "- Cả 2 lần đều trả về phiếu P1 (lọc theo CẢ tên và mã hàng)\n- Không trả về P2\n"
         "- Phiếu có nhiều hàng chỉ cần 1 hàng khớp là được trả về",
         "whereHas('products') theo code hoặc product_name"),

        ("006", "Lọc theo Người tạo", "P0",
         "User 'Nguyễn Văn A' tạo 4 phiếu; người khác tạo 6 phiếu, tất cả trong phạm vi xem",
         "1. Chọn Người tạo = 'Nguyễn Văn A'\n2. Quan sát",
         "created_by = id của A",
         "- Trả về đúng 4 phiếu\n- Cột 'Người tạo' mọi dòng đều là 'Nguyễn Văn A'",
         "where created_by = giá trị chọn"),

        ("007", "Lọc theo Người tiếp nhận", "P0",
         "Có 3 phiếu đã bị 'Không duyệt' bởi user F, 5 phiếu chưa ai xử lý",
         "1. Chọn Người tiếp nhận = user F\n2. Quan sát",
         "approver = id của F",
         "- Trả về đúng 3 phiếu\n- Cột 'Người tiếp nhận' của các dòng đều là user F, cột 'Ngày tiếp nhận' có giá trị\n"
         "- Phiếu chưa ai xử lý không xuất hiện",
         "where approver_id = giá trị chọn"),

        ("008", "Lọc theo khoảng ngày tạo", "P0",
         "Có phiếu tạo ngày 01/08/2026, 05/08/2026, 10/08/2026",
         "1. Chọn Ngày tạo từ = 05/08/2026, Ngày tạo đến = 10/08/2026\n2. Quan sát",
         "start_date = 2026-08-05, end_date = 2026-08-10",
         "- Trả về 2 phiếu (05/08 và 10/08)\n- KHÔNG trả về phiếu 01/08",
         "created_at >= start_date AND created_at <= end_date + 1 ngày"),

        ("009", "Ngày tạo đến bao trọn cả ngày được chọn", "P0",
         "Có phiếu tạo lúc 23:45 ngày 10/08/2026",
         "1. Chọn Ngày tạo đến = 10/08/2026 (bỏ trống ô 'từ')\n2. Tìm phiếu tạo lúc 23:45",
         "end_date = 2026-08-10, phiếu 10/08 23:45",
         "- Phiếu tạo lúc 23:45 ngày 10/08 VẪN nằm trong kết quả\n"
         "- Nếu bị loại thì là lỗi so sánh mốc cuối ngày → ghi Failed",
         "Port nguyên ERP: so <= (end_date + 1 ngày) để bao trọn cả ngày"),

        ("010", "Chỉ lọc ngày tạo từ", "P1",
         "Có phiếu tạo 01/08, 05/08, 10/08/2026",
         "1. Chọn Ngày tạo từ = 05/08/2026, để trống ô 'đến'\n2. Quan sát",
         "start_date = 2026-08-05",
         "- Trả về phiếu 05/08 và 10/08\n- Không trả về phiếu 01/08",
         "Chỉ áp điều kiện >= start_date"),

        ("011", "Lọc theo Công ty (chỉ với quyền tổng công ty)", "P0",
         "User C có quyền tổng công ty; công ty 1 có 12 phiếu, công ty 2 có 7 phiếu",
         "1. Đăng nhập user C, mở bộ lọc nâng cao\n2. Chọn Công ty = công ty 2\n3. Quan sát",
         "company_id = 2",
         "- Ô 'Công ty' có hiển thị (chỉ với quyền tổng công ty)\n- Trả về đúng 7 phiếu của công ty 2\n"
         "- Sau khi chọn công ty, dropdown Phòng ban chỉ còn phòng ban thuộc công ty 2",
         "where company_id = giá trị chọn; V2BaseCompanyDepartmentFilter tự lọc cascade"),

        ("012", "Lọc theo Phòng ban", "P0",
         "User B có quyền theo công ty; phòng ban P1 có 4 phiếu, P2 có 8 phiếu, cùng công ty của B",
         "1. Mở bộ lọc nâng cao\n2. Chọn Phòng ban = P1\n3. Quan sát",
         "department_id = P1",
         "- Trả về đúng 4 phiếu của phòng ban P1\n- Kết hợp đúng với phạm vi quyền (không lộ phiếu công ty khác)",
         "where department_id = giá trị chọn, cộng dồn với nhánh quyền"),

        ("013", "Kết hợp nhiều bộ lọc (AND)", "P0",
         "Có phiếu Chờ duyệt của A tạo 05/08; phiếu Chờ duyệt của B tạo 05/08; phiếu Đang tạo của A",
         "1. Chọn Trạng thái = 'Chờ duyệt'\n2. Chọn Người tạo = A\n3. Chọn Ngày tạo từ = 05/08/2026\n4. Quan sát",
         "3 điều kiện",
         "- Chỉ trả về phiếu Chờ duyệt của A tạo ngày 05/08\n- Các điều kiện cộng dồn bằng AND",
         "Mọi filter nối AND, chồng lên nhánh quyền"),

        ("014", "Nút Đặt lại bộ lọc", "P0",
         "Đã nhập mã + trạng thái + người tạo + khoảng ngày, đang ở trang 3",
         "1. Bấm nút Đặt lại\n2. Quan sát ô lọc và danh sách",
         "—",
         "- Mọi ô lọc về rỗng (kể cả 2 ô ngày)\n- Danh sách tải lại đầy đủ từ trang 1\n- Chỉ 1 request phát sinh",
         "handleReset(): filters về initialStateForm, watcher gọi API"),

        ("015", "Ghi nhớ bộ lọc khi rời trang rồi quay lại", "P1",
         "Đang lọc Trạng thái = 'Chờ duyệt', Ngày tạo từ = 01/08/2026",
         "1. Chuyển sang màn khác\n2. Quay lại trong vòng 10 phút\n3. Quan sát ô lọc và kết quả",
         "Rời trang < 10 phút",
         "- Bộ lọc khôi phục đúng cả trạng thái và ngày\n- Trạng thái mở/thu gọn panel cũng khôi phục\n"
         "- Chỉ 1 request khi vào lại màn",
         "filterStateMixin key finance_product_transfer_requests, hạn 10 phút"),

        ("016", "Chống gọi API trùng lặp", "P1",
         "Đang ở màn danh sách, mở tab Network",
         "1. Bấm Tìm kiếm 3 lần liên tiếp thật nhanh, không đổi tham số\n2. Đếm request",
         "3 lần bấm < 1 giây",
         "- Chỉ phát sinh 1 request\n- Danh sách không nhấp nháy",
         "DedupeLoadMixin"),
    ]),

    # ---------------------------------------------------------------- III
    ("III", "STATS / THỐNG KÊ ĐẦU TRANG", [
        ("001", "Không áp dụng cho feature này", "P2",
         "Đang ở màn /finance/product-transfer-requests",
         "1. Quan sát vùng phía trên bảng dữ liệu",
         "—",
         "- Màn KHÔNG có thẻ thống kê (giống ERP)\n- Chỉ có panel lọc và bảng dữ liệu\n"
         "- Con số duy nhất ngoài bảng là tổng phiếu ở vùng phân trang",
         "Section giữ lại theo chuẩn tài liệu — ERP cũng không có stats"),
    ]),

    # ---------------------------------------------------------------- IV
    ("IV", "DANH SÁCH / GRID DỮ LIỆU", [
        ("001", "Thứ tự mặc định — phiếu mới nhất lên đầu", "P0",
         "Trong phạm vi xem có ≥ 5 phiếu tạo ở thời điểm khác nhau",
         "1. Vào màn danh sách\n2. Đối chiếu thứ tự với created_at trong DB\n3. So sánh với thứ tự bên ERP",
         "—",
         "- Phiếu tạo gần nhất đứng đầu, giảm dần theo ngày tạo\n- Thứ tự KHỚP với màn ERP cùng bộ lọc",
         "orderBy created_at desc — port nguyên ERP"),

        ("002", "Không có cột nào sắp xếp được", "P1",
         "Đang ở màn danh sách",
         "1. Click lần lượt tiêu đề cả 8 cột\n2. Quan sát thứ tự và tab Network",
         "—",
         "- Không cột nào đổi thứ tự\n- Không hiện icon sắp xếp\n- Không phát sinh request thừa",
         "tableColumns không khai sortable"),

        ("003", "Phân trang — chuyển trang", "P0",
         "Phạm vi xem có 25 phiếu, 10 dòng/trang",
         "1. Ở trang 1 ghi lại mã phiếu dòng đầu\n2. Sang trang 2\n3. Quan sát STT và dữ liệu",
         "total = 25, pageSize = 10",
         "- STT dòng đầu trang 2 = 11\n- Dữ liệu khác hoàn toàn trang 1\n- Bộ lọc hiện tại được giữ nguyên",
         "Phân trang không đụng filters → không reset về trang 1"),

        ("004", "Đổi số dòng trên mỗi trang", "P0",
         "Phạm vi xem có ≥ 60 phiếu; đang ở trang 3, 10 dòng/trang",
         "1. Đổi số dòng/trang sang 50\n2. Quan sát danh sách và số request",
         "pageSize: 10 → 50",
         "- Hiển thị tối đa 50 dòng, tự về trang 1, STT bắt đầu từ 1\n"
         "- CHỈ 1 request phát sinh (không gọi lại lần 2 do lệch kiểu dữ liệu)\n- Dòng 'Hiển thị x–y / z' đúng",
         "FE ép Number cho perPage/total/currentPage"),

        ("005", "Badge trạng thái đúng màu", "P0",
         "Có phiếu ở trạng thái 'Đã tiếp nhận' (1), 'Chờ duyệt' (2), 'Đang tạo' (3), 'Đã phân bổ' (12)",
         "1. Quan sát cột Trạng thái của 4 phiếu\n2. Đối chiếu màu với ERP",
         "4 trạng thái khác nhau",
         "- 'Đã tiếp nhận' và 'Đã phân bổ': badge màu tích cực (success)\n"
         "- 'Chờ duyệt' và 'Đang tạo': badge màu cảnh báo (danger)\n- Màu khớp với quy ước bên ERP",
         "status_type từ STATUSES port nguyên ERP"),

        ("006", "Nút Sửa/Xóa bật đúng theo điều kiện", "P0",
         "Cùng 1 màn có: phiếu nháp do chính mình tạo, phiếu nháp trạng thái khác của mình, phiếu Chờ duyệt của mình, phiếu của người khác",
         "1. Rê chuột lần lượt vào nút Sửa và Xóa của từng dòng",
         "4 tình huống",
         "- Phiếu nháp của mình: cả 2 nút BẬT\n"
         "- Phiếu Chờ duyệt của mình: cả 2 nút MỜ, tooltip 'Bạn không có quyền sửa/xóa phiếu này'\n"
         "- Phiếu của người khác (mọi trạng thái): cả 2 nút MỜ\n- Nút In luôn bật ở mọi dòng",
         "is_can_edit/is_can_delete = status 3 AND created_by = mình"),

        ("007", "Nút Tổng hợp chỉ hiện khi được duyệt", "P0",
         "User F (Kế toán kho công ty 1) xem danh sách có: phiếu Chờ duyệt công ty 1, phiếu Đã tiếp nhận công ty 1",
         "1. Quan sát cột Hành động của 2 dòng",
         "User F: Kế toán kho",
         "- Phiếu Chờ duyệt: có nút 'Tổng hợp' (icon dấu tích kép)\n"
         "- Phiếu Đã tiếp nhận: KHÔNG có nút Tổng hợp\n"
         "- Với user thường: không dòng nào có nút Tổng hợp\n- Khớp đúng với ERP (ERP cũng chỉ hiện khi canApprove)",
         "is_can_approve = status 2 + Kế toán kho/Super admin + cùng company_id"),

        ("008", "Hiển thị người/ngày tiếp nhận", "P1",
         "Phiếu vừa bị 'Không duyệt' bởi user F ngày hôm nay; phiếu khác chưa ai xử lý",
         "1. Quan sát cột 'Người tiếp nhận' và 'Ngày tiếp nhận' của 2 phiếu",
         "1 phiếu có approver, 1 phiếu chưa",
         "- Phiếu đã xử lý: hiện tên user F và ngày hôm nay (dạng dd/mm/yyyy)\n"
         "- Phiếu chưa xử lý: cả 2 ô hiện '—'",
         "approver_name / approved_time format d/m/Y"),

        ("009", "Định dạng ngày tạo", "P1",
         "Phiếu tạo ngày 05/08/2026",
         "1. Quan sát cột 'Ngày tạo'",
         "created_at = 2026-08-05",
         "- Hiển thị '05/08/2026' (không kèm giờ)\n- Khớp định dạng bên ERP",
         "format d/m/Y"),

        ("010", "Đối chiếu danh sách 2 cổng ERP ↔ HRM", "P0",
         "Cùng 1 user có quyền trên cả ERP và HRM; không lọc gì",
         "1. Mở màn ERP `admin/warehouse/product_transfer_requests?type=all`, ghi lại tổng số và 10 mã đầu\n"
         "2. Mở màn HRM cùng user\n3. So sánh tổng số, thứ tự và trạng thái từng phiếu",
         "Cùng user, không lọc",
         "- Tổng số phiếu GIỐNG NHAU\n- 10 mã đầu giống nhau và cùng thứ tự\n- Trạng thái từng phiếu khớp\n"
         "- Nếu lệch: ghi rõ mã phiếu lệch và phía nào thừa/thiếu",
         "searchByFilter HRM port nguyên nhánh type=all của ERP"),
    ]),

    # ---------------------------------------------------------------- V
    ("V", "CHỨC NĂNG CHÍNH (CRUD / ACTION)", [
        # --- Form
        ("001", "Mở màn tạo phiếu mới", "P0",
         "User đã đăng nhập, đang ở màn danh sách",
         "1. Bấm 'Thêm mới'\n2. Quan sát toàn bộ trang",
         "—",
         "- Chuyển sang /finance/product-transfer-requests/create\n"
         "- Khối 'Thông tin chung': Ngày lập (hôm nay, chỉ đọc), Người lập (tên mình, chỉ đọc), Ghi chú (nhập được)\n"
         "- Khối danh sách hàng hóa rỗng với dòng 'Chưa có hàng hóa' + dropdown 'Xem tồn theo kho'\n"
         "- Khối 'File đính kèm (PDF)' có dấu bắt buộc\n- Cuối trang: nút 'Lưu', 'Lưu & Gửi duyệt', 'Hủy'",
         "Ngày lập/Người lập lấy từ user đăng nhập, không cho sửa"),

        ("002", "Thêm hàng hóa qua popup", "P0",
         "Đang ở màn tạo phiếu; có hàng 'Bulong M10' mã 'BL-M10'",
         "1. Bấm nút thêm hàng hóa\n2. Tìm 'Bulong', tick 2 hàng\n3. Bấm nút thêm\n4. Quan sát bảng ngoài",
         "2 hàng hóa",
         "- Popup chọn hàng mở ra (dùng chung popup của màn báo giá), chỉ có tab Hàng hóa, KHÔNG có đường 'Thêm hàng tạm'\n"
         "- 2 dòng hàng được thêm với tên, mã, ĐVT mặc định, giá niêm yết\n"
         "- Popup KHÔNG tự đóng, người dùng tự bấm Đóng\n- Mỗi dòng hàng tự có sẵn 1 dòng con khách hàng trống",
         "Popup dùng chung QuotationProductSearchModal (goods-only, hide-manual-create)"),

        ("003", "Chặn thêm trùng hàng hóa", "P0",
         "Phiếu đã có hàng 'Bulong M10'",
         "1. Mở popup, tick lại 'Bulong M10' và 1 hàng mới\n2. Bấm thêm\n3. Quan sát",
         "1 trùng + 1 mới",
         "- Chỉ thêm 1 dòng mới\n- Hiện cảnh báo 'Hàng hóa đã có trong phiếu: Bulong M10'\n"
         "- Bảng KHÔNG có 2 dòng cùng 1 hàng hóa",
         "FE chặn trùng + BE cũng chặn ('Hàng hóa bị trùng trong phiếu')"),

        ("004", "Đổi ĐVT của dòng hàng", "P0",
         "Hàng 'Bulong M10' có 2 ĐVT: 'Cái' (gốc) và 'Hộp' (hệ số 100), giá khác nhau",
         "1. Thêm hàng vào phiếu, quan sát ĐVT mặc định\n2. Mở dropdown ĐVT, đổi sang 'Hộp'\n3. Quan sát cột giá",
         "2 ĐVT",
         "- ĐVT mặc định là đơn vị gốc ('Cái')\n- Dropdown liệt kê đơn vị gốc trước rồi đến các đơn vị theo hệ số tăng dần\n"
         "- Đổi sang 'Hộp': giá niêm yết cập nhật theo ĐVT mới",
         "productUnits: sắp xếp is_base desc rồi unit_coefficient asc"),

        ("005", "Xem tồn kho theo kho / nhóm kho", "P0",
         "Phiếu có 2 dòng hàng; hệ thống có nhóm kho và kho lẻ",
         "1. Mở dropdown 'Xem tồn theo kho'\n2. Chọn 1 nhóm kho\n3. Quan sát cột tồn của 2 dòng hàng\n4. Đổi sang 1 kho cụ thể",
         "1 nhóm kho, 1 kho",
         "- Dropdown liệt kê nhóm kho, kho trong nhóm có tiền tố '|--', kho lẻ đứng riêng\n"
         "- Sau khi chọn: các cột tồn kho / khả dụng / giữ hàng / hàng gửi được nạp cho từng dòng hàng\n"
         "- Đổi kho: số liệu tồn cập nhật lại theo phạm vi kho mới\n- Chọn lại option rỗng 'Xem tồn': số tồn về 0, không gọi API",
         "stockOptions + stock — port rút gọn ProductStockService của ERP"),

        ("006", "Thêm dòng con khách hàng", "P0",
         "Phiếu có 1 dòng hàng 'Bulong M10'",
         "1. Bấm nút thêm dòng con của hàng đó\n2. Chọn Khách hàng qua popup, nhập SL = 10, Ngày cần = ngày mai, Ghi chú = 'gấp'\n"
         "3. Bấm thêm dòng con thứ 2, khai khách hàng khác SL = 5",
         "2 dòng con",
         "- Mỗi dòng con gồm: Khách hàng, SL, Ngày cần, Ghi chú và nút xóa dòng\n"
         "- Popup chọn khách hàng tự đóng sau khi chọn\n"
         "- Sau khi lưu, SL của dòng hàng = 15 (tổng 2 dòng con)",
         "qty dòng hàng = tổng qty details (tính ở server khi lưu)"),

        ("007", "Xóa dòng con khách hàng", "P0",
         "Dòng hàng có 2 dòng con",
         "1. Bấm nút xóa dòng con thứ 2\n2. Quan sát\n3. Thử xóa nốt dòng con cuối cùng",
         "2 → 1 → 0 dòng con",
         "- Xóa dòng con thứ 2: bảng còn 1 dòng con\n"
         "- Xóa dòng con CUỐI CÙNG: hệ thống cảnh báo không cho xóa (mỗi hàng hóa phải có ít nhất 1 khách hàng cần hàng)",
         "BE rule: products.*.details min 1"),

        ("008", "Xóa dòng hàng hóa", "P0",
         "Phiếu có 3 dòng hàng",
         "1. Bấm nút xóa dòng hàng thứ 2\n2. Đọc hộp xác nhận\n3. Xác nhận",
         "3 → 2 dòng hàng",
         "- Có hộp xác nhận trước khi xóa\n- Sau xác nhận: dòng hàng và toàn bộ dòng con của nó biến mất\n"
         "- STT các dòng còn lại đánh lại liên tục",
         "askRemoveProduct + BaseConfirmModal"),

        # --- Lưu
        ("009", "Lưu nháp phiếu mới", "P0",
         "Đang ở màn tạo, đã khai 1 hàng hóa với 1 dòng con hợp lệ, đã đính 1 file PDF",
         "1. Bấm nút 'Lưu'\n2. Quan sát thông báo và danh sách",
         "status = 3",
         "- Toast 'Yêu cầu của bạn đã được lưu. Bạn cần gửi để yêu cầu được xử lý' (khớp message ERP)\n"
         "- Quay về danh sách, phiếu mới đứng đầu với trạng thái 'Đang tạo'\n"
         "- Mã phiếu dạng PYCCH-xxxxx (5 chữ số)\n- Nút Sửa và Xóa của dòng này BẬT",
         "Mã sinh sau khi có id: PYCCH- + id đệm 5 số, cùng công thức ERP"),

        ("010", "Lưu & Gửi duyệt phiếu mới", "P0",
         "Đang ở màn tạo, dữ liệu hợp lệ; công ty có user F giữ quyền Kế toán kho",
         "1. Bấm 'Lưu & Gửi duyệt'\n2. Quan sát thông báo và danh sách\n3. Đăng nhập user F kiểm tra chuông thông báo",
         "status = 2",
         "- Toast 'Yêu cầu của bạn đã được gửi' (khớp message ERP)\n"
         "- Phiếu mới có trạng thái 'Chờ duyệt'; nút Sửa và Xóa của dòng này bị MỜ\n"
         "- User F nhận thông báo chuông: '<tên người tạo> vừa tạo yêu cầu chuyển hàng: PYCCH-xxxxx', bấm vào mở đúng màn chi tiết phiếu",
         "notifyAccountants: gửi cho MỌI nhân viên có quyền Kế toán kho CÙNG CÔNG TY với phiếu"),

        ("011", "Gửi duyệt phiếu nháp đã lưu", "P0",
         "Phiếu nháp 'PYCCH-00123' của chính mình",
         "1. Bấm Sửa phiếu đó\n2. Không đổi gì, bấm 'Lưu & Gửi duyệt'\n3. Quan sát",
         "status 3 → 2",
         "- Toast 'Yêu cầu của bạn đã được gửi'\n- Trạng thái đổi thành 'Chờ duyệt'\n"
         "- Kế toán kho cùng công ty nhận được thông báo\n- Nút Sửa/Xóa của phiếu chuyển sang mờ",
         "Update với status = 2 cũng bắn notify (giống ERP)"),

        # --- Sửa
        ("012", "Mở màn sửa và kiểm tra dữ liệu nạp lại", "P0",
         "Phiếu nháp của mình có 2 hàng hóa, tổng 3 dòng con, 2 file PDF",
         "1. Bấm Sửa\n2. Đối chiếu từng khối với dữ liệu đã lưu",
         "id phiếu nháp",
         "- Ghi chú, danh sách hàng hóa (ĐVT, giá), toàn bộ dòng con (khách hàng, SL, ngày cần, ghi chú) nạp đúng\n"
         "- 2 file PDF hiển thị và mở được, mỗi file có nút xóa\n- Khối File đính kèm KHÔNG còn dấu bắt buộc",
         "Update: attachments nullable (giữ file cũ)"),

        ("013", "Sửa phiếu nháp và lưu lại", "P0",
         "Phiếu nháp của mình có 1 hàng hóa",
         "1. Bấm Sửa\n2. Thêm 1 hàng hóa mới với 1 dòng con hợp lệ, sửa Ghi chú\n3. Bấm 'Lưu'\n4. Mở lại màn Sửa",
         "Thêm 1 hàng",
         "- Lưu thành công, giữ trạng thái 'Đang tạo'\n- Mở lại: có đủ 2 hàng hóa và ghi chú mới\n"
         "- Không phát sinh dòng hàng/dòng con trùng lặp trong DB",
         "syncProducts: xóa hết rồi tạo lại (port nguyên ERP)"),

        ("014", "Không sửa được phiếu đã gửi duyệt", "P0",
         "Phiếu 'Chờ duyệt' do chính mình tạo",
         "1. Quan sát nút Sửa trên dòng\n2. Truy cập thẳng URL /finance/product-transfer-requests/<id>/edit",
         "status = 2",
         "- Nút Sửa bị mờ với tooltip 'Bạn không có quyền sửa phiếu này'\n"
         "- Vào thẳng URL: hiện thông báo 'Chỉ sửa được phiếu Đang tạo do chính bạn lập' và không cho lưu",
         "canEdit = status 3 AND created_by = mình"),

        # --- File
        ("015", "Đính kèm file PDF khi tạo", "P0",
         "Đang ở màn tạo; có file yeucau.pdf",
         "1. Bấm 'Thêm file', chọn yeucau.pdf\n2. Quan sát\n3. Khai đủ dữ liệu và bấm Lưu\n4. Mở lại màn Sửa",
         "1 file PDF",
         "- Thẻ file hiện với icon PDF, tên file và nút x để bỏ\n- Lưu thành công\n"
         "- Mở lại màn Sửa: file nằm ở danh sách file đã lưu, bấm vào mở được (URL S3)",
         "Upload S3 prefix product_transfer_requests"),

        ("016", "Chọn file không phải PDF", "P0",
         "Có file anh.png",
         "1. Bấm 'Thêm file', chọn anh.png\n2. Quan sát",
         "File .png",
         "- Cảnh báo 'File \"anh.png\" không phải PDF'\n- File không được thêm\n"
         "- Nếu vẫn gửi lên được, server trả 422 'File đính kèm phải là file PDF'",
         "FE kiểm đuôi + BE rule mimes:pdf"),

        ("017", "Xóa file đã lưu khỏi phiếu", "P0",
         "Phiếu nháp của mình có 2 file PDF đã lưu",
         "1. Bấm Sửa\n2. Bấm nút x trên file thứ nhất, xác nhận\n3. Quan sát\n4. Kiểm tra file trên S3 và cột attachments trong DB",
         "2 file → 1 file",
         "- Toast 'Xóa file thành công', thẻ file biến mất ngay\n"
         "- Cột attachments trong DB chỉ còn URL file thứ 2\n"
         "- File trên S3 đã bị xóa THẬT (khác ERP: bên ERP xóa file không hoạt động vì xóa nhầm đường dẫn cục bộ)",
         "deleteFile: xóa object S3 qua CmcS3Helper — HRM sửa lỗi ERP"),

        ("018", "Xóa file của phiếu không được sửa", "P1",
         "Phiếu 'Chờ duyệt' có file đính kèm",
         "1. Lấy id phiếu và URL file\n2. Gọi API xóa file bằng token của chính người tạo",
         "status = 2",
         "- API trả 403 'Chỉ sửa được phiếu Đang tạo do chính bạn lập'\n- File vẫn còn trên phiếu và trên S3",
         "deleteFile yêu cầu canEdit()"),

        ("019", "Thêm file vào phiếu đã có file", "P1",
         "Phiếu nháp có 1 file PDF",
         "1. Bấm Sửa, thêm 1 file PDF mới\n2. Bấm Lưu\n3. Mở lại màn Sửa",
         "1 cũ + 1 mới",
         "- Sau khi lưu có ĐỦ 2 file (file cũ không bị ghi đè)\n"
         "- Chuỗi attachments trong DB đúng định dạng, KHÔNG có dấu phẩy thừa ở đầu/cuối (lỗi này có ở bản ERP)",
         "uploadAttachments nối chuỗi — HRM sửa lỗi nối phần tử null của ERP"),

        # --- Xóa
        ("020", "Xóa phiếu nháp của mình", "P0",
         "Phiếu nháp 'PYCCH-00123' của mình có 2 hàng hóa và 3 dòng con",
         "1. Bấm nút Xóa trên dòng\n2. Đọc hộp xác nhận\n3. Bấm 'Xóa'\n4. Kiểm tra 3 bảng trong DB",
         "Phiếu nháp của mình",
         "- Hộp xác nhận: \"Bạn có chắc muốn xóa phiếu yêu cầu chuyển hàng 'PYCCH-00123'?\"\n"
         "- Toast 'Xóa thành công', dòng biến mất\n"
         "- CẢ 3 bảng đều sạch: phiếu, 2 dòng hàng, 3 dòng con — không còn bản ghi mồ côi",
         "boot deleting cascade: details → products → phiếu"),

        ("021", "Hủy thao tác xóa", "P1",
         "Phiếu nháp của mình",
         "1. Bấm Xóa\n2. Bấm 'Hủy'",
         "—",
         "- Hộp xác nhận đóng, phiếu vẫn còn\n- Không phát sinh request DELETE",
         "Chỉ gọi API khi bấm nút xác nhận"),

        # --- Chi tiết
        ("022", "Xem chi tiết phiếu", "P0",
         "Phiếu 'PYCCH-00123' có 2 hàng hóa, 3 dòng con, 1 file, ghi chú",
         "1. Bấm vào mã phiếu\n2. Quan sát toàn bộ màn chi tiết",
         "id phiếu",
         "- Hiện mã phiếu + badge trạng thái ở đầu trang\n"
         "- Khối 'Thông tin chung': người tạo, ngày tạo, người tiếp nhận, ngày tiếp nhận, ghi chú\n"
         "- Khối 'File đính kèm': mở được file\n"
         "- Khối 'Danh sách hàng hóa': mỗi hàng gộp ô với các dòng con (khách hàng, SL, ngày cần, ghi chú)\n"
         "- Cuối trang: nút In, Quay lại; nút Sửa nếu được phép",
         "Đối chiếu ERP show.blade.php — cùng bố cục"),

        ("023", "Cột 'Được nhận' chỉ hiện khi Đã phân bổ", "P1",
         "Phiếu A ở trạng thái 'Đã phân bổ' (12); phiếu B ở trạng thái 'Chờ duyệt'",
         "1. Mở chi tiết phiếu A, quan sát bảng hàng hóa\n2. Mở chi tiết phiếu B, so sánh",
         "status 12 vs status 2",
         "- Phiếu A: bảng có thêm cột 'Được nhận' hiển thị số lượng đã phân bổ\n"
         "- Phiếu B: KHÔNG có cột 'Được nhận'\n- Khớp đúng hành vi ERP",
         "Mirror ERP show.blade: cột allocated_qty chỉ khi status = 12"),

        ("024", "Xem chi tiết phiếu không thuộc phạm vi", "P0",
         "Phiếu id = X thuộc công ty khác, người dùng không phải người tạo, không Kế toán kho cùng công ty, không Super admin",
         "1. Truy cập thẳng /finance/product-transfer-requests/X",
         "Phiếu ngoài phạm vi",
         "- Hiện thông báo không có quyền xem (API trả 403 'Bạn không có quyền xem phiếu này')\n"
         "- Không lộ bất kỳ thông tin nào của phiếu\n- Không lỗi trắng màn hình",
         "canView(): người tạo / Kế toán kho cùng công ty / Super admin"),

        ("025", "Xem chi tiết phiếu không tồn tại", "P2",
         "Không tồn tại phiếu id = 999999",
         "1. Truy cập /finance/product-transfer-requests/999999",
         "id không tồn tại",
         "- Hiện thông báo không tìm thấy (API trả 404)\n- Không lỗi trắng màn hình",
         "findOrFail → Handler chung trả 404"),

        # --- Reject
        ("026", "Không duyệt phiếu Chờ duyệt", "P0",
         "User F (Kế toán kho công ty 1) mở chi tiết phiếu Chờ duyệt của công ty 1 do user A tạo",
         "1. Nhập 'Ghi chú duyệt' = 'Thiếu chứng từ, đề nghị bổ sung'\n2. Bấm 'Không duyệt'\n3. Xác nhận\n"
         "4. Quan sát phiếu\n5. Đăng nhập user A kiểm tra chuông",
         "comment có nội dung",
         "- Hộp xác nhận trước khi thực hiện\n- Toast thành công, trang tải lại\n"
         "- Trạng thái phiếu chuyển về 'Đang tạo' (3); Người tiếp nhận = user F, Ngày tiếp nhận = hôm nay; ghi chú duyệt được lưu\n"
         "- User A nhận thông báo: '<tên F> vừa từ chối yêu cầu chuyển hàng: PYCCH-xxxxx'\n"
         "- User A mở lại phiếu thì SỬA được (đã về nháp)",
         "reject: status = 3 + comment + approver_id + approved_time; notify người tạo"),

        ("027", "Không duyệt khi bỏ trống ghi chú", "P0",
         "User F đang ở màn chi tiết phiếu Chờ duyệt, chưa nhập ghi chú",
         "1. Bấm 'Không duyệt' ngay",
         "comment rỗng",
         "- Hiện lỗi inline dưới ô ghi chú: 'Vui lòng nhập ghi chú duyệt'\n"
         "- KHÔNG mở hộp xác nhận, KHÔNG gửi request\n- Trạng thái phiếu không đổi",
         "FE validate trước; BE cũng rule comment required"),

        ("028", "Ghi chú duyệt hiển thị cho người không có quyền duyệt", "P1",
         "Phiếu đã bị 'Không duyệt' với ghi chú 'Thiếu chứng từ'; đăng nhập bằng người tạo",
         "1. Mở chi tiết phiếu\n2. Quan sát khối ghi chú duyệt",
         "Người tạo, không phải Kế toán kho",
         "- Hiện khối 'Ghi chú duyệt' ở dạng CHỈ ĐỌC với nội dung 'Thiếu chứng từ'\n"
         "- Không có ô nhập, không có nút 'Không duyệt'",
         "v-else-if data.comment → readonly"),

        ("029", "Không duyệt phiếu không ở trạng thái Chờ duyệt", "P0",
         "Phiếu ở trạng thái 'Đã tiếp nhận' (1); user F là Kế toán kho cùng công ty",
         "1. Mở chi tiết phiếu\n2. Gọi trực tiếp API reject với comment hợp lệ",
         "status = 1",
         "- Màn chi tiết KHÔNG hiện ô ghi chú duyệt và nút 'Không duyệt'\n"
         "- Gọi API trả 403 'Bạn không có quyền thực hiện thao tác này'\n- Phiếu không đổi",
         "canApprove yêu cầu status = 2"),

        ("030", "Nút Tổng hợp mở màn ERP", "P0",
         "User F là Kế toán kho, đang xem phiếu Chờ duyệt cùng công ty",
         "1. Bấm nút 'Tổng hợp' (ở danh sách hoặc màn chi tiết)\n2. Quan sát tab mới",
         "id phiếu",
         "- Mở TAB MỚI tới màn tạo phiếu đề nghị xuất hàng bên ERP, kèm tham số id phiếu chuyển hàng\n"
         "- Màn ERP nạp sẵn dữ liệu hàng hóa của phiếu\n- Màn HRM không đổi trạng thái gì",
         "Màn Yêu cầu xuất hàng chưa port sang HRM — mở tab ERP theo thiết kế"),

        # --- In
        ("031", "In phiếu yêu cầu chuyển hàng", "P0",
         "Phiếu 'PYCCH-00123' có 2 hàng hóa, hàng thứ nhất có 2 khách hàng",
         "1. Bấm nút In trên dòng (hoặc nút In ở màn chi tiết)\n2. Quan sát tab mới",
         "id phiếu",
         "- Mở TAB MỚI tới /finance/product-transfer-requests/<id>/print\n"
         "- Bản xem trước có: logo + tên + địa chỉ công ty CỦA NGƯỜI TẠO, số phiếu, ngày yêu cầu, người yêu cầu, ghi chú\n"
         "- Bảng chi tiết: STT / Hàng hóa (kèm 'Mã: ...') / ĐVT / SL gộp ô theo số khách hàng, rồi 4 cột Khách hàng / SL / Ngày cần / Ghi chú\n"
         "- Nội dung khớp với bản in cùng phiếu bên ERP",
         "Dùng CHUNG mẫu in ERP id 87 nên 2 cổng ra cùng biểu mẫu"),

        ("032", "Bấm nút In mở hộp thoại in", "P0",
         "Đang ở màn xem trước bản in",
         "1. Bấm nút 'In'\n2. Quan sát hộp thoại in",
         "—",
         "- Mở hộp thoại in\n- Bản in KHÔNG có nút 'In', bảng đủ viền, font Times New Roman\n"
         "- Ô gộp (rowspan) của hàng hóa không bị cắt ngang khi sang trang mới\n- Không tràn lề phải",
         "Nhóm mỗi hàng hóa vào 1 khối riêng để tránh vỡ ô gộp khi in nhiều trang"),

        ("033", "In phiếu khi ghi chú trống", "P2",
         "Phiếu không nhập ghi chú",
         "1. In phiếu đó\n2. Quan sát ô Ghi chú trên bản in",
         "note = null",
         "- Ô Ghi chú hiển thị '_____' (đúng như ERP)\n- Không hiển thị chữ 'null'",
         "GHI_CHU = note hoặc '_____' — port nguyên ERP"),

        ("034", "In phiếu có dòng con thiếu ngày cần", "P2",
         "Phiếu có dòng con với date_needed = NULL (dữ liệu cũ từ ERP)",
         "1. In phiếu đó\n2. Quan sát cột 'Ngày cần' của dòng đó\n3. So sánh với bản in bên ERP",
         "date_needed = null",
         "- Bản in HRM để TRỐNG ô ngày cần\n"
         "- Bản in ERP hiển thị NGÀY HÔM NAY (sai dữ liệu) — chênh lệch này là chủ đích, HRM đúng hơn\n"
         "- Ghi nhận vào biên bản test, không coi là lỗi HRM",
         "HRM sửa lỗi ERP: date null → in rỗng thay vì ngày hôm nay"),

        ("035", "In phiếu không thuộc phạm vi xem", "P0",
         "Phiếu id = X ngoài phạm vi xem của user",
         "1. Truy cập thẳng /finance/product-transfer-requests/X/print",
         "Phiếu ngoài phạm vi",
         "- Hiện thông báo lỗi không có quyền (API trả 403)\n- Nút In bị vô hiệu\n- Không lộ nội dung phiếu",
         "printData cũng chặn bằng canView() như ERP"),

        # --- Export
        ("036", "Xuất Excel danh sách", "P0",
         "Phạm vi xem có 30 phiếu",
         "1. Bấm 'Xuất Excel'\n2. Mở file tải về",
         "30 phiếu",
         "- Tải về file 'danh_sach_yeu_cau_chuyen_hang.xlsx' (tên khớp ERP)\n- Toast 'Xuất Excel thành công'\n"
         "- File có 7 cột: STT, Mã yêu cầu, Người tạo, Ngày tạo, Người tiếp nhận, Ngày tiếp nhận, Trạng thái\n"
         "- Đủ 30 dòng, KHÔNG bị giới hạn theo trang hiện tại",
         "exportData: cùng searchByFilter nhưng get() tất cả"),

        ("037", "Xuất Excel CÓ áp bộ lọc đang chọn", "P0",
         "Phạm vi xem 30 phiếu; lọc Trạng thái = 'Chờ duyệt' còn 5 phiếu",
         "1. Lọc Trạng thái = 'Chờ duyệt'\n2. Bấm 'Xuất Excel'\n3. Đếm số dòng trong file",
         "Lọc còn 5 phiếu",
         "- File Excel chứa ĐÚNG 5 phiếu đang lọc (khác màn Danh mục gói bảo dưỡng — màn này CÓ áp bộ lọc)\n"
         "- Phạm vi quyền vẫn được giữ (không lộ phiếu ngoài phạm vi)",
         "FE gửi toàn bộ filters khi export; BE dùng cùng searchByFilter"),

        ("038", "Xuất Excel có lọc khoảng ngày", "P1",
         "Lọc Ngày tạo từ = 01/08/2026 đến 10/08/2026",
         "1. Đặt bộ lọc ngày\n2. Bấm 'Xuất Excel'\n3. Mở file, đọc phần đầu",
         "start_date và end_date",
         "- File có dòng 'Từ ngày 01/08/2026 đến ngày 10/08/2026'\n"
         "- Chỉ lọc 1 đầu ngày thì hiện 'Từ ngày ...' hoặc 'Đến ngày ...' tương ứng",
         "exportFilterText port đủ 3 nhánh của ERP"),

        ("039", "Xuất Excel khi phạm vi rỗng", "P2",
         "Lọc bằng điều kiện chắc chắn không có kết quả",
         "1. Lọc mã 'PYCCH-99999999'\n2. Bấm 'Xuất Excel'\n3. Mở file",
         "0 phiếu",
         "- Vẫn tải về file hợp lệ với dòng tiêu đề, phần dữ liệu rỗng\n- Không lỗi 500, không file hỏng",
         "Export với danh sách rỗng vẫn chạy"),

        ("040", "Đối chiếu file Excel 2 cổng", "P1",
         "Cùng 1 user, cùng bộ lọc trên cả ERP và HRM",
         "1. Xuất Excel ở ERP với bộ lọc X\n2. Xuất Excel ở HRM với bộ lọc tương đương\n3. So sánh số dòng và nội dung",
         "Cùng bộ lọc",
         "- Số dòng và nội dung 7 cột GIỐNG NHAU\n"
         "- Chênh lệch chấp nhận được: ô tên người tạo/tiếp nhận khi nhân viên đã bị xóa — HRM để trống, ERP có thể lỗi",
         "HRM null-safe hơn ERP ở tên nhân viên và trạng thái lạ"),
    ]),

    # ---------------------------------------------------------------- VI
    ("VI", "EDGE CASES & VALIDATION", [
        ("001", "Lưu phiếu không có hàng hóa nào", "P0",
         "Đang ở màn tạo, chưa thêm hàng hóa nào",
         "1. Bấm 'Lưu'",
         "products rỗng",
         "- Hiện lỗi 'Bắt buộc phải có ít nhất 1 hàng hóa' (toast + inline dưới bảng hàng hóa)\n"
         "- KHÔNG gửi request lên server\n- Trang không điều hướng",
         "Rule products required|array|min:1"),

        ("002", "Lỗi chỉ hiện SAU lần bấm Lưu đầu tiên", "P0",
         "Vừa mở màn tạo phiếu",
         "1. Quan sát ngay khi vào form\n2. Click vào ô SL rồi click ra ngoài (không nhập)\n3. Bấm Lưu",
         "—",
         "- Bước 1 và 2: KHÔNG ô nào viền đỏ / hiện lỗi\n- Bước 3: lỗi hiện đồng loạt ở các ô còn thiếu",
         "Cờ touched — chuẩn CLAUDE.md"),

        ("003", "Bỏ trống trường bắt buộc của dòng con", "P0",
         "Đã thêm 1 hàng hóa với 1 dòng con trống",
         "1. Bấm 'Lưu' khi chưa khai gì ở dòng con",
         "Dòng con trống",
         "- Báo 'Bắt buộc chọn' ở ô Khách hàng\n- Báo 'Bắt buộc nhập' ở ô SL và ô Ghi chú\n"
         "- Báo 'Bắt buộc chọn' ở ô Ngày cần\n- Toast 'Vui lòng kiểm tra lại dữ liệu nhập', trang cuộn tới ô lỗi đầu tiên",
         "Rule: customer_id/qty/note/date_needed đều required"),

        ("004", "Ghi chú dòng con là BẮT BUỘC", "P0",
         "Đã khai đủ Khách hàng, SL, Ngày cần; để trống Ghi chú của dòng con",
         "1. Bấm 'Lưu'",
         "note dòng con rỗng",
         "- Báo 'Bắt buộc nhập' ở ô Ghi chú của dòng con\n- Không lưu được\n"
         "- Lưu ý: ghi chú CẤP PHIẾU thì không bắt buộc, ghi chú CẤP DÒNG CON thì bắt buộc (đúng như ERP)",
         "products.*.details.*.note required — khác với note cấp phiếu (nullable)"),

        ("005", "Số lượng nhỏ hơn 1", "P0",
         "Đang khai dòng con",
         "1. Nhập SL = 0\n2. Bấm 'Lưu'\n3. Thử lại với SL = -5",
         "qty = 0 rồi -5",
         "- Báo 'Số lượng tối thiểu là 1' ở ô SL\n- Không lưu được ở cả 2 trường hợp",
         "Rule qty numeric|min:1"),

        ("006", "Số lượng vượt giới hạn", "P1",
         "Đang khai dòng con",
         "1. Nhập SL = 1000000000 (1 tỷ)\n2. Bấm 'Lưu'",
         "qty = 1.000.000.000",
         "- Báo 'Số lượng không được vượt quá 999.999.999'\n- Không lưu được\n- Nhập 999999999 thì lưu được",
         "Rule qty max:999999999"),

        ("007", "Ngày cần hàng là ngày hôm nay hoặc quá khứ (tạo mới)", "P0",
         "Đang tạo phiếu mới, hôm nay = 07/08/2026",
         "1. Chọn Ngày cần = 07/08/2026 (hôm nay), bấm Lưu\n2. Đổi thành 06/08/2026, bấm Lưu\n3. Đổi thành 08/08/2026, bấm Lưu",
         "Hôm nay / hôm qua / ngày mai",
         "- Bước 1 và 2: báo 'Ngày cần hàng phải sau ngày hôm nay', không lưu được\n"
         "- Bước 3: lưu thành công",
         "Rule after:today khi tạo mới — port nguyên ERP"),

        ("008", "Sửa phiếu nháp cũ có dòng con ngày quá khứ — GIỮ NGUYÊN ngày", "P0",
         "Phiếu nháp lập từ tuần trước, có dòng con Ngày cần = ngày đã qua",
         "1. Bấm Sửa phiếu đó\n2. KHÔNG đụng vào ô Ngày cần của dòng cũ\n3. Chỉ sửa Ghi chú rồi bấm 'Lưu'",
         "date_needed thuộc quá khứ, giữ nguyên",
         "- LƯU THÀNH CÔNG, không báo lỗi ngày\n"
         "- Đây là điểm HRM NỚI SO VỚI ERP: bên ERP kiểm ngày cho MỌI dòng nên phiếu nháp cũ không sửa nổi\n"
         "- Đối chiếu: thao tác tương tự trên ERP sẽ báo lỗi 'Không hợp lệ' ở ô ngày",
         "Rule nới #6: dòng cũ giữ nguyên ngày thì bỏ qua kiểm tra quá khứ"),

        ("009", "Sửa phiếu nháp cũ và ĐỔI ngày sang quá khứ", "P0",
         "Cùng phiếu ở TC_06.008",
         "1. Bấm Sửa\n2. Đổi Ngày cần của dòng cũ sang một ngày quá khứ KHÁC\n3. Bấm 'Lưu'",
         "Đổi sang ngày quá khứ khác",
         "- Báo 'Ngày cần hàng phải sau ngày hôm nay', KHÔNG lưu được\n"
         "- Chứng minh rule nới chỉ miễn cho dòng GIỮ NGUYÊN ngày, không phải miễn hết",
         "Chỉ bỏ qua khi id detail khớp VÀ ngày không đổi"),

        ("010", "Thêm dòng con mới vào phiếu cũ với ngày quá khứ", "P0",
         "Phiếu nháp cũ đang mở màn Sửa",
         "1. Thêm 1 dòng con MỚI, chọn Ngày cần = hôm qua\n2. Bấm 'Lưu'",
         "Dòng mới, ngày quá khứ",
         "- Báo lỗi ở đúng dòng con MỚI\n- Các dòng cũ giữ nguyên ngày không bị báo lỗi\n- Không lưu được",
         "Dòng không có id → coi là dòng mới → kiểm tra chặt"),

        ("011", "Gửi id dòng con của phiếu KHÁC qua API", "P0",
         "Phiếu nháp X của mình; dòng con id = D thuộc phiếu Y của người khác",
         "1. Gọi API sửa phiếu X, gửi kèm dòng con có id = D và ngày quá khứ",
         "id dòng con lạ",
         "- Báo lỗi ngày quá khứ như dòng mới (không được bỏ qua kiểm tra)\n"
         "- Không lộ thông tin ngày cần của phiếu người khác qua việc có/không báo lỗi",
         "existingDetailDates chỉ nạp dòng con thuộc CHÍNH phiếu đang sửa"),

        ("012", "Tạo phiếu không đính kèm file PDF", "P0",
         "Đang tạo phiếu mới, đã khai đủ hàng hóa và dòng con",
         "1. Không đính file nào\n2. Bấm 'Lưu'",
         "attachments rỗng",
         "- Báo 'Bắt buộc phải đính kèm ít nhất 1 file PDF'\n- Không lưu được\n"
         "- Áp dụng cho cả nút 'Lưu' lẫn 'Lưu & Gửi duyệt'",
         "Store: attachments required|array|min:1 — giống ERP"),

        ("013", "Sửa phiếu KHÔNG bắt buộc đính file mới", "P0",
         "Phiếu nháp đã có 1 file PDF",
         "1. Bấm Sửa\n2. Không thêm file mới, sửa ghi chú\n3. Bấm 'Lưu'",
         "Không thêm file",
         "- Lưu thành công (không bắt buộc file mới khi sửa)\n- File cũ vẫn còn nguyên",
         "Update: attachments nullable — khác store"),

        ("014", "Ghi chú phiếu vượt 255 ký tự", "P1",
         "Đang ở màn tạo",
         "1. Nhập Ghi chú = chuỗi 300 ký tự\n2. Khai đủ dữ liệu còn lại, bấm Lưu",
         "note 300 ký tự",
         "- Báo 'Ghi chú không được nhập quá 255 ký tự'\n- Không lưu được",
         "Rule note max:255"),

        ("015", "Ghi chú dòng con vượt 255 ký tự", "P1",
         "Đang khai dòng con",
         "1. Nhập Ghi chú dòng con = chuỗi 300 ký tự\n2. Bấm Lưu",
         "note dòng con 300 ký tự",
         "- Báo 'Ghi chú không được nhập quá 255 ký tự' ở đúng dòng con\n- Không lưu được",
         "Rule products.*.details.*.note max:255"),

        ("016", "Gửi trùng hàng hóa qua API", "P0",
         "Có hàng hóa id = 100",
         "1. Gọi API tạo phiếu với mảng products chứa 2 phần tử cùng product_id = 100, dữ liệu còn lại hợp lệ",
         "2 dòng cùng product_id",
         "- API trả 422 với thông báo 'Hàng hóa bị trùng trong phiếu'\n- Không tạo phiếu\n"
         "- ERP cũng chặn nhưng chỉ báo chung 'Thông tin không hợp lệ!' — HRM báo rõ hơn",
         "withValidator: kiểm trùng product_id"),

        ("017", "Gửi ĐVT không thuộc hàng hóa", "P0",
         "ĐVT id = 5 không được khai cho hàng hóa id = 100",
         "1. Gọi API tạo phiếu với product_id = 100 và unit_id = 5",
         "unit không thuộc product",
         "- API trả 422 với thông báo 'Đơn vị tính không thuộc hàng hóa đã chọn'\n- KHÔNG lỗi 500\n"
         "- ERP trong tình huống này báo lỗi hệ thống 500 — HRM xử lý tốt hơn, ghi nhận là cải tiến",
         "syncProducts kiểm product_units trước khi ghi"),

        ("018", "Hàng hóa chưa có giá niêm yết theo ĐVT", "P1",
         "Hàng hóa có ĐVT nhưng chưa khai giá bán lẻ cho ĐVT đó",
         "1. Thêm hàng đó vào phiếu, chọn ĐVT chưa có giá\n2. Bấm Lưu",
         "Thiếu dòng giá",
         "- API trả 422 'Hàng hóa chưa có giá niêm yết cho đơn vị tính đã chọn'\n- Không tạo phiếu, không lỗi 500",
         "syncProducts kiểm product_unit_prices trước khi ghi"),

        ("019", "Hàng hóa / khách hàng bị xóa giữa chừng", "P2",
         "Đang mở form với hàng hóa id = 100; song song đó hàng này bị xóa ở cổng khác",
         "1. Sau khi hàng bị xóa, bấm Lưu phiếu",
         "Hàng bị xóa giữa chừng",
         "- API trả 422 với thông báo rõ ràng ('Hàng hóa không tồn tại' hoặc 'Hàng hóa hoặc đơn vị tính không còn tồn tại')\n"
         "- Không tạo phiếu dở dang: 3 bảng đều không có dữ liệu rác",
         "Toàn bộ thao tác lưu bọc trong giao dịch — rollback khi lỗi"),

        ("020", "Giá lưu trong DB so với giá hiển thị trên form", "P1",
         "Công ty của người lập có hệ số giá ≠ 1 cho hàng 'Bulong M10'",
         "1. Thêm hàng vào phiếu, ghi lại cột giá hiển thị trên form\n2. Lưu phiếu\n"
         "3. Kiểm tra cột price của dòng hàng trong DB\n4. So sánh 2 con số",
         "Hệ số công ty ≠ 1",
         "- Giá trên form là giá ĐÃ áp hệ số công ty\n- Giá lưu trong DB là giá THÔ (chưa áp hệ số)\n"
         "- Hai con số lệch nhau là ĐÚNG nguyên bản ERP — ghi nhận để nghiệp vụ xác nhận có muốn thống nhất không",
         "syncProducts lưu giá thô như ERP; productUnits hiển thị giá đã áp hệ số"),

        ("021", "Chống double-click nút Lưu", "P0",
         "Đang ở màn tạo với dữ liệu hợp lệ",
         "1. Bấm 'Lưu' liên tiếp 3 lần thật nhanh\n2. Kiểm tra danh sách và bảng DB",
         "3 lần bấm < 1 giây",
         "- Nút Lưu chuyển sang trạng thái đang xử lý sau lần bấm đầu\n"
         "- Chỉ tạo ĐÚNG 1 phiếu, không có mã PYCCH trùng lặp\n- Bảng con không bị nhân đôi",
         "isSubmitSave chặn gọi lại"),

        ("022", "Bấm Lưu và Lưu & Gửi duyệt liên tiếp", "P1",
         "Đang ở màn tạo với dữ liệu hợp lệ",
         "1. Bấm 'Lưu' rồi ngay lập tức bấm 'Lưu & Gửi duyệt'",
         "2 nút khác nhau",
         "- Chỉ 1 hành động được thực hiện, chỉ 1 phiếu được tạo\n- Không tạo 2 phiếu với 2 trạng thái khác nhau",
         "Cả 2 nút cùng dùng cờ isSubmitSave"),

        ("023", "Nhập ký tự đặc biệt và tiếng Việt có dấu", "P1",
         "Đang khai phiếu",
         "1. Ghi chú phiếu = 'Giao gấp & thanh toán 100% (đợt 1) #A'\n2. Ghi chú dòng con tương tự\n3. Lưu rồi mở chi tiết và In",
         "Chuỗi có &, %, #, ()",
         "- Lưu thành công\n- Màn chi tiết hiển thị đúng nguyên văn, không bị mã hoá HTML\n"
         "- Bản in hiển thị đúng, không vỡ bảng",
         "Bản in escape HTML — HRM bổ sung so với ERP (ERP in thô)"),

        ("024", "Nhập thẻ script vào ghi chú (XSS)", "P0",
         "Đang khai phiếu",
         "1. Ghi chú = '<script>alert(1)</script>'\n2. Lưu\n3. Mở chi tiết\n4. Mở màn In",
         "note = '<script>alert(1)</script>'",
         "- KHÔNG xuất hiện hộp thoại alert ở cả màn chi tiết lẫn màn in\n"
         "- Chuỗi hiển thị nguyên văn dạng text\n- Không lỗi JS console",
         "Bản in dùng v-html nên giá trị phải được escape ở server — HRM đã bổ sung"),

        ("025", "Ghi chú chứa ký tự $ khi in", "P2",
         "Phiếu có ghi chú 'Giá $100 và $2 phụ phí'",
         "1. In phiếu đó\n2. Đọc ô Ghi chú trên bản in",
         "note chứa ký tự $",
         "- Bản in hiển thị đầy đủ 'Giá $100 và $2 phụ phí'\n- KHÔNG bị mất ký tự hay mất đoạn văn bản",
         "Escape ký tự $ và \\ trước khi thay vào mẫu in"),

        ("026", "Mất kết nối mạng khi đang lưu", "P2",
         "Đang ở màn tạo với dữ liệu hợp lệ",
         "1. Ngắt mạng (DevTools → Offline)\n2. Bấm Lưu\n3. Quan sát",
         "Network: Offline",
         "- Hiện toast lỗi, thanh loading tắt, nút Lưu bật lại\n"
         "- Trang KHÔNG điều hướng, toàn bộ hàng hóa/dòng con/file đã khai còn nguyên",
         "finally: reset cờ + tắt loading; không điều hướng khi lỗi"),

        ("027", "Xóa phiếu đã bị người khác xử lý", "P1",
         "Phiếu nháp của mình đang mở trên màn danh sách; song song đó phiếu được gửi duyệt ở cổng ERP",
         "1. Trên màn HRM (chưa reload), bấm Xóa phiếu đó, xác nhận\n2. Quan sát",
         "Trạng thái đã đổi ở cổng khác",
         "- API trả 403 'Chỉ xóa được phiếu Đang tạo do chính bạn lập'\n"
         "- Hiện toast lỗi rõ ràng, danh sách vẫn dùng được\n- Phiếu không bị xóa",
         "canDelete đọc trạng thái mới nhất từ DB"),

        ("028", "Phiếu nhiều hàng hóa và nhiều dòng con", "P1",
         "Đang ở màn tạo",
         "1. Thêm 10 hàng hóa, mỗi hàng 3 dòng con (tổng 30 dòng con)\n2. Lưu\n3. Mở lại màn Sửa và đối chiếu\n4. In phiếu",
         "10 hàng × 3 dòng con",
         "- Lưu thành công trong thời gian chấp nhận được\n"
         "- Mở lại đủ 10 hàng và 30 dòng con, dữ liệu đúng vị trí\n"
         "- SL từng dòng hàng = tổng 3 dòng con của nó\n- Bản in gộp ô đúng, không vỡ khi sang trang",
         "syncProducts tạo lại toàn bộ; bản in rowspan theo số khách hàng"),

        ("029", "Đối chiếu dữ liệu sau khi lưu ở HRM, xem lại ở ERP", "P0",
         "Có quyền truy cập cả HRM và ERP",
         "1. Tạo phiếu ở HRM với 2 hàng hóa, 3 dòng con, 1 file PDF, gửi duyệt\n"
         "2. Mở phiếu đó ở màn chi tiết bên ERP\n3. So sánh từng trường",
         "Phiếu tạo từ HRM",
         "- ERP hiển thị đủ: mã, trạng thái Chờ duyệt, người tạo, ghi chú, 2 hàng hóa với đúng ĐVT và SL, 3 dòng con đúng khách hàng/SL/ngày cần/ghi chú\n"
         "- File đính kèm mở được từ ERP\n- Không có trường nào trống bất thường",
         "2 cổng dùng chung 3 bảng — dữ liệu phải đọc được 2 chiều"),

        ("030", "Đối chiếu dữ liệu tạo ở ERP, sửa ở HRM", "P0",
         "Có quyền truy cập cả 2 cổng",
         "1. Tạo phiếu nháp ở ERP với 1 hàng hóa\n2. Mở màn Sửa phiếu đó ở HRM\n"
         "3. Thêm 1 hàng hóa, lưu\n4. Mở lại ở ERP",
         "Phiếu tạo từ ERP",
         "- HRM nạp đúng dữ liệu phiếu từ ERP\n- Sau khi sửa ở HRM, ERP hiển thị đủ 2 hàng hóa\n"
         "- Trạng thái, người tạo, công ty/phòng ban của phiếu không bị thay đổi ngoài ý muốn",
         "Sửa ở HRM không được phá dữ liệu do ERP tạo"),
    ]),

    # ---------------------------------------------------------------- VII
    ("VII", "CÔ LẬP DỮ LIỆU & BẢO MẬT", [
        ("001", "Gọi API khi chưa đăng nhập", "P0",
         "Chưa có token hoặc đã đăng xuất",
         "1. Gọi GET /api/v1/finance/product-transfer-requests không gắn Authorization\n"
         "2. Gọi GET /v1/finance/product-transfer-requests/{id}/print-data\n3. Gọi GET .../export",
         "Không có Bearer token",
         "- Cả 3 request trả HTTP 401\n- Không lộ bất kỳ dữ liệu phiếu nào",
         "Route group middleware auth:api"),

        ("002", "Truy cập URL màn hình khi chưa đăng nhập", "P0",
         "Trình duyệt đã đăng xuất",
         "1. Gõ URL /finance/product-transfer-requests\n2. Gõ URL /finance/product-transfer-requests/1",
         "—",
         "- Cả 2 đều chuyển về màn đăng nhập\n- Không hiển thị dữ liệu dù chỉ trong tích tắc",
         "Middleware auth phía FE"),

        ("003", "Xuất Excel bị giới hạn theo phạm vi quyền", "P0",
         "User A không có quyền xem theo cấp, chỉ có 3 phiếu của mình; hệ thống có 30 phiếu",
         "1. Đăng nhập user A\n2. Bấm 'Xuất Excel' (không lọc gì)\n3. Đếm số dòng trong file",
         "User A: phạm vi 3 phiếu",
         "- File chỉ có 3 phiếu của A\n- KHÔNG lộ 27 phiếu ngoài phạm vi\n"
         "- Đúng như ERP: export dùng chung điều kiện phạm vi với danh sách",
         "exportData dùng cùng searchByFilter với index"),

        ("004", "Không lộ nội dung phiếu qua API in", "P0",
         "Phiếu id = X ngoài phạm vi xem của user A",
         "1. Lấy token user A\n2. Gọi GET /v1/finance/product-transfer-requests/X/print-data",
         "Phiếu ngoài phạm vi",
         "- API trả 403, body KHÔNG chứa nội dung mẫu in hay bất kỳ trường nào của phiếu",
         "printData chặn bằng canView() như show"),

        ("005", "Thông báo gửi đúng người", "P0",
         "Công ty 1 có 2 Kế toán kho (F1, F2); công ty 2 có 1 Kế toán kho (F3)",
         "1. User A (công ty 1) tạo phiếu và gửi duyệt\n2. Kiểm tra chuông của F1, F2, F3",
         "Phiếu công ty 1",
         "- F1 và F2 nhận được thông báo\n- F3 (công ty 2) KHÔNG nhận được\n"
         "- Nội dung: '<tên A> vừa tạo yêu cầu chuyển hàng: <mã>' và bấm vào mở đúng màn chi tiết phiếu",
         "Danh sách người nhận lọc theo công ty của phiếu"),

        ("006", "Lỗi gửi thông báo không làm hỏng thao tác chính", "P2",
         "Giả lập lỗi ở kênh gửi thông báo (tắt Redis hoặc chặn kênh đẩy)",
         "1. Tạo phiếu và bấm 'Lưu & Gửi duyệt'\n2. Quan sát kết quả lưu\n3. Kiểm tra log hệ thống",
         "Kênh thông báo lỗi",
         "- Phiếu VẪN được lưu thành công với trạng thái 'Chờ duyệt'\n"
         "- Không hiện lỗi chặn người dùng\n- Lỗi gửi thông báo được ghi vào log để theo dõi",
         "notify bọc try/catch + ghi log, không làm fail giao dịch"),

        ("007", "Xử lý khi API danh sách trả lỗi 500", "P1",
         "Giả lập GET danh sách trả HTTP 500",
         "1. Giả lập lỗi\n2. Reload màn",
         "Response: 500",
         "- Hiện toast 'Lỗi khi tải dữ liệu'\n- Bảng rỗng, loading đã tắt\n- Panel lọc vẫn thao tác được để thử lại",
         "catch: chỉ bỏ qua toast khi 403"),

        ("008", "Đối chiếu phạm vi dữ liệu 2 cổng theo từng mức quyền", "P0",
         "Chuẩn bị 4 user tương ứng 4 mức quyền: tổng công ty, công ty, phòng ban, không quyền",
         "1. Với TỪNG user: mở màn ERP type=all và màn HRM, ghi lại tổng số phiếu\n2. So sánh từng cặp",
         "4 user × 2 cổng",
         "- Với cả 4 user, tổng số phiếu ở ERP và HRM GIỐNG NHAU\n"
         "- Danh sách mã phiếu trùng khớp\n- Nếu lệch: ghi rõ user nào, mã phiếu nào, phía nào thừa/thiếu",
         "Đây là phép kiểm chứng quan trọng nhất của việc port searchByFilter"),
    ]),

    # ---------------------------------------------------------------- VIII
    ("VIII", "E2E FLOW", [
        ("001", "Luồng đầy đủ: tạo nháp → sửa → gửi duyệt → không duyệt → sửa lại → gửi lại", "P0",
         "User A (người lập) và user F (Kế toán kho cùng công ty)",
         "1. User A: tạo phiếu với 2 hàng hóa, mỗi hàng 2 khách hàng, đính 1 file PDF → bấm 'Lưu'\n"
         "2. User A: bấm Sửa, thêm 1 hàng hóa nữa, thêm 1 file PDF → 'Lưu'\n"
         "3. User A: bấm 'Lưu & Gửi duyệt'\n"
         "4. User F: nhận thông báo, mở chi tiết, nhập ghi chú duyệt 'Thiếu chứng từ' → 'Không duyệt'\n"
         "5. User A: nhận thông báo, mở phiếu, đọc ghi chú duyệt, bấm Sửa, bổ sung file → 'Lưu & Gửi duyệt'\n"
         "6. User F: mở lại phiếu, bấm 'Tổng hợp'",
         "1 phiếu qua đủ vòng đời",
         "- B1: trạng thái 'Đang tạo', mã PYCCH-xxxxx, nút Sửa/Xóa bật\n"
         "- B2: đủ 3 hàng hóa và 2 file\n- B3: trạng thái 'Chờ duyệt', nút Sửa/Xóa mờ, F nhận chuông\n"
         "- B4: trạng thái về 'Đang tạo', Người/Ngày tiếp nhận = F/hôm nay, ghi chú duyệt được lưu, A nhận chuông\n"
         "- B5: A sửa được (đã về nháp), gửi lại thành công → 'Chờ duyệt'\n"
         "- B6: mở tab ERP màn tạo phiếu đề nghị xuất hàng với dữ liệu phiếu này",
         "Bao trùm toàn bộ vòng đời phiếu ở phần HRM ghi được (status 2 ↔ 3)"),

        ("002", "Luồng tạo phiếu với xem tồn kho", "P0",
         "Hàng 'Bulong M10' có tồn ở kho K1 và K2 khác nhau",
         "1. Tạo phiếu mới, thêm hàng 'Bulong M10'\n2. Chọn 'Xem tồn theo kho' = nhóm kho chứa K1 và K2\n"
         "3. Ghi lại số tồn / khả dụng\n4. Đổi sang riêng kho K1, ghi lại lần nữa\n"
         "5. Khai dòng con với SL lớn hơn tồn khả dụng\n6. Lưu nháp",
         "2 phạm vi kho",
         "- B2 và B4: số liệu tồn khác nhau đúng theo phạm vi kho đã chọn\n"
         "- B3: cột 'Khả dụng' không âm (tối thiểu bằng 0)\n"
         "- B5-B6: hệ thống VẪN cho lưu (màn này không chặn theo tồn kho) — ghi nhận đúng hành vi ERP",
         "Tồn kho chỉ để tham khảo khi lập phiếu, không phải điều kiện chặn"),

        ("003", "Luồng phân quyền: 3 user 3 mức quyền cùng nhìn 1 phiếu", "P0",
         "Phiếu 'Chờ duyệt' của user A thuộc công ty 1, phòng ban P1",
         "1. User A (người lập) mở phiếu, ghi lại các nút\n"
         "2. User F (Kế toán kho công ty 1) mở phiếu, ghi lại các nút\n"
         "3. User G (Super admin công ty 2) mở phiếu, ghi lại các nút\n"
         "4. User H (không quyền, khác phòng ban) thử mở phiếu",
         "4 vai trò",
         "- A: xem được, KHÔNG có nút Sửa (đã gửi duyệt), không có nút Không duyệt\n"
         "- F: xem được, CÓ ô ghi chú duyệt + nút 'Không duyệt' + 'Tổng hợp'\n"
         "- G: xem được nhưng KHÔNG có nút Không duyệt/Tổng hợp (khác công ty)\n"
         "- H: 403, không mở được\n- Kết quả khớp 1-1 với cùng phép thử trên ERP",
         "canView / canApprove / canEdit độc lập nhau"),

        ("004", "Luồng lọc — phân trang — xuất Excel — in", "P1",
         "Phạm vi xem có ≥ 30 phiếu đủ nhiều trạng thái",
         "1. Lọc Trạng thái = 'Chờ duyệt' + Ngày tạo từ 01/08/2026\n2. Đổi số dòng/trang = 10, sang trang 2\n"
         "3. Bấm 'Xuất Excel', đối chiếu số dòng với tổng đang hiển thị\n"
         "4. Mở In 1 phiếu bất kỳ ở trang 2\n5. Bấm Đặt lại",
         "Bộ lọc kép",
         "- B2: trang 2 chỉ chứa phiếu khớp bộ lọc, STT bắt đầu từ 11, bộ lọc không mất\n"
         "- B3: file Excel có đúng số dòng bằng tổng đang hiển thị, có dòng 'Từ ngày 01/08/2026'\n"
         "- B4: bản in đúng phiếu đã chọn\n- B5: bộ lọc về rỗng, danh sách hiển thị lại từ trang 1",
         "Bộ lọc, phân trang, export, in phải nhất quán trên cùng 1 tập dữ liệu"),

        ("005", "Luồng song song 2 cổng ERP ↔ HRM trên cùng 1 phiếu", "P0",
         "Có quyền truy cập cả 2 cổng; user F là Kế toán kho",
         "1. Tạo phiếu nháp ở HRM (2 hàng hóa) → Lưu\n"
         "2. Mở phiếu đó ở ERP, sửa thêm 1 hàng hóa → lưu ở ERP\n"
         "3. Quay lại HRM, mở màn Sửa, kiểm tra có đủ 3 hàng hóa\n"
         "4. Ở HRM bấm 'Lưu & Gửi duyệt'\n5. Ở ERP kiểm tra trạng thái và danh sách phiếu chờ duyệt\n"
         "6. Ở ERP thực hiện 'Không duyệt', rồi quay lại HRM xem trạng thái",
         "1 phiếu, thao tác xen kẽ 2 cổng",
         "- B3: HRM thấy đủ 3 hàng hóa do ERP thêm\n- B5: ERP thấy trạng thái 'Chờ duyệt'\n"
         "- B6: HRM thấy trạng thái về 'Đang tạo' kèm ghi chú duyệt và người tiếp nhận từ ERP\n"
         "- Không phát sinh bản ghi trùng, không mất dòng hàng nào ở cả 2 chiều",
         "Kiểm chứng 2 cổng ghi/đọc chung 3 bảng không xung đột"),

        ("006", "Luồng phiếu do chuỗi kho ERP đẩy trạng thái", "P1",
         "Phiếu đã được Kế toán kho 'Tổng hợp' bên ERP và đi tiếp chuỗi nghiệp vụ kho",
         "1. Sau khi Tổng hợp ở ERP, theo dõi phiếu ở màn HRM qua các mốc\n"
         "2. Ghi lại trạng thái hiển thị tại từng mốc\n3. Khi phiếu ở trạng thái 'Đã phân bổ', mở chi tiết ở HRM",
         "Trạng thái 1, 4, 5, 6, ..., 12",
         "- HRM hiển thị đúng tên và màu badge cho từng trạng thái mà ERP đẩy sang\n"
         "- HRM KHÔNG cho sửa/xóa các phiếu này\n"
         "- Ở trạng thái 'Đã phân bổ': màn chi tiết hiện thêm cột 'Được nhận' với số lượng đã phân bổ",
         "HRM chỉ ghi status 2 và 3; các trạng thái còn lại chỉ hiển thị"),
    ]),
]

# =========================================================================
# STYLES
# =========================================================================
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DESC_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
DESC_LABEL_FILL = PatternFill("solid", fgColor="FFF2CC")
DESC_BODY_FONT = Font(name="Calibri", size=11)
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")

SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

DATA_FONT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

COL_WIDTHS = {
    'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
    'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
    'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22
}

# =========================================================================
# BUILD
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)").font = Font(bold=True, size=12)
ws.merge_cells("B1:O1")
ws.row_dimensions[1].height = 22

for idx, (label, body) in enumerate(DESCRIPTION_BLOCK, start=2):
    a = ws.cell(idx, 1, label)
    a.font = DESC_LABEL_FONT
    a.fill = DESC_LABEL_FILL
    a.alignment = WRAP_TOP_LEFT
    a.border = BORDER
    b = ws.cell(idx, 2, body)
    b.font = DESC_BODY_FONT
    b.alignment = WRAP_TOP_LEFT
    b.border = BORDER
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=15)
    ws.row_dimensions[idx].height = max(40, body.count("\n") * 16 + 30)

t = ws.cell(11, 1, "Testcase _ %s" % FEATURE_NAME)
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.merge_cells("B11:E11")
ws.merge_cells("F11:H11")
fs = ws.cell(11, 6, "TEST SUMMARY")
fs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
fs.fill = TITLE_FILL
fs.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[11].height = 28

summary_rows = [
    (11, "Số trường hợp kiểm thử đạt (P):", '=COUNTIF(L18:N500,"Passed")'),
    (12, "Số trường hợp kiểm thử không đạt (F):", '=COUNTIF(L18:N500,"Failed")'),
    (13, "Số trường hợp kiểm thử đang xem xét (PE):", '=COUNTIF(L18:N500,"Pending")'),
    (14, "Số trường hợp kiểm thử chưa thực hiện:", '=COUNTIF(L18:N500,"Not Executed")'),
    (15, "Tổng số trường hợp kiểm thử:", '=COUNTIF(L18:N500,"<>")'),
]
for r, label, formula in summary_rows:
    lc = ws.cell(r, 9, label)
    lc.font = SUMMARY_LABEL_FONT
    lc.fill = SUMMARY_LABEL_FILL
    lc.alignment = Alignment(vertical="center", horizontal="right")
    lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    vc = ws.cell(r, 12, formula)
    vc.font = SUMMARY_VALUE_FONT
    vc.fill = SUMMARY_LABEL_FILL
    vc.alignment = SUMMARY_VALUE_ALIGN
    vc.border = BORDER
    ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=15)
    if r > 11:
        ws.row_dimensions[r].height = 22

ws.row_dimensions[16].height = 8

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "Giải thích nghiệp vụ", "KQ thực tế",
    "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3",
    "Ghi chú",
]
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(17, i, h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws.row_dimensions[17].height = 36

current_row = 18
data_row_idx = 0


def write_section_row(title):
    global current_row
    cell = ws.cell(current_row, 3, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = BORDER
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
    for col in (1, 2):
        ws.cell(current_row, col).fill = SECTION_FILL
        ws.cell(current_row, col).border = BORDER
    ws.row_dimensions[current_row].height = 26
    current_row += 1


def write_tc(tc_id, function, priority, precondition, steps, test_data, expected,
             business_note, module=MODULE_NAME, group=""):
    global current_row, data_row_idx
    values = [
        module, group, tc_id, function, priority,
        precondition, steps, test_data,
        expected, business_note, "",
        "Not Executed", "Not Executed", "Not Executed",
        "",
    ]
    fill = DATA_FONT_FILL_EVEN if data_row_idx % 2 == 1 else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(current_row, i, v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = WRAP_TOP_LEFT if i != 5 else WRAP_TOP_CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
    longest = max(len(str(v)) for v in values)
    newlines = max(str(v).count("\n") for v in values)
    ws.row_dimensions[current_row].height = max(30, min(210, max(longest // 4, newlines * 15 + 20)))
    current_row += 1
    data_row_idx += 1


total_tc = 0
p0_tc = 0

if HAS_ROLE_SECTION:
    write_section_row("Phân quyền & truy cập")
    for suffix, func, prio, pre, steps, td, exp, note in ROLE_TCS:
        write_tc("TC-ROLE-%s" % suffix, func, prio, pre, steps, td, exp, note,
                 group="Phân quyền & truy cập")
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

ROMANS = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section_row("%s. %s" % (roman, title))
    sec_idx = ROMANS.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp, note in tcs:
        tc_id = "TC_%02d.%03d" % (sec_idx, int(tc_num))
        write_tc(tc_id, func, prio, pre, steps, td, exp, note, group=title)
        total_tc += 1
        if prio == "P0":
            p0_tc += 1

dv = DataValidation(
    type="list",
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
    showDropDown=False,
)
dv.add("L18:N%d" % (current_row + 100))
ws.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print("Generated: %s" % OUTPUT_FILE)
print("Rows: 1-10 description, 11-15 summary, 17 header, 18-%d data" % (current_row - 1))
print("Total TC: %d | P0: %d (%.0f%%)" % (total_tc, p0_tc, p0_tc * 100.0 / total_tc))
