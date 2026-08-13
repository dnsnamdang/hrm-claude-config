# -*- coding: utf-8 -*-
"""
Sinh file testcase Excel cho man "Danh muc dich vu sua chua va chi phi khac"
(phan he CSKH).

Bam theo FORM cua file mau: D:\\CompanyProject\\Document\\TC mau phan bomlist.xlsx
  - Row 1      : tieu de "MO TA TINH NANG"
  - Row 2-10   : 9 muc mo ta
  - Row 11-15  : title + TEST SUMMARY (2 khoi DNS / TP)
  - Row 17     : header cot
  - Row 18+    : section (merge C:Q) + data
  - Font Times New Roman 12, border thin, dropdown 2 khoi check DNS (K/L/M) va TP (O/P/Q)

⚠️ NGUYEN TAC NGON NGU (user chot 2026-08-12):
   Tai lieu nay danh cho QA va bo phan nghiep vu, KHONG phai dev.
   TUYET DOI khong viet ten bang / ten cot DB, id quyen, ten ham, ten route, ma HTTP.
   Chi dung dung nhan hien tren man hinh va cau chu nguoi dung hieu duoc.
"""
import sys

# Console Windows mac dinh cp1252 -> print() chuoi tieng Viet se nem UnicodeEncodeError.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = r"d:\CompanyProject\hrm\hrm-claude-config\hrm\.plans\gop-db\customer-care-cost-catalog\testcase.xlsx"
SHEET_NAME = "Trang tính1"
FEATURE_NAME = "Danh mục dịch vụ sửa chữa và chi phí khác - Cập nhật ngày 12/08/2026"
MODULE_NAME = "DV sửa chữa & CP khác"

# ============================================================================
# 9 MUC MO TA
# ============================================================================
DESCRIPTION_BLOCK = [
    ("1. Mục đích tính năng",
     "Kiểm thử TOÀN BỘ màn hình “Danh mục dịch vụ sửa chữa và chi phí khác” của phân hệ CSKH "
     "(vào theo đường dẫn: CSKH → nhóm menu “Danh mục - Dịch vụ” → “Danh mục dịch vụ sửa chữa và "
     "chi phí khác”).\n"
     " ► Phạm vi kiểm thử: danh sách, tìm kiếm, bộ lọc, sắp xếp, phân trang, Tạo mới, Sửa, Xem, "
     "Khóa, Mở khóa, Xóa, Xuất Excel, phân quyền và toàn bộ ràng buộc nhập liệu.\n"
     " ► Danh mục này khai báo 3 thông số dùng cho nghiệp vụ phía sau: % Tính giá vốn, % VAT và "
     "ĐM giảm giá. Đây là nguồn dữ liệu cho Báo giá hãng và Hợp đồng hãng.\n"
     " ► Màn hình dùng CHUNG dữ liệu với màn danh mục chi phí tương ứng bên phần mềm ERP: sửa ở "
     "cổng nào thì cổng còn lại cũng thấy ngay."),

    ("2. Đối tượng được tính / hiển thị",
     "► Danh sách hiển thị toàn bộ dịch vụ sửa chữa và chi phí khác đã khai báo. Dữ liệu thực tế "
     "trên cổng dev đang có 526 dòng.\n"
     " ► 2 trạng thái: “Hoạt động” (nhãn xanh) và “Khóa” (nhãn đỏ).\n"
     " ► 2 phân loại: “Có tính doanh thu” (nhãn xanh) và “Chi phí khác” (nhãn đỏ). Phân loại này "
     "do ô tích “Dịch vụ có tính doanh thu” trên form quyết định.\n"
     " ► Cột “ĐM giảm giá” hiển thị định mức giảm giá của CÔNG TY ĐANG CHỌN của người đăng nhập. "
     "Cùng một dịch vụ, người của công ty A và công ty B có thể thấy hai mức khác nhau — đây là "
     "hành vi đúng. Công ty chưa khai báo thì hiển thị “—”.\n"
     " ► 9 cột danh sách: STT | Tên dịch vụ / chi phí | Phân loại | ĐM giảm giá | % Tính giá vốn | "
     "% VAT | Trạng thái | Cập nhật | Hành động."),

    ("3. Đối tượng bị ẩn / không tính",
     "► Các khoản “Chi phí phải trả” và “Chi phí bán hàng” KHÔNG thuộc màn này (vẫn quản lý bên "
     "phần mềm ERP), nên không xuất hiện trong danh sách.\n"
     " ► Nút Tạo mới / Sửa / Xóa / Khóa / Mở khóa CHỈ hiện với người có quyền quản lý danh mục. "
     "Người chỉ được xem thì chỉ thấy nút Xem và nút Xuất Excel.\n"
     " ► 2 dịch vụ hệ thống là “Chi phí đi lại” và “Chi phí vận chuyển” bị khóa cứng: không sửa, "
     "không xóa, không khóa, không mở khóa được. Nút vẫn hiển thị nhưng bị làm mờ.\n"
     " ► Dòng đang ở trạng thái Khóa: nút Sửa và nút Xóa bị làm mờ, phải Mở khóa trước.\n"
     " ► Cửa sổ Tạo mới KHÔNG có ô “Trạng thái” (dịch vụ mới luôn ở trạng thái Hoạt động); ô này "
     "chỉ xuất hiện ở cửa sổ Sửa.\n"
     " ► Nút “Lưu & Tiếp tục” chỉ có ở cửa sổ Tạo mới, không có ở cửa sổ Sửa và Xem."),

    ("4. Bộ lọc thời gian áp dụng cho",
     "► Màn hình KHÔNG có bộ lọc theo khoảng thời gian.\n"
     " ► Cột “Cập nhật” chỉ để hiển thị: dòng trên là ngày giờ sửa gần nhất, dòng dưới là người "
     "sửa (dạng MÃ NV - Họ tên). Dịch vụ chưa từng được sửa thì hiển thị người tạo.\n"
     " ► Cột “Cập nhật” có sắp xếp được.\n"
     " ► Khi mới vào màn, danh sách sắp xếp theo ngày tạo giảm dần (dịch vụ mới nhất nằm đầu).\n"
     " ► Hệ thống ghi nhớ bộ lọc đang áp dụng trong 10 phút: rời màn rồi quay lại trong khoảng "
     "thời gian đó thì bộ lọc cũ còn nguyên; quá 10 phút thì bộ lọc tự trở về mặc định."),

    ("5. Cấu trúc dữ liệu / cây phân cấp",
     "► Danh mục PHẲNG, không có cấu trúc cha - con, không gom nhóm.\n"
     " ► Mỗi dòng gồm: Tên dịch vụ / chi phí, Phân loại, % Tính giá vốn, % VAT, Trạng thái, và "
     "ĐM giảm giá.\n"
     " ► Riêng ĐM giảm giá được lưu TÁCH RIÊNG THEO TỪNG CÔNG TY: một dịch vụ có thể có nhiều mức "
     "giảm giá, mỗi công ty một mức. Người dùng chỉ nhìn thấy và chỉ sửa được mức của công ty mình "
     "đang chọn.\n"
     " ► Dịch vụ trong danh mục này được sử dụng ở 2 nơi: Báo giá hãng và Hợp đồng hãng. Đây là "
     "căn cứ để hệ thống quyết định một dịch vụ được xóa hẳn hay chỉ được chuyển sang Khóa."),

    ("6. Quy tắc cộng dồn / deduplicate",
     "► KHÔNG cộng dồn — mỗi dòng là một dịch vụ độc lập, màn hình không có ô tổng.\n"
     " ► Chống trùng: Tên dịch vụ / chi phí phải là DUY NHẤT trong toàn danh mục, không phân biệt "
     "dịch vụ đó đang Hoạt động hay đang Khóa. Khi Sửa thì không tính chính dòng đang sửa là trùng.\n"
     " ► Mỗi công ty chỉ có DUY NHẤT một mức ĐM giảm giá cho một dịch vụ, lưu lại nhiều lần cũng "
     "không sinh dòng thừa.\n"
     " ► Bỏ trống hoặc nhập 0 ở ĐM giảm giá nghĩa là “không có định mức”: hệ thống gỡ mức giảm giá "
     "của CÔNG TY ĐANG CHỌN và KHÔNG đụng tới mức của các công ty khác.\n"
     " ► Xóa hẳn một dịch vụ thì mức giảm giá của MỌI công ty cũng bị xóa theo. Còn Khóa thì giữ "
     "nguyên toàn bộ mức giảm giá."),

    ("7. Phân quyền cấp",
     "• Quyền “Quản lý dịch vụ sửa chữa và chi phí khác” — xem danh sách, xem chi tiết, Tạo mới, "
     "Sửa, Xóa, Khóa, Mở khóa.\n"
     " • Quyền “Xem dịch vụ sửa chữa và chi phí khác” — CHỈ xem danh sách, xem chi tiết, tìm kiếm "
     "và Xuất Excel.\n"
     " • Mục menu “Danh mục dịch vụ sửa chữa và chi phí khác” chỉ hiện khi tài khoản có ít nhất "
     "một trong hai quyền trên. Không có quyền nào thì mục menu bị ẩn và vào thẳng đường dẫn cũng "
     "bị chặn.\n"
     " • Mọi thao tác làm thay đổi dữ liệu (Tạo mới, Sửa, Xóa, Khóa, Mở khóa) đều yêu cầu quyền "
     "“Quản lý dịch vụ sửa chữa và chi phí khác”.\n"
     " • Màn hình KHÔNG phân quyền theo cấp công ty / phòng ban / bộ phận: mọi người có quyền đều "
     "nhìn thấy đủ 526 dòng. Công ty đang chọn CHỈ ảnh hưởng tới giá trị cột “ĐM giảm giá”.\n"
     " ⚠️ Quyền được cấp theo vai trò (chức danh), không cấp trực tiếp cho từng nhân viên."),

    ("8. Cách tính các ô thống kê",
     "► Màn hình KHÔNG có khối thống kê đầu trang. Chỉ có dòng đếm ở chân bảng:\n"
     " ► Ô “Hiển thị a–b / N dịch vụ / chi phí”: a và b là số thứ tự dòng đầu và dòng cuối của "
     "trang đang xem; N là TỔNG số dịch vụ khớp bộ lọc (không phải số dòng của trang).\n"
     " ► Các ô phần trăm trên lưới (% Tính giá vốn, % VAT, ĐM giảm giá) hiển thị đúng giá trị đã "
     "lưu, tối đa 2 chữ số thập phân, tự bỏ số 0 thừa (80,00 hiển thị thành “80%”). Không có giá "
     "trị thì hiển thị “—”.\n"
     " ► File Excel xuất ra luôn ghi phần trăm với đúng 2 chữ số thập phân (ví dụ 80,00) — khác "
     "cách rút gọn trên màn hình. Đây là hành vi đúng của hiện trạng, không phải lỗi."),

    ("9. Ghi chú đọc bảng",
     "► Toàn bộ Tạo mới / Sửa / Xem đều mở trong một cửa sổ nhỏ ngay trên trang, không chuyển sang "
     "màn khác.\n"
     " ► Lỗi nhập liệu hiển thị ngay dưới từng ô (viền đỏ + chữ đỏ) kèm thông báo đỏ “Bạn chưa "
     "nhập đầy đủ thông tin” ở góc màn hình. Cửa sổ không bị đóng và dữ liệu đã nhập vẫn còn.\n"
     " ► ⚠️ Ở 3 ô phần trăm, DẤU PHẨY được hiểu là DẤU THẬP PHÂN: nhập “12,5” sẽ lưu là 12,5 phần "
     "trăm, KHÔNG phải 125. Đây là chủ đích, khác với màn Tiền tệ.\n"
     " ► ⚠️ “% Tính giá vốn” KHÔNG bị chặn trần 100 (chỉ cần từ 0 trở lên) — dữ liệu thật đang có "
     "dòng 321%. Riêng “% VAT” và “ĐM giảm giá” bị chặn tối đa 100.\n"
     " ► Nút Xóa không xóa ngay: hệ thống kiểm tra trước xem dịch vụ đã được dùng ở Báo giá hãng "
     "hay Hợp đồng hãng chưa. Nếu đã dùng, hộp thoại đổi thành “Xác nhận khóa” và dịch vụ chỉ bị "
     "chuyển sang trạng thái Khóa chứ không bị xóa.\n"
     " ► Phân trang mặc định 10 dòng/trang, chọn được 5 / 10 / 20 / 50.\n"
     " ► Các trường hợp ghi “gọi thẳng chức năng, bỏ qua giao diện” là kiểm thử bảo mật, cần công "
     "cụ kiểm thử API (Postman hoặc tương đương) — dành cho tester kỹ thuật."),
]

# ============================================================================
# SECTION PHAN QUYEN
# ============================================================================
ROLE_TCS = [
    ("00", "Không có quyền nào của màn — menu và màn hình bị chặn", "P0",
     "Tài khoản A đăng nhập được nhưng KHÔNG có quyền “Quản lý dịch vụ sửa chữa và chi phí khác” "
     "và cũng KHÔNG có quyền “Xem dịch vụ sửa chữa và chi phí khác”",
     "1. Đăng nhập tài khoản A, vào phân hệ CSKH\n2. Mở nhóm menu “Danh mục - Dịch vụ”\n"
     "3. Dán thẳng đường dẫn màn hình vào trình duyệt",
     "Tài khoản A: không có quyền nào",
     "- Nhóm “Danh mục - Dịch vụ” KHÔNG có mục “Danh mục dịch vụ sửa chữa và chi phí khác”\n"
     "- Vào thẳng đường dẫn: hệ thống chặn, bảng không hiển thị dòng dữ liệu nào"),

    ("01", "Chỉ có quyền “Xem dịch vụ sửa chữa và chi phí khác”", "P0",
     "Tài khoản B chỉ được cấp quyền “Xem dịch vụ sửa chữa và chi phí khác”",
     "1. Đăng nhập B, vào màn hình\n2. Quan sát thanh công cụ phía trên bảng\n"
     "3. Quan sát cột Hành động và cột Trạng thái",
     "Tài khoản B: chỉ quyền xem",
     "- Vào được màn hình, bảng hiển thị đủ 526 dòng, bộ lọc dùng bình thường\n"
     "- Thanh công cụ CHỈ có nút “Xuất Excel”, KHÔNG có nút “Tạo mới”\n"
     "- Cột Hành động CHỈ có nút Xem; không có nút Sửa, không có nút Xóa\n"
     "- Cột Trạng thái chỉ có nhãn trạng thái, KHÔNG có nút ổ khóa"),

    ("02", "Có quyền “Quản lý dịch vụ sửa chữa và chi phí khác”", "P0",
     "Tài khoản C được cấp quyền “Quản lý dịch vụ sửa chữa và chi phí khác”",
     "1. Đăng nhập C, vào màn hình\n2. Quan sát thanh công cụ, cột Hành động, cột Trạng thái",
     "Tài khoản C: quyền quản lý",
     "- Thanh công cụ có đủ nút “Tạo mới” và “Xuất Excel”\n"
     "- Cột Hành động có đủ 3 nút: Xem, Sửa, Xóa\n"
     "- Cột Trạng thái có nút ổ khóa để Khóa / Mở khóa"),

    ("03", "Chỉ quyền Xem — bấm nút Xem vẫn mở được chi tiết", "P0",
     "Tài khoản B (chỉ quyền xem); danh sách có ít nhất 1 dịch vụ",
     "1. Đăng nhập B\n2. Bấm nút Xem ở một dòng bất kỳ",
     "Dịch vụ “Dịch vụ đánh bóng xy lanh”",
     "- Cửa sổ “Xem dịch vụ / chi phí” mở ra bình thường\n"
     "- Mọi ô đều mờ, không nhập được; chân cửa sổ chỉ có nút “Đóng”"),

    ("04", "Chỉ quyền Xem — gọi thẳng chức năng Tạo mới, bỏ qua giao diện", "P0",
     "Tài khoản B (chỉ quyền xem), đã đăng nhập hợp lệ",
     "1. Dùng công cụ kiểm thử API gọi thẳng chức năng Tạo mới với dữ liệu hợp lệ\n"
     "2. Vào lại danh sách kiểm tra",
     "Tên = “TC bypass”, % Tính giá vốn = 10, % VAT = 8",
     "- Hệ thống từ chối, báo không có quyền\n- KHÔNG có dịch vụ mới nào được tạo ra"),

    ("05", "Chỉ quyền Xem — gọi thẳng chức năng Sửa, bỏ qua giao diện", "P0",
     "Tài khoản B (chỉ quyền xem); tồn tại dịch vụ X đang Hoạt động",
     "1. Dùng công cụ kiểm thử API gọi thẳng chức năng Sửa dịch vụ X\n"
     "2. Mở lại dịch vụ X trên màn hình để đối chiếu",
     "Tài khoản B, dịch vụ X",
     "- Hệ thống từ chối, báo không có quyền\n"
     "- Dịch vụ X giữ nguyên mọi thông tin, cột “Cập nhật” không đổi"),

    ("06", "Chỉ quyền Xem — gọi thẳng chức năng Xóa / Khóa / Mở khóa", "P0",
     "Tài khoản B (chỉ quyền xem); dịch vụ X đang Hoạt động",
     "1. Dùng công cụ kiểm thử API gọi lần lượt 3 chức năng: Xóa, Khóa, Mở khóa cho dịch vụ X",
     "3 lần gọi, cùng tài khoản B",
     "- CẢ 3 lần đều bị hệ thống từ chối, báo không có quyền\n"
     "- Dịch vụ X vẫn còn và vẫn ở trạng thái Hoạt động"),

    ("07", "Không có quyền nào — gọi thẳng chức năng xem danh sách và xem chi tiết", "P0",
     "Tài khoản A (không có quyền nào của màn), đã đăng nhập hợp lệ",
     "1. Dùng công cụ kiểm thử API gọi chức năng lấy danh sách\n"
     "2. Gọi chức năng xem chi tiết một dịch vụ",
     "Tài khoản A",
     "- Cả 2 lần đều bị từ chối, báo không có quyền\n"
     "- Kết quả trả về KHÔNG để lộ bất kỳ thông tin nào của danh mục"),

    ("08", "ĐM giảm giá hiển thị theo công ty đang chọn của người đăng nhập", "P0",
     "Dịch vụ “Dịch vụ đánh bóng xy lanh” đã khai báo ĐM giảm giá: công ty 1 là 5%, công ty 4 là 12%. "
     "Tài khoản C thuộc công ty 1, tài khoản D thuộc công ty 4, cả hai đều có quyền quản lý",
     "1. Đăng nhập C, xem cột “ĐM giảm giá” của dòng đó\n"
     "2. Đăng nhập D, xem lại đúng dòng đó",
     "C ở công ty 1; D ở công ty 4",
     "- Tài khoản C thấy “5%”\n- Tài khoản D thấy “12%”\n"
     "- Cùng một dịch vụ nhưng 2 người thấy 2 giá trị khác nhau — đây là hành vi đúng, không phải lỗi"),

    ("09", "Công ty chưa khai báo ĐM giảm giá — hiển thị gạch ngang", "P1",
     "Dịch vụ “Hướng dẫn, bàn giao máy hàn rút tồn” chưa được khai báo ĐM giảm giá cho công ty của "
     "người đăng nhập",
     "1. Đăng nhập, xem cột “ĐM giảm giá” của dòng đó",
     "Chưa khai báo định mức",
     "- Cột “ĐM giảm giá” hiển thị “—”\n- KHÔNG hiển thị 0% và không để trống"),
]

# ============================================================================
# SECTIONS NGHIEP VU
# ============================================================================
SECTIONS = [
 ("I", "HIỂN THỊ TRANG & TRUY CẬP", [
  ("001", "Truy cập màn hình từ menu", "P0",
   "Tài khoản C có quyền quản lý danh mục",
   "1. Vào phân hệ CSKH\n2. Bấm nhóm menu “Danh mục - Dịch vụ”\n"
   "3. Bấm “Danh mục dịch vụ sửa chữa và chi phí khác”",
   "Tài khoản C",
   "- Mở đúng màn hình danh mục\n"
   "- Tiêu đề trên thanh trên cùng ghi “Danh mục dịch vụ sửa chữa và chi phí khác”"),

  ("002", "Bố cục trang khi vào lần đầu", "P0",
   "Tài khoản C vào màn hình lần đầu trong ngày (chưa từng đặt bộ lọc)",
   "1. Vào màn hình\n2. Quan sát bố cục",
   "—",
   "- Trang có 2 khối: panel “Bộ lọc dịch vụ sửa chữa và chi phí khác” ở trên và bảng danh sách ở dưới\n"
   "- Bộ lọc nâng cao đang THU GỌN, chỉ hiện ô tìm nhanh cùng 2 nút “Tìm kiếm” và “Làm mới”\n"
   "- KHÔNG có khối thống kê ở đầu trang"),

  ("003", "Hiển thị đủ và đúng thứ tự các cột", "P0",
   "Danh sách có ít nhất 1 dịch vụ",
   "1. Vào danh sách\n2. Đọc tiêu đề từng cột từ trái sang phải",
   "—",
   "- Đúng 9 cột theo thứ tự: STT | Tên dịch vụ / chi phí | Phân loại | ĐM giảm giá | % Tính giá vốn | "
   "% VAT | Trạng thái | Cập nhật | Hành động\n"
   "- Cột STT và cột Tên đứng yên (dính bên trái) khi cuộn ngang"),

  ("004", "Biểu tượng sắp xếp chỉ có ở đúng 6 cột", "P1",
   "Đang ở màn danh sách",
   "1. Quan sát biểu tượng 2 mũi tên trên từng tiêu đề cột",
   "—",
   "- CÓ biểu tượng sắp xếp ở: Tên dịch vụ / chi phí, ĐM giảm giá, % Tính giá vốn, % VAT, "
   "Trạng thái, Cập nhật\n"
   "- KHÔNG có ở: STT, Phân loại, Hành động"),

  ("005", "Nhãn Trạng thái Hoạt động / Khóa", "P0",
   "Danh sách có dịch vụ đang Hoạt động và dịch vụ đang Khóa (ví dụ “test01111” đang Khóa)",
   "1. Quan sát cột Trạng thái của 2 dòng đó",
   "1 dòng Hoạt động, 1 dòng Khóa",
   "- Dòng Hoạt động: nhãn xanh “Hoạt động” kèm biểu tượng dấu tích\n"
   "- Dòng Khóa: nhãn đỏ “Khóa” kèm biểu tượng ổ khóa"),

  ("006", "Nhãn Phân loại", "P0",
   "Danh sách có cả dịch vụ có tính doanh thu và khoản chi phí khác",
   "1. Quan sát cột Phân loại",
   "“Trang test” có tính doanh thu; “test01111” là chi phí khác",
   "- Dịch vụ có tính doanh thu: nhãn xanh “Có tính doanh thu” kèm biểu tượng biểu đồ\n"
   "- Chi phí khác: nhãn đỏ “Chi phí khác” kèm biểu tượng nhãn giá"),

  ("007", "Cột Cập nhật hiển thị thời điểm và người sửa", "P1",
   "Dịch vụ X được nhân viên “10410008 - Cao Đình Hòe” sửa lần cuối lúc 16/07/2026 15:12",
   "1. Quan sát cột Cập nhật của dòng X",
   "Sửa lần cuối 16/07/2026 15:12 bởi Cao Đình Hòe",
   "- Dòng trên: “16/07/2026 15:12”\n- Dòng dưới: “bởi 10410008 - Cao Đình Hòe” (tên in đậm)"),

  ("008", "Dịch vụ chưa từng sửa — cột Cập nhật lấy người tạo", "P1",
   "Dịch vụ Y do “DNS01 - DNS Admin” tạo và CHƯA từng được sửa lần nào",
   "1. Quan sát cột Cập nhật của dòng Y",
   "Chưa từng sửa",
   "- Vẫn hiển thị “bởi DNS01 - DNS Admin” (lấy theo người tạo), không được để trống"),

  ("009", "Định dạng phần trăm — tự bỏ số 0 thừa", "P1",
   "Dịch vụ có % Tính giá vốn là 80,00 và % VAT là 8,00",
   "1. Quan sát cột “% Tính giá vốn” và “% VAT”",
   "80,00 và 8,00",
   "- Hiển thị “80%” và “8%”, KHÔNG hiển thị “80,00%”"),

  ("010", "Giá trị trống hiển thị gạch ngang", "P1",
   "Dịch vụ chưa khai báo ĐM giảm giá",
   "1. Quan sát cột “ĐM giảm giá”",
   "Chưa khai báo",
   "- Hiển thị “—”, không hiển thị 0% và không để trống"),

  ("011", "Danh sách rỗng khi bộ lọc không khớp", "P1",
   "Danh sách đang có dữ liệu",
   "1. Nhập “ZZZZZ999” vào ô tìm nhanh\n2. Bấm “Tìm kiếm”",
   "Từ khóa: ZZZZZ999",
   "- Bảng hiển thị dòng chữ “Không có dữ liệu phù hợp bộ lọc.”\n- Không báo lỗi, không treo trang"),

  ("012", "Ghi nhớ bộ lọc trong 10 phút", "P2",
   "Đã lọc Trạng thái = “Khóa”, sau đó chuyển sang màn khác",
   "1. Quay lại màn danh mục trong vòng 10 phút\n2. Mở bộ lọc nâng cao",
   "Quay lại sau 2 phút",
   "- Ô “Trạng thái” vẫn đang chọn “Khóa”, danh sách vẫn lọc theo điều kiện cũ\n"
   "- Danh sách chỉ nạp đúng 1 lần, không bị nháy nạp lại 2 lần"),

  ("013", "Bộ lọc hết hạn sau 10 phút", "P2",
   "Đã lọc Trạng thái = “Khóa”, rời màn quá 10 phút",
   "1. Quay lại màn danh mục sau 10 phút",
   "Quay lại sau 15 phút",
   "- Bộ lọc trở về trắng, danh sách hiển thị lại toàn bộ 526 dòng"),
 ]),

 ("II", "BỘ LỌC & TÌM KIẾM", [
  ("001", "Tìm nhanh theo tên dịch vụ", "P0",
   "Tồn tại dịch vụ tên “Dịch vụ đánh bóng xy lanh”",
   "1. Nhập “đánh bóng” vào ô tìm nhanh\n2. Bấm nút “Tìm kiếm”",
   "Từ khóa: đánh bóng",
   "- Danh sách chỉ còn các dịch vụ có tên chứa “đánh bóng”\n- Danh sách nhảy về trang 1"),

  ("002", "Ô tìm nhanh KHÔNG tự lọc khi đang gõ", "P0",
   "Đang ở màn danh sách 526 dòng",
   "1. Gõ “đánh bóng” vào ô tìm nhanh\n2. KHÔNG bấm gì thêm, chờ 5 giây\n3. Quan sát bảng",
   "Đã gõ nhưng chưa bấm Tìm kiếm",
   "- Bảng GIỮ NGUYÊN 526 dòng, chưa lọc gì\n"
   "- Phải bấm “Tìm kiếm” (hoặc nhấn Enter) thì danh sách mới lọc lại"),

  ("003", "Mở và đóng bộ lọc nâng cao", "P1",
   "Bộ lọc đang thu gọn",
   "1. Bấm “Tìm kiếm nâng cao”\n2. Bấm lại nút đó (chữ đã đổi)",
   "—",
   "- Lần 1: hiện 4 ô lọc (Tên dịch vụ / chi phí, Phân loại, Trạng thái, Người cập nhật), nút đổi "
   "chữ thành “Ẩn tìm kiếm nâng cao”\n- Lần 2: 4 ô ẩn đi, nút đổi lại thành “Tìm kiếm nâng cao”"),

  ("004", "Lọc theo ô Tên dịch vụ / chi phí", "P0",
   "Tồn tại dịch vụ có tên chứa “Thiết kế”",
   "1. Mở bộ lọc nâng cao\n2. Nhập “Thiết kế” vào ô Tên dịch vụ / chi phí\n3. Bấm “Tìm kiếm”",
   "Tên: Thiết kế",
   "- Danh sách chỉ còn dịch vụ có tên chứa “Thiết kế”"),

  ("005", "Lọc Phân loại = Dịch vụ có tính doanh thu", "P0",
   "Danh sách có cả 2 phân loại",
   "1. Mở bộ lọc nâng cao\n2. Chọn Phân loại = “Dịch vụ có tính doanh thu”",
   "Phân loại: Có tính doanh thu",
   "- Danh sách CHỈ còn dòng có nhãn “Có tính doanh thu”\n"
   "- Không còn dòng nào có nhãn “Chi phí khác”"),

  ("006", "Lọc Phân loại = Chi phí khác", "P0",
   "Danh sách có dịch vụ thuộc nhóm Chi phí khác (ví dụ “test01111”)",
   "1. Chọn Phân loại = “Chi phí khác”\n2. Quan sát kết quả",
   "Phân loại: Chi phí khác",
   "- Danh sách CHỈ còn dòng có nhãn “Chi phí khác”\n"
   "- ⚠️ Điểm dễ lỗi: tuyệt đối KHÔNG được trả về toàn bộ danh sách như khi chưa chọn bộ lọc"),

  ("007", "Lọc Trạng thái = Khóa", "P0",
   "Danh sách có dịch vụ đang Khóa (ví dụ “test01111”)",
   "1. Chọn Trạng thái = “Khóa”",
   "Trạng thái: Khóa",
   "- Danh sách CHỈ còn dòng có nhãn “Khóa”\n"
   "- ⚠️ Điểm dễ lỗi: không được coi lựa chọn này là “chưa chọn gì” rồi trả về cả danh sách"),

  ("008", "Lọc Trạng thái = Hoạt động", "P0",
   "Danh sách có cả dòng Hoạt động và dòng Khóa",
   "1. Chọn Trạng thái = “Hoạt động”",
   "Trạng thái: Hoạt động",
   "- Danh sách CHỈ còn dòng có nhãn “Hoạt động”"),

  ("009", "Lọc theo Người cập nhật", "P1",
   "Nhân viên “10410008 - Cao Đình Hòe” đã sửa một số dịch vụ",
   "1. Chọn Người cập nhật = “10410008 - Cao Đình Hòe”",
   "Người cập nhật: Cao Đình Hòe",
   "- Danh sách chỉ còn các dịch vụ do nhân viên đó sửa lần cuối"),

  ("010", "Lọc Người cập nhật bắt cả dịch vụ chưa từng sửa", "P1",
   "Nhân viên “DNS01 - DNS Admin” TẠO dịch vụ Z và Z chưa từng được sửa lần nào",
   "1. Chọn Người cập nhật = “DNS01 - DNS Admin”\n2. Tìm dịch vụ Z trong kết quả",
   "Dịch vụ Z chưa từng sửa",
   "- Dịch vụ Z VẪN xuất hiện trong kết quả: dịch vụ chưa từng sửa thì tính theo người tạo"),

  ("011", "Kết hợp nhiều điều kiện lọc", "P0",
   "Dữ liệu đa dạng về trạng thái, phân loại và tên",
   "1. Chọn Trạng thái = Hoạt động\n2. Chọn Phân loại = Có tính doanh thu\n"
   "3. Nhập “Thiết kế” vào ô Tên dịch vụ / chi phí\n4. Bấm “Tìm kiếm”",
   "3 điều kiện cùng lúc",
   "- Kết quả thỏa ĐỒNG THỜI cả 3 điều kiện, không phải thỏa 1 trong 3"),

  ("012", "Đổi ô lọc nâng cao là danh sách tự lọc lại", "P1",
   "Đang ở trang 3 của danh sách",
   "1. Mở bộ lọc nâng cao, chọn Trạng thái = “Khóa”\n2. KHÔNG bấm “Tìm kiếm”",
   "Trạng thái: Khóa",
   "- Danh sách tự lọc lại ngay khi chọn, không cần bấm “Tìm kiếm”\n- Danh sách nhảy về trang 1"),

  ("013", "Nút “Làm mới” xóa lọc và nạp lại danh sách", "P0",
   "Đang lọc Trạng thái = Khóa kèm từ khóa “test”, danh sách chỉ còn vài dòng",
   "1. Bấm nút “Làm mới”\n2. Quan sát các ô lọc và bảng",
   "—",
   "- Ô tìm nhanh và cả 4 ô lọc nâng cao đều trở về trắng\n"
   "- Bảng NẠP LẠI đủ 526 dòng và về trang 1\n"
   "- ⚠️ Không được chỉ xóa ô lọc mà vẫn giữ nguyên kết quả cũ trên bảng"),

  ("014", "Làm mới khi chỉ dùng ô tìm nhanh", "P1",
   "Đang lọc bằng ô tìm nhanh với từ khóa “test”",
   "1. Bấm “Làm mới”",
   "Chỉ có từ khóa tìm nhanh",
   "- Bảng vẫn phải nạp lại toàn bộ danh sách, không được đứng yên"),

  ("015", "Tìm với ký tự đặc biệt", "P2",
   "Danh sách có dữ liệu; không có dịch vụ nào tên chứa ký tự “%”",
   "1. Nhập “%” vào ô tìm nhanh\n2. Bấm “Tìm kiếm”",
   "Từ khóa: %",
   "- Trả về danh sách RỖNG\n"
   "- ⚠️ Không được hiểu “%” là ký tự đại diện rồi trả về toàn bộ danh sách"),

  ("016", "Tìm với toàn khoảng trắng", "P2",
   "Danh sách có dữ liệu",
   "1. Nhập vài dấu cách vào ô tìm nhanh\n2. Bấm “Tìm kiếm”",
   "Từ khóa: chỉ có dấu cách",
   "- Danh sách trả về bình thường như khi không lọc, không báo lỗi"),
 ]),

 ("III", "DANH SÁCH, SẮP XẾP & PHÂN TRANG", [
  ("001", "Thứ tự mặc định khi vào màn", "P0",
   "Danh sách 526 dòng, có dịch vụ vừa được tạo hôm nay",
   "1. Vào màn hình lần đầu\n2. Quan sát thứ tự các dòng",
   "Chưa bấm sắp xếp cột nào",
   "- Sắp xếp theo ngày tạo giảm dần: dịch vụ mới tạo nhất nằm ở đầu danh sách"),

  ("002", "Sắp xếp theo Tên dịch vụ / chi phí", "P0",
   "Danh sách có từ 20 dòng trở lên",
   "1. Bấm tiêu đề cột “Tên dịch vụ / chi phí”\n2. Bấm lần nữa",
   "Sắp xếp theo cột Tên",
   "- Lần 1: sắp tăng dần A → Z\n- Lần 2: sắp giảm dần Z → A\n"
   "- Biểu tượng mũi tên trên tiêu đề đổi chiều tương ứng"),

  ("003", "Sắp xếp theo ĐM giảm giá", "P0",
   "Có dịch vụ ĐM giảm giá 5%, 32%, 95% và có dịch vụ chưa khai báo",
   "1. Bấm tiêu đề cột “ĐM giảm giá”\n2. Bấm lần nữa",
   "Sắp xếp theo cột ĐM giảm giá",
   "- Sắp đúng theo giá trị số của công ty đang chọn: 5 → 32 → 95\n"
   "- Các dòng chưa khai báo gom về một đầu danh sách, không gây lỗi"),

  ("004", "Sắp xếp theo % Tính giá vốn", "P1",
   "Có dịch vụ 0%, 1%, 60%, 80% và 321%",
   "1. Bấm tiêu đề cột “% Tính giá vốn” để sắp tăng dần",
   "Sắp xếp theo cột % Tính giá vốn",
   "- Thứ tự đúng: 0 → 1 → 60 → 80 → 321\n"
   "- ⚠️ Điểm dễ lỗi: 321 phải đứng SAU 80 (so sánh theo số, không phải theo chữ)"),

  ("005", "Sắp xếp theo % VAT", "P1",
   "Có dịch vụ VAT 0%, 8%, 21% và 100%",
   "1. Bấm tiêu đề cột “% VAT”",
   "Sắp xếp theo cột % VAT",
   "- Thứ tự tăng dần đúng theo giá trị số"),

  ("006", "Sắp xếp theo Trạng thái", "P1",
   "Có cả dòng Hoạt động và dòng Khóa",
   "1. Bấm tiêu đề cột “Trạng thái”",
   "Sắp xếp theo cột Trạng thái",
   "- Các dòng cùng trạng thái được gom lại với nhau, không báo lỗi"),

  ("007", "Sắp xếp theo Cập nhật", "P1",
   "Có dòng vừa sửa gần đây và dòng sửa đã lâu",
   "1. Bấm tiêu đề cột “Cập nhật” 2 lần",
   "Sắp xếp theo cột Cập nhật",
   "- Sắp đúng theo thời điểm cập nhật ở cả 2 chiều"),

  ("008", "Thứ tự sắp xếp giữ nguyên khi chuyển trang", "P1",
   "Đang sắp xếp theo Tên tăng dần",
   "1. Bấm sang trang 2\n2. Quan sát thứ tự và biểu tượng trên tiêu đề cột",
   "Đang sắp theo Tên, sang trang 2",
   "- Trang 2 tiếp nối đúng thứ tự A → Z của trang 1\n"
   "- Biểu tượng sắp xếp trên cột Tên vẫn giữ nguyên"),

  ("009", "Bấm sắp xếp thì quay về trang 1", "P2",
   "Đang ở trang 3",
   "1. Bấm tiêu đề cột “Tên dịch vụ / chi phí”",
   "Đang ở trang 3",
   "- Danh sách nhảy về trang 1 với thứ tự mới"),

  ("010", "Phân trang mặc định 10 dòng", "P0",
   "Danh sách có 526 dòng",
   "1. Vào màn hình\n2. Đọc dòng đếm ở chân bảng",
   "526 dòng",
   "- Bảng hiện đúng 10 dòng\n- Chân bảng ghi “Hiển thị 1–10 / 526 dịch vụ / chi phí”\n"
   "- Có 53 trang"),

  ("011", "Chuyển trang", "P0",
   "Danh sách 526 dòng, 10 dòng mỗi trang",
   "1. Bấm sang trang 2\n2. Bấm sang trang 3\n3. Bấm nút về trang cuối",
   "Trang 2, 3 và trang cuối",
   "- Mỗi trang hiện đúng 10 dòng với dữ liệu KHÁC nhau, dòng đếm đổi thành 11–20 rồi 21–30\n"
   "- Trang cuối hiển thị 6 dòng\n- Bộ lọc đang áp dụng KHÔNG bị xóa khi chuyển trang"),

  ("012", "Đổi số dòng mỗi trang", "P0",
   "Danh sách 526 dòng",
   "1. Chọn “Số dòng/trang” = 50\n2. Quan sát bảng và số trang",
   "50 dòng mỗi trang",
   "- Bảng hiện 50 dòng, chân bảng ghi “Hiển thị 1–50 / 526”, tổng số trang còn 11\n"
   "- Danh sách nhảy về trang 1"),

  ("013", "Danh sách lựa chọn số dòng mỗi trang", "P2",
   "Đang ở màn danh sách",
   "1. Mở ô chọn “Số dòng/trang”",
   "—",
   "- Có đúng 4 lựa chọn: 5, 10, 20, 50; đang chọn sẵn 10"),

  ("014", "Một thao tác chỉ nạp danh sách một lần", "P1",
   "Đang ở màn danh sách",
   "1. Bấm sang trang 2\n2. Quan sát bảng có bị nháy nạp lại 2 lần không",
   "1 lần bấm chuyển trang",
   "- Danh sách chỉ nạp đúng 1 lần, không nháy 2 lần"),

  ("015", "Số thứ tự liên tục qua các trang", "P1",
   "Đang để 10 dòng mỗi trang",
   "1. Xem cột STT ở trang 1\n2. Sang trang 2 xem cột STT",
   "Trang 1 rồi trang 2",
   "- Trang 1: STT từ 1 đến 10\n- Trang 2: STT từ 11 đến 20, không quay lại đếm từ 1"),
 ]),

 ("IV", "TẠO MỚI / SỬA / XEM", [
  ("001", "Mở cửa sổ Tạo mới", "P0",
   "Tài khoản C có quyền quản lý",
   "1. Bấm nút “Tạo mới”\n2. Quan sát cửa sổ mở ra",
   "—",
   "- Cửa sổ tiêu đề “Thêm dịch vụ / chi phí” mở ra ngay trên trang\n"
   "- Có 4 ô nhập: Tên dịch vụ / chi phí (bắt buộc), % Tính giá vốn (bắt buộc), % VAT (bắt buộc), "
   "ĐM giảm giá (%)\n- Có ô tích “Dịch vụ có tính doanh thu” ĐANG ĐƯỢC TÍCH SẴN\n"
   "- KHÔNG có ô “Trạng thái”\n- Chân cửa sổ có 3 nút: “Lưu”, “Lưu & Tiếp tục”, “Đóng”"),

  ("002", "Giá trị điền sẵn khi mở cửa sổ Tạo mới", "P0",
   "Vừa bấm “Tạo mới”",
   "1. Quan sát giá trị từng ô",
   "—",
   "- Tên, % Tính giá vốn, % VAT và ĐM giảm giá đều TRỐNG (chỉ có chữ gợi ý mờ)\n"
   "- Ô “Dịch vụ có tính doanh thu” ĐÃ ĐƯỢC TÍCH\n- Chưa có ô nào báo lỗi đỏ"),

  ("003", "Tạo mới dịch vụ hợp lệ đầy đủ", "P0",
   "Tài khoản C có quyền quản lý; chưa tồn tại dịch vụ tên “TC Dịch vụ kiểm thử 01”",
   "1. Bấm “Tạo mới”\n2. Nhập Tên = “TC Dịch vụ kiểm thử 01”\n3. % Tính giá vốn = 70\n"
   "4. % VAT = 8\n5. ĐM giảm giá = 5\n6. Giữ tích “Dịch vụ có tính doanh thu”\n7. Bấm “Lưu”",
   "Tên: TC Dịch vụ kiểm thử 01; giá vốn 70; VAT 8; giảm giá 5",
   "- Thông báo xanh “Thêm mới thành công”\n- Cửa sổ đóng lại, danh sách nạp lại, dịch vụ mới nằm ĐẦU danh sách\n"
   "- Dòng mới: Phân loại “Có tính doanh thu”, ĐM giảm giá 5%, % Tính giá vốn 70%, % VAT 8%, "
   "Trạng thái “Hoạt động”, cột Cập nhật ghi hôm nay bởi chính tài khoản C"),

  ("004", "Tạo mới bỏ trống ĐM giảm giá", "P0",
   "Chưa tồn tại dịch vụ tên “TC Dịch vụ kiểm thử 02”",
   "1. Tạo mới, nhập Tên, % Tính giá vốn = 50, % VAT = 10\n2. ĐỂ TRỐNG ô ĐM giảm giá\n3. Bấm “Lưu”",
   "ĐM giảm giá: để trống",
   "- Lưu thành công, KHÔNG báo lỗi ở ô ĐM giảm giá\n"
   "- Cột “ĐM giảm giá” của dòng mới hiển thị “—”"),

  ("005", "Tạo mới bỏ tích “Dịch vụ có tính doanh thu”", "P0",
   "Chưa tồn tại dịch vụ tên “TC Chi phí khác 01”",
   "1. Tạo mới, nhập Tên = “TC Chi phí khác 01”, % Tính giá vốn = 0, % VAT = 0\n"
   "2. BỎ TÍCH ô “Dịch vụ có tính doanh thu”\n3. Bấm “Lưu”",
   "Bỏ tích ô có tính doanh thu",
   "- Lưu thành công\n- Dòng mới hiển thị nhãn đỏ “Chi phí khác” ở cột Phân loại"),

  ("006", "Trạng thái sau khi tạo luôn là Hoạt động", "P0",
   "Vừa tạo mới một dịch vụ",
   "1. Xem cột Trạng thái của dòng vừa tạo",
   "—",
   "- Trạng thái là “Hoạt động”; người dùng không chọn được trạng thái lúc tạo mới"),

  ("007", "Nút “Lưu & Tiếp tục” giữ cửa sổ mở", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập đủ thông tin hợp lệ cho dịch vụ “TC Liên tục 01”\n2. Bấm “Lưu & Tiếp tục”\n"
   "3. Quan sát cửa sổ",
   "—",
   "- Thông báo “Thêm mới thành công”\n- Cửa sổ VẪN MỞ, tiêu đề vẫn là “Thêm dịch vụ / chi phí”\n"
   "- Mọi ô đã được xóa trắng, ô tích trở về trạng thái tích sẵn, nhập tiếp được ngay"),

  ("008", "Bấm Lưu 2 lần liên tiếp không tạo 2 dịch vụ", "P0",
   "Đang mở cửa sổ Tạo mới với dữ liệu hợp lệ",
   "1. Bấm “Lưu” rồi bấm ngay lần thứ 2 thật nhanh\n2. Đếm số dòng trùng tên trong danh sách",
   "Bấm nhanh 2 lần",
   "- Chỉ tạo ĐÚNG 1 dịch vụ; lần bấm thứ 2 không có tác dụng"),

  ("009", "Mở cửa sổ Sửa — dữ liệu nạp đúng", "P0",
   "Tồn tại dịch vụ “Dịch vụ đánh bóng xy lanh”: giá vốn 60%, VAT 8%, ĐM giảm giá 5%, có tính "
   "doanh thu, đang Hoạt động",
   "1. Bấm nút Sửa (biểu tượng bút chì) ở dòng đó\n2. Đối chiếu từng ô với dữ liệu trên lưới",
   "Dịch vụ đánh bóng xy lanh",
   "- Cửa sổ tiêu đề “Sửa dịch vụ / chi phí”\n"
   "- Trên tiêu đề có dòng “Cập nhật: 16/07/2026 15:12” và “Bởi: 10410008 - Cao Đình Hòe”\n"
   "- Tên, % Tính giá vốn = 60, % VAT = 8, ĐM giảm giá = 5, Trạng thái = “Hoạt động”, ô tích ĐÃ TÍCH"),

  ("010", "Cửa sổ Sửa CÓ ô Trạng thái", "P0",
   "Đang mở cửa sổ Sửa",
   "1. Quan sát khu vực bên phải ô ĐM giảm giá và chân cửa sổ",
   "—",
   "- CÓ ô chọn “Trạng thái” với 2 lựa chọn “Hoạt động” và “Khóa”\n"
   "- Chân cửa sổ chỉ có “Lưu” và “Đóng”, KHÔNG có “Lưu & Tiếp tục”"),

  ("011", "Sửa và lưu thành công", "P0",
   "Dịch vụ “TC Dịch vụ kiểm thử 01” đang Hoạt động",
   "1. Bấm Sửa\n2. Đổi Tên thành “TC Dịch vụ kiểm thử 01 - đã sửa”\n3. Đổi % VAT = 10\n4. Bấm “Lưu”",
   "Đổi tên và đổi VAT thành 10",
   "- Thông báo “Cập nhật thành công”, cửa sổ đóng lại\n"
   "- Dòng trên lưới đổi tên và % VAT thành 10%\n"
   "- Cột “Cập nhật” đổi sang thời điểm hiện tại và tên người đang đăng nhập"),

  ("012", "Sửa để bỏ trống ĐM giảm giá", "P0",
   "Dịch vụ X đang có ĐM giảm giá 5% cho công ty đang chọn",
   "1. Bấm Sửa X\n2. XÓA TRẮNG ô “ĐM giảm giá”\n3. Bấm “Lưu”\n4. Xem lại trên lưới",
   "Xóa trắng ĐM giảm giá",
   "- Lưu thành công\n- Cột “ĐM giảm giá” của X đổi thành “—”\n"
   "- Mức giảm giá của các công ty KHÁC không bị ảnh hưởng"),

  ("013", "Sửa nhập ĐM giảm giá = 0", "P1",
   "Dịch vụ X đang có ĐM giảm giá 5%",
   "1. Sửa X, nhập ĐM giảm giá = 0\n2. Bấm “Lưu”",
   "ĐM giảm giá: 0",
   "- Lưu thành công\n- Cột “ĐM giảm giá” hiển thị “—”, KHÔNG hiển thị 0% "
   "(nhập 0 được hiểu là không có định mức)"),

  ("014", "Sửa để chuyển sang Khóa ngay trong cửa sổ", "P0",
   "Dịch vụ X đang Hoạt động",
   "1. Bấm Sửa X\n2. Đổi ô Trạng thái = “Khóa”\n3. Bấm “Lưu”",
   "Trạng thái: Khóa",
   "- Lưu thành công\n- Dòng X đổi sang nhãn “Khóa”\n"
   "- Nút Sửa và nút Xóa của dòng X chuyển sang mờ; nút ổ khóa đổi thành biểu tượng mở khóa"),

  ("015", "Sửa mà giữ nguyên tên thì không báo trùng", "P0",
   "Dịch vụ “TC Dịch vụ kiểm thử 02” đang tồn tại",
   "1. Bấm Sửa, GIỮ NGUYÊN tên, chỉ đổi % VAT = 5\n2. Bấm “Lưu”",
   "Không đổi tên",
   "- Lưu thành công, KHÔNG báo lỗi “Đã tồn tại trên hệ thống”"),

  ("016", "Mở cửa sổ Xem — mọi ô chỉ đọc", "P0",
   "Tồn tại dịch vụ bất kỳ",
   "1. Bấm nút Xem (biểu tượng con mắt)\n2. Thử gõ vào từng ô",
   "—",
   "- Cửa sổ tiêu đề “Xem dịch vụ / chi phí”\n"
   "- 4 ô nhập, ô Trạng thái và ô tích đều mờ, không gõ và không chọn được\n"
   "- Chân cửa sổ CHỈ có nút “Đóng”, không có nút “Lưu”"),

  ("017", "Bấm Xem trên dòng đang Khóa", "P1",
   "Dịch vụ “test01111” đang ở trạng thái Khóa",
   "1. Bấm nút Xem ở dòng đó",
   "Dòng đang Khóa",
   "- Vẫn mở được cửa sổ Xem với đầy đủ thông tin, ô Trạng thái hiển thị “Khóa”"),

  ("018", "Nút Sửa bị mờ khi dòng đang Khóa", "P0",
   "Dịch vụ “test01111” đang Khóa; tài khoản có quyền quản lý",
   "1. Rê chuột lên nút Sửa của dòng đó\n2. Thử bấm",
   "Dòng đang Khóa",
   "- Nút Sửa hiển thị nhưng MỜ, không bấm được\n"
   "- Chú thích khi rê chuột: “Chi phí đang bị khóa, hãy mở khóa trước khi sửa”"),

  ("019", "Đóng cửa sổ thì xóa dữ liệu nhập dở", "P1",
   "Đang mở cửa sổ Tạo mới và đã nhập vài ô",
   "1. Bấm “Đóng”\n2. Bấm lại “Tạo mới”",
   "—",
   "- Cửa sổ mở lại ở trạng thái trắng hoàn toàn, không còn dữ liệu của lần trước"),

  ("020", "Đóng cửa sổ bằng dấu X ở góc", "P2",
   "Đang mở cửa sổ Sửa và đã đổi vài ô",
   "1. Bấm dấu “×” ở góc phải trên cửa sổ",
   "—",
   "- Cửa sổ đóng, thay đổi KHÔNG được lưu, danh sách giữ nguyên"),

  ("021", "Dịch vụ hệ thống “Chi phí đi lại” không sửa được", "P0",
   "Tồn tại dòng tên đúng “Chi phí đi lại”",
   "1. Tìm dòng “Chi phí đi lại”\n2. Rê chuột lên nút Sửa\n"
   "3. Dùng công cụ kiểm thử API gọi thẳng chức năng Sửa cho dòng này",
   "Dịch vụ hệ thống",
   "- Nút Sửa MỜ, chú thích “Chi phí này không được phép sửa”\n"
   "- Gọi thẳng chức năng Sửa cũng bị hệ thống từ chối, dữ liệu không đổi"),

  ("022", "Dịch vụ hệ thống “Chi phí vận chuyển” bị chặn mọi thao tác", "P0",
   "Tồn tại dòng tên đúng “Chi phí vận chuyển”",
   "1. Rê chuột lên nút Sửa và nút Xóa của dòng đó\n2. Rê chuột lên nút ổ khóa",
   "Dịch vụ hệ thống",
   "- Cả 3 nút Sửa, Xóa và ổ khóa đều MỜ\n"
   "- Chú thích lần lượt: “Chi phí này không được phép sửa”, “Chi phí này không được phép xóa”, "
   "“Chi phí này không được phép khóa”"),
 ]),

 ("V", "KHÓA / MỞ KHÓA", [
  ("001", "Vị trí nút Khóa / Mở khóa", "P0",
   "Tài khoản C có quyền quản lý",
   "1. Quan sát cột “Trạng thái” của từng dòng",
   "—",
   "- Nút ổ khóa nằm NGAY TRONG cột Trạng thái, bên phải nhãn trạng thái (KHÔNG nằm ở cột Hành động)\n"
   "- Dòng Hoạt động: biểu tượng ổ khóa đóng; dòng Khóa: biểu tượng ổ khóa mở"),

  ("002", "Khóa một dịch vụ đang Hoạt động", "P0",
   "Dịch vụ “TC Dịch vụ kiểm thử 02” đang Hoạt động",
   "1. Bấm nút ổ khóa ở dòng đó\n2. Quan sát hộp thoại\n3. Bấm “Khóa”",
   "Khóa dịch vụ TC Dịch vụ kiểm thử 02",
   "- Hộp thoại tiêu đề “Xác nhận khóa”, nội dung “Bạn có chắc muốn khóa 'TC Dịch vụ kiểm thử 02'?”, "
   "có 2 nút “Hủy” và “Khóa”\n"
   "- Sau khi xác nhận: thông báo “Khóa thành công”, danh sách nạp lại, nhãn đổi thành “Khóa”"),

  ("003", "Bấm Hủy ở hộp xác nhận khóa", "P0",
   "Dịch vụ X đang Hoạt động",
   "1. Bấm nút ổ khóa\n2. Bấm “Hủy”",
   "—",
   "- Hộp thoại đóng lại, trạng thái của X giữ nguyên “Hoạt động”"),

  ("004", "Mở khóa một dịch vụ đang Khóa", "P0",
   "Dịch vụ “TC Dịch vụ kiểm thử 02” đang Khóa",
   "1. Bấm nút ổ khóa (biểu tượng mở) ở dòng đó\n2. Quan sát hộp thoại\n3. Bấm “Mở khóa”",
   "Mở khóa dịch vụ đang Khóa",
   "- Hộp thoại tiêu đề “Xác nhận mở khóa”, nút xác nhận ghi “Mở khóa”\n"
   "- Thông báo “Mở khóa thành công”, nhãn đổi lại thành “Hoạt động”\n"
   "- Nút Sửa và nút Xóa của dòng đó sáng trở lại"),

  ("005", "Mở khóa giữ nguyên ĐM giảm giá cũ", "P0",
   "Dịch vụ X có ĐM giảm giá 12% rồi bị Khóa",
   "1. Mở khóa X\n2. Xem cột “ĐM giảm giá”",
   "Trước khi khóa: 12%",
   "- Sau khi mở khóa, ĐM giảm giá vẫn là 12%\n"
   "- Thao tác khóa và mở khóa chỉ đổi trạng thái, không làm mất định mức đã khai báo"),

  ("006", "Khóa rồi mở khóa không làm đổi thông tin khác", "P1",
   "Dịch vụ X: giá vốn 70%, VAT 8%, có tính doanh thu",
   "1. Khóa X\n2. Mở khóa X\n3. Đối chiếu lại các cột",
   "—",
   "- % Tính giá vốn, % VAT và Phân loại giữ nguyên\n"
   "- Chỉ cột “Cập nhật” đổi sang người và thời điểm mới nhất"),

  ("007", "Dịch vụ hệ thống không khóa được", "P0",
   "Dòng “Chi phí đi lại” đang Hoạt động",
   "1. Rê chuột lên nút ổ khóa của dòng đó\n"
   "2. Dùng công cụ kiểm thử API gọi thẳng chức năng Khóa cho dòng này",
   "Dịch vụ hệ thống",
   "- Nút ổ khóa MỜ, chú thích “Chi phí này không được phép khóa”\n"
   "- Gọi thẳng chức năng Khóa cũng bị từ chối, trạng thái không đổi"),

  ("008", "Khóa một dịch vụ vốn đã Khóa", "P1",
   "Dịch vụ X đang ở trạng thái Khóa",
   "1. Dùng công cụ kiểm thử API gọi thẳng chức năng Khóa cho X",
   "X đang Khóa sẵn",
   "- Hệ thống báo “Chi phí đang bị khóa” và không thực hiện gì thêm"),

  ("009", "Mở khóa một dịch vụ vốn đang Hoạt động", "P1",
   "Dịch vụ X đang Hoạt động",
   "1. Dùng công cụ kiểm thử API gọi thẳng chức năng Mở khóa cho X",
   "X đang Hoạt động sẵn",
   "- Hệ thống báo “Chi phí đang hoạt động” và không thực hiện gì thêm"),

  ("010", "Người chỉ có quyền Xem không thấy nút ổ khóa", "P0",
   "Tài khoản B chỉ có quyền xem",
   "1. Đăng nhập B, quan sát cột Trạng thái",
   "Tài khoản B",
   "- Cột Trạng thái CHỈ có nhãn trạng thái, không hiển thị nút ổ khóa"),
 ]),

 ("VI", "XÓA (XÓA HẲN HOẶC CHUYỂN SANG KHÓA)", [
  ("001", "Xóa dịch vụ chưa phát sinh chứng từ", "P0",
   "Dịch vụ “TC Dịch vụ kiểm thử 01” đang Hoạt động và CHƯA được dùng ở Báo giá hãng hay "
   "Hợp đồng hãng nào",
   "1. Bấm nút Xóa (biểu tượng thùng rác) ở dòng đó\n2. Quan sát hộp thoại\n3. Bấm “Xóa”",
   "Dịch vụ chưa dùng ở đâu",
   "- Hộp thoại tiêu đề “Xác nhận xóa”, nội dung “Bạn có chắc muốn xóa 'TC Dịch vụ kiểm thử 01'?”, "
   "nút xác nhận ghi “Xóa”\n"
   "- Thông báo “Xóa thành công”, dịch vụ BIẾN MẤT khỏi danh sách\n"
   "- ĐM giảm giá của dịch vụ đó ở MỌI công ty cũng bị xóa theo"),

  ("002", "Xóa dịch vụ đã dùng ở Báo giá hãng — chỉ bị Khóa", "P0",
   "Dịch vụ Y đang Hoạt động và ĐÃ được dùng trong ít nhất 1 Báo giá hãng",
   "1. Bấm nút Xóa ở dòng Y\n2. Đọc kỹ hộp thoại\n3. Bấm nút xác nhận",
   "Dịch vụ Y đã dùng ở Báo giá hãng",
   "- Hộp thoại đổi tiêu đề thành “Xác nhận khóa”, nút xác nhận ghi “Khóa”\n"
   "- Nội dung ghi rõ “Chi phí đang được sử dụng ở: Báo giá hãng. Hệ thống sẽ chuyển sang trạng thái "
   "Khóa thay vì xóa.”\n"
   "- Sau xác nhận: thông báo “Khóa thành công”, dòng Y VẪN CÒN trong danh sách nhưng nhãn đổi "
   "thành “Khóa”\n- ĐM giảm giá được GIỮ NGUYÊN"),

  ("003", "Xóa dịch vụ đã dùng ở Hợp đồng hãng — chỉ bị Khóa", "P0",
   "Dịch vụ Z đã được dùng trong ít nhất 1 Hợp đồng hãng",
   "1. Bấm nút Xóa ở dòng Z\n2. Đọc thông báo trong hộp thoại",
   "Dịch vụ Z đã dùng ở Hợp đồng hãng",
   "- Thông báo ghi rõ “Chi phí đang được sử dụng ở: Hợp đồng hãng.”\n"
   "- Sau xác nhận: dịch vụ chuyển sang Khóa, không bị xóa"),

  ("004", "Dịch vụ dùng ở cả Báo giá hãng và Hợp đồng hãng", "P1",
   "Dịch vụ W đã được dùng ở CẢ Báo giá hãng và Hợp đồng hãng",
   "1. Bấm nút Xóa ở dòng W\n2. Đọc thông báo",
   "Dịch vụ W dùng ở cả 2 nơi",
   "- Thông báo liệt kê ĐỦ cả 2 nguồn: “Báo giá hãng, Hợp đồng hãng”"),

  ("005", "Bấm Hủy ở hộp xác nhận xóa", "P0",
   "Dịch vụ X đang Hoạt động",
   "1. Bấm nút Xóa\n2. Bấm “Hủy”",
   "—",
   "- Hộp thoại đóng lại, dịch vụ vẫn còn nguyên, trạng thái không đổi"),

  ("006", "Nút Xóa bị mờ khi dòng đang Khóa", "P0",
   "Dịch vụ “test01111” đang Khóa",
   "1. Rê chuột lên nút Xóa của dòng đó",
   "Dòng đang Khóa",
   "- Nút Xóa MỜ, chú thích “Chi phí đã bị khóa”"),

  ("007", "Dịch vụ hệ thống không xóa được", "P0",
   "Dòng “Chi phí đi lại”",
   "1. Rê chuột lên nút Xóa\n2. Dùng công cụ kiểm thử API gọi thẳng chức năng Xóa cho dòng này",
   "Dịch vụ hệ thống",
   "- Nút Xóa MỜ, chú thích “Chi phí này không được phép xóa”\n"
   "- Gọi thẳng chức năng Xóa cũng bị từ chối"),

  ("008", "Bấm Xóa liên tiếp nhiều dòng", "P1",
   "Danh sách có nhiều dòng xóa được",
   "1. Bấm nút Xóa ở dòng 1\n2. Bấm ngay nút Xóa ở dòng 2 khi hệ thống còn đang kiểm tra",
   "2 lần bấm liên tiếp",
   "- Trong lúc hệ thống đang kiểm tra dòng 1, nút Xóa của các dòng khác bị vô hiệu\n"
   "- Chỉ mở đúng 1 hộp thoại, không mở chồng 2 hộp"),

  ("009", "Xóa xong tổng số dòng được cập nhật", "P1",
   "Danh sách đang có 526 dòng",
   "1. Xóa hẳn 1 dịch vụ\n2. Đọc dòng đếm ở chân bảng",
   "—",
   "- Tổng số đổi thành 525 ngay, không cần tải lại trang thủ công"),

  ("010", "Xóa dịch vụ vừa bị người khác xóa", "P2",
   "Dịch vụ X vừa được người dùng khác xóa xong",
   "1. Bấm Xóa dòng X trên màn hình chưa được tải lại",
   "Dịch vụ đã không còn",
   "- Hệ thống báo dữ liệu đã thay đổi, không treo trang, không hiện lỗi hệ thống"),
 ]),

 ("VII", "XUẤT EXCEL", [
  ("001", "Xuất Excel toàn bộ danh sách", "P0",
   "Danh sách 526 dòng, chưa áp bộ lọc nào",
   "1. Bấm nút “Xuất Excel”\n2. Mở file tải về",
   "Không lọc",
   "- Thông báo “Xuất Excel thành công”\n"
   "- File tên “danh_muc_dich_vu_sua_chua_va_chi_phi_khac.xlsx”\n"
   "- File chứa ĐỦ 526 dòng, không phải chỉ 10 dòng của trang đang xem"),

  ("002", "File Excel đúng 8 cột", "P0",
   "Đã tải được file Excel",
   "1. Đọc hàng tiêu đề trong file",
   "—",
   "- Đúng 8 cột theo thứ tự: STT | Tên dịch vụ / chi phí | Phân loại | ĐM giảm giá (%) | "
   "% Tính giá vốn | % VAT | Trạng thái | Người tạo\n"
   "- Phía trên có thông tin công ty và tiêu đề “Danh mục dịch vụ sửa chữa và chi phí khác”"),

  ("003", "Xuất Excel theo đúng bộ lọc đang áp dụng", "P0",
   "Đang lọc Trạng thái = “Khóa”, kết quả còn N dòng",
   "1. Bấm “Xuất Excel”\n2. Đếm số dòng dữ liệu trong file",
   "Đang lọc Trạng thái Khóa",
   "- File CHỈ chứa đúng N dòng đang Khóa, không xuất cả 526 dòng"),

  ("004", "Xuất Excel giữ đúng thứ tự sắp xếp", "P1",
   "Đang sắp xếp theo Tên tăng dần",
   "1. Bấm “Xuất Excel”\n2. Đối chiếu thứ tự dòng trong file với trên màn hình",
   "Sắp theo Tên tăng dần",
   "- Thứ tự dòng trong file khớp thứ tự đang hiển thị trên màn hình"),

  ("005", "Cột phần trăm trong Excel ghi 2 chữ số thập phân", "P1",
   "Dịch vụ có % Tính giá vốn là 80 (màn hình hiển thị “80%”)",
   "1. Mở file, xem ô tương ứng",
   "Giá vốn 80%",
   "- Trong file ghi “80,00”\n"
   "- Đây là hành vi đúng của hiện trạng (file luôn ghi 2 chữ số thập phân), không phải lỗi"),

  ("006", "Ô ĐM giảm giá để trống khi chưa khai báo", "P1",
   "Dịch vụ chưa khai báo ĐM giảm giá cho công ty đang chọn",
   "1. Mở file, xem cột “ĐM giảm giá (%)” của dòng đó",
   "Chưa khai báo",
   "- Ô để TRỐNG, không ghi 0,00 và không ghi “—”"),

  ("007", "Cột Người tạo trong Excel", "P1",
   "Dịch vụ do “DNS01 - DNS Admin” tạo, sau đó người khác sửa",
   "1. Mở file, xem cột “Người tạo”",
   "Người tạo khác người sửa",
   "- Ghi “DNS01 - DNS Admin” (mã nhân viên - họ tên)\n"
   "- ⚠️ Đây là NGƯỜI TẠO, khác với cột “Cập nhật” trên màn hình (là người sửa gần nhất)"),

  ("008", "Xuất Excel khi kết quả lọc rỗng", "P2",
   "Đang lọc bằng từ khóa không khớp dòng nào",
   "1. Bấm “Xuất Excel”\n2. Mở file",
   "Từ khóa: ZZZZZ999",
   "- Vẫn tải được file, có tiêu đề và hàng tiêu đề cột nhưng không có dòng dữ liệu, không báo lỗi"),

  ("009", "Người chỉ có quyền Xem vẫn xuất được Excel", "P1",
   "Tài khoản B chỉ có quyền xem",
   "1. Đăng nhập B, bấm “Xuất Excel”",
   "Tài khoản B",
   "- Nút “Xuất Excel” hiển thị và tải file thành công"),

  ("010", "File tải về mở được bình thường", "P1",
   "Đang đăng nhập bình thường",
   "1. Bấm “Xuất Excel”\n2. Mở file bằng Excel",
   "—",
   "- File mở được, đúng định dạng Excel\n"
   "- KHÔNG phải file lỗi hay trang đăng nhập bị tải nhầm"),
 ]),

 ("VIII", "RÀNG BUỘC NHẬP LIỆU", [
  ("001", "Bỏ trống toàn bộ trường bắt buộc", "P0",
   "Đang mở cửa sổ Tạo mới, chưa nhập gì",
   "1. Bấm “Lưu” ngay",
   "Để trống tất cả",
   "- Thông báo đỏ “Bạn chưa nhập đầy đủ thông tin”\n"
   "- Hiện chữ đỏ “Bắt buộc phải nhập” ngay dưới CẢ 3 ô: Tên dịch vụ / chi phí, % Tính giá vốn, % VAT\n"
   "- Cửa sổ KHÔNG bị đóng, không tạo dịch vụ nào"),

  ("002", "Bỏ trống Tên dịch vụ / chi phí", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Để trống Tên, nhập % Tính giá vốn = 50 và % VAT = 8\n2. Bấm “Lưu”",
   "Tên: để trống",
   "- Báo lỗi “Bắt buộc phải nhập” ngay dưới ô Tên, ô viền đỏ"),

  ("003", "Tên chỉ gồm khoảng trắng", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập Tên chỉ gồm vài dấu cách\n2. Nhập đủ 2 ô còn lại\n3. Bấm “Lưu”",
   "Tên: chỉ có dấu cách",
   "- Hệ thống coi như để trống và báo “Bắt buộc phải nhập”"),

  ("004", "Tên vượt quá 255 ký tự", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập Tên dài 256 ký tự\n2. Nhập đủ các ô khác\n3. Bấm “Lưu”",
   "Tên dài 256 ký tự",
   "- Báo lỗi “Tối đa 255 ký tự” dưới ô Tên"),

  ("005", "Tên dài đúng 255 ký tự", "P1",
   "Đang mở cửa sổ Tạo mới, chưa có dịch vụ nào trùng tên đó",
   "1. Nhập Tên dài đúng 255 ký tự\n2. Nhập đủ các ô khác\n3. Bấm “Lưu”",
   "Tên dài đúng 255 ký tự",
   "- Lưu THÀNH CÔNG (255 ký tự là hợp lệ)"),

  ("006", "Tên trùng dịch vụ đã có", "P0",
   "Đã tồn tại dịch vụ tên “Dịch vụ đánh bóng xy lanh”",
   "1. Tạo mới với Tên = “Dịch vụ đánh bóng xy lanh”\n2. Nhập đủ các ô khác\n3. Bấm “Lưu”",
   "Tên trùng dịch vụ đã có",
   "- Báo lỗi “Đã tồn tại trên hệ thống” dưới ô Tên\n- Không tạo ra dòng thứ 2"),

  ("007", "Sửa thành trùng tên dịch vụ khác", "P0",
   "Tồn tại 2 dịch vụ A và B khác tên nhau",
   "1. Sửa A, đổi tên A thành đúng tên của B\n2. Bấm “Lưu”",
   "Đổi tên A thành tên B",
   "- Báo lỗi “Đã tồn tại trên hệ thống”, không lưu"),

  ("008", "Trùng tên với dịch vụ đang bị Khóa", "P1",
   "Dịch vụ “test01111” đang ở trạng thái Khóa",
   "1. Tạo mới với Tên = “test01111”\n2. Bấm “Lưu”",
   "Trùng tên dịch vụ đang Khóa",
   "- Vẫn báo “Đã tồn tại trên hệ thống”\n"
   "- Quy tắc chống trùng tên áp dụng cho cả dịch vụ đang Khóa"),

  ("009", "Bỏ trống % Tính giá vốn", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập Tên hợp lệ, để trống % Tính giá vốn, nhập % VAT = 8\n2. Bấm “Lưu”",
   "% Tính giá vốn: để trống",
   "- Báo lỗi “Bắt buộc phải nhập” dưới ô % Tính giá vốn"),

  ("010", "% Tính giá vốn nhập chữ", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % Tính giá vốn = “abc”\n2. Bấm “Lưu”",
   "% Tính giá vốn: abc",
   "- Báo lỗi “Phải là số”"),

  ("011", "% Tính giá vốn nhập số âm", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % Tính giá vốn = -5\n2. Bấm “Lưu”",
   "% Tính giá vốn: -5",
   "- Báo lỗi “Không được nhỏ hơn 0”"),

  ("012", "% Tính giá vốn nhập 0", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % Tính giá vốn = 0, các ô khác hợp lệ\n2. Bấm “Lưu”",
   "% Tính giá vốn: 0",
   "- Lưu THÀNH CÔNG, lưới hiển thị “0%”\n"
   "- Giá trị 0 là hợp lệ, không được coi là bỏ trống"),

  ("013", "% Tính giá vốn lớn hơn 100", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % Tính giá vốn = 321\n2. Nhập các ô khác hợp lệ\n3. Bấm “Lưu”",
   "% Tính giá vốn: 321",
   "- Lưu THÀNH CÔNG, lưới hiển thị “321%”\n"
   "- ⚠️ Ô này CHỈ chặn giá trị âm, KHÔNG chặn trần 100 (dữ liệu thật đang có dòng 321%)"),

  ("014", "% Tính giá vốn nhập dấu phẩy thập phân", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % Tính giá vốn = “12,5”\n2. Nhập các ô khác hợp lệ\n3. Bấm “Lưu”\n"
   "4. Mở lại bằng chức năng Sửa để kiểm tra giá trị đã lưu",
   "% Tính giá vốn: 12,5",
   "- Lưu đúng giá trị 12,5 phần trăm, lưới hiển thị “12,5%”\n"
   "- ⚠️ ĐÂY LÀ TRƯỜNG HỢP DỄ SAI NHẤT: nếu lưu thành 125 là LỖI NẶNG — sai gấp 10 lần mà hệ thống "
   "không báo gì"),

  ("015", "Bỏ trống % VAT", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Để trống % VAT, nhập đủ các ô khác\n2. Bấm “Lưu”",
   "% VAT: để trống",
   "- Báo lỗi “Bắt buộc phải nhập” dưới ô % VAT"),

  ("016", "% VAT lớn hơn 100", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % VAT = 101\n2. Bấm “Lưu”",
   "% VAT: 101",
   "- Báo lỗi “Tối đa 100” dưới ô % VAT\n"
   "- ⚠️ Khác ô % Tính giá vốn: ô VAT CÓ chặn trần 100"),

  ("017", "% VAT đúng 100", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % VAT = 100\n2. Bấm “Lưu”",
   "% VAT: 100",
   "- Lưu THÀNH CÔNG, lưới hiển thị “100%”"),

  ("018", "% VAT nhập số âm", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % VAT = -1\n2. Bấm “Lưu”",
   "% VAT: -1",
   "- Báo lỗi ở ô % VAT, không lưu"),

  ("019", "% VAT nhập chữ", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập % VAT = “tám”\n2. Bấm “Lưu”",
   "% VAT: tám",
   "- Báo lỗi “Phải là số”"),

  ("020", "ĐM giảm giá lớn hơn 100", "P0",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập ĐM giảm giá = 150\n2. Nhập các ô khác hợp lệ\n3. Bấm “Lưu”",
   "ĐM giảm giá: 150",
   "- Báo lỗi “Tối đa 100” dưới ô ĐM giảm giá"),

  ("021", "ĐM giảm giá nhập số âm", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập ĐM giảm giá = -3\n2. Bấm “Lưu”",
   "ĐM giảm giá: -3",
   "- Báo lỗi ở ô ĐM giảm giá, không lưu"),

  ("022", "ĐM giảm giá nhập chữ", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập ĐM giảm giá = “abc”\n2. Bấm “Lưu”",
   "ĐM giảm giá: abc",
   "- Báo lỗi “Phải là số”"),

  ("023", "ĐM giảm giá nhập dấu phẩy thập phân", "P1",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập ĐM giảm giá = “7,5”\n2. Bấm “Lưu”\n3. Mở lại bằng chức năng Sửa để kiểm tra",
   "ĐM giảm giá: 7,5",
   "- Lưu đúng giá trị 7,5 phần trăm, lưới hiển thị “7,5%” (không phải 75)"),

  ("024", "Sửa lại giá trị sai thì lỗi biến mất", "P1",
   "Đang ở cửa sổ Sửa và đang có lỗi ở ô % VAT",
   "1. Bấm Lưu với % VAT = 101 (thấy báo lỗi)\n2. Sửa % VAT = 8\n3. Bấm “Lưu” lần nữa",
   "101 rồi sửa thành 8",
   "- Lần 2 lưu thành công, chữ đỏ ở ô % VAT biến mất"),

  ("025", "Thao tác lên khoản chi phí không thuộc màn này", "P0",
   "Tồn tại một khoản “Chi phí phải trả” (thuộc danh mục bên ERP, không thuộc màn này)",
   "1. Dùng công cụ kiểm thử API gọi lần lượt các chức năng Xem chi tiết, Sửa, Xóa, Khóa cho khoản đó",
   "Bản ghi thuộc danh mục khác",
   "- TẤT CẢ đều bị từ chối với thông báo “Bản ghi không thuộc danh mục dịch vụ sửa chữa và chi phí khác”\n"
   "- Dữ liệu của khoản chi phí đó không bị thay đổi"),

  ("026", "Xem chi tiết dịch vụ không tồn tại", "P2",
   "Dịch vụ đã bị xóa hoặc chưa từng tồn tại",
   "1. Bấm Xem trên màn hình chưa được tải lại sau khi người khác đã xóa dịch vụ đó",
   "Dịch vụ không còn",
   "- Thông báo “Dữ liệu đã thay đổi, vui lòng tải lại”, cửa sổ không mở, không treo trang"),

  ("027", "Tên có ký tự đặc biệt và tiếng Việt có dấu", "P2",
   "Đang mở cửa sổ Tạo mới",
   "1. Nhập Tên = “Dịch vụ kiểm tra & sửa chữa (100%) - Điện/Nước”\n2. Bấm “Lưu”",
   "Tên có ký tự &, %, ngoặc, gạch chéo và dấu tiếng Việt",
   "- Lưu thành công\n"
   "- Tên hiển thị nguyên vẹn trên lưới, trong cửa sổ Sửa và trong file Excel, không bị biến dạng"),
 ]),

 ("IX", "CÔ LẬP DỮ LIỆU & THAO TÁC ĐỒNG THỜI", [
  ("001", "Hai người cùng sửa một dịch vụ", "P2",
   "Tài khoản C và D cùng mở cửa sổ Sửa của dịch vụ X, cả hai đều có quyền quản lý",
   "1. C đổi Tên rồi Lưu\n2. D (không tải lại màn hình) đổi % VAT rồi Lưu\n"
   "3. Mở lại dịch vụ X để đối chiếu",
   "2 người thao tác song song",
   "- Cả 2 lần lưu đều cho kết quả rõ ràng (thành công hoặc có thông báo lỗi)\n"
   "- Dịch vụ X không bị mất thông tin một cách âm thầm; cột Cập nhật ghi người lưu sau"),

  ("002", "Sửa dịch vụ vừa bị người khác xóa", "P2",
   "Tài khoản C đang mở cửa sổ Sửa dịch vụ X; tài khoản D xóa X ở màn hình khác",
   "1. C bấm “Lưu”",
   "Dịch vụ đã bị xóa",
   "- Hệ thống báo dữ liệu đã thay đổi, không treo trang, không hiện lỗi hệ thống"),

  ("003", "Thu hồi quyền khi người dùng đang mở cửa sổ", "P1",
   "Tài khoản C đang mở cửa sổ Sửa với quyền quản lý",
   "1. Quản trị gỡ quyền “Quản lý dịch vụ sửa chữa và chi phí khác” của C\n"
   "2. C (không tải lại trang) bấm “Lưu”\n3. C tải lại trang",
   "Gỡ quyền lúc cửa sổ đang mở",
   "- Bấm Lưu: hệ thống từ chối, dữ liệu KHÔNG bị ghi\n"
   "- Sau khi tải lại: nút Tạo mới, Sửa, Xóa và ổ khóa biến mất"),

  ("004", "Đổi công ty đang chọn rồi quay lại màn hình", "P1",
   "Tài khoản E được gán nhiều công ty; dịch vụ X có ĐM giảm giá khác nhau ở 2 công ty",
   "1. Ở công ty 1, xem cột “ĐM giảm giá” của X\n2. Đổi sang công ty 4\n3. Vào lại màn hình, xem X",
   "Đổi công ty từ 1 sang 4",
   "- Giá trị “ĐM giảm giá” đổi theo công ty vừa chọn\n"
   "- Tên, % Tính giá vốn, % VAT và Trạng thái KHÔNG đổi"),

  ("005", "Sửa ĐM giảm giá không ảnh hưởng công ty khác", "P0",
   "Dịch vụ X: công ty 1 là 5%, công ty 4 là 12%. Tài khoản C thuộc công ty 1",
   "1. C sửa X, đổi ĐM giảm giá thành 20\n2. Lưu\n"
   "3. Đăng nhập tài khoản thuộc công ty 4 và xem lại X",
   "Công ty 1 đổi 5% thành 20%",
   "- Công ty 1 thấy 20%\n- Công ty 4 VẪN thấy 12%, không bị ghi đè"),

  ("006", "Dữ liệu dùng chung với phần mềm ERP", "P1",
   "Cùng một dịch vụ tồn tại trên cả 2 cổng",
   "1. Sửa % VAT của dịch vụ đó trên cổng HRM\n"
   "2. Mở màn danh mục chi phí tương ứng bên ERP và tìm dịch vụ đó",
   "Sửa bên HRM, kiểm tra bên ERP",
   "- Bên ERP hiển thị đúng giá trị vừa sửa bên HRM\n"
   "- Hai cổng dùng chung một danh mục, không cần thao tác đồng bộ thủ công"),
 ]),
]

# ============================================================================
# STYLE
# ============================================================================
FONT_NAME = "Times New Roman"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_L = Alignment(wrap_text=True, vertical="center", horizontal="left")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")

WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
DESC_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
GREEN_FILL = PatternFill("solid", fgColor="00FF00")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")

COL_WIDTHS = {
    'A': 26.9, 'B': 27.1, 'C': 16, 'D': 26.6, 'E': 9,
    'F': 22.8, 'G': 18.6, 'H': 22, 'I': 43.9, 'J': 41.6,
    'K': 14, 'L': 14, 'M': 14, 'N': 20,
    'O': 11, 'P': 11, 'Q': 11,
}

LAST_COL = 17  # Q

wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w


def put(r, c, value, *, bold=False, size=12, color=None, fill=None,
        align=WRAP_L, border=True):
    cell = ws.cell(r, c, value)
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    cell.alignment = align
    if fill is not None:
        cell.fill = fill
    if border:
        cell.border = BORDER
    return cell


# --- ROW 1 ---
put(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)", bold=True, border=False)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
ws.row_dimensions[1].height = 22

# --- ROW 2-10 ---
for idx, (label, bodytext) in enumerate(DESCRIPTION_BLOCK, start=2):
    put(idx, 1, label, bold=True, fill=DESC_FILL)
    put(idx, 2, bodytext)
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=14)
    ws.row_dimensions[idx].height = max(45, bodytext.count("\n") * 17 + 32)

# --- ROW 11-15 : title + summary ---
put(11, 1, f"Testcase _ {FEATURE_NAME}", bold=True, size=15, color="FFFFFFFF",
    fill=TITLE_FILL, border=False)
ws.merge_cells("A11:E11")
put(11, 6, "TEST SUMMARY", bold=True, size=15, color="FFFFFFFF", fill=TITLE_FILL,
    align=WRAP_C, border=False)
ws.merge_cells("F11:H15")
ws.row_dimensions[11].height = 30

SUMMARY = [
    (11, "Số trường hợp DNS kiểm thử đạt (P):", '=COUNTIF(K18:M1000,"Passed")',
         "Số trường hợp TP kiểm thử đạt (P):", '=COUNTIF(O18:Q1000,"P")'),
    (12, "Số trường hợp DNS kiểm thử không đạt (F)", '=COUNTIF(K18:M1000,"Failed")',
         "Số trường hợp TP kiểm thử không đạt (F)", '=COUNTIF(O18:Q1000,"F")'),
    (13, "Số trường hợp DNS kiểm thử đang xem xét (PE)", '=COUNTIF(K18:M1000,"Pending")',
         "Số trường hợp TP kiểm thử đang xem xét (PE)", '=COUNTIF(O18:Q1000,"PE")'),
    (14, "Số trường hợp DNS kiểm thử chưa thực hiện", '=COUNTIF(K18:K1000,"Not Executed")',
         "Số trường hợp TP kiểm thử chưa thực hiện", '=COUNTIF(O18:O1000,"")'),
    (15, "Tổng số trường hợp DNS kiểm thử", '=COUNTIF(K18:K1000,"<>")',
         "Tổng số trường hợp TP kiểm thử", '=COUNTIF(C18:C1000,"TC*")'),
]
for r, dns_label, dns_formula, tp_label, tp_formula in SUMMARY:
    put(r, 9, dns_label, fill=GREEN_FILL, border=False)
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
    put(r, 11, dns_formula, bold=True, fill=GREEN_FILL, align=WRAP_C)
    ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
    put(r, 13, tp_label, fill=WHITE_FILL, border=False)
    ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=14)
    put(r, 15, tp_formula, bold=True, fill=WHITE_FILL, align=WRAP_C)
    ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=17)
    ws.row_dimensions[r].height = 24

ws.row_dimensions[16].height = 8

# --- ROW 17 : header ---
HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "KQ thực tế",
    "DNS check lần 1", "DNS check lần 2", "DNS check lần 3", "Ghi chú",
    "TP check lần 1", "TP check lần 2", "TP check lần 3",
]
for i, hd in enumerate(HEADERS, start=1):
    put(17, i, hd, bold=True, color="FFFFFFFF", fill=HEADER_FILL, align=WRAP_C)
ws.row_dimensions[17].height = 40

# --- DATA ---
current_row = 18


def write_section(title):
    global current_row
    put(current_row, 1, None, fill=SECTION_FILL)
    put(current_row, 2, None, fill=SECTION_FILL)
    put(current_row, 3, title, bold=True, color="FF1F4E79", fill=SECTION_FILL)
    ws.merge_cells(start_row=current_row, start_column=3,
                   end_row=current_row, end_column=LAST_COL)
    ws.row_dimensions[current_row].height = 26
    current_row += 1


def write_tc(group, tc_id, function, priority, precondition, steps, test_data, expected):
    global current_row
    values = [
        MODULE_NAME, group, tc_id, function, priority,
        precondition, steps, test_data, expected, "",
        "Not Executed", "Not Executed", "Not Executed", "",
        "", "", "",
    ]
    for i, v in enumerate(values, start=1):
        align = WRAP_C if i in (3, 5, 11, 12, 13, 15, 16, 17) else WRAP_L
        put(current_row, i, v, fill=WHITE_FILL, align=align)
    longest = max(len(str(v)) for v in values)
    ws.row_dimensions[current_row].height = max(34, min(200, longest // 3))
    current_row += 1


write_section("Phân quyền & truy cập")
for suffix, func, prio, pre, steps, td, exp in ROLE_TCS:
    write_tc("Phân quyền & truy cập", f"TC-ROLE-{suffix}", func, prio, pre, steps, td, exp)

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
for roman, title, tcs in SECTIONS:
    write_section(f"{roman}. {title}")
    sec_idx = ROMAN.index(roman) + 1
    for tc_num, func, prio, pre, steps, td, exp in tcs:
        write_tc(title, f"TC_{sec_idx:02d}.{int(tc_num):03d}", func, prio, pre, steps, td, exp)

last_data_row = current_row - 1

# --- DATA VALIDATION ---
dv_dns = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                        allow_blank=True, showDropDown=False)
dv_dns.add(f"K18:M{last_data_row + 50}")
ws.add_data_validation(dv_dns)

dv_tp = DataValidation(type="list", formula1='"P,F,PE"',
                       allow_blank=True, showDropDown=False)
dv_tp.add(f"O18:Q{last_data_row + 50}")
ws.add_data_validation(dv_tp)

wb.save(OUTPUT_FILE)

# ---------------------------------------------------------------- kiem tra
# Chan thuat ngu ky thuat lot vao tai lieu danh cho QA / nghiep vu.
import re

BANNED = [
    r"`[a-z_]{3,}`",                      # ten bang / ten cot dat trong dau `
    r"\bBE\b", r"\bFE\b",
    r"\bHTTP\b", r"\b(4\d{2}|5\d{2}) Forbidden\b",
    r"trả (400|403|404|422)", r"\b(400|403|404|422)\b",
    r"permission id", r"\bAPI /", r"/api/v1", r"localStorage",
    r"number_format", r"meta\.", r"sort_by", r"per_page",
    r"role_has_permissions", r"current_company_role",
]
blob = []
for label, bodytext in DESCRIPTION_BLOCK:
    blob.append(bodytext)
for tc in ROLE_TCS:
    blob.extend(tc[3:])
for _, _, tcs in SECTIONS:
    for tc in tcs:
        blob.extend(tc[3:])
text_all = "\n".join(blob)
found = {}
for pat in BANNED:
    hits = re.findall(pat, text_all)
    if hits:
        found[pat] = len(hits)
if found:
    print("!!! CON THUAT NGU KY THUAT:", found)
else:
    print("OK - khong con thuat ngu ky thuat trong noi dung")

total = sum(len(s[2]) for s in SECTIONS) + len(ROLE_TCS)
p0 = sum(1 for s in SECTIONS for tc in s[2] if tc[2] == "P0") + \
     sum(1 for tc in ROLE_TCS if tc[2] == "P0")
print("Da tao:", OUTPUT_FILE)
print("Tong TC:", total, "| P0:", p0, "(%.0f%%)" % (p0 * 100.0 / total))
print("Dong du lieu: 18 ->", last_data_row)
