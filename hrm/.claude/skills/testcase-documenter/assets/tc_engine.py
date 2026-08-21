# -*- coding: utf-8 -*-
"""Engine dung file Excel testcase chuan ERP TPE (form 17 cot, 2 khoi summary DNS/TP).

Tach ra tu `gen_testcase_mau.py` de nhieu feature dung chung: file generator cua tung
man chi con 3 khoi CONFIG (DESCRIPTION_BLOCK / ROLE_TCS / SECTIONS) roi goi `build(...)`.

Cach dung trong `.plans/<feature>/gen_testcase.py`:

    import os, sys
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    "..", "..", "..", ".claude", "skills",
                                    "testcase-documenter", "assets"))
    from tc_engine import build

    build(output_file=..., sheet_name="Trang tinh1", feature_name=..., module_name=...,
          description_block=DESCRIPTION_BLOCK, role_tcs=ROLE_TCS, sections=SECTIONS)

Cau truc du lieu:
    description_block : list[(nhan, noi dung)] - dung 9 muc
    role_tcs          : list[(hau to, chuc nang, priority, tien dieu kien, buoc, test data, expected)]
    sections          : list[(so La Ma, ten section, [ (so tc, chuc nang, priority,
                                                        tien dieu kien, buoc, test data, expected) ])]
"""
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

try:  # console Windows mac dinh cp1252 -> print tieng Viet se no UnicodeEncodeError
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ---------------------------------------------------------------- style
FONT_NAME = "Times New Roman"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_L = Alignment(wrap_text=True, vertical="center", horizontal="left")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")

WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
DESC_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
GREEN_FILL = PatternFill("solid", fgColor="00FF00")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")

COL_WIDTHS = {
    'A': 26.9, 'B': 27.1, 'C': 16, 'D': 26.6, 'E': 9,
    'F': 22.8, 'G': 18.6, 'H': 22, 'I': 43.9, 'J': 41.6,
    'K': 14, 'L': 14, 'M': 14, 'N': 20,
    'O': 11, 'P': 11, 'Q': 11,
}

LAST_COL = 17  # Q

HEADERS = [
    "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
    "Tiền điều kiện", "Bước thực hiện", "Test Data",
    "Expected Result (chi tiết)", "KQ thực tế",
    "DNS check lần 1", "DNS check lần 2", "DNS check lần 3", "Ghi chú",
    "TP check lần 1", "TP check lần 2", "TP check lần 3",
]

SUMMARY = [
    (11, "Số trường hợp DNS kiểm thử đạt (P):", '=COUNTIF(K18:M1000,"Passed")',
         "Số trường hợp TP kiểm thử đạt (P):", '=COUNTIF(O18:Q1000,"P")'),
    (12, "Số trường hợp DNS kiểm thử không đạt (F)", '=COUNTIF(K18:M1000,"Failed")',
         "Số trường hợp TP kiểm thử không đạt (F)", '=COUNTIF(O18:Q1000,"F")'),
    (13, "Số trường hợp DNS kiểm thử đang xem xét (PE)", '=COUNTIF(K18:M1000,"Pending")',
         "Số trường hợp TP kiểm thử đang xem xét (PE)", '=COUNTIF(O18:Q1000,"PE")'),
    (14, "Số trường hợp DNS kiểm thử chưa thực hiện", '=COUNTIF(K18:K1000,"Not Executed")',
         "Số trường hợp TP kiểm thử chưa thực hiện", '=COUNTIF(O18:O1000,"")'),
    (15, "Tổng số trường hợp DNS kiểm thử", '=COUNTIF(K18:K1000,"<>")',
         "Tổng số trường hợp TP kiểm thử", '=COUNTIF(C18:C1000,"TC*")'),
]

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]

# Chan thuat ngu ky thuat lot vao tai lieu danh cho QA / nghiep vu.
BANNED = [
    r"`[a-z_]{3,}`",                      # ten bang / ten cot dat trong dau `
    r"\bBE\b", r"\bFE\b",
    r"\bHTTP\b", r"\b(4\d{2}|5\d{2}) Forbidden\b",
    r"trả (400|403|404|422)", r"\b(400|403|404|422)\b",
    r"permission id", r"\bAPI /", r"/api/v1", r"localStorage",
    r"number_format", r"meta\.", r"sort_by", r"per_page",
    r"role_has_permissions", r"current_company_role",
]


def check_terms(description_block, role_tcs, sections):
    """In canh bao neu con thuat ngu code. Tra ve dict pattern -> so lan."""
    blob = [body for _, body in description_block]
    for tc in role_tcs:
        blob.extend(tc[3:])
    for _, _, tcs in sections:
        for tc in tcs:
            blob.extend(tc[3:])
    text_all = "\n".join(blob)
    found = {}
    for pat in BANNED:
        hits = re.findall(pat, text_all)
        if hits:
            found[pat] = len(hits)
    if found:
        print("!!! CON THUAT NGU KY THUAT:", found)
    else:
        print("OK - khong con thuat ngu ky thuat trong noi dung")
    return found


def build(output_file, sheet_name, feature_name, module_name,
          description_block, role_tcs, sections, role_section_title="Phân quyền & truy cập"):
    assert len(description_block) == 9, "Khoi mo ta phai du 9 muc"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    def put(r, c, value, bold=False, size=12, color=None, fill=None,
            align=WRAP_L, border=True):
        cell = ws.cell(r, c, value)
        cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
        cell.alignment = align
        if fill is not None:
            cell.fill = fill
        if border:
            cell.border = BORDER
        return cell

    # --- ROW 1 ---
    put(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)", bold=True, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.row_dimensions[1].height = 22

    # --- ROW 2-10 : 9 muc mo ta ---
    for idx, (label, bodytext) in enumerate(description_block, start=2):
        put(idx, 1, label, bold=True, fill=DESC_FILL)
        put(idx, 2, bodytext)
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=14)
        ws.row_dimensions[idx].height = max(45, bodytext.count("\n") * 17 + 32)

    # --- ROW 11-15 : title + 2 khoi summary ---
    put(11, 1, "Testcase _ %s" % feature_name, bold=True, size=15, color="FFFFFFFF",
        fill=TITLE_FILL, border=False)
    ws.merge_cells("A11:E11")
    put(11, 6, "TEST SUMMARY", bold=True, size=15, color="FFFFFFFF", fill=TITLE_FILL,
        align=WRAP_C, border=False)
    ws.merge_cells("F11:H15")
    ws.row_dimensions[11].height = 30

    for r, dns_label, dns_formula, tp_label, tp_formula in SUMMARY:
        put(r, 9, dns_label, fill=GREEN_FILL, border=False)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        put(r, 11, dns_formula, bold=True, fill=GREEN_FILL, align=WRAP_C)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
        put(r, 13, tp_label, fill=WHITE_FILL, border=False)
        ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=14)
        put(r, 15, tp_formula, bold=True, fill=WHITE_FILL, align=WRAP_C)
        ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=17)
        ws.row_dimensions[r].height = 24

    ws.row_dimensions[16].height = 8

    # --- ROW 17 : header 17 cot ---
    for i, hd in enumerate(HEADERS, start=1):
        put(17, i, hd, bold=True, color="FFFFFFFF", fill=HEADER_FILL, align=WRAP_C)
    ws.row_dimensions[17].height = 40

    # --- DATA ---
    state = {"row": 18}

    def write_section(title):
        r = state["row"]
        put(r, 1, None, fill=SECTION_FILL)
        put(r, 2, None, fill=SECTION_FILL)
        put(r, 3, title, bold=True, color="FF1F4E79", fill=SECTION_FILL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
        ws.row_dimensions[r].height = 26
        state["row"] = r + 1

    def write_tc(group, tc_id, function, priority, precondition, steps, test_data, expected):
        r = state["row"]
        values = [
            module_name, group, tc_id, function, priority,
            precondition, steps, test_data, expected, "",
            "Not Executed", "Not Executed", "Not Executed", "",
            "", "", "",
        ]
        for i, v in enumerate(values, start=1):
            align = WRAP_C if i in (3, 5, 11, 12, 13, 15, 16, 17) else WRAP_L
            put(r, i, v, fill=WHITE_FILL, align=align)
        longest = max(len(str(v)) for v in values)
        ws.row_dimensions[r].height = max(34, min(200, longest // 3))
        state["row"] = r + 1

    seen_ids = set()

    def claim(tc_id):
        assert tc_id not in seen_ids, "Trung TC ID: %s" % tc_id
        seen_ids.add(tc_id)
        return tc_id

    if role_tcs:
        write_section(role_section_title)
        for suffix, func, prio, pre, steps, td, exp in role_tcs:
            write_tc(role_section_title, claim("TC-ROLE-%s" % suffix),
                     func, prio, pre, steps, td, exp)

    for roman, title, tcs in sections:
        write_section("%s. %s" % (roman, title))
        sec_idx = ROMAN.index(roman) + 1
        for tc_num, func, prio, pre, steps, td, exp in tcs:
            tc_id = claim("TC_%02d.%03d" % (sec_idx, int(tc_num)))
            write_tc(title, tc_id, func, prio, pre, steps, td, exp)

    last_data_row = state["row"] - 1

    dv_dns = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"',
                            allow_blank=True, showDropDown=False)
    dv_dns.add("K18:M%d" % (last_data_row + 50))
    ws.add_data_validation(dv_dns)

    dv_tp = DataValidation(type="list", formula1='"P,F,PE"',
                           allow_blank=True, showDropDown=False)
    dv_tp.add("O18:Q%d" % (last_data_row + 50))
    ws.add_data_validation(dv_tp)

    wb.save(output_file)

    check_terms(description_block, role_tcs, sections)

    total = sum(len(s[2]) for s in sections) + len(role_tcs)
    p0 = sum(1 for s in sections for tc in s[2] if tc[2] == "P0") + \
        sum(1 for tc in role_tcs if tc[2] == "P0")
    print("Da tao:", output_file)
    print("Tong TC:", total, "| P0:", p0, "(%.0f%%)" % (p0 * 100.0 / total))
    print("Dong du lieu: 18 ->", last_data_row)
    return total
