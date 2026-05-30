"""
Generator Excel testcase cho màn "Lịch làm việc của tôi" (/assign/my-todo)
Format theo mẫu Testcase_baocao.xlsx:
- Block mô tả tính năng (dòng 1-9)
- Block TEST SUMMARY (dòng 11-15)
- Header 15 cột (dòng 16)
- Section rows + Test case rows
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT = "docs/srs/lich-lam-viec-testcases.xlsx"
FEATURE_TITLE = "Testcase _ Lịch làm việc của tôi (My To Do)"
MODULE = "Giao việc"
URL = "/assign/my-todo"

# ============ STYLES ============
thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
font_desc_label = Font(name="Calibri", size=11, bold=True)
font_desc_text = Font(name="Calibri", size=11)
font_summary_label = Font(name="Calibri", size=11, bold=True, color="1F4E79")
font_summary_value = Font(name="Calibri", size=11, bold=True)
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=11, bold=True, color="1F4E79")
font_body = Font(name="Calibri", size=11)

fill_title = PatternFill("solid", fgColor="DCE6F1")
fill_summary = PatternFill("solid", fgColor="FFF2CC")
fill_header = PatternFill("solid", fgColor="4472C4")
fill_section = PatternFill("solid", fgColor="D6E4F0")
fill_even = PatternFill("solid", fgColor="F7F9FC")

align_wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
align_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
align_section = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

# ============ DATA ============
DESCRIPTION_ROWS = [
    ("1. Mục đích tính năng",
     "Cung cấp cho mỗi user 1 màn hình tổng hợp công việc cá nhân theo thời gian, "
     "tích hợp dữ liệu từ Task, Issue, Phiếu giao việc, Phiếu giao công tác, Meeting "
     "và Todo cá nhân (Personal Todo) do user tự tạo. Giúp user nắm bắt nhanh khối lượng "
     "việc trong ngày, tuần, quá hạn và lập kế hoạch cá nhân (danh sách + sub-task)."),
    ("2. Đối tượng được hiển thị",
     "► TASK: assignee_id = user_id, KHÔNG bao gồm status Nháp(1), Hoàn thành(8), Huỷ(9).\n"
     "► ISSUE: assignee_id = user_id, KHÔNG bao gồm closed, completed, rejected.\n"
     "► PHIẾU GIAO VIỆC (AssignJob): status = Đã duyệt, user là thành viên qua employee_info_id.\n"
     "► PHIẾU CÔNG TÁC (AssignRequest): type = PHIEU_CONG_TAC, status = Đã lập phiếu công tác, user là thành viên.\n"
     "► MEETING: status ∈ {Lên lịch, Chốt lịch}, user là company_members.\n"
     "► PERSONAL TODO: user_id = auth user, parent_id IS NULL (todo cấp 1)."),
    ("3. Đối tượng bị ẩn / không tính",
     "► Task ở các status Nháp, Hoàn thành, Huỷ.\n"
     "► Issue đã đóng / hoàn thành / từ chối.\n"
     "► Meeting ở status Đang tạo, Hoàn thành, Hủy.\n"
     "► Phiếu chưa duyệt / từ chối / huỷ.\n"
     "► Dữ liệu của user khác — màn luôn cô lập theo user_id."),
    ("4. Bộ lọc thời gian áp dụng cho",
     "Lọc theo NGÀY HẠN (due_date) của từng item. Có 4 nhóm thời gian luôn hiển thị: "
     "Hôm nay, Ngày mai, Tuần này, Quá hạn. Ngoài ra có Tuần sau, Sắp tới, Không hạn. "
     "Click chọn 1 ngày trên mini calendar → chuyển sang flat mode chỉ hiển thị item có due_date = ngày đó."),
    ("5. Cấu trúc danh sách cá nhân",
     "User → PersonalTodoList → PersonalTodo (cấp 1) → Sub-task (cấp 2, tối đa 1 cấp).\n"
     "Lần đầu truy cập tự sinh 1 list mặc định 'Mac dinh' nếu user chưa có list nào."),
    ("6. Quy tắc cộng dồn / deduplicate",
     "Stats trên đầu trang đếm theo tất cả nhóm sau filter (không phụ thuộc nhóm đang mở/đóng).\n"
     "Nếu user vừa là assignee vừa là watcher cùng 1 entity → chỉ hiện 1 lần, "
     "ưu tiên: Được giao(1) > Cần duyệt(2) > Theo dõi(3) > Tham gia(4) > Cá nhân(5)."),
    ("7. Phân quyền cấp",
     "Không yêu cầu role/permission đặc biệt — mọi user đã đăng nhập đều truy cập được.\n"
     "Dữ liệu được cô lập theo user_id: mỗi user chỉ thấy việc của mình. "
     "Personal todo / list chỉ chính chủ thao tác được; user khác KHÔNG đọc/sửa/xoá được."),
    ("8. Cách tính 4 ô thống kê trên đầu trang",
     "► 'Ngày hạn' của 1 việc tương ứng với:\n"
     "   • Công việc (Task): cột Ngày hạn / Hạn hoàn thành.\n"
     "   • Vấn đề (Issue): Hạn xử lý.\n"
     "   • Phiếu giao việc / Phiếu công tác: ngày phải hoàn thành ghi trên phiếu.\n"
     "   • Cuộc họp (Meeting): ngày diễn ra cuộc họp.\n"
     "   • Việc cá nhân (Personal Todo): ngày hạn do user nhập khi tạo.\n"
     "► Quá hạn = số việc có Ngày hạn TRƯỚC hôm nay.\n"
     "► Hôm nay = số việc có Ngày hạn = hôm nay.\n"
     "► Tuần này = số việc có Ngày hạn từ HÔM NAY đến CHỦ NHẬT của tuần hiện tại "
     "(BAO GỒM cả những việc thuộc 'Hôm nay').\n"
     "► Tổng việc = tổng số việc đang hiển thị trên trang sau bộ lọc dropdown (Loại / Trạng thái), "
     "tính cả việc KHÔNG có ngày hạn; KHÔNG tính bước con (sub-task).\n"
     "► Việc KHÔNG có Ngày hạn: bỏ qua 3 ô Quá hạn / Hôm nay / Tuần này, chỉ vào Tổng việc.\n"
     "► Các việc đã hoàn thành / huỷ / đóng đã được hệ thống loại từ trước nên không tính vào Quá hạn "
     "(Task ở trạng thái Hoàn thành/Huỷ/Nháp, Issue đã Đóng/Hoàn thành/Từ chối, "
     "Meeting đã Hoàn thành/Huỷ, Phiếu chưa duyệt → đều không xuất hiện).\n"
     "► Gõ ô Tìm kiếm KHÔNG làm đổi số 4 ô (tìm kiếm chỉ lọc danh sách bên dưới, không ảnh hưởng stats)."),
    ("9. Ghi chú đọc bảng",
     "Nhóm thời gian luôn show 4 nhóm chính dù rỗng (Hôm nay, Ngày mai, Tuần này, Quá hạn). "
     "Click checkbox chỉ áp dụng cho Personal Todo; entity hệ thống không có checkbox toggle. "
     "Action buttons (Xem/Sửa/Nhập KQ/Duyệt/Xử lý/Lịch sử) chỉ hiện trên Task & Issue khi hover."),
]

# ============ TEST CASES ============
# Mỗi entry: (section_title, [(tc_func, priority, precond, steps, test_data, expected, note), ...])

SECTIONS = []

# --- SECTION I: HIỂN THỊ TRANG ---
SECTIONS.append(("I. HIỂN THỊ TRANG & TRUY CẬP", [
    ("Truy cập trang Lịch làm việc của tôi", "P0",
     "User đã đăng nhập hệ thống",
     "1. Truy cập URL /assign/my-todo\n2. Quan sát layout",
     "User bất kỳ",
     "Trang load thành công với layout 2 cột:\n- Cột trái: header stats (Quá hạn, Hôm nay, Tuần này, Tổng), filter bar (Loại / Trạng thái / Tìm kiếm / Tạo todo), khu vực danh sách nhóm theo thời gian.\n- Cột phải (300px): mini calendar tháng + khu vực 'Danh sách cá nhân' kèm nút '+ Tạo danh sách mới'.",
     "BR — Mọi user đều truy cập được"),
    ("Hiển thị 4 ô stats trên header", "P1",
     "Đang ở /assign/my-todo",
     "1. Quan sát hàng stats trên cùng\n2. Kiểm tra label + icon 4 ô",
     "—",
     "Header hiển thị đúng 4 ô theo thứ tự: 'Quá hạn' (icon cảnh báo đỏ), 'Hôm nay' (icon mặt trời cam), "
     "'Tuần này' (icon lịch xanh), 'Tổng việc' (icon list).", ""),
    ("Ô 'Quá hạn' đếm việc có Ngày hạn TRƯỚC hôm nay", "P0",
     "User có: 2 Task ngày hạn = hôm qua, 1 Issue ngày hạn = tuần trước, 3 việc ngày hạn = hôm nay",
     "1. Truy cập /assign/my-todo\n2. Đọc số ô 'Quá hạn'",
     "Hôm nay = 22/05/2026",
     "Ô 'Quá hạn' = 3 (2 Task hôm qua + 1 Issue tuần trước). "
     "Việc ngày hạn = hôm nay KHÔNG tính vào Quá hạn.",
     "Công thức: due_date < today"),
    ("Ô 'Hôm nay' đếm việc có Ngày hạn ĐÚNG hôm nay", "P0",
     "User có: 2 Task ngày hạn = hôm nay, 1 Meeting ngày hạn = hôm nay, 1 Task ngày hạn = ngày mai",
     "1. Truy cập /assign/my-todo\n2. Đọc số ô 'Hôm nay'",
     "Hôm nay = 22/05/2026",
     "Ô 'Hôm nay' = 3 (2 Task + 1 Meeting cùng ngày 22/05). "
     "Việc ngày mai KHÔNG tính vào Hôm nay.",
     "Công thức: due_date = today"),
    ("Ô 'Tuần này' BAO GỒM cả việc của 'Hôm nay'", "P0",
     "Hôm nay = thứ 5. User có: 2 việc hôm nay, 1 việc thứ 6, 1 việc CN, 1 việc thứ 3 tuần sau",
     "1. Truy cập /assign/my-todo\n2. Đọc số ô 'Tuần này'",
     "Hôm nay = thứ 5 22/05/2026",
     "Ô 'Tuần này' = 4 (2 hôm nay + 1 thứ 6 + 1 CN). Việc thứ 3 tuần sau KHÔNG tính. "
     "Range: từ hôm nay (thứ 5) → CN tuần này (25/05).",
     "Công thức: today ≤ due_date ≤ endOfWeek"),
    ("Ô 'Tuần này' khi hôm nay là CN", "P1",
     "Hôm nay = Chủ nhật. User có 2 việc hôm nay, 1 việc thứ 2 tuần sau",
     "1. Truy cập trang vào ngày CN\n2. Đọc số 'Tuần này'",
     "Hôm nay = CN 25/05/2026",
     "Ô 'Tuần này' = 2 (chỉ việc hôm nay). Range chỉ còn đúng 1 ngày là hôm nay. "
     "Việc thứ 2 tuần sau KHÔNG tính.", ""),
    ("Ô 'Tuần này' KHÔNG bao gồm việc quá hạn", "P1",
     "Hôm nay = thứ 4. User có: 1 việc thứ 2 đầu tuần (đã qua), 2 việc thứ 4 (hôm nay)",
     "1. Đọc số 'Tuần này'",
     "—",
     "Ô 'Tuần này' = 2. Việc thứ 2 đã qua KHÔNG tính vào Tuần này (nó vào ô Quá hạn).", ""),
    ("Ô 'Tổng việc' = tổng tất cả việc đang hiển thị", "P0",
     "User có: 3 Task, 2 Issue, 1 Meeting, 4 Personal Todo (2 có ngày hạn, 2 không có ngày hạn)",
     "1. Truy cập /assign/my-todo\n2. Đọc số 'Tổng việc'",
     "—",
     "Ô 'Tổng việc' = 10. Tính cả 2 todo không có ngày hạn. KHÔNG tính bước con (sub-task).",
     "Tổng = items.length sau bộ lọc"),
    ("Việc KHÔNG có ngày hạn chỉ vào 'Tổng việc'", "P1",
     "User có 5 Personal Todo: 3 có ngày hạn (hôm nay), 2 không có ngày hạn",
     "1. Đọc 4 ô stats",
     "—",
     "Quá hạn = 0, Hôm nay = 3, Tuần này = 3, Tổng việc = 5. "
     "2 todo không có ngày hạn chỉ được đếm vào Tổng việc.", ""),
    ("Sub-task KHÔNG tính vào 'Tổng việc'", "P1",
     "User có 2 todo cấp 1, mỗi todo có 3 sub-task",
     "1. Đọc 'Tổng việc'",
     "—",
     "Ô 'Tổng việc' = 2 (chỉ tính todo cấp 1). 6 sub-task KHÔNG tính.", ""),
    ("Stats KHÔNG đổi khi gõ ô tìm kiếm", "P1",
     "User có 10 việc, stats hiển thị 10",
     "1. Gõ 'abc' vào ô tìm kiếm (không trùng việc nào)\n2. Quan sát stats",
     "search='abc'",
     "Danh sách bên dưới rỗng nhưng 4 ô stats vẫn = giá trị cũ (10). "
     "Tìm kiếm chỉ lọc danh sách hiển thị, không tính lại stats.",
     "Stats tính trên this.items (raw), không phải filteredItems"),
    ("Stats CẬP NHẬT khi đổi dropdown Loại / Trạng thái", "P0",
     "User có 10 việc gồm 4 Task, đang xem tất cả",
     "1. Chọn Loại = 'Task'\n2. Quan sát stats",
     "type=task",
     "API gọi lại với type=task, items chỉ còn 4 Task. 4 ô stats cập nhật theo 4 Task này.", ""),
    ("Stats khi user không có việc nào", "P1",
     "User mới, danh sách rỗng",
     "1. Truy cập /assign/my-todo\n2. Quan sát stats",
     "—",
     "4 ô đều = 0: Quá hạn 0 / Hôm nay 0 / Tuần này 0 / Tổng việc 0.", ""),
    ("Tự động tạo danh sách mặc định khi user lần đầu vào", "P0",
     "User chưa từng truy cập My To Do (chưa có list nào)",
     "1. Truy cập /assign/my-todo\n2. Kiểm tra cột phải",
     "User mới",
     "Hệ thống tự tạo 1 danh sách tên 'Mac dinh' (sort_order=0) và hiển thị ở sidebar bên phải. "
     "API GET /assign/my-todo trả về list này trong field lists[].",
     "BR-02"),
    ("Cô lập dữ liệu theo user", "P0",
     "User A có 3 Personal Todo, User B có 5 Personal Todo",
     "1. Login user A, truy cập /assign/my-todo\n2. Logout, login user B, truy cập /assign/my-todo",
     "user_a, user_b",
     "User A chỉ thấy 3 todo của A (kèm Task/Issue/Meeting của A). "
     "User B chỉ thấy 5 todo của B. "
     "Không user nào thấy được dữ liệu của user kia, kể cả cùng phòng ban/công ty.",
     "BR-05"),
    ("Hiển thị item entity hệ thống với viền màu theo loại", "P1",
     "User có cả Task, Issue, AssignJob, AssignBusiness, Meeting, Personal Todo",
     "1. Quan sát từng item trong danh sách",
     "—",
     "Mỗi item có viền trái 3px màu phân biệt: Task=#4a7cfb (xanh dương), Issue=#e25c5c (đỏ), "
     "AssignJob=#22c55e (xanh lá), AssignBusiness=#d4a03c (vàng), Meeting=#0ea5e9 (xanh nhạt), "
     "Personal=#8e99a4 (xám). Icon trái khớp loại item.", ""),
    ("Hiển thị Role badge trên mỗi item", "P1",
     "User có item với các vai trò khác nhau",
     "1. Quan sát badge bên cạnh title của từng item",
     "—",
     "Mỗi item hiển thị 1 role badge: 'Được giao' (#4a7cfb), 'Tham gia' (#22c55e), "
     "'Cần duyệt' (#d4a03c), 'Theo dõi' (#8b5cf6), 'Cá nhân' (#6b7280). "
     "Màu nền badge khớp với role (EFF6FF/F0FDF4/FFFBEB/F5F3FF/F3F4F6).", ""),
    ("Hiển thị Status badge khớp với loại entity", "P1",
     "User có Task status=4 'Đang làm', Issue 'in_progress', Meeting 'CHOT_LICH'",
     "1. Quan sát hàng status bên dưới title từng item",
     "—",
     "Task hiển thị 'Đang làm' (màu primary). Issue hiển thị 'Đang xử lý' (warning). "
     "Meeting hiển thị 'Chốt lịch' (success). Personal todo hiện 'Việc cần làm' (secondary).", ""),
    ("Item quá hạn hiển thị ngày hạn màu đỏ", "P1",
     "User có Task due_date = hôm qua",
     "1. Tìm item Task quá hạn trong nhóm 'Quá hạn'\n2. Kiểm tra góc phải item",
     "—",
     "Ngày hạn (DD/MM) hiển thị màu đỏ (#e25c5c) và item nằm trong nhóm 'Quá hạn'.", ""),
    ("Item hôm nay hiển thị giờ HH:mm", "P2",
     "User có Meeting due_date=today, due_time=14:00",
     "1. Mở nhóm 'Hôm nay'\n2. Kiểm tra góc phải item Meeting",
     "—",
     "Góc phải item hiển thị '14:00' (không hiện ngày, vì là hôm nay).", ""),
]))

# --- SECTION II: NHÓM THỜI GIAN ---
SECTIONS.append(("II. NHÓM THỜI GIAN (TIME GROUPS)", [
    ("Nhóm 'Hôm nay' chứa item có due_date = hôm nay", "P0",
     "User có 1 Task due_date=today, 1 Personal Todo due_date=today",
     "1. Mở nhóm 'Hôm nay'\n2. Đếm số item",
     "—",
     "Nhóm 'Hôm nay' header hiển thị tiêu đề + ngày DD/MM hôm nay + tổng (2). Liệt kê đúng 2 item.", ""),
    ("Nhóm 'Ngày mai' chứa item có due_date = tomorrow", "P1",
     "User có Task due_date = ngày mai",
     "1. Mở nhóm 'Ngày mai'\n2. Kiểm tra item bên trong",
     "—",
     "Nhóm 'Ngày mai' hiển thị item Task. Màu chip = #d4a03c (vàng).", ""),
    ("Nhóm 'Tuần này' chứa item từ ngày kia → cuối tuần", "P1",
     "User có 3 item trong tuần (sau 'Ngày mai'), 1 item tuần sau",
     "1. Mở nhóm 'Tuần này'\n2. Đếm item",
     "—",
     "Nhóm 'Tuần này' chỉ có 3 item, không bao gồm item tuần sau.", ""),
    ("Nhóm 'Quá hạn' chứa item có due_date < today và chưa xong", "P0",
     "User có 2 Task có due_date < today, status chưa xong",
     "1. Mở nhóm 'Quá hạn'\n2. Kiểm tra item",
     "—",
     "Nhóm 'Quá hạn' (màu đỏ #e25c5c) hiển thị 2 Task. Ngày hạn hiển thị đỏ.", ""),
    ("Nhóm 'Không hạn' chứa item có due_date = null", "P1",
     "User có Personal Todo không nhập ngày hạn",
     "1. Mở nhóm 'Không hạn'",
     "—",
     "Nhóm 'Không hạn' hiển thị todo, góc phải item hiện '—' hoặc bỏ trống thay vì ngày.", ""),
    ("4 nhóm cơ bản luôn hiển thị dù rỗng", "P1",
     "User mới, chưa có item nào",
     "1. Truy cập /assign/my-todo\n2. Quan sát các nhóm",
     "—",
     "Vẫn hiển thị header 4 nhóm: 'Hôm nay (0)', 'Ngày mai (0)', 'Tuần này (0)', 'Quá hạn (0)'.", ""),
    ("Mặc định mở nhóm 'Hôm nay', thu gọn các nhóm khác", "P2",
     "User có item ở nhiều nhóm",
     "1. Truy cập /assign/my-todo\n2. Quan sát trạng thái mặc định các nhóm",
     "—",
     "Nhóm 'Hôm nay' mở (icon ▼). Các nhóm còn lại thu gọn (icon ▶).", ""),
    ("Toggle expand/collapse 1 nhóm", "P1",
     "Đang ở /assign/my-todo",
     "1. Click header nhóm 'Tuần này'\n2. Click lại",
     "—",
     "Lần 1: nhóm mở rộng, icon ▼, hiện danh sách item. Lần 2: thu gọn, icon ▶, ẩn item. Không gọi API.", ""),
]))

# --- SECTION III: BỘ LỌC & TÌM KIẾM ---
SECTIONS.append(("III. BỘ LỌC & TÌM KIẾM", [
    ("Lọc theo Loại = Task", "P0",
     "User có nhiều loại item",
     "1. Mở dropdown 'Loại'\n2. Chọn 'Công việc (Task)'\n3. Quan sát danh sách",
     "type=task",
     "Gọi API GET /assign/my-todo?type=task. Danh sách CHỈ hiển thị item type=task. "
     "Các loại khác (Issue, Meeting, Personal...) bị ẩn.", ""),
    ("Lọc theo Loại = Việc cá nhân", "P1",
     "User có Personal Todo + nhiều loại khác",
     "1. Chọn dropdown 'Loại' = 'Việc cá nhân'",
     "type=personal",
     "Chỉ hiển thị Personal Todo. Các entity hệ thống bị ẩn.", ""),
    ("Lọc theo Trạng thái = Đã xong (chỉ áp với personal)", "P0",
     "User có 5 Personal Todo (2 đã xong) + 3 Task chưa xong",
     "1. Chọn dropdown 'Trạng thái' = 'Đã xong'",
     "status=done",
     "Chỉ hiện 2 Personal Todo đã xong. Task/Issue/Meeting KHÔNG xuất hiện (vì filter done chỉ áp với personal todo).",
     "BR-10"),
    ("Lọc theo Trạng thái = Chưa xong (mặc định)", "P1",
     "User có item ở nhiều trạng thái",
     "1. Truy cập /assign/my-todo (lần đầu)",
     "status=pending",
     "Mặc định filter status=pending: hiển thị Task chưa xong, Issue chưa xong, Personal Todo chưa hoàn thành.", ""),
    ("Tìm kiếm text trên title (client-side)", "P1",
     "Có item title chứa 'báo cáo'",
     "1. Gõ 'báo cáo' vào ô tìm kiếm",
     "search=báo cáo",
     "Danh sách lọc ngay (không gọi API): chỉ giữ item có title chứa 'báo cáo' (case-insensitive). "
     "Các nhóm vẫn hiển thị, chỉ item match được giữ.", ""),
    ("Xóa text tìm kiếm → quay lại danh sách đầy đủ", "P2",
     "Đang có text trong ô search",
     "1. Xóa nội dung ô search",
     "—",
     "Danh sách hiển thị lại tất cả item như chưa tìm kiếm.", ""),
    ("Kết hợp filter Loại + Trạng thái", "P1",
     "User có Task chưa xong, Task đã xong, Personal Todo đã xong",
     "1. Chọn Loại = 'Task'\n2. Chọn Trạng thái = 'Đã xong'",
     "type=task&status=done",
     "Danh sách rỗng (vì status=done chỉ giữ personal todo, không có Task). Hiển thị empty state.",
     "BR-10"),
    ("Reload sau khi đổi filter", "P1",
     "Đang ở /assign/my-todo",
     "1. Đổi dropdown Loại\n2. Kiểm tra network request",
     "—",
     "Mỗi lần đổi dropdown (Loại / Trạng thái) → gọi lại API GET /assign/my-todo với query params tương ứng. "
     "Loading indicator hiện trong lúc gọi.", ""),
]))

# --- SECTION IV: MINI CALENDAR ---
SECTIONS.append(("IV. MINI CALENDAR (CỘT PHẢI)", [
    ("Calendar hiển thị tháng hiện tại", "P1",
     "Đang ở /assign/my-todo",
     "1. Quan sát mini calendar cột phải",
     "—",
     "Calendar mặc định tháng hiện tại. Header: '◀ Tháng [N] [YYYY] ▶'. "
     "Hiển thị đầy đủ T2-CN trong tháng. Ô hôm nay highlight (border hoặc bg).", ""),
    ("Calendar hiển thị dot trên ngày có item", "P1",
     "User có 3 item due_date=02/05, 1 item due_date=15/05",
     "1. Mở calendar tháng 05",
     "—",
     "Ô 02/05 và 15/05 có chấm dot (dấu hiệu có item). Các ngày khác không có dot.", ""),
    ("Click 1 ngày trên calendar → flat mode lọc theo ngày", "P0",
     "User có item ở nhiều ngày",
     "1. Click ô ngày 03/05 trên calendar",
     "date=2026-05-03",
     "Danh sách chuyển sang flat mode (KHÔNG nhóm theo thời gian). "
     "Chỉ hiển thị item có due_date = 03/05. "
     "Gọi API GET /assign/my-todo?date=2026-05-03. Ô 03/05 highlight.", ""),
    ("Click lại cùng ngày → bỏ chọn, quay về grouped mode", "P1",
     "Đã click ngày 03/05",
     "1. Click lại ô 03/05",
     "—",
     "Bỏ chọn ngày: danh sách quay về grouped mode (hiển thị các nhóm Hôm nay/Ngày mai/Tuần này...).", ""),
    ("Chuyển tháng calendar", "P2",
     "Đang xem tháng 05/2026",
     "1. Click ▶ (next)\n2. Quan sát header",
     "—",
     "Calendar chuyển sang tháng 06/2026. Dot dữ liệu update theo tháng mới (gọi API GET ?month=2026-06).", ""),
]))

# --- SECTION V: CRUD DANH SÁCH CÁ NHÂN ---
SECTIONS.append(("V. CRUD DANH SÁCH CÁ NHÂN (PersonalTodoList)", [
    ("Mở modal Tạo danh sách mới", "P0",
     "Đang ở /assign/my-todo",
     "1. Click '+ Tạo danh sách mới' ở sidebar phải",
     "—",
     "Mở modal 'TodoListFormModal' có 2 field: 'Tên danh sách' (required), 'Mô tả' (optional). "
     "Nút 'Lưu' (disable khi tên rỗng) + 'Huỷ'.", ""),
    ("Tạo danh sách thành công", "P0",
     "Đang ở modal tạo list",
     "1. Nhập tên = 'Việc cần làm hôm nay'\n2. Click Lưu",
     "name='Việc cần làm hôm nay'",
     "POST /assign/my-todo/lists trả 200. Modal đóng. Sidebar xuất hiện list mới ở cuối danh sách. "
     "sort_order = max(sort_order) + 1. Toast 'Tạo danh sách thành công'.",
     "BR-08"),
    ("Tạo list rỗng name → báo lỗi validation", "P0",
     "Đang ở modal tạo list",
     "1. Để trống Tên\n2. Click Lưu",
     "name=''",
     "Nút Lưu disable HOẶC submit báo lỗi: field tên có viền đỏ + text 'Vui lòng nhập tên danh sách'. "
     "Không gọi API.", ""),
    ("Tạo list với name > 255 ký tự → lỗi", "P2",
     "Đang ở modal tạo list",
     "1. Nhập tên 256 ký tự\n2. Click Lưu",
     "name=256 ký tự",
     "API trả 422 hoặc FE validate trước: hiện lỗi 'Tên không quá 255 ký tự'.", ""),
    ("Sửa tên danh sách", "P1",
     "Có ít nhất 1 list cá nhân",
     "1. Click icon 'Sửa' bên cạnh tên list\n2. Đổi tên → Lưu",
     "id=1, name='Tên mới'",
     "PUT /assign/my-todo/lists/1 trả 200. Sidebar cập nhật tên ngay. Toast thành công.", ""),
    ("Xóa danh sách (confirm + cascade xóa todos)", "P0",
     "Có list 'Test' chứa 3 todo + 2 sub-task",
     "1. Click icon 'Xoá' bên cạnh list\n2. Confirm dialog 'Xóa danh sách? Tất cả việc sẽ bị xóa.'\n3. Bấm Xóa",
     "id=1",
     "DELETE /assign/my-todo/lists/1 trả 200. List biến mất khỏi sidebar. "
     "Tất cả todo + sub-task trong list bị xóa cascade. Nếu đang xem detail list này → quay về trang chính.",
     "BR-03"),
    ("Huỷ confirm xóa list", "P2",
     "Đã mở confirm xóa",
     "1. Bấm 'Huỷ'",
     "—",
     "Đóng dialog. List vẫn còn nguyên.", ""),
    ("User không xóa được list của user khác", "P0",
     "List id=99 thuộc user B",
     "Login user A, gọi DELETE /assign/my-todo/lists/99",
     "user A, id=99",
     "API trả 404 hoặc 403. List vẫn còn. Không bị xóa.",
     "BR-05"),
    ("Kéo thả sắp xếp list", "P1",
     "Có 3 list",
     "1. Hover list 'B' → icon drag handle\n2. Kéo lên trước list 'A'",
     "[{id:b,sort_order:0},{id:a,sort_order:1},{id:c,sort_order:2}]",
     "UI update ngay (optimistic). POST /assign/my-todo/lists/reorder. Refresh trang vẫn giữ thứ tự mới.", ""),
    ("Kéo thả lỗi API → rollback", "P2",
     "Đang kéo thả, API lỗi 500",
     "1. Mô phỏng API lỗi\n2. Kéo thả list",
     "—",
     "Hiện toast lỗi. Sidebar reload lại danh sách từ server (về thứ tự cũ).", ""),
    ("Hiển thị bộ đếm 'completed/total' trên mỗi list", "P1",
     "List 'A' có 5 todo, 2 đã xong",
     "1. Quan sát label list 'A' ở sidebar",
     "—",
     "Hiển thị '2/5' bên phải tên list. Chỉ tính todo cấp 1 (parent_id IS NULL), không tính sub-task.",
     "BR-09"),
]))

# --- SECTION VI: CRUD PERSONAL TODO ---
SECTIONS.append(("VI. CRUD PERSONAL TODO", [
    ("Mở modal Tạo todo từ filter bar", "P0",
     "Đang ở /assign/my-todo",
     "1. Click nút 'Tạo todo' trên filter bar",
     "—",
     "Mở TodoFormModal gồm: Tiêu đề (required), Mô tả, Danh sách (dropdown, default = list đầu), "
     "Ngày hạn (datepicker), Giờ hạn (time picker). Nút Lưu + Huỷ.", ""),
    ("Tạo todo thành công", "P0",
     "Đang ở modal tạo todo",
     "1. Nhập Tiêu đề = 'Gọi NCC'\n2. Chọn list = 'Mac dinh'\n3. Chọn ngày hạn = today\n4. Lưu",
     "title='Gọi NCC', list_id=1, due_date=today",
     "POST /assign/my-todo/todos trả 200. Modal đóng. Todo xuất hiện trong nhóm 'Hôm nay'. "
     "sort_order = max + 1 trong list.",
     "BR-08"),
    ("Tạo todo title rỗng → lỗi", "P0",
     "Đang ở modal",
     "1. Bỏ trống title\n2. Lưu",
     "title=''",
     "Title viền đỏ 'is-invalid' + text 'Vui lòng nhập tiêu đề'. Không gọi API.", ""),
    ("Tạo todo title > 500 ký tự → lỗi", "P2",
     "Đang ở modal",
     "1. Nhập title 501 ký tự\n2. Lưu",
     "title=501 ký tự",
     "Lỗi 422: 'Tiêu đề không quá 500 ký tự'.", ""),
    ("Tạo todo có giờ hạn", "P1",
     "Đang ở modal",
     "1. Nhập title, chọn date=today, time=17:00\n2. Lưu",
     "due_time=17:00",
     "Todo tạo thành công. Hiển thị trong nhóm 'Hôm nay', góc phải item hiện '17:00'.", ""),
    ("Sửa todo", "P1",
     "Có todo cá nhân",
     "1. Click 'Sửa' trên todo (hoặc click vào item)\n2. Đổi title\n3. Lưu",
     "id=5",
     "PUT /assign/my-todo/todos/5 trả 200. UI cập nhật title ngay.", ""),
    ("Đổi list của todo (move giữa list)", "P1",
     "Todo X ở list A",
     "1. Mở modal sửa\n2. Đổi dropdown list → list B\n3. Lưu",
     "list_id changes",
     "Todo X chuyển sang list B. Sidebar: list A giảm 1, list B tăng 1.", ""),
    ("Xóa todo (cascade sub-task)", "P0",
     "Todo cha có 2 sub-task",
     "1. Click icon Xoá trên todo cha\n2. Confirm",
     "id=10",
     "DELETE /assign/my-todo/todos/10 trả 200. Todo cha + 2 sub-task biến mất.",
     "BR-04"),
    ("Toggle hoàn thành todo cá nhân", "P0",
     "Todo X is_completed=0",
     "1. Click checkbox bên trái todo X",
     "id=5",
     "PATCH /assign/my-todo/todos/5/toggle. is_completed=1, completed_at=now(). "
     "UI: title gạch ngang, checkbox checked.",
     "BR-07"),
    ("Toggle bỏ hoàn thành", "P1",
     "Todo X is_completed=1",
     "1. Click checkbox X (đang tick)",
     "id=5",
     "Toggle về 0, completed_at=null. Title hết gạch ngang.",
     "BR-07"),
    ("User không sửa/xóa được todo của user khác", "P0",
     "Todo id=99 thuộc user B",
     "Login user A, gọi DELETE /assign/my-todo/todos/99",
     "user A, id=99",
     "API trả 404/403. Todo vẫn còn.",
     "BR-05"),
    ("Tạo todo có due_date sai format → lỗi", "P2",
     "Đang ở modal",
     "1. Nhập date không hợp lệ qua API",
     "due_date='abc'",
     "API trả 422 với message 'due_date phải có format Y-m-d'.", ""),
]))

# --- SECTION VII: SUB-TASK ---
SECTIONS.append(("VII. SUB-TASK (BƯỚC CON)", [
    ("Thêm sub-task inline trong list detail", "P0",
     "Đang xem chi tiết 1 list, có todo X",
     "1. Click '+ Thêm bước con' dưới todo X\n2. Nhập 'Bước 1' + Enter",
     "parent_id=5, title='Bước 1'",
     "POST /assign/my-todo/todos với parent_id=X. Sub-task xuất hiện thụt vào dưới todo X. "
     "Todo X hiển thị thêm indicator có sub-task.",
     "BR-01, BR-08"),
    ("Tạo sub-task cấp 2 → BE chặn", "P0",
     "Có sub-task Y (parent_id != null)",
     "Gọi API POST /assign/my-todo/todos với parent_id=Y",
     "parent_id=Y (Y là sub-task)",
     "API trả 400/422 với message 'Không thể tạo sub-task của sub-task'. Không tạo record.",
     "BR-01"),
    ("Toggle hoàn thành sub-task", "P1",
     "Sub-task Y is_completed=0",
     "1. Click checkbox bên cạnh sub-task Y",
     "id=Y",
     "PATCH .../todos/Y/toggle. UI: checkbox tick, title gạch ngang.",
     "BR-07"),
    ("Xóa sub-task", "P1",
     "Sub-task Y dưới todo X",
     "1. Click icon Xoá sub-task Y",
     "id=Y",
     "DELETE /assign/my-todo/todos/Y. Sub-task biến mất, todo X vẫn còn.", ""),
    ("Hoàn thành tất cả sub-task → todo cha KHÔNG tự hoàn thành", "P1",
     "Todo X có 3 sub-task, đánh hoàn thành cả 3",
     "1. Tick lần lượt 3 sub-task",
     "—",
     "3 sub-task is_completed=1. Todo X vẫn is_completed=0 (không có logic auto-complete cha).", ""),
    ("Trong list detail, sub-task không xuất hiện ở danh sách todo cấp 1", "P2",
     "Todo X có 2 sub-task Y, Z",
     "1. Mở list detail",
     "—",
     "Cấp 1 chỉ hiện todo X. Y, Z chỉ hiện khi expand X (lồng vào trong X).", ""),
]))

# --- SECTION VIII: MÀN CHI TIẾT LIST ---
SECTIONS.append(("VIII. MÀN CHI TIẾT DANH SÁCH", [
    ("Mở chi tiết list", "P0",
     "Có ít nhất 1 list",
     "1. Click tên list trên sidebar",
     "list_id=1",
     "GET /assign/my-todo/lists/1/todos. Vùng chính chuyển sang TodoListDetail: "
     "header (tên + mô tả + stats 'X/Y') + danh sách todo (pending + completed) + input thêm mới ở cuối.", ""),
    ("Toggle quay về trang tổng hợp", "P1",
     "Đang ở TodoListDetail của list A",
     "1. Click lại tên list A trên sidebar",
     "—",
     "Vùng chính quay lại TodoMainList (grouped view). Sidebar hết highlight list A.", ""),
    ("Inline add todo bằng Enter", "P0",
     "Đang ở TodoListDetail",
     "1. Focus input '+ Thêm việc mới'\n2. Gõ 'Việc A' + Enter",
     "title='Việc A'",
     "POST /assign/my-todo/todos với list_id hiện tại. Todo xuất hiện cuối danh sách pending. "
     "Input clear, sẵn sàng nhập tiếp.", ""),
    ("Phân nhóm Pending / Completed", "P1",
     "List có 3 todo pending + 2 todo completed",
     "1. Mở chi tiết list",
     "—",
     "Hiển thị block 'Đang làm' chứa 3 todo + block 'Đã hoàn thành' (collapse được) chứa 2 todo.", ""),
    ("Sửa list từ header detail", "P1",
     "Đang ở TodoListDetail",
     "1. Click 'Sửa' trên header",
     "—",
     "Mở TodoListFormModal pre-fill tên + mô tả hiện tại. Lưu → cập nhật header ngay.", ""),
    ("Xóa list từ header detail", "P0",
     "Đang ở TodoListDetail",
     "1. Click 'Xoá' trên header → confirm",
     "—",
     "DELETE /lists/{id}. Quay về trang tổng hợp. List biến mất khỏi sidebar.", ""),
]))

# --- SECTION IX: TƯƠNG TÁC TASK/ISSUE ---
SECTIONS.append(("IX. ACTION BUTTONS TRÊN TASK / ISSUE", [
    ("Hover Task hiện action buttons", "P1",
     "Có Task trong danh sách",
     "1. Di chuột hover lên item Task",
     "—",
     "Hiện 5 nút bên phải: 'Xem', 'Sửa' (nếu can_edit), 'Nhập KQ' (nếu can_import_result), "
     "'Duyệt' (nếu can_approve), 'Lịch sử'. Nút ẩn khi không hover.", ""),
    ("Click 'Xem' Task status 1-3 → mở CreateTaskModal view", "P0",
     "Task status=3",
     "1. Hover Task\n2. Click 'Xem'",
     "task status=3",
     "Mở CreateTaskModal ở chế độ view (form chỉ đọc).", ""),
    ("Click 'Xem' Task status 4-9 → mở ImportResultModal", "P1",
     "Task status=4 (Đang làm)",
     "1. Click 'Xem'",
     "—",
     "Mở ImportResultModal ở chế độ view.", ""),
    ("Click 'Sửa' → mở CreateTaskModal edit", "P1",
     "Task có can_edit=true",
     "1. Click 'Sửa'",
     "—",
     "Mở CreateTaskModal chế độ edit (form sửa được). Sau lưu, danh sách reload.", ""),
    ("Click 'Nhập KQ' → mở ImportResultModal", "P0",
     "Task có can_import_result=true",
     "1. Click 'Nhập KQ'",
     "—",
     "Mở ImportResultModal cho phép nhập kết quả + đính kèm.", ""),
    ("Click 'Duyệt' → mở ImportResultModal duyệt", "P1",
     "Task có can_approve=true",
     "1. Click 'Duyệt'",
     "—",
     "Mở ImportResultModal chế độ duyệt (có nút Đồng ý / Từ chối).", ""),
    ("Click 'Lịch sử' → mở TaskHistoryModal", "P1",
     "Task bất kỳ",
     "1. Click 'Lịch sử'",
     "—",
     "Mở TaskHistoryModal hiện log thay đổi trạng thái + ý kiến.", ""),
    ("Action buttons không hiện khi không có quyền", "P0",
     "Task can_edit=false, can_import_result=false, can_approve=false",
     "1. Hover Task",
     "—",
     "Chỉ hiện 'Xem' + 'Lịch sử'. 3 nút Sửa/Nhập KQ/Duyệt ẩn.", ""),
    ("Hover Issue hiện action buttons", "P1",
     "Có Issue trong danh sách",
     "1. Hover Issue",
     "—",
     "Hiện: Xem, Sửa (nếu can_edit), Xử lý (nếu can_handle), Lịch sử.", ""),
    ("Click Xử lý Issue → mở CreateIssueModal", "P1",
     "Issue can_handle=true",
     "1. Click 'Xử lý'",
     "—",
     "Mở CreateIssueModal chế độ xử lý.", ""),
    ("Click AssignJob/AssignBusiness/Meeting → mở tab mới", "P1",
     "Có item AssignJob",
     "1. Click vào item",
     "—",
     "window.open(item.url, '_blank'). Tab mới mở trang chi tiết tương ứng. "
     "Không mở modal inline.", ""),
]))

# --- SECTION X: DEDUPLICATE & EDGE CASES ---
SECTIONS.append(("X. DEDUPLICATE & EDGE CASES", [
    ("Deduplicate: user vừa assignee vừa watcher", "P0",
     "Task X có user_id ở cả assignee và watcher",
     "1. Truy cập /assign/my-todo\n2. Tìm Task X",
     "user là assignee + watcher cùng entity",
     "Task X CHỈ xuất hiện 1 lần trong danh sách với role badge 'Được giao' (ưu tiên 1, không phải 'Theo dõi' = 3).",
     "BR-06"),
    ("Deduplicate ưu tiên Approver hơn Theo dõi", "P1",
     "Issue có user là approver + watcher",
     "1. Tìm Issue trong danh sách",
     "—",
     "Issue hiện 1 lần, role = 'Cần duyệt' (ưu tiên 2).", "BR-06"),
    ("Pagination tổng hợp (nếu > 100 items)", "P2",
     "User có 150 item",
     "1. Truy cập trang\n2. Cuộn xuống cuối danh sách",
     "—",
     "Trang load đủ items (nếu thiết kế load all) hoặc có lazy load/infinite scroll tuỳ implementation. "
     "Không bị crash UI.", ""),
    ("Empty state khi user không có việc gì", "P1",
     "User mới hoàn toàn, không có Task/Issue/Personal",
     "1. Truy cập /assign/my-todo",
     "—",
     "Hiển thị 4 nhóm cơ bản với '(0)'. Phần body từng nhóm hiện text 'Không có việc nào'. "
     "Stats header = 0/0/0/0.", ""),
    ("Trùng nhiều entity vẫn không double-count trong stats", "P2",
     "User dedup 2 record về 1",
     "1. Quan sát số liệu stats",
     "—",
     "Stats đếm theo items sau dedup, không double count.",
     "BR-06"),
    ("Phân biệt Personal Todo và Task qua icon", "P2",
     "User có cả Personal và Task tiêu đề giống nhau",
     "1. Quan sát icon trái",
     "—",
     "Personal có checkbox, Task không có. Viền màu khác (xám vs xanh dương).", ""),
    ("Tạo todo trong list của user khác → BE chặn", "P0",
     "List id=99 thuộc user B",
     "Login user A, POST /todos với list_id=99",
     "list_id=99 (user B)",
     "API trả 404/403. Không tạo todo.",
     "BR-05"),
    ("Concurrent: 2 tab cùng toggle 1 todo", "P2",
     "Mở 2 tab trên cùng todo X (is_completed=0)",
     "1. Tab A toggle\n2. Tab B toggle ngay sau",
     "—",
     "Sau cả 2 lần toggle, todo X is_completed = 0 (toggle 2 lần). Không lỗi. "
     "Reload trang khớp với state cuối.", ""),
    ("Calendar với tháng không có item nào", "P2",
     "Mở tháng tương lai 12/2027 (chưa có item)",
     "1. Chuyển calendar đến 12/2027",
     "—",
     "Calendar không có dot. API GET ?month=2027-12 trả calendar_summary rỗng. UI không lỗi.", ""),
    ("Refresh trang vẫn giữ filter URL", "P2",
     "Đã chọn filter type=task",
     "1. F5 refresh",
     "—",
     "Filter UI reset về mặc định (vì state không lưu trong URL/localStorage — tuỳ design).", ""),
]))

# --- SECTION XI: E2E FLOW ---
SECTIONS.append(("XI. E2E FLOW TỔNG QUÁT", [
    ("E2E Tạo list + Tạo todo + Tạo sub-task + Toggle + Xoá", "P0",
     "User đã đăng nhập",
     "1. Truy cập /assign/my-todo\n2. Tạo list 'Sprint 1'\n3. Click vào list 'Sprint 1'\n"
     "4. Inline add 3 todo: 'A', 'B', 'C'\n5. Mở todo 'A', thêm 2 sub-task\n"
     "6. Tick hoàn thành todo 'B'\n7. Xóa todo 'C'\n8. Quay về trang tổng hợp\n"
     "9. Kiểm tra item A, B, sub-task trong các nhóm",
     "user full flow",
     "Mọi action thành công, không lỗi. Sidebar list 'Sprint 1' hiện '1/2' (1 done / 2 active cấp 1). "
     "Trang tổng hợp hiển thị 'A' + sub-task lồng trong nhóm phù hợp. 'B' hiển thị gạch ngang. 'C' biến mất hoàn toàn.",
     ""),
    ("E2E Lọc + Click ngày calendar + Toggle filter", "P1",
     "User có nhiều item",
     "1. Chọn Loại = Task\n2. Click ngày 03/05 trên calendar\n3. Đổi Loại = Tất cả\n4. Click lại 03/05",
     "—",
     "Mỗi bước UI cập nhật đúng. Cuối cùng quay về grouped view tất cả loại.", ""),
    ("E2E Sửa list từ sidebar + Đổi tên realtime", "P1",
     "Có list 'A'",
     "1. Hover list → icon Sửa\n2. Click Sửa\n3. Đổi tên = 'A2' → Lưu",
     "—",
     "Modal đóng, sidebar đổi 'A' → 'A2' ngay. Nếu đang mở detail của list này, header cũng update.", ""),
]))


# ============ GENERATOR ============
def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LichLamViec"

    # Column widths (15 cột A-O)
    col_widths = {
        'A': 22, 'B': 22, 'C': 16, 'D': 42, 'E': 10,
        'F': 32, 'G': 55, 'H': 22, 'I': 65, 'J': 35,
        'K': 18, 'L': 16, 'M': 16, 'N': 16, 'O': 22,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: Title block header
    ws.cell(1, 1, "MÔ TẢ TÍNH NĂNG (đọc trước khi xem testcase)")
    ws.cell(1, 1).font = font_title
    ws.cell(1, 1).fill = fill_title
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 26

    # Rows 2-9: Description
    for i, (label, content) in enumerate(DESCRIPTION_ROWS, start=2):
        ws.cell(i, 1, label).font = font_desc_label
        ws.cell(i, 1).alignment = align_wrap
        ws.cell(i, 2, content).font = font_desc_text
        ws.cell(i, 2).alignment = align_wrap
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=15)
        # Auto-height heuristic
        lines = content.count("\n") + 1
        ws.row_dimensions[i].height = max(28, min(180, lines * 16))

    # Row 10: blank
    ws.row_dimensions[10].height = 6

    # Row 11: Title + TEST SUMMARY block
    ws.cell(11, 1, FEATURE_TITLE).font = font_title
    ws.cell(11, 1).fill = fill_title
    ws.merge_cells("A11:E11")
    ws.cell(11, 6, "TEST SUMMARY").font = font_summary_label
    ws.cell(11, 6).fill = fill_summary
    ws.cell(11, 6).alignment = align_center
    ws.merge_cells("F11:H11")
    ws.cell(11, 9, "Số trường hợp kiểm thử đạt (P):").font = font_summary_label
    ws.cell(11, 9).fill = fill_summary
    ws.merge_cells("I11:K11")
    ws.cell(11, 12, '=COUNTIF(L18:L500,"Passed")').font = font_summary_value
    ws.cell(11, 12).fill = fill_summary

    ws.cell(12, 9, "Số trường hợp kiểm thử không đạt (F):").font = font_summary_label
    ws.cell(12, 9).fill = fill_summary
    ws.merge_cells("I12:K12")
    ws.cell(12, 12, '=COUNTIF(L18:L500,"Failed")').font = font_summary_value
    ws.cell(12, 12).fill = fill_summary

    ws.cell(13, 9, "Số trường hợp kiểm thử đang xem xét (PE):").font = font_summary_label
    ws.cell(13, 9).fill = fill_summary
    ws.merge_cells("I13:K13")
    ws.cell(13, 12, '=COUNTIF(L18:L500,"Pending")').font = font_summary_value
    ws.cell(13, 12).fill = fill_summary

    ws.cell(14, 9, "Số trường hợp kiểm thử chưa thực hiện:").font = font_summary_label
    ws.cell(14, 9).fill = fill_summary
    ws.merge_cells("I14:K14")
    ws.cell(14, 12, '=COUNTIF(L18:L500,"Not Executed")').font = font_summary_value
    ws.cell(14, 12).fill = fill_summary

    ws.cell(15, 9, "Tổng số trường hợp kiểm thử:").font = font_summary_label
    ws.cell(15, 9).fill = fill_summary
    ws.merge_cells("I15:K15")
    ws.cell(15, 12, '=COUNTIF(L18:L500,"<>")').font = font_summary_value
    ws.cell(15, 12).fill = fill_summary

    for r in range(11, 16):
        ws.row_dimensions[r].height = 22

    # Row 17: Header
    headers = [
        "Module", "Nhóm chức năng", "TC ID", "Chức năng", "Priority",
        "Tiền điều kiện", "Bước thực hiện", "Test Data", "Expected Result (chi tiết)", "Giải thích nghiệp vụ",
        "KQ thực tế", "trạng thái check lần 1", "trạng thái check lần 2", "trạng thái check lần 3", "Ghi chú",
    ]
    HEADER_ROW = 17
    for c, val in enumerate(headers, start=1):
        cell = ws.cell(HEADER_ROW, c, val)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    ws.row_dimensions[HEADER_ROW].height = 36

    # Data Validation (cột L,M,N) — Status dropdown
    dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
    dv.add(f"L18:N500")
    ws.add_data_validation(dv)

    # Write sections + test cases
    current_row = HEADER_ROW + 1
    section_idx = 0
    tc_global = 0

    for section_title, tcs in SECTIONS:
        section_idx += 1
        # Section row: merge C:O
        ws.cell(current_row, 3, section_title).font = font_section
        ws.cell(current_row, 3).fill = fill_section
        ws.cell(current_row, 3).alignment = align_section
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=15)
        # border for section row
        for c in range(1, 16):
            ws.cell(current_row, c).border = border_all
            ws.cell(current_row, c).fill = fill_section
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for tc_idx, tc in enumerate(tcs, start=1):
            tc_global += 1
            func, prio, precond, steps, test_data, expected, note = tc
            tc_id = f"TC_{section_idx:02d}.{tc_idx:03d}"
            values = [
                "Lịch làm việc",   # A Module
                section_title.split(". ", 1)[-1][:40],  # B Nhóm chức năng
                tc_id,                                  # C TC ID
                func,                                   # D Chức năng
                prio,                                   # E Priority
                precond,                                # F Tiền điều kiện
                steps,                                  # G Bước thực hiện
                test_data or "—",                       # H Test Data
                expected,                               # I Expected Result
                note or "",                             # J Giải thích nghiệp vụ
                "",                                     # K KQ thực tế
                "Not Executed",                         # L check lần 1
                "",                                     # M check lần 2
                "",                                     # N check lần 3
                "",                                     # O Ghi chú
            ]
            fill_row = fill_even if tc_global % 2 == 0 else None
            for c, val in enumerate(values, start=1):
                cell = ws.cell(current_row, c, val)
                cell.font = font_body
                cell.alignment = align_wrap if c not in (3, 5, 12, 13, 14) else align_center
                cell.border = border_all
                if fill_row:
                    cell.fill = fill_row
            # Estimate row height
            longest = max(len(str(steps)), len(str(expected)), len(str(precond)))
            ws.row_dimensions[current_row].height = max(40, min(200, longest // 3))
            current_row += 1

    # Freeze panes: keep header row visible
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    wb.save(OUTPUT)
    print(f"✓ Generated: {OUTPUT}")
    print(f"  Total test cases: {tc_global}")
    print(f"  Sections: {section_idx}")


if __name__ == "__main__":
    generate()
