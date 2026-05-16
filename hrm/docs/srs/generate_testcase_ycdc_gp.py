import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "YCDC Giai phap"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Yêu cầu điều chỉnh giải pháp (Assign/YCĐC GP)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'YCĐC Giải pháp'
PREFIX = 'YCDC'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. GIAO DIỆN — MÀN KD (Dự án TKT)
# =============================================
write_section_row('I. GIAO DIỆN — MÀN KD (Dự án TKT > Tab Giải pháp > Tab con YCĐC GP)')

write_tc(1, 1,
    'Hiển thị tab con "Yêu cầu điều chỉnh GP"',
    'P0',
    'User đăng nhập, vào chi tiết Dự án TKT có GP',
    '1. Vào chi tiết Dự án TKT\n2. Click tab "Giải pháp"',
    '',
    'Tab "Giải pháp" tách thành 2 tab con:\n- "Thông tin giải pháp" (mặc định)\n- "Yêu cầu điều chỉnh GP"')

write_tc(1, 2,
    'Hiển thị bảng danh sách YCĐC',
    'P0',
    'Đang ở tab "Yêu cầu điều chỉnh GP"',
    '1. Click tab con "Yêu cầu điều chỉnh GP"',
    '',
    'Bảng hiển thị đầy đủ cột: STT, Mã yêu cầu, Version, Người yêu cầu, Ngày gửi, Trạng thái\nKhông có cột Hành động (màn KD)')

write_tc(1, 3,
    'Hiển thị nút "Tạo yêu cầu" khi user là KD phụ trách + GP đã duyệt',
    'P0',
    'User = main_sale_employee_id, GP status ∈ {11,13,15,17}',
    '1. Vào tab YCĐC GP',
    'GP status = 11 (Đã duyệt GP)',
    'Nút "Tạo yêu cầu" hiển thị ở góc phải trên')

write_tc(1, 4,
    'Ẩn nút "Tạo yêu cầu" khi user KHÔNG phải KD phụ trách',
    'P0',
    'User ≠ main_sale_employee_id của dự án',
    '1. Đăng nhập bằng tài khoản khác\n2. Vào cùng dự án TKT\n3. Vào tab YCĐC GP',
    '',
    'Không hiển thị nút "Tạo yêu cầu"')

write_tc(1, 5,
    'Ẩn nút "Tạo yêu cầu" khi GP chưa đến status cho phép',
    'P0',
    'User = KD phụ trách, GP status < 11 (VD: Nháp, Chờ duyệt)',
    '1. Vào dự án có GP ở status Nháp\n2. Vào tab YCĐC GP',
    'GP status = 1 (Nháp)',
    'Không hiển thị nút "Tạo yêu cầu"')

write_tc(1, 6,
    'Hiển thị empty state khi chưa có YCĐC',
    'P1',
    'Dự án TKT có GP nhưng chưa có YCĐC nào',
    '1. Vào tab YCĐC GP',
    '',
    'Hiển thị: "Chưa có yêu cầu điều chỉnh giải pháp nào."')

write_tc(1, 7,
    'Hiển thị badge trạng thái đúng màu',
    'P0',
    'Có phiếu YCĐC ở cả 3 trạng thái',
    '1. Quan sát cột Trạng thái',
    '3 phiếu status 1, 2, 3',
    'Đã gửi = xanh dương (#2563EB)\nTiếp nhận = xanh lá (#16A34A)\nTừ chối = đỏ (#DC2626)')

write_tc(1, 8,
    'Hiển thị cột Version đúng',
    'P1',
    'Có phiếu YCĐC được tạo ở version V1 và V2',
    '1. Quan sát cột Version',
    '',
    'Hiển thị "V1", "V2" tương ứng với version tại thời điểm tạo phiếu')

write_tc(1, 9,
    'Click mã yêu cầu mở popup chi tiết',
    'P0',
    'Có ít nhất 1 phiếu YCĐC',
    '1. Click vào mã yêu cầu (VD: YCDCGP.00001)',
    '',
    'Mở popup "Chi tiết yêu cầu điều chỉnh giải pháp" dạng bảng\nHiển thị: Mã, GP, Người yêu cầu, Ngày gửi, Trạng thái, Nội dung')

write_tc(1, 10,
    'Popup chi tiết hiển thị dạng bảng text gọn',
    'P1',
    'Mở popup chi tiết 1 phiếu',
    '1. Click mã yêu cầu\n2. Quan sát nội dung popup',
    '',
    'Popup hiển thị dạng bảng (table bordered), mỗi dòng: label bold + giá trị text\nKhông dùng input/textarea disabled')

write_tc(1, 11,
    'Popup chi tiết hiện lý do từ chối khi status = Từ chối',
    'P0',
    'Có phiếu YCĐC bị từ chối',
    '1. Click mã phiếu bị từ chối',
    'Phiếu status = 3',
    'Popup hiện thêm dòng "Lý do từ chối" + "Người xử lý — Ngày xử lý"')

write_tc(1, 12,
    'Popup chi tiết hiện file đính kèm',
    'P1',
    'Phiếu YCĐC có file đính kèm',
    '1. Click mã phiếu có file đính kèm',
    '',
    'Popup hiện thêm dòng "File đính kèm" với bảng file')

write_tc(1, 13,
    'Phân trang hoạt động',
    'P1',
    'Có > 20 phiếu YCĐC',
    '1. Quan sát phân trang ở dưới bảng\n2. Click trang 2',
    '',
    'Bảng chuyển sang trang 2, STT tính tiếp (21, 22...)')


# =============================================
# II. TẠO YÊU CẦU ĐIỀU CHỈNH
# =============================================
write_section_row('II. TẠO YÊU CẦU ĐIỀU CHỈNH')

write_tc(2, 1,
    'Mở popup tạo yêu cầu',
    'P0',
    'User = KD phụ trách, GP status = 11',
    '1. Click nút "Tạo yêu cầu"',
    '',
    'Popup "Tạo yêu cầu điều chỉnh giải pháp" hiện ra\nGiải pháp: readonly, hiện mã + tên + version\nNội dung: textarea trống\nFile đính kèm: FileAttachmentTable editable\nFooter: nút "Gửi" + "Đóng"')

write_tc(2, 2,
    'Tạo YCĐC thành công (happy path)',
    'P0',
    'Popup tạo đang mở',
    '1. Nhập nội dung: "Khách hàng yêu cầu thay đổi module A"\n2. Click "Gửi"',
    'Nội dung = "Khách hàng yêu cầu thay đổi module A"',
    'Toast: "Gửi yêu cầu điều chỉnh thành công"\nPopup đóng\nDanh sách reload, hiện phiếu mới với mã YCDCGP.XXXXX, status "Đã gửi"')

write_tc(2, 3,
    'Tạo YCĐC với file đính kèm',
    'P1',
    'Popup tạo đang mở',
    '1. Nhập nội dung điều chỉnh\n2. Thêm file đính kèm\n3. Click "Gửi"',
    'File: document.pdf',
    'Tạo thành công\nXem chi tiết → hiện file đính kèm')

write_tc(2, 4,
    'Validate: nội dung trống',
    'P0',
    'Popup tạo đang mở',
    '1. Để trống nội dung\n2. Click "Gửi"',
    'Nội dung = ""',
    'Hiện lỗi: "Nội dung điều chỉnh không được để trống"\nKhông gọi API')

write_tc(2, 5,
    'Validate: nội dung chỉ chứa khoảng trắng',
    'P1',
    'Popup tạo đang mở',
    '1. Nhập "   " (3 dấu cách)\n2. Click "Gửi"',
    'Nội dung = "   "',
    'Hiện lỗi: "Nội dung điều chỉnh không được để trống"')

write_tc(2, 6,
    'Mã phiếu tự sinh đúng format',
    'P0',
    'Tạo YCĐC thành công',
    '1. Tạo phiếu YCĐC\n2. Quan sát mã yêu cầu trong danh sách',
    '',
    'Mã format: YCDCGP.XXXXX (5 chữ số zero-pad)\nVD: YCDCGP.00001, YCDCGP.00002')

write_tc(2, 7,
    'Version được snapshot tại thời điểm tạo',
    'P0',
    'GP đang ở version V2',
    '1. Tạo YCĐC\n2. Xem chi tiết phiếu vừa tạo',
    'GP current_version = V2',
    'Cột Version hiện "V2"\nChi tiết hiện GP: "GP.XXXXX - Tên GP (V2)"')

write_tc(2, 8,
    'Button "Gửi" hiện loading khi đang submit',
    'P1',
    'Popup tạo đang mở, đã nhập nội dung',
    '1. Click "Gửi"\n2. Quan sát button',
    '',
    'Button "Gửi" hiện loading spinner, không click được lần 2\nButton "Đóng" bị disabled')

write_tc(2, 9,
    'Đóng popup → reset form',
    'P1',
    'Popup tạo đang mở, đã nhập nội dung',
    '1. Nhập nội dung\n2. Click "Đóng"\n3. Mở lại popup',
    '',
    'Form reset: nội dung trống, file trống, không lỗi')

write_tc(2, 10,
    'Tạo phiếu mới sau khi phiếu trước bị từ chối',
    'P0',
    'Có phiếu YCĐC đã bị từ chối',
    '1. Mở tab YCĐC GP (đã có phiếu bị từ chối)\n2. Click "Tạo yêu cầu"\n3. Nhập nội dung mới\n4. Click "Gửi"',
    '',
    'Tạo phiếu mới thành công\nDanh sách hiện cả phiếu cũ (Từ chối) và phiếu mới (Đã gửi)')


# =============================================
# III. GIAO DIỆN — MÀN TP/PM (Quản lý GP)
# =============================================
write_section_row('III. GIAO DIỆN — MÀN TP/PM (Quản lý GP > Tab YC Điều chỉnh)')

write_tc(3, 1,
    'Hiển thị tab "YC Điều chỉnh" trên màn quản lý GP',
    'P0',
    'User đăng nhập, vào quản lý GP /assign/solutions/:id/manager',
    '1. Vào màn quản lý giải pháp\n2. Quan sát các tab',
    '',
    'Có tab "YC Điều chỉnh" với icon ri-edit-2-line')

write_tc(3, 2,
    'Bảng danh sách có cột Hành động',
    'P0',
    'Vào tab "YC Điều chỉnh" trên màn TP/PM',
    '1. Click tab "YC Điều chỉnh"\n2. Quan sát bảng',
    '',
    'Bảng có 7 cột: STT, Mã yêu cầu, Version, Người yêu cầu, Ngày gửi, Trạng thái, Hành động')

write_tc(3, 3,
    'Cột Hành động hiện buttons luôn (không chỉ khi hover)',
    'P0',
    'Có phiếu YCĐC status "Đã gửi"',
    '1. Quan sát cột Hành động (không hover)',
    '',
    'Các icon button luôn hiển thị:\n- 👁 Xem chi tiết\n- ✓ Tiếp nhận (xanh lá)\n- ✗ Từ chối (đỏ)')

write_tc(3, 4,
    'Không hiện nút "Tạo yêu cầu" trên màn TP/PM',
    'P0',
    'User là TP/PM, vào tab YC Điều chỉnh',
    '1. Quan sát góc phải trên',
    '',
    'Không hiển thị nút "Tạo yêu cầu" (isProjectCreator = false)')

write_tc(3, 5,
    'Cột Hành động ẩn Tiếp nhận/Từ chối khi phiếu đã xử lý',
    'P0',
    'Có phiếu status "Tiếp nhận" hoặc "Từ chối"',
    '1. Quan sát cột Hành động của phiếu đã xử lý',
    'Phiếu status = 2 hoặc 3',
    'Chỉ hiện icon 👁 Xem, không hiện ✓ Tiếp nhận / ✗ Từ chối')

write_tc(3, 6,
    'Popup chi tiết có nút Tiếp nhận / Từ chối',
    'P0',
    'Phiếu status = "Đã gửi", user là TP/PM',
    '1. Click 👁 xem chi tiết phiếu "Đã gửi"',
    '',
    'Footer popup hiện 3 nút: "Tiếp nhận" (xanh) | "Từ chối" (đỏ) | "Đóng"')

write_tc(3, 7,
    'Popup chi tiết ẩn nút duyệt khi phiếu đã xử lý',
    'P1',
    'Phiếu status = "Tiếp nhận" hoặc "Từ chối"',
    '1. Click 👁 xem chi tiết phiếu đã xử lý',
    'Phiếu status = 2',
    'Footer popup chỉ hiện nút "Đóng"')


# =============================================
# IV. TIẾP NHẬN YÊU CẦU
# =============================================
write_section_row('IV. TIẾP NHẬN YÊU CẦU')

write_tc(4, 1,
    'Tiếp nhận từ cột Hành động (happy path)',
    'P0',
    'User là TP/PM, phiếu status = "Đã gửi"',
    '1. Click icon ✓ (Tiếp nhận) trên dòng phiếu\n2. Popup xác nhận hiện: "Bạn có chắc muốn tiếp nhận...?"\n3. Click "Đồng ý"',
    '',
    'Toast: "Đã tiếp nhận yêu cầu điều chỉnh"\nDanh sách reload, badge chuyển sang "Tiếp nhận" (xanh lá)\nPhiếu không còn nút Tiếp nhận/Từ chối')

write_tc(4, 2,
    'Tiếp nhận từ popup chi tiết',
    'P0',
    'Mở popup chi tiết phiếu "Đã gửi"',
    '1. Click nút "Tiếp nhận" trong popup chi tiết\n2. Popup chi tiết đóng\n3. Popup xác nhận hiện\n4. Click "Đồng ý"',
    '',
    'Toast: "Đã tiếp nhận yêu cầu điều chỉnh"\nDanh sách reload')

write_tc(4, 3,
    'Hủy tiếp nhận',
    'P1',
    'Popup xác nhận đang hiện',
    '1. Click "Không" trên popup xác nhận',
    '',
    'Popup đóng, phiếu vẫn giữ status "Đã gửi"')

write_tc(4, 4,
    'Ghi nhận người xử lý + thời gian',
    'P0',
    'Vừa tiếp nhận thành công',
    '1. Click mã phiếu vừa tiếp nhận → xem chi tiết',
    '',
    'Hiện dòng "Người xử lý": tên TP/PM — dd/mm/yyyy')


# =============================================
# V. TỪ CHỐI YÊU CẦU
# =============================================
write_section_row('V. TỪ CHỐI YÊU CẦU')

write_tc(5, 1,
    'Từ chối từ cột Hành động (happy path)',
    'P0',
    'User là TP/PM, phiếu status = "Đã gửi"',
    '1. Click icon ✗ (Từ chối) trên dòng phiếu\n2. Popup "Từ chối yêu cầu điều chỉnh" hiện ra\n3. Nhập lý do: "Yêu cầu không phù hợp"\n4. Click "Gửi"',
    'Lý do = "Yêu cầu không phù hợp"',
    'Toast: "Đã từ chối yêu cầu điều chỉnh"\nDanh sách reload, badge chuyển sang "Từ chối" (đỏ)')

write_tc(5, 2,
    'Từ chối từ popup chi tiết',
    'P0',
    'Mở popup chi tiết phiếu "Đã gửi"',
    '1. Click nút "Từ chối" (đỏ) trong popup chi tiết\n2. Popup chi tiết đóng → popup từ chối mở\n3. Nhập lý do\n4. Click "Gửi"',
    '',
    'Toast: "Đã từ chối yêu cầu điều chỉnh"\nDanh sách reload')

write_tc(5, 3,
    'Validate: lý do từ chối trống',
    'P0',
    'Popup từ chối đang mở',
    '1. Để trống lý do\n2. Click "Gửi"',
    'Lý do = ""',
    'Hiện lỗi: "Lý do từ chối không được để trống"\nKhông gọi API')

write_tc(5, 4,
    'Validate: lý do chỉ chứa khoảng trắng',
    'P1',
    'Popup từ chối đang mở',
    '1. Nhập "   " (3 dấu cách)\n2. Click "Gửi"',
    'Lý do = "   "',
    'Hiện lỗi: "Lý do từ chối không được để trống"')

write_tc(5, 5,
    'Lý do từ chối hiện trong popup chi tiết',
    'P0',
    'Phiếu vừa bị từ chối',
    '1. Click mã phiếu bị từ chối → xem chi tiết',
    '',
    'Hiện dòng "Lý do từ chối": nội dung lý do\nHiện dòng "Người xử lý": tên — ngày')

write_tc(5, 6,
    'Hủy từ chối',
    'P1',
    'Popup từ chối đang mở',
    '1. Click "Đóng"',
    '',
    'Popup đóng, phiếu vẫn giữ status "Đã gửi"')


# =============================================
# VI. THÔNG BÁO (NOTIFICATION)
# =============================================
write_section_row('VI. THÔNG BÁO (NOTIFICATION)')

write_tc(6, 1,
    'PM nhận thông báo khi KD tạo YCĐC',
    'P0',
    'GP có pm_id = user PM',
    '1. KD tạo phiếu YCĐC\n2. Đăng nhập bằng tài khoản PM',
    '',
    'PM nhận thông báo:\nTitle: "Yêu cầu điều chỉnh giải pháp mới"\nContent: "Có yêu cầu điều chỉnh giải pháp: YCDCGP.XXXXX - GP: GP.XXXXX - Tên GP"')

write_tc(6, 2,
    'TP phòng nhận thông báo khi KD tạo YCĐC',
    'P0',
    'GP thuộc phòng nhận có TP quản lý',
    '1. KD tạo phiếu YCĐC\n2. Đăng nhập bằng tài khoản TP',
    '',
    'TP nhận thông báo tương tự PM')

write_tc(6, 3,
    'Click thông báo → chuyển đúng màn hình + tab',
    'P0',
    'TP/PM nhận thông báo',
    '1. Click vào thông báo',
    '',
    'Chuyển đến: /assign/solutions/{id}/manager?active_tab=adjustment-requests\nTab "YC Điều chỉnh" được active')

write_tc(6, 4,
    'Không gửi trùng thông báo nếu PM = TP',
    'P1',
    'PM phụ trách GP đồng thời là TP phòng nhận',
    '1. KD tạo phiếu YCĐC\n2. Đăng nhập bằng tài khoản PM/TP',
    '',
    'Chỉ nhận 1 thông báo (không trùng)')


# =============================================
# VII. PHÂN QUYỀN & BẢO MẬT
# =============================================
write_section_row('VII. PHÂN QUYỀN & BẢO MẬT')

write_tc(7, 1,
    'User không phải KD phụ trách → không tạo được YCĐC qua API',
    'P0',
    'User ≠ main_sale_employee_id',
    '1. Gọi API POST .../solution-adjustment-requests với user khác',
    '',
    'HTTP 422: "Chỉ nhân viên KD phụ trách dự án mới được tạo yêu cầu điều chỉnh"')

write_tc(7, 2,
    'User không phải TP/PM → không tiếp nhận được qua API',
    'P0',
    'User không phải pm_id và không phải TP phòng nhận',
    '1. Gọi API PUT .../accept với user khác',
    '',
    'is_can_accept = false trên FE\nBE: Entity kiểm tra isCurrentUserTPOrPM()')

write_tc(7, 3,
    'Tạo YCĐC cho GP status không hợp lệ → bị chặn',
    'P0',
    'GP status = 1 (Nháp) hoặc 3 (Chờ duyệt)',
    '1. Gọi API POST .../solution-adjustment-requests\n   Body: { solution_id: <GP nháp>, content: "test" }',
    'GP status = 1',
    'HTTP 422: "Giải pháp chưa được duyệt, không thể tạo yêu cầu điều chỉnh"')

write_tc(7, 4,
    'Tạo YCĐC cho GP không thuộc dự án → bị chặn',
    'P0',
    'GP thuộc dự án A, nhưng gọi API trên dự án B',
    '1. POST /prospective-projects/{B}/solution-adjustment-requests\n   Body: { solution_id: <GP of A>, content: "test" }',
    '',
    'HTTP 422: "Giải pháp không thuộc dự án này"')


# =============================================
# VIII. E2E FLOW
# =============================================
write_section_row('VIII. E2E FLOW')

write_tc(8, 1,
    'E2E: KD tạo → TP tiếp nhận',
    'P0',
    '2 tài khoản: KD (main_sale) + TP/PM',
    '1. [KD] Vào Dự án TKT > Tab GP > Tab con YCĐC\n2. [KD] Tạo yêu cầu: nhập nội dung + đính kèm file\n3. [KD] Gửi → thành công\n4. [TP/PM] Nhận thông báo → click\n5. [TP/PM] Vào tab YC Điều chỉnh → thấy phiếu "Đã gửi"\n6. [TP/PM] Click Tiếp nhận → Đồng ý\n7. [KD] Reload tab → phiếu chuyển "Tiếp nhận"',
    '',
    'Toàn bộ flow hoạt động:\n- Phiếu tạo đúng mã YCDCGP.XXXXX\n- Notification gửi đúng người\n- Click thông báo vào đúng trang\n- Tiếp nhận chuyển status + ghi nhận người xử lý\n- KD thấy status cập nhật')

write_tc(8, 2,
    'E2E: KD tạo → PM từ chối → KD tạo lại',
    'P0',
    '2 tài khoản: KD + PM',
    '1. [KD] Tạo YCĐC → thành công\n2. [PM] Từ chối: nhập lý do "Không phù hợp"\n3. [KD] Xem chi tiết phiếu → thấy "Từ chối" + lý do\n4. [KD] Tạo phiếu YCĐC mới → thành công\n5. [PM] Nhận thông báo phiếu mới',
    '',
    'KD thấy lý do từ chối\nKD có thể tạo phiếu mới\nPM nhận thông báo phiếu mới')

write_tc(8, 3,
    'E2E: Nhiều phiếu YCĐC cùng 1 GP',
    'P1',
    'GP đã duyệt, KD phụ trách',
    '1. [KD] Tạo YCĐC lần 1 → "Đã gửi"\n2. [KD] Tạo YCĐC lần 2 → "Đã gửi"\n3. [TP] Tiếp nhận phiếu 1\n4. [TP] Từ chối phiếu 2\n5. Quan sát danh sách',
    '',
    'Danh sách hiện 2 phiếu:\n- Phiếu 1: Tiếp nhận (xanh lá)\n- Phiếu 2: Từ chối (đỏ)\nMã phiếu tăng dần')


# === DATA VALIDATION ===
dv = DataValidation(type='list', formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
dv.error = 'Chỉ chọn: Passed, Failed, Pending, Not Executed'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add(f'L8:L{current_row}')

# === SAVE ===
output_path = 'docs/srs/solution-adjustment-request-testcases.xlsx'
wb.save(output_path)
print(f'Saved: {output_path} ({current_row - 7} test cases)')
