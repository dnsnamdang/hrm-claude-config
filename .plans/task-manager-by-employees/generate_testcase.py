import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BC Phan bo NV"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 18, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 50, 'H': 22, 'I': 12, 'J': 55,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Báo cáo phân bổ nguồn lực theo nhân viên (QLDA_BC_V2_11)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER FUNCTIONS ===
current_row = 7

MODULE = 'Giao việc'
GROUP = 'BC phân bổ NV'
PREFIX = 'TMBE'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. TRANG BÁO CÁO - HIỂN THỊ & LAYOUT
# =============================================
write_section_row('I. TRANG BÁO CÁO - HIỂN THỊ & LAYOUT (/assign/report/task-manager-by-employees)')

write_tc(1, 1,
    'Hiển thị trang báo cáo khi truy cập',
    'P0',
    'User đã đăng nhập, có quyền xem báo cáo Giao việc',
    '1. Truy cập menu Giao việc > Báo cáo > Phân bổ nguồn lực theo nhân viên',
    '',
    'Trang hiển thị đúng:\n- Tiêu đề: "Báo cáo phân bổ nguồn lực theo nhân viên"\n- Bộ lọc hiển thị đầy đủ\n- Bảng hiển thị với các cột: STT, Nhân viên, Task, Giờ làm, Giờ task, % sử dụng, Trạng thái, Lịch Gantt')

write_tc(1, 2,
    'Hiển thị sticky columns khi scroll ngang',
    'P1',
    'Trang đã load dữ liệu, có nhiều ngày trên Gantt',
    '1. Kéo scroll ngang sang phải trên bảng báo cáo',
    'Chế độ thời gian: Tháng hiện tại (nhiều ngày)',
    '7 cột đầu (STT → Trạng thái) giữ cố định bên trái khi scroll\nCột Lịch Gantt scroll tự do')

write_tc(1, 3,
    'Hiển thị khi không có dữ liệu',
    'P1',
    'User đã đăng nhập',
    '1. Chọn bộ lọc sao cho không có nhân viên nào\n2. Nhấn Tìm kiếm',
    'Phòng ban/Bộ phận không có nhân viên',
    'Bảng hiển thị trống hoặc thông báo "Không có dữ liệu"\nSummary hiển thị: 0 phòng ban, 0 nhân viên, 0 task')

write_tc(1, 4,
    'Hiển thị range label theo chế độ thời gian',
    'P1',
    'Trang đã load',
    '1. Chọn chế độ thời gian "Tuần hiện tại"\n2. Nhấn Tìm kiếm',
    '',
    'Header hiển thị đúng khoảng thời gian đã chọn (VD: "07/04/2026 - 13/04/2026")')


# =============================================
# II. BỘ LỌC
# =============================================
write_section_row('II. BỘ LỌC')

write_tc(2, 1,
    'Filter theo Công ty',
    'P0',
    'Có ít nhất 2 công ty trong hệ thống, mỗi công ty có nhân viên',
    '1. Chọn một công ty trong bộ lọc\n2. Nhấn Tìm kiếm',
    'Công ty: Công ty A',
    'Chỉ hiển thị nhân viên thuộc Công ty A\nPhòng ban chỉ hiển thị các phòng thuộc Công ty A')

write_tc(2, 2,
    'Filter theo Phòng ban',
    'P0',
    'Đã chọn công ty, công ty có nhiều phòng ban',
    '1. Chọn Công ty\n2. Chọn Phòng ban\n3. Nhấn Tìm kiếm',
    'Phòng ban: Phòng Dự án',
    'Chỉ hiển thị nhân viên thuộc Phòng Dự án')

write_tc(2, 3,
    'Filter theo Bộ phận',
    'P1',
    'Đã chọn công ty + phòng ban, phòng ban có nhiều bộ phận',
    '1. Chọn Công ty > Phòng ban > Bộ phận\n2. Nhấn Tìm kiếm',
    'Bộ phận: Bộ phận Dev',
    'Chỉ hiển thị nhân viên thuộc Bộ phận Dev')

write_tc(2, 4,
    'Filter theo Nhân viên',
    'P1',
    'Hệ thống có nhiều nhân viên',
    '1. Chọn nhân viên cụ thể trong dropdown\n2. Nhấn Tìm kiếm',
    'Nhân viên: Nguyễn Văn A',
    'Chỉ hiển thị duy nhất nhân viên Nguyễn Văn A')

write_tc(2, 5,
    'Filter theo Dự án',
    'P1',
    'Có nhiều dự án, mỗi dự án có task giao cho nhân viên',
    '1. Chọn dự án cụ thể\n2. Nhấn Tìm kiếm',
    'Dự án: PRJ_ERP_01',
    'Chỉ hiển thị task thuộc dự án PRJ_ERP_01\nGiờ task và % sử dụng chỉ tính từ task thuộc dự án đó')

write_tc(2, 6,
    'Chế độ thời gian - Hôm nay',
    'P0',
    'Trang đã load',
    '1. Chọn chế độ thời gian: "Hôm nay"\n2. Nhấn Tìm kiếm',
    '',
    'Từ ngày = Đến ngày = ngày hôm nay\nGantt chỉ hiển thị 1 cột ngày\nField Từ ngày/Đến ngày bị disabled (không cho sửa)')

write_tc(2, 7,
    'Chế độ thời gian - Tuần hiện tại',
    'P0',
    'Trang đã load',
    '1. Chọn chế độ thời gian: "Tuần hiện tại"\n2. Nhấn Tìm kiếm',
    '',
    'Từ ngày = Thứ 2 tuần này, Đến ngày = Chủ nhật tuần này\nGantt hiển thị 7 cột ngày (T2 → CN)\nField Từ ngày/Đến ngày bị disabled')

write_tc(2, 8,
    'Chế độ thời gian - Tháng hiện tại',
    'P0',
    'Trang đã load',
    '1. Chọn chế độ thời gian: "Tháng hiện tại"\n2. Nhấn Tìm kiếm',
    '',
    'Từ ngày = ngày 1 tháng này, Đến ngày = ngày cuối tháng này\nGantt hiển thị đủ số ngày trong tháng\nField Từ ngày/Đến ngày bị disabled')

write_tc(2, 9,
    'Chế độ thời gian - Tuỳ chọn',
    'P0',
    'Trang đã load',
    '1. Chọn chế độ thời gian: "Tuỳ chọn"\n2. Chọn Từ ngày và Đến ngày\n3. Nhấn Tìm kiếm',
    'Từ: 01/04/2026, Đến: 15/04/2026',
    'Field Từ ngày/Đến ngày được bật cho phép chọn\nGantt hiển thị 15 cột ngày từ 01/04 → 15/04')

write_tc(2, 10,
    'Filter khoảng công suất - Dưới 60%',
    'P1',
    'Có nhân viên với utilization < 60%, 60-100%, > 100%',
    '1. Chọn khoảng công suất: "< 60%"\n2. Nhấn Tìm kiếm',
    '',
    'Chỉ hiển thị nhân viên có % sử dụng < 60%\nNhân viên có % >= 60% bị ẩn')

write_tc(2, 11,
    'Filter khoảng công suất - 60% đến 100%',
    'P1',
    'Có nhân viên với utilization trong khoảng 60-100%',
    '1. Chọn khoảng công suất: "60% - 100%"\n2. Nhấn Tìm kiếm',
    '',
    'Chỉ hiển thị nhân viên có 60% <= % sử dụng <= 100%')

write_tc(2, 12,
    'Filter khoảng công suất - Trên 100%',
    'P1',
    'Có nhân viên với utilization > 100%',
    '1. Chọn khoảng công suất: "> 100%"\n2. Nhấn Tìm kiếm',
    '',
    'Chỉ hiển thị nhân viên có % sử dụng > 100%')

write_tc(2, 13,
    'Checkbox "Chỉ NV có task"',
    'P1',
    'Có nhân viên không có task nào trong kỳ',
    '1. Tick checkbox "Chỉ NV có task"\n2. Nhấn Tìm kiếm',
    '',
    'Chỉ hiển thị nhân viên có ít nhất 1 task trong khoảng thời gian\nNhân viên 0 task bị ẩn')

write_tc(2, 14,
    'Nút Làm mới (Reset filter)',
    'P1',
    'Đã chọn nhiều filter',
    '1. Chọn nhiều bộ lọc (công ty, phòng ban, dự án, khoảng công suất...)\n2. Nhấn nút "Làm mới"',
    '',
    'Tất cả filter được reset về giá trị mặc định\nChế độ thời gian reset về "Tháng hiện tại" hoặc mặc định ban đầu\nBảng dữ liệu load lại')

write_tc(2, 15,
    'Validate bắt buộc Từ ngày / Đến ngày',
    'P0',
    'Trang đã load',
    '1. Chọn chế độ "Tuỳ chọn"\n2. Để trống Từ ngày hoặc Đến ngày\n3. Nhấn Tìm kiếm',
    '',
    'Hiển thị thông báo lỗi: "Vui lòng chọn khoảng thời gian (Từ ngày, Đến ngày)"\nKhông gọi API')


# =============================================
# III. BẢNG DỮ LIỆU & PHÂN TRANG
# =============================================
write_section_row('III. BẢNG DỮ LIỆU & PHÂN TRANG')

write_tc(3, 1,
    'Hiển thị dòng phòng ban (summary row)',
    'P0',
    'Có dữ liệu, phòng ban có nhiều nhân viên',
    '1. Tìm kiếm với filter phù hợp\n2. Quan sát dòng phòng ban',
    '',
    'Dòng phòng ban hiển thị:\n- Tên phòng ban\n- Tổng số task (sum tất cả NV)\n- Tổng giờ làm (sum capacity)\n- Tổng giờ task\n- % sử dụng trung bình\n- Nền màu xám để phân biệt')

write_tc(3, 2,
    'Expand/Collapse phòng ban',
    'P0',
    'Bảng hiển thị dữ liệu có nhiều phòng ban',
    '1. Click vào icon mở rộng ở dòng phòng ban\n2. Quan sát danh sách nhân viên\n3. Click lại icon để thu gọn',
    '',
    'Click mở rộng → hiển thị danh sách nhân viên bên dưới phòng ban\nClick thu gọn → ẩn danh sách nhân viên\nIcon thay đổi giữa expand/collapse')

write_tc(3, 3,
    'Hiển thị dòng nhân viên',
    'P0',
    'Phòng ban đã expand',
    '1. Expand phòng ban\n2. Quan sát dòng nhân viên',
    '',
    'Mỗi dòng nhân viên hiển thị:\n- STT trong phòng ban\n- Tên nhân viên + Mã NV + Chức vụ\n- Số task\n- Giờ làm (capacity từ ca làm việc)\n- Giờ task (SUM estimated_hours)\n- % sử dụng\n- Trạng thái (badge màu)')

write_tc(3, 4,
    'Trạng thái utilization - Cần phân thêm task (< 60%)',
    'P0',
    'Có nhân viên với utilization < 60%',
    '1. Quan sát cột Trạng thái của nhân viên có % < 60%',
    'NV có 10h task / 40h capacity = 25%',
    'Badge hiển thị "Cần phân thêm task" với màu amber (#f59e0b)')

write_tc(3, 5,
    'Trạng thái utilization - Phân bổ hợp lý (60-100%)',
    'P0',
    'Có nhân viên với utilization 60-100%',
    '1. Quan sát cột Trạng thái của nhân viên có 60% <= % <= 100%',
    'NV có 32h task / 40h capacity = 80%',
    'Badge hiển thị "Phân bổ hợp lý" với màu xanh lá (#16a34a)')

write_tc(3, 6,
    'Trạng thái utilization - Vượt tải (> 100%)',
    'P0',
    'Có nhân viên với utilization > 100%',
    '1. Quan sát cột Trạng thái của nhân viên có % > 100%',
    'NV có 50h task / 40h capacity = 125%',
    'Badge hiển thị "Vượt tải" với màu đỏ (#ef4444)')

write_tc(3, 7,
    'Tính giờ làm (capacity) từ ca làm việc',
    'P0',
    'Nhân viên đã được phân ca làm việc trong kỳ',
    '1. Tìm kiếm với khoảng thời gian cụ thể\n2. So sánh cột "Giờ làm" với phân ca thực tế của NV',
    'NV A được phân ca 8h/ngày x 5 ngày = 40h\nNV B được phân ca 4h/ngày x 5 ngày = 20h',
    'Cột "Giờ làm" = tổng labour_hour từ phân ca thực tế, KHÔNG phải cố định 8h/ngày\nNV A: 40h, NV B: 20h')

write_tc(3, 8,
    'Nhân viên chưa được phân ca - capacity = 0',
    'P1',
    'Có nhân viên chưa được phân ca trong kỳ',
    '1. Tìm kiếm khoảng thời gian mà NV chưa có phân ca\n2. Quan sát cột Giờ làm và % sử dụng',
    '',
    'Cột Giờ làm hiển thị 0\n% sử dụng hiển thị 0%\nKhông bị lỗi chia cho 0')

write_tc(3, 9,
    'Phân trang - Chuyển trang',
    'P1',
    'Có nhiều phòng ban (> 10)',
    '1. Quan sát pagination ở cuối bảng\n2. Click trang 2',
    '',
    'Bảng hiển thị các phòng ban ở trang 2\nPagination cập nhật trang hiện tại')

write_tc(3, 10,
    'Phân trang - Thay đổi số dòng/trang',
    'P1',
    'Có nhiều phòng ban',
    '1. Thay đổi dropdown số dòng/trang (5, 10, 20, 50)\n2. Quan sát bảng',
    'Chọn: 5 dòng/trang',
    'Bảng hiển thị tối đa 5 phòng ban\nSố trang thay đổi tương ứng')

write_tc(3, 11,
    'Click vào số task → mở modal chi tiết',
    'P0',
    'Nhân viên có task trong kỳ',
    '1. Click vào số task của nhân viên',
    '',
    'Modal chi tiết task mở ra\nHiển thị danh sách task của nhân viên đó')

write_tc(3, 12,
    'Click vào giờ task → mở modal chi tiết',
    'P1',
    'Nhân viên có task trong kỳ',
    '1. Click vào giờ task của nhân viên',
    '',
    'Modal chi tiết task mở ra\nHiển thị danh sách task của nhân viên đó')

write_tc(3, 13,
    'Task loại trừ trạng thái Nháp và Huỷ',
    'P0',
    'Nhân viên có task ở các trạng thái khác nhau (Nháp, Đang thực hiện, Huỷ...)',
    '1. Tìm kiếm và quan sát số task của nhân viên\n2. So sánh với DB',
    'NV có 5 task: 1 Nháp, 3 Đang thực hiện, 1 Huỷ',
    'Chỉ đếm 3 task (Đang thực hiện)\nTask Nháp và Huỷ bị loại trừ khỏi báo cáo')


# =============================================
# IV. BIỂU ĐỒ GANTT
# =============================================
write_section_row('IV. BIỂU ĐỒ GANTT')

write_tc(4, 1,
    'Hiển thị Gantt chart đúng vị trí theo ngày',
    'P0',
    'Nhân viên có task với start_date và due_date trong kỳ',
    '1. Expand phòng ban\n2. Quan sát Gantt chart của nhân viên',
    'Task A: 01/04 → 05/04\nTask B: 03/04 → 08/04',
    'Thanh Gantt của mỗi task bắt đầu và kết thúc đúng cột ngày\nTask chồng ngày được xếp trên các lane khác nhau (không đè lên nhau)')

write_tc(4, 2,
    'Màu thanh Gantt - Bình thường (xanh lá)',
    'P0',
    'Nhân viên có task chưa đến hạn',
    '1. Quan sát thanh Gantt của task có due_date > hôm nay',
    'Task due_date: 20/04/2026 (tương lai)',
    'Thanh Gantt màu xanh lá (#22c55e)')

write_tc(4, 3,
    'Màu thanh Gantt - Sắp hạn (cam)',
    'P0',
    'Nhân viên có task đến hạn hôm nay',
    '1. Quan sát thanh Gantt của task có due_date = hôm nay',
    'Task due_date: ngày hôm nay',
    'Thanh Gantt màu cam (#f97316)')

write_tc(4, 4,
    'Màu thanh Gantt - Quá hạn (đỏ)',
    'P0',
    'Nhân viên có task quá hạn chưa hoàn thành',
    '1. Quan sát thanh Gantt của task có due_date < hôm nay và status ≠ Hoàn thành',
    'Task due_date: 05/04/2026 (quá khứ), status: Đang thực hiện',
    'Thanh Gantt màu đỏ (#ef4444)')

write_tc(4, 5,
    'Task quá hạn nhưng đã hoàn thành → không đỏ',
    'P1',
    'Nhân viên có task quá hạn nhưng status = Hoàn thành (8)',
    '1. Quan sát thanh Gantt của task đã hoàn thành dù quá due_date',
    'Task due_date: 05/04 (quá khứ), status: Hoàn thành',
    'Thanh Gantt KHÔNG hiển thị màu đỏ (hiển thị màu xanh bình thường)')

write_tc(4, 6,
    'Highlight ngày hôm nay',
    'P1',
    'Khoảng thời gian bao gồm ngày hôm nay',
    '1. Chọn "Tuần hiện tại" hoặc "Tháng hiện tại"\n2. Quan sát Gantt header',
    '',
    'Cột ngày hôm nay có nền màu xanh nhạt để phân biệt với các ngày khác')

write_tc(4, 7,
    'Nút Ẩn/Hiện Gantt',
    'P1',
    'Bảng đã load dữ liệu',
    '1. Click nút "Ẩn Gantt" ở header bảng\n2. Quan sát bảng\n3. Click "Hiện Gantt"',
    '',
    'Click "Ẩn Gantt" → cột Lịch Gantt bị ẩn, bảng chỉ hiển thị 7 cột data\nClick "Hiện Gantt" → cột Lịch Gantt hiện lại')

write_tc(4, 8,
    'Click thanh Gantt → mở modal chi tiết',
    'P0',
    'Nhân viên có task hiển thị trên Gantt',
    '1. Click vào thanh Gantt của 1 task',
    '',
    'Modal chi tiết task mở ra\nHiển thị thông tin task được click')

write_tc(4, 9,
    'Task vượt ra ngoài khoảng thời gian (clamp)',
    'P1',
    'Task có start_date < from_date hoặc due_date > to_date',
    '1. Chọn khoảng thời gian 07/04 → 13/04\n2. Quan sát task có start_date: 01/04, due_date: 20/04',
    'Task: start 01/04, due 20/04\nKỳ: 07/04 → 13/04',
    'Thanh Gantt bị cắt (clamp) vào khoảng hiển thị\nBắt đầu từ cột 07/04 và kéo đến cột 13/04')

write_tc(4, 10,
    'Gantt hiển thị header ngày đúng (thứ + ngày)',
    'P1',
    'Bảng đã load dữ liệu',
    '1. Quan sát header Gantt chart',
    'Tuần 07/04 → 13/04',
    'Mỗi cột ngày hiển thị:\n- Hàng 1: Tên thứ (T2, T3... CN)\n- Hàng 2: Số ngày (7, 8, 9...)')


# =============================================
# V. MODAL CHI TIẾT TASK
# =============================================
write_section_row('V. MODAL CHI TIẾT TASK')

write_tc(5, 1,
    'Hiển thị modal chi tiết đúng layout',
    'P0',
    'Nhân viên có task trong kỳ',
    '1. Click vào số task hoặc thanh Gantt của nhân viên\n2. Quan sát modal',
    '',
    'Modal hiển thị:\n- Tiêu đề: Tên nhân viên\n- Phụ đề: "Chi tiết task trong kỳ"\n- Bảng task với cột: STT, Mã task, Tên task, Dự án, Ngày bắt đầu, Hạn, Trạng thái, Tiến độ, Giờ ước tính')

write_tc(5, 2,
    'Hiển thị footer tổng hợp trong modal',
    'P0',
    'Modal chi tiết đang mở',
    '1. Cuộn xuống cuối modal\n2. Quan sát footer',
    '',
    'Footer hiển thị:\n- Tổng số task (đếm đúng)\n- Tổng giờ ước tính (SUM estimated_hours)')

write_tc(5, 3,
    'Format dữ liệu trong modal',
    'P1',
    'Modal chi tiết đang mở',
    '1. Quan sát format của từng cột trong modal',
    '',
    '- Mã task: font monospace\n- Ngày: format DD-MM-YYYY\n- Trạng thái: badge có màu theo status\n- Tiến độ: hiển thị %\n- Giờ: format số thập phân 1 chữ số')

write_tc(5, 4,
    'Đóng modal',
    'P2',
    'Modal chi tiết đang mở',
    '1. Click nút X ở góc trên phải\n2. Hoặc click nút "Đóng" ở footer',
    '',
    'Modal đóng lại\nBảng báo cáo phía sau không bị thay đổi')

write_tc(5, 5,
    'Trạng thái task hiển thị đúng màu badge',
    'P1',
    'Nhân viên có task ở nhiều trạng thái khác nhau',
    '1. Mở modal chi tiết\n2. Quan sát cột Trạng thái',
    'Task có status: Đang thực hiện, Hoàn thành, Tạm dừng',
    'Mỗi status hiển thị badge đúng màu:\n- Đang thực hiện: blue\n- Hoàn thành: green\n- Tạm dừng: yellow\n- Chờ phê duyệt: orange\n- Từ chối KQ: red')


# =============================================
# VI. XUẤT EXCEL & IN BÁO CÁO
# =============================================
write_section_row('VI. XUẤT EXCEL & IN BÁO CÁO')

write_tc(6, 1,
    'Xuất Excel - Export đúng dữ liệu',
    'P0',
    'Bảng đã hiển thị dữ liệu với filter',
    '1. Nhấn nút "Xuất Excel"\n2. Mở file Excel được tải về',
    '',
    'File Excel chứa đúng dữ liệu theo filter đã chọn\nCấu trúc: Phòng ban → Nhân viên → Task\nCác cột: STT, Nhân viên, Task, Giờ làm, Giờ task, % sử dụng, Trạng thái')

write_tc(6, 2,
    'Xuất Excel - Không có dữ liệu',
    'P1',
    'Bảng không có dữ liệu (filter không khớp)',
    '1. Nhấn nút "Xuất Excel"',
    '',
    'Hiển thị thông báo hoặc file Excel trống/không tải về')

write_tc(6, 3,
    'In báo cáo',
    'P2',
    'Bảng đã hiển thị dữ liệu',
    '1. Nhấn nút "In"\n2. Quan sát bản xem trước in',
    '',
    'Bản in hiển thị đúng layout:\n- Tiêu đề báo cáo\n- Bảng dữ liệu (không có Gantt hoặc Gantt được format phù hợp)')


# === DATA VALIDATION cho cột Status ===
dv = DataValidation(
    type='list',
    formula1='"Passed,Failed,Pending,Not Executed"',
    allow_blank=True,
)
dv.error = 'Chọn 1 trong 4 giá trị'
dv.errorTitle = 'Giá trị không hợp lệ'
ws.add_data_validation(dv)
dv.add('L8:L500')

# === SAVE ===
output_path = r'd:\CompanyProject\hrm\hrm-claude-config\.plans\task-manager-by-employees\Testcase_BC_Phan_Bo_NV.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Total testcases: {current_row - 7 - 6}')  # trừ 6 section rows
