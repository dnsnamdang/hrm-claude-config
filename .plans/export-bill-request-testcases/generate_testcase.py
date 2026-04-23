import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Testcase Export Bill Request"

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
data_font = Font(size=11)
data_alignment = Alignment(vertical='top', wrap_text=True)
summary_font = Font(size=11)

col_widths = {'A': 18, 'B': 20, 'C': 16, 'D': 42, 'E': 10, 'F': 30, 'G': 55, 'H': 25, 'I': 12, 'J': 60, 'K': 15, 'L': 14, 'M': 20}
for c, w in col_widths.items():
    ws.column_dimensions[c].width = w

ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase - Đề nghị xuất hoá đơn (Export Bill Request)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

for i, (label, formula) in enumerate([
    ('Passed:', '=COUNTIF(L8:L500,"Passed")'),
    ('Failed:', '=COUNTIF(L8:L500,"Failed")'),
    ('Pending:', '=COUNTIF(L8:L500,"Pending")'),
    ('Not Executed:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng:', '=COUNTA(L8:L500)')]):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority', 'Tiền điều kiện', 'Bước thực hiện', 'Test Data', '', 'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

M = 'IncomeExpenditure'
G = 'Đề nghị xuất hoá đơn'
P = 'EBR'
r = 7

def sec(row, title):
    ws.merge_cells(f'C{row}:M{row}')
    c = ws.cell(row=row, column=3, value=title)
    c.font = section_font
    c.fill = section_fill
    for col in range(1, 14):
        ws.cell(row=row, column=col).border = thin_border

def tc(row, tc_id, func, pri, pre='', steps='', td='', exp=''):
    vals = [M, G, tc_id, func, pri, pre, steps, td, '', exp, '', 'Not Executed', '']
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = data_font
        c.alignment = data_alignment
        c.border = thin_border

# ============== I. TRANG DANH SÁCH ==============
sec(r, 'I. TRANG DANH SÁCH ĐỀ NGHỊ XUẤT HOÁ ĐƠN (/admin/income-expenditure/export-bill-request)'); r += 1

tc(r, f'{P}_001.001', 'Hiển thị trang danh sách (tab Của tôi)', 'P0',
   'Đăng nhập, có quyền truy cập module',
   '1. Vào menu Thu/Chi > Đề nghị xuất hoá đơn\n2. Tab mặc định "Của tôi"',
   'type=index',
   'Hiển thị: tiêu đề, nút "Tạo mới", bộ lọc (mã phiếu, từ ngày, đến ngày, trạng thái, công ty/phòng ban/bộ phận), bảng danh sách với các phiếu do chính user tạo'); r += 1

tc(r, f'{P}_001.002', 'Hiển thị tab Tất cả', 'P0',
   'Có quyền xem tất cả',
   '1. Click tab "Tất cả" (URL ?type=all)',
   'type=all',
   'Hiển thị tất cả phiếu theo phạm vi quyền (tổng công ty/công ty/phòng ban). Không bị filter theo created_by'); r += 1

tc(r, f'{P}_001.003', 'Hiển thị tab Đã duyệt (bởi tôi)', 'P0',
   'User có quyền duyệt, đã từng duyệt ≥ 1 phiếu',
   '1. Click tab "Đã duyệt"',
   'type=approved',
   'Hiển thị các phiếu có approver_id = user hiện tại. KHÔNG còn bug "approved_id" → dữ liệu trả về đầy đủ, không rỗng'); r += 1

tc(r, f'{P}_001.004', 'Lọc theo mã phiếu', 'P1',
   '',
   '1. Nhập mã (partial) vào ô tìm\n2. Nhấn tìm kiếm',
   'code: "EBR001"',
   'Chỉ hiện phiếu có code LIKE %EBR001%'); r += 1

tc(r, f'{P}_001.005', 'Lọc theo trạng thái', 'P0',
   '',
   '1. Chọn "Chờ duyệt" trong filter status\n2. Nhấn tìm',
   'status: CHO_DUYET',
   'Chỉ hiện phiếu status = 2 (Chờ duyệt)'); r += 1

tc(r, f'{P}_001.006', 'Lọc theo khoảng thời gian (bao gồm cả ngày đầu và cuối)', 'P0',
   'Có phiếu tạo đúng ngày startDate và endDate',
   '1. Chọn startDate = 2026-04-01\n2. Chọn endDate = 2026-04-16\n3. Nhấn tìm',
   'startDate, endDate',
   'Hiển thị phiếu có created_at nằm TRONG và BAO GỒM cả ngày 2026-04-01 và 2026-04-16 (sau fix bug: whereDate >= / <=)'); r += 1

tc(r, f'{P}_001.007', 'Lọc theo công ty/phòng ban/bộ phận', 'P1',
   '',
   '1. Chọn công ty\n2. Chọn phòng ban\n3. Chọn bộ phận\n4. Nhấn tìm',
   'company_id, department_id, part_id',
   'Chỉ hiện phiếu khớp đúng các filter'); r += 1

tc(r, f'{P}_001.008', 'Lọc theo người tạo', 'P1',
   '',
   '1. Chọn người tạo trong filter',
   'created_by',
   'Chỉ hiện phiếu có created_by khớp'); r += 1

tc(r, f'{P}_001.009', 'Nút thao tác theo trạng thái', 'P0',
   'Có phiếu ở các trạng thái',
   '1. Kiểm tra cột Thao tác theo từng trạng thái',
   '',
   'DANG_TAO (1): Xem, Sửa, Xoá. CHO_DUYET (2): Xem, Duyệt, Từ chối. DA_DUYET (3): Xem (không sửa/xoá). KHONG_DUYET (4): Xem, Sửa, Xoá'); r += 1

tc(r, f'{P}_001.010', 'Phân trang và sort mặc định', 'P1',
   '',
   '1. Xem bảng danh sách',
   '',
   'Sắp xếp mới nhất trước (created_at DESC). Phân trang hoạt động'); r += 1

# ============== II. TẠO MỚI ==============
sec(r, 'II. TẠO MỚI PHIẾU ĐỀ NGHỊ XUẤT HOÁ ĐƠN'); r += 1

tc(r, f'{P}_002.001', 'Mở form tạo mới', 'P0',
   'Có quyền tạo',
   '1. Nhấn "Tạo mới" hoặc vào /create',
   '',
   'Chuyển sang form. Có nút "Lưu" (status=1) và "Lưu và gửi duyệt" (status=2). Các field: contractable, customer, handover_date, note, summary, tabs, attachments (optional)'); r += 1

tc(r, f'{P}_002.002', 'Chọn hợp đồng khách hàng', 'P0',
   '',
   '1. Click chọn hợp đồng\n2. Chọn 1 hợp đồng trong modal',
   'contractable_type, contractable_id',
   'Auto fill: customer_id, customer_code, customer_name, tax_code. contractable_code hiển thị đúng'); r += 1

tc(r, f'{P}_002.003', 'Thêm tab và chọn sản phẩm/chi phí/dịch vụ', 'P0',
   'Đã chọn hợp đồng',
   '1. Thêm tab\n2. Trong tab, thêm dòng sản phẩm (hoặc cost/service)\n3. Nhập qty, price',
   '',
   'Các dòng hiện đầy đủ. amount_after_discount, vat_cost, amount_after_vat tính tự động'); r += 1

tc(r, f'{P}_002.004', 'Validate tab rỗng', 'P0',
   '',
   '1. Không thêm tab nào\n2. Nhấn Lưu',
   'tabs = []',
   'Hiện lỗi validateTab yêu cầu có ít nhất 1 tab với ít nhất 1 dòng'); r += 1

tc(r, f'{P}_002.005', 'Lưu nháp không có attachments', 'P0',
   'Đã điền đủ contractable, customer, handover_date, ít nhất 1 tab',
   '1. KHÔNG upload attachments\n2. Nhấn "Lưu" (status=1)',
   'attachments: null',
   'Sau fix bug 10: lưu nháp thành công (attachments = nullable). Trả status=1 (DANG_TAO)'); r += 1

tc(r, f'{P}_002.006', 'Lưu và gửi duyệt - thành công', 'P0',
   'Đã điền đủ + có attachments',
   '1. Nhấn "Lưu và gửi duyệt"',
   'status=2, attachments=[file.pdf]',
   'Tạo phiếu status=CHO_DUYET. Gửi notification đến user có quyền "Kế toán thanh toán". Redirect về danh sách'); r += 1

tc(r, f'{P}_002.007', 'Counter exporting_bill_qty +qty (status=1)', 'P0',
   'contractable có sản phẩm với counter ban đầu = X',
   '1. Tạo phiếu status=1 với 5 qty sản phẩm A\n2. Kiểm tra counter sản phẩm A trên objectable',
   'qty=5',
   'objectable.exporting_bill_qty = X + 5 (tăng 5). exported_bill_qty không đổi'); r += 1

tc(r, f'{P}_002.008', 'Counter cost flag (status=1)', 'P0',
   'contractable có cost với exporting_bill = 0',
   '1. Tạo phiếu status=1 có cost A',
   '',
   'cost A: exporting_bill = 1, exported_bill không đổi'); r += 1

tc(r, f'{P}_002.009', 'Rollback transaction khi exception (sau fix dd)', 'P0',
   'Simulate exception trong syncTabs',
   '1. Tạo phiếu\n2. Gây lỗi (ví dụ objectable không tồn tại)',
   '',
   'Sau fix bug 1 (xoá dd($e)): DB::rollBack() được gọi → không có phiếu được tạo, counter không bị tăng. Response JSON error trả về FE'); r += 1

tc(r, f'{P}_002.010', 'Upload attachments đa dạng format', 'P1',
   '',
   '1. Upload .pdf, .xlsx, .docx, .jpg, .png',
   '',
   'Upload thành công. S3 path được lưu. Attachments hiển thị trong show'); r += 1

tc(r, f'{P}_002.011', 'Upload attachment sai format', 'P1',
   '',
   '1. Upload file .exe hoặc .zip',
   '',
   'Validate lỗi: chỉ chấp nhận pdf,png,jpg,docx,doc,xls,xlsx,jpeg'); r += 1

tc(r, f'{P}_002.012', 'Upload file > 10MB', 'P1',
   '',
   '1. Upload file PDF > 10MB',
   '',
   'Validate lỗi: max:10240'); r += 1

# ============== III. SỬA ==============
sec(r, 'III. SỬA PHIẾU ĐỀ NGHỊ XUẤT HOÁ ĐƠN'); r += 1

tc(r, f'{P}_003.001', 'Mở form sửa với status DANG_TAO', 'P0',
   'Phiếu status=1 do chính user tạo',
   '1. Click Sửa',
   '',
   'Mở form với data đã lưu. canEdit=true → cho phép sửa'); r += 1

tc(r, f'{P}_003.002', 'Không cho sửa phiếu status CHO_DUYET', 'P0',
   'Phiếu status=2',
   '1. Vào URL /edit/{id}',
   '',
   'canEdit=false → không cho vào form sửa (redirect hoặc báo lỗi)'); r += 1

tc(r, f'{P}_003.003', 'Không cho sửa phiếu status DA_DUYET', 'P0',
   'Phiếu status=3',
   '1. Vào URL /edit/{id}',
   '',
   'canEdit=false'); r += 1

tc(r, f'{P}_003.004', 'Nút "Quay lại" đúng URL (sau fix bug 8)', 'P1',
   '',
   '1. Mở form sửa\n2. Click "Quay lại"',
   '',
   'Redirect về ?type=all (KHÔNG phải ?_type=all như trước) → hiển thị đúng tab Tất cả'); r += 1

tc(r, f'{P}_003.005', 'Sửa qty từ 5 xuống 3 (test reverse counter)', 'P0',
   'Phiếu status=1 với qty=5, counter objectable.exporting_bill_qty đã +5',
   '1. Mở sửa\n2. Đổi qty từ 5 xuống 3\n3. Save status=1',
   'qty_old=5, qty_new=3',
   'Sau fix bug 3: counter reverse -5, rồi +3 = NET +3 (không bị +5+3 = +8 như trước). exported_bill_qty không đổi'); r += 1

tc(r, f'{P}_003.006', 'Sửa từ DANG_TAO → CHO_DUYET (test counter)', 'P0',
   'Phiếu status=1 với qty=5, counter +5',
   '1. Mở sửa\n2. Không đổi data\n3. Click "Lưu và gửi duyệt"',
   'status_new=2',
   'Counter reverse -5 (theo status cũ = DANG_TAO), status mới = CHO_DUYET không cộng → counter NET = 0. Phiếu chuyển status=2, gửi notification'); r += 1

tc(r, f'{P}_003.007', 'Sửa phiếu KHONG_DUYET → gửi lại duyệt', 'P1',
   'Phiếu status=4 (bị từ chối)',
   '1. Mở sửa\n2. Sửa note\n3. Lưu và gửi duyệt',
   'status_new=2',
   'canEdit=true cho KHONG_DUYET. Chuyển status=2. Counter xử lý đúng'); r += 1

tc(r, f'{P}_003.008', 'Sửa tab - thêm/xoá tab', 'P0',
   '',
   '1. Mở sửa\n2. Thêm tab mới\n3. Xoá 1 tab cũ\n4. Save',
   '',
   'syncTabs: reverse counter tabs cũ → delete all → insert tabs mới → apply counter mới. Counter NET đúng với tabs mới'); r += 1

tc(r, f'{P}_003.009', 'Sửa attachments - thêm/xoá file', 'P1',
   '',
   '1. Mở sửa\n2. Upload thêm file\n3. Xoá 1 file cũ\n4. Save',
   '',
   'Attachments mới lưu vào S3. File cũ còn trong attachments string nếu không xoá'); r += 1

# ============== IV. XEM CHI TIẾT (SHOW) ==============
sec(r, 'IV. XEM CHI TIẾT PHIẾU'); r += 1

tc(r, f'{P}_004.001', 'Hiển thị chi tiết đầy đủ', 'P0',
   'Phiếu đã tồn tại',
   '1. Click Xem trên danh sách',
   '',
   'Hiển thị: header (mã, trạng thái, người tạo, ngày tạo), thông tin hợp đồng/khách hàng, tabs (sản phẩm/chi phí/dịch vụ), attachments (download link), người duyệt/ngày duyệt (nếu đã duyệt), summary tổng tiền'); r += 1

tc(r, f'{P}_004.002', 'Hiển thị nút theo status', 'P0',
   '',
   '1. Xem chi tiết phiếu các status khác nhau',
   '',
   'DANG_TAO: Sửa + Xoá. CHO_DUYET (nếu canApprove): Duyệt + Từ chối. DA_DUYET: chỉ In/Export. KHONG_DUYET: Sửa + Xoá'); r += 1

tc(r, f'{P}_004.003', 'In phiếu khi objectable đã bị xoá (bug 9)', 'P1',
   'Phiếu có tab_product.objectable_id trỏ đến record đã bị xoá',
   '1. Vào show phiếu\n2. Gọi print function dùng getBillIncomeWithExchangeRateTableAttribute',
   '',
   'Sau fix bug 9: `continue` khi objectable null → KHÔNG crash fatal error. Rows còn lại hiển thị bình thường (row null bị skip)'); r += 1

# ============== V. DUYỆT / TỪ CHỐI ==============
sec(r, 'V. DUYỆT / TỪ CHỐI'); r += 1

tc(r, f'{P}_005.001', 'Duyệt phiếu thành công', 'P0',
   'Phiếu status=CHO_DUYET, user có quyền "Duyệt đề nghị xuất hóa đơn"',
   '1. Vào show\n2. Click Duyệt',
   '',
   'canApprove=true. Update status=3 (DA_DUYET), approved_time=now, approver_id=user. Trả success'); r += 1

tc(r, f'{P}_005.002', 'Không có quyền duyệt', 'P0',
   'User không có quyền duyệt',
   '1. Vào show phiếu CHO_DUYET\n2. Click Duyệt (nếu có)',
   '',
   'canApprove=false → không hiện nút hoặc gọi API trả "Không có quyền"'); r += 1

tc(r, f'{P}_005.003', 'Không duyệt được phiếu status khác CHO_DUYET', 'P0',
   'Phiếu status=3 (DA_DUYET)',
   '1. Gọi API approve trực tiếp',
   '',
   'canApprove=false → lỗi "Không có quyền"'); r += 1

tc(r, f'{P}_005.004', 'Từ chối phiếu (sau fix bug 4)', 'P0',
   'Phiếu CHO_DUYET, user có quyền',
   '1. Click Từ chối',
   '',
   'canReject=true (method mới, logic giống canApprove nhưng tách semantics). Update status=KHONG_DUYET (4), approved_time=now, approver_id=user'); r += 1

tc(r, f'{P}_005.005', 'Không từ chối được phiếu DA_DUYET', 'P0',
   'Phiếu status=3',
   '1. Gọi API reject',
   '',
   'canReject=false → lỗi "Không có quyền"'); r += 1

tc(r, f'{P}_005.006', 'Duyệt: counter chuyển từ exporting → exported', 'P0',
   'Phiếu CHO_DUYET có 5 qty (nhưng vì CHO_DUYET không cộng counter → exporting_bill_qty chưa tăng)',
   '1. Duyệt phiếu',
   '',
   'Note: approve() chỉ update status, KHÔNG chạy syncTabs. Counter không thay đổi ở đây. (Counter đã được set khi tạo phiếu status=1 hoặc khi lưu status=3 trực tiếp)'); r += 1

# ============== VI. XOÁ ==============
sec(r, 'VI. XOÁ PHIẾU'); r += 1

tc(r, f'{P}_006.001', 'Xoá phiếu DANG_TAO (sau fix bug 2)', 'P0',
   'Phiếu status=1 do user tạo',
   '1. Click Xoá trên danh sách',
   '',
   'canDelete=true. DB::beginTransaction: xoá tabs (products, services, costs, tab) → xoá parent. Commit. Hiển thị success'); r += 1

tc(r, f'{P}_006.002', 'Xoá phiếu KHONG_DUYET', 'P0',
   'Phiếu status=4',
   '1. Click Xoá',
   '',
   'canDelete=true. Xoá thành công'); r += 1

tc(r, f'{P}_006.003', 'Chặn xoá phiếu CHO_DUYET (bảo mật sau fix bug 2)', 'P0',
   'Phiếu status=2',
   '1. Gọi trực tiếp URL /delete/{id}',
   '',
   'Sau fix bug 2: canDelete=false → trả về "Không có quyền xóa phiếu này!" với alert-type=error. KHÔNG xoá tabs (khác với bug cũ: xoá bất chấp)'); r += 1

tc(r, f'{P}_006.004', 'Chặn xoá phiếu DA_DUYET', 'P0',
   'Phiếu status=3',
   '1. Gọi URL /delete/{id}',
   '',
   'canDelete=false → không cho xoá'); r += 1

tc(r, f'{P}_006.005', 'Rollback khi lỗi xoá (sau fix bug 2)', 'P0',
   'Simulate lỗi DB khi xoá 1 trong 5 bảng',
   '1. Gây lỗi trong middle của delete flow',
   '',
   'Sau fix: DB::rollBack() → không có bảng nào bị xoá (atomicity). Trước fix: orphan rows có thể còn lại'); r += 1

# ============== VII. COUNTER LOGIC (BUG 3 SCENARIOS) ==============
sec(r, 'VII. COUNTER LOGIC (syncTabs) - SAU FIX BUG 3'); r += 1

tc(r, f'{P}_007.001', 'Tạo mới status=1: cộng counter', 'P0',
   'objectable.exporting_bill_qty = 0',
   '1. Tạo phiếu status=1 với qty=10',
   '',
   'Sau tạo: exporting_bill_qty = 10. exported_bill_qty = 0'); r += 1

tc(r, f'{P}_007.002', 'Tạo mới status=3 (nếu có luồng): cộng exported', 'P0',
   'objectable.exporting_bill_qty = 0',
   '1. Tạo phiếu status=3 với qty=10',
   '',
   'exporting_bill_qty = -10 (giảm 10), exported_bill_qty = 10'); r += 1

tc(r, f'{P}_007.003', 'Update từ DANG_TAO 10 → DANG_TAO 5 (giảm qty)', 'P0',
   'Phiếu status=1, qty=10, counter exporting=+10',
   '1. Mở sửa\n2. Đổi qty xuống 5\n3. Save status=1',
   '',
   'Counter: reverse -10 (theo oldStatus=1), rồi +5 (status mới=1) = NET +5 (tăng từ ban đầu 0). exported_bill_qty = 0'); r += 1

tc(r, f'{P}_007.004', 'Update từ DANG_TAO 5 → DANG_TAO 10 (tăng qty)', 'P0',
   'Phiếu status=1, qty=5, counter exporting=+5',
   '1. Sửa qty lên 10\n2. Save status=1',
   '',
   'Counter reverse -5, rồi +10 = NET +10'); r += 1

tc(r, f'{P}_007.005', 'Update từ DANG_TAO → CHO_DUYET (status thay đổi)', 'P0',
   'Phiếu status=1, qty=5, counter +5',
   '1. Sửa không đổi data\n2. Save status=2',
   '',
   'Counter reverse -5 (theo oldStatus=1), status mới=2 không cộng gì → NET = 0'); r += 1

tc(r, f'{P}_007.006', 'Update từ CHO_DUYET → CHO_DUYET (không tác động counter)', 'P0',
   'Phiếu status=2, counter đã về 0 sau lần submit',
   '1. Sửa ở status=2 (nếu có quyền admin)\n2. Save status=2',
   '',
   'oldStatus=2 → không reverse (logic chỉ reverse khi 1 hoặc 3). status mới=2 không cộng. Counter giữ nguyên'); r += 1

tc(r, f'{P}_007.007', 'Update từ KHONG_DUYET → CHO_DUYET', 'P0',
   'Phiếu status=4 (bị từ chối trước đó), counter đã về 0',
   '1. Sửa, gửi lại duyệt (status=2)',
   '',
   'oldStatus=4 → không reverse. Status mới=2 không cộng → counter giữ nguyên 0'); r += 1

tc(r, f'{P}_007.008', 'Update xoá tabs cũ, insert tabs hoàn toàn mới', 'P0',
   'Phiếu status=1 có 2 tab với 10 qty (counter +10)',
   '1. Mở sửa\n2. Xoá tất cả tab cũ\n3. Thêm tab mới với 3 qty\n4. Save status=1',
   '',
   'Reverse counter cho 2 tab cũ = -10. Insert tab mới = +3. NET +3'); r += 1

tc(r, f'{P}_007.009', 'Update với cost flag', 'P0',
   'Phiếu status=1 có cost A (exporting_bill = 1)',
   '1. Mở sửa\n2. Bỏ cost A\n3. Thêm cost B\n4. Save status=1',
   '',
   'cost A reverse: exporting_bill = 0. cost B apply: exporting_bill = 1'); r += 1

tc(r, f'{P}_007.010', 'Objectable bị xoá giữa chừng', 'P1',
   'Phiếu status=1, tab.objectable đã bị xoá',
   '1. Sửa và save',
   '',
   'reverseCountersFromExistingTabs skip bằng `continue` khi objectable null → không crash. Tab được xoá + tạo mới bình thường'); r += 1

# ============== VIII. VALIDATION + EDGE CASES ==============
sec(r, 'VIII. VALIDATION + EDGE CASES'); r += 1

tc(r, f'{P}_008.001', 'Bắt buộc contractable_id', 'P0',
   '',
   '1. Không chọn hợp đồng\n2. Submit',
   'contractable_id=null',
   'Validate lỗi: "contractable_id required"'); r += 1

tc(r, f'{P}_008.002', 'Bắt buộc customer_id', 'P0',
   '',
   '1. Submit không có customer',
   '',
   'Validate lỗi'); r += 1

tc(r, f'{P}_008.003', 'Bắt buộc handover_date', 'P0',
   '',
   '1. Submit không nhập ngày bàn giao',
   '',
   'Validate lỗi'); r += 1

tc(r, f'{P}_008.004', 'Submit status invalid', 'P1',
   '',
   '1. Submit status=99 (giá trị không có trong constants)',
   '',
   'Validate hoặc ignore, KHÔNG tạo phiếu status không hợp lệ'); r += 1

tc(r, f'{P}_008.005', 'Loading submit không reset sau lỗi AJAX (sau fix bug 7)', 'P1',
   '',
   '1. Ngắt mạng khi submit\n2. AJAX error',
   '',
   'Sau fix: $scope.loading.submit = false → nút "Lưu"/"Lưu và gửi duyệt" enabled trở lại. User submit lại được'); r += 1

tc(r, f'{P}_008.006', 'Amount tính chính xác khi có discount/vat', 'P0',
   '',
   '1. Nhập qty=10, price=1000, vat=10%, discount=100',
   '',
   'amount_after_discount = (10*1000) - 100 = 9900. vat_cost = 9900*10% = 990. amount_after_vat = 9900 + 990 = 10890'); r += 1

tc(r, f'{P}_008.007', 'Ngày bàn giao trong tương lai', 'P1',
   '',
   '1. Chọn handover_date > today',
   '',
   'Tuỳ business rule: cho hoặc không. Ghi rõ requirement'); r += 1

tc(r, f'{P}_008.008', 'Tax code format validation', 'P2',
   '',
   '1. Nhập tax_code không đúng định dạng VN (10 số)',
   '',
   'Validate format nếu có rule (phụ thuộc requirement)'); r += 1

# ============== IX. PHÂN QUYỀN ==============
sec(r, 'IX. PHÂN QUYỀN'); r += 1

tc(r, f'{P}_009.001', 'User không có quyền truy cập module', 'P0',
   'User không có permission liên quan',
   '1. Truy cập URL index',
   '',
   'Redirect hoặc 403'); r += 1

tc(r, f'{P}_009.002', 'User không có quyền xem tab "Tất cả"', 'P1',
   '',
   '1. Click tab Tất cả',
   '',
   'Chỉ hiện phiếu trong phạm vi quyền (công ty/phòng ban/bộ phận). Không hiện phiếu ngoài phạm vi'); r += 1

tc(r, f'{P}_009.003', 'User chỉ có quyền duyệt — không tạo', 'P1',
   '',
   '1. User A chỉ có quyền "Duyệt"\n2. Vào /create',
   '',
   'Chặn hoặc hiện không có button tạo'); r += 1

tc(r, f'{P}_009.004', 'User sửa phiếu của user khác (status DANG_TAO)', 'P0',
   'Phiếu status=1 do user B tạo, user A là admin',
   '1. User A vào /edit/{id}',
   '',
   'canEdit() check created_by == auth user → user A không phải creator → chặn. Trừ khi logic mở rộng cho role admin'); r += 1

# ============== X. NOTIFICATION ==============
sec(r, 'X. NOTIFICATION'); r += 1

tc(r, f'{P}_010.001', 'Notification khi gửi duyệt', 'P0',
   'Có user B có quyền "Kế toán thanh toán"',
   '1. User A tạo phiếu status=2 (hoặc sửa status=1 → 2)',
   '',
   'User B nhận notification "Bạn có một phiếu đề nghị xuất hóa đơn cần duyệt từ [tên user A]" với link đến show phiếu'); r += 1

tc(r, f'{P}_010.002', 'Không gửi notification khi lưu nháp (status=1)', 'P1',
   '',
   '1. Lưu phiếu status=1',
   '',
   'Không gửi notification. Chỉ gửi khi status == CHO_DUYET'); r += 1

tc(r, f'{P}_010.003', 'Notification link clickable', 'P1',
   '',
   '1. Click notification',
   '',
   'Redirect đúng show phiếu tương ứng'); r += 1

# Lưu file
output_file = '/Users/nguyentrancu/DEV/code/HRM/hrm-claude-config/.plans/export-bill-request-testcases/Testcase_ExportBillRequest.xlsx'
wb.save(output_file)
print(f'File created: {output_file}')
print(f'Total test cases: ~{r - 7}')
