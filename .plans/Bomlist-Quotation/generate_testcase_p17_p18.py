import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cascade Status + Chot GP"

# === STYLES ===
thin_border = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)

title_font = Font(bold=True, size=14, color='FF1F4E79')
header_font = Font(bold=True, size=11, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

section_font = Font(bold=True, size=11, color='FF1F4E79')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')

data_font = Font(size=11, color='FF000000')
data_alignment = Alignment(vertical='top', wrap_text=True)

summary_font = Font(size=11, color='FF000000')

# === COLUMN WIDTHS ===
col_widths = {
    'A': 14, 'B': 22, 'C': 16, 'D': 40, 'E': 10,
    'F': 30, 'G': 55, 'H': 22, 'I': 12, 'J': 60,
    'K': 15, 'L': 14, 'M': 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === ROW 1: Title + Test Summary ===
ws.merge_cells('A1:E1')
ws['A1'] = 'Testcase _ Cascade Status + Chốt giải pháp (Assign/Bomlist-Quotation P17-P18)'
ws['A1'].font = title_font

ws.merge_cells('F1:I1')
ws['F1'] = 'TEST SUMMARY'
ws['F1'].font = summary_font

summary_labels = [
    ('Số trường hợp kiểm thử đạt (P):', '=COUNTIF(L8:L500,"Passed")'),
    ('Số trường hợp kiểm thử không đạt (F):', '=COUNTIF(L8:L500,"Failed")'),
    ('Số trường hợp kiểm thử đang xem xét (PE):', '=COUNTIF(L8:L500,"Pending")'),
    ('Số trường hợp kiểm thử chưa thực hiện:', '=COUNTIF(L8:L500,"Not Executed")'),
    ('Tổng số trường hợp kiểm thử:', '=COUNTA(L8:L500)'),
]
for i, (label, formula) in enumerate(summary_labels):
    ws.cell(row=i+1, column=10, value=label).font = summary_font
    ws.cell(row=i+1, column=11, value=formula).font = summary_font

# === ROW 6: Header ===
headers = ['Module', 'Nhóm chức năng', 'TC ID', 'Chức năng', 'Priority',
           'Tiền điều kiện', 'Bước thực hiện', 'Test Data', 'Test Data',
           'Expected Result (chi tiết)', 'KQ thực tế', 'Status', 'Ghi chú']
ws.row_dimensions[6].height = 30
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# === HELPER ===
current_row = 7
MODULE = 'Giao việc'
GROUP = 'Cascade Status + Chốt GP'
PREFIX = 'CSG'


def write_section_row(title):
    global current_row
    ws.merge_cells(f'C{current_row}:M{current_row}')
    cell = ws.cell(row=current_row, column=3, value=title)
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1


def write_tc(section_num, tc_num, func, priority, precondition, steps, test_data, expected):
    global current_row
    tc_id = f'{PREFIX}_{section_num:03d}.{tc_num:03d}'
    values = [MODULE, GROUP, tc_id, func, priority,
              precondition, steps, test_data, '', expected, '', 'Not Executed', '']
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    current_row += 1


# =============================================
# I. PHASE 17 — CASCADE STATUS KHI GỬI YCBG
# =============================================
write_section_row('I. CASCADE STATUS KHI GỬI YCBG (Phase 17)')

write_tc(1, 1,
    'Gửi YCBG → Dự án TKT đổi sang "Dự toán"',
    'P0',
    'Dự án TKT có giải pháp, có BOM tổng hợp hoàn thành.\nUser là NV KD phụ trách dự án.\nĐã tạo YCBG ở trạng thái Nháp.',
    '1. Vào chi tiết dự án TKT > tab Hồ sơ\n2. Mở YCBG đang ở trạng thái "Đang tạo"\n3. Click "Gửi"\n4. Quay lại danh sách dự án TKT, kiểm tra cột trạng thái',
    'YCBG status = 1 (Đang tạo)',
    'Dự án TKT hiển thị trạng thái "Dự toán" (status=6)')

write_tc(1, 2,
    'Gửi YCBG → Giải pháp đổi sang "Chờ làm giá"',
    'P0',
    'Như TC CSG_001.001',
    '1. Sau khi gửi YCBG thành công\n2. Vào chi tiết giải pháp\n3. Kiểm tra trạng thái giải pháp',
    '',
    'Giải pháp hiển thị trạng thái "Chờ làm giá" (status=15)')

write_tc(1, 3,
    'Gửi YCBG → YCBG đổi sang "Chờ XD giá"',
    'P0',
    'Như TC CSG_001.001',
    '1. Sau khi gửi YCBG thành công\n2. Kiểm tra trạng thái YCBG trong tab Hồ sơ',
    '',
    'YCBG hiển thị trạng thái "Chờ XD giá" (status=2)\nCó timestamp sent_at')

write_tc(1, 4,
    'Gửi YCBG nhiều lần → status vẫn đúng (idempotent)',
    'P1',
    'Dự án TKT đang ở trạng thái "Dự toán".\nĐã có YCBG trước đó.\nTạo YCBG mới ở trạng thái Nháp.',
    '1. Tạo YCBG mới cho cùng dự án\n2. Gửi YCBG mới\n3. Kiểm tra status dự án TKT + giải pháp',
    'Dự án đã có status=6',
    'Dự án TKT vẫn ở "Dự toán" (status=6)\nGiải pháp vẫn ở "Chờ làm giá" (status=15)\nKhông lỗi, không thay đổi gì bất thường')

write_tc(1, 5,
    'Gửi YCBG → Thông báo đến người xây dựng giá',
    'P1',
    'Như TC CSG_001.001\nCó user có quyền "Xây dựng giá Bom giải pháp"',
    '1. Gửi YCBG thành công\n2. Đăng nhập user có quyền "Xây dựng giá Bom giải pháp"\n3. Kiểm tra thông báo',
    '',
    'User nhận được thông báo mới về YCBG vừa gửi')


# =============================================
# II. PHASE 17 — CASCADE STATUS KHI DUYỆT BÁO GIÁ
# =============================================
write_section_row('II. CASCADE STATUS KHI DUYỆT BÁO GIÁ (Phase 17)')

write_tc(2, 1,
    'Tự duyệt (Level 1) → Dự án đổi "Thương thảo giá"',
    'P0',
    'Báo giá ở trạng thái "Đang tạo".\nCấp duyệt = Level 1 (tự duyệt).\nĐã nhập đầy đủ giá > 0.',
    '1. Vào báo giá > click "Gửi duyệt"\n2. Hệ thống xác nhận cấp tự duyệt → click "Tự duyệt"\n3. Kiểm tra trạng thái dự án TKT',
    'price_approval_level = 1',
    'Báo giá → "Đã duyệt" (status=4)\nDự án TKT → "Thương thảo giá" (status=7)\nGiải pháp → "Đã duyệt giá" (status=13)')

write_tc(2, 2,
    'Tự duyệt (Level 1) → Giải pháp đổi "Đã duyệt giá"',
    'P0',
    'Như TC CSG_002.001',
    '1. Sau khi tự duyệt thành công\n2. Vào chi tiết giải pháp\n3. Kiểm tra trạng thái',
    '',
    'Giải pháp hiển thị "Đã duyệt giá" (status=13)\nMàu badge: #059669')

write_tc(2, 3,
    'TP duyệt (Level 2) → cascade status đúng',
    'P0',
    'Báo giá ở trạng thái "Chờ TP duyệt" (status=2).\nCấp duyệt = Level 2.\nUser có quyền "Trưởng phòng duyệt giá Bom giải pháp".',
    '1. Đăng nhập TP\n2. Vào danh sách báo giá chờ duyệt\n3. Click duyệt báo giá\n4. Kiểm tra trạng thái dự án + giải pháp',
    'price_approval_level = 2',
    'Báo giá → "Đã duyệt" (status=4)\nDự án TKT → "Thương thảo giá" (status=7)\nGiải pháp → "Đã duyệt giá" (status=13)\nYCBG → "Đã có báo giá" (status=4)')

write_tc(2, 4,
    'BGĐ duyệt (Level 3) → cascade status đúng',
    'P0',
    'Báo giá ở trạng thái "Chờ BGĐ duyệt" (status=3).\nCấp duyệt = Level 3.\nTP đã duyệt.\nUser có quyền "Ban giám đốc duyệt giá Bom giải pháp".',
    '1. Đăng nhập BGĐ\n2. Vào danh sách báo giá chờ duyệt\n3. Click duyệt báo giá\n4. Kiểm tra trạng thái dự án + giải pháp',
    'price_approval_level = 3',
    'Báo giá → "Đã duyệt" (status=4)\nDự án TKT → "Thương thảo giá" (status=7)\nGiải pháp → "Đã duyệt giá" (status=13)\nYCBG → "Đã có báo giá" (status=4)')

write_tc(2, 5,
    'TP duyệt Level 3 → KHÔNG cascade (chỉ chuyển BGĐ)',
    'P0',
    'Báo giá ở trạng thái "Chờ TP duyệt" (status=2).\nCấp duyệt = Level 3.',
    '1. Đăng nhập TP\n2. Duyệt báo giá\n3. Kiểm tra trạng thái dự án + giải pháp',
    'price_approval_level = 3',
    'Báo giá → "Chờ BGĐ duyệt" (status=3)\nDự án TKT KHÔNG đổi status (vẫn giữ nguyên)\nGiải pháp KHÔNG đổi status\nChưa cascade vì chưa phải bước duyệt cuối')

write_tc(2, 6,
    'Từ chối báo giá → KHÔNG rollback status dự án/GP',
    'P1',
    'Báo giá ở trạng thái "Chờ TP duyệt" hoặc "Chờ BGĐ duyệt".\nDự án đang ở "Dự toán" (status=6).',
    '1. TP/BGĐ từ chối báo giá\n2. Kiểm tra trạng thái dự án TKT + giải pháp',
    '',
    'Báo giá → "Đang tạo" (status=1)\nDự án TKT giữ nguyên status hiện tại\nGiải pháp giữ nguyên status hiện tại\nKhông rollback')


# =============================================
# III. PHASE 18 — BUTTON CHỐT GIẢI PHÁP
# =============================================
write_section_row('III. BUTTON "CHỐT GIẢI PHÁP" — HIỂN THỊ (Phase 18)')

write_tc(3, 1,
    'Button hiện khi NV KD + có hồ sơ Đã duyệt',
    'P0',
    'User là NV KD phụ trách (main_sale_employee_id).\nDự án chưa đóng.\nCó ít nhất 1 hồ sơ giải pháp status "Đã duyệt".',
    '1. Đăng nhập NV KD\n2. Vào chi tiết dự án TKT',
    '',
    'Footer hiển thị button "Chốt giải pháp" với icon ri-check-double-line\nButton màu primary (xanh)')

write_tc(3, 2,
    'Button hiện khi có hồ sơ "Hết hiệu lực"',
    'P0',
    'User là NV KD phụ trách.\nChỉ có hồ sơ status "Hết hiệu lực" (không có "Đã duyệt").',
    '1. Đăng nhập NV KD\n2. Vào chi tiết dự án TKT',
    'Hồ sơ status = expired',
    'Button "Chốt giải pháp" vẫn hiển thị')

write_tc(3, 3,
    'Button ẨN khi không phải NV KD phụ trách',
    'P0',
    'User KHÔNG phải main_sale_employee_id.\nDự án có hồ sơ Đã duyệt.',
    '1. Đăng nhập user khác (PM, Leader, admin...)\n2. Vào chi tiết dự án TKT',
    '',
    'KHÔNG thấy button "Chốt giải pháp" ở footer')

write_tc(3, 4,
    'Button ẨN khi không có hồ sơ đủ điều kiện',
    'P1',
    'User là NV KD phụ trách.\nTất cả hồ sơ đều ở status Nháp/Chờ duyệt/Không duyệt.',
    '1. Vào chi tiết dự án TKT',
    'Hồ sơ status = draft/pending/rejected',
    'KHÔNG thấy button "Chốt giải pháp"')

write_tc(3, 5,
    'Button ẨN khi dự án đã đóng',
    'P1',
    'Dự án đã đóng (status=11).\nCó hồ sơ Đã duyệt.',
    '1. Vào chi tiết dự án đã đóng',
    '',
    'KHÔNG thấy button "Chốt giải pháp"\nHiển thị banner đỏ "Dự án đã đóng"')


# =============================================
# IV. PHASE 18 — POPUP CHỐT GIẢI PHÁP
# =============================================
write_section_row('IV. POPUP CHỐT GIẢI PHÁP — GIAO DIỆN (Phase 18)')

write_tc(4, 1,
    'Popup hiển thị đúng cấu trúc',
    'P0',
    'Button "Chốt giải pháp" hiển thị.',
    '1. Click button "Chốt giải pháp"',
    '',
    'Popup title: "Chốt giải pháp"\nBảng hồ sơ: cột Radio, Mã hồ sơ, Version GP, Trạng thái, Ngày duyệt\nTextarea: "Ghi chú chốt giải pháp"\nButton: "Lưu & gửi thông báo" (primary) + "Đóng" (light)\nSize: lg')

write_tc(4, 2,
    'Bảng chỉ hiện hồ sơ Đã duyệt + Hết hiệu lực',
    'P0',
    'Có hồ sơ ở nhiều trạng thái: Nháp, Chờ duyệt, Đã duyệt, Không duyệt, Hết hiệu lực.',
    '1. Click "Chốt giải pháp"\n2. Kiểm tra danh sách hồ sơ trong popup',
    '5 hồ sơ ở 5 status khác nhau',
    'Chỉ hiện hồ sơ "Đã duyệt" và "Hết hiệu lực"\nKhông hiện Nháp, Chờ duyệt, Không duyệt')

write_tc(4, 3,
    'Badge trạng thái hiển thị đúng màu',
    'P1',
    'Popup đang mở, có cả 2 loại hồ sơ.',
    '1. Quan sát cột Trạng thái trong bảng',
    '',
    'Đã duyệt = xanh lá (#16A34A)\nHết hiệu lực = xám (#6B7280)')

write_tc(4, 4,
    'Button "Lưu & gửi thông báo" disabled khi chưa chọn hồ sơ',
    'P0',
    'Popup vừa mở, chưa chọn hồ sơ nào.',
    '1. Kiểm tra trạng thái button "Lưu & gửi thông báo"',
    '',
    'Button bị disabled (không click được)\nGhi chú có thể để trống (không bắt buộc)')

write_tc(4, 5,
    'Chọn hồ sơ bằng radio → enable button',
    'P0',
    'Popup đang mở, có ≥1 hồ sơ.',
    '1. Click vào 1 dòng hồ sơ bất kỳ (hoặc radio)\n2. Kiểm tra trạng thái button',
    '',
    'Radio được chọn, dòng highlight\nButton "Lưu & gửi thông báo" enable\nChỉ chọn được 1 hồ sơ (radio, không checkbox)')

write_tc(4, 6,
    'Nhập ghi chú tối đa 1000 ký tự',
    'P2',
    'Popup đang mở.',
    '1. Nhập > 1000 ký tự vào textarea ghi chú\n2. Kiểm tra',
    'Chuỗi > 1000 ký tự',
    'Textarea giới hạn maxlength=1000\nKhông nhập thêm được khi đạt giới hạn')


# =============================================
# V. PHASE 18 — CHỐT GIẢI PHÁP — CHỨC NĂNG
# =============================================
write_section_row('V. CHỐT GIẢI PHÁP — CHỨC NĂNG (Phase 18)')

write_tc(5, 1,
    'Chốt thành công → toast success + đóng popup',
    'P0',
    'Popup mở, đã chọn 1 hồ sơ, đã nhập ghi chú.',
    '1. Chọn 1 hồ sơ "Đã duyệt"\n2. Nhập ghi chú: "Khách hàng đồng ý"\n3. Click "Lưu & gửi thông báo"',
    'review_profile_id hợp lệ',
    'Toast success: "Đã chốt giải pháp thành công"\nPopup tự đóng\nTrang reload dữ liệu mới')

write_tc(5, 2,
    'Hồ sơ được chọn → status "Đã chốt"',
    'P0',
    'Vừa chốt thành công.',
    '1. Vào tab Hồ sơ\n2. Kiểm tra trạng thái hồ sơ vừa được chọn',
    '',
    'Hồ sơ đổi sang trạng thái "Đã chốt" (finalized)\nBadge màu #4F46E5\nfinalized_at có giá trị (datetime)\nfinalized_note = ghi chú đã nhập')

write_tc(5, 3,
    'YCGP → status "Đã chốt giải pháp"',
    'P0',
    'Vừa chốt thành công.',
    '1. Kiểm tra trạng thái Yêu cầu làm giải pháp (RequestSolution)',
    '',
    'YCGP đổi sang "Đã chốt giải pháp" (status=11)\nBadge màu #4F46E5')

write_tc(5, 4,
    'Giải pháp → status "Chốt giải pháp"',
    'P0',
    'Vừa chốt thành công.',
    '1. Vào chi tiết giải pháp\n2. Kiểm tra trạng thái',
    '',
    'Giải pháp đổi sang "Chốt giải pháp" (status=17)\nBadge màu #4F46E5')

write_tc(5, 5,
    'Dự án TKT → status "Thương thảo dự án hợp đồng"',
    'P0',
    'Vừa chốt thành công.',
    '1. Quay lại danh sách dự án TKT\n2. Kiểm tra cột trạng thái',
    '',
    'Dự án đổi sang "Thương thảo dự án hợp đồng" (status=8)')

write_tc(5, 6,
    'Thông báo gửi đến PM giải pháp',
    'P0',
    'Giải pháp có pm_id.\nVừa chốt thành công.',
    '1. Đăng nhập tài khoản PM giải pháp\n2. Kiểm tra thông báo',
    '',
    'PM nhận thông báo: "[Tên NV KD] đã chốt giải pháp dự án [Tên dự án]. Hồ sơ: [Mã HS]."\nClick thông báo → vào chi tiết dự án TKT')

write_tc(5, 7,
    'Thông báo gửi đến người tạo giải pháp',
    'P0',
    'Giải pháp có created_by (khác PM).\nVừa chốt thành công.',
    '1. Đăng nhập tài khoản người tạo giải pháp\n2. Kiểm tra thông báo',
    '',
    'Người tạo nhận thông báo tương tự PM\nURL dẫn đến /assign/prospective-projects/{id}/manager')

write_tc(5, 8,
    'NV KD tự mình KHÔNG nhận thông báo',
    'P1',
    'NV KD vừa thực hiện chốt giải pháp.',
    '1. Kiểm tra thông báo của chính NV KD',
    '',
    'NV KD KHÔNG nhận thông báo cho chính mình (loại current user)')

write_tc(5, 9,
    'Chốt với ghi chú trống (không bắt buộc)',
    'P1',
    'Popup mở, đã chọn hồ sơ, ghi chú để trống.',
    '1. Chọn 1 hồ sơ\n2. Để trống ghi chú\n3. Click "Lưu & gửi thông báo"',
    'note = null',
    'Chốt thành công\nfinalized_note = null trong DB\nCascade status đầy đủ')

write_tc(5, 10,
    'Chốt hồ sơ "Hết hiệu lực" (không chỉ "Đã duyệt")',
    'P1',
    'Chỉ có hồ sơ ở status "Hết hiệu lực".',
    '1. Click "Chốt giải pháp"\n2. Chọn hồ sơ "Hết hiệu lực"\n3. Click "Lưu & gửi thông báo"',
    'Hồ sơ status = expired',
    'Chốt thành công\nHồ sơ đổi sang "Đã chốt"\nCascade status đầy đủ (YCGP, GP, Dự án)')

write_tc(5, 11,
    'Sau chốt → button "Chốt giải pháp" biến mất',
    'P0',
    'Vừa chốt thành công, trang đã reload.',
    '1. Kiểm tra footer sau khi popup đóng',
    '',
    'Button "Chốt giải pháp" KHÔNG còn hiển thị\n(Vì hồ sơ đã chuyển sang finalized → không còn hồ sơ approved/expired)')


# =============================================
# VI. PHASE 18 — EDGE CASES + BẢO MẬT
# =============================================
write_section_row('VI. EDGE CASES + BẢO MẬT (Phase 18)')

write_tc(6, 1,
    'API finalize-solution bị gọi bởi user không phải NV KD',
    'P0',
    'User KHÔNG phải main_sale_employee_id.\nGọi API trực tiếp.',
    '1. Gọi POST /assign/prospective-projects/{id}/finalize-solution\n   với review_profile_id hợp lệ',
    'User khác NV KD',
    'Response 403: "Chỉ NV KD phụ trách mới được chốt giải pháp."')

write_tc(6, 2,
    'API với review_profile_id không thuộc solution của dự án',
    'P0',
    'NV KD gọi API với profile_id thuộc dự án khác.',
    '1. Gọi POST /assign/prospective-projects/{id}/finalize-solution\n   với review_profile_id của dự án khác',
    'review_profile_id invalid',
    'Response 422: "Hồ sơ không hợp lệ hoặc không ở trạng thái Đã duyệt / Hết hiệu lực."')

write_tc(6, 3,
    'API với review_profile_id ở status draft/pending/rejected',
    'P0',
    'NV KD gọi API với profile_id status = draft.',
    '1. Gọi POST /assign/prospective-projects/{id}/finalize-solution\n   với review_profile_id status=draft',
    'Hồ sơ status != approved/expired',
    'Response 422: "Hồ sơ không hợp lệ..."')

write_tc(6, 4,
    'API với review_profile_id không tồn tại',
    'P1',
    'NV KD gọi API.',
    '1. Gọi POST với review_profile_id = 999999',
    'review_profile_id = 999999',
    'Response 422: "Hồ sơ không tồn tại" (validation rule exists)')

write_tc(6, 5,
    'API với note > 1000 ký tự',
    'P1',
    'NV KD gọi API.',
    '1. Gọi POST với note = chuỗi 1001 ký tự',
    'note length = 1001',
    'Response 422: "Ghi chú tối đa 1000 ký tự"')

write_tc(6, 6,
    'Dự án chưa có giải pháp → API trả lỗi',
    'P1',
    'Dự án TKT chưa được tạo giải pháp.',
    '1. Gọi POST /assign/prospective-projects/{id}/finalize-solution',
    '',
    'Response 422: "Dự án chưa có giải pháp."')

write_tc(6, 7,
    'Chốt đồng thời (concurrent) → chỉ 1 request thành công',
    'P2',
    '2 tab trình duyệt mở cùng popup chốt GP.',
    '1. Tab A: chọn hồ sơ X → click "Lưu & gửi thông báo"\n2. Tab B: chọn cùng hồ sơ X → click "Lưu & gửi thông báo" ngay sau',
    '',
    'Tab A: thành công\nTab B: lỗi 422 (hồ sơ đã chuyển sang finalized → không còn approved/expired)')


# === DATA VALIDATION ===
dv = DataValidation(type="list", formula1='"Passed,Failed,Pending,Not Executed"', allow_blank=True)
dv.error = "Vui lòng chọn: Passed, Failed, Pending hoặc Not Executed"
dv.errorTitle = "Giá trị không hợp lệ"
ws.add_data_validation(dv)
dv.add(f'L7:L{current_row}')

# === SAVE ===
output_path = 'docs/srs/bomlist-quotation-p17-p18-testcases.xlsx'
wb.save(output_path)
print(f'Saved to {output_path} — {current_row - 7} test cases')
